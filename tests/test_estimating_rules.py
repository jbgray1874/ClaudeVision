"""
test_estimating_rules.py — lock the costing rules against regression.

WHY THIS EXISTS. In one working session, seven of ten commits fixed defects that earlier
commits in the same session had introduced or exposed. Every one was found by a person
reading a diff or a spreadsheet, days or hours after it shipped. Nothing in the repository
would have caught any of them.

Each test below is a defect that reached a real estimate. The number in it is the number the
job actually produced. They are deliberately written against the PURE functions — material
normalisation, thickness selection, route gating, precedence — so the whole suite runs in a
second with no SolidWorks, no Excel, no SQL and no drawings.

Run:
    python tests/test_estimating_rules.py          # plain, no dependencies
    pytest tests/test_estimating_rules.py          # if pytest is installed
"""
from __future__ import annotations

import os
import sys
import traceback

# NO LIVE ANYTHING. This suite is meant to run in a second with no SolidWorks, no Excel, no
# SQL and no drawings — and it was opening a connection to the production SDILive server
# just by importing learning_engine, which tests its pure functions. A suite that needs the
# production database is not isolated, cannot run in CI, and makes every fixture depend on a
# machine being reachable. Set BEFORE any src import, because the connection was made at
# module import time.
os.environ.setdefault("SDI_OFFLINE", "1")

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "src"))

_FAILS = []

# Under pytest a failure MUST raise, or pytest reports a green run over a broken engine:
# these helpers only ever appended to a list, so `pytest tests/test_estimating_rules.py`
# passed against every mutation the plain runner caught. The plain runner keeps collecting,
# so it can still report every failure in a test rather than stopping at the first.
_COLLECT_ONLY = False


def _fail(msg: str) -> None:
    _FAILS.append(msg)
    if not _COLLECT_ONLY:
        raise AssertionError(msg)


def eq(actual, expected, what):
    if actual != expected:
        _fail(f"{what}\n      expected: {expected!r}\n      actual:   {actual!r}")


def ok(cond, what):
    if not cond:
        _fail(what)


# ── material identity — job 0348837 (Horti Crate) ────────────────────────────────────
def test_joinery_part_numbers_are_not_steel():
    """'-J01' was an unconditional return MILD_STEEL ahead of every material check, so the
    Crate's timber panels were routed to laser, weld and powder. J reads as JOINERY."""
    from json_normaliser import normalise_material_for_part as f
    base = {"surface_finishes": [], "textual_operations": [],
            "description": "HORTI CRATE LOWER FRONT PANEL"}
    eq(f({**base, "part_number": "11225-01-J01", "materials": ["FSC PINE"]}),
       "TIMBER", "-J part stated FSC PINE must resolve TIMBER")
    eq(f({**base, "part_number": "11225-01-J08", "materials": ["MR MDF"]}),
       "MDF", "-J part stated MR MDF must resolve MDF")
    # '-M' remains SDI's metal convention, and still wins when nothing contradicts it.
    eq(f({"part_number": "12120-01-01M", "materials": [], "surface_finishes": [],
          "textual_operations": [], "description": ""}),
       "MILD_STEEL", "-M with no stated material stays MILD_STEEL")
    # ...but yields to a stated timber, as the '-SA' rule always did.
    eq(f({"part_number": "12345-01-02M", "materials": ["FSC PINE"], "surface_finishes": [],
          "textual_operations": [], "description": "TIMBER RAIL"}),
       "TIMBER", "-M stated as pine must not stay steel")


def test_species_resolve_to_a_family():
    """The vocabulary knew TIMBER/WOOD/MDF/PLYWOOD — words a title block never uses. It
    states the species, so a stated material resolved to nothing and part-number hints won."""
    from json_normaliser import normalise_material_for_part as f
    b = {"part_number": "X-01", "surface_finishes": [], "textual_operations": [], "description": ""}
    for stated, expect in (("FSC PINE", "TIMBER"), ("PINE", "TIMBER"), ("SPRUCE", "TIMBER"),
                           ("MR MDF", "MDF"), ("MARINE PLY", "PLYWOOD")):
        eq(f({**b, "materials": [stated]}), expect, f"{stated} must resolve {expect}")
    # A veneered board is a board, whatever species is on its face.
    eq(f({**b, "materials": ["OAK VENEER MDF"]}), "OAK_VENEER_MDF",
       "OAK VENEER MDF must not collapse to solid oak")
    # A pine note on a steel part must not flip it.
    eq(f({**b, "materials": ["MILD STEEL"], "description": "PINE PACKER ON MILD STEEL FRAME"}),
       "MILD_STEEL", "a species mentioned alongside steel must not flip the part")


# ── thickness — 0.5mm TIMBER reached the Crate's labour rows ──────────────────────────
def test_board_thickness_floor():
    """0.5mm sheet steel is ordinary stock; 0.5mm timber is not a thing. Sheet board and
    solid timber have different physical minimums, so one floor cannot serve both."""
    from estimator import _safe_thickness_mm as thk
    def t(mat, vals, dxf=""):
        return thk({"normalized_material": mat, "thicknesses_mm": vals, "dxf_source_file": dxf})
    eq(t("TIMBER", [0.5]), None, "0.5mm on timber must be rejected")
    eq(t("MDF", [0.5]), None, "0.5mm on MDF must be rejected")
    eq(t("FSC PINE", [4.0]), None, "4mm solid timber is below the 6mm floor")
    eq(t("MR MDF", [4.0]), 4.0, "4mm MDF is real sheet board and must pass")
    eq(t("OAK VENEER MDF", [4.0]), 4.0, "a veneered board takes the BOARD floor, not timber's")
    eq(t("OAK", [4.0]), None, "solid oak at 4mm takes the timber floor")
    eq(t("TIMBER", [18.0]), 18.0, "a real board gauge passes")
    # Metal and acrylic are untouched.
    eq(t("MILD_STEEL", [0.5]), 0.5, "0.5mm steel is ordinary stock")
    eq(t("ACRYLIC", [3.0]), 3.0, "3mm acrylic passes")
    # Ordering: the tolerance-table strip must run BEFORE the board floor, or the 3.0 that
    # is also part of that table survives and looks like a real 3mm board.
    eq(t("MDF", [0.5, 1.0, 1.5, 2.0, 3.0]), None,
       "a board whose only values ARE the tolerance table has no thickness")
    eq(t("MDF", [0.5, 1.0, 1.5, 2.0, 3.0, 18.0]), 18.0,
       "the real gauge survives the tolerance table")
    eq(t("MILD_STEEL", [0.5, 1.0, 1.5, 2.0, 3.0]), 0.5,
       "for METAL the tolerance values are real gauges and the fallback still applies")


# ── route gating ─────────────────────────────────────────────────────────────────────
def test_timber_is_never_welded():
    """A weld cue in border boilerplate booked Weld (CO2) AND its chained Dress Welds
    against the Crate's timber panels. Both departments from one legend phrase."""
    from estimator import estimate_process_times
    p = {"part_number": "11225-01-J01", "normalized_material": "TIMBER",
         "materials": ["FSC PINE"], "textual_operations": ["welding", "saw", "glue"],
         "normalized_thickness_mm": 18.0, "overall_length_mm": 600, "overall_width_mm": 400,
         "manufacturing_features": {"welding_required": True, "bend_count": 0},
         "risk_flags": ["weld_required"], "geometry_rollup": {}}
    estimate_process_times(p)
    ok("welding" not in (p.get("textual_operations") or []), "weld op removed from timber")
    ok("dress_welds" not in (p.get("textual_operations") or []), "dress not chained onto timber")
    # The SIGNAL must go with the op, or the report tells the estimator to verify weld
    # content on a part whose sheet has no weld line.
    eq(p["manufacturing_features"].get("welding_required"), False, "weld signal cleared")
    ok("weld_required" not in p.get("risk_flags", []), "weld risk flag cleared")

    s = {"part_number": "X-01M", "normalized_material": "MILD_STEEL", "materials": ["MILD STEEL"],
         "textual_operations": ["welding"], "normalized_thickness_mm": 1.5,
         "overall_length_mm": 300, "overall_width_mm": 200,
         "manufacturing_features": {"welding_required": True}, "risk_flags": ["weld_required"],
         "geometry_rollup": {}}
    estimate_process_times(s)
    ok("welding" in (s.get("textual_operations") or []), "steel keeps its weld")
    eq(s["manufacturing_features"].get("welding_required"), True, "steel keeps its weld signal")


def test_measured_flats_get_no_blank_allowance():
    """A DXF or cut-list flat IS the developed blank. The allowance estimates one from
    FORMED dimensions, and adding it to a measured blank inflated 03M from 45x20 to 51x26 —
    47% more material area, on every DXF-backed job."""
    from document_builder import _build_normalized_geometry as g
    def blank(L, W, t, bends, **marks):
        p = {"manufacturing_features": {"bend_count": bends}, "overall_length_mm": L,
             "overall_width_mm": W, "normalized_thickness_mm": t, "angles_deg": []}
        p.update(marks)
        b = g(p)["bounding_box_flat_mm"]
        return (b["length"], b["width"])
    eq(blank(45.0, 20.0, 1.5, 0, flat_pattern_detected=True), (45.0, 20.0),
       "12120-01-03M measured flat must be unchanged")
    eq(blank(126.39, 82.2, 1.5, 2, geometry_source="dxf"), (126.39, 82.2),
       "12120-01-01M measured flat must be unchanged")
    eq(blank(120.0, 80.0, 2.0, 0), (120.0, 80.0),
       "a part with no bends does not develop — formed size IS blank size")
    # PDF-only WITH bends is where the allowance is a genuine estimate, and still applies.
    eq(blank(120.0, 80.0, 2.0, 2), (128.0, 88.0),
       "formed dimensions with bends still take the allowance")


# ── make vs buy ──────────────────────────────────────────────────────────────────────
def test_bought_in_carries_no_fabrication_labour():
    """The UPC sticker was a catalogue line at GBP 1.05 in one path and a null-numbered
    record carrying weld, dress, powder and glue in another. Same item, two identities."""
    from bought_in_policy import (is_bought_in, has_fabrication_evidence,
                                  bought_in_conflict, strip_fabrication_ops)
    sticker = {"part_number": "UPC STICKER; CLEAR", "page_roles": ["bought_in"],
               "textual_operations": ["welding", "dress_welds", "powder_coating",
                                      "glue", "handling"]}
    ok(is_bought_in(sticker), "a bought_in page role classifies")
    strip_fabrication_ops(sticker)
    eq(sticker["textual_operations"], ["handling"],
       "fabrication removed, handling KEPT — we do fit purchased components")

    fab = {"part_number": "12120-01-01M", "dxf_augmented": True,
           "geometry_source": "dxf_flat_pattern",
           "textual_operations": ["laser_cutting", "folding", "handling"]}
    ok(not is_bought_in(fab), "a fabricated part is not bought-in")
    eq(strip_fabrication_ops(fab), [], "a fabricated part keeps its route")

    # A DRAWING is not fabrication evidence — we draw bought-ins to locate them.
    drawn = {"part_number": "BI-KNURLEDKNOB", "dxf_source_file": "BI-KNURLEDKNOB.DXF",
             "geometry_source": "dxf_matched_no_geometry", "dxf_measured_outline": False}
    ok(not has_fabrication_evidence(drawn),
       "a matched-but-unmeasured DXF is not fabrication evidence")
    ok(not bought_in_conflict(drawn), "and must not raise a false make/buy conflict")
    # Nor is an extents fallback, which also sets flat_pattern_detected.
    ok(not has_fabrication_evidence({"flat_pattern_detected": True,
                                     "dxf_measured_outline": False}),
       "the extents fallback is not measured geometry")
    # A genuine conflict IS surfaced.
    ok(bought_in_conflict({"part_number": "BI-BRACKET", "dxf_augmented": True,
                           "geometry_source": "dxf_flat_pattern"}),
       "bought-in identity plus a measured flat must conflict, not auto-resolve")


# ── precedence ───────────────────────────────────────────────────────────────────────
def test_a_weaker_source_cannot_replace_a_stronger_one():
    """The PDF GA-tree pass overwrote quantities from the SolidWorks assembly BOM, and
    knowledge-base rules replaced native material. Both silently."""
    from source_precedence import rank, apply_field
    ok(rank("knowledge_base (92%)") > rank("solidworks_api"),
       "a person outranks the model")
    ok(rank("solidworks_api") > rank("dxf_flat_pattern") > rank("bom_tree") > rank("llm_full_extract"),
       "model > measured DXF > printed table > transcription")
    eq(rank("wibble"), 0, "an unknown source fills gaps only")

    native = {"quantity": 2, "quantity_source": "solidworks_api"}
    ok(not apply_field(native, "quantity", 1, "bom_tree"), "GA tree must not overwrite native qty")
    eq(native["quantity"], 2, "native quantity kept")
    ok(any("NOT applied" in f for f in native.get("review_flags", [])),
       "and the disagreement is recorded, not swallowed")

    weak = {"quantity": 3, "quantity_source": "llm_full_extract"}
    ok(apply_field(weak, "quantity", 5, "bom_tree"), "a stronger source may still correct a weaker one")
    gap = {}
    ok(apply_field(gap, "quantity", 4, "bom_tree"), "anything may fill a gap")

    from learning_engine import _is_reliable_material
    for src in ("solidworks_api", "dxf_flat_pattern", "knowledge_base (90%)"):
        ok(_is_reliable_material({"normalized_material": "MILD_STEEL", "material_source": src}),
           f"{src} must be protected from override rules")


# ── what we describe must be what we priced ──────────────────────────────────────────
def test_deliverables_describe_only_costed_operations():
    """The quote promised powder coating and weld dressing on a lacquered timber crate the
    sheet charges neither for; the report congratulated the engine for powder-coating eight
    timber panels."""
    from costed_facts import (costed_operations, costed_finish_label,
                              part_numbers_with_operation, operations_for_part)
    src = {
        "workbook_labour": {"rows": [
            {"wb_operation": "Assemble/pack (Acrylic)", "engine_operations": ["handling"],
             "part_numbers": ["11225-01-J01"], "workbook_row": 96},
            {"wb_operation": "CNC / Joinery machining", "engine_operations": ["cnc_routing"],
             "part_numbers": ["11225-01-J01"], "workbook_row": 97},
            {"wb_operation": "Spray / Wet Paint", "engine_operations": ["wet_spray"],
             "part_numbers": ["11225-01-J01"], "workbook_row": 98},
            {"wb_operation": "P.Coat", "engine_operations": ["powder_coating"],
             "part_numbers": ["11225-01-J01"], "workbook_row": 99},
        ]},
        "final_estimate": {"labour_rows": [
            {"operation": "Assemble/pack (Acrylic)", "workbook_row": 96, "total_value_gbp": 0.29},
            {"operation": "CNC / Joinery machining", "workbook_row": 97, "total_value_gbp": 10.19},
            {"operation": "Spray / Wet Paint", "workbook_row": 98, "total_value_gbp": 5.87},
            # priced at nothing by the sheet — not part of the job
            {"operation": "P.Coat", "workbook_row": 99, "batch_hours": 2.8, "total_value_gbp": 0.0},
        ]},
        # the PRE-FILTER engine fields still carry powder and weld; they must not win
        "estimate_summary": {"part_estimates": [
            {"part_number": "11225-01-J01", "labour_estimate": {"costs_gbp": {
                "powder_coating": 3.1, "welding": 2.0, "dress_welds": 1.0}}}]},
    }
    ops = sorted(costed_operations(src))
    eq(ops, ["cnc_routing", "handling", "wet_spray"], "only what the sheet charged")
    eq(costed_finish_label(src), "Wet-spray painted", "finish named from what was charged")
    eq(part_numbers_with_operation(src, "powder_coating", "p.coat"), [],
       "a zero-valued powder row means no part is charged powder")
    eq(operations_for_part(src, "11225-01-J01"), ["handling", "cnc_routing", "wet_spray"],
       "per-part view agrees with the job view")


def test_a_department_resolves_to_one_operation_not_every_alias():
    """Calculated rows carry cost but not identity. Falling back to inverting the department
    expanded every synonym, so the quote listed 'Assembly' and 'Assemble', 'Fold' and
    'Folding', 'Weld' and 'Welding'."""
    from costed_facts import costed_operations
    src = {"final_estimate": {"labour_rows": [
        {"operation": "Assemble/pack (Metal)", "workbook_row": 96, "total_value_gbp": 0.36},
        {"operation": "Fold", "workbook_row": 97, "total_value_gbp": 1.55},
        {"operation": "Weld (CO2)", "workbook_row": 98, "total_value_gbp": 1.56},
    ]}}
    ops = sorted(costed_operations(src))
    eq(len(ops), 3, f"one operation per department, got {ops}")
    for pair in (("assemble", "assembly"), ("fold", "folding"), ("weld", "welding")):
        ok(not all(p in ops for p in pair), f"{pair} must not both appear")
    # A legacy row that wrote the DEPARTMENT into engine_operation must still resolve.
    legacy = {"workbook_labour": {"rows": [
        {"wb_operation": "Spray / Wet Paint", "engine_operation": "Spray / Wet Paint",
         "qty_per_unit": 12, "part_numbers": ["J01"]}]}}
    from costed_facts import costed_finish_label
    eq(costed_finish_label(legacy), "Wet-spray painted",
       "a legacy row shape must still resolve, not read as an unknown operation")


# ── SolidWorks native ────────────────────────────────────────────────────────────────
def test_native_bom_and_geometry_rules():
    """Job 12120. GA double-count, the folded-box-as-flat guard, and flag-don't-zero."""
    from source_connectors.solidworks import (normalize_native_extract,
                                              apply_native_to_pre_estimate)
    recs = [
        {"title": "12120-01-GA", "doctype": 2, "bom": [
            {"part_number": "12120-01-103", "qty": 1.0},
            {"part_number": "12120-01-01M", "qty": 2.0},
            {"part_number": "12120-01-05M", "qty": 4.0},
            {"part_number": "M4 Male Grip Knob", "qty": 8.0}]},
        {"title": "12120-01-103", "doctype": 2, "bom": [{"part_number": "12120-01-01M", "qty": 1.0}]},
        {"title": "12120-01-01M", "doctype": 1, "route_signals": {
            "material": "Mild Steel [CR4]", "is_sheet_metal": True, "bend_count": 4,
            "flat_length_mm": 126.39, "flat_width_mm": 82.2, "thickness_mm": 1.5,
            "bbox_mm": [79.0, 64.5, 21.5]}},
        # cut list returning the FOLDED bounding box on a bent part is not a flat pattern
        {"title": "12120-01-09M", "doctype": 1, "route_signals": {
            "material": "Mild Steel [CR4]", "is_sheet_metal": True, "bend_count": 3,
            "flat_length_mm": 100.0, "flat_width_mm": 50.0, "bbox_mm": [100.0, 50.0, 30.0]}},
        # material but no geometry at all — must be flagged, never costed at zero
        {"title": "12120-01-05M", "doctype": 1, "route_signals": {"material": "Mild Steel [CR4]"}},
        {"title": "M4 Male Grip Knob", "doctype": 1, "route_signals": {"likely_bought_in": True}},
        # cut-outs were extracted from the cut list and then discarded before costing
        {"title": "12120-01-06M", "doctype": 1, "route_signals": {
            "material": "Mild Steel [CR4]", "is_sheet_metal": True, "bend_count": 2,
            "flat_length_mm": 96.49, "flat_width_mm": 39.09, "thickness_mm": 1.2,
            "cut_out_count": 3, "bbox_mm": [30.9, 39.09, 35.0]}},
    ]
    job = normalize_native_extract(recs)
    eq({r.part_number: r.quantity for r in job.bom}["12120-01-01M"], 2.0,
       "BOM quantity read from `qty` (reading only `quantity` pinned every row to 1)")
    eq(job.part_signals["12120-01-09M"].flat_length_mm, None,
       "a folded part whose 'flat' equals its bounding box is the FOLDED box, not a blank")
    eq(job.part_signals["12120-01-01M"].flat_length_mm, 126.39,
       "a genuine developed blank, larger than the envelope, survives")

    parts = [{"part_number": p} for p in
             ("12120-01-GA", "12120-01-103", "12120-01-01M", "12120-01-05M",
              "M4 Male Grip Knob", "12120-01-06M")]
    # 06M already carries a weaker pierce count — the sort of figure a PDF pass leaves
    # behind. The model says nine; the drawing text guessed two.
    for _p in parts:
        if _p["part_number"] == "12120-01-06M":
            _p["manufacturing_features"] = {"pierce_count": 2}
            _p["geometry_rollup"] = {"estimated_pierce_count": 2}
    out = apply_native_to_pre_estimate(parts, job)
    by = {p["part_number"]: p for p in parts}
    # The GA is absent from its own BOM; without indexing the assemblies it is costed as a leaf.
    ok(by["12120-01-GA"].get("is_assembly_parent"), "the top assembly is a parent, not a leaf")
    ok(by["12120-01-103"].get("is_assembly_parent"), "a sub-assembly is a parent too")
    eq(by["12120-01-01M"]["blank_length_mm"], 126.39, "native flat applied where no DXF")
    eq(by["12120-01-01M"]["quantity"], 2, "quantity from the full-depth assembly BOM")
    eq(by["12120-01-01M"]["quantity_source"], "solidworks_api",
       "and its SOURCE recorded, so a later pass cannot silently overwrite it")
    ok(by["M4 Male Grip Knob"].get("is_bought_in"), "an imported body with no fab features is bought in")
    ok(by["12120-01-05M"].get("native_material_without_geometry"),
       "material with no geometry is FLAGGED — a GBP 0 line must not read as free")
    ok(out["assembly_parent"] >= 2 and out["bought_in"] >= 1, "counts reported")
    # Cut-outs reach costing: each internal profile needs its own pierce, as does the outer.
    _mf = by["12120-01-06M"].get("manufacturing_features") or {}
    eq(_mf.get("cut_out_count"), 3, "cut-out count must reach the engine, not be discarded")
    eq(_mf.get("pierce_count"), 4, "3 cut-outs plus the outer profile = 4 pierces")
    # The model is rank 90 and must SUPERSEDE the weaker count, not merely fill a gap: any
    # earlier positive number, however weak, otherwise locks the stronger evidence out.
    eq(_mf.get("pierce_count_source"), "solidworks_api", "and the source is recorded")
    # This is the field the laser reads. Asserting only on manufacturing_features tests the
    # staging value and says nothing about what the part was actually charged.
    eq((by["12120-01-06M"].get("geometry_rollup") or {}).get("estimated_pierce_count"), 4,
       "the pierce count COSTING reads must carry the native figure, not the PDF's 2")
    ok(any("pierce_count" in str(f) for f in (by["12120-01-06M"].get("review_flags") or [])),
       "replacing a weaker figure is flagged, not done silently")


# ── DXF ──────────────────────────────────────────────────────────────────────────────
def test_dxf_blocks_are_exploded():
    """A SolidWorks flat-pattern export wraps the profile in a block. Iterating model space
    alone found one INSERT and no geometry, so the blank came back 0 and the part fell
    through to the drawing's dimension TEXT — 4 of 7 parts on 12120."""
    try:
        import ezdxf
    except ImportError:
        print("      (skipped: ezdxf not installed)")
        return
    import tempfile
    from pathlib import Path
    from dxf_reader import extract_flat_pattern_data

    def build(path, L, W, blocked, holes=0):
        d = ezdxf.new(); d.header["$INSUNITS"] = 4
        msp = d.modelspace()
        pts = [(0, 0, L, 0), (L, 0, L, W), (L, W, 0, W), (0, W, 0, 0)]
        target = d.blocks.new(name="FLAT") if blocked else msp
        for a in pts:
            target.add_line(a[:2], a[2:], dxfattribs={"layer": "0" if blocked else "SLD-0"})
        for i in range(holes):
            target.add_circle((10 + i * 10, W / 2), 2.5, dxfattribs={"layer": "0" if blocked else "SLD-0"})
        if blocked:
            msp.add_blockref("FLAT", (0, 0), dxfattribs={"layer": "SLD-0"})
        d.saveas(path)
        return Path(path)

    tmp = tempfile.mkdtemp()
    for pn, L, W in (("01M", 126.39, 82.2), ("08M", 79.0, 37.79)):
        r = extract_flat_pattern_data(build(os.path.join(tmp, f"{pn}.dxf"), L, W, True))
        eq((r.get("blank_length_mm"), r.get("blank_width_mm")), (max(L, W), min(L, W)),
           f"{pn}: a blocked profile must measure the same as a loose one")
    # Holes inside a block were invisible: 06M reported zero, under-pricing the laser.
    r = extract_flat_pattern_data(build(os.path.join(tmp, "holes.dxf"), 96.49, 39.09, True, holes=8))
    eq(int(r.get("estimated_hole_count") or r.get("hole_count") or 0), 8,
       f"circles inside a block must be counted (got {r.get('estimated_hole_count')})")
    # Every hole is a pierce, and so is the outer profile. The non-recursive parser never
    # entered the block, reported zero, and the laser's pierce time vanished with it.
    eq(r.get("estimated_pierce_count"), 9, "8 holes plus the outer profile = 9 pierces")


def test_annotation_circles_are_not_holes():
    """hole_diams filtered sub-1mm circles and annotation layers; hole_count then counted
    every circle in the file regardless. The two described different sets, and a symbol
    layer inflated the laser. One filtered list must feed both."""
    try:
        import ezdxf
    except ImportError:
        print("      (skipped: ezdxf not installed)")
        return
    import tempfile
    from pathlib import Path
    from dxf_reader import extract_flat_pattern_data

    d = ezdxf.new(); d.header["$INSUNITS"] = 4
    msp = d.modelspace()
    L, W = 120.0, 60.0
    for a in [(0, 0, L, 0), (L, 0, L, W), (L, W, 0, W), (0, W, 0, 0)]:
        msp.add_line(a[:2], a[2:], dxfattribs={"layer": "SLD-0"})
    for i in range(3):                                    # three real holes
        msp.add_circle((20 + i * 20, W / 2), 3.0, dxfattribs={"layer": "SLD-0"})
    for i in range(5):                                    # weld/finish symbols — not holes
        msp.add_circle((10 + i * 5, 5), 2.0, dxfattribs={"layer": "SYMBOLS(BENCHMARK)"})
    msp.add_circle((60, 50), 0.2, dxfattribs={"layer": "SLD-0"})   # centre mark, not a hole
    path = os.path.join(tempfile.mkdtemp(), "annot.dxf")
    d.saveas(path)

    r = extract_flat_pattern_data(Path(path))
    eq(int(r.get("estimated_hole_count") or r.get("hole_count") or 0), 3,
       "only circles that survive the hole filter may be counted as holes")
    eq(r.get("estimated_pierce_count"), 4, "3 holes plus the outer profile = 4 pierces")
    ok(all(dia >= 1.0 for dia in (r.get("hole_diameters_mm") or [])),
       "hole diameters and hole count must describe the same set of circles")


def test_annotation_layers_are_skipped_by_effective_layer():
    """An entity drawn on layer '0' INSIDE a block sits on the layer its INSERT sits on —
    that is how these exports are written. Reading the entity's own layer returns '0' for
    every one of them, so circles inside a SYMBOLS(BENCHMARK) block passed the skip filter
    untouched and were priced as holes. The outline walk always resolved this; the feature
    walk did not."""
    try:
        import ezdxf
    except ImportError:
        print("      (skipped: ezdxf not installed)")
        return
    import tempfile
    from pathlib import Path
    from dxf_reader import extract_flat_pattern_data

    d = ezdxf.new(); d.header["$INSUNITS"] = 4
    msp = d.modelspace()
    L, W = 120.0, 60.0
    for a in [(0, 0, L, 0), (L, 0, L, W), (L, W, 0, W), (0, W, 0, 0)]:
        msp.add_line(a[:2], a[2:], dxfattribs={"layer": "SLD-0"})
    msp.add_circle((30, 30), 3.0, dxfattribs={"layer": "SLD-0"})       # one real hole
    sym = d.blocks.new(name="WELDSYM")                                  # symbols on layer 0 …
    for i in range(6):
        sym.add_circle((i * 4, 0), 2.0, dxfattribs={"layer": "0"})
    # … placed on an annotation layer. Their effective layer is the INSERT's.
    msp.add_blockref("WELDSYM", (10, 5), dxfattribs={"layer": "SYMBOLS(BENCHMARK)"})
    path = os.path.join(tempfile.mkdtemp(), "inherit.dxf")
    d.saveas(path)

    r = extract_flat_pattern_data(Path(path))
    eq(int(r.get("estimated_hole_count") or r.get("hole_count") or 0), 1,
       "a circle inheriting an annotation layer from its block is not a hole")
    eq(r.get("estimated_pierce_count"), 2, "1 hole plus the outer profile = 2 pierces")


def test_non_circular_cut_outs_are_pierced():
    """A pierce is a closed CONTOUR, not a round hole. Counting circles + 1 prices a slot,
    a rectangular aperture and a D-cut at nothing — and every one of them costs the laser a
    pierce and its own cutting time. Cut-outs arrive in three shapes and all three count: a
    circle, a closed polyline, and a loop assembled from separate lines and arcs."""
    try:
        import ezdxf
    except ImportError:
        print("      (skipped: ezdxf not installed)")
        return
    import tempfile
    from pathlib import Path
    from dxf_reader import extract_flat_pattern_data

    d = ezdxf.new(); d.header["$INSUNITS"] = 4
    msp = d.modelspace()
    L, W = 200.0, 100.0
    for a in [(0, 0, L, 0), (L, 0, L, W), (L, W, 0, W), (0, W, 0, 0)]:
        msp.add_line(a[:2], a[2:], dxfattribs={"layer": "SLD-0"})
    msp.add_circle((30, 50), 5.0, dxfattribs={"layer": "SLD-0"})            # 1: round hole
    msp.add_lwpolyline([(60, 40), (90, 40), (90, 60), (60, 60)],            # 2: rectangular
                       close=True, dxfattribs={"layer": "SLD-0"})
    # 3: a slot with radiused ends — two lines and two arcs, no single closed entity.
    msp.add_line((120, 45), (150, 45), dxfattribs={"layer": "SLD-0"})
    msp.add_line((150, 55), (120, 55), dxfattribs={"layer": "SLD-0"})
    msp.add_arc((150, 50), 5.0, -90, 90, dxfattribs={"layer": "SLD-0"})
    msp.add_arc((120, 50), 5.0, 90, 270, dxfattribs={"layer": "SLD-0"})
    # A drawing frame and title block are closed rectangles too. Counting every closed
    # contour in the file would charge the laser a pierce for the border.
    for box in ([(-50, -50), (260, -50), (260, 160), (-50, 160)],
                [(180, -45), (255, -45), (255, -10), (180, -10)]):
        msp.add_lwpolyline(box, close=True, dxfattribs={"layer": "TITLE_FRAME"})
    path = os.path.join(tempfile.mkdtemp(), "cutouts.dxf")
    d.saveas(path)

    r = extract_flat_pattern_data(Path(path))
    eq(r.get("estimated_pierce_count"), 4,
       "3 cut-outs (round, rectangular, slot) plus the outer profile = 4 pierces")
    eq(r.get("closed_contour_count"), 4, "each closed contour counted once")
    ok(not r.get("pierce_count_incomplete"),
       "every segment chained into a loop — nothing left open to warn about")
    # The outer profile is one of the contours, not an extra on top of them.
    eq(int(r.get("estimated_hole_count") or r.get("hole_count") or 0), 1,
       "only the circle is a HOLE; the slot and aperture are cut-outs, not drilled")


# ── workbook join — job 12120 ────────────────────────────────────────────────────────
def test_labour_rows_join_to_the_parts_that_produced_them():
    """The calculated cost of a labour row must land on the parts that row was grouped
    from. wb_populate used to publish the route record BEFORE the write loop and then
    back-fill sheet rows by matching on DEPARTMENT NAME in insertion order. One department,
    one group — fine. Two gauges of Fold, and the 1.2mm cost lands on the 1.5mm parts (and
    the Laser rows swap with them). 12120 has exactly that shape.

    The invariant: a row's workbook_row, parts and gauge all come off the same group."""
    from wb_populate import build_workbook_labour
    from costed_facts import operations_for_part, part_numbers_with_operation

    # Insertion order deliberately unlike sheet order, and unlike sorted-key order: any
    # implementation that pairs rows to sheet rows positionally will mis-associate.
    groups = {
        "Fold|1.5":  {"wb_op": "Fold",  "engine_ops": ["folding"], "thickness": 1.5,
                      "material": "Mild Steel", "qty": 2, "parts": ["12120-01-04M"],
                      "workbook_row": 44},
        "Laser|1.2": {"wb_op": "Laser", "engine_ops": ["laser_cutting"], "thickness": 1.2,
                      "material": "Mild Steel", "qty": 3, "parts": ["12120-01-01M", "12120-01-06M"],
                      "workbook_row": 41},
        "Fold|1.2":  {"wb_op": "Fold",  "engine_ops": ["folding"], "thickness": 1.2,
                      "material": "Mild Steel", "qty": 3, "parts": ["12120-01-01M", "12120-01-06M"],
                      "workbook_row": 43},
        "Laser|1.5": {"wb_op": "Laser", "engine_ops": ["laser_cutting"], "thickness": 1.5,
                      "material": "Mild Steel", "qty": 2, "parts": ["12120-01-04M"],
                      "workbook_row": 42},
        # Never written to the sheet — must not appear in the route record at all.
        "Weld|-":    {"wb_op": "Weld",  "engine_ops": ["welding"], "thickness": None,
                      "material": "Mild Steel", "qty": 1, "parts": ["12120-01-09M"]},
    }
    rec = build_workbook_labour(groups, ["12120-01-99B"])
    rows = rec["rows"]

    eq([r["workbook_row"] for r in rows], [41, 42, 43, 44], "rows must be in sheet order")
    ok(all(r["wb_operation"] != "Weld" for r in rows),
       "a group with no sheet row was never priced and must not be published as route")
    by_row = {r["workbook_row"]: r for r in rows}
    for g in groups.values():
        if not g.get("workbook_row"):
            continue
        r = by_row[g["workbook_row"]]
        eq((r["wb_operation"], r["thickness_mm"], r["part_numbers"]),
           (g["wb_op"], g["thickness"], g["parts"]),
           f"row {g['workbook_row']}: identity must come off the group that wrote it")
    # The bug was invisible on identity alone until cost was joined on. £30 was booked for
    # the 1.5mm fold; it must not be reported against the 1.2mm parts.
    summary = {
        "workbook_labour": rec,
        "final_estimate": {"schema": "final_estimate.v2", "labour_rows": [
            {"workbook_row": 41, "operation": "Laser", "batch_hours": 0.4, "total_value_gbp": 18.0},
            {"workbook_row": 42, "operation": "Laser", "batch_hours": 0.2, "total_value_gbp": 9.0},
            {"workbook_row": 43, "operation": "Fold",  "batch_hours": 0.5, "total_value_gbp": 22.0},
            {"workbook_row": 44, "operation": "Fold",  "batch_hours": 0.7, "total_value_gbp": 30.0},
        ]},
    }
    from costed_facts import _workbook_rows
    priced = {r["total_value_gbp"]: r["part_numbers"] for r in (_workbook_rows(summary) or [])}
    eq(priced.get(30.0), ["12120-01-04M"], "the 1.5mm fold cost belongs to the 1.5mm part")
    eq(priced.get(22.0), ["12120-01-01M", "12120-01-06M"], "the 1.2mm fold cost belongs to the 1.2mm parts")
    ok("folding" in operations_for_part(summary, "12120-01-04M"),
       "a part in a priced fold group must show folding")
    eq(sorted(part_numbers_with_operation(summary, "folding")),
       ["12120-01-01M", "12120-01-04M", "12120-01-06M"], "every folded part, once")


# ── read-back — job 12120 ────────────────────────────────────────────────────────────
def test_every_material_block_is_read_back_not_just_the_bom():
    """Fabricated material lives in three blocks below the BOM, and each names its value
    column differently — "Cost" on tube, "Cost Per Part" on steel and other-sheet. Reading
    all of them through the BOM's "Total Value" returned rows carrying no value; asking for
    a block named "wire" when CELL_MAP defines "tube" skipped that block entirely, in
    silence. On 12120 the read-back summed to GBP 9.64 of the sheet's own GBP 10.07 total,
    the missing 43p being exactly the fabricated material. A snapshot that will not
    reconcile to its own total is not something an ERP export can be built on."""
    from wb_populate import CELL_MAP
    from wep_readback_from_xlsx import read_final_rows

    class Cell:
        def __init__(self, v): self.Value = v

    class FakeSheet:
        """Only what _read_block touches: .Cells(row, col).Value."""
        def __init__(self, grid): self.grid = grid
        def Cells(self, r, c): return Cell(self.grid.get((r, c)))

    grid = {}

    def block(name, headers, rows):
        b = CELL_MAP[name]
        hr = b["first_row"] - 1
        for c, text in headers.items():
            grid[(hr, c)] = text
        for i, vals in enumerate(rows):
            for c, v in vals.items():
                grid[(b["first_row"] + i, c)] = v

    block("bom",
          {3: "Bill Of Materials", 10: "Price", 11: "Qty Per Unit", 12: "Scrap", 13: "Total Value"},
          [{3: "M6 Pem Stud", 10: 0.12, 11: 4, 13: 9.64}])
    # Value column here is "Cost", not "Total Value".
    block("tube", {3: "Part Description", 5: "Qty Per Unit", 6: "Gauge",
                   7: "Length", 8: "Price Per M", 11: "Cost"},
          [{3: "25x25x1.5 SHS Leg", 5: 2, 6: 1.5, 7: 300.0, 8: 4.10, 11: 0.18}])
    # And "Cost Per Part" here.
    block("steel", {3: "Part Description", 5: "Qty Per Unit", 6: "Part Length",
                    7: "Part Width", 8: "Gauge", 13: "Cost Per Part"},
          [{3: "12120-01-01M", 5: 1, 6: 126.39, 7: 82.2, 8: 1.2, 13: 0.15}])
    block("other_sheet", {3: "Part Description", 4: "Qty Per Unit", 5: "Part Length",
                          6: "Part Width", 7: "Thickness", 13: "Cost Per Part"},
          [{3: "5mm Acrylic Window", 4: 1, 5: 200.0, 6: 100.0, 7: 5.0, 13: 0.10}])

    rows = read_final_rows(FakeSheet(grid), 24).get("material_rows") or []
    got = {r.get("block") for r in rows}
    for name in ("bom", "tube", "steel", "other_sheet"):
        ok(name in got, f"the {name} block must be read back — CELL_MAP defines it")
    for r in rows:
        ok(r.get("total_value_gbp") is not None,
           f"{r.get('block')} row carries no value: its value column was not mapped")
    total = sum(float(r.get("total_value_gbp") or 0) for r in rows)
    eq(round(total, 2), 10.07, "material rows must reconcile to the sheet's own total")


def test_the_resolver_sees_nested_fields_and_explicit_zeros():
    """Two things the connector-local writes kept getting wrong, and the reason they stayed
    brittle: the fields that drive cost do not live at the top of a part record, and zero is
    an answer rather than a gap."""
    from source_precedence import apply_field, source_of, value_of, MISSING

    # Nested. A resolver that only sees top-level keys cannot arbitrate the pierce count.
    p = {}
    ok(apply_field(p, "geometry_rollup.estimated_pierce_count", 9, "solidworks_api"),
       "a nested field must be writable through the resolver")
    eq(p["geometry_rollup"]["estimated_pierce_count"], 9, "written where costing reads it")
    eq(source_of(p, "geometry_rollup.estimated_pierce_count"), "solidworks_api",
       "and its source recorded alongside it, not at the top of the record")
    ok(not apply_field(p, "geometry_rollup.estimated_pierce_count", 2, "llm_extract"),
       "a weaker source must not replace it")
    eq(p["geometry_rollup"]["estimated_pierce_count"], 9, "the stronger value survives")
    ok(any("estimated_pierce_count" in str(f) for f in p.get("review_flags") or []),
       "and the disagreement is recorded, not discarded")

    # Explicit zero. `if cut_out_count:` read the model saying "none" as nobody having
    # looked, which is how a weaker count survived against the strongest source there is.
    z = {}
    ok(apply_field(z, "manufacturing_features.cut_out_count", 0, "solidworks_api"),
       "zero is a value and must be written")
    eq(value_of(z, "manufacturing_features.cut_out_count"), 0, "and readable as zero")
    ok(not apply_field(z, "manufacturing_features.cut_out_count", 4, "llm_extract"),
       "a recorded zero is defended like any other value — it is not an opening")
    eq(value_of({}, "manufacturing_features.cut_out_count"), MISSING,
       "absent is MISSING, which is not the same fact as zero")
    # Empty containers and None really are gaps.
    e = {"normalized_material": ""}
    ok(apply_field(e, "normalized_material", "MILD_STEEL", "inference"),
       "an empty string is a gap anything may fill")


def test_arbitrated_fields_are_not_written_directly():
    """The resolver only protects a datum if every writer goes through it. One direct
    assignment anywhere reintroduces last-writer-wins for that field, silently.

    So the rule is checked on the SOURCE, not just the behaviour: a fixture that exercises
    apply_field proves the resolver works, not that anyone calls it. Reverting a converted
    call site to a direct write passed every behavioural test in this file.

    PARSED, NOT PATTERN-MATCHED. The first version of this guard was a regex, and a regex
    reads text rather than code: it missed single-quoted keys, dict literals, keyword
    arguments, .update(), aliased records and any assignment split over two lines. It was
    already blind to empty_part_record(quantity=...), which is exactly how BOM quantities
    were reaching parts unattributed. The AST sees the shapes themselves.

    This guards the modules already converted. It is not the whole pipeline yet — the
    remaining writers are listed at the end of this file's docstring and each one is a place
    a stronger source can still be overwritten in silence."""
    import ast
    from pathlib import Path

    root = Path(__file__).resolve().parents[1] / "src"
    # STAGE 5 COMPLETE: every module that writes an arbitrated field. Adding a module here
    # is how a conversion gets locked in, and the guard is what stops the next writer someone
    # adds from silently reintroducing last-writer-wins on a job nobody has seen yet.
    RESOLVER_CLEAN = ["bom_tree.py", "part_index.py", "learning_engine.py",
                      "source_connectors/solidworks.py",
                      "document_builder.py", "json_normaliser.py", "file_scan.py",
                      "drawing_job_merge.py", "dxf_reader.py.py", "estimator.py"]
    ARBITRATED = {"quantity", "normalized_material", "pierce_count",
                  "estimated_pierce_count", "normalized_thickness_mm"}
    # A single record that holds arbitrated evidence. NOT `parts`, which is a collection:
    # `parts[pn] = <record>` inserts a whole record rather than writing a field, while
    # `parts[pn]["quantity"] = x` is still caught by the literal-key rule below.
    RECORDS = {"part", "_part", "p", "_p", "nr", "target", "pe"}
    # Constructors that build a part record: passing an arbitrated field as a keyword sets it
    # with no source, which is a direct write wearing a different hat.
    CONSTRUCTORS = {"empty_part_record", "_empty_part_record"}

    class Guard(ast.NodeVisitor):
        def __init__(self, rel, allowed):
            self.rel, self.allowed, self.hits = rel, allowed, []

        def _report(self, node, what):
            if node.lineno not in self.allowed:
                self.hits.append(f"{self.rel}:{node.lineno}  {what}")

        def _is_record(self, node):
            if isinstance(node, ast.Name):
                return node.id in RECORDS
            if isinstance(node, ast.Subscript):      # parts[pn]["quantity"] = ...
                return self._is_record(node.value)
            if isinstance(node, ast.Attribute):
                return node.attr in RECORDS
            return False

        def visit_Assign(self, node):
            for tgt in node.targets:
                if not isinstance(tgt, ast.Subscript):
                    continue
                key = tgt.slice
                # record["arbitrated"] = ...  (quote style is irrelevant to the parser)
                if isinstance(key, ast.Constant) and key.value in ARBITRATED:
                    self._report(node, f'{ast.unparse(tgt)} = ...')
                # record[<computed>] = ...  — names no field, so the field name cannot be
                # the test. This is how the override rules wrote material.
                elif not isinstance(key, ast.Constant) and self._is_record(tgt.value):
                    self._report(node, f'{ast.unparse(tgt)} = ...  (computed key)')
            self.generic_visit(node)

        def visit_Call(self, node):
            fn = node.func
            name = fn.attr if isinstance(fn, ast.Attribute) else getattr(fn, "id", "")
            # record.update({"quantity": ...}) / record.update(quantity=...)
            if name == "update" and isinstance(fn, ast.Attribute) and self._is_record(fn.value):
                for kw in node.keywords:
                    if kw.arg in ARBITRATED:
                        self._report(node, f"update({kw.arg}=...)")
                for a in node.args:
                    if isinstance(a, ast.Dict):
                        for k in a.keys:
                            if isinstance(k, ast.Constant) and k.value in ARBITRATED:
                                self._report(node, f"update({{{k.value!r}: ...}})")
            # empty_part_record(quantity=...) — born with a value and no source.
            if name in CONSTRUCTORS:
                for kw in node.keywords:
                    if kw.arg in ARBITRATED and not (
                            isinstance(kw.value, ast.Constant) and kw.value.value is None):
                        self._report(node, f"{name}({kw.arg}=...) with no source")
            self.generic_visit(node)

    offenders = []
    for rel in RESOLVER_CLEAN:
        path = root / rel
        if not path.exists():
            continue
        text = path.read_text(encoding="utf-8")
        # A deliberate direct write is allowed when the author marked the line.
        allowed = {n for n, line in enumerate(text.splitlines(), 1)
                   if "precedence: direct-write ok" in line}
        g = Guard(rel, allowed)
        g.visit(ast.parse(text))
        offenders.extend(g.hits)
    eq(sorted(offenders), [],
       "these writes bypass the resolver — a stronger source can be overwritten in silence")


def test_agreeing_evidence_upgrades_provenance():
    """A source that CONFIRMS the value already present has told us something real: the datum
    now rests on stronger evidence than it did. Callers used to skip the resolver when the
    numbers matched — "nothing to change" — so the value kept the weaker source's name and a
    later medium-ranked pass could still displace a figure the model had confirmed."""
    from source_precedence import apply_field, source_of
    p = {"quantity": 2, "quantity_source": "bom_tree"}
    changed = apply_field(p, "quantity", 2, "solidworks_api")
    eq(source_of(p, "quantity"), "solidworks_api",
       "agreement from a stronger source must upgrade the recorded source")
    ok(not changed,
       "but the VALUE did not change, and an audit message must not claim it did")
    # And that upgrade is what protects it.
    ok(not apply_field(p, "quantity", 7, "drawing_deterministic"),
       "once the model has confirmed it, a weaker source cannot displace it")
    eq(p["quantity"], 2, "the confirmed value stands")
    # A weaker source agreeing must not DOWNGRADE the provenance.
    q = {"quantity": 2, "quantity_source": "solidworks_api"}
    apply_field(q, "quantity", 2, "llm_extract")
    eq(source_of(q, "quantity"), "solidworks_api", "agreement never weakens provenance")


def test_equal_rank_conflicts_are_never_resolved_by_running_order():
    """Two title-block readings of the same rank disagreeing is not refinement, it is a
    conflict. Letting the later one win made the answer depend on the order pages happened to
    be read in — silently, because the write succeeded so nothing was flagged."""
    from source_precedence import apply_field
    p = {}
    apply_field(p, "normalized_material", "MILD_STEEL", "drawing_deterministic")
    ok(not apply_field(p, "normalized_material", "ALUMINIUM", "drawing_deterministic"),
       "an equal-ranked disagreement must not overwrite")
    eq(p["normalized_material"], "MILD_STEEL", "the first observation is kept")
    ok(any("equal standing" in str(f) for f in p.get("review_flags") or []),
       "and the conflict is flagged for a person, not silently resolved")
    # Confidence is a real reason to prefer the newcomer; page order is not.
    q = {}
    apply_field(q, "normalized_material", "MILD_STEEL", "llm_extract", confidence=0.4)
    ok(apply_field(q, "normalized_material", "ALUMINIUM", "llm_extract", confidence=0.9),
       "a strictly higher confidence at equal rank is a reason, not an accident")
    eq(q["normalized_material"], "ALUMINIUM", "and it wins")


def test_the_knowledge_base_is_not_gated_before_arbitration():
    """A local reliability test refused to offer a knowledge-base correction whenever the
    part already carried anything ranked at or above an override rule — so rank-100 KB data,
    the one signal carrying knowledge the drawing does not, was never offered against a
    rank-50, 70 or 90 value. That inverts the ranking table it exists to serve."""
    from source_precedence import apply_field
    for existing_source in ("override_rule:x", "drawing_deterministic", "solidworks_api"):
        p = {"normalized_material": "MILD_STEEL", "material_source": existing_source}
        ok(apply_field(p, "normalized_material", "ALUMINIUM", "knowledge_base (95%)",
                       confidence=0.95),
           f"a person's correction must outrank {existing_source}")
        eq(p["normalized_material"], "ALUMINIUM", "and be applied")


def test_bom_quantities_are_attributed_when_the_record_is_born():
    """Records were constructed with quantity=<BOM value> and no source, and the apply path
    further on only ran for quantities of None or 1 — so every quantity ABOVE one, which is
    most of them, was never attributed and never protected."""
    from document_builder import _empty_part_record
    from source_precedence import apply_field, source_of
    r = _empty_part_record("X-01", description="d", quantity=None)
    eq(r["quantity"], None, "constructed with no quantity, so it is a gap the resolver fills")
    apply_field(r, "quantity", 4, "bom_tree")
    eq((r["quantity"], source_of(r, "quantity")), (4, "bom_tree"), "attributed at birth")
    ok(not apply_field(r, "quantity", 9, "llm_extract"),
       "and defended from a weaker source thereafter — including when it is above one")


def test_the_rules_suite_touches_no_live_service():
    """The suite opened a connection to the production SDILive server merely by importing
    learning_engine for its pure functions. A suite that needs the production database is not
    isolated, cannot run in CI, and makes every fixture depend on a machine being up."""
    import learning_engine
    eq(os.environ.get("SDI_OFFLINE"), "1", "the suite declares itself offline before importing")
    ok(not learning_engine._DB_AVAILABLE,
       "and the module honours it rather than dialling out at import time")


def test_solidworks_submits_agreement_at_the_call_site():
    """Behavioural, on the real call path. Asserting on apply_field proves the resolver
    upgrades provenance on agreement; it says nothing about whether the connector SUBMITS
    when the values already match. Restoring the `if values differ` guard passed every
    resolver fixture in this file."""
    from source_connectors.solidworks import (normalize_native_extract,
                                              apply_native_to_pre_estimate)
    recs = [
        {"title": "ASM-01", "doctype": 2, "bom": [{"part_number": "AAA-01M", "qty": 2.0}]},
        {"title": "AAA-01M", "doctype": 1, "route_signals": {
            "material": "Mild Steel [CR4]", "is_sheet_metal": True, "bend_count": 1,
            "flat_length_mm": 100.0, "flat_width_mm": 50.0, "thickness_mm": 1.5,
            "bbox_mm": [60.0, 50.0, 20.0]}},
    ]
    job = normalize_native_extract(recs)
    # The engine already has the RIGHT answers, from weaker sources.
    parts = [{"part_number": "AAA-01M",
              "quantity": 2, "quantity_source": "bom_tree",
              "normalized_material": "MILD_STEEL", "material_source": "drawing_deterministic",
              "normalized_thickness_mm": 1.5, "thickness_source": "drawing_deterministic"}]
    apply_native_to_pre_estimate(parts, job)
    p = parts[0]
    eq(p.get("quantity_source"), "solidworks_api",
       "the model confirming a quantity must upgrade its provenance, not skip the resolver")
    eq(p.get("material_source"), "solidworks_api", "same for a confirmed material")
    eq(p.get("thickness_source"), "solidworks_api", "and a confirmed thickness")
    eq((p["quantity"], p["normalized_material"], p["normalized_thickness_mm"]),
       (2, "MILD_STEEL", 1.5), "and none of the values changed")
    # That upgrade is the whole point: the values are now defended.
    from source_precedence import apply_field
    ok(not apply_field(p, "quantity", 9, "drawing_deterministic"),
       "a confirmed quantity is no longer displaceable by a medium-ranked pass")


def test_the_knowledge_base_reaches_the_resolver_at_the_call_site():
    """Behavioural, on the real call path. The local reliability gate sat BEFORE the
    resolver, so a rank-100 correction was never offered at all — and a fixture that calls
    apply_field directly cannot see that, because it starts after the gate."""
    import learning_engine as LE

    class _StubDB:
        def lookup_part(self, pn):
            return {"material": "ALUMINIUM", "thickness_mm": 3.0, "confidence": 0.95}
        def get_active_overrides(self):
            return []
        def fire_override(self, _id):
            pass

    _orig_db, _orig_avail = LE.db, LE._DB_AVAILABLE
    try:
        LE.db, LE._DB_AVAILABLE = _StubDB(), True
        eng = LE.LearningEngine.__new__(LE.LearningEngine)   # no DB work in __init__
        eng._overrides_cache = []
        # A part already carrying a strong, WRONG answer. The old gate refused to offer the
        # knowledge base against anything ranked at or above an override rule.
        summary = {"manufacturing_writeup": {"parts": [
            {"part_number": "AAA-01M", "normalized_material": "MILD_STEEL",
             "material_source": "solidworks_api"}]}}
        eng.pre_scan(summary)
        part = summary["manufacturing_writeup"]["parts"][0]
        eq(part["normalized_material"], "ALUMINIUM",
           "a person's correction outranks the model and must reach the part")
        ok("knowledge_base" in str(part.get("material_source")),
           f"and be attributed to it (got {part.get('material_source')!r})")
    finally:
        LE.db, LE._DB_AVAILABLE = _orig_db, _orig_avail


def test_a_plate_is_not_folded():
    """12120's 04M is 60 x 34.04 x 1.5mm at 1.5mm gauge — a flat plate, bends 0 on the cut
    list — and it was routed into the Fold 1.5mm group off the drawing. A part one thickness
    thick has nowhere for a bend to be.

    Zero bends ALONE must not be enough to strip the fold: that is exactly the case
    formed_but_no_bend_features exists for, where a Base Flange hides real folds from the
    feature tree (12120's 02M, 06M and 08M all count 0 and genuinely fold). The solid has to
    corroborate it."""
    from source_connectors.solidworks import (normalize_native_extract,
                                              apply_native_to_pre_estimate)

    def _job_for(pn, bends, bbox, flat_l, flat_w, thk=1.5, formed=False):
        rs = {"material": "Mild Steel [CR4]", "is_sheet_metal": True, "bend_count": bends,
              "flat_length_mm": flat_l, "flat_width_mm": flat_w, "thickness_mm": thk,
              "bbox_mm": bbox}
        if formed:
            rs["formed_but_no_bend_features"] = True
        return normalize_native_extract([{"title": pn, "doctype": 1, "route_signals": rs}])

    # 04M: plate. The fold came off the drawing and must go.
    parts = [{"part_number": "04M", "textual_operations": ["laser_cutting", "folding"],
              "manufacturing_features": {"bend_count": 1},
              "risk_flags": ["fold_count_uncertain"]}]
    apply_native_to_pre_estimate(parts, _job_for("04M", 0, [60.0, 34.04, 1.5], 60.0, 34.04))
    ok("folding" not in (parts[0].get("textual_operations") or []),
       "a plate must not be folded")
    eq((parts[0].get("manufacturing_features") or {}).get("bend_count"), 0,
       "and its bend count must say so")
    ok(not [f for f in (parts[0].get("risk_flags") or []) if "fold" in str(f).lower()],
       "with the fold risk flag cleared, not left telling an estimator to check a fold")
    # A DURABLE marker, not just a removal. The connector runs early and a later pass
    # re-derives folding from the drawing text; without this the op grows back, which is
    # exactly what 12120's 04M did on the next run.
    ok(parts[0].get("native_flat_solid"),
       "the model's statement that this is a plate must persist for the costing-time gate")

    # 01M: genuinely folded — envelope stands far taller than the gauge. Untouched.
    parts = [{"part_number": "01M", "textual_operations": ["laser_cutting", "folding"],
              "manufacturing_features": {"bend_count": 6}}]
    apply_native_to_pre_estimate(parts, _job_for("01M", 6, [79.0, 64.5, 21.5], 126.39, 82.2))
    ok("folding" in (parts[0].get("textual_operations") or []),
       "a folded part keeps its fold")
    eq((parts[0].get("manufacturing_features") or {}).get("bend_count"), 6, "and its count")

    # 08M: cut list counts 0 bends but the solid is 17mm deep at 1.2mm gauge — it folds.
    # Stripping on a zero count alone would silently un-fold three of 12120's parts.
    parts = [{"part_number": "08M", "textual_operations": ["laser_cutting", "folding"],
              "manufacturing_features": {"bend_count": 1}}]
    apply_native_to_pre_estimate(parts, _job_for("08M", 0, [79.0, 23.0, 17.0], 79.0, 37.79,
                                                 thk=1.2))
    ok("folding" in (parts[0].get("textual_operations") or []),
       "a zero bend COUNT on a part whose solid is deeper than its gauge is not evidence "
       "of a plate — the fold must survive")


def test_a_plate_stays_unfolded_after_the_drawing_text_pass():
    """The connector strips the fold from a plate. Then document_builder re-derives folding
    from the drawing's own text — "fold or bend work indicated" — and 12120's 04M was back
    in the Fold 1.5mm group on the very next run. The op was removed once and grew back.

    Removing it once is not the same as it staying removed. The model's statement that the
    part is a plate has to survive every pass that can re-add the op, and be enforced again
    at costing time, which is the last point before an op becomes money."""
    from estimator import estimate_process_times
    plate = {"part_number": "04M", "normalized_material": "MILD_STEEL",
             "native_flat_solid": True,          # the model: 60 x 34.04 x 1.5 at 1.5mm gauge
             # …and a later pass has put the fold back from the drawing text.
             "textual_operations": ["laser_cutting", "folding"],
             "inferred_operations": ["folding"],
             "manufacturing_features": {"bend_count": 2},
             "risk_flags": ["fold_count_uncertain"],
             "normalized_thickness_mm": 1.5, "overall_length_mm": 60, "overall_width_mm": 34,
             "geometry_rollup": {"estimated_cut_length_mm": 224.0}}
    estimate_process_times(plate)
    for _f in ("textual_operations", "inferred_operations"):
        ok("folding" not in (plate.get(_f) or []),
           f"{_f} still folds a part the model says is a plate")
    eq((plate.get("manufacturing_features") or {}).get("bend_count"), 0, "and no bends")
    ok(not [f for f in (plate.get("risk_flags") or []) if "fold" in str(f).lower()],
       "no fold risk flag telling an estimator to check a fold that cannot exist")

    # A part the model has NOT called a plate keeps whatever the drawing says.
    folded = dict(plate, part_number="01M", native_flat_solid=False,
                  textual_operations=["laser_cutting", "folding"],
                  inferred_operations=["folding"],
                  manufacturing_features={"bend_count": 6}, risk_flags=[])
    estimate_process_times(folded)
    ok("folding" in (folded.get("textual_operations") or []),
       "a genuinely folded part is untouched — the gate is about plates, not about folding")


def test_every_native_disagreement_reaches_the_resolver():
    """Three branches decided the outcome themselves instead of submitting the observation.
    Same-family material ("keep the drawing value"), a differing thickness on a DXF-backed
    part, and a quantity counter that incremented whether or not the write was accepted.
    Each is a rank judgement made locally, and rank is not a connector's to decide."""
    from source_connectors.solidworks import (normalize_native_extract,
                                              apply_native_to_pre_estimate)
    recs = [
        {"title": "ASM-9", "doctype": 2, "bom": [{"part_number": "BBB-01M", "qty": 3.0}]},
        {"title": "BBB-01M", "doctype": 1, "route_signals": {
            "material": "AISI 304", "is_sheet_metal": True, "bend_count": 1,
            "flat_length_mm": 120.0, "flat_width_mm": 60.0, "thickness_mm": 2.0,
            "bbox_mm": [80.0, 60.0, 25.0]}},
    ]
    job = normalize_native_extract(recs)
    # Same FAMILY, different grade: the drawing says mild steel, the model says stainless.
    # (Different families in the engine's map, but both metals — the old branch kept the
    # drawing value for anything not crossing metal/non-metal.)
    parts = [{"part_number": "BBB-01M",
              "normalized_material": "MILD_STEEL", "material_source": "llm_extract",
              "normalized_thickness_mm": 3.0, "thickness_source": "llm_extract",
              "dxf_source_file": "BBB-01M_3mm.dxf", "dxf_measured_outline": True,
              "blank_length_mm": 120.0, "blank_width_mm": 60.0}]
    apply_native_to_pre_estimate(parts, job)
    p = parts[0]
    eq(p["normalized_material"], "STAINLESS_STEEL",
       "the model must beat a rank-40 reading even within the same broad family")
    eq(p["material_source"], "solidworks_api", "and be attributed")
    eq(p["normalized_thickness_mm"], 2.0,
       "a DXF-backed part must still receive the model's gauge — a DXF filename is a filename")
    eq(p["thickness_source"], "solidworks_api", "and be attributed")

    # A rank-100 correction must survive all three, and nothing may claim it changed.
    parts2 = [{"part_number": "BBB-01M",
               "quantity": 1, "quantity_source": "estimator_confirmed",
               "normalized_material": "ALUMINIUM", "material_source": "estimator_confirmed"}]
    counts = apply_native_to_pre_estimate(parts2, job)
    eq(parts2[0]["quantity"], 1, "an estimator's quantity stands against the model")
    eq(parts2[0]["normalized_material"], "ALUMINIUM", "and so does their material")
    eq(counts["qty"], 0, "a rejected write must not be counted as applied")
    eq(counts["material"], 0, "nor a rejected material")

    # Cut-outs, same rule. The flag claimed the previous pierce count had been replaced
    # whatever the resolver decided, so a rank-100 value could survive while the audit trail
    # said SolidWorks had overwritten it.
    recs3 = [{"title": "CCC-01M", "doctype": 1, "route_signals": {
        "material": "Mild Steel [CR4]", "is_sheet_metal": True, "bend_count": 1,
        "flat_length_mm": 100.0, "flat_width_mm": 50.0, "thickness_mm": 1.5,
        "cut_out_count": 8, "bbox_mm": [60.0, 50.0, 20.0]}}]
    job3 = normalize_native_extract(recs3)
    parts3 = [{"part_number": "CCC-01M",
               "manufacturing_features": {"pierce_count": 2,
                                          "pierce_count_source": "estimator_confirmed"},
               "geometry_rollup": {"estimated_pierce_count": 2,
                                   "estimated_pierce_count_source": "estimator_confirmed"}}]
    apply_native_to_pre_estimate(parts3, job3)
    eq((parts3[0]["geometry_rollup"] or {}).get("estimated_pierce_count"), 2,
       "an estimator's pierce count stands against the model")
    ok(not any("replaced by" in str(f) for f in (parts3[0].get("review_flags") or [])),
       "and nothing may claim SolidWorks replaced it")


def test_source_and_confidence_move_together():
    """Updating the source while leaving the weaker source's confidence behind produces a
    datum labelled with strong evidence and scored with weak — a record that reads as better
    than anything actually supplied."""
    from source_precedence import apply_field, source_of, confidence_of
    p = {}
    apply_field(p, "normalized_material", "MILD_STEEL", "llm_extract", confidence=0.4)
    # Stronger source agrees but offers no confidence: the stale 0.4 must not be inherited.
    apply_field(p, "normalized_material", "MILD_STEEL", "solidworks_api")
    eq(source_of(p, "normalized_material"), "solidworks_api", "provenance upgraded")
    eq(confidence_of(p, "normalized_material"), None,
       "and the weaker source's confidence cleared, not left attached to a stronger name")
    # Corroboration at equal rank raises confidence without changing the source.
    q = {}
    apply_field(q, "normalized_material", "MDF", "drawing_deterministic", confidence=0.5)
    apply_field(q, "normalized_material", "MDF", "title_block", confidence=0.8)
    eq(confidence_of(q, "normalized_material"), 0.8,
       "two independent readings agreeing is a real strengthening")
    eq(source_of(q, "normalized_material"), "drawing_deterministic",
       "but nothing was replaced, so the source is unchanged")


def test_automatic_com_execution_is_opt_in():
    """SAFETY, not caution. The analyser calls Dispatch("SldWorks.Application"), which
    ATTACHES to a SolidWorks already running on the machine. SolidWorks does not open a
    document twice, so OpenDoc6 on a file a designer has open returns THEIR document — and
    the analyser then closes every title it touched, taking unsaved work with it.

    Making acquisition automatic was right. Making it the default on any machine that runs
    an estimate was not. It is enabled only where SolidWorks belongs to this process."""
    import ast
    from pathlib import Path
    src = (Path(__file__).resolve().parents[1] / "src" / "file_scan.py").read_text(encoding="utf-8")
    tree = ast.parse(src)
    assigns = [n for n in ast.walk(tree)
               if isinstance(n, ast.Assign)
               and any(getattr(t, "id", "") == "_sw_run" for t in n.targets)]
    ok(assigns, "file_scan must decide whether to run the analyser explicitly")
    src_txt = ast.unparse(assigns[0].value)
    # Opt-IN reads "flag in {truthy}". Opt-OUT reads "flag not in {falsy}" and would run by
    # default on a designer's workstation.
    ok(" not in " not in src_txt,
       f"automatic COM execution must be opt-in, not opt-out: {src_txt}")
    ok("SDI_SW_RUN_ANALYSER" in src_txt, "and gated on the documented switch")


def test_the_analyser_never_closes_a_document_it_did_not_open():
    """The one irreversible thing this tool could do. close_all() closed every title it had
    touched; a document the designer already had open is not ours to close."""
    import ast
    from pathlib import Path
    src = (Path(__file__).resolve().parents[1] / "tools" / "solidworks"
           / "sw_native_analyse.py").read_text(encoding="utf-8")
    ok("_is_already_open" in src,
       "the analyser must ask what was open BEFORE it opens anything")
    ok("_borrowed_titles" in src,
       "and keep borrowed documents apart from the ones it opened")
    tree = ast.parse(src)
    fn = next((n for n in ast.walk(tree)
               if isinstance(n, ast.FunctionDef) and n.name == "close_all"), None)
    ok(fn is not None, "close_all must exist")
    body = ast.unparse(fn)
    ok("_borrowed_titles" not in body or "CloseDoc" not in body.split("_borrowed_titles")[1],
       "close_all must never call CloseDoc on a borrowed title")

    # The decisive line: a title is added to the close-list ONLY when we opened it. Checked
    # structurally because COM cannot be exercised here, and dropping the guard is a silent
    # one-word change that puts a designer's open documents back on the close-list.
    opener = next((n for n in ast.walk(tree)
                   if isinstance(n, ast.FunctionDef) and n.name == "open"
                   and "OpenDoc6" in ast.unparse(n)), None)
    ok(opener is not None, "the document-opening helper must exist")
    guarded = False
    for node in ast.walk(opener) if opener is not None else []:
        if isinstance(node, ast.If) and "_already" in ast.unparse(node.test):
            blk = "".join(ast.unparse(b) for b in node.body)
            if "_open_titles" in blk and "append" in blk:
                guarded = True
    ok(guarded,
       "the append to _open_titles must be guarded on the document NOT already being open — "
       "otherwise close_all closes documents this process never opened")


def test_an_extract_that_read_nothing_is_not_a_successful_read():
    """Per-file failures are written as error-only records and the analyser still exits
    zero, so an extraction where every file failed produced a non-empty list that read
    downstream as a successful read."""
    from invariants import check_job
    j = _job()
    j["solidworks_native"] = {"extract_incomplete": True, "files_read": 0, "files_failed": 9}
    r = check_job(j, write_back=False)
    ok("native_extract_incomplete" in [v["code"] for v in r["violations"]],
       "an extraction that read nothing must block")
    ok(not r["may_quote_firm"], "and must stop the price being firm")
    # PARTIAL COVERAGE BLOCKS until the failures are shown to be irrelevant. Nothing here
    # knows whether the file that failed was a fixture or a released component of the
    # assembly being priced — and if it was the latter the job is undercosted by whatever
    # that part contributes. "Some files failed" is not a thing to wave through.
    j2 = _job()
    j2["solidworks_native"] = {"files_read": 6, "files_failed": 2}
    r2 = check_job(j2, write_back=False)
    ok("native_extract_partial" in [v["code"] for v in r2["violations"]], "partial is flagged")
    ok(not r2["may_quote_firm"],
       "and blocks: an unread file may be a released component")
    # Shown to be outside the BOM closure, it drops to a warning and the price stands.
    j3 = _job()
    j3["solidworks_native"] = {"files_read": 6, "files_failed": 2,
                               "failed_outside_bom_closure": True}
    r3 = check_job(j3, write_back=False)
    ok(r3["may_quote_firm"],
       "failures proven irrelevant to the priced assembly do not block")


def test_freshness_without_a_manifest_says_it_is_weak():
    """The connector compared a fingerprint the analyser never wrote, so the check silently
    degraded to file mtime — which a copy, a restore or a touch defeats, and which cannot see
    a model that has been deleted or renamed at all."""
    from invariants import check_job
    j = _job()
    j["solidworks_native"] = {"manifest_absent": True, "freshness_check": "mtime_only"}
    codes = [v["code"] for v in check_job(j, write_back=False)["violations"]]
    ok("native_freshness_unverified" in codes,
       f"a timestamp-only check must not read as a verified one: {codes}")


def test_the_extract_payload_carries_a_manifest_and_still_reads_old_files():
    """The analyser wrote a bare list while the consumer looked for a dict, so the
    fingerprint was never compared. The new payload must carry one — and an extract already
    sitting in a job folder must not stop working because the writer gained a header."""
    import json as _json, tempfile
    from pathlib import Path
    from source_connectors.solidworks import load_native_extract, load_native_payload

    d = Path(tempfile.mkdtemp())
    v2 = d / "v2.json"
    v2.write_text(_json.dumps({
        "schema": "sw_native_extract.v2",
        "_manifest": {"native_files_fingerprint": "abc123", "files_read": 2, "files_failed": 0},
        "records": [{"title": "A-01M"}, {"title": "A-02M"}]}), encoding="utf-8")
    eq(len(load_native_extract(v2)), 2, "records read from the v2 payload")
    eq(load_native_payload(v2)["_manifest"]["native_files_fingerprint"], "abc123",
       "and the fingerprint is actually reachable, which it never was before")

    legacy = d / "legacy.json"
    legacy.write_text(_json.dumps([{"title": "B-01M"}]), encoding="utf-8")
    eq(len(load_native_extract(legacy)), 1, "a pre-manifest extract still reads")
    ok(load_native_payload(legacy)["manifest_absent"],
       "but reports that it has no manifest, rather than an empty one that reads as checked")


def test_archived_and_lock_files_are_not_counted_as_unread_models():
    """native_files_state counted every .SLD* under the folder, including superseded models
    and SolidWorks lock files the analyser itself skips — so an archived model from two
    revisions ago could block a drawing-only job for being 'not read'."""
    import tempfile
    from pathlib import Path
    from source_connectors.solidworks import native_files_state

    d = Path(tempfile.mkdtemp())
    (d / "live.SLDPRT").write_bytes(b"x")
    (d / "~$live.SLDPRT").write_bytes(b"x")            # SolidWorks lock file
    (d / "Archive").mkdir()
    (d / "Archive" / "old_rev.SLDPRT").write_bytes(b"x")
    (d / "OBSOLETE").mkdir()
    (d / "OBSOLETE" / "dead.SLDASM").write_bytes(b"x")
    st = native_files_state(d)
    eq(st["count"], 1, "only the live model counts")
    eq(st["files"], ["live.sldprt"], "archived revisions and lock files are not evidence")


def test_an_extract_supplied_from_elsewhere_says_it_cannot_be_verified():
    """The models routinely live somewhere other than the drawings, which is what
    SDI_SW_EXTRACT_JSON is for. Fingerprinting the JOB folder then compares a fingerprint of
    the models against a folder containing none — every freshly generated extract would read
    as STALE. And with no manifest recording where it came from, nothing can be checked at
    all; that must not read as a pass."""
    from invariants import check_job
    j = _job()
    j["solidworks_native"] = {"freshness_unverifiable": True,
                              "fingerprint_folder": None, "files_read": 22}
    r = check_job(j, write_back=False)
    codes = [v["code"] for v in r["violations"]]
    ok("native_freshness_unverifiable" in codes,
       f"an unverifiable extract must say so: {codes}")
    ok("native_extract_stale" not in codes,
       "and must NOT be reported as stale — that is a different, checkable fact")
    # A warning does not stop a firm quote, so "we cannot verify this" recorded as a warning
    # let a diagnostic extract produce a firm price while saying it could not be checked.
    ok(not r["may_quote_firm"],
       "unverifiable freshness must stop the price being firm, not merely annotate it")

    # An unreachable source drive is not evidence of staleness — only that we could not look.
    j2 = _job()
    j2["solidworks_native"] = {"source_unreachable": True, "fingerprint_folder": "K:\\models"}
    r2 = check_job(j2, write_back=False)
    codes2 = [v["code"] for v in r2["violations"]]
    ok("native_source_unreachable" in codes2, f"say the drive was unreachable: {codes2}")
    ok("native_extract_stale" not in codes2,
       "a missing network drive must not be reported as a stale extract")
    ok(not r2["may_quote_firm"], "and nothing verified means nothing firm")

    # Files that moved under the extraction are not a snapshot of anything.
    j3 = _job()
    j3["solidworks_native"] = {"changed_during_extraction": True, "fingerprint_before": "aaa"}
    r3 = check_job(j3, write_back=False)
    ok("native_models_changed_during_extraction" in [v["code"] for v in r3["violations"]],
       "a mid-run save must invalidate the extract")
    ok(not r3["may_quote_firm"], "and block")


def _load_analyser():
    """Import the SolidWorks analyser for its PURE helpers, on any platform.

    COM is imported lazily inside it, so the decision logic — which folders are the live
    design, whether a cut-list zero is a value — loads without pywin32. The module must be
    registered in sys.modules before exec, or dataclass field resolution cannot find it."""
    import importlib.util
    from pathlib import Path
    path = Path(__file__).resolve().parents[1] / "tools" / "solidworks" / "sw_native_analyse.py"
    spec = importlib.util.spec_from_file_location("sw_native_analyse", path)
    mod = importlib.util.module_from_spec(spec)
    sys.modules["sw_native_analyse"] = mod
    spec.loader.exec_module(mod)
    return mod


def test_exclusion_matches_whole_words_not_substrings():
    """'folder' contains 'old'. A substring rule therefore excluded 'CAD Folder' — and the
    consequence was not a missed folder but a silent one: the analyser READ those files while
    the manifest omitted them, so any later change to them was invisible to the freshness
    check. Discovery and fingerprinting must also share one rule, or they disagree about what
    the extract covers."""
    swa = _load_analyser()
    for keep in ("CAD Folder", "Folder", "Boldon Screens", "Goldsmith", "Models", "Released"):
        ok(not swa._is_excluded_dir(keep), f"{keep!r} is live design and must be kept")
    for drop in ("Archive", "OBSOLETE", "Old Revs", "rev-old", "_BAK", "Superseded",
                 "Do Not Use", "WIP", ".git"):
        ok(swa._is_excluded_dir(drop), f"{drop!r} is not the live design and must be dropped")

    # The consumer's copy must agree, or it counts files the analyser never opened.
    from source_connectors.solidworks import _is_excluded_dir as consumer_rule
    for name in ("CAD Folder", "Archive", "Old Revs", "Models", "_BAK", "Boldon Screens"):
        eq(consumer_rule(name), swa._is_excluded_dir(name),
           f"analyser and consumer must agree about {name!r}")


def test_a_linked_table_cell_is_read_until_something_answers():
    """A BOM table's cells are LINKED: Text2 can return an empty string while
    DisplayedText2 holds the resolved value. Treating the first non-None answer as final
    accepted "" and never asked the call that would have answered — which is why both of
    12120's drawings came back with zero BOM rows AND zero errors, reading as 'this drawing
    has no BOM' rather than 'we did not manage to read it'."""
    swa = _load_analyser()

    class _EmptyFirst:
        """The FIRST accessor tried answers successfully with an empty string; a later one
        holds the value. Whichever accessor happens to be first, an empty success must not
        end the search — that is the whole defect, and a fake that answers on the first call
        tests the ORDER rather than the rule."""
        def __init__(self, real): self.real = real
        def DisplayedText2(self, r, c, *a): return ""
        def Text2(self, r, c, *a): return ""
        def Text(self, r, c): return self.real.get((r, c), "")

    t = _EmptyFirst({(0, 0): "PART NO", (1, 0): "12120-01-01M"})
    eq(swa._table_text(t, 1, 0), "12120-01-01M",
       "an empty success must not end the search while a later accessor has the value")
    eq(swa._table_text(t, 5, 5), "", "a genuinely empty cell is still empty")

    class _LinkedCell:
        """The live shape: linked BOM cells resolve through DisplayedText2 while Text2
        returns nothing."""
        def __init__(self, real): self.real = real
        def DisplayedText2(self, r, c, *a): return self.real.get((r, c), "")
        def Text2(self, r, c, *a): return ""
    eq(swa._table_text(_LinkedCell({(1, 0): "12120-01-04M"}), 1, 0), "12120-01-04M",
       "and a linked cell resolves through DisplayedText2")

    class _OnlyText:
        """An older build with no DisplayedText2 at all."""
        def Text(self, r, c): return "LEGACY-01"
    eq(swa._table_text(_OnlyText(), 1, 0), "LEGACY-01", "the oldest form still works")


def test_discovery_and_fingerprinting_share_one_exclusion_rule():
    """Discovery kept its own substring list ("previous", "old versions") while the
    fingerprint used the whole-word rule. Two lists means the extract and the manifest can
    disagree about what was covered — a folder read by one and omitted by the other — which
    is the same silent divergence as "folder" matching "old", from the other direction."""
    swa = _load_analyser()
    ok(swa.ARCHIVE_FOLDER_TOKENS is swa._EXCLUDED_DIR_TOKENS,
       "discovery must not have a second token list")
    # Unifying must not WIDEN what gets read: everything the discovery list excluded still is.
    for name in ("previous", "Previous Revisions", "Old Versions", "WIP", "Do Not Use",
                 "Archive", "superseded", "backup"):
        ok(swa._is_excluded_dir(name), f"{name!r} was excluded before and must stay excluded")
    for name in ("CAD Folder", "Models", "Released", "Rev A", "Boldon"):
        ok(not swa._is_excluded_dir(name), f"{name!r} is live design")


def test_a_zero_count_from_the_cut_list_survives_parsing():
    """'Cut Outs = 0' says this part is a plain blank with one pierce. Parsed with the LENGTH
    reader — which rejects non-positive values, correctly, because a blank cannot be 0mm long
    — it became None, and the resolver's explicit-zero handling never saw a zero to defend.
    The confusion was fixed downstream and reintroduced one layer up."""
    swa = _load_analyser()
    eq(swa._num_count("0"), 0, "an explicit zero count is a value")
    eq(swa._num_count(0), 0, "including as a number")
    eq(swa._num_count("3"), 3, "and a real count reads normally")
    eq(swa._num_count(""), None, "empty is absent")
    eq(swa._num_count(None), None, "and so is nothing at all")
    # Lengths keep the old rule: a 0mm blank is a failed read, not a flat part.
    eq(swa._num_mm("0"), None, "a zero LENGTH is still rejected")
    eq(swa._num_mm("126.39"), 126.39, "and a real length still reads")


def test_the_solidworks_version_is_captured_before_the_session_closes():
    """shutdown() clears session.sw, so reading the version afterwards recorded an empty
    string on every extract — a manifest field that always looked like 'we did not ask'."""
    import ast
    from pathlib import Path
    src = (Path(__file__).resolve().parents[1] / "tools" / "solidworks"
           / "sw_native_analyse.py").read_text(encoding="utf-8")
    ok("_sw_version_string(session)" not in src.split('"solidworks_version"')[1][:200],
       "the manifest must not call _sw_version_string AFTER shutdown — read the captured value")
    tree = ast.parse(src)
    main_fn = next((n for n in ast.walk(tree)
                    if isinstance(n, ast.FunctionDef) and n.name == "main"), None)
    ok(main_fn is not None, "main() must exist")
    body = ast.unparse(main_fn)
    _cap = body.index("_sw_version = ") if "_sw_version = " in body else -1
    _shut = body.index("session.shutdown()") if "session.shutdown()" in body else -1
    ok(_cap != -1, "the version must be captured into a variable")
    ok(_cap < _shut, "and captured BEFORE shutdown, while the session is still alive")


def test_the_pipeline_actually_runs_the_analyser():
    """Source-level, because the scan path cannot be exercised without a job folder — and a
    behavioural fixture that stops short of the call site is exactly what let this sit.

    native_extract_for_job defaults to run=False, and file_scan called it without the
    argument, so the pipeline CONSUMED an extract but never produced one. On a machine with
    SolidWorks the strongest source available was used or skipped depending on whether
    somebody had remembered to run a separate script."""
    import ast
    from pathlib import Path

    src = (Path(__file__).resolve().parents[1] / "src" / "file_scan.py").read_text(encoding="utf-8")
    calls = [n for n in ast.walk(ast.parse(src))
             if isinstance(n, ast.Call)
             and getattr(n.func, "id", getattr(n.func, "attr", "")) == "native_extract_for_job"]
    ok(calls, "file_scan must call native_extract_for_job at all")
    for c in calls:
        kw = {k.arg: k.value for k in c.keywords}
        ok("run" in kw, f"line {c.lineno}: the call must pass run= explicitly, not inherit "
                        f"the run=False default")
        ok(not (isinstance(kw.get("run"), ast.Constant) and kw["run"].value is False),
           f"line {c.lineno}: run=False hard-coded — the analyser would never generate an "
           f"extract, only consume one somebody else made")


def test_native_models_present_but_unread_is_loud():
    """The pipeline consumed an extract only if one already existed, so on a machine WITH
    SolidWorks the strongest source available was used or skipped depending on whether a
    person had remembered to run a separate script — and skipping it looked exactly like a
    job with no models at all."""
    from invariants import check_job
    j = _job()
    j["solidworks_native"] = {"found": False, "native_present_but_unread": True,
                              "native_files_present": 7, "reason": "no extract generated"}
    r = check_job(j, write_back=False)
    ok("native_models_not_read" in [v["code"] for v in r["violations"]],
       "models in the folder but unread must block")
    ok(not r["may_quote_firm"], "and stop the price being firm")

    # Stale is its own fact: an extract taken before the design changed.
    j2 = _job()
    j2["solidworks_native"] = {"extract_stale": True, "native_files_present": 7}
    r2 = check_job(j2, write_back=False)
    ok("native_extract_stale" in [v["code"] for v in r2["violations"]],
       "an extract older than the models must block")
    # A job with no models at all is not a violation — most jobs are drawings only.
    ok(check_job(_job(), write_back=False)["may_quote_firm"],
       "a job with no native models is unaffected")


def test_the_analyser_reports_why_it_could_not_run():
    """Failures were swallowed whole, so an analyser that could not start — no SolidWorks, no
    licence, a bad path — was indistinguishable from a job with no models."""
    from source_connectors.solidworks import _run_analyser
    err = _run_analyser("/nonexistent/job/folder",
                        analyser="/nonexistent/analyser.py",
                        python_exe="definitely-not-a-python-executable")
    ok(err, "a failure to launch must return a reason, not None")
    ok("analyser" in str(err).lower(), f"and the reason must name what failed: {err!r}")


def test_native_files_are_fingerprinted_for_freshness():
    """An extract is a photograph of the models at the moment it was taken. Reading one
    without checking it still describes the files on disk is how a job gets costed from last
    month's geometry with nothing on screen to say so."""
    import tempfile
    from pathlib import Path
    from source_connectors.solidworks import native_files_state

    d = Path(tempfile.mkdtemp())
    eq(native_files_state(d)["count"], 0, "an empty folder has no native files")
    (d / "a.SLDPRT").write_bytes(b"x" * 10)
    (d / "b.sldasm").write_bytes(b"y" * 20)
    (d / "c.pdf").write_bytes(b"z")              # not a native model
    st = native_files_state(d)
    eq(st["count"], 2, "parts and assemblies counted, drawings-and-PDFs distinguished")
    ok(st["fingerprint"], "and fingerprinted")
    before = st["fingerprint"]
    (d / "a.SLDPRT").write_bytes(b"x" * 999)     # the design changed
    ok(native_files_state(d)["fingerprint"] != before,
       "a changed model must change the fingerprint, or staleness cannot be detected")


def test_late_passes_cannot_clobber_the_assembly_quantity():
    """The PDF passes run late and used to write straight to the record. part_index treated a
    quantity of ONE as an empty slot, so a part the model says there is one of was open to
    replacement by whatever a table happened to say."""
    from source_precedence import apply_field
    p = {"part_number": "01M", "quantity": 1, "quantity_source": "solidworks_api"}
    ok(not apply_field(p, "quantity", 4, "bom_tree"),
       "a PDF BOM reading must not displace the assembly the shop builds from")
    eq(p["quantity"], 1, "a quantity of one is a value, not an empty slot")
    # With nothing recorded, the same write is welcome — it is filling a gap, not a fight.
    q = {"part_number": "02M"}
    ok(apply_field(q, "quantity", 4, "bom_tree"), "an unclaimed quantity may be filled")
    eq(q["quantity_source"], "bom_tree", "and the filler is named")


def test_an_assembly_parent_does_not_report_a_missing_rate_for_work_it_does_not_do():
    """Job 12120's 101 and 103 both carried missing_labour_rate:folding, and 103 also
    hole_machining — reading as "work identified and not priced", which is under-costing.
    The truth was the reverse. An assembly does not fold; its children do, and they carried
    the fold. Suppressing it on the parent is correct.

    The flag came from an inconsistent record. estimate_process_times returns five time
    maps, unit_times_min and times_min derived from the other two and built BEFORE the
    assembly-parent strip runs. The strip popped folding from two of them, so the part said
    "does not fold" in the maps costing reads and "folds" in the map the risk check reads,
    and requested_ops - costed_ops reported a rate that was never missing. It sent an
    estimator hunting for a configuration error that did not exist."""
    # estimate_part, not estimate_process_times: the strip runs in the caller, after the
    # time maps have been built. Testing the producer alone would start before the code
    # under test — the same gap that let a reverted call site pass earlier in this file.
    from estimator import estimate_part
    parent = {
        "part_number": "12120-01-101", "normalized_material": "MILD_STEEL",
        "is_assembly_parent": True, "is_sub_assembly": True,
        "textual_operations": ["folding", "hole_machining", "welding", "assembly"],
        "normalized_thickness_mm": 1.5, "overall_length_mm": 300, "overall_width_mm": 200,
        "manufacturing_features": {"bend_count": 2, "hole_count": 4},
        "geometry_rollup": {"estimated_cut_length_mm": 900.0},
        "quantity": 1,
    }
    _est = estimate_part(parent) or {}
    proc = _est.get("process_estimate") or parent.get("process_estimate") or {}
    ok(proc, "estimate_part must return a process estimate")
    maps = ("setup_times_min", "run_times_min_per_unit", "unit_times_min", "times_min")
    present = {m: sorted(set(proc.get(m) or {}) & {"folding", "hole_machining"})
               for m in maps if isinstance(proc.get(m), dict)}
    for m, ops in present.items():
        eq(ops, [], f"{m} still lists fabrication ops on an assembly parent")
    # The DERIVED maps must agree with each other. setup_times_min legitimately holds only
    # operations that have a setup, so it is not comparable — but unit_times_min and
    # times_min are both built from the same union and any divergence between them means one
    # was edited and the other was not, which is exactly the defect above.
    _u = set(proc.get("unit_times_min") or {})
    _t = set(proc.get("times_min") or {})
    eq(sorted(_u ^ _t), [],
       "unit_times_min and times_min disagree about which operations this part has")
    _flags = [f for f in (_est.get("risk_flags") or [])
              if str(f).startswith("missing_labour_rate:")]
    eq([f for f in _flags if "folding" in f or "hole_machining" in f], [],
       "no missing-rate flag for work the parent does not do")

    # A real fabricated part is untouched — the strip is about parents, not about folding.
    leaf = dict(parent, part_number="12120-01-01M", is_assembly_parent=False,
                is_sub_assembly=False)
    _lest = estimate_part(leaf) or {}
    lproc = _lest.get("process_estimate") or leaf.get("process_estimate") or {}
    ok(any("folding" in (lproc.get(m) or {}) for m in maps),
       "a leaf part that folds must keep its fold")


def test_the_same_catalogue_always_yields_the_same_price():
    """THE DEFECT NO OTHER FIXTURE COULD CATCH. Job 12120 priced three times on identical
    inputs — same drawings, same SolidWorks extract, same quantity — at GBP 27.67, GBP 29.39
    and GBP 32.86. An 18.8% swing. Labour was identical to the penny every run; the whole
    movement was bought-in lines.

    The knurled knob went 1.45 -> 1.90 -> 1.45. It came BACK, so it was never a catalogue
    being updated: it was the same two rows being chosen between. The rank tuple was
    (-priority, penalty, -confidence), two rows for one part code tie on all three, and
    `sorted(...)[0]` then fell through to Python's stable sort — which preserves whatever
    order the connector returned, and a SQL query with no ORDER BY does not promise one.

    What made this so dangerous is that every run RECONCILED. Rows summed to subtotals,
    subtotals to the unit price, every money invariant passing. The engine was internally
    consistent and externally unrepeatable."""
    from price_sources import PriceCandidate, _candidate_rank_tuple, _price_disagreement

    def _c(src, price, **meta):
        return PriceCandidate(source=src, kind="material", price=price, currency="GBP",
                              unit="each", confidence=0.9, metadata=meta)

    rules = {"source_priority": {"udef_sqlserver": 100}, "freshness_penalty": {}}
    rows = [_c("udef_sqlserver", 1.90, row_id=7), _c("udef_sqlserver", 1.45, row_id=2)]

    # The same set in EITHER order must give the same answer. That is the whole property.
    a = sorted(rows, key=lambda c: _candidate_rank_tuple(c, rules))[0]
    b = sorted(list(reversed(rows)), key=lambda c: _candidate_rank_tuple(c, rules))[0]
    eq(a.price, b.price, "row order must not change the price — this is the 1.45/1.90 flip")
    eq(a.price, 1.45, "and of two tied catalogue rows the cheaper is the defensible one")

    # Identical rows in rank AND price still order deterministically.
    same = [_c("udef_sqlserver", 2.00, row_id=1), _c("udef_sqlserver", 2.00, row_id=9)]
    eq(sorted(same, key=lambda c: _candidate_rank_tuple(c, rules))[0].metadata,
       sorted(list(reversed(same)), key=lambda c: _candidate_rank_tuple(c, rules))[0].metadata,
       "two rows agreeing on price must still resolve to the same ROW every time")

    # DETERMINISM IS NOT CORRECTNESS. Picking 1.45 every time is repeatable and still hides a
    # catalogue holding two different prices for one code. The spread is reported.
    d = _price_disagreement(rows)
    ok(d, "a catalogue that answers twice with different prices must be reported")
    eq((d["low_gbp"], d["high_gbp"]), (1.45, 1.90), "with both figures named")
    ok(d["spread_pct"] > 20, f"and the spread quantified ({d['spread_pct']}%)")
    ok(_price_disagreement([_c("udef_sqlserver", 1.45)]) is None, "one answer is no conflict")
    ok(_price_disagreement([_c("a", 1.45), _c("b", 1.4501)]) is None,
       "a rounding-level difference is not a disagreement")


def test_an_ai_guessed_price_cannot_make_a_quote_firm():
    """Job 12120 priced three times on identical inputs at GBP 27.67, GBP 29.39 and GBP 32.86.
    Labour was identical to the penny every run; the steel never moved. SQL missed on
    THUM620, the knurled knob and the screen cable every time, so the lookup fell through to
    an LLM market estimate — confidence 0.3-0.4, INDICATIVE by its own description — and the
    cable came back at 4.54, then 6.00, then 8.54.

    No tie-break fixes that. An LLM is being asked what a part costs and answering
    differently each time, which is what it does. Filling a gap that way is defensible;
    putting it on a quote as the applied unit cost is not, because the number cannot be
    reproduced, audited, or defended to a customer who asks how it was reached.

    And every one of those runs RECONCILED. No other check here can see it, because each run
    is individually consistent.

    THE SHAPE IS THE TEST. The first version of this fixture hand-built
    part_estimates[*].price_source at the top level, and passed — while the same check
    reported CLEAR on the real 12120 JSON, where all three AI prices sit at
    estimate_summary.part_estimates[*].cost_breakdown.system_cost.source, stamped
    source_name=llm_market_estimate but source_type=external. So this fixture is built to the
    real nesting: a top-level `parts` list carrying geometry and no money, the priced records
    under estimate_summary, and the guessed price two levels inside a cost breakdown."""
    from invariants import check_job

    def _system_cost(code, source_name, price, applied_to_total=True, source_type="external"):
        """One bought-in line, in the shape the estimator actually writes."""
        return {
            "part_number": code, "quantity": 1, "quantity_source": "bom_tree",
            "cost_breakdown": {"system_cost": {
                "unit_cost_gbp": price, "matched_part_code": code,
                "applied_to_total": applied_to_total,
                "source": {"source_name": source_name, "source_type": source_type,
                           "applied": True, "affects_total": applied_to_total,
                           "source_rank": 0, "confidence": 0.35,
                           "selected": {"source": source_name, "price": price}},
            }},
        }

    j = _job()
    # Geometry records: where the old check looked, and where no price has ever lived.
    j["parts"] = [{"part_number": "12120-01-01M", "geometry_source": "dxf_flat_pattern",
                   "dxf_measured_outline": True, "blank_length_mm": 126.39,
                   "blank_width_mm": 82.2, "blank_area_mm2": 10389.3}]
    # Priced records: where the money actually is.
    j["estimate_summary"] = {"part_estimates": [
        _system_cost("THUM620", "udef_sqlserver", 1.16),
        _system_cost("BI-SCREENCABLE", "llm_market_estimate", 8.54),
        _system_cost("BI-KNURLEDKNOB", "llm_market_estimate", 1.90),
    ]}
    r = check_job(j, write_back=False)
    codes = [v["code"] for v in r["violations"]]
    ok("price_not_reproducible" in codes, f"an AI-guessed price must block: {codes}")
    ok(not r["may_quote_firm"], "and stop the quote being firm")
    _v = next(v for v in r["violations"] if v["code"] == "price_not_reproducible")
    eq(_v["detail"]["count"], 2, "both guessed lines counted, not just the first")
    # The violation ends by telling someone to add these codes to the catalogue. On the live
    # 12120 JSON it named estimate_summary.part_estimates[11], which nobody can add to
    # anything — the instruction was unfollowable.
    eq(sorted(_v["detail"]["parts"]), ["BI-KNURLEDKNOB", "BI-SCREENCABLE"],
       "naming the codes to catalogue, not array indices")
    ok("part_estimates[" not in _v["message"], "and not in the sentence a human reads")
    _sources = {str(l["source"]) for l in _v["detail"]["lines"]}
    eq(_sources, {"llm_market_estimate"}, "and only the guessed ones")
    ok("udef_sqlserver" not in str(_v["detail"]),
       "a line priced from the catalogue is not implicated")

    # A guessed price the estimator resolved but did NOT add to the total cannot move the
    # total, so it is not what makes the job unrepeatable.
    j3 = _job()
    j3["estimate_summary"] = {"part_estimates": [
        _system_cost("BI-SCREENCABLE", "llm_market_estimate", 8.54, applied_to_total=False)]}
    ok("price_not_reproducible" not in
       [v["code"] for v in check_job(j3, write_back=False)["violations"]],
       "a resolved-but-unapplied guess does not block")

    # A document written before affects_total existed carries only the looser `applied` on
    # the stamp, with the real answer on the sibling beside it. Reading the stamp alone
    # reported a GBP 75.00 weldment line as reaching a GBP 32.86 unit price, which it plainly
    # had not — the enclosing block said applied_to_total was false.
    j5 = _job()
    _legacy = _system_cost("12120-01-101", "llm_market_estimate", 75.00,
                           applied_to_total=False)
    _legacy["cost_breakdown"]["system_cost"]["source"].pop("affects_total")
    j5["estimate_summary"] = {"part_estimates": [_legacy]}
    ok("price_not_reproducible" not in
       [v["code"] for v in check_job(j5, write_back=False)["violations"]],
       "the sibling flag is believed over the stamp's looser one on older documents")

    # Every line from a real source: nothing to report.
    j2 = _job()
    j2["estimate_summary"] = {"part_estimates": [
        _system_cost("THUM620", "udef_sqlserver", 1.16)]}
    ok("price_not_reproducible" not in
       [v["code"] for v in check_job(j2, write_back=False)["violations"]],
       "catalogue prices raise nothing")

    # A job with no price stamps anywhere has not been checked — it has not passed.
    j4 = _job()
    j4["part_estimates"][0].pop("material_estimate")
    r4 = check_job(j4, write_back=False)
    ok("price_reproducibility_not_evaluated" in [v["code"] for v in r4["violations"]],
       "and a job with no priced lines reads as unverified, never as clear")


def test_the_estimator_stamps_what_kind_of_thing_priced_each_line():
    """The gate above can only fire on what the estimator wrote down.

    An LLM market estimate reached the sheet stamped source_type='external' — the same word a
    SQL catalogue hit carries — because the type was derived from the connector that answered
    rather than from what the source IS. Nothing downstream could tell them apart. This tests
    the writer, not the classifier: the classifier was right all along and never consulted."""
    from estimator import _build_price_source_metadata

    def _stamp(source, **sel):
        sel.setdefault("price", 8.54)
        return _build_price_source_metadata(
            {"selected": dict(sel, source=source)}, fallback_source="x", applied=True)

    guessed = _stamp("llm_market_estimate")
    eq(guessed["source_class"], "ai_estimate", "an LLM estimate is named as one")
    eq(guessed["reproducible"], False, "and declared unrepeatable")
    ok(guessed["source_type"] != "external",
       f"never labelled like a catalogue hit (got {guessed['source_type']!r})")
    ok("indicative" in str(guessed["review_reason"]).lower(),
       "carrying the sentence an estimator needs to read")

    catalogued = _stamp("udef_sqlserver")
    eq(catalogued["source_class"], "catalogue", "a SQL row is a catalogue price")
    eq(catalogued["reproducible"], True, "and repeats")

    # The connector reached through the pricing service reports its mode in metadata; the
    # source name alone would not give it away.
    viaservice = _build_price_source_metadata(
        {"selected": {"source": "web", "price": 4.54,
                      "metadata": {"pricing_mode": "web_ai_fallback"}}},
        fallback_source="x", applied=True)
    eq(viaservice["source_class"], "ai_estimate",
       "an AI mode is caught even when the source name is just the connector")

    # And the stamp must be findable without knowing where it was written.
    import price_provenance
    found = dict(price_provenance.iter_price_stamps(
        {"estimate_summary": {"part_estimates": [
            {"cost_breakdown": {"system_cost": {"source": guessed}}}]}}))
    eq(list(found), ["estimate_summary.part_estimates[0].cost_breakdown.system_cost.source"],
       "the walker finds a stamp by marker, at whatever depth it was written")


# ── invariants — the checks that make a wrong answer loud ────────────────────────────
def _job(**over):
    """A job that passes every invariant, so each test can break exactly one thing."""
    job = {
        "final_estimate": {
            "schema": "final_estimate.v2",
            "totals": {"material_gbp": 10.07, "labour_gbp": 52.0, "unit_gbp": 62.07},
            "material_rows": [{"workbook_row": 11, "description": "Pem", "total_value_gbp": 9.64},
                              {"workbook_row": 63, "description": "01M", "total_value_gbp": 0.43}],
            "labour_rows": [{"workbook_row": 41, "operation": "Laser", "total_value_gbp": 22.0},
                            {"workbook_row": 43, "operation": "Fold", "total_value_gbp": 30.0}],
            "adapter_problems": [],
        },
        "workbook_labour": {
            "schema": "workbook_labour_rows.v2",
            "rows": [{"workbook_row": 41, "route_group_id": "rg_a", "wb_operation": "Laser",
                      "engine_operations": ["laser_cutting"], "part_numbers": ["01M"]},
                     {"workbook_row": 43, "route_group_id": "rg_b", "wb_operation": "Fold",
                      "engine_operations": ["folding"], "part_numbers": ["01M"]}],
        },
        "part_estimates": [{"part_number": "01M", "operations": ["laser_cutting", "folding"],
                            "normalized_material": "MILD_STEEL", "material_source": "dxf",
                            "quantity": 2, "quantity_source": "solidworks_api",
                            "geometry_source": "dxf_flat_pattern", "dxf_measured_outline": True,
                            "blank_length_mm": 126.39, "blank_length_mm_source": "dxf",
                            "blank_width_mm": 82.2, "blank_area_mm2": 10389.3,
                            # A clean job has priced something, from somewhere nameable. A
                            # job with no price stamp at all is not clean, it is unchecked —
                            # so the baseline here carries one, from a catalogue.
                            "material_estimate": {"price_source": {
                                "schema": "price_source.v1", "source_name": "udef_sqlserver",
                                "source_class": "catalogue", "reproducible": True,
                                "applied": True, "affects_total": True}}}],
    }
    for k, v in over.items():
        job[k] = v
    return job


def test_a_clean_job_passes_every_invariant():
    from invariants import check_job
    r = check_job(_job(), write_back=False)
    ok(r["ok"], f"a consistent job must pass: {[v['code'] for v in r['violations']]}")
    eq(r["blocking"], 0, "no blocking violations on a clean job")


def test_rows_must_sum_to_the_workbook_total():
    """The defect this is built from: material rows summed to GBP 9.64 against the sheet's
    own GBP 10.07 and nothing objected. A snapshot that will not reconcile to its own total
    cannot be exported or quoted from."""
    from invariants import check_job
    j = _job()
    j["final_estimate"]["material_rows"] = [{"workbook_row": 11, "total_value_gbp": 9.64}]
    codes = [v["code"] for v in check_job(j, write_back=False)["violations"]]
    ok("material_rows_do_not_sum_to_total" in codes,
       f"a 43p shortfall must be a blocking violation, got {codes}")
    # An Excel error reads back as null. Summing it as zero manufactures agreement out of
    # missing data — the one thing these checks exist to prevent.
    j2 = _job()
    j2["final_estimate"]["labour_rows"] = [{"workbook_row": 41, "total_value_gbp": 52.0},
                                           {"workbook_row": 43, "total_value_gbp": None}]
    codes2 = [v["code"] for v in check_job(j2, write_back=False)["violations"]]
    ok("labour_rows_incomplete" in codes2,
       f"a row that did not calculate is missing data, not zero: {codes2}")


def test_the_unit_price_must_equal_its_parts():
    """Job 12120: material rows reconciled to GBP 13.43, labour rows to GBP 12.30, both
    checks passed — and the sheet's Total Unit Cost Price was GBP 27.67. GBP 1.94, seven per
    cent of the price, belonging to nothing on the sheet. Every row reconciled to its own
    subtotal and nothing ever asked whether the subtotals reconciled to the price."""
    from invariants import check_job
    j = _job()
    j["final_estimate"]["totals"] = {"material_gbp": 13.43, "labour_gbp": 12.30,
                                     "unit_gbp": 27.67}
    j["final_estimate"]["material_rows"] = [{"workbook_row": 11, "total_value_gbp": 13.43}]
    j["final_estimate"]["labour_rows"] = [{"workbook_row": 41, "total_value_gbp": 22.0},
                                          {"workbook_row": 43, "total_value_gbp": -9.70}]
    r = check_job(j, write_back=False)
    codes = [v["code"] for v in r["violations"]]
    ok("unit_price_does_not_equal_its_parts" in codes,
       f"an unexplained GBP 1.94 in the price must block: {codes}")
    ok(not r["may_quote_firm"], "and stop the price being firm")

    # Adding up is the normal case and must stay silent.
    j2 = _job()
    j2["final_estimate"]["totals"] = {"material_gbp": 10.07, "labour_gbp": 52.0,
                                      "unit_gbp": 62.07}
    ok("unit_price_does_not_equal_its_parts" not in
       [v["code"] for v in check_job(j2, write_back=False)["violations"]],
       "a price that equals its parts raises nothing")

    # A declared component is accounted for by NAME, not left in the gap.
    j3 = _job()
    j3["final_estimate"]["totals"] = {"material_gbp": 13.43, "labour_gbp": 12.30,
                                      "other_gbp": 1.94, "unit_gbp": 27.67}
    ok("unit_price_does_not_equal_its_parts" not in
       [v["code"] for v in check_job(j3, write_back=False)["violations"]],
       "an uplift declared as its own figure is a component, not a discrepancy")


def test_the_unit_price_uplift_must_be_attributed_not_assumed():
    """12120's unit price is (material + labour) / 0.9299 — a 7.54% uplift, identical to four
    decimal places across two runs with different material totals. It is real and deliberate:
    the template's overhead absorption.

    But the RESIDUAL IS NOT AN ANSWER. Stamping other_gbp = unit - material - labour would
    make the reconciliation invariant tautological — it could never fail again, which is
    worse than not having it. The uplift is declared only when the unit cell's own FORMULA
    accounts for it.

    Reading the formula also catches what is already true here: config documents the divisor
    as 0.92, which gives GBP 27.97 on this job, not GBP 27.67. The live template does not do
    what the comment says, and a constant taken from config would have declared a wrong
    number with total confidence."""
    from wep_readback_from_xlsx import read_unit_price_composition

    class _Sheet:
        """Only what the reader touches: Cells(r,c).Value/.Formula and Range(ref).Value."""
        def __init__(self, formula, refs=None, unit=27.67):
            self.formula, self.refs, self.unit = formula, refs or {}, unit
        class _C:
            def __init__(self, v, f=""): self.Value, self.Formula = v, f
        def Cells(self, r, c):
            if r == 5 and c == 1: return self._C("Total Unit Cost Price")
            if r == 5 and c == 8: return self._C(self.unit, self.formula)
            return self._C(None)
        def Range(self, ref): return self._C(self.refs.get(ref))

    # The real shape: a divisor written into the formula.
    v = read_unit_price_composition(_Sheet("=((M59+M103)/(1-M107))/0.93",
                                           refs={"M107": 0.0}), 13.43, 12.30, 27.67, 20, 12)
    ok(v["explained"], f"a formula that accounts for the gap must be attributed: {v}")
    eq(round(v["other_gbp"], 2), 1.94, "and the uplift declared as a figure")
    ok("0.93" in str(v["basis"]), f"named, not merely observed: {v['basis']}")

    # Same arithmetic expressed as a rebate cell instead — also attributable.
    v = read_unit_price_composition(_Sheet("=(M59+M103)/(1-M107)", refs={"M107": 0.07}),
                                    13.43, 12.30, 27.67, 20, 12)
    ok(v["explained"] and "M107" in str(v["basis"]), f"a referenced cell is named too: {v}")

    # A formula that does NOT account for the gap must leave it undeclared, so the invariant
    # still fires. This is the case that stops the check becoming self-fulfilling.
    v = read_unit_price_composition(_Sheet("=(M59+M103)/0.92", refs={}),
                                    13.43, 12.30, 27.67, 20, 12)
    ok(not v["explained"],
       "0.92 gives 27.97, not 27.67 — an uplift the formula cannot explain stays undeclared")
    eq(v["other_gbp"], None, "and no figure is invented for it")

    # A sheet that simply adds up needs no uplift at all.
    v = read_unit_price_composition(_Sheet("=M59+M103", unit=25.73), 13.43, 12.30, 25.73, 20, 12)
    ok(v["explained"] and v["other_gbp"] == 0.0, "nothing to explain is explained")


def test_other_gbp_is_never_stamped_as_a_bare_residual():
    """Source-level, because the stamp runs through Excel COM and cannot be exercised here.

    other_gbp = unit - material - labour would satisfy the reconciliation invariant by
    construction: the check could never fail again on any job, and a silently-wrong unit
    price would sail through it forever. The figure may only be stamped when the unit cell's
    formula ACCOUNTED for the gap. A behavioural fixture on the composition reader proves the
    reader is honest; it says nothing about whether the caller respects it."""
    import ast
    from pathlib import Path
    src = (Path(__file__).resolve().parents[1] / "src"
           / "wep_readback_from_xlsx.py").read_text(encoding="utf-8")
    tree = ast.parse(src)
    found = False
    for node in ast.walk(tree):
        if not isinstance(node, ast.Dict):
            continue
        for k, v in zip(node.keys, node.values):
            if isinstance(k, ast.Constant) and k.value == "other_gbp":
                found = True
                _v = ast.unparse(v)
                ok("-" not in _v,
                   f"other_gbp is being computed as a residual ({_v}) — that makes the "
                   f"reconciliation invariant unfailable")
        # The conditional spread is the sanctioned form: present only when explained.
        for v in node.values:
            pass
    _txt = "".join(src.split())
    ok('"other_gbp":_comp.get("other_gbp")}if_comp.get("explained")' in _txt or not found,
       "other_gbp must be stamped only when the formula explained the uplift")


def test_every_priced_row_must_join_exactly_once():
    from invariants import check_job
    j = _job()
    j["final_estimate"]["labour_rows"].append(
        {"workbook_row": 99, "operation": "Weld", "total_value_gbp": 8.0})
    codes = [v["code"] for v in check_job(j, write_back=False)["violations"]]
    ok("priced_row_without_identity" in codes,
       f"a priced row joining to nothing belongs to no parts: {codes}")
    j2 = _job()
    j2["workbook_labour"]["rows"].append(
        {"workbook_row": 41, "route_group_id": "rg_c", "wb_operation": "Laser"})
    codes2 = [v["code"] for v in check_job(j2, write_back=False)["violations"]]
    ok("priced_row_with_ambiguous_identity" in codes2,
       f"one cost must not be reported against two groups: {codes2}")


def test_a_block_that_could_not_be_read_is_a_failure_not_a_footnote():
    """Returning no rows for a missing header is indistinguishable from a block with
    nothing in it, and the two mean opposite things."""
    from invariants import check_job
    j = _job()
    j["final_estimate"]["adapter_problems"] = [
        {"block": "tube", "code": "header_row_not_found", "message": "template moved"}]
    codes = [v["code"] for v in check_job(j, write_back=False)["violations"]]
    ok("workbook_block_not_read" in codes, f"an unread block must block: {codes}")


def test_measured_geometry_must_actually_be_measured():
    """'Measured' unlocks the credibility gate and the blank-allowance skip. A part claiming
    it with no outline is doing all that on the strength of a matched filename."""
    from invariants import check_job
    j = _job()
    j["part_estimates"][0].update({"blank_length_mm": 0, "blank_width_mm": 0,
                                   "blank_area_mm2": 0})
    codes = [v["code"] for v in check_job(j, write_back=False)["violations"]]
    ok("measured_geometry_without_outline" in codes,
       f"an empty measurement must not pass as measured: {codes}")
    # A matched-but-unmeasured file is an honest state and says so in its own name.
    j2 = _job()
    j2["part_estimates"][0].update({"geometry_source": "dxf_matched_no_geometry",
                                    "dxf_measured_outline": False,
                                    "blank_length_mm": 0, "blank_width_mm": 0,
                                    "blank_area_mm2": 0})
    codes2 = [v["code"] for v in check_job(j2, write_back=False)["violations"]]
    ok("measured_geometry_without_outline" not in codes2,
       "a file that says it measured nothing is honest, not a violation")


def test_an_unknown_schema_is_never_read_as_if_known():
    from invariants import check_job
    j = _job()
    j["final_estimate"]["schema"] = "final_estimate.v9"
    codes = [v["code"] for v in check_job(j, write_back=False)["violations"]]
    ok("unknown_schema" in codes,
       f"a contract whose shape may have moved must not be read anyway: {codes}")


def test_the_checks_run_against_the_stamped_document_not_a_stale_view():
    """One run produced two verdicts. The console said "2 blocking, 6 unverified"; the quote
    for the same job said "3 consistency check(s) FAILED", the extra one being the unit-price
    gap.

    The read-back writes final_estimate to the JSON ON DISK, not to the in-memory summary.
    Checking the summary therefore ran six checks against a job carrying no final_estimate at
    all. Fail-closed did its job — every one reported itself UNVERIFIED rather than passing —
    but two views of one job that disagree is the exact defect this layer exists to stop.

    The document checked must be the one the read-back stamped."""
    from invariants import check_job

    # The in-memory view, mid-run: no final_estimate yet.
    pre = {"part_estimates": [{"part_number": "01M"}]}
    r_pre = check_job(pre, write_back=False)
    _unev = [v["code"] for v in r_pre["violations"] if v["severity"] == "unverified"]
    ok(len(_unev) >= 4,
       f"a summary with no final_estimate cannot verify the reconciliation checks: {_unev}")
    ok(not r_pre["may_quote_firm"], "and must not be quotable")

    # The stamped document: the same checks now have something to read, and find the real
    # problem — which the earlier view could not have reported.
    post = dict(pre, **{"final_estimate": {
        "schema": "final_estimate.v2",
        "totals": {"material_gbp": 15.03, "labour_gbp": 12.30, "unit_gbp": 29.39},
        "material_rows": [{"workbook_row": 11, "total_value_gbp": 15.03}],
        "labour_rows": [{"workbook_row": 41, "total_value_gbp": 12.30}],
        "adapter_problems": []}})
    r_post = check_job(post, write_back=False)
    _codes = [v["code"] for v in r_post["violations"]]
    ok("unit_price_does_not_equal_its_parts" in _codes,
       f"the stamped document exposes the real gap: {_codes}")
    ok("totals_reconcile_not_evaluated" not in _codes,
       "and that check is no longer unverified — it had data to read")
    # The two verdicts differ, which is precisely why only ONE may be reported.
    ok(r_pre["blocking"] != r_post["blocking"],
       "the two views genuinely disagree — checking both and printing one is the bug")


def test_main_checks_the_document_it_stamps():
    """Source-level, because the scan path cannot be exercised here — and a behavioural
    fixture on check_job alone would start after the decision being tested, which is the
    call-site gap that has caught this suite repeatedly."""
    import ast
    from pathlib import Path
    src = (Path(__file__).resolve().parents[1] / "src" / "main.py").read_text(encoding="utf-8")
    calls = [n for n in ast.walk(ast.parse(src))
             if isinstance(n, ast.Call)
             and getattr(n.func, "id", "") in ("_check_job", "check_job")]
    ok(calls, "main must run the invariants")
    eq(len(calls), 1,
       "the invariants must be run ONCE. Two calls means two verdicts for one job, and the "
       "console and the quote reported different numbers of failures for exactly that reason")
    _arg = ast.unparse(calls[0].args[0]) if calls[0].args else ""
    ok(_arg != "summary",
       f"the checks must read the document the read-back stamped, not the in-memory "
       f"summary that has no final_estimate (got {_arg!r})")


def test_a_job_that_was_never_checked_is_not_a_pass():
    """THE FAILURE THIS ALMOST SHIPPED WITH. If the read-back dies — Excel COM falls over, the
    workbook will not open — there is no final_estimate, so every reconciliation check found
    nothing to complain about and the job came back ok: true. Nothing was wrong because
    nothing was looked at, and that is indistinguishable on a console from a job that
    reconciled. A check that could not run must FAIL CLOSED."""
    from invariants import check_job
    r = check_job({"part_estimates": [{"part_number": "01M"}]}, write_back=False)
    ok(r["unverified"] > 0, "checks with no data to read must report themselves unverified")
    ok(not r["verified"], "and the job as a whole is not verified")
    ok(not r["may_quote_firm"],
       "an unverified job must not be quotable as a firm price, whatever `ok` says")
    codes = [v["code"] for v in r["violations"]]
    ok(any(c.endswith("_not_evaluated") for c in codes), f"named as unevaluated: {codes}")

    # And the clean job must still be firm — fail-closed must not mean fail-always.
    ok(check_job(_job(), write_back=False)["may_quote_firm"],
       "a complete, consistent job is still releasable")


def test_the_quote_says_so_when_the_engine_cannot_stand_behind_it():
    """A gate nothing consumes is a log line. The quote was generated unconditionally and
    never read the invariant record, so a job with failing checks produced a document that
    looked exactly like one that passed."""
    from client_quote_html import _invariant_banner
    eq(_invariant_banner({"invariants": {"may_quote_firm": True}}), "",
       "a job that passes carries no banner")
    for job, why in (
        ({}, "no invariant record at all — the checks did not run"),
        ({"invariants": {"may_quote_firm": False, "violations": [
            {"severity": "blocking", "message": "material rows do not sum to the total"}]}},
         "a blocking failure"),
        ({"invariants": {"may_quote_firm": False, "violations": [
            {"severity": "unverified", "message": "read-back did not run"}]}},
         "an unverified job"),
    ):
        b = _invariant_banner(job)
        ok("PROVISIONAL" in b, f"{why} must be stated on the quote itself")
        ok("firm price" in b, f"{why} must say it is not a firm price")


def test_the_report_and_the_quote_never_disagree_about_the_same_job():
    """On 12120 the quote said "3 consistency check(s) FAILED" while the report, for the same
    job, said "Estimate completed cleanly ... no blocking data-sufficiency failure". Two
    documents describing one estimate and contradicting each other about whether it could be
    trusted — and the one an estimator reads first was the one saying everything was fine.

    They were answering different questions. estimate_status is DATA SUFFICIENCY: did the
    engine have enough to reach a number. The invariants are the later question: does that
    number hold together. Both are worth reporting; neither may be reported as the other."""
    from job_report_html import build_report_html
    from client_quote_html import _invariant_banner

    def _job_with(inv):
        j = {"estimate_summary": {
                "estimate_status": "ok",
                "workbook_equivalent_pricing": {"m105_total_unit_cost_gbp": 29.39,
                                                "m59_material_subtotal_gbp": 15.03,
                                                "m103_labour_subtotal_gbp": 12.30},
                "part_estimates": []}}
        if inv is not None:
            j["invariants"] = inv
        return j

    failing = _job_with({"may_quote_firm": False, "blocking": 3, "unverified": 0,
                         "checks_run": ["a"] * 11, "violations": [
                             {"severity": "blocking",
                              "code": "unit_price_does_not_equal_its_parts",
                              "message": "Material + labour does not equal the unit price."}]})
    rep = build_report_html(failing)
    ok("PROVISIONAL" in _invariant_banner(failing), "the quote says provisional")
    ok("PROVISIONAL and must not be released" in rep,
       "and the report must say the same, not 'completed cleanly'")
    ok("completed cleanly" not in rep,
       "a job with failing checks must never be described as having completed cleanly")
    ok("unit_price_does_not_equal_its_parts" in rep,
       "and the report names WHICH check failed — the quote can only carry a banner")

    passing = _job_with({"may_quote_firm": True, "blocking": 0, "unverified": 0,
                         "checks_run": ["a"] * 11, "violations": []})
    ok(_invariant_banner(passing) == "", "a clean job carries no banner")
    ok("PROVISIONAL" not in build_report_html(passing),
       "and its report is not marked provisional either")

    # Checks that never ran must not read as checks that passed, in EITHER document.
    unrun = _job_with(None)
    ok("PROVISIONAL" in _invariant_banner(unrun), "no invariant record: the quote says so")
    _r = build_report_html(unrun)
    ok("completed cleanly" not in _r,
       "and the report must not claim a clean completion it never verified")
    ok("did not run" in _r, "it must say the checks did not run")


def test_a_failing_check_names_the_records_it_objected_to():
    """"5 part(s) claim measured geometry but carry no usable outline" gives an estimator a
    number and nothing to act on — they cannot open five unnamed parts. Every check already
    collects which records it objected to; printing the message and discarding the evidence
    made the report describe a problem instead of locating it."""
    from job_report_html import build_report_html
    j = {"estimate_summary": {"estimate_status": "ok", "part_estimates": [],
         "workbook_equivalent_pricing": {"m105_total_unit_cost_gbp": 32.86,
                                         "m59_material_subtotal_gbp": 18.25,
                                         "m103_labour_subtotal_gbp": 12.30}},
         "invariants": {"may_quote_firm": False, "blocking": 1, "unverified": 0,
                        "checks_run": ["a"] * 12, "violations": [
             {"severity": "blocking", "code": "measured_geometry_without_outline",
              "message": "5 part(s) claim measured geometry but carry no usable outline.",
              "detail": {"count": 5, "parts": [
                  {"part_number": "12120-01-101", "geometry_source": "dxf_flat_pattern"},
                  {"part_number": "12120-01-103", "geometry_source": "dxf"}]}}]}}
    h = build_report_html(j)
    ok("12120-01-101" in h, "the report must name the parts a check objected to")
    ok("dxf_flat_pattern" in h, "and say what each of them claimed")
    ok("count=5" in h, "scalar detail is carried through too")


def test_the_customer_quote_carries_no_engineering_diagnostics():
    """The invariant banner was written for an estimator and rendered onto a document that
    leaves the building: "2 consistency check(s) FAILED", "the credibility gate has judged
    the measured coverage too low", "5 SolidWorks model file(s) ... were not read. Run
    tools/solidworks/sw_native_analyse.py". Naming an internal script on a customer's
    quotation is worse than saying nothing.

    THE WARNING IS NOT LOST, and that is the condition for removing it: the INTERNAL job
    report still leads with it and renders the full invariants section, and the price on
    the quote stays labelled indicative — so nothing on the page reads as firm."""
    from client_quote_html import build_quote_html

    _failing = {"estimate_summary": {}, "invariants": {"may_quote_firm": False,
                "violations": [
        {"severity": "blocking", "message":
         "5 SolidWorks model file(s) are in this job folder but were not read: run "
         "tools/solidworks/sw_native_analyse.py on a machine with SolidWorks."},
        {"severity": "blocking", "message":
         "The credibility gate has judged the measured coverage too low to stand behind."}]}}
    _html = build_quote_html(_failing, job_stem="2085 - SolidWorks")

    for _leak in ("consistency check", "credibility gate", "sw_native_analyse",
                  "SolidWorks model file", "invariant"):
        ok(_leak not in _html,
           f"the customer document must not carry {_leak!r}")

    # THE PRICE MUST STILL NOT READ AS FIRM. Without this the fixture would pass against a
    # version that simply presented an unverified number as a quotation.
    ok("indicative" in _html.lower(),
       "the price must still be labelled indicative on an unverified job")

    # AND THE ESTIMATOR MUST STILL BE TOLD, on the internal document.
    from job_report_html import _invariants_section
    _internal = _invariants_section(_failing)
    ok("sw_native_analyse" in _internal or "SolidWorks model file" in _internal,
       "the internal report keeps the detail the quote no longer carries")


def test_same_area_is_not_the_same_blank():
    """A 200 x 25 DXF and a 100 x 50 model flat have identical area. On area alone they
    'agree' — but they nest differently, may need a different stock width, and may not fit
    the same machine. Both SIDES have to agree."""
    from geometry_arbitration import arbitrate_flat
    v = arbitrate_flat(200.0, 25.0, 100.0, 50.0)
    ok(not v["agree"], "equal area with different sides is not agreement")
    ok(v["unreconciled"], "and neither measurement is obviously the broken one")
    # Orientation is a drawing convention, not a fact about the part.
    v = arbitrate_flat(82.2, 126.39, 126.39, 82.2)
    ok(v["agree"], "a transposed export is the same blank and must not be flagged")


def test_a_collapsed_polyline_is_not_a_cut_out():
    """A 100 x 0 closed polyline — a line drawn back on itself, which is what a collapsed or
    duplicated edge looks like — spans 100mm on its long side. Testing the LARGER dimension
    passes it straight through as an aperture the laser pierces."""
    try:
        import ezdxf
    except ImportError:
        print("      (skipped: ezdxf not installed)")
        return
    import tempfile
    from pathlib import Path
    from dxf_reader import extract_flat_pattern_data

    d = ezdxf.new(); d.header["$INSUNITS"] = 4
    msp = d.modelspace()
    L, W = 150.0, 80.0
    for a in [(0, 0, L, 0), (L, 0, L, W), (L, W, 0, W), (0, W, 0, 0)]:
        msp.add_line(a[:2], a[2:], dxfattribs={"layer": "SLD-0"})
    msp.add_circle((40, 40), 4.0, dxfattribs={"layer": "SLD-0"})       # one real hole
    # Long, but with no width at all.
    msp.add_lwpolyline([(20, 60), (120, 60), (120, 60), (20, 60)],
                       close=True, dxfattribs={"layer": "SLD-0"})
    path = os.path.join(tempfile.mkdtemp(), "collapsed.dxf")
    d.saveas(path)

    r = extract_flat_pattern_data(Path(path))
    eq(r.get("estimated_pierce_count"), 2,
       "1 hole plus the outer profile = 2; a zero-width loop is not an aperture")


def test_a_readback_that_found_nothing_still_says_why():
    """The condition was "we read some rows", which throws the record away in the one case it
    exists for: every header moved, no block yields rows, and adapter_problems holds the
    explanation. Discarded, that is indistinguishable from a read-back that never ran."""
    from wep_readback_from_xlsx import should_stamp_final_estimate as f
    ok(f({"labour_rows": [{"workbook_row": 41}], "material_rows": [], "adapter_problems": []}),
       "rows were read — stamp it")
    ok(f({"labour_rows": [], "material_rows": [], "adapter_problems": [
        {"block": "steel", "code": "header_row_not_found"}]}),
       "NO rows but a recorded reason is the case this exists for")
    ok(not f({"labour_rows": [], "material_rows": [], "adapter_problems": []}),
       "nothing read and nothing to report is genuinely empty")
    ok(not f(None), "no read-back at all")


def test_a_route_the_sheet_charged_nothing_for():
    """An unmapped department calculates to zero, reconciles perfectly against the total, and
    gives the work away. Rows at zero are skipped as 'not part of the priced job' — but an
    ACCEPTED route row at zero is the opposite: the engine decided this work happens."""
    from invariants import check_job
    j = _job()
    j["final_estimate"]["labour_rows"] = [
        {"workbook_row": 41, "operation": "Laser", "total_value_gbp": 52.0},
        {"workbook_row": 43, "operation": "Fold", "total_value_gbp": 0.0}]
    codes = [v["code"] for v in check_job(j, write_back=False)["violations"]]
    ok("accepted_route_priced_at_zero" in codes,
       f"work the engine routed but the sheet did not charge for must block: {codes}")


def test_geometry_the_engine_could_not_reconcile_blocks_a_firm_price():
    """arbitrate_flat marks an oversized DXF unreconciled, which was the right call — but a
    review flag alone changed nothing, so the 400x300-against-60x34 case could still leave as
    a firm quote."""
    from invariants import check_job
    j = _job()
    j["part_estimates"][0]["flat_unreconciled"] = True
    j["part_estimates"][0]["flat_arbitration"] = {"unreconciled": True, "reason": "much larger"}
    r = check_job(j, write_back=False)
    ok("geometry_unreconciled" in [v["code"] for v in r["violations"]],
       "an unresolved size disagreement must block")
    ok(not r["may_quote_firm"], "and must stop the price being firm")


def test_the_money_tolerance_does_not_scale_with_the_estimate():
    """A 0.5% relative tolerance passes a GBP 50 error on a GBP 10,000 job. The only
    legitimate slack is Excel's per-row rounding, which scales with ROW COUNT."""
    from invariants import _money_agrees
    ok(not _money_agrees(10000.00, 10050.00, 12), "GBP 50 out on a big job is an error")
    ok(_money_agrees(100.00, 100.03, 12), "twelve rows may each round by half a penny")
    ok(not _money_agrees(100.00, 100.50, 2), "two rows cannot account for 50p")


def test_attribution_is_checked_against_the_precedence_contract():
    """The checker invented its own source-field names, so a correctly attributed job warned
    on every part while thickness attribution went unchecked entirely."""
    from invariants import _source_key_for
    eq(_source_key_for("normalized_material"), "material_source",
       "the contract writes material_source, not normalized_material_source")
    eq(_source_key_for("quantity"), "quantity_source", "quantity")
    eq(_source_key_for("normalized_thickness_mm"), "thickness_source",
       "and thickness is normalized_thickness_mm -> thickness_source")


def test_a_broken_check_reports_that_it_verified_nothing():
    """A checker that throws must not read as a clean pass — that is worse than no check."""
    import invariants
    j = _job()
    r = invariants.check_job({"final_estimate": j["final_estimate"],
                              "workbook_labour": "not a dict at all"}, write_back=False)
    ok(isinstance(r["violations"], list), "a malformed job must not crash the checker")
    ok("checks_run" in r and r["checks_run"], "the checks it ran are reported")


# ── geometry arbitration — the 04M rule, keyed on geometry alone ─────────────────────
def test_an_incomplete_dxf_loses_to_the_model():
    """12120's 04M measured 43.00 x 20.04mm from a DXF whose outer profile was not in the
    file, against a model flat of 60.00 x 34.04mm — a quarter of the area, and it was costed.
    A developed blank cannot be SMALLER than the flat it develops: material is consumed going
    round a bend, never created. So a materially smaller DXF is missing geometry.

    The rule is geometric. It knows nothing about 04M, this job, or any part number."""
    from geometry_arbitration import arbitrate_flat, DXF, NATIVE

    v = arbitrate_flat(43.00, 20.04, 60.00, 34.04)
    eq(v["winner"], NATIVE, "an incomplete DXF must not be costed")
    ok(v["dxf_incomplete"], "and must be named as incomplete, not merely 'different'")
    ok("60" in v["reason"] and "43" in v["reason"], "both measurements quoted in the reason")

    # Two honest measurements of the same blank agree, and the DXF stays authoritative.
    v = arbitrate_flat(126.39, 82.2, 126.4, 82.0)
    eq(v["winner"], DXF, "agreement keeps the direct measurement of the file")
    ok(v["agree"] and not v["unreconciled"], "and is not flagged for review")

    # Larger is also a disagreement, but swapping in the model trades one unverified number
    # for another. Keep the DXF, mark it unreconciled, send it to a human.
    v = arbitrate_flat(400.0, 300.0, 60.0, 34.0)
    eq(v["winner"], DXF, "a too-large DXF is still the only direct measurement of the file")
    ok(v["unreconciled"], "but the two are unreconciled and must reach a person")

    # One-sided evidence is not a conflict, and must not be reported as agreement.
    eq(arbitrate_flat(None, None, 60.0, 34.0)["winner"], NATIVE, "model only")
    eq(arbitrate_flat(60.0, 34.0, None, None)["winner"], DXF, "DXF only")
    ok(not arbitrate_flat(None, None, None, None)["agree"],
       "no measurement at all is not agreement")


def test_degenerate_geometry_is_not_a_pierce():
    """Three ways of drawing the same thing must clear the same bar. A microscopic closed
    polyline is a duplicated vertex, not an aperture; a zero-length line chains into a
    self-loop whose one node has degree two, which is exactly the closed-contour test. Both
    invent pierces the laser is then charged for."""
    try:
        import ezdxf
    except ImportError:
        print("      (skipped: ezdxf not installed)")
        return
    import tempfile
    from pathlib import Path
    from dxf_reader import extract_flat_pattern_data

    d = ezdxf.new(); d.header["$INSUNITS"] = 4
    msp = d.modelspace()
    L, W = 150.0, 80.0
    for a in [(0, 0, L, 0), (L, 0, L, W), (L, W, 0, W), (0, W, 0, 0)]:
        msp.add_line(a[:2], a[2:], dxfattribs={"layer": "SLD-0"})
    msp.add_circle((40, 40), 4.0, dxfattribs={"layer": "SLD-0"})          # one real hole
    # A closed polyline spanning microns — a duplicated vertex from the export.
    msp.add_lwpolyline([(70, 40), (70.002, 40), (70.002, 40.002), (70, 40.002)],
                       close=True, dxfattribs={"layer": "SLD-0"})
    for i in range(4):                                                    # zero-length lines
        msp.add_line((100 + i, 40), (100 + i, 40), dxfattribs={"layer": "SLD-0"})
    path = os.path.join(tempfile.mkdtemp(), "degenerate.dxf")
    d.saveas(path)

    r = extract_flat_pattern_data(Path(path))
    eq(r.get("estimated_pierce_count"), 2,
       "1 real hole plus the outer profile = 2; degenerate loops are not apertures")
    eq(r.get("closed_contour_count"), 2, "and they are not contours either")


def test_the_complete_reader_wins_over_the_inflating_one():
    """The two DXF readers are not interchangeable. The flat reader walks topologically —
    blocks exploded, layers inherited, contours closed. The raw reader never enters a block
    and counts every short closed polyline as a pierce whether or not it is the outer
    profile, so it double-counts. Taking max() unconditionally hands the decision to
    whichever reader is more wrong upward.

    The raw figure is used only where the flat walk ADMITS it is incomplete."""
    from drawing_job_merge import _arbitrate_pierces
    v = _arbitrate_pierces({"estimated_pierce_count": 4, "pierce_count_incomplete": False},
                           {"estimated_pierce_count": 9})
    eq(v["value"], 4, "a complete topological walk is the answer, not a floor to be raised")
    ok(not v["uncertain"], "and it is a measurement, recorded as one")

    # An incomplete walk yields a FLOOR. Taking the larger of two unreliable readings is a
    # choice between guesses, and it must be recorded as a choice — value, source and
    # uncertainty kept apart — not laundered into a confident measured number.
    v = _arbitrate_pierces({"estimated_pierce_count": 4, "pierce_count_incomplete": True},
                           {"estimated_pierce_count": 9})
    eq(v["value"], 9, "where the walk could not close its loops, the larger is the floor")
    ok(v["uncertain"], "but it is NOT a measurement and must not be presented as one")
    ok("floor" in str(v.get("note", "")).lower(), "and says so in words a person can read")

    v = _arbitrate_pierces({"estimated_pierce_count": 0, "pierce_count_incomplete": False},
                           {"estimated_pierce_count": 6})
    eq(v["value"], 6, "a reader that saw nothing at all defers to one that saw something")
    ok(v["uncertain"], "the raw parser's known upward bias makes that figure provisional")
    eq(_arbitrate_pierces({}, {}), None, "no evidence is None, never 0")


def test_zero_cut_outs_is_evidence_not_absence():
    """A cut list reporting ZERO cut-outs has said something definite: a plain blank, one
    outer profile, one pierce — and the model is the strongest source there is. Reading 0 as
    'no data' let a weaker PDF count survive against explicit model evidence, which is the
    silent-overwrite failure running the other way."""
    from source_connectors.solidworks import (normalize_native_extract,
                                              apply_native_to_pre_estimate)
    recs = [{"title": "AAA-01M", "doctype": 1, "route_signals": {
        "material": "Mild Steel [CR4]", "is_sheet_metal": True, "bend_count": 1,
        "flat_length_mm": 100.0, "flat_width_mm": 50.0, "thickness_mm": 1.5,
        "cut_out_count": 0, "bbox_mm": [60.0, 50.0, 20.0]}}]
    job = normalize_native_extract(recs)
    parts = [{"part_number": "AAA-01M",
              "manufacturing_features": {"pierce_count": 5},      # a PDF-derived guess
              "geometry_rollup": {"estimated_pierce_count": 5}}]
    apply_native_to_pre_estimate(parts, job)
    eq((parts[0].get("geometry_rollup") or {}).get("estimated_pierce_count"), 1,
       "no cut-outs means one pierce for the outer profile — not the PDF's 5")
    eq((parts[0].get("manufacturing_features") or {}).get("cut_out_count"), 0,
       "and zero is recorded as a value, not left absent")


def test_route_group_id_is_stable_and_distinguishes_gauges():
    """Row numbers move when the template changes; a group's identity does not. Without a
    stable id, two runs of the same job cannot be compared row for row."""
    from wb_populate import route_group_id
    a = route_group_id("Fold", "Mild Steel", 1.2, ["01M", "06M"])
    b = route_group_id("Fold", "Mild Steel", 1.2, ["06M", "01M"])   # order must not matter
    c = route_group_id("Fold", "Mild Steel", 1.5, ["04M"])
    eq(a, b, "the same group must have the same id whatever order its parts arrive in")
    ok(a != c, "two gauges of the same operation are different groups")
    ok(a.startswith("rg_"), "ids are recognisable on sight")


# ── runner ───────────────────────────────────────────────────────────────────────────
def test_a_measured_cut_path_is_not_a_measured_blank():
    """Five parts on 12120 blocked as "measured geometry without an outline" for run after
    run. The cause was one flag carrying two different claims.

    A DXF can yield a measured cut PATH without yielding a measured BLANK — the cut layer
    holds geometry the reader can follow, but nothing closing into an outline with an area.
    Both are real measurement and they license completely different things: knowing the
    blank means no blank allowance is needed; knowing the cut length says nothing whatever
    about the blank. dxf_measured_outline was set from EITHER, so those five skipped their
    blank allowance on the strength of a measurement about something else.

    bought_in_policy already documented the contract this broke — "dxf_augmented is set ONLY
    where an outline was actually measured" — so the reader was right and the writer wrong.

    Drives the merge, not the flags: it is the writer that was wrong."""
    import drawing_job_merge as djm
    from pathlib import Path

    _saved_geom, _saved_flat = djm.build_geometry_summary_for_dxf, djm.extract_flat_pattern_data
    try:
        # A cut layer the reader could follow, that never closed into an outline.
        djm.build_geometry_summary_for_dxf = lambda p: (
            {}, {"estimated_cut_length_mm": 842.0, "drawing_extents_mm": []}, 0.5)
        djm.extract_flat_pattern_data = lambda p: {"flat_pattern_detected": False}
        part = {"part_number": "12120-01-05M"}
        djm.apply_dxf_geometry_to_part(part, Path("05M_1.5mm_MILD_STEEL.DXF"))

        eq(part["dxf_measured_cut_length"], True, "the cut path was measured, and says so")
        eq(part["dxf_measured_outline"], False, "the blank was not, and says that too")
        eq(part["dxf_augmented"], False,
           "so the blank extents are not trusted and the allowance applies")
        eq(part["geometry_source"], "dxf_cut_length_only",
           "and the source names exactly what was read")
        ok(any("blank allowance" in str(f) for f in part.get("review_flags") or []),
           "with a flag saying so in words")

        # It is still a part we fabricate — something is cutting that path.
        from bought_in_policy import has_fabrication_evidence
        eq(has_fabrication_evidence(part), True,
           "a measured cut path is still evidence we make it")

        # And the invariant stops objecting, because the claim is now true.
        from invariants import check_job
        j = _job()
        j["part_estimates"] = [dict(part, blank_length_mm=None, blank_width_mm=None,
                                    quantity=1, quantity_source="bom_tree")]
        ok("measured_geometry_without_outline" not in
           [v["code"] for v in check_job(j, write_back=False)["violations"]],
           "an honest cut-length-only part is not a false claim")

        # A real flat pattern still claims the blank, and still must carry one.
        # Zero for anything the reader asks for and this stub has not named: the point here
        # is which flags come out, not to re-declare the flat-pattern contract field by field.
        class _Flat(dict):
            def __missing__(self, k):
                return 0.0
        djm.extract_flat_pattern_data = lambda p: _Flat(
            flat_pattern_detected=True, blank_area_mm2=10389.3, blank_length_mm=126.39,
            blank_width_mm=82.2, perimeter_mm=842.0, weight_kg=0.12, hole_count=4,
            bend_count=2, hole_diameters_mm=[])
        whole = {"part_number": "12120-01-01M"}
        djm.apply_dxf_geometry_to_part(whole, Path("01M_1.5mm_MILD_STEEL.DXF"))
        eq(whole["dxf_measured_outline"], True, "a measured blank still claims one")
        eq(whole["dxf_augmented"], True, "and is still trusted for extents")
    finally:
        djm.build_geometry_summary_for_dxf = _saved_geom
        djm.extract_flat_pattern_data = _saved_flat


def test_a_blank_is_found_wherever_the_writer_put_it():
    """Four parts on 12120 blocked the quote as "claims measured geometry but carries no
    usable outline" while their blanks sat on the populated sheet in front of the estimator:
    126.39 x 82.2, 45 x 20, 33.3 x 27.8, 79 x 37.79. All four read geometry_source
    dxf_flat_pattern, which by construction requires a measured area.

    drawing_job_merge writes a measured flat pattern to part["normalized_geometry"] and
    mirrors the extents to overall_length_mm / overall_width_mm. It does not write
    blank_length_mm to the part root — and the root plus geometry_rollup were the only two
    places the check looked.

    A false positive here is not harmless: it blocks a firm quote, and it sent a defect hunt
    after the wrong cause for several runs."""
    from invariants import check_job

    def _codes(part):
        j = _job()
        j["part_estimates"] = [dict(part, part_number="12120-01-01M", quantity=1,
                                    quantity_source="bom_tree",
                                    normalized_material="MILD_STEEL",
                                    material_source="dxf")]
        return [v["code"] for v in check_job(j, write_back=False)["violations"]]

    measured = {"geometry_source": "dxf_flat_pattern", "dxf_measured_outline": True,
                "flat_pattern_detected": True,
                "overall_length_mm": 126.39, "overall_width_mm": 82.2,
                "normalized_geometry": {"blank_length_mm": 126.39, "blank_width_mm": 82.2,
                                        "blank_area_mm2": 10389.3,
                                        "geometry_source": "dxf_flat_pattern"}}
    ok("measured_geometry_without_outline" not in _codes(measured),
       "a blank in normalized_geometry is a blank")

    # 12120's DXF-sourced parts carry their blank ONLY as overall_length_mm /
    # overall_width_mm and never write blank_area_mm2 — 01M is 126.393 x 82.197 with area
    # None. Four parts failed for a field that was merely never written, while the two with
    # native flats passed, which is what gave the game away. Two extents ARE an outline.
    extents_only = {"geometry_source": "dxf_flat_pattern", "dxf_measured_outline": True,
                    "flat_pattern_detected": True,
                    "overall_length_mm": 126.393, "overall_width_mm": 82.197}
    ok("measured_geometry_without_outline" not in _codes(extents_only),
       "extents with no stored area are still an outline")

    # One extent is not. A part with a length and no width has not been measured.
    half = {"geometry_source": "dxf_flat_pattern", "dxf_measured_outline": True,
            "overall_length_mm": 126.393}
    ok("measured_geometry_without_outline" in _codes(half),
       "and half an outline is still caught")

    # The check must still catch the thing it exists for: a measurement claim with no
    # measurement behind it, in any of the three places.
    empty = {"geometry_source": "dxf_flat_pattern", "dxf_measured_outline": True,
             "normalized_geometry": {"geometry_source": "dxf_flat_pattern"}}
    ok("measured_geometry_without_outline" in _codes(empty),
       "and an empty claim is still caught")

    # geometry_rollup remains a valid home for it.
    rolled = {"geometry_source": "dxf_flat_pattern", "dxf_measured_outline": True,
              "geometry_rollup": {"blank_length_mm": 45.0, "blank_width_mm": 20.0,
                                  "blank_area_mm2": 900.0}}
    ok("measured_geometry_without_outline" not in _codes(rolled),
       "wherever it was written, it counts")


def test_a_guessed_price_says_so_on_the_estimating_sheet():
    """The supplier column was blank on every BOM line, so a catalogue price and an AI market
    estimate looked identical to the estimator reading the workbook.

    On 12120 that mattered: the three guessed lines were 96% of material cost, and the
    knurled knob was quoted at GBP 1.25, GBP 11.52 and GBP 1.77 across three runs of
    identical inputs. A figure nobody can reproduce must not sit in a price column looking
    like a quote.

    Keyed on the source CLASS, never on a part code, so a job nobody has seen yet is labelled
    by the same rule."""
    from wb_populate import _price_origin, _INDICATIVE_TAG

    def _line(source_name, applied=True):
        return {"part_number": "BI-KNURLEDKNOB", "quantity": 2, "unit_cost_gbp": 1.77,
                "cost_breakdown": {"system_cost": {
                    "applied_to_total": applied,
                    "source": {"source_name": source_name, "applied": True,
                               "affects_total": applied, "source_rank": 0,
                               "selected": {"source": source_name, "price": 1.77}}}}}

    label, indicative = _price_origin(_line("llm_market_estimate"))
    eq(indicative, True, "a guessed line is marked indicative")
    ok("INDICATIVE" in label, f"and named as one in the supplier column (got {label!r})")

    # NAMED, not just categorised. "An AI estimate" is a category; "xAI Grok" answers the
    # question an estimator actually asks about a number they cannot reproduce. The lookup
    # records the provider, so a switch needs no edit here.
    _named = _line("llm_market_estimate")
    _named["cost_breakdown"]["system_cost"]["source"]["llm_provider"] = "xai"
    label, _ = _price_origin(_named)
    ok("Grok" in label, f"the engine that answered is named (got {label!r})")

    _unknown = _line("llm_market_estimate")
    _unknown["cost_breakdown"]["system_cost"]["source"]["llm_provider"] = "somenewvendor"
    label, _ = _price_origin(_unknown)
    ok("somenewvendor" in label,
       f"a provider this map has not met still names itself (got {label!r})")

    label, indicative = _price_origin(_line("udef_sqlserver"))
    eq(indicative, False, "a catalogue line is not")
    eq(label, "", "and needs no warning next to it")

    # A price that was resolved but never added cannot mislead anyone about the total.
    _, indicative = _price_origin(_line("llm_market_estimate", applied=False))
    eq(indicative, False, "a resolved-but-unapplied guess does not tag the line")

    # The tag has to be readable as what it is, in a narrow cell, by someone skim-reading.
    ok("NOT A QUOTE" in _INDICATIVE_TAG and "INDICATIVE" in _INDICATIVE_TAG,
       f"the tag says plainly what it is: {_INDICATIVE_TAG!r}")

    # The cell is truncated to 120 characters on write and the tag is appended last, so on a
    # long description the warning is the first thing that would be cut. It must survive.
    _long = "BI-SCREENCABLE  " + ("SCREEN CABLE ASSEMBLY WITH MOULDED BOOT AND STRAIN RELIEF " * 3)
    _room = 120 - len(_INDICATIVE_TAG) - 2
    _cell = f"{_long[:_room].rstrip()}  {_INDICATIVE_TAG}"[:120]
    ok(_cell.endswith(_INDICATIVE_TAG),
       f"the warning survives truncation on a long description (got ...{_cell[-40:]!r})")
    ok(len(_cell) <= 120, "and the cell still fits")


def test_a_plate_modelled_as_an_extrude_is_still_a_plate():
    """04M stayed in the Fold 1.5mm group across every run, and the fixture above passed the
    whole time — because it declares is_sheet_metal True and a 1.5mm cut-list thickness. The
    real part has neither.

    is_sheet_metal is true only where the feature tree holds a sheet-metal feature:
    SMBaseFlange, an EdgeFlange, a SketchBend. Nobody models a flat plate that way; you
    extrude it. Both the analyser's last-resort thickness and the connector's fold gate were
    keyed on that flag, so the parts the gate exists to catch were exactly the ones it could
    not see.

    Nothing about the physics changed: a part one thickness thick has nowhere for a bend to
    be, however it was drawn."""
    from source_connectors.solidworks import (normalize_native_extract,
                                              apply_native_to_pre_estimate)

    # An extruded plate, as SolidWorks actually reports one: no sheet-metal feature, and a
    # thickness the analyser had to infer from the solid.
    job = normalize_native_extract([{"title": "04M", "doctype": 1, "route_signals": {
        "material": "Mild Steel [CR4]", "is_sheet_metal": False, "bend_count": 0,
        "thickness_mm": 1.5, "bbox_mm": [60.0, 34.04, 1.5]}}])
    parts = [{"part_number": "04M", "textual_operations": ["laser_cutting", "folding"],
              "manufacturing_features": {"bend_count": 1}}]
    apply_native_to_pre_estimate(parts, job)
    ok("folding" not in (parts[0].get("textual_operations") or []),
       "a plate is not folded just because it was extruded rather than flanged")
    eq(parts[0].get("native_flat_solid"), True,
       "and the verdict is durable, so a later drawing-text pass cannot re-add the fold")

    # A formed part with no countable bends must still fold — that is the case zero-bends
    # alone can never be trusted for, and dropping is_sheet_metal must not weaken it.
    formed = normalize_native_extract([{"title": "06M", "doctype": 1, "route_signals": {
        "material": "Mild Steel [CR4]", "is_sheet_metal": False, "bend_count": 0,
        "thickness_mm": 1.2, "bbox_mm": [96.49, 39.09, 35.0],
        "formed_but_no_bend_features": True}}])
    fparts = [{"part_number": "06M", "textual_operations": ["laser_cutting", "folding"]}]
    apply_native_to_pre_estimate(fparts, formed)
    ok("folding" in (fparts[0].get("textual_operations") or []),
       "a formed part still folds")
    ok(not fparts[0].get("native_flat_solid"), "and is never called a plate")

    # A solid that stands taller than its material folds, whatever its feature tree says.
    tall = normalize_native_extract([{"title": "01M", "doctype": 1, "route_signals": {
        "material": "Mild Steel [CR4]", "is_sheet_metal": False, "bend_count": 0,
        "thickness_mm": 1.5, "bbox_mm": [79.0, 64.5, 21.5]}}])
    tparts = [{"part_number": "01M", "textual_operations": ["laser_cutting", "folding"]}]
    apply_native_to_pre_estimate(tparts, tall)
    ok("folding" in (tparts[0].get("textual_operations") or []),
       "21.5mm of envelope on 1.5mm material is not a plate")


def test_a_thickness_is_inferred_for_a_plate_with_no_sheet_metal_feature():
    """The analyser's last-resort thickness — the smallest bbox dimension of an unformed
    blank — was gated on is_sheet_metal too, so an extruded plate reached the connector with
    no thickness at all and could not pass the flat-solid test whatever the gate allowed.

    A part with no sheet-metal feature has to earn it on shape: a 10mm pin measuring
    10 x 10 x 30 also has a small minimum, and is not 10mm thick.

    Drives the analyser's own function. The first version of this fixture restated the rule
    inline and asserted against its own copy, which proves only that the test agrees with
    itself — the same mistake this suite keeps having to relearn."""
    _infer = _load_analyser().infer_thickness_from_bbox

    eq(_infer([60.0, 34.04, 1.5], False), 1.5, "an extruded plate gets its thickness")
    eq(_infer([10.0, 10.0, 30.0], False), None, "a pin does not")
    eq(_infer([79.0, 64.5, 21.5], False), None, "nor does a formed envelope")
    eq(_infer([60.0, 34.04, 1.5], False, bend_count=2), None,
       "nor does anything with bends — its envelope is not its material")
    eq(_infer([54.7, 45.0, 1.5], True), 1.5, "and sheet-metal parts are unaffected")
    eq(_infer([60.0, 34.04], False), None, "a two-dimensional box says nothing about a solid")
    eq(_infer(None, False), None, "and neither does no box at all")
    eq(_infer([600.0, 340.0, 20.0], False), None,
       "20mm is plate-shaped but out of sheet range — that is a fabrication, not a gauge")


def test_a_measured_zero_bend_count_outvotes_a_drawing_callout():
    """04M carried native_flat_solid=True and inferred_operations=['folding'] at the same
    time, and GBP 0.16 of folding on its priced row. The op was stripped at extraction,
    stripped again at costing, and came back both times.

    The cause is the engine's own rule broken in one line:

        bends = manufacturing_features.get("bend_count") or max(len(angles_deg), ...)

    bend_count = 0 stamped by solidworks_api is a MEASUREMENT — the model was read and has no
    bends. Zero is falsy, so it fell through to the PDF's 30-degree callout. Removing the op
    in more places would be whack-a-mole; the evidence that regenerates it has to stop
    outvoting the measurement.

    Falling through is still right where the zero came from something that cannot see bends —
    parts whose folds live only in a PDF callout must keep folding."""
    from estimator import _model_measured_zero_bends, estimate_process_times

    measured_flat = {"part_number": "04M", "normalized_material": "MILD_STEEL",
                     "native_flat_solid": True, "angles_deg": ["30"],
                     "normalized_thickness_mm": 1.5,
                     "manufacturing_features": {"bend_count": 0,
                                                "bend_count_source": "solidworks_api"},
                     "textual_operations": ["laser_cutting", "powder_coating", "handling"]}
    eq(_model_measured_zero_bends(measured_flat), True,
       "a zero from the model is a measurement")

    # The whole point: drive costing and check no fold is priced.
    _res = estimate_process_times(dict(measured_flat), quantity=1)
    _ops = list((_res or {}).get("times_min", {}) or (_res or {}).get("unit_times_min", {}) or {})
    ok(not [o for o in _ops if "fold" in str(o).lower()],
       f"a measured plate is not costed to fold (ops priced: {_ops})")

    # A zero from something that cannot see bends must STILL fall through — this is the
    # fold-shadowing fix, and breaking it would drop real folds off PDF-only parts.
    pdf_only = {"part_number": "09M", "normalized_material": "MILD_STEEL",
                "angles_deg": ["90"], "normalized_thickness_mm": 1.5,
                "manufacturing_features": {"bend_count": 0},
                "textual_operations": ["laser_cutting"]}
    eq(_model_measured_zero_bends(pdf_only), False,
       "a bare zero with no source is not a measurement")
    _res2 = estimate_process_times(dict(pdf_only), quantity=1)
    _ops2 = list((_res2 or {}).get("times_min", {}) or (_res2 or {}).get("unit_times_min", {}) or {})
    ok([o for o in _ops2 if "fold" in str(o).lower()],
       f"a PDF-only fold is still costed (ops priced: {_ops2})")

    # Where the drawing says fold and the model counted no bends, the op stands — someone
    # wrote it on the drawing — but the COUNT used to time it comes from the measurement, not
    # from however many angles the PDF happens to carry. Two callouts would otherwise buy
    # press-brake time for bends the model says are not there.
    contested = {"part_number": "07M", "normalized_material": "MILD_STEEL",
                 "normalized_thickness_mm": 1.5, "angles_deg": ["30", "90"],
                 "textual_operations": ["laser_cutting", "folding"],
                 "manufacturing_features": {"bend_count": 0,
                                            "bend_count_source": "solidworks_api"}}
    _t = estimate_process_times(dict(contested), quantity=1)
    _tm = (_t or {}).get("times_min") or (_t or {}).get("unit_times_min") or {}
    _unsourced = dict(contested, manufacturing_features={"bend_count": 0})
    _t2 = estimate_process_times(_unsourced, quantity=1)
    _tm2 = (_t2 or {}).get("times_min") or (_t2 or {}).get("unit_times_min") or {}
    ok(_tm.get("folding", 0) < _tm2.get("folding", 0),
       f"a measured zero times fewer bends than two PDF callouts "
       f"({_tm.get('folding')} vs {_tm2.get('folding')} min)")

    # An absent count is absence, not a measured zero.
    eq(_model_measured_zero_bends({"manufacturing_features": {
        "bend_count": None, "bend_count_source": "solidworks_api"}}), False,
       "no count read is not a count of none")
    # A real count is a real count.
    eq(_model_measured_zero_bends({"manufacturing_features": {
        "bend_count": 2, "bend_count_source": "solidworks_api"}}), False,
       "two bends is not zero bends")


def test_a_plate_is_never_charged_to_fold():
    """The op was removed in three places and the money still reached the sheet, because
    each strip ran before the next pass re-inferred the fold. Removing it in a fourth place
    is another chance to miss it.

    This asks the only question that settles it, at the point where it is settled: is the
    engine CHARGING to fold something it measured as flat? Any future path that resurrects
    the op fails here, whatever route it took."""
    from invariants import check_job

    j = _job()
    j["parts"] = [{"part_number": "12120-01-04M", "native_flat_solid": True,
                   "geometry_source": "dxf_cut_length_only"}]
    j["estimate_summary"] = {"part_estimates": [
        {"part_number": "12120-01-04M",
         "labour_estimate": {"costs_gbp": {"laser_cutting": 0.68, "folding": 0.16}}}]}
    r = check_job(j, write_back=False)
    codes = [v["code"] for v in r["violations"]]
    ok("plate_charged_for_folding" in codes,
       f"charging a measured plate to fold must block: {codes}")
    ok(not r["may_quote_firm"], "and stop the quote being firm")
    _v = next(v for v in r["violations"] if v["code"] == "plate_charged_for_folding")
    ok("12120-01-04M" in str(_v["detail"]), "naming the part and the money")

    # The verdict and the money live on different records for the same part; the check has
    # to join them, which is why 04M read native_flat_solid=None on its costed record.
    j2 = _job()
    j2["parts"] = [{"part_number": "12120-01-04M", "native_flat_solid": True}]
    j2["estimate_summary"] = {"part_estimates": [
        {"part_number": "12120-01-04M",
         "cost_breakdown": {"labour": {"costs_gbp": {"folding": 0.16}}}}]}
    ok("plate_charged_for_folding" in
       [v["code"] for v in check_job(j2, write_back=False)["violations"]],
       "wherever the cost was written, it is found")

    # A part that genuinely folds is untouched.
    j3 = _job()
    j3["parts"] = [{"part_number": "12120-01-01M"}]
    j3["estimate_summary"] = {"part_estimates": [
        {"part_number": "12120-01-01M",
         "labour_estimate": {"costs_gbp": {"folding": 1.96}}}]}
    ok("plate_charged_for_folding" not in
       [v["code"] for v in check_job(j3, write_back=False)["violations"]],
       "a part that folds is charged to fold")


def test_the_report_does_not_call_a_guess_a_catalogue_price():
    """Under a green "Sound" tag the report said every bought-in was "identified and priced
    from catalogue/historical sources" — on a job where three of those prices were AI market
    estimates and a fourth was zero. Recognising a part and pricing it are different
    achievements, and only the first had gone right.

    The first version of this fixture checked whether a helper existed and passed when it did
    not, asserting nothing at all. The row builder was extracted so this can drive it."""
    from job_report_html import bought_in_strength_row

    def _bi(pn, price, ai=False):
        src = {"source_name": "llm_market_estimate" if ai else "udef_sqlserver",
               "applied": True, "affects_total": True, "source_rank": 0,
               "selected": {"source": "llm_market_estimate" if ai else "udef_sqlserver",
                            "price": price}}
        return {"part_number": pn, "unit_cost_gbp": price,
                "cost_breakdown": {"system_cost": {"unit_cost_gbp": price,
                                                   "applied_to_total": True, "source": src}}}

    clean = bought_in_strength_row([_bi("BI-SELFCLINCHNUT", 0.03)])
    ok("Sound" in clean, "catalogue-priced bought-ins still read as sound")
    ok("AI market estimate" not in clean, "with nothing to warn about")

    mixed = bought_in_strength_row([_bi("BI-SELFCLINCHNUT", 0.03),
                                    _bi("BI-KNURLEDKNOB", 9.52, ai=True),
                                    _bi("BI-PEMSTUD", 0.0)])
    ok("Sound" not in mixed, "a mixed set is not reported as sound")
    ok("1 priced by an AI market estimate" in mixed, "the guessed line is counted and named")
    ok("1 carrying no price at all" in mixed, "and so is the unpriced one")
    ok("Identification is not pricing" in mixed, "with the distinction spelled out")

    eq(bought_in_strength_row([]), "", "and no bought-ins says nothing at all")


def test_the_report_counts_bought_ins_the_same_way_the_checks_do():
    """The report said 4 bought-ins with 2 AI-priced, and the invariant reading the same job
    said 3 AI-priced. Both were right about what they looked at: the report used a local
    "starts with BI-" test, and THUM620 does not start with BI- — while bought_in_policy has
    listed THUM as a bought-in family all along.

    Two counts of the same thing on one page is exactly the defect bought_in_policy was
    written to end. One module answers the question."""
    import job_report_html as jr
    from bought_in_policy import is_bought_in

    thumbscrew = {"part_number": "THUM620", "description": "M4x10mm MUSHROOM THUMBSCREW"}
    eq(is_bought_in(thumbscrew), True, "the policy knows a thumbscrew is bought in")

    _src = open(jr.__file__, encoding="utf-8").read()
    ok("from bought_in_policy import is_bought_in" in _src,
       "and the report asks the policy rather than testing the prefix itself")
    _seg = _src[_src.index("# ONE MODULE ANSWERS"):]
    _seg = _seg[:_seg.index("rows +=")]
    ok('startswith("BI-")' not in _seg.split("except ImportError:")[0],
       "with the prefix left only as an import fallback")


def test_reproducible_and_firm_are_different_questions():
    """This module answered one question where there are two. A public distributor list price
    repeats perfectly, every run, forever — and it is not a quote: no contract behind it, no
    validity date, no commitment to honour it. Treating reproducible as sufficient would let
    a list price sit on a customer quote looking exactly like a negotiated one.

        reproducible — same inputs, same answer
        firm         — we will stand behind it, and it has not expired

    Nothing here judges a supplier; it classifies what KIND of thing a price is, so a
    supplier nobody has onboarded yet is judged by the same rule."""
    import price_provenance as pp

    def _blk(src, **kw):
        b = {"source_name": src, "applied": True, "affects_total": True, "source_rank": 0,
             "selected": {"source": src, "price": 1.0}}
        b.update(kw)
        return b

    eq(pp.source_class_of("udef_sqlserver"), pp.CONTRACT, "UDEF is an agreed rate")
    eq(pp.source_class_of("historical_quote_material_line"), pp.PURCHASE_HISTORY,
       "a historical quote line is what we last paid")
    eq(pp.source_class_of("estimating_supplier_catalog_url"), pp.CATALOGUE,
       "a supplier's published list is a catalogue")
    eq(pp.source_class_of("argus_crc_index"), pp.COMMODITY_INDEX, "an index is a benchmark")
    eq(pp.source_class_of("llm_market_estimate"), pp.AI_ESTIMATE, "and a guess is a guess")

    # A catalogue price is reproducible and never firm — that is the whole distinction.
    cat = pp.price_firmness(_blk("estimating_supplier_catalog_url"), today="2026-07-29")
    eq((cat["reproducible"], cat["firm"]), (True, False),
       "a list price repeats and still commits nobody")
    ok("commitment" in cat["reason"], "and says why, not just that it failed")

    # An agreed rate is firm only while it is in date.
    live = pp.price_firmness(_blk("udef_sqlserver", price_valid_to="2026-12-31"),
                             today="2026-07-29")
    eq(live["firm"], True, "an unexpired contract price is firm")
    dead = pp.price_firmness(_blk("udef_sqlserver", price_valid_to="2026-01-01"),
                             today="2026-07-29")
    eq(dead["firm"], False, "an expired one is not")
    ok("expired on 2026-01-01" in dead["reason"], "naming the date it lapsed")

    # THE STATE OF THE ENGINE TODAY: no source carries a validity date, so every line is
    # unfirmable. That is the honest answer, and it is why the check reports rather than
    # blocks — a gate that fails every job on every run is one people learn to click past.
    bare = pp.price_firmness(_blk("udef_sqlserver"), today="2026-07-29")
    eq((bare["reproducible"], bare["firm"]), (True, False),
       "a contract price with no validity date cannot be shown to be firm")
    ok("price_valid_to" in bare["reason"], "and names the missing field, so it can be asked for")


def test_a_price_nobody_committed_to_is_reported_not_assumed():
    """The firmness verdict has to reach a human, grouped by what is actually missing —
    "not firm" on its own tells an estimator nothing they can act on."""
    from invariants import check_job

    def _line(pn, src, **kw):
        src_blk = {"source_name": src, "applied": True, "affects_total": True,
                   "source_rank": 0, "selected": {"source": src, "price": 1.0}}
        src_blk.update(kw)
        return {"part_number": pn, "cost_breakdown": {"system_cost": {
            "applied_to_total": True, "source": src_blk}}}

    j = _job()
    j["estimate_summary"] = {"part_estimates": [
        _line("A", "udef_sqlserver"),
        _line("B", "estimating_supplier_catalog_url"),
        _line("C", "udef_sqlserver", price_valid_to="2099-12-31")]}
    r = check_job(j, write_back=False)
    v = next((x for x in r["violations"] if x["code"] == "price_not_firm"), None)
    ok(v is not None, "lines nobody committed to are reported")
    eq(v["severity"], "warning",
       "as a warning while no source carries validity — a gate that fails everything is noise")
    _named = {str(l.get("part")) for l in v["detail"]["lines"]}
    ok("C" not in _named, f"the in-date contract line is not implicated (named: {_named})")
    ok({"A", "B"} <= _named, f"the two that cannot be stood behind are (named: {_named})")
    ok(len(v["detail"]["reasons"]) == 2,
       f"and the reasons are distinguished, not lumped: {v['detail']['reasons']}")
    ok(r["ok"], "a warning does not make the job wrong")

    # A LABOUR RATE IS NOT A SUPPLIER PRICE. Every operation on every part stamps its rate,
    # so on the live 12120 run this reported 83 lines of which 80 were labour — burying the
    # three an estimator can act on. Firmness asks whether someone OUTSIDE SDI has committed
    # to a price; our own rate card is a different question.
    j2 = _job()
    j2["estimate_summary"] = {"part_estimates": [
        _line("A", "udef_sqlserver"),
        {"part_number": "A", "labour_estimate": {"rate_sources": {
            "folding": {"source_name": "sqlserver", "applied": True, "affects_total": True,
                        "source_rank": 0,
                        "selected": {"source": "sqlserver", "kind": "labour_rate",
                                     "price": 40.47}}}}}]}
    v2 = next((x for x in check_job(j2, write_back=False)["violations"]
               if x["code"] == "price_not_firm"), None)
    ok(v2 is not None, "material lines are still reported")
    ok(not [l for l in v2["detail"]["lines"] if "rate_sources" in str(l.get("where"))],
       f"but labour rates are not counted: {v2['detail']['count']} line(s)")


def test_a_failing_check_shows_its_lines_not_just_a_count():
    """The firmness advisory rendered as "count=83" and nothing else, because the report's
    detail builder only knew the key names other checks happened to use. A verdict an
    estimator cannot act on is the exact failure section 8 exists to prevent."""
    import job_report_html as jr
    _src = open(jr.__file__, encoding="utf-8").read()
    _seg = _src[_src.index("_DETAIL_LISTS = ("):]
    _seg = _seg[:_seg.index("_detail = (")]
    ok('"lines"' in _seg, "a check reporting `lines` has its records rendered")
    ok('_item.get("part")' in _seg, "and a line named by `part` is labelled, not blanked")
    ok('_item.get("reason")' in _seg, "with the reason it objected")


def test_firmness_severity_follows_intent_and_coverage():
    """A single global constant flipped when the first supplier feed lands would claim every
    other material was integrated too. Firm pricing arrives one supplier at a time.

        indicative                          -> warning, whatever the source
        firm intent, class has a connector  -> BLOCKING: the price is missing or stale
        firm intent, class has none         -> BLOCKING: nothing could have been firm

    Those are different failures and an estimator can act on each one differently."""
    from invariants import check_job
    import config

    def _steel_line(pn, src):
        return {"part_number": pn, "normalized_material": "MILD_STEEL",
                "cost_breakdown": {"system_cost": {"applied_to_total": True, "source": {
                    "source_name": src, "applied": True, "affects_total": True,
                    "source_rank": 0, "selected": {"source": src, "price": 1.0}}}}}

    def _job_with(intent):
        j = _job()
        j["quote_intent"] = intent
        j["estimate_summary"] = {"part_estimates": [_steel_line("01M", "udef_sqlserver")]}
        return j

    codes = lambda j: [v["code"] for v in check_job(j, write_back=False)["violations"]]

    ind = check_job(_job_with("indicative"), write_back=False)
    _v = next(v for v in ind["violations"] if v["code"] == "price_not_firm")
    eq(_v["severity"], "warning", "an indicative estimate is not lying by using a list price")
    ok(ind["ok"], "and the job is not marked wrong")
    ok("indicative estimate" in _v["message"], "the message says which kind of job this is")

    # Firm intent, and no connector exists for sheet steel yet.
    _saved = dict(getattr(config, "FIRM_PRICING_COVERAGE", {}) or {})
    try:
        firm = check_job(_job_with("firm"), write_back=False)
        ok("no_firm_pricing_source" in [v["code"] for v in firm["violations"]],
           "a firm quote on a class with no connector is refused for that reason")
        ok(not firm["ok"], "and blocks")
        _n = next(v for v in firm["violations"] if v["code"] == "no_firm_pricing_source")
        ok("Uptonsteel" in _n["message"],
           "naming the source that would make it firm, so the gap is actionable")

        # Same job once sheet steel HAS a firm-capable connector: the complaint changes from
        # "nothing could have been firm" to "this particular price is not".
        config.FIRM_PRICING_COVERAGE = dict(_saved)
        config.FIRM_PRICING_COVERAGE["sheet_steel"] = {"firm_capable": True,
                                                       "intended_source": "Uptonsteel"}
        covered = check_job(_job_with("firm"), write_back=False)
        _cc = [v["code"] for v in covered["violations"]]
        ok("no_firm_pricing_source" not in _cc, "coverage removes the not-configured refusal")
        _p = next(v for v in covered["violations"] if v["code"] == "price_not_firm")
        eq(_p["severity"], "blocking", "but a firm quote still blocks on an unfirm price")

        # And turning steel on must not claim plastic is integrated.
        eq((config.FIRM_PRICING_COVERAGE.get("plastic_sheet") or {}).get("firm_capable"), False,
           "one supplier at a time — plastic is untouched by the steel switch")
    finally:
        config.FIRM_PRICING_COVERAGE = _saved


def test_an_old_invoice_is_not_an_agreement():
    """History is evidence of a completed transaction, not a commitment to repeat it. Putting
    a validity date on an old invoice must not create an agreement that never existed — the
    agreement REFERENCE is what does the work."""
    import price_provenance as pp

    def _hist(**kw):
        b = {"source_name": "historical_quote_material_line", "applied": True,
             "affects_total": True, "source_rank": 0,
             "selected": {"source": "historical_quote_material_line", "price": 24.15}}
        b.update(kw)
        return b

    plain = pp.price_firmness(_hist(), today="2026-07-29")
    eq((plain["class"], plain["firm"]), (pp.PURCHASE_HISTORY, False),
       "what we last paid is not, by itself, firm")

    dated = pp.price_firmness(_hist(price_valid_to="2099-01-01"), today="2026-07-29")
    eq(dated["firm"], False, "and a date on an invoice does not make it an agreement")

    covered = pp.price_firmness(_hist(price_valid_to="2099-01-01",
                                      quote_reference="CONTRACT-4471"), today="2026-07-29")
    eq((covered["class"], covered["firm"]), (pp.CONTRACT, True),
       "a purchase still covered by a live agreement IS a contract price")

    lapsed = pp.price_firmness(_hist(price_valid_to="2020-01-01",
                                     quote_reference="CONTRACT-4471"), today="2026-07-29")
    eq(lapsed["firm"], False, "once the agreement lapses it is history again")


def test_a_pdf_only_job_still_reads_its_routes():
    """EVERY FIX THIS SESSION WAS DRIVEN BY A SOLIDWORKS-BACKED JOB. The measured-zero bend
    rule, the outline/cut-length split, the fold gate, the fabrication-evidence guard — all
    of them were written while looking at a pack with models, DXFs and a cut list.

    Most packs do not have that. A PDF-only job has no native record, no DXF outline and no
    cut list, and it still has to produce a BOM and a route: folds come from drawing
    callouts, and nothing about a measurement it never had may take them away. This is the
    case that would degrade silently, because it is not the one anybody is looking at."""
    from bought_in_policy import has_fabrication_evidence, is_bought_in
    from estimator import _model_measured_zero_bends, estimate_process_times

    def _ops(part):
        r = estimate_process_times(dict(part), quantity=1)
        return sorted((r or {}).get("times_min") or (r or {}).get("unit_times_min") or {})

    # Fabricated from the drawing alone: folds stated as callouts, no model, no DXF.
    pdf_part = {"part_number": "X-01M", "normalized_material": "MILD_STEEL",
                "normalized_thickness_mm": 1.5, "angles_deg": ["90", "90"],
                "fold_count_textual": 2, "flat_pattern_detected": True,
                "textual_operations": ["laser_cutting", "folding", "powder_coating"]}
    eq(has_fabrication_evidence(pdf_part), True, "a drawing-derived flat is still evidence we make it")
    eq(is_bought_in(pdf_part), False, "and it is not mistaken for a purchased part")
    eq(_model_measured_zero_bends(pdf_part), False,
       "with no model to count bends, nothing claims a measured zero")
    _o = _ops(pdf_part)
    ok("folding" in _o, f"so the drawing's folds are still costed: {_o}")
    ok("laser_cutting" in _o and "powder_coating" in _o, "along with the rest of its route")

    # The weakest case the engine has to survive: a part with no geometry at all.
    bare = {"part_number": "X-02M", "normalized_material": "MILD_STEEL",
            "textual_operations": ["laser_cutting"]}
    _b = _ops(bare)
    ok("folding" not in _b, f"a part with no fold evidence does not gain one: {_b}")
    ok("laser_cutting" in _b, "and keeps what the drawing did say")

    # A PDF-only part claims no measured geometry, so the outline check must not object to it.
    from invariants import check_job
    j = _job()
    j["part_estimates"] = [dict(pdf_part, quantity=1, quantity_source="bom_tree",
                                material_source="pdf", geometry_source="pdf_text")]
    _codes = [v["code"] for v in check_job(j, write_back=False)["violations"]]
    ok("measured_geometry_without_outline" not in _codes,
       f"a part that never claimed measurement is not accused of faking it: {_codes}")
    ok("plate_charged_for_folding" not in _codes,
       "and with no model to call it a plate, its fold stands")


def test_dwg_flat_patterns_are_converted_not_ignored():
    """DWG is DXF's binary sibling — the same geometry in a different container — and the
    engine read neither it nor anything else outside .pdf/.dxf/.sldXXX. A customer sending
    DWG flat patterns got an estimate built from the PDF alone, with transcribed blanks and
    inferred cut lengths, while the measured outline sat unread in the same folder.

    The conversion runs an external program, so it is driven here through an injected runner:
    a function that can only be tested on a machine with a particular tool installed is a
    function nobody tests."""
    import cad_inputs
    from pathlib import Path
    import tempfile

    with tempfile.TemporaryDirectory() as _tmp:
        job = Path(_tmp)
        (job / "12120-01-01M_1.5mm_MILD_STEEL.dwg").write_text("x")
        (job / "12120-01-02M_1.5mm_MILD_STEEL.DWG").write_text("x")
        (job / "12120-01-GA.pdf").write_text("x")
        (job / "assembly.STEP").write_text("x")

        # The converter is absent: say so, name the cost, and never raise into the run.
        _none = cad_inputs.convert_dwgs(job, converter=None)
        eq(len(_none["found"]), 2, "both DWGs are found whatever their case")
        eq(_none["converted"], [], "nothing is converted without the tool")
        ok("ODA File Converter" in _none["reason"],
           "and the reason names what to install rather than failing silently")
        ok("if any are part flat patterns" in _none["reason"],
           "hedged correctly — a job folder's DWGs are not all part flats, and 12120's only "
           "DWG is a GA sheet")
        ok("flat patterns are unread" not in _none["reason"],
           "so it does not promise measured blanks from files nobody has opened")

        # With a converter, the produced DXFs are reported as ours, not as customer input.
        out = job / "_dxf_from_dwg"
        def _fake(cmd):
            out.mkdir(parents=True, exist_ok=True)
            for n in ("12120-01-01M_1.5mm_MILD_STEEL.dxf", "12120-01-02M_1.5mm_MILD_STEEL.dxf"):
                (out / n).write_text("0\nSECTION\n")
            return 0
        _done = cad_inputs.convert_dwgs(job, out, converter="x", runner=_fake)
        eq(len(_done["converted"]), 2, "both DWGs become DXFs")
        eq(_done["reason"], "", "and there is nothing to explain")

        inv = cad_inputs.inventory(job, converted=[Path(p) for p in _done["converted_paths"]])
        ok("12120-01-GA.pdf" in inv["read"], "the PDF is read")
        eq(sorted(inv["converted"]), sorted(_done["converted"]),
           "a DXF we made is not reported as one the customer supplied")
        ok(any(f.lower().endswith(".step") for f in inv["unread"]),
           f"and the STEP file is named as present and unread: {inv['unread']}")

        # A converter that runs and produces nothing is a different failure and says so.
        _empty = cad_inputs.convert_dwgs(job, job / "_none", converter="x",
                                         runner=lambda cmd: 0)
        ok("no DXF" in _empty["reason"], f"reported, not silent: {_empty['reason']}")

    # CONVERTED IS NOT THE SAME AS READ. Folder discovery globs the job folder and a "DXF"
    # subfolder only — it does not recurse — so a converted file written anywhere else is
    # produced and then never opened, which looks exactly like the feature working. main.py
    # hands them over explicitly instead of hoping they are found.
    _main = open(__import__("main").__file__, encoding="utf-8").read()
    _seg = _main[_main.index("_cad_conv, _cad_inv = {}, {}"):]
    _seg = _seg[:_seg.index("scan_label = job_folder.name")]
    ok("convert_dwgs(job_folder)" in _seg, "conversion runs before the folder is scanned")
    ok("attach_dxf_paths=_converted_dxf" in _seg,
       "and the DXFs it produced are handed to the scan, not left to discovery")
    ok('summary["cad_inputs"] = _cad_inv' in _seg,
       "with the inventory stamped so the report and the gate can read it")

    # A JOB FOLDER'S DWGs ARE WHATEVER THE CUSTOMER SENT. attach_dxf_paths deliberately skips
    # the flat-part filter because a human naming a file has already made that judgement.
    # Nobody made it here, so a converted GA sheet handed over unfiltered would be read as a
    # part's flat pattern.
    # The exact expression, not just the name: the name also appears in the line that
    # REPORTS the rejects, so a looser assertion passes with the filter itself deleted.
    ok("_converted_dxf = [p for p in _converted_dxf if is_flat_part_dxf(p)]" in _seg,
       "converted files face the same test discovery applies to supplied ones")
    from drawing_job_merge import is_flat_part_dxf
    from pathlib import Path as _P
    eq(is_flat_part_dxf(_P("12120-01-GA-_1.5mm.dxf")), False, "a GA sheet is not a flat pattern")
    eq(is_flat_part_dxf(_P("12120-01-01M_1.5mm_MILD_STEEL.dxf")), True, "a part flat is")


def test_a_file_nobody_opened_is_named_in_the_report():
    """The engine ignored everything outside .pdf/.dxf/.sldXXX without a word. An estimator
    deciding whether to trust a number deserves to know a file they supplied was never opened
    — especially a DWG, which is measured geometry we could have used."""
    from invariants import check_job

    j = _job()
    j["cad_inputs"] = {"schema": "cad_inputs.v1", "present": True,
                       "read": ["12120-01-GA.pdf"], "solidworks": [], "converted": [],
                       "unread": ["01M.dwg", "02M.dwg", "assembly.STEP"], "unknown": []}
    r = check_job(j, write_back=False)
    _v = next((v for v in r["violations"] if v["code"] == "cad_files_not_read"), None)
    ok(_v is not None, "unread CAD files are reported")
    eq(_v["severity"], "warning", "as a warning — the number still stands")
    eq(_v["detail"]["dwg_count"], 2, "counting the DWGs separately")
    ok("01M.dwg" in _v["message"], "naming them so they can be found")
    ok("ODA File Converter" in _v["message"], "and saying what would make the DWGs readable")
    ok("skipped by design" in _v["message"],
       "while STEP is explained rather than presented as an omission")

    # Nothing unread: nothing to say.
    j2 = _job()
    j2["cad_inputs"] = {"schema": "cad_inputs.v1", "present": True,
                        "read": ["a.pdf", "b.dxf"], "solidworks": [], "converted": [],
                        "unread": [], "unknown": []}
    ok("cad_files_not_read" not in
       [v["code"] for v in check_job(j2, write_back=False)["violations"]],
       "a folder with nothing unread raises nothing")

    # A run that was not folder-based has no folder to inventory, and must not be marked
    # unverified for ever on the strength of a question that did not apply to it.
    r3 = check_job(_job(), write_back=False)
    ok("cad_inputs_not_evaluated" not in [v["code"] for v in r3["violations"]],
       "and a single-file run is not held against itself")


def test_dxfs_are_found_in_a_job_named_subfolder_and_read_once():
    """M&S job 2085 keeps its flat in TWO places: the job folder, and "2085 - DXFs_DEV1".

    Discovery matched a subfolder named literally "DXF", so the job-named one was invisible.
    On 2085 that was survivable only because a root copy happened to exist — had the flats
    lived solely in that subfolder, every part would have been sized from drawing text and
    nothing would have said so.

    Widening discovery then creates the opposite problem: the same file found twice, read
    twice, and a part that looks ambiguous when nothing about it is."""
    import drawing_job_merge as djm
    from pathlib import Path
    import tempfile

    with tempfile.TemporaryDirectory() as _tmp:
        job = Path(_tmp)
        sub = job / "2085 - DXFs_DEV1"
        sub.mkdir()
        _flat = b"0\nSECTION\n2\nENTITIES\n0\nENDSEC\n0\nEOF\n"
        (job / "2085-01 - Bracket Plate_1.2mm MS.DXF").write_bytes(_flat)
        (sub / "2085-01 - Bracket Plate_1.2mm MS.DXF").write_bytes(_flat)      # same bytes
        (sub / "2085-04 - Gusset_1.2mm MS.DXF").write_bytes(_flat + b"\n")     # only in the sub

        found = djm.discover_flat_dxf_files_in_folder(job)
        _names = sorted(p.name for p in found)
        ok(any("Gusset" in n for n in _names),
           f"a flat that exists only in the job-named subfolder is found: {_names}")

        kept = djm.collect_dxf_paths_for_job(job, {}, auto_discover_dxf=True)
        _plate = [p for p in kept if "Bracket Plate" in p.name]
        eq(len(_plate), 1, f"and the file copied to both places is read once: {_plate}")
        ok(any("Gusset" in p.name for p in kept), "without losing the one that is unique")

        # Same name, DIFFERENT bytes is a real ambiguity — a superseded revision beside a
        # current one — and both must survive for the candidate scoring to choose and flag.
        (sub / "2085-01 - Bracket Plate_1.2mm MS.DXF").write_bytes(_flat + b"0\nLINE\n")
        kept2 = djm.collect_dxf_paths_for_job(job, {}, auto_discover_dxf=True)
        eq(len([p for p in kept2 if "Bracket Plate" in p.name]), 2,
           "two files sharing a name and differing inside are both kept, not picked by "
           "accident of filesystem order")


def test_transcription_and_inference_are_separate_passes():
    """M&S 2085's LLM extract came back with every part field null — material 0, thickness 0,
    tube 0 — on a pack whose text plainly says MATERIAL: MILD STEEL, 1.2 WALL and 12.7.

    The model was not underperforming. It was obeying its instruction: every value must be
    PRINTED, and if it is not printed, use null. MILD STEEL is printed once at assembly level
    and the tubes are dimensioned on the views, so a strict transcriber correctly returns
    nothing — and the job books no material for two of three parts, no operation at all for
    either tube, and GBP 2.00 of labour on a welded bracket.

    That rule stays: it is what makes the first pass trustworthy. What was missing is the
    judgement an estimator applies after reading the same page, so there is a second pass with
    the opposite rule, ranked below everything real."""
    import llm_full_extract as lfe
    from source_precedence import SOURCE_RANK

    # One schema, two rules. The consumer reads one shape; `inferred` per route and `source`
    # per BOM row separate a reading from a judgement item by item.
    ok(lfe._SCHEMA in lfe._PROMPT and lfe._SCHEMA in lfe._INFER_PROMPT,
       "both passes return the same shape")
    ok(lfe._COMMON_RULES in lfe._PROMPT and lfe._COMMON_RULES in lfe._INFER_PROMPT,
       "and share the material-family and mixed-assembly rules")

    ok("TRANSCRIBE, NEVER INVENT" in lfe._PROMPT, "pass one may not invent")
    ok("inferred=false" in lfe._PROMPT, "so every route it returns is a reading")
    ok("Return an empty routes\nlist rather than a plausible one" in lfe._PROMPT,
       "and it is told to return nothing rather than something plausible")
    ok("opposite of transcription" in lfe._INFER_PROMPT, "pass two is the other thing")
    ok("inferred=true" in lfe._INFER_PROMPT, "and everything it returns says so")
    # Whitespace-insensitive: the prompt is wrapped for readability and a line break inside
    # the sentence is not a change of meaning.
    _flat = " ".join(lfe._INFER_PROMPT.split())
    ok("A null is honest" in _flat,
       "with permission to decline, which is what stops it filling every field")
    ok("cannot tell, leave it null" in _flat, "and an explicit instruction to do so")

    # Multi-material: five families plus bought-in, and the mixed-assembly rule.
    for fam in ("metal", "acrylic", "timber", "wire", "tube", "bought_in"):
        ok(fam in lfe._SCHEMA, f"'{fam}' is a material family the schema knows")
    ok("Keep components PURE" in lfe._COMMON_RULES,
       "components are pure; only the assembly is mixed")
    ok("those part numbers only" in lfe._COMMON_RULES,
       "and a finish is not spread across parts it is not applied to")
    # The prompt used to name departments by TITLE ("Laser (Acrylic)", "Tubebend"). It now
    # asks for the rate table's own CODES, because a title is one paraphrase away from a
    # string no LOOKUP resolves, and an unresolved department costs nothing in silence.
    for dept in ("LASA", "LINE", "TBEN", "EDGE", "ROBO"):
        ok(dept in lfe._COMMON_RULES, f"the shop's own department code is used: {dept}")

    # The rank is the whole safety argument: inference must lose to everything real.
    _inf = SOURCE_RANK.get(lfe.INFERENCE_SOURCE)
    ok(_inf is not None, f"'{lfe.INFERENCE_SOURCE}' is a known source")
    for stronger in ("solidworks_api", "dxf", "drawing_deterministic", "bom_tree",
                     "llm_full_extract"):
        ok(SOURCE_RANK[stronger] > _inf,
           f"{stronger} ({SOURCE_RANK[stronger]}) outranks inference ({_inf})")

    # Only parts with nothing to cost from are asked about again.
    _missing = lfe.parts_missing_detail([
        {"part_number": "2085-02", "description": "OUTER TUBE"},
        {"part_number": "2085-01", "description": "PLATE", "material": "MILD STEEL",
         "thickness_mm": 1.2},
        {"part_number": "BI-KNOB", "description": "KNOB", "is_bought_in": True},
    ])
    _asked = {m["part_number"] for m in _missing}
    ok("2085-02" in _asked, "a part with no material and no size is asked about")
    ok("2085-01" not in _asked, "one the drawing described is left as the drawing described it")
    ok("BI-KNOB" not in _asked, "and a bought-in is not something we route")

    eq(lfe.infer_missing_details("ctx", [], []), {}, "nothing missing, nothing asked")

    # BUILT IS NOT WIRED. Three times this session a capability went in with nothing calling
    # it. A capability nothing calls looks exactly like one that does not work.
    _fs = open(__import__("file_scan").__file__, encoding="utf-8").read()
    ok("infer_missing_details(" in _fs, "the inference pass is actually called by the scan")
    ok("parts_missing_detail(" in _fs, "on the parts the first pass left empty")
    ok('_job["inferred_parts"]' in _fs, "and its result is kept on the job")
    # KEPT IS NOT MERGED. The line above passed while the pass was a no-op: the result was
    # stored at job["inferred_parts"], and apply_full_job_to_pre_estimate reads job["parts"]
    # and job["routes"] and nothing else. "Kept on the job" was true and worthless.
    ok("merge_inference(_job, _inf)" in _fs,
       "and MERGED into the parts and routes the fold actually reads")
    _lj = open(__import__("source_connectors.llm_full_job", fromlist=["x"]).__file__,
               encoding="utf-8").read()
    ok("apply_routes_to_parts(parts, job)" in _lj,
       "and the extracted route is folded onto the parts it names")
    ok('"inference" if route.get("inferred")' in _lj,
       "with the model's own inferred flag deciding the source rank, so a concluded "
       "operation never outranks a measured one")


def test_the_dxf_model_pass_judges_but_never_measures():
    """ezdxf reads a DXF exactly — entities, loops, areas, cut lengths, hole diameters. What it
    cannot say is what the geometry MEANS: which layer is the cut profile and which the bend
    lines, whether a 7mm circle is a clearance hole or a keyhole, whether the file is one part
    or a nest of six. That is process judgement and it is what a model is for.

    The model is asked for the geometry too, deliberately — and its numbers are NOT used. The
    measurement stands. A second independent read that differs materially means one of them is
    wrong about a file we are about to cost, and neither silently winning is the right answer.

    It is sent the extraction, never the file: a DXF is tens of thousands of coordinate
    triples, and a model handed those produces a confident cut length with nothing behind it —
    the exact failure removed from prices, not to be reintroduced through geometry.

    And it is NOT sent the cut lengths. Those are the numbers that cost money, so its own
    figure is worth having — but only if it was reached independently. An earlier version of
    this fixture asserted the measured cut length WAS in the payload, which made the
    "disagreement" check compare a number against itself.
    """
    import dxf_llm_interpret as dli
    from source_precedence import SOURCE_RANK

    measured = {
        "layers": ["CUT", "BEND", "DIMS"], "entity_counts": {"LINE": 42, "CIRCLE": 3},
        "bounding_box_mm": [90.88, 80.0], "blank_mm": [90.88, 80.0],
        "closed_contour_count": 4, "hole_diameters_mm": [7.0, 4.0, 4.0],
        "hole_count": 3, "cut_length_mm": 374.31,
        "text_entities": ["2085-01", "1.2 MS"],
    }

    # The payload is the extraction, not the file — minus what is being cross-checked.
    _payload = dli.build_payload(measured, "2085-01 - Bracket Plate_1.2mm MS.DXF")
    ok("CUT" in _payload and "BEND" in _payload, "the layer names it must classify are shown")
    ok("2085-01" in _payload, "and the text it must read material and thickness from")
    ok("374.31" not in _payload,
       "but NOT the measured cut length — the second opinion has to be independent to count")

    def _reply(_prompt):
        return {
            "geometry_interpretation": {
                "estimated_total_cut_length_mm": 374.0, "overall_width_mm": 90.9,
                "overall_height_mm": 80.0, "is_flat_pattern": True, "profile_type": "closed"},
            "holes": [{"diameter_mm": 7.0, "count": 1, "type": "keyhole"},
                      {"diameter_mm": 4.0, "count": 2, "type": "round"}],
            "manufacturing": {"recommended_process": "laser", "complexity": "simple",
                              "secondary_processes": ["folding"],
                              "is_nested": False, "part_count": 1,
                              "material_inferred": "MILD STEEL", "thickness_inferred_mm": 1.2,
                              "profile_role_by_layer": {"CUT": "cut", "BEND": "bend",
                                                        "DIMS": "dimension"},
                              "operations_implied": ["laser_cutting", "folding"]},
            "warnings": [], "extraction_confidence": "high"}

    got = dli.interpret(measured, caller=_reply)
    eq(got["profile_role_by_layer"]["BEND"], "bend", "the bend layer is identified")
    eq(got["hole_types"][0]["type"], "keyhole",
       "and a 7mm circle is called what it is, which a diameter alone cannot say")
    eq(got["recommended_process"], "laser", "with the process judgement taken")
    eq(got["disagreements"], [], "agreeing reads raise nothing")
    ok("overall_width_mm" in got["not_cross_checked"],
       "and the fields it WAS shown are declared as not cross-checked, not passed off as agreement")

    # Everything it judged is ranked below every measurement.
    _inf = SOURCE_RANK[got["interpretation_source"]]
    for stronger in ("solidworks_api", "dxf", "drawing_deterministic"):
        ok(SOURCE_RANK[stronger] > _inf,
           f"{stronger} outranks the model's interpretation ({_inf})")

    # A materially different read is a finding, not a coin toss.
    def _wrong(_prompt):
        r = _reply(_prompt)
        r["geometry_interpretation"]["estimated_total_cut_length_mm"] = 748.0      # double
        return r
    _dis = dli.interpret(measured, caller=_wrong)["disagreements"]
    eq([d["field"] for d in _dis], ["estimated_total_cut_length_mm"],
       "the disagreement is named")
    eq(_dis[0]["measured"], 374.31, "with the measured value kept as the answer")
    ok(_dis[0]["difference_pct"] > 90, "and the size of the gap quantified")

    # A model that cannot be reached leaves the geometry costing the part exactly as today.
    eq(dli.interpret(measured, caller=lambda _p: None), {}, "no model, no interpretation")
    eq(dli.interpret({}, caller=_reply), {}, "and nothing measured, nothing to interpret")


def test_the_dxf_interpretation_fills_gaps_and_shouts_about_a_nest():
    """The merge, and the precedence in it. ezdxf wins every measured length; the model wins
    process, hole typing and material inference, and only where the part has nothing.

    The nest case is the one judgement here that moves a price on its own: a measured blank
    and cut length cover the whole sheet, so a nest of six costed as one part is wrong by
    about six. It is never applied silently — a person is told and asked."""
    import dxf_llm_interpret as dli

    interp = {"found": True, "material_inferred": "MILD STEEL", "thickness_inferred_mm": 1.2,
              "recommended_process": "laser", "is_nested": True, "part_count": 6,
              "disagreements": [{"field": "estimated_total_cut_length_mm", "measured": 374.31,
                                 "model_read": 748.0, "difference_pct": 99.8}],
              "warnings": ["outer profile does not close"]}

    # A part with nothing: the interpretation fills it and says it was inferred.
    empty = {"part_number": "2085-01"}
    dli.apply_to_part(empty, interp)
    eq(empty["normalized_material"], "MILD STEEL", "a gap is filled")
    eq(empty["material_source"], "inference", "and stamped inference, not measurement")
    eq(empty["normalized_thickness_mm"], 1.2, "thickness likewise")
    ok(any("NEST of 6" in str(f) for f in empty["review_flags"]),
       "a nest is raised loudly — costing it as one part is wrong by about six")
    ok(any("cross-check" in str(f) for f in empty["review_flags"]),
       "and so is a materially different second read")
    ok(any("does not close" in str(f) for f in empty["review_flags"]),
       "with the model's own warnings carried to a person")

    # A part that was MEASURED: nothing the model says displaces it.
    measured_part = {"part_number": "2085-01", "normalized_material": "STAINLESS_STEEL",
                     "normalized_thickness_mm": 2.0, "material_source": "solidworks_api"}
    dli.apply_to_part(measured_part, interp)
    eq(measured_part["normalized_material"], "STAINLESS_STEEL",
       "a measured material is not overwritten by a read one")
    eq(measured_part["material_source"], "solidworks_api", "and keeps its own source")
    eq(measured_part["normalized_thickness_mm"], 2.0, "nor is a measured thickness")

    eq(dli.apply_to_part({}, {"found": False}), {"filled": 0, "flags": 0},
       "no interpretation, no change")

    # BUILT IS NOT WIRED — this module shipped complete and called by nobody for two commits.
    _djm = open(__import__("drawing_job_merge").__file__, encoding="utf-8").read()
    ok("from dxf_llm_interpret import" in _djm,
       "the interpretation runs where a DXF is actually measured")
    ok("_dxf_apply(part, _interp)" in _djm, "and is folded onto that part")


def test_the_drawing_border_is_not_a_bill_of_materials_line():
    """M&S 2085's sheet is gridded 1-20 across and A-I down. The table reader swallowed
    "...14 15 16 17 18 19 20" sitting immediately before the ITEM NO. header and emitted
    part_number "1415", description "16 17 18 19", quantity 20 — then priced it by AI market
    estimate at GBP 219.21 of a GBP 273.98 unit cost. Eighty percent of the job, from the
    picture frame.

    Every drawing has a border and every border is numbered, so this is not one customer's
    quirk. The test is deliberately narrow in both directions at once — dropping a real BOM
    line is silent and far worse than costing a phantom one, which at least shows up in the
    total and gets challenged."""
    from bom_pipeline import is_drawing_furniture

    eq(is_drawing_furniture("1415", "16 17 18 19"), True, "the border grid is not a part")
    eq(is_drawing_furniture("12", "1 2 3"), True, "nor is any other run of frame numbers")

    # Real rows, none of which may be lost.
    for code, desc in (("2085-01", "BRACKET PLATE"), ("2085-02", "OUTER TUBE"),
                       ("THUM620", "M4x10mm MUSHROOM THUMBSCREW"),
                       ("BI-PEMSTUD", "M4 THREADED PEM STUD (LENGTH: 30mm)"),
                       ("12120-01-01M", "MOUNTING BRACKET")):
        eq(is_drawing_furniture(code, desc), False, f"a real BOM row survives: {code} {desc}")

    # BOTH tests must hold, because either alone is unsafe: a customer may genuinely number
    # parts "1415", and "M6 x 20" is a perfectly good description made mostly of digits.
    eq(is_drawing_furniture("1415", "BRACKET"), False,
       "a digits-only code with a real description is a real part")
    eq(is_drawing_furniture("2085-01", "16 17 18 19"), False,
       "and a structured part number is never furniture, whatever sits beside it")
    eq(is_drawing_furniture("", "16 17"), False, "an empty code decides nothing")

    # THE PREDICATE IS NOT THE FIX. A correct test that nothing calls changes no estimate,
    # and this suite has now watched four capabilities ship unplugged in one session. The
    # flattening loop needs a real PDF to drive, so the call site is asserted by its exact
    # expression — the name alone also appears in the definition and the finding text.
    import bom_pipeline
    _src = open(bom_pipeline.__file__, encoding="utf-8").read()
    ok("if is_drawing_furniture(code, desc):" in _src,
       "the BOM flattener actually applies it")
    ok("bom_row_is_drawing_furniture" in _src,
       "and a dropped row is recorded as a finding, not discarded in silence")


def test_a_bom_line_only_one_reader_saw_is_not_costed_in_silence():
    """The BOM is read twice on purpose, and where only one reader sees a row it is emitted
    and FLAGGED — a vision pass missing a real line is as likely as a table reader inventing
    one. On 2085 the phantom came back A_ONLY with "vision did not corroborate — review"
    against it, and was then priced at 80% of the job. The flag reached a JSON field and
    nothing downstream weighed it.

    Not dropped here. Named, with what it is worth."""
    from invariants import check_job

    def _job_with(share_value, corroborated_value=0.10):
        j = _job()
        j["document_analysis"] = {"bom_rows": [
            {"part_number": "2085-01", "description": "BRACKET PLATE", "quantity": 1,
             "bom_source": "BOTH", "bom_confidence": "HIGH", "bom_flag": ""},
            {"part_number": "1415", "description": "16 17 18 19", "quantity": 20,
             "bom_source": "A_ONLY", "bom_confidence": "MED",
             "bom_flag": "A-only (vision did not corroborate) — review"},
        ]}
        j["final_estimate"]["material_rows"] = [
            {"workbook_row": 11, "part_code": "2085-01",
             "total_value_gbp": corroborated_value},
            {"workbook_row": 12, "part_code": "1415", "total_value_gbp": share_value},
        ]
        j["final_estimate"]["totals"] = {"material_gbp": corroborated_value + share_value,
                                         "labour_gbp": 2.00,
                                         "unit_gbp": round((corroborated_value + share_value
                                                            + 2.00) / 0.93, 2)}
        return j

    r = check_job(_job_with(219.21), write_back=False)
    v = next((x for x in r["violations"]
              if x["code"] == "uncorroborated_bom_line_costed"), None)
    ok(v is not None, "a flagged row carrying money is reported")
    eq(v["severity"], "blocking", "and dominating the material total blocks the quote")
    ok("1415" in v["message"], "naming it")
    ok("219.21" in v["message"], "with what it is worth")
    ok("2085-01" not in str(v["detail"]["lines"]), "the corroborated row is not implicated")

    # Flagged but trivial — GBP 0.40 against GBP 50 of corroborated material. Worth saying,
    # not worth stopping a quote for: the severity follows what the line is worth, not the
    # fact that it was flagged, or every job with one odd row would block.
    _small = check_job(_job_with(0.40, corroborated_value=50.00), write_back=False)
    _v2 = next(x for x in _small["violations"]
               if x["code"] == "uncorroborated_bom_line_costed")
    eq(_v2["severity"], "warning", "a small uncorroborated line is a warning, not a block")

    # Both readers agreed on everything: nothing to say.
    j3 = _job()
    j3["document_analysis"] = {"bom_rows": [
        {"part_number": "2085-01", "description": "BRACKET PLATE", "quantity": 1,
         "bom_source": "BOTH", "bom_confidence": "HIGH", "bom_flag": ""}]}
    ok("uncorroborated_bom_line_costed" not in
       [x["code"] for x in check_job(j3, write_back=False)["violations"]],
       "a BOM both readers saw raises nothing")


def main() -> int:
    global _COLLECT_ONLY
    _COLLECT_ONLY = True          # collect every failure in a test, don't stop at the first
    tests = [(n, f) for n, f in sorted(globals().items())
             if n.startswith("test_") and callable(f)]

    # EVERY FIXTURE IN THIS FILE MUST ACTUALLY RUN. globals() only holds what has been
    # defined by the time main() executes, so a fixture written below the __main__ block is
    # collected by nobody and reports nothing — it simply is not there. One was, and it sat
    # "passing" through two mutation checks that should have failed, which is worse than no
    # fixture at all: a green suite asserting something it never executed.
    import re as _re
    _declared = {m.group(1) for m in
                 _re.finditer(r"^def (test_\w+)", open(__file__, encoding="utf-8").read(),
                              _re.MULTILINE)}
    _missing = sorted(_declared - {n for n, _ in tests})
    if _missing:
        # Reported here rather than through _fail(): the failure count is tallied inside the
        # loop below, so anything recorded before it is discarded and the suite still prints
        # OK — which is the same "green but unexecuted" problem in a different place.
        print(f"\nFAILED — {len(_missing)} fixture(s) are defined but never run. They sit "
              f"below the __main__ block, so nothing collects them; move them above it:")
        for _name in _missing:
            print(f"      - {_name}")
        return 1

    print(f"\n{len(tests)} regression fixture(s) — each locks a defect that reached a real estimate\n")
    errors = 0
    for name, fn in tests:
        before = len(_FAILS)
        try:
            fn()
        except Exception:
            _FAILS.append(f"{name} raised:\n{traceback.format_exc()}")
        new = len(_FAILS) - before
        print(f"  {'FAIL' if new else 'pass'}  {name}")
        for f in _FAILS[before:]:
            print(f"      - {f}")
        errors += new
    print(f"\n{'FAILED' if errors else 'OK'} — {errors} failure(s)\n")
    return 1 if errors else 0


def test_a_part_we_make_is_never_sent_to_the_market_for_a_price():
    """12120-01-101 and -103 are our own weldments — job-prefixed, fabricated here, their
    material carried by their child BOM lines. Both were priced by an LLM market estimate,
    and both figures reached the total.

    The gate meant to stop that reads is_assembly_parent, is_sub_assembly, reliability_flags
    and the part's operations. The caller handed it a three-key stub — number, description,
    material — so every guard but two text tests was blind, and neither text test matches a
    number like 12120-01-101. Testing the gate directly passed; the path through the caller
    was never exercised. That is the fifth time this exact shape has bitten.

    So this drives the CALLER, with a pricing service that records what it was asked."""
    import estimator

    asked = []

    class _RecordingService:
        def _select_anchor_price_source(self, part):
            asked.append(dict(part))
            from pricing_service import PricingService
            if not PricingService._web_ai_fallback_allowed(self, part, {"enable_web_ai_fallback": True}):
                return {"source": "fallback", "unit_price_gbp": 0.0, "confidence": 0.0}
            return {"source": "llm_market_estimate", "unit_price_gbp": 7.20,
                    "source_type": "web_ai_fallback", "confidence": 0.35}
        _web_ai_calls = 0

    _saved = estimator._PRICING_SERVICE_SINGLETON
    estimator._PRICING_SERVICE_SINGLETON = _RecordingService()
    try:
        # A part we fabricate, described the way the real ones are: the number does not end
        # -GA and the description says nothing about a weldment or an assembly, so the two
        # text tests that were the gate's only working guards both miss it. What identifies
        # it is that we weld it — evidence no purchased component can carry.
        #
        # The first version of this fixture wrote "BRACKET WELDMENT LH", which the old text
        # test caught, so both mutations passed and the fixture proved nothing. Same failure
        # as the defect it is testing.
        made = {"part_number": "12120-01-101", "description": "BRACKET LH",
                "normalized_material": "MILD_STEEL",
                "textual_operations": ["welding", "dress_welds"]}
        got = estimator._resolve_part_system_cost(made)
        ok(asked, "the pricing service was reached at all")
        ok("textual_operations" in asked[-1],
           "the caller passes the whole part, not a three-key stub")
        ok(got.get("applied_unit_cost") in (None, 0.0, 0),
           f"a part we make gets no market price (got {got.get('applied_unit_cost')})")

        # A genuine bought-in with no fabrication evidence still reaches the fallback: this
        # gate must not become a blanket ban on filling catalogue gaps.
        asked.clear()
        bought = {"part_number": "BI-SCREENCABLE", "description": "SCREEN CABLE ASSY",
                  "textual_operations": ["handling"]}
        got2 = estimator._resolve_part_system_cost(bought)
        eq(got2.get("applied_unit_cost"), 7.20,
           "a real bought-in is still priced by the fallback")
    finally:
        estimator._PRICING_SERVICE_SINGLETON = _saved


def test_the_inference_pass_is_not_told_never_to_invent_first():
    """The second pass came back all nulls, and the reason was in the message it was sent.

    _call_llm hardcoded the TRANSCRIPTION prompt — "ABSOLUTE RULE FOR THIS PASS: TRANSCRIBE,
    NEVER INVENT ... Return an empty routes list rather than a plausible one" — plus a system
    message of "Never invent", and then appended whatever the caller passed as if it were the
    drawing pack. So the inference pass, whose entire purpose is to conclude what the drawing
    does not print, was told to refuse before it was asked to answer. It obeyed the first
    instruction. Estimators got better results pasting the same pack into a chat window,
    because the chat window was not being told that.

    This asserts what actually goes ON THE WIRE. The first version of this fixture stubbed
    _call_llm and inspected its arguments, which proved nothing: the defect was INSIDE
    _call_llm, so replacing it made the mutation pass. The stub goes at the openai boundary
    instead, and the real _call_llm builds the real message.
    """
    import sys as _sys
    import types as _types
    import llm_full_extract as lfe

    sent = {}

    class _FakeCompletions:
        def create(self, model=None, messages=None, temperature=None):
            sent["system"] = messages[0]["content"]
            sent["user"] = messages[1]["content"]
            eq(len(messages), 2, "one system message and one user message, nothing else")
            _msg = _types.SimpleNamespace(content='{"parts": [], "routes": []}')
            return _types.SimpleNamespace(choices=[_types.SimpleNamespace(message=_msg)])

    class _FakeClient:
        def __init__(self, **_kw):
            self.chat = _types.SimpleNamespace(completions=_FakeCompletions())

    _fake_openai = _types.ModuleType("openai")
    _fake_openai.OpenAI = _FakeClient
    _saved_mod = _sys.modules.get("openai")
    _saved_key = os.environ.get("XAI_API_KEY")
    _saved_ref = os.environ.get("SDI_LLM_EXTRACT_REFRESH")
    _sys.modules["openai"] = _fake_openai
    os.environ["XAI_API_KEY"] = "test-key-not-used"
    # THE CACHE WOULD ANSWER THIS BEFORE THE MODEL DOES. Extract results are now cached on
    # content, so a fixture that must observe a LIVE call has to bypass it — otherwise it
    # silently stops testing anything the moment a real run has populated the cache, which
    # is exactly what happened while this was being written.
    os.environ["SDI_LLM_EXTRACT_REFRESH"] = "1"
    try:
        lfe.infer_missing_details("PACK TEXT", [], [{"part_number": "2085-02"}])
    finally:
        if _saved_mod is None:
            _sys.modules.pop("openai", None)
        else:
            _sys.modules["openai"] = _saved_mod
        if _saved_key is None:
            os.environ.pop("XAI_API_KEY", None)
        else:
            os.environ["XAI_API_KEY"] = _saved_key
        if _saved_ref is None:
            os.environ.pop("SDI_LLM_EXTRACT_REFRESH", None)
        else:
            os.environ["SDI_LLM_EXTRACT_REFRESH"] = _saved_ref

    ok(sent, "the call was made at all")
    ok("NEVER INVENT" not in sent.get("user", ""),
       "the inference pass is not sent the transcription pass's absolute rule")
    ok("Return an empty routes list rather than a plausible one" not in sent.get("user", ""),
       "the inference pass is not told to return an empty routes list")
    ok("the opposite of transcription" in sent.get("user", ""),
       "it IS sent its own prompt")
    ok("Never invent." != sent.get("system"),
       "the inference pass does not run under the transcriber's system message")
    ok("IMPLIES" in str(sent.get("system", "")),
       "the inference pass runs under a system message that asks it to infer")


def test_the_inference_pass_returns_the_routes_it_was_asked_for():
    """_INFER_PROMPT says "Every route you return must have inferred=true", and
    apply_routes_to_parts already ranks an inferred route below a stated one — but
    infer_missing_details returned {"parts": [...]} and dropped `routes` on the floor. The
    entire mechanism built for inferred routes could never fire. M&S 2085's welded three-part
    bracket booked GBP 2.00 of labour with no operation at all against either tube.

    Also accepts `bom` for the component list: the shared schema calls it that, and returning
    nothing because the model used the schema's own word is the most expensive shape of bug
    here — it is indistinguishable from a model that had nothing to say.
    """
    import llm_full_extract as lfe

    _saved = lfe._call_llm
    lfe._call_llm = lambda user_content, model, system=None: (
        '{"bom": [{"part_number": "2085-02", "material": "MILD STEEL",'
        ' "thickness_or_section": "25x25x1.5 SHS"}],'
        ' "routes": [{"sequence": 10, "operation": "saw", "part_numbers": ["2085-02"],'
        '  "inferred": false, "confidence": "high"}]}')
    try:
        got = lfe.infer_missing_details("PACK", [], [{"part_number": "2085-02"}])
    finally:
        lfe._call_llm = _saved

    eq(len(got.get("routes") or []), 1, "the route survives the pass")
    eq(got["routes"][0]["inferred"], True,
       "a route from the inference pass is inferred whatever it claimed")
    eq(len(got.get("parts") or []), 1, "a component list returned as `bom` is still read")
    eq(got["parts"][0]["source"], "inference", "and it is stamped inference")
    ok(got.get("found"), "a pass that produced routes and parts reports that it found something")


def test_inference_fills_a_hole_and_never_overwrites_a_reading():
    """merge_inference is the rule, so it is what gets tested.

    A transcribed value is a reading; an inferred one is a conclusion. Per datum, not per
    part: a part whose material was printed but whose section was not must take the inferred
    section and keep the printed material. What was filled is recorded in field_sources, so
    the distinction survives into the estimate instead of being a fact about which pass ran.
    """
    from llm_full_extract import merge_inference

    job = {"parts": [{"part_number": "2085-02", "material": "MILD STEEL"}],
           "routes": [{"operation": "welding", "part_numbers": ["2085-01"]}]}
    inf = {"parts": [{"part_number": "2085-02", "material": "ALUMINIUM",
                      "tube_section": "25x25x1.5 SHS", "cut_length_mm": 340},
                     {"part_number": "2085-03", "material": "MILD STEEL"}],
           "routes": [{"operation": "saw", "part_numbers": ["2085-02"]},
                      {"operation": "welding", "part_numbers": ["2085-01"]}]}
    counts = merge_inference(job, inf)

    p = next(x for x in job["parts"] if x["part_number"] == "2085-02")
    eq(p["material"], "MILD STEEL", "a printed material is not overwritten by an inferred one")
    eq(p["tube_section"], "25x25x1.5 SHS", "but an empty section is filled")
    eq(p["field_sources"].get("tube_section"), "inference", "and recorded as inference")
    ok("material" not in p.get("field_sources", {}),
       "the material the drawing printed is not relabelled inference")
    eq(counts["parts_added"], 1, "a part the transcription never listed is added whole")

    ops = [(r["operation"], tuple(r.get("part_numbers") or [])) for r in job["routes"]]
    ok(("saw", ("2085-02",)) in ops, "the inferred route is added")
    eq(len([o for o in ops if o[0] == "welding"]), 1,
       "an operation already read for that part is not repeated as inferred")


def test_an_inferred_route_reaches_the_part_and_says_it_was_inferred():
    """The three previous attempts at this all built the mechanism and wired none of it: the
    pass ran, stamped everything correctly, and stored the result at job["inferred_parts"] —
    which apply_full_job_to_pre_estimate does not read. It reads job["parts"] and
    job["routes"]. So the inference sat beside the job it was meant to complete.

    This drives the whole chain: merge, then fold, then check the part.
    """
    from llm_full_extract import merge_inference
    from source_connectors.llm_full_job import apply_full_job_to_pre_estimate

    job = {"found": True, "source": "llm_full_extract",
           "parts": [{"part_number": "2085-02", "description": "TUBE"}],
           "routes": []}
    merge_inference(job, {
        "parts": [{"part_number": "2085-02", "material": "MILD STEEL",
                   "tube_section": "25x25x1.5 SHS", "cut_length_mm": 340}],
        "routes": [{"operation": "saw", "part_numbers": ["2085-02"]},
                   {"operation": "welding", "part_numbers": ["2085-02"]}]})

    parts = [{"part_number": "2085-02", "description": "TUBE"}]
    counts = apply_full_job_to_pre_estimate(parts, job)
    part = parts[0]

    ok("saw" in (part.get("textual_operations") or []),
       f"the inferred saw operation reaches the part (got {part.get('textual_operations')})")
    eq((part.get("operation_sources") or {}).get("saw"), "inference",
       "and carries source 'inference', not 'llm_full_extract'")
    eq(part.get("material_source"), "inference",
       "an inferred material is stamped inference, not left indistinguishable from a reading")
    eq((part.get("section_stock") or {}).get("length_mm"), 340.0,
       "the inferred cut length drives the tube path")
    ok(any("INFERRED" in str(f) for f in (part.get("review_flags") or [])),
       "and the estimator is told on the part which values were concluded, not read")
    ok(counts["operations"] >= 2, "both inferred operations were folded in")

def test_the_schema_says_bom_and_the_engine_reads_parts():
    """M&S 2085's whole sheet, in one disconnection.

    When the multi-material schema went in, the component list was renamed `bom`. Three
    consumers key on job["parts"]: apply_full_job_to_pre_estimate, overlay_drawing_facts and
    parts_missing_detail. All three went silently to zero. No material, no thickness and no
    section reached any part; and because the missing-detail list came back empty, the
    inference pass — written precisely for this job — was never asked to run at all.

    What reached the estimator looked like a costing bug: two tubes with no material and no
    operation, the outer one priced at GBP 86.04 by a market estimate because nothing said we
    make it, and GBP 2.00 of labour on a welded three-part bracket.
    """
    from llm_full_extract import normalize_job, parts_missing_detail

    job = {"found": True, "bom": [
        {"part_number": "2085-01", "description": "BRACKET PLATE", "qty": 1,
         "material_family": "metal", "material": "MILD STEEL",
         "thickness_or_section": "1.2mm"},
        {"part_number": "2085-02", "description": "OUTER TUBE", "qty": 1,
         "material_family": "tube", "material": "", "thickness_or_section": "25x25x1.5 SHS",
         "cut_length_mm": 340},
        {"part_number": "BI-KNOB", "description": "KNURLED KNOB", "qty": 2,
         "material_family": "bought_in", "is_bought_in": True},
        # A section on a METAL row. The family branch does not save this one: without the
        # section test, "40 x 40 x 2mm SHS" reads as a 2mm sheet thickness and the part is
        # nested and lasered as flat plate instead of sawn from box.
        {"part_number": "2085-04", "description": "LEG", "qty": 4,
         "material_family": "metal", "thickness_or_section": "40 x 40 x 2mm SHS"},
    ]}
    normalize_job(job)

    eq(len(job["parts"]), 4, "every BOM row becomes a part record the engine can read")
    plate = next(p for p in job["parts"] if p["part_number"] == "2085-01")
    tube = next(p for p in job["parts"] if p["part_number"] == "2085-02")
    knob = next(p for p in job["parts"] if p["part_number"] == "BI-KNOB")

    eq(plate["thickness_mm"], 1.2, "'1.2mm' on a metal row is a sheet thickness")
    eq(plate["tube_section"], None, "and not a section")
    eq(tube["tube_section"], "25x25x1.5 SHS", "'25x25x1.5 SHS' is a section")
    eq(tube["thickness_mm"], None, "and not a thickness")
    eq(tube["cut_length_mm"], 340, "the cut length comes through, or the section cannot be costed")
    eq(tube["is_bought_in"], False, "a tube is stock we saw, not a component we purchase")
    eq(knob["is_bought_in"], True, "and a bought_in family still is one")

    leg = next(p for p in job["parts"] if p["part_number"] == "2085-04")
    eq(leg["tube_section"], "40 x 40 x 2mm SHS",
       "a section is a section whatever family the row is in")
    eq(leg["thickness_mm"], None,
       "it is NOT a 2mm sheet — that would nest and laser a part we saw from box")

    _asked = {m["part_number"] for m in parts_missing_detail(job["parts"])}
    ok("2085-02" in _asked,
       "the tube with no material is now what the inference pass gets asked about")


def test_a_job_already_in_the_parts_shape_is_left_alone():
    """The bridge must not rewrite a job that already arrived in the shape the engine reads —
    an older cached extract, or a caller that built the job itself."""
    from llm_full_extract import normalize_job
    job = {"found": True, "parts": [{"part_number": "X-1", "material": "MILD STEEL"}],
           "bom": [{"part_number": "SOMETHING-ELSE"}]}
    normalize_job(job)
    eq(len(job["parts"]), 1, "the existing parts list is untouched")
    eq(job["parts"][0]["part_number"], "X-1", "and not replaced from bom")


def test_an_unidentified_tube_is_not_a_purchase():
    """A GA states MILD STEEL once, at assembly level. Its tubes therefore carry no material
    of their own, and an unidentified material falls through to BOUGHT_IN by default. Being
    bought-in, strip_fabrication_ops removed the saw and the weld, so the part had no route
    and was priced by a market estimate instead of costed.

    The drawing DID say what it was: material_family "tube". That classification was read and
    consumed by nothing. A stated family is a positive statement and outranks the absence of
    a material — but it must not outrank catalogue identity, which is the module's whole
    founding rule, so both directions are asserted here.
    """
    from bought_in_policy import is_bought_in, bought_in_reason, strip_fabrication_ops

    tube = {"part_number": "2085-02", "description": "OUTER TUBE",
            "material_family": "tube", "normalized_material": "BOUGHT_IN",
            "textual_operations": ["saw", "welding", "handling"]}
    ok(not is_bought_in(tube), "a tube whose material went unidentified is still a tube")
    eq(bought_in_reason(tube), "", "and nothing claims otherwise")
    eq(strip_fabrication_ops(tube), [], "so its saw and weld survive")
    ok("welding" in tube["textual_operations"], "the weld is still on the part")

    # Catalogue identity is NOT overridable by a family. A supplier's own part can perfectly
    # well be a metal one; we still buy it.
    catalogue = {"part_number": "BI-KNURLEDKNOB", "material_family": "metal"}
    ok(is_bought_in(catalogue), "a BI- code is bought-in whatever family it is in")
    ok("code family" in bought_in_reason(catalogue), "and says which rule decided")

    flagged = {"part_number": "2085-09", "material_family": "tube", "is_bought_in": True}
    ok(is_bought_in(flagged), "an explicit bought-in flag is not overridden by a family")

    # And a genuinely unidentified part is unchanged: this must not become a blanket amnesty.
    unknown = {"part_number": "2085-77", "normalized_material": "BOUGHT_IN"}
    ok(is_bought_in(unknown), "no family and no material is still bought-in")
    ok("defaulted" in bought_in_reason(unknown), "and says so, rather than just returning True")


def test_the_family_reaches_the_part_that_the_makebuy_rule_reads():
    """bought_in_policy reads part["material_family"]. The extract produced it. Nothing
    carried it from one to the other — which is the same shape of gap as the one above, one
    layer down, and would have left the fix inert."""
    from llm_full_extract import normalize_job
    from source_connectors.llm_full_job import apply_full_job_to_pre_estimate
    from bought_in_policy import is_bought_in

    job = {"found": True, "source": "llm_full_extract", "routes": [], "bom": [
        {"part_number": "2085-02", "description": "OUTER TUBE", "qty": 1,
         "material_family": "tube", "thickness_or_section": "25x25x1.5 SHS",
         "cut_length_mm": 340}]}
    normalize_job(job)

    parts = [{"part_number": "2085-02", "description": "OUTER TUBE",
              "normalized_material": "BOUGHT_IN", "textual_operations": ["saw"]}]
    apply_full_job_to_pre_estimate(parts, job)

    eq(parts[0].get("material_family"), "tube", "the family reaches the part record")
    ok(not is_bought_in(parts[0]),
       "so the make/buy rule can see it and the tube keeps its route")
    eq((parts[0].get("section_stock") or {}).get("length_mm"), 340.0,
       "and the section is costed at the length the drawing gives")


def test_a_part_on_two_pages_is_one_part():
    """Both 2085 tubes came out at qty 2 the first run the quantity rollup was able to fire.

    The extract's `bom` is the WHOLE PACK flattened, and the rollup summed every row sharing a
    part number. A tube listed once in the GA's parts table and once more from its own detail
    page's title block became two tubes. Before the bom/parts bridge existed the rollup could
    never fire, so this was latent the whole time and surfaced the moment it was joined up.

    Only an explicit_bom_table row carries a PRINTED quantity. A title-block, note or filename
    row says the part EXISTS — its qty is a default, and adding it is inventing stock.
    """
    from source_connectors.llm_full_job import _rollup_quantities

    job = {"bom": [
        {"part_number": "2085-02", "qty": 1, "source": "explicit_bom_table"},
        {"part_number": "2085-02", "qty": 1, "source": "title_block"},   # its own detail page
        {"part_number": "2085-03", "qty": 1, "source": "title_block"},   # detail page only
        {"part_number": "2085-03", "qty": 1, "source": "filename"},      # and its DXF
    ]}
    got = _rollup_quantities(job)
    eq(got.get("2085-02"), 1, "a table line plus a title block is one tube, not two")
    eq(got.get("2085-03"), 1, "and two non-table mentions are still one tube")

    # A REAL BOM table that genuinely lists a part on two lines still sums — that is a
    # printed quantity twice over, and dropping it would lose stock we are told to buy.
    job2 = {"bom": [
        {"part_number": "FIXING236", "qty": 4, "source": "explicit_bom_table"},
        {"part_number": "FIXING236", "qty": 2, "source": "explicit_bom_table"},
    ]}
    eq(_rollup_quantities(job2).get("FIXING236"), 6,
       "two printed BOM lines for the same fixing are six fixings")

    # Sub-assembly children still multiply by the parent's GA quantity.
    job3 = {"bom": [{"part_number": "2085-SUB", "qty": 2, "source": "explicit_bom_table"}],
            "assemblies": [{"part_number": "2085-SUB",
                            "children": [{"part_number": "2085-09", "qty": 3}]}]}
    eq(_rollup_quantities(job3).get("2085-09"), 6,
       "three per sub-assembly, two sub-assemblies, is six")


def test_round_tube_is_a_section_too():
    """Round tube had no detection path at all, and it is the commonest tube we buy.

    PATH 1 needs 'a x b x t'. PATH 2 needs two unqualified EXT dimensions. A round tube has
    ONE dimension and a wall, so both fall through and return None — no section, no tube
    stock form, no saw and no weld.

    M&S 2085's GA is a single page reading "12.7  1.2 WALL" for the outer tube and "10.0" for
    the inner — 12.7 less two 1.2 walls is a 10.3 bore, so they telescope. Neither tube got a
    section, neither reached the labour path, and a welded three-part bracket booked GBP 2.00
    of labour with no operation against either of them.
    """
    from document_builder import _detect_section_stock

    # The actual text off 2085's GA, as pdfplumber extracted it.
    got = _detect_section_stock("4.0 7.56 INT DRAIN 12.7 1.2 WALL 10.0 12.7 80.0 21.0 INT")
    ok(got is not None, "a round tube is detected at all")
    eq(got["profile_form"], "CHS", "and identified as round, not square")
    eq(got["outside_diameter_mm"], 12.7, "with the diameter beside the wall callout read")
    eq(got["t"], 1.2, "and the wall")
    ok(got.get("review_section_profile"),
       "flagged for verification — this reads a layout convention, not a printed callout")

    # A RECTANGULAR section must never be read as round. PATH 2 still wins where there is a
    # pair of sides, which is the whole reason PATH 3 sits after it.
    rect = _detect_section_stock("60.0 EXT 30.0 EXT 1.5 WALL 1072.0 EXT")
    eq(rect["profile_form"], "RHS", "two EXT sides is still a rectangular section")
    eq(sorted([rect["a"], rect["b"]]), [30.0, 60.0], "with both sides read")

    # And the canonical form is untouched.
    canon = _detect_section_stock("30 x 60 x 1.50mm TUBE 1125")
    eq(canon["profile_form"], "RHS", "the canonical cutting-list form still reads as before")

    # Nothing hollow, nothing returned.
    eq(_detect_section_stock("80.0 70.2 34.0 R1.0 MILD STEEL"), None,
       "no wall callout, no section — an honest gap, not a guessed tube")


def test_a_round_tube_does_not_weigh_what_a_square_one_weighs():
    """The section mass formula was A = outer - inner on the sides of a RECTANGLE, and it was
    the only formula. Round tube carries its outside diameter in a and b, so 12.7 x 1.2 CHS
    computed as a 12.7 square gives 55.2mm2 against a true 43.4mm2 — 27% heavy on every metre
    of every round tube we buy. profile_form is not decoration; the mass switches on it."""
    import math
    D, t = 12.7, 1.2
    d = D - 2 * t
    square = (D * D) - (d * d)
    round_ = math.pi / 4.0 * ((D * D) - (d * d))

    # Reproduce both branches exactly as estimator._resolve_material_cost computes them.
    def _csa(profile_form):
        inner = max(0.0, D - 2.0 * t)
        if str(profile_form or "").upper() == "CHS":
            return max(0.0, math.pi / 4.0 * ((D ** 2) - (inner ** 2)))
        return max(0.0, (D * D) - (inner * inner))

    eq(round(_csa("CHS"), 1), round(round_, 1), "a round tube uses the circular area")
    eq(round(_csa("RHS"), 1), round(square, 1), "a rectangular one is unchanged")
    ok(_csa("RHS") / _csa("CHS") > 1.25,
       f"and the gap is real: {_csa('RHS') / _csa('CHS'):.2f}x, not a rounding difference")

    _src = open(__import__("estimator").__file__, encoding="utf-8").read()
    ok('_ss.get("profile_form") or "").upper() == "CHS"' in _src,
       "the estimator actually branches on it — a formula nothing selects is not a fix")


def test_the_top_assembly_is_not_a_second_bill_of_materials():
    """Every part on 2085 came out at qty 2 — both tubes AND the plate.

    The pack is a single page: one GA whose parts table reads 2085-01 x1, 2085-02 x1,
    2085-03 x1. The extract also correctly describes the top assembly as having those three
    as its children. The rollup then added the child quantity on top of the GA line, and the
    whole job doubled. Filtering by explicit_bom_table did not touch this: the duplication is
    between `bom` and `assemblies`, not within `bom`.

    A genuine SUB-assembly's children are not GA lines — that is what makes them children —
    so they still roll up and still multiply. Only the echo is dropped.
    """
    from source_connectors.llm_full_job import _rollup_quantities

    job = {"bom": [
        {"part_number": "2085-01", "qty": 1, "source": "explicit_bom_table"},
        {"part_number": "2085-02", "qty": 1, "source": "explicit_bom_table"},
        {"part_number": "2085-03", "qty": 1, "source": "explicit_bom_table"},
    ], "assemblies": [
        {"part_number": "2085", "children": [{"part_number": "2085-01", "qty": 1},
                                             {"part_number": "2085-02", "qty": 1},
                                             {"part_number": "2085-03", "qty": 1}]},
    ]}
    got = _rollup_quantities(job)
    eq(got.get("2085-01"), 1, "the plate is one plate")
    eq(got.get("2085-02"), 1, "the outer tube is one tube")
    eq(got.get("2085-03"), 1, "and so is the inner")

    # A REAL sub-assembly still multiplies: its children are not GA lines.
    job2 = {"bom": [{"part_number": "1448-GA", "qty": 2, "source": "explicit_bom_table"}],
            "assemblies": [{"part_number": "1448-GA",
                            "children": [{"part_number": "1448-01", "qty": 3}]}]}
    eq(_rollup_quantities(job2).get("1448-01"), 6,
       "three per sub-assembly, two sub-assemblies on the GA, is six")


def test_a_route_we_read_is_a_route_we_cost():
    """Every rule building labour_parts decided by CLASSIFICATION — is this steel, board,
    wire, a tube, a board panel. A part the engine could not classify got no labour row no
    matter what it was doing, so an operation read off the drawing, or concluded from it, was
    dropped on the floor at the last step before the sheet.

    M&S 2085 spent four runs proving it. Its tubes are round; the section detector only knew
    rectangular; no section meant no `tube` stock form; and no stock form meant that even a
    saw and a weld sitting on the part in black and white would never have reached the labour
    block. Fixing the detector fixes that job. This fixes the shape of it.

    Gated on FABRICATION operations, never handling/assembly, because bought_in_policy strips
    fabrication ops from anything purchased — so a part still carrying one is something we
    make. That is what kept BI-ADHESIVECABLE from getting an Assemble/pack line.
    """
    from bought_in_policy import FABRICATION_OPS

    # Mirror the selection exactly as wb_populate applies it.
    def _routed(bom_parts, already=()):
        _already = {id(p) for p in already}
        out = []
        for p in bom_parts:
            if id(p) in _already:
                continue
            ops = set()
            for k in ("textual_operations", "operations", "inferred_operations"):
                v = p.get(k)
                if isinstance(v, list):
                    ops |= {str(o).strip().lower() for o in v if str(o).strip()}
            if ops & FABRICATION_OPS:
                out.append(p)
        return out

    tube = {"part_number": "2085-02", "textual_operations": ["saw", "welding"]}
    bought = {"part_number": "BI-ADHESIVECABLE", "textual_operations": ["handling", "assembly"]}
    unclassified = {"part_number": "2085-77"}
    got = _routed([tube, bought, unclassified])

    eq([p["part_number"] for p in got], ["2085-02"],
       "a part carrying a saw and a weld is costed, whatever the engine made of its stock")
    ok(bought not in got,
       "a purchased item with only handling/assembly is NOT pulled in — that was the old bug")
    ok(unclassified not in got, "and a part with no route at all adds nothing")

    ok("saw" in FABRICATION_OPS and "welding" in FABRICATION_OPS,
       "saw and welding are fabrication")
    ok("handling" not in FABRICATION_OPS and "assembly" not in FABRICATION_OPS,
       "handling and assembly are not — we do fit bought-in parts and that time is real")

    # BUILT IS NOT WIRED.
    _wb = open(__import__("wb_populate").__file__, encoding="utf-8").read()
    ok("from bought_in_policy import FABRICATION_OPS as _FAB_OPS" in _wb,
       "wb_populate actually applies this when building labour_parts")
    ok("labour_parts.append(p)" in _wb, "and appends the routed part to the labour block")


def test_every_fabricated_part_is_in_the_bill_of_materials():
    """The template costs a nested sheet part in the Sheet Steel block, so it has never
    appeared in the "Bill of Materials (Per Unit)" list. Arithmetically correct, practically
    wrong: on 2085 an estimator opens the sheet, reads the bill of materials, and the main
    part of the job — the bracket plate everything welds into — is not in it.

    Listed now, and listed first. The price is deliberately ZERO because Total Material Cost
    (M92) sums BOM + Wire + Sheet Steel + Other Sheet — a priced duplicate here would double
    that part's material and the sheet would be wrong in a way nobody would spot.
    """
    _wb = open(__import__("wb_populate").__file__, encoding="utf-8").read()
    ok('"_bom_cross_reference": True' in _wb,
       "fabricated parts are added to the BOM list as identifiable cross-reference rows")
    ok('"unit_cost_gbp": 0.0' in _wb,
       "at zero — the Sheet Steel block already carries the money")
    ok("bom_parts = _xref_rows + list(bom_parts)" in _wb,
       "and first, so the bill of materials reads as the parts list it claims to be")
    ok('costed in {_blk_name} below' in _wb,
       "with the block carrying the cost named on the line, so the zero is explained")

    # The total formula is what makes the zero mandatory. If the sheet ever stops summing
    # the Sheet Steel block into M92, this decision has to be revisited — so it is asserted.
    ok("M92 (Total Material Cost) sums every material block" in _wb,
       "the material total still sums every block, which is why the duplicate must be free")


def test_every_word_we_ask_for_is_a_word_we_can_cost():
    """The route vocabulary is a CONTRACT, and it was only honoured on one side.

    The prompts tell the model to return operations from a named list. Five of those words —
    tube_cut, tube_bending, hole_machining, tapping, edge_banding — had no entry in
    wb_populate.OP_NAME_MAP, so the model would return exactly the word it was told to use
    and the workbook would not know which department it belonged to.

    tube_cut is the operation M&S 2085's two tubes need. It was on the asking side of the
    contract and missing from the paying side, so that route could never have been costed
    however perfectly it was read. Routes are what this engine is FOR; a word we ask for and
    cannot pay for is work silently deleted at the last inch.
    """
    from llm_full_extract import ROUTE_OPERATIONS
    from wb_populate import OP_NAME_MAP

    unmapped = [op for op in ROUTE_OPERATIONS if op not in OP_NAME_MAP]
    eq(unmapped, [],
       f"every operation the prompts ask for maps to a department (unmapped: {unmapped})")

    eq(OP_NAME_MAP["tube_cut"], "Tube", "cutting a tube to length is the Tube department")
    eq(OP_NAME_MAP["tube_bending"], "Tubebend",
       "and bending one is Tubebend — SDI uses a tube-bender, not a press brake")

    # ONE list, read by both prompts. Written out longhand in each, they drift, and the
    # drift is invisible until a route goes missing from a sheet.
    import dxf_llm_interpret as dli
    import llm_full_extract as lfe
    _joined = ", ".join(ROUTE_OPERATIONS)
    ok(_joined in dli._PROMPT,
       "the DXF prompt is BUILT from the shared list, not a copy of it that can drift")
    # The PDF passes had NO vocabulary at all — they asked for routes and left `operation`
    # free text, so "Cut to length" or "MIG weld" came back and matched no department. Both
    # PDF passes share _COMMON_RULES, so stating it once covers transcription and inference.
    ok(_joined in lfe._COMMON_RULES,
       "and both PDF passes are told the same list — free-text operations match no department")
    for _name, _prompt in (("pdf transcription", lfe._PROMPT), ("pdf inference", lfe._INFER_PROMPT)):
        ok("tube_cut" in _prompt, f"the {_name} prompt offers 'tube_cut'")

    # ONE OPERATION, ONE ROUTE LINE. A weld that joins three parts is one line naming three
    # part numbers, not three lines — otherwise the same work is booked three times.
    ok("ONE OPERATION, ONE ROUTE LINE" in lfe._COMMON_RULES,
       "and told not to repeat a joining operation per part")


def test_the_finish_grok_read_reaches_the_part():
    """finish and colour were read, projected onto the LLM part row, and then dropped.

    The extract asks for them per BOM row and Grok returns them — 2085's GA states
    "SURFACE FINISH: POWDER COATED  COLOUR: RAL9006 - WHITE ALUMINIUM". project_row carried
    them across and apply_full_job_to_pre_estimate never looked, so they never reached
    normalized_finish — which is exactly the field the powder gate reads. On 2085 the gate's
    assembly-pointer path saved it; on a pack stating the finish per part in the BOM table,
    the coat would simply not have been costed.
    """
    from llm_full_extract import normalize_job
    from source_connectors.llm_full_job import apply_full_job_to_pre_estimate

    job = {"found": True, "source": "llm_full_extract", "routes": [], "bom": [
        {"part_number": "2085-01", "description": "BRACKET PLATE", "qty": 1,
         "material_family": "metal", "material": "MILD STEEL",
         "thickness_or_section": "1.2mm",
         "finish": "POWDER COATED", "colour": "RAL9006 - WHITE ALUMINIUM"}]}
    normalize_job(job)
    parts = [{"part_number": "2085-01"}]
    apply_full_job_to_pre_estimate(parts, job)

    eq(parts[0].get("normalized_finish"), "POWDER COATED",
       "the finish reaches the field the powder gate reads")
    ok("POWDER COATED" in (parts[0].get("surface_finishes") or []),
       "and its fallback field too")
    eq(parts[0].get("normalized_colour"), "RAL9006 - WHITE ALUMINIUM", "with the colour")

    # A finish the engine already established is not overwritten by a read one.
    parts2 = [{"part_number": "2085-01", "normalized_finish": "RAW"}]
    apply_full_job_to_pre_estimate(parts2, job)
    eq(parts2[0]["normalized_finish"], "RAW", "an established finish stands")


def test_labour_rows_come_out_in_manufacturing_order():
    """2085's sheet read Assemble/pack, Laser, P.Coat. Pack before cut.

    The rows were emitted with sorted(_groups.keys()), and the keys start with the
    department name — so the route came out ALPHABETICALLY. That is not a route, it is a
    word list, and an estimator reading down it cannot sanity-check a sequence that is not
    in sequence.

    The extract already returns `sequence` per route and it was thrown away. Where it is
    present it wins — it is the model reading THIS drawing's order of work. Where it is
    absent, the shop's own order applies: cut, form, weld, coat, pack.
    """
    import re as _re
    _wb = open(__import__("wb_populate").__file__, encoding="utf-8").read()
    _blk = _wb[_wb.index("_SHOP_ORDER = {"):]
    _blk = _blk[:_blk.index("}")]
    order = dict(_re.findall(r'"([^"]+)":\s*(\d+)', _blk))

    for _earlier, _later in (("Laser (Metal)", "Fold"), ("Fold", "Weld (CO2)"),
                             ("Weld (CO2)", "P.Coat"),
                             ("P.Coat", "Assemble/pack (Metal)"),
                             ("Tube", "Weld (CO2)"), ("Saw", "Fold")):
        ok(int(order[_earlier]) < int(order[_later]),
           f"{_earlier} is done before {_later}")

    # Alphabetical would put Assemble/pack first and Laser second — the exact order the
    # 2085 sheet came out in. This is what the fixture is defending against.
    ok(int(order["Assemble/pack (Metal)"]) > int(order["Laser (Metal)"]),
       "pack is not the first operation on the sheet, whatever the alphabet says")

    ok('sorted(_groups.keys(), key=_group_order)' in _wb,
       "and the emit loop actually uses the ordering — a table nothing sorts by is a comment")
    ok('_read = _g.get("route_sequence")' in _wb,
       "with the drawing's own sequence taking precedence when the extract gave one")
    ok('g["route_sequence"] = (_rs if g.get("route_sequence") is None' in _wb,
       "which means it has to be carried onto the group in the first place")


def test_every_department_we_emit_exists_in_the_rate_table():
    """The rate table (H173:K204) is a CLOSED vocabulary and the only one that pays.

    A string it cannot resolve produces a blank department, a zero rate and a zero cost —
    and a zero-cost labour line is indistinguishable, on the sheet, from work nobody
    identified. Silent, always reduces the price, nothing in the output says it happened.

    The engine had two vocabularies and neither was the rate table. Diffed against it,
    OP_NAME_MAP pointed four operations at "Grinding / Deburr" — GRIN — and there IS no GRIN
    row, so every deburr and every linish on every job has been costing nothing.
    """
    from department_codes import code_for, DEPARTMENT_CODES, CODE_TITLES
    from wb_populate import OP_NAME_MAP, OP_NAME_MAP_ACRYLIC

    # EXACT, not merely resolvable. The workbook's LOOKUP matches the title string in
    # column H character for character, so a title that is close enough for our own alias
    # table to understand still produces a zero rate on the sheet. "CNC / Joinery machining"
    # resolved fine here and would have paid nothing there; the row is called "CNC Joinery".
    _exact = {t for t, _c in CODE_TITLES.values()}
    for _name, _m in (("OP_NAME_MAP", OP_NAME_MAP), ("OP_NAME_MAP_ACRYLIC", OP_NAME_MAP_ACRYLIC)):
        bad = {op: t for op, t in _m.items() if t not in _exact}
        eq(bad, {}, f"every {_name} title is column-H exact (bad: {bad})")

    # The specific ones that were free: GRIN is not a department at all.
    eq(code_for(OP_NAME_MAP["deburring"]), "MANM",
       "deburring pointed at GRIN, which does not exist; BENC is a JOINERY bench, so metal "
       "deburring goes to manual labour (metal)")
    # Confirmed with the estimators: general edge work after laser/saw/punch is MANM;
    # linishing a WELD BEAD is its own department. Costing a dressed weld as deburring
    # books shop time against the wrong machine at the wrong rate.
    for _op in ("deburring", "fettling", "metal deburr", "edge deburr"):
        eq(code_for(_op), "MANM", f"'{_op}' is manual labour (metal)")
    for _op in ("linishing", "weld dress", "dress welds"):
        eq(code_for(_op), "DRES", f"'{_op}' is weld dressing, not deburring")
    eq(code_for("finishing"), None,
       "'finishing' is ambiguous — it can mean deburring or powder coating, so it stays "
       "loud rather than quietly moving money between departments")
    ok(code_for("bench work") == "BENC" and code_for("deburring") != "BENC",
       "BENC is left for joinery")
    eq(OP_NAME_MAP["hole_machining"], "Drill (Acrylic)",
       "and drilling/tapping uses the DRIL row, whose title says Acrylic but is the row the "
       "shop books metal drilling against")
    eq(code_for("Assemble/pack (Acrylic)"), "PACP",
       "acrylic assembly is PACP — it was forced to PACM on a guess")
    eq(CODE_TITLES["SALV"][0], "Salvagnini",
       "SALV is a panel bender, not salvage/rework — the guess would have sent reworked "
       "parts to a forming machine")
    ok("GRIN" not in DEPARTMENT_CODES and "TAP" not in DEPARTMENT_CODES
       and "COUN" not in DEPARTMENT_CODES and "HAND" not in DEPARTMENT_CODES,
       "and GRIN/TAP/COUN/HAND are confirmed absent, which is why they cost nothing")

    # Every title we would ever write maps back into the closed set, and every one is now
    # confirmed against the live table rather than plausibly guessed.
    for _code, (_title, _confirmed) in CODE_TITLES.items():
        ok(_code in DEPARTMENT_CODES, f"{_code} is a real rate-table code")
        eq(code_for(_title), _code, f"'{_title}' resolves back to {_code}")
        ok(_confirmed, f"{_code}'s title was read off the rate table, not inferred")
    eq(len(CODE_TITLES), len(DEPARTMENT_CODES),
       "every code in the table has a title, and no title exists for a code that does not")


def test_free_text_from_a_model_still_lands_on_a_department():
    """A model writing about manufacturing produces English — "Cut to length", "MIG weld",
    "Laser cut" — sensible to a person and invisible to every LOOKUP. The prompt now asks
    for the code, but a prompt is a request, not a guarantee, so the alias table catches
    what comes back anyway. A wrong code an estimator can correct; an unrecognised one is
    silently free."""
    from wb_populate import _map_operation
    from department_codes import code_for, unresolved_operations

    for _text, _code in (("tube_cut", "TUBE"), ("Cut to length", "SAW"), ("MIG weld", "WELD"),
                         ("CO2 weld", "WELD"), ("Laser cut", "LASM"), ("press brake", "FOLD"),
                         ("Saw tube", "SAW"), ("edge banding", "EDGE"),
                         ("tapping", "DRIL"), ("hole_machining", "DRIL"),
                         ("Assemble/pack (Acrylic)", "PACP"), ("Salvagnini", "SALV"),
                         ("CNC Joinery", "CNCJ"), ("Wet Spray", "SPRY")):
        eq(code_for(_text), _code, f"'{_text}' is understood as {_code}")

    # A model answering in codes, as instructed, is understood too.
    for _c in ("TUBE", "WELD", "LASM", "FOLD", "P/C", "PACM"):
        eq(code_for(_c), _c, f"the code {_c} is its own alias")

    # And it reaches an actual workbook title through the real mapping function.
    eq(_map_operation("MIG weld", False), "Weld (CO2)",
       "free text resolves all the way to a title the sheet can price")
    eq(_map_operation("tube_cut", False), "Tube", "and so does the token that started this")

    # Unrecognised stays unrecognised. None is the point: it is the difference between
    # "we could not price this" and a zero that reads as "this work does not exist".
    eq(_map_operation("frobnicate", False), None, "nonsense is not quietly mapped to something")
    eq(unresolved_operations(["welding", "frobnicate", "MIG weld"]), ["frobnicate"],
       "and can be reported rather than swallowed")


def test_the_prompt_asks_for_the_vocabulary_that_pays():
    """The prompts were asking for free text, so the model had no way to answer costably."""
    import llm_full_extract as lfe
    from department_codes import DEPARTMENT_CODES

    for _code in ("TUBE", "WELD", "LASM", "FOLD", "TBEN", "DRIL", "P/C", "PACM"):
        ok(_code in lfe._COMMON_RULES, f"the prompt names {_code}")
    ok(", ".join(sorted(DEPARTMENT_CODES)) in lfe._COMMON_RULES,
       "the whole closed list is built from the code table, not copied where it can drift")
    ok("silently free" in lfe._COMMON_RULES,
       "and the model is told what happens when it invents one, which is the part that "
       "makes the rule stick")


def test_title_variants_resolve_without_guessing():
    """Exact == on a workbook title is brittle. Sheets drift by a space, a slash, an "&" for
    "and", a capital, a bracket — and each drift is another silent zero.

    Normalised, NOT fuzzy. No Levenshtein and no token-set ratio anywhere near this: those
    can land on the wrong department and still produce a cost, which is worse than a loud
    None. "Laser (Acrylic)" and "Laser (Metal)" are one token apart and must never be
    confused; a near-miss department is never right, so a number from one is never traceable.
    """
    from department_codes import code_for, CODE_TITLES, _norm_title

    # Punctuation, spacing and case drift — absorbed.
    for _variant, _code in (
            ("Assemble / pack (Acrylic)", "PACP"),
            ("assemble  and  pack ( metal )", "PACM"),
            ("CNC / Joinery machining", "CNCJ"),
            ("Drilling / Tapping", "DRIL"),
            ("Spray / Wet Paint", "SPRY"),
            ("Salvage / Rework", "SALV"),
            ("Packaging - Carton", "PACP"),
            ("P Coat", "P/C"),
            ("WET SPRAY", "SPRY"),
            ("Weld  (CO2)", "WELD"),
            ("MCJ", "MC J")):
        eq(code_for(_variant), _code, f"'{_variant}' resolves to {_code}")

    # Still loud on anything we do not actually recognise. GRIN is the important one: it was
    # never a department, so it must NOT become costable — a saved job that names it is read
    # back through LEGACY_TITLES, which is a different question from what to charge.
    eq(code_for("Grinding / Deburr"), None, "a department that never existed stays unresolved")
    eq(code_for("Some invented op"), None, "and so does anything invented")
    eq(code_for(""), None, "and an empty operation is not a department")

    # NORMALISATION MUST NOT MERGE TWO DEPARTMENTS. The whole reason fuzzy is refused is
    # that a wrong-but-costed department is worse than none; a collision here would do
    # exactly that, silently, by dict order.
    _seen = {}
    for _c, (_t, _ok) in CODE_TITLES.items():
        _n = _norm_title(_t)
        ok(_n not in _seen,
           f"{_c} and {_seen.get(_n)} both normalise to {_n!r} — the match would be ambiguous")
        _seen[_n] = _c
    ok(_norm_title("Laser (Acrylic)") != _norm_title("Laser (Metal)"),
       "the acrylic and metal lasers stay distinct through normalisation")
    ok(_norm_title("Manual labour (Acrylic)") != _norm_title("Manual labour (Metal)"),
       "and so do the two manual-labour departments")

    # The canonical string is still what gets WRITTEN, so the workbook's own LOOKUP hits.
    from department_codes import title_for
    eq(title_for("CNC / Joinery machining"), "CNC Joinery",
       "a drifted title in resolves as the canonical title out")
    eq(title_for("Packaging - Carton"), "Assemble/pack (Acrylic)",
       "including an old engine title we wrote ourselves")


def test_a_section_callout_belongs_to_the_part_it_describes():
    """M&S 2085 is one sheet: a GA carrying the parts table and every detail view. Its parts
    therefore all share the same page text, so the "12.7  1.2 WALL" printed for the tubes was
    offered to every part on the drawing — and the BRACKET PLATE took it.

    The plate left the Sheet Steel block, its nesting and laser calculator went blank, and
    the one measured part in the job stopped being costed as sheet at all. The run reported
    "0 steel" where it had reported 1.

    Two guards, both on evidence: measured flat geometry means sheet, and the
    lower-confidence WALL detections need the part to look like section stock on its own
    account. The canonical cutting-list form is unambiguous and still applies to anything.
    """
    from document_builder import _detect_section_stock, _apply_post_build_fixes

    _ga = "4.0 7.56 INT DRAIN 12.7 1.2 WALL 10.0 12.7 80.0 MATERIAL: MILD STEEL"
    ok(_detect_section_stock(_ga) is not None,
       "the tubes' section is still read off the drawing")

    def _run(part):
        """Drive the REAL post-build pass over a one-page pack. The first version of this
        fixture re-implemented the guard and its mutation passed — the same defect shape as
        the bug being tested."""
        summary = {"pages": [{"page": 1, "text": _ga, "page_number": 1}]}
        p = dict(part, pages=[1], page_roles=["assembly"], materials=["MILD STEEL"])
        _apply_post_build_fixes([p], summary)
        return p.get("section_stock")

    eq(_run({"part_number": "2085-01", "description": "BRACKET PLATE",
             "material_family": "metal", "flat_pattern_detected": True}), None,
       "a part with a measured flat pattern is sheet — it does not become tube because a "
       "tube is dimensioned elsewhere on the same drawing")

    ok(_run({"part_number": "2085-02", "description": "OUTER TUBE",
             "material_family": "tube"}) is not None,
       "the tube still gets its section")

    eq(_run({"part_number": "2085-09", "description": "BACK PANEL",
             "material_family": "metal"}), None,
       "and neither does an unmeasured plate, on shared page text")

    # THE FLAT-GEOMETRY GUARD ON ITS OWN. The two guards overlap on a part called "BRACKET
    # PLATE", so removing either still gives the right answer there and a mutation passes.
    # This is the case only measurement settles: a flat part whose NAME reads like section.
    # On a job literally called "Rail Nav Bracket" that is not a contrived example.
    eq(_run({"part_number": "2085-11", "description": "RAIL MOUNT BRACKET",
             "material_family": "metal", "flat_pattern_detected": True}), None,
       "a measured flat pattern beats a description that merely sounds like section stock")
    ok(_run({"part_number": "2085-12", "description": "RAIL MOUNT BRACKET",
             "material_family": "metal"}) is not None,
       "and without the measurement, the same description DOES take the section — which is "
       "what makes the guard above the thing doing the work")

    # PATH 1 is a cutting-list line naming the part's own stock. Unambiguous, and it applies
    # regardless — narrowing that would lose every tube the reader currently gets right.
    ok(_detect_section_stock("30 x 60 x 1.50mm TUBE 1125")["detection_path"]
       == "canonical_profile",
       "a canonical cutting-list callout is not one of the corroborated paths")


def test_a_routed_operation_with_no_cost_model_still_reaches_the_sheet():
    """The last gate, and the one every fix so far was blocked behind.

    The labour loop built its operation list from labour_estimate.costs_gbp alone — the
    operations the estimator managed to put a NUMBER against. An operation on the part's
    route with no time model behind it was not merely uncosted, it was invisible: no row, no
    flag, nothing on the sheet saying the job includes it.

    2085 proved every other link works. Both tubes carry tube_cut and tube_bending, the
    plate carries folding, all three map to real departments (Tube, Tubebend, Fold), and the
    invariant reported them "named but not priced" while the sheet showed five labour rows
    on one part. The route was right the whole way down and stopped here.
    """
    from wb_populate import _map_operation, routed_operations_without_cost as _routed

    def _ops_for(pe):
        """Calls the REAL helper. The first version of this mirrored the logic instead, and
        the mutation that deleted the fabrication-only guard passed clean."""
        costs = (pe.get("labour_estimate") or {}).get("costs_gbp") or {}
        return list(costs.keys()) + _routed(pe, costs)

    tube = {"part_number": "2085-02",
            "labour_estimate": {"costs_gbp": {"welding": 0.31}},
            "textual_operations": ["tube_cut", "tube_bending", "welding", "handling"]}
    got = _ops_for(tube)
    ok("tube_cut" in got, "a routed tube cut with no cost model still gets a row")
    ok("tube_bending" in got, "and so does the bend")
    eq(got.count("welding"), 1, "an operation the estimator costed is not duplicated")
    ok("handling" not in got,
       "handling with no cost is NOT a missing row — that is how a bought-in got an "
       "Assemble/pack line it should never have had")

    # And they map to departments the rate table can pay.
    eq(_map_operation("tube_cut", False, "tube"), "Tube", "tube_cut is the Tube department")
    eq(_map_operation("tube_bending", False, "tube"), "Tubebend", "and the bend is Tubebend")
    eq(_map_operation("folding", False, ""), "Fold", "the plate's fold is Fold")

    # A part with neither a cost nor a route still produces nothing.
    eq(_ops_for({"part_number": "X"}), [], "nothing routed and nothing costed is no row")

    _wb = open(__import__("wb_populate").__file__, encoding="utf-8").read()
    ok("_routed_extra = routed_operations_without_cost(" in _wb
       and "ops = ops + _routed_extra" in _wb,
       "the labour loop actually takes the union — this is the fifth time a fix has been "
       "built and not wired, and it is the one that decides whether the route is costed")
    ok("throughput default so the work appears on the sheet" in _wb,
       "and says so, because a time nobody modelled must not read as a time somebody agreed")


def test_a_measured_zero_is_not_undone_by_a_reading():
    """The fold, and the reason it must stay off 2085-01.

    A DXF flat pattern that measures zero bend lines is a MEASUREMENT: the part does not
    fold. That strip existed, but it deleted the operation and left nothing to say it had —
    so the LLM route, which reads formed walls off the drawing views, simply added `folding`
    back. Harmless while a routed operation could not reach the sheet. Not harmless now: a
    routed operation without a cost model gets a labour row, so the resurrected fold would
    be CHARGED FOR. An explicit zero overwritten by a conclusion is the exact failure that
    took three commits on 12120.

    And not silent. The model saw walls, the reader measured no bends, and one of them is
    wrong about a part being costed — commonly a flat exported as a block, where the bend
    layer is real but unreadable. That belongs in front of an estimator.
    """
    from source_connectors.llm_full_job import apply_routes_to_parts
    from wb_populate import routed_operations_without_cost as _routed

    part = {"part_number": "2085-01", "description": "BRACKET PLATE",
            "flat_pattern_detected": True,
            "operations_ruled_out": {"folding": "DXF flat pattern measured 0 bend lines"}}
    job = {"routes": [
        {"operation": "folding", "part_numbers": ["2085-01"], "inferred": True},
        {"operation": "laser_cutting", "part_numbers": ["2085-01"], "inferred": False},
    ]}
    apply_routes_to_parts([part], job)

    ok("folding" not in (part.get("textual_operations") or []),
       "the measurement stands — a fold the DXF ruled out is not added back by a reading")
    ok("laser_cutting" in (part.get("textual_operations") or []),
       "and everything the measurement did NOT answer still comes through")
    ok(any("not showing its bend lines" in str(f) for f in (part.get("review_flags") or [])),
       "with the disagreement put in front of a person, because a flat exported as a block "
       "measures zero bends and folds perfectly well in the shop")

    # The last gate before a labour row refuses it too. Belt and braces, because this is
    # the one that spends money.
    eq(_routed({"labour_estimate": {"costs_gbp": {}},
                "textual_operations": ["folding", "tube_cut"],
                "operations_ruled_out": {"folding": "DXF measured 0 bend lines"}}),
       ["tube_cut"],
       "a ruled-out operation is not a missing row — it is an answered question")

    # A part with no ruling is unaffected: this must not become a blanket suppression.
    plain = {"part_number": "2085-77"}
    apply_routes_to_parts([plain], {"routes": [
        {"operation": "folding", "part_numbers": ["2085-77"], "inferred": True}]})
    ok("folding" in (plain.get("textual_operations") or []),
       "a part nothing measured still takes the inferred fold")

    # AND THE RULING HAS TO BE RECORDED IN THE FIRST PLACE.
    _djm = open(__import__("drawing_job_merge").__file__, encoding="utf-8").read()
    ok('setdefault("operations_ruled_out", {})["folding"]' in _djm,
       "the zero-bend strip records WHAT it ruled out, not just that it removed something")


def test_the_route_and_the_cost_live_on_different_records():
    """2085's tube_cut was visible to the invariant and invisible to the sheet, and both
    were right about the record they were reading.

    estimator.estimate_part builds a costed part_estimate and never copies
    textual_operations onto it. estimate_summary.part_estimates — which is what wb_populate
    walks — therefore has no route at all. The route stays on summary["parts"], which is
    where invariants._parts falls through to and finds it.

    So `operation_named_but_not_priced: tube_cut` was reported by a check reading one
    collection while the labour block read another, and every fix aimed at the labour block
    was aimed past the problem. Third time this wrong-record shape has cost a day: _parts()
    handing geometry to a price check, the schema saying `bom` while the engine read `parts`,
    and now this.
    """
    from wb_populate import route_operations_by_part, routed_operations_without_cost

    summary = {"parts": [
        {"part_number": "2085-02", "textual_operations": ["tube_cut", "welding"]},
        {"part_number": "2085-01", "textual_operations": ["folding", "laser_cutting"],
         "operations_ruled_out": {"folding": "DXF flat pattern measured 0 bend lines"}},
    ]}
    routes = route_operations_by_part(summary)
    eq(routes["2085-02"], ["tube_cut", "welding"], "the tube's route is found on the raw record")
    ok("folding" not in routes["2085-01"],
       "and a ruling by measurement travels WITH the route — otherwise the sheet re-adds "
       "exactly what the measurement removed")
    ok("laser_cutting" in routes["2085-01"], "while everything unruled comes through")

    # manufacturing_writeup.parts is the other place a route can live; both are read.
    summary2 = {"manufacturing_writeup": {"parts": [
        {"part_number": "X-1", "textual_operations": ["saw"]}]}}
    eq(route_operations_by_part(summary2).get("X-1"), ["saw"],
       "the writeup parts are read too — a route in either place reaches the sheet")

    # The costed record carries the COST and nothing else; the union needs both halves.
    costed = {"part_number": "2085-02", "labour_estimate": {"costs_gbp": {"welding": 1.0}}}
    eq(routed_operations_without_cost(costed, None, routes["2085-02"]), ["tube_cut"],
       "the routed operation with no cost model is found, which it cannot be from the "
       "costed record alone")
    eq(routed_operations_without_cost(costed, None, None), [],
       "and without the bridge there is nothing to find — which is exactly what the sheet "
       "was seeing")

    # BUILT IS NOT WIRED — sixth time. Both call sites must use it.
    _wb = open(__import__("wb_populate").__file__, encoding="utf-8").read()
    ok("_route_by_pn = route_operations_by_part(summary)" in _wb,
       "the bridge is built once in populate_workbook")
    ok("_ops = set(_route_by_pn.get(" in _wb,
       "the labour_parts gate uses it, so a routed part reaches the labour block at all")
    ok("_route_by_pn.get(_pn.strip().upper())" in _wb,
       "and the union uses it, so its operations become rows")


def test_the_bom_price_column_is_a_material_column():
    """M&S 2085's tubes showed GBP 19.25 each in the Bill of Materials Price column, and not
    one penny of it is tube.

    The canonical JSON breaks it down: GBP 17.80 powder-coating labour, GBP 0.71 weld,
    GBP 0.33 dress welds, GBP 0.42 handling. Material: unpriced. The price chain fell back to
    extended_total_cost_gbp — the part's WHOLE unit cost, labour included — so a fabricated
    part with no material price displayed its labour as material.

    It read as an impossible material price (60 metres of stock per bracket, which is what I
    spent a round deriving from it) and it is about to become worse than misleading: now that
    a routed operation reaches the labour block, those same operations get their own rows and
    the labour is counted twice.
    """
    from wb_populate import _bom_line_price

    # A fabricated tube: no material price, a labour bundle. The bundle must NOT appear.
    tube = {"part_number": "2085-02", "quantity": 1,
            "labour_estimate": {"costs_gbp": {"powder_coating": 17.80, "welding": 0.71,
                                              "dress_welds": 0.33, "handling": 0.42}},
            "extended_total_cost_gbp": 19.25}
    eq(_bom_line_price(tube), None,
       "a part whose material nobody could price is UNPRICED — not priced at its labour")

    # Material that IS known still comes through, by every legitimate route.
    eq(_bom_line_price({"part_number": "X", "quantity": 1, "unit_cost_gbp": 3.13}), 3.13,
       "a bought-in unit price is used")
    eq(_bom_line_price({"part_number": "Y", "quantity": 2,
                        "extended_material_cost_gbp": 8.0,
                        "labour_estimate": {"costs_gbp": {"welding": 5.0}},
                        "extended_total_cost_gbp": 40.0}), 4.0,
       "extended MATERIAL divided by qty — never the total, even when a total exists")

    # The whole-total fallback survives where it was written to work: a genuine bought-in
    # with no labour at all, whose price is only recorded as an extended figure.
    eq(_bom_line_price({"part_number": "BI-KNOB", "quantity": 4,
                        "extended_total_cost_gbp": 10.0}), 2.5,
       "a bought-in with no labour still resolves from its extended total")

    # "HAS NO LABOUR" IS NOT "IS A BOUGHT-IN", and the difference cuts both ways.
    # A bought-in DOES carry handling — bought_in_policy leaves it deliberately, because
    # fitting a purchased component is real bench time — so the first version of this rule
    # blanked a priced fixing. Material is the unit total less the labour on it, which needs
    # no classification and is right for both kinds of part.
    eq(_bom_line_price({"part_number": "FIXING236", "quantity": 1,
                        "unit_total_cost_gbp": 4.20,
                        "labour_estimate": {"costs_gbp": {"handling": 0.42}}}), 3.78,
       "a bought-in with handling keeps its material price, less the handling")

    # ONE CHAIN. The row was written by a hand-copy of this helper carrying the same bug, so
    # fixing the helper alone would have corrected an overflow sum nobody sees and left the
    # sheet showing GBP 19.25.
    _wb = open(__import__("wb_populate").__file__, encoding="utf-8").read()
    eq(_wb.count("price = _bom_line_price(pe)"), 1,
       "the BOM row uses the helper rather than a copy of it")
    ok("ext = _safe(pe.get(\"extended_total_cost_gbp\"))" not in _wb,
       "and the duplicated whole-total fallback is gone")


def test_a_tube_is_not_laser_profiled():
    """2085's tubes carried laser_cutting, inherited from the shared assembly page the
    plate's route was read off. A tube has no flat blank to profile — it is sawn to length
    and welded in — and with routed operations now reaching the labour block that booked a
    laser cut on both tubes.

    THE FIRST VERSION OF THIS FIXTURE PASSED WHILE THE FIX WAS INERT. It handed
    _is_spurious_operation a stock_form of "tube", which is precisely the fact the real
    record does not have: estimator stamps stock_form only inside the branches that produce
    a PRICE, and these tubes could not be priced. So the fixture supplied the classification
    the production data was missing, and proved nothing about the production data.

    This drives the real resolution: evidence off the raw record, through the bridge.
    """
    from wb_populate import _is_spurious_operation, tube_part_numbers

    # The records as they ACTUALLY are on this job: costed record with an empty
    # material_estimate, section_stock only on the raw one.
    summary = {
        "parts": [
            {"part_number": "2085-02", "description": "OUTER TUBE",
             "section_stock": {"a": 12.7, "b": 12.7, "t": 1.2, "profile_form": "CHS"}},
            {"part_number": "2085-01", "description": "BRACKET PLATE",
             "flat_pattern_detected": True},
        ],
        "estimate_summary": {"part_estimates": [
            {"part_number": "2085-02", "material_estimate": {}},
            {"part_number": "2085-01", "material_estimate": {"stock_form": "sheet"}},
        ]},
    }
    _tubes = tube_part_numbers(summary)
    ok("2085-02" in _tubes,
       "a tube is identified by its SECTION, not by whether its material happened to price")
    ok("2085-01" not in _tubes, "and a flat plate is not")

    _sf = "tube" if "2085-02" in _tubes else None

    # A TUBE IS NOT LASER-PROFILED — BUT IT IS STILL CUT.
    #
    # The first version of this rule DROPPED laser and guillotine on tube stock. That
    # removed the profile cut which never happens AND the cut-to-length which always does,
    # and 2085's tubes came out of it with no cutting operation of any kind: the sheet read
    # as though nobody cut them. The operation is misplaced, not impossible, so it is
    # relocated to the tube department rather than deleted.
    from wb_populate import _map_operation
    for _op in ("laser_cutting", "laser", "guillotine"):
        ok(not _is_spurious_operation(_op, _sf, "MILD_STEEL"),
           f"'{_op}' on a tube is misplaced, not impossible — dropping it makes the cut free")
        _dept = _map_operation(_op, False, _sf)
        ok(_dept and "laser" not in _dept.lower(),
           f"'{_op}' must never book a LASER against a tube, got {_dept!r}")
        eq(_dept, "Tube", f"'{_op}' on a tube is a tube-department cut")

    # Punch stays spurious: a hole through a tube wall is hole work, not a cut to length,
    # and remapping it would invent a second cut on every punched tube.
    ok(_is_spurious_operation("punch", _sf, "MILD_STEEL"),
       "punching a tube is not cutting it to length")

    for _op in ("saw", "tube_cut", "welding", "tube_bending", "powder_coating"):
        ok(not _is_spurious_operation(_op, _sf, "MILD_STEEL"),
           f"'{_op}' is real work on a tube and must survive")

    # The plate is unaffected — this is keyed on evidence, not on a part number.
    ok(not _is_spurious_operation("laser_cutting", None, "MILD_STEEL"),
       "a sheet part is still lasered")

    # material_family alone is enough where no section was read.
    eq(tube_part_numbers({"parts": [{"part_number": "X-1", "material_family": "tube"}]}),
       {"X-1"}, "the BOM family identifies a tube the reader found no profile for")

    # And the call site uses it, rather than reading a field the real record lacks.
    _wb = open(__import__("wb_populate").__file__, encoding="utf-8").read()
    ok("_tube_pns = tube_part_numbers(summary)" in _wb, "the evidence set is built")
    ok("if not _sf and _pn.strip().upper() in _tube_pns:" in _wb,
       "and consulted where stock_form is blank — which is the case this exists for")



def test_section_stock_is_not_priced_like_sheet():
    """The section/tube path costs by mass x MATERIAL_PRICE_GBP_PER_KG — the FLAT-PRODUCT
    rate, GBP 0.80/kg, matching the sheet's own GBP 900/tonne. Small-diameter tube is not
    sold on that basis.

    12.7 x 1.2mm mild steel retails around GBP 2.48/m at 0.34 kg/m. That is GBP 7.29/kg,
    roughly EIGHT TIMES the flat rate. On 2085 it is the difference between GBP 0.05 and
    GBP 0.37 for the two tubes; on a job built from section it is the difference between a
    quote and a loss, and it would have been silent — the number looks like a material cost
    because it is one, just the wrong one.

    The rate is left as None on purpose. Inventing one is how "Salvage / Rework" happened.
    While it is None the engine prices at the flat rate and FLAGS EVERY LINE, which is
    visible and wrong rather than invisible and wrong.
    """
    import math
    import config

    # The CHS mass formula is confirmed against the same supplier figure: 0.34 kg/m.
    D, t = 12.7, 1.2
    csa = math.pi / 4.0 * ((D ** 2) - ((D - 2 * t) ** 2))
    kg_per_m = csa * 7850 / 1e6
    ok(abs(kg_per_m - 0.34) < 0.005,
       f"the round-tube mass matches the supplier's own kg/m (got {kg_per_m:.4f})")

    # And the rate does not.
    _flat = 0.80
    _real = 2.48 / 0.34
    ok(_real / _flat > 5,
       f"the flat-product rate under-reads section by {_real / _flat:.1f}x — not a rounding gap")

    ok(hasattr(config, "SECTION_STOCK_PRICE_GBP_PER_KG"),
       "there is a lever for the real rate")
    eq(getattr(config, "SECTION_STOCK_PRICE_GBP_PER_KG"), None,
       "left unset — a guessed rate is what put a panel bender in the salvage department")

    _est = open(__import__("estimator").__file__, encoding="utf-8").read()
    # The exact READ, not just the name — the name also appears in the flag text, so a
    # substring check passed while the lever was ignored entirely.
    ok('getattr(config, "SECTION_STOCK_PRICE_GBP_PER_KG", None)' in _est,
       "the estimator actually reads the lever")
    ok("price_per_kg = _sec_rate" in _est, "and uses it as the rate when it is set")
    ok("this line UNDER-READS" in _est,
       "and flags every section line while it is unset, so the gap is visible on the part")
    ok("if isinstance(_sec_rate, dict):" in _est,
       "a per-material rate is accepted too — steel, stainless and aluminium tube do not "
       "cost the same")


def test_the_model_supplies_the_cut_length_the_drawing_never_printed():
    """A tube's cut length is the one number the section path cannot do without, and on a
    GA-only pack nothing prints it. M&S 2085 dimensions its tubes on the views, so both came
    through with a profile and no length — unpriced, and their coated area zero, which
    under-reads the powder too.

    The model has it. bbox_mm is already captured per part by the native analyser and
    already reaches NativePart; for a straight section its LONGEST dimension IS the cut
    length. Measured, rank solidworks_api, not inferred — the answer was in the extract the
    whole time and nothing read it.
    """
    from source_connectors.solidworks import apply_native_to_pre_estimate, NativeJob, NativePart

    def _run(part, bbox, pn="2085-02"):
        nat = NativePart(part_number=pn, bbox_mm=bbox)
        job = NativeJob(part_signals={pn: nat}, found=True)
        apply_native_to_pre_estimate([part], job)
        return part

    # A tube with a profile and no length takes the model's longest dimension.
    tube = _run({"part_number": "2085-02", "description": "OUTER TUBE",
                 "section_stock": {"a": 12.7, "b": 12.7, "t": 1.2, "profile_form": "CHS"}},
                [150.0, 12.7, 12.7])
    eq((tube.get("section_stock") or {}).get("length_mm"), 150.0,
       "the cut length comes from the model's bounding box")
    eq((tube.get("section_stock") or {}).get("length_source"), "solidworks_api",
       "and is stamped as measured, not inferred")
    ok(any("does not print one" in str(f) for f in (tube.get("review_flags") or [])),
       "with the estimator told where it came from")

    # A length the drawing DID state is never overwritten by an envelope.
    stated = _run({"part_number": "2085-02",
                   "section_stock": {"a": 25.0, "b": 25.0, "t": 1.5, "length_mm": 340.0}},
                  [900.0, 25.0, 25.0])
    eq((stated.get("section_stock") or {}).get("length_mm"), 340.0,
       "a stated cut length stands — the bounding box does not overwrite a reading")

    # A part with no section is not given one. This must not invent tube.
    plate = _run({"part_number": "2085-01", "description": "BRACKET PLATE"},
                 [90.88, 80.0, 1.2], pn="2085-01")
    eq(plate.get("section_stock"), None,
       "a part with no profile is not turned into a section by having a bounding box")

    # A BENT tube's bounding box is its ENVELOPE, not its developed length, so it
    # under-reads — which is how a job gets quoted short. Used, and said out loud.
    bent = _run({"part_number": "2085-02",
                 "section_stock": {"a": 12.7, "b": 12.7, "t": 1.2},
                 "textual_operations": ["tube_bending", "welding"]},
                [150.0, 60.0, 12.7])
    eq((bent.get("section_stock") or {}).get("length_mm"), 150.0, "the envelope is used")
    ok(any("UNDER-READS" in str(f) for f in (bent.get("review_flags") or [])),
       "and flagged as the envelope of a bent part, not its developed length")


def test_a_native_extract_for_another_job_is_refused():
    """M&S 2085 was costed against 12120_sw_extract_v7.json — top assembly 12120-01-GA, a
    different customer's job, bound through a persistent SDI_SW_EXTRACT_JSON. The connector
    took it, matched nothing, applied nothing, and stamped a native_extract_partial BLOCKER
    describing 12120's unreadable model onto 2085's estimate. The estimator was told a
    released component might be missing, about a pack with nothing to do with the job.

    Zeros are survivable. What it opens is not: SDI numbers sequentially, so -01/-02/-03
    collide across jobs constantly, and a colliding pair would have had one job's bounding
    boxes, materials and bend counts written onto the other's parts — silently, at the
    HIGHEST rank in the waterfall.

    The test is evidence, not names: a folder can be called anything and a job renumbered,
    but an extract describing THIS job matches at least one of its parts.
    """
    from source_connectors.solidworks import (extract_is_for_this_job, NativeJob, NativePart,
                                              apply_native_to_pre_estimate)

    _parts = [{"part_number": "2085-02"}, {"part_number": "2085-03"}]
    _foreign = NativeJob(found=True, meta={"top_assembly": "12120-01-GA"},
                         part_signals={"12120-01-01M": NativePart(part_number="12120-01-01M"),
                                       "12120-01-103": NativePart(part_number="12120-01-103")})
    _v = extract_is_for_this_job(_parts, _foreign)
    ok(not _v["belongs"], "an extract matching none of this job's parts does not belong to it")
    eq(_v["matched"], 0, "and reports that nothing matched")
    eq(_v["top_assembly"], "12120-01-GA",
       "naming the job it DOES describe, so the estimator can see what went wrong")

    _own = NativeJob(found=True, meta={"top_assembly": "2085-GA"},
                     part_signals={"2085-02": NativePart(part_number="2085-02")})
    ok(extract_is_for_this_job(_parts, _own)["belongs"],
       "a single genuine match is enough — a partial extract is still this job's")

    # A pack with no part signals at all is not evidence of the wrong job, just of a thin
    # extract. Refusing that would throw away the BOM rows it may still carry.
    ok(extract_is_for_this_job(_parts, NativeJob(found=True))["belongs"],
       "an empty extract is not a foreign one")

    # THE INDEX IS SHARED. This check must run the SAME matcher the fold runs, or it
    # verifies its own copy rather than the code — the mistake made twice in fixtures today.
    _sw = open(__import__("source_connectors.solidworks", fromlist=["x"]).__file__,
               encoding="utf-8").read()
    eq(_sw.count("exact, tail, lead = _native_match_index(job)"), 2,
       "the identity check and the fold build their match index the same way, once — "
       "including the leading-code tier, or the guard refuses extracts the fold could match")

    # And the call site actually refuses, rather than logging and carrying on.
    _fs = open(__import__("file_scan").__file__, encoding="utf-8").read()
    ok('if not _sw_id.get("belongs"):' in _fs, "the scan checks before applying")
    ok("_sw_job = None" in _fs,
       "and drops the extract, so neither its geometry NOR its blockers reach this job")
    ok('"refused_wrong_job": True' in _fs, "recording why, on the job")


def test_the_same_pack_produces_the_same_route():
    """M&S 2085 was run twice with identical inputs and identical code. The first returned a
    route including welding and dress_welds; the second did not. Labour fell from GBP 11.14
    to GBP 5.19 and the unit cost from GBP 12.13 to GBP 5.73 — a 53% drop representing no
    manufacturing saving whatsoever, only the model reconsidering an inference it had already
    made. temperature=0 does not make a model deterministic.

    This is the route-side twin of the price problem this engine already refuses to tolerate:
    a figure that will not repeat cannot be quoted. Prices got a reproducibility gate; routes
    got nothing, because until the route reached the sheet nobody could see it move.

    The extract is keyed on the CONTENT it came from — document text, model, prompt — so a
    redrawn pack, a model change or a prompt edit all miss and re-ask, and a re-run of the
    same job on the same code returns the same answer by construction.
    """
    import tempfile
    import llm_full_extract as lfe

    _saved_dir = os.environ.get("SDI_LLM_EXTRACT_CACHE_DIR")
    _saved_ref = os.environ.get("SDI_LLM_EXTRACT_REFRESH")
    os.environ["SDI_LLM_EXTRACT_CACHE_DIR"] = tempfile.mkdtemp()
    os.environ.pop("SDI_LLM_EXTRACT_REFRESH", None)
    try:
        k = lfe._cache_key("full", "PACK TEXT", "grok-4.3", lfe._PROMPT, lfe.SYSTEM_TRANSCRIBE)
        eq(lfe._cache_read(k), None, "nothing cached yet")

        _route = {"found": True, "routes": [{"operation": "welding", "inferred": True},
                                            {"operation": "tube_cut", "inferred": True}]}
        lfe._cache_write(k, _route)
        _back = lfe._cache_read(k)
        eq([r["operation"] for r in _back["routes"]], ["welding", "tube_cut"],
           "the SAME route comes back — the weld does not get reconsidered")
        ok(_back.get("_from_cache"), "and it is marked as a reuse, not a fresh read")

        # A CHANGED PACK MUST MISS. Otherwise a redrawn drawing silently keeps the old route,
        # which is a worse failure than the instability being fixed.
        eq(lfe._cache_read(lfe._cache_key("full", "PACK TEXT REVISED", "grok-4.3",
                                          lfe._PROMPT, lfe.SYSTEM_TRANSCRIBE)), None,
           "a revised pack is a different key")
        eq(lfe._cache_read(lfe._cache_key("full", "PACK TEXT", "grok-9",
                                          lfe._PROMPT, lfe.SYSTEM_TRANSCRIBE)), None,
           "and so is a different model")
        eq(lfe._cache_read(lfe._cache_key("full", "PACK TEXT", "grok-4.3",
                                          lfe._PROMPT + " extra rule",
                                          lfe.SYSTEM_TRANSCRIBE)), None,
           "and so is an edited prompt — a rule change must be re-asked, not remembered")

        os.environ["SDI_LLM_EXTRACT_REFRESH"] = "1"
        eq(lfe._cache_read(k), None, "and a deliberate refresh always re-asks")
    finally:
        os.environ.pop("SDI_LLM_EXTRACT_REFRESH", None)
        if _saved_ref is not None:
            os.environ["SDI_LLM_EXTRACT_REFRESH"] = _saved_ref
        if _saved_dir is None:
            os.environ.pop("SDI_LLM_EXTRACT_CACHE_DIR", None)
        else:
            os.environ["SDI_LLM_EXTRACT_CACHE_DIR"] = _saved_dir

    # BOTH PASSES. The inferred route is the unstable half — welding on 2085 was a
    # conclusion, not a reading, and it is the one that vanished.
    # THE EXACT KEY THE PRODUCTION PATH BUILDS, not one this fixture composes. Computing the
    # key here and asserting on it tests the fixture's arithmetic: dropping `ctx` or `_PROMPT`
    # from extract_full_job's key left this green, because the fixture went on hashing them.
    # That is the mirror mistake, for the third time today.
    _src = open(lfe.__file__, encoding="utf-8").read()
    ok('_key = _cache_key("full", ctx, model, _PROMPT, SYSTEM_TRANSCRIBE)' in _src,
       "the transcription key covers the pack text, the model AND the prompt — drop the text "
       "and a redrawn drawing silently keeps the old route; drop the prompt and a rule change "
       "is never re-asked")
    ok("_cache_write(_key, obj)" in _src, "and the result is stored")
    ok('_ikey = _cache_key("infer", payload, model, SYSTEM_INFER)' in _src,
       "the inference key covers its whole payload — which is where the weld came from")
    ok("_cache_write(_ikey, _res)" in _src, "and it is stored too")


def test_the_raw_model_response_is_kept():
    """"The model did not say it" and "we did not read it" produce identical artefacts.

    2085's weld appeared, vanished, and came back across three runs of the same drawing.
    Settling which of the two it was took reading a commit diff and noticing that no line of
    the prompt, parser, route-fold or department map had changed between the runs — so the
    same code on the same input could not have produced different routes, therefore the input
    differed. That reasoning is sound and it is not a method: it works once, by luck of the
    commit boundaries falling in a useful place.

    _llm_extract.json holds the PARSED job, so it can never answer the question. The raw
    response can, directly, and it costs nothing to keep.
    """
    import llm_full_extract as lfe

    _src = open(lfe.__file__, encoding="utf-8").read()
    ok('obj["_raw_response"] = (raw or "")[:20000]' in _src,
       "the transcription pass keeps what the model actually returned")
    ok('"_raw_response": (raw or "")[:20000],' in _src,
       "and so does the inference pass, which is the unstable half")
    ok('obj["_prompt_fingerprint"]' in _src,
       "with a fingerprint of the prompt it was answering, so a response cannot be compared "
       "against the wrong question")

    _fs = open(__import__("file_scan").__file__, encoding="utf-8").read()
    ok('"raw_response": _job.get("_raw_response")' in _fs,
       "and it reaches the sidecar an estimator can actually open")
    # BOTH RAWS. The first version surfaced only the transcription response — and welding is
    # a CONCLUSION, so it comes from the inference pass. The deliverable answered a question
    # nobody was asking and left the one that cost a commit-diff argument unanswerable.
    ok('_job["_inferred_raw_response"] = _inf.get("_raw_response")' in _fs,
       "the inference raw is carried onto the job")
    ok('"inferred_raw_response": _job.get("_inferred_raw_response")' in _fs,
       "and surfaced — it is where the weld actually is")
    ok('"inference_from_cache"' in _fs,
       "with its own cache flag, because the two passes cache independently")
    ok('"from_cache": bool(_job.get("_from_cache"))' in _fs,
       "which also records whether this run re-asked or reused — otherwise a stable number "
       "and a cached one look the same")


def test_a_model_file_named_code_plus_description_still_matches():
    """SolidWorks files are routinely named "<code> - <description>", so the document title
    arrives as "2085-02 - Outer Tube" and the BOM's "2085-02" matches neither exactly nor by
    trailing segment.

    On M&S 2085 that threw away a GENUINE extract. The model carries 34mm and 20mm bounding
    boxes for the two tubes — the cut lengths the drawing never printed, which is the whole
    reason 7e7cdbd exists — and both were discarded because a description was appended to
    the number. Worse, zero matches is exactly what the wrong-job guard treats as foreign, so
    the job's own extract was refused as somebody else's.

    A separator must follow the code, and ambiguity is refused exactly as the tail rule
    refuses it. Putting one job's geometry on another's part is what this module guards.
    """
    from source_connectors.solidworks import (_native_match_index, _match_native,
                                              extract_is_for_this_job, NativeJob, NativePart)

    job = NativeJob(found=True, meta={"top_assembly": "2085 - Assembly"}, part_signals={
        "2085-01 - Bracket Plate": NativePart(part_number="2085-01 - Bracket Plate"),
        "2085-02 - Outer Tube": NativePart(part_number="2085-02 - Outer Tube",
                                           bbox_mm=[34.0, 12.7, 12.7]),
        "2085-03 - Inner Tube": NativePart(part_number="2085-03 - Inner Tube",
                                           bbox_mm=[20.0, 10.0, 10.0])})
    e, t, l = _native_match_index(job)

    eq(_match_native({"part_number": "2085-02"}, e, t, l), "2085-02 - Outer Tube",
       "the BOM's code finds the model file that leads with it")
    eq(_match_native({"part_number": "2085-03"}, e, t, l), "2085-03 - Inner Tube",
       "and so does the other tube")

    # A PARTIAL CODE MUST NOT CLAIM IT. A separator has to follow, or "2085-0" swallows
    # everything under 2085 and one part's geometry lands on another.
    eq(_match_native({"part_number": "2085-0"}, e, t, l), None,
       "a truncated code matches nothing")

    # A CODE WITH NO DESCRIPTION MUST YIELD NO ALIAS AT ALL.
    #
    # The first version split the whitespace-stripped KEY at every separator, so
    # "12120-01-103" aliased to both "12120" and "12120-01" — and "12120-01" is a real
    # sub-assembly on that very job. It would have taken a leaf part's bounding box,
    # material and bend count. The fix for 2085 reintroduced, by the front door, exactly the
    # cross-contamination this module exists to prevent.
    _plain = NativeJob(found=True, part_signals={
        "12120-01-103": NativePart(part_number="12120-01-103")})
    _e5, _t5, _l5 = _native_match_index(_plain)
    eq(_l5, {}, "a hyphenated code with no description produces no leading alias")
    eq(_match_native({"part_number": "12120-01"}, _e5, _t5, _l5), None,
       "so a sub-assembly cannot take one of its own children's geometry")
    eq(_match_native({"part_number": "12120"}, _e5, _t5, _l5), None,
       "nor can the job number")
    eq(_match_native({"part_number": "12120-01-103"}, _e5, _t5, _l5), "12120-01-103",
       "while the part itself still matches exactly")

    # THE BOUNDARY IS " - ", AND NOTHING LOOSER.
    #
    # "first whitespace, if letters follow" was tried and had to be withdrawn. It aliases
    # "1450 GA" to "1450" — and GA is part of a legitimate SDI code, not a description — and
    # "M4 Male Grip Knob" to "M4", where M4 is a thread spec a BOM line could plausibly
    # carry. Either would put one part's geometry onto another, silently, at the highest
    # rank in the waterfall.
    #
    # Space-dash-space is what the files actually use, and it cannot occur inside a part
    # number. A title without it yields no alias, which is the safe direction: an unmatched
    # part costs a measurement, a mismatched one costs a wrong price nobody can see.
    for _title, _query in (("1450 GA", "1450"),
                           ("M4 Male Grip Knob", "M4"),
                           ("2085-02 Outer Tube", "2085-02"),
                           ("12120 01 103", "12120")):
        _j = NativeJob(found=True, part_signals={_title: NativePart(part_number=_title)})
        _ex, _tx, _lx = _native_match_index(_j)
        eq(_lx, {}, f"'{_title}' has no ' - ' boundary, so it aliases nothing")
        eq(_match_native({"part_number": _query}, _ex, _tx, _lx), None,
           f"and '{_query}' does not claim it")

    # ISOLATE THE SEPARATOR RULE. With three parts a truncated code is ambiguous and gets
    # dropped anyway, so removing the separator requirement still gives the right answer and
    # a mutation passes. It takes a SINGLE-part extract to show what the rule actually does:
    # there, "2085-0" would uniquely claim "2085-02 - Outer Tube" and put that tube's
    # geometry on whatever part happened to be numbered 2085-0.
    _one = NativeJob(found=True, part_signals={
        "2085-02 - Outer Tube": NativePart(part_number="2085-02 - Outer Tube",
                                           bbox_mm=[34.0, 12.7, 12.7])})
    _e1, _t1, _l1 = _native_match_index(_one)
    eq(_match_native({"part_number": "2085-02"}, _e1, _t1, _l1), "2085-02 - Outer Tube",
       "the full code still matches on a single-part extract")
    eq(_match_native({"part_number": "2085-0"}, _e1, _t1, _l1), None,
       "but a truncated one does not, even with nothing to be ambiguous against — the alias "
       "is the whole code before the description, not any prefix of it")
    # AN ASSEMBLY NUMBER MUST NOT CLAIM A PART. "2085" leads "2085-02 - Outer Tube" and is
    # followed by a separator, so it matches on the rule alone — putting the outer tube's
    # bounding box and material onto the assembly that CONTAINS it. Usually the ambiguity
    # rule catches it, because a job number leads several parts; on a single-part extract
    # nothing does. The assembly documents are known, so they are excluded explicitly.
    _asm = NativeJob(found=True, assembly_pns=["2085"], part_signals={
        "2085-02 - Outer Tube": NativePart(part_number="2085-02 - Outer Tube",
                                           bbox_mm=[34.0, 12.7, 12.7])})
    _e4, _t4, _l4 = _native_match_index(_asm)
    eq(_match_native({"part_number": "2085"}, _e4, _t4, _l4), "2085",
       "the assembly matches ITSELF exactly")
    eq(_match_native({"part_number": "2085-02"}, _e4, _t4, _l4), "2085-02 - Outer Tube",
       "and the part still finds its own model")
    ok("2085" not in _l4,
       "but the assembly number is not a leading key for the parts inside it — it is indexed "
       "EXACTLY as a document, and the lead tier excludes anything already in exact")
    eq(_match_native({"part_number": "2085-99"}, e, t, l), None, "and an absent one likewise")

    # AMBIGUITY IS REFUSED, not resolved by dict order — the same discipline as the tail rule.
    _amb = NativeJob(found=True, part_signals={
        "2085-02 - Outer Tube": NativePart(part_number="2085-02 - Outer Tube"),
        "2085-02 - Outer Tube Rev B": NativePart(part_number="2085-02 - Outer Tube Rev B")})
    _e2, _t2, _l2 = _native_match_index(_amb)
    eq(_match_native({"part_number": "2085-02"}, _e2, _t2, _l2), None,
       "a code claimed by two documents is dropped, not guessed between")

    # An exact title still wins outright, and the earlier trailing-segment rule survives.
    _mix = NativeJob(found=True, part_signals={
        "12120-01-01M": NativePart(part_number="12120-01-01M"),
        "2085-02": NativePart(part_number="2085-02")})
    _e3, _t3, _l3 = _native_match_index(_mix)
    eq(_match_native({"part_number": "2085-02"}, _e3, _t3, _l3), "2085-02", "exact wins")
    eq(_match_native({"part_number": "01M"}, _e3, _t3, _l3), "12120-01-01M",
       "and the trailing-segment match still works")

    # AND THE GUARD NOW ACCEPTS THE JOB'S OWN EXTRACT. It was refusing this one, correctly by
    # its own rule and wrongly in fact, because nothing could match.
    ok(extract_is_for_this_job([{"part_number": "2085-02"}, {"part_number": "2085-03"}],
                               job)["belongs"],
       "the genuine extract is recognised as this job's once its parts can be matched")


def test_a_chained_operation_inherits_the_scope_of_what_spawned_it():
    """estimator.py adds dress_welds automatically wherever it finds welding, so that row
    has no route line of its own and therefore no scope of its own.

    It is the same event: if the weld happens once per assembly, so does dressing it.
    Without the inheritance the weld would be charged once and its dressing once per part —
    worse than either answer applied consistently, and the kind of split that makes a sheet
    impossible to reason about.
    """
    from wb_populate import operation_scope_for

    _job_scope = {"welding": "assembly"}
    eq(operation_scope_for({}, "welding", _job_scope), "assembly", "the weld's own scope")
    eq(operation_scope_for({}, "dress_welds", _job_scope), "assembly",
       "and its dressing inherits it, having no route line to carry one")

    # An operation with its OWN scope keeps it — inheritance only fills a gap.
    eq(operation_scope_for({"operation_scope": {"dress_welds": "part"}},
                           "dress_welds", _job_scope), "part",
       "an explicit scope on the part beats the inherited one")

    # Nothing said, nothing invented.
    eq(operation_scope_for({}, "dress_welds", {}), None, "no weld scope, no inherited scope")
    eq(operation_scope_for({}, "laser_cutting", _job_scope), None,
       "and an unrelated operation inherits nothing from the weld")
    eq(operation_scope_for({}, "", _job_scope), None, "an empty operation has no scope")

    # And the loop uses the helper rather than a copy of it.
    _wb = open(__import__("wb_populate").__file__, encoding="utf-8").read()
    ok("_sc = operation_scope_for(pe, _op_l, _scope_by_op)" in _wb,
       "the group loop resolves scope through the shared helper")


def test_qty_per_unit_is_read_by_the_workbook_not_merely_stored():
    """The schema asks the LLM for qty_per_unit, the route fold carries it onto the part,
    and until now the workbook threw it away — an assembly-scoped operation was charged
    exactly once whatever the route said.

    Once per product is the normal case and stays the default. But a product with two welded
    frames in it is two weldings, and a field that is asked for, transported and then ignored
    reads as consumed when it is not.
    """
    from wb_populate import assembly_scoped_qty

    eq(assembly_scoped_qty({}), 1, "nothing said means once per product")
    eq(assembly_scoped_qty({"qty_per_unit_by_scope": 1}), 1, "and one means once")
    eq(assembly_scoped_qty({"qty_per_unit_by_scope": 2}), 2,
       "two welded frames per product is two weldings, not one")

    # Never below one: the operation is on the route, so it happens.
    eq(assembly_scoped_qty({"qty_per_unit_by_scope": 0}), 1, "zero is not a saving")
    eq(assembly_scoped_qty({"qty_per_unit_by_scope": -3}), 1, "nor is a negative")
    eq(assembly_scoped_qty({"qty_per_unit_by_scope": "rubbish"}), 1,
       "an unparseable qty falls back to once rather than raising mid-sheet")

    # And the sheet resolves it through the helper rather than a copy of it.
    _wb = open(__import__("wb_populate").__file__, encoding="utf-8").read()
    ok('g["qty"] = assembly_scoped_qty(g)' in _wb,
       "the labour loop takes its assembly qty from the shared helper")

    # The fold must actually put the number where the helper looks for it, or the helper
    # is correct about a field nothing populates.
    ok('g["qty_per_unit_by_scope"]' in _wb,
       "and the group loop records qty_per_unit from the part's route")


def test_a_coated_tube_contributes_coated_area():
    """The coated-area sum had two contributors, sheet and wire, and section stock matched
    neither — the sheet loop excludes it on stock_form AND again on a "TUBE" description
    guard. So every powder-coated tube, box section and angle on every job contributed zero
    area, an under-charge that grows with how much section the job carries.

    Powder lands on the outside, so the coated surface is outer perimeter x cut length. The
    wall thickness does not enter it.
    """
    from wb_populate import section_coated_area_m2

    # 2085's outer tube: 12.7 CHS x 1.2 wall, 34 mm long, two off per product.
    _chs = {"part_number": "2085-02", "quantity": 2,
            "section_stock": {"a": 12.7, "b": 12.7, "t": 1.2, "profile_form": "CHS"},
            "material_estimate": {"stock_form": "tube",
                                  "stock_estimate": {"section_length_mm": 34.0}}}
    _got = section_coated_area_m2(_chs)
    _want = (3.14159265 * 12.7 / 1000.0) * (34.0 / 1000.0) * 2
    ok(abs(_got - _want) < 1e-9, f"pi x D x L x qty, got {_got}")
    ok(_got > 0, "a coated tube is not a zero-area part")

    # A square section is not a round one: 2(a+b), not pi x D.
    _shs = {"quantity": 1, "section_stock": {"a": 30.0, "b": 30.0, "t": 2.0, "profile_form": "SHS"},
            "material_estimate": {"stock_estimate": {"section_length_mm": 1000.0}}}
    eq(round(section_coated_area_m2(_shs), 6), round(2.0 * 0.060 * 1.0, 6),
       "a 30x30 box is 120 mm of perimeter per metre, not 94.2")

    # The wall does not change the OUTSIDE, so it does not change the coated area.
    _thick = {"quantity": 1, "section_stock": {"a": 30.0, "b": 30.0, "t": 5.0, "profile_form": "SHS"},
              "material_estimate": {"stock_estimate": {"section_length_mm": 1000.0}}}
    eq(round(section_coated_area_m2(_thick), 6), round(section_coated_area_m2(_shs), 6),
       "powder lands on the outside — wall thickness is irrelevant to coverage")

    # Acrylic tube is diamond polished, never coated.
    _acr = dict(_chs, normalized_material="ACRYLIC")
    eq(section_coated_area_m2(_acr), 0.0, "acrylic is polished, not powder coated")

    # No profile or no length means unmeasured, not zero-sized: contribute nothing rather
    # than invent a diameter.
    eq(section_coated_area_m2({"section_stock": {"a": 12.7}}), 0.0, "no length, no claim")
    eq(section_coated_area_m2({"section_stock": {},
                               "material_estimate": {"stock_estimate": {"section_length_mm": 34.0}}}),
       0.0, "no profile, no claim")
    eq(section_coated_area_m2({}), 0.0, "a part with no section stock is not a section")

    # And the powder sum actually adds it, or the helper is right about a total nobody uses.
    _wb = open(__import__("wb_populate").__file__, encoding="utf-8").read()
    ok("_sheet_powder_area_m2 + _wire_powder_area_m2 + _section_powder_area_m2" in _wb,
       "the coated total sums section area alongside sheet and wire")


def test_every_module_resolves_the_names_it_uses():
    """A name a module never binds is a NameError waiting for the one run that reaches it.

    12120 produced no workbook because populate_workbook called getattr(config, ...) in a
    module that imports settings as `from config import X as _X` and never binds the bare
    name `config`. Every fixture passed, because not one of them CALLS populate_workbook —
    the whole suite tested helpers around a function that could no longer run.

    Importing a module does not catch this: the bad line is inside a function body, so it
    only raises when that branch executes on a real job. Nor can fixtures realistically call
    every function on every path. A static scope check can, and does it for every module at
    once, including the ones written after this was.

    Scope is the import closure from main.py — what a real run actually loads — rather than
    every .py in src/, which would drag in one-shot patch scripts and paste-in fragments that
    were never modules and whose noise is where a real finding would hide.
    """
    import ast, builtins as _bi, os as _os

    _SRC = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), "src")
    _BUILTIN = set(dir(_bi)) | {"__file__", "__name__", "__doc__", "__spec__",
                                "__package__", "__builtins__", "__loader__"}
    _SCOPES = (ast.FunctionDef, ast.AsyncFunctionDef, ast.Lambda, ast.ClassDef)
    _COMPS = (ast.ListComp, ast.SetComp, ast.DictComp, ast.GeneratorExp)

    def _parse(path):
        # utf-8-sig: one module carries a BOM, and a BOM is not a syntax error.
        return ast.parse(open(path, encoding="utf-8-sig").read())

    def _bindings(node):
        """Names bound directly in this scope — not those bound inside nested scopes."""
        out, args = set(), getattr(node, "args", None)
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.Lambda)) and args:
            for a in (list(getattr(args, "posonlyargs", [])) + list(args.args)
                      + list(args.kwonlyargs)):
                out.add(a.arg)
            if args.vararg:
                out.add(args.vararg.arg)
            if args.kwarg:
                out.add(args.kwarg.arg)
        body = node.body if isinstance(node.body, list) else [node.body]
        stack = list(body)
        while stack:
            s = stack.pop()
            if isinstance(s, _SCOPES):
                out.add(getattr(s, "name", ""))
                continue                      # a nested scope's body is not this one's
            if isinstance(s, _COMPS):
                continue                      # nor is a comprehension's
            if isinstance(s, (ast.Import, ast.ImportFrom)):
                for a in s.names:
                    out.add((a.asname or a.name).split(".")[0])
            elif isinstance(s, ast.Name) and isinstance(s.ctx, (ast.Store, ast.Del)):
                out.add(s.id)
            elif isinstance(s, ast.ExceptHandler) and s.name:
                out.add(s.name)
            elif isinstance(s, (ast.Global, ast.Nonlocal)):
                out.update(s.names)
            stack.extend(ast.iter_child_nodes(s))
        return out

    def _unresolved(path):
        tree = _parse(path)
        # A module that writes its own globals (dxf_reader re-exports an implementation with
        # globals().update) binds names no static reader can see. Those modules are exempt --
        # the alternative is naming files in an allowlist, which is where the next real
        # finding would hide.
        _src_text = open(path, encoding="utf-8-sig").read()
        if "globals().update" in _src_text:
            return []
        found = []

        def walk(node, scopes, where):
            if isinstance(node, _SCOPES):
                _decs = getattr(node, "decorator_list", []) or []
                for d in _decs:
                    walk(d, scopes, where)                  # decorators read the OUTER scope
                inner = scopes + [_bindings(node)]
                for child in ast.iter_child_nodes(node):
                    if child in _decs:
                        continue
                    walk(child, inner, getattr(node, "name", where))
                return
            if isinstance(node, _COMPS):
                _t = set()
                for gen in node.generators:
                    for n2 in ast.walk(gen.target):
                        if isinstance(n2, ast.Name):
                            _t.add(n2.id)
                for child in ast.iter_child_nodes(node):
                    walk(child, scopes + [_t], where)
                return
            if isinstance(node, ast.Name) and isinstance(node.ctx, ast.Load):
                if node.id not in _BUILTIN and not any(node.id in s for s in scopes):
                    found.append((node.lineno, node.id, where))
                return
            for child in ast.iter_child_nodes(node):
                walk(child, scopes, where)

        walk(tree, [_bindings(tree)], "<module>")
        seen, out = set(), []
        for ln, name, where in sorted(found):
            if (name, where) in seen:
                continue
            seen.add((name, where))
            out.append((ln, name, where))
        return out

    # What a real run loads, walked transitively from the entry point.
    _local = {f[:-3] for f in _os.listdir(_SRC) if f.endswith(".py")}

    def _imports_of(mod):
        _p = _os.path.join(_SRC, mod + ".py")
        if not _os.path.exists(_p):
            return set()
        try:
            _t = _parse(_p)
        except SyntaxError:
            return set()
        _out = set()
        for n in ast.walk(_t):
            if isinstance(n, ast.Import):
                for a in n.names:
                    _out.add(a.name.split(".")[0])
            elif isinstance(n, ast.ImportFrom) and n.module and n.level == 0:
                _out.add(n.module.split(".")[0])
        return {m for m in _out if m in _local}

    _seen, _stack = set(), ["main"]
    while _stack:
        _m = _stack.pop()
        if _m in _seen or _m not in _local:
            continue
        _seen.add(_m)
        _stack.extend(_imports_of(_m))

    ok("main" in _seen, "the entry point itself must be in the checked set")
    ok("wb_populate" in _seen and "estimator" in _seen,
       "the workbook and the estimator are on the path a real run loads")
    ok(len(_seen) > 30, f"the closure looks too small to be the real one ({len(_seen)} modules)")

    _bad = []
    for _m in sorted(_seen):
        _p = _os.path.join(_SRC, _m + ".py")
        try:
            for _ln, _name, _where in _unresolved(_p):
                _bad.append(f"{_m}.py:{_ln}: '{_name}' in {_where}()")
        except SyntaxError as _e:
            _bad.append(f"{_m}.py: will not parse — {_e}")
    ok(not _bad, "names used but never bound (each is a NameError on the run that "
                 "reaches it):\n      " + "\n      ".join(_bad))


def test_a_weldment_parent_is_a_parent_on_both_axes():
    """There were two parent tests and they disagreed.

    estimate_material concludes "weldment parent" from a description token, a -WA/-SA suffix
    or the upstream flag, and suppresses the parent's sheet material. The phantom-fab strip
    in estimate_part reads ONLY the upstream flag, which is stamped by a rule requiring the
    part number to be a strict prefix of >=2 others.

    A sub-assembly whose constituents are SIBLINGS has no prefix children, so that rule
    cannot fire. 12120-01-101 STAND WELD ASSY is built from 12120-01-02M and -03M; it was
    recognised as a parent by its description, correctly carried no material, and still
    billed a laser and a fold read off the assembly drawing's linework.

    Same conclusion, suppressed on one axis and not the other. This drives estimate_part --
    the caller -- because that is where the two axes have to agree.
    """
    from estimator import estimate_part

    def _part(desc):
        return {"part_number": "12120-01-101", "description": desc, "quantity": 1,
                "normalized_material": "MILD STEEL", "normalized_thickness_mm": 1.5,
                "material": "MILD STEEL", "thickness_mm": 1.5,
                "overall_length_mm": 300, "overall_width_mm": 120,
                "normalized_geometry": {"blank_length_mm": 300, "blank_width_mm": 120,
                                        "bend_count": 2, "cut_length_mm": 900, "hole_count": 4},
                "bend_count": 2, "cut_length_mm": 900,
                "manufacturing_interpretation": {
                    "stock_form": "sheet",
                    "operations": ["laser_cutting", "folding", "welding"]},
                "textual_operations": ["laser_cutting", "folding", "welding"]}

    def _ops(desc):
        p = _part(desc)
        out = estimate_part(p, job_quantity=180)
        _rt = (out.get("process_estimate") or {}).get("run_times_min_per_unit") or {}
        return p, set(_rt)

    _parent, _p_ops = _ops("STAND WELD ASSY")
    _leaf, _l_ops = _ops("STAND BRACKET")

    ok(_parent.get("is_assembly_parent"),
       "a description the material path calls a weldment parent must reach the fab strip")
    ok(not _leaf.get("is_assembly_parent"),
       "and an ordinary bracket is still a leaf — the rule must not swallow everything")

    # The leaf is the control: without it, a strip that removed EVERYTHING would pass.
    ok({"laser_cutting", "folding"} <= _l_ops,
       f"the leaf must keep the ops the parent is being stripped of, got {sorted(_l_ops)}")
    ok(not ({"laser_cutting", "folding"} & _p_ops),
       f"a parent has no flat blank, so it cuts and folds nothing, got {sorted(_p_ops)}")

    # WELDING THE PARENT IS THE WORK. Stripping assembly ops would be the opposite error,
    # and would silently delete the only operation the parent legitimately owns.
    ok("welding" in _p_ops, f"the weld is why the assembly exists, got {sorted(_p_ops)}")
    ok("dress_welds" in _p_ops, "and dressing that weld travels with it")
    ok("handling" in _p_ops, "the parent is still handled")

    # The material axis, which already worked, must not have regressed.
    _m = estimate_part(_part("STAND WELD ASSY"), job_quantity=180).get("material_estimate") or {}
    eq(_m.get("cost_per_part_gbp"), 0.0, "the parent's material is carried by its children")
    eq(_m.get("blank_length_mm"), None, "and it has no blank of its own")


def test_the_fallback_sheet_admits_it_is_a_fallback():
    """xlsx_output runs only when populate_workbook could not produce the SDI template, and
    the sheet it writes has always been indistinguishable from the real one — same title,
    same header boxes, same totals block.

    It is not the same. It has no route grouping, so it emits one row per part per operation
    where the template collapses Weld / Dress / P.Coat / Assemble-pack into ONE row per job.
    Read as an estimate it shows every assembly operation charged at every level of the
    tree, which reads as a cost-model defect and is not one.

    Two full review cycles on 12120 were spent diagnosing assembly-scope failures from this
    sheet before anyone checked which writer produced it. The sheet has to say so itself.
    """
    import os as _os
    import wb_populate

    # Read the source rather than import it: xlsx_output needs openpyxl at import time, and
    # a machine without it must still be able to check that the warning is present.
    _SRCDIR = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), "src")
    _src = open(_os.path.join(_SRCDIR, "xlsx_output.py"), encoding="utf-8").read()
    _title = [l for l in _src.splitlines() if "SDI Displays Limited" in l]
    ok(_title, "the fallback still writes a title cell")
    _banner = "\n".join(_src.split("SDI Displays Limited")[1].splitlines()[:6]).upper()
    ok("FALLBACK" in _banner,
       "the fallback sheet's own title must carry the word FALLBACK, on the sheet, where "
       "someone reading the estimate will see it — not only in the console")
    ok("OVERSTATED" in _banner or "OVERSTATE" in _banner,
       "and must say which way the number is wrong, since labour is the part that inflates")

    # The claim the banner makes has to be true of the template writer, or the warning is
    # telling estimators something the code does not do.
    _wb = open(_os.path.join(_SRCDIR, "wb_populate.py"), encoding="utf-8").read()
    _one_row = _wb.split("_ONE_ROW_PER_JOB")[1].split("}")[0]
    for _op in ("Assemble/pack (Metal)", "Weld (CO2)", "Dress Welds", "P.Coat"):
        ok(_op in _one_row,
           f"the banner says the template groups {_op} once per job — it must actually do it")

    # And an unreachable share must not be a dead end.
    _was = _os.environ.get("SDI_WB_TEMPLATE")
    try:
        _os.environ["SDI_WB_TEMPLATE"] = r"C:\local\Blank Estimate Sheet.xlsx"
        eq(wb_populate.template_path(), r"C:\local\Blank Estimate Sheet.xlsx",
           "SDI_WB_TEMPLATE points the populater at a local copy")
        _os.environ["SDI_WB_TEMPLATE"] = '  "C:\\quoted\\path.xlsx"  '
        eq(wb_populate.template_path(), r"C:\quoted\path.xlsx",
           "a pasted Windows path keeps its quotes and spaces — strip them, not the path")
        _os.environ["SDI_WB_TEMPLATE"] = ""
        eq(wb_populate.template_path(), wb_populate.CELL_MAP["template_path"],
           "and an empty override falls back to the share, not to an empty path")
    finally:
        if _was is None:
            _os.environ.pop("SDI_WB_TEMPLATE", None)
        else:
            _os.environ["SDI_WB_TEMPLATE"] = _was

    # The CALL SITE, not merely the definition -- "template_path()" alone also matches the
    # `def` line, so a helper wired to nothing passed this check the first time it was run.
    ok("tpl = template_path()" in _wb,
       "populate_workbook must resolve the template through the override, not from CELL_MAP "
       "directly — otherwise the override exists and does nothing")


def test_what_the_drawing_shows_is_not_what_we_supply():
    """A drawing routinely depicts things it is not quoting for and says so beside them.

    12120's GA carries "SCREEN & CABLE SHOWN FOR REFERENCE". The prose recogniser saw the
    SAFE electrical head-word 'cable', minted BI-SCREENCABLE, priced it from history and
    booked handling labour against it. Nothing read the half of the sentence saying it is
    not ours.

    These are drawing conventions, not one customer's wording, so the rule inherits.
    """
    from bought_in_recogniser import is_reference_only, recognise_bought_in_in_prose
    import bought_in_recogniser as _bir

    _note = "1. SCREEN & CABLE SHOWN FOR REFERENCE\n2. FIT 4x M4 THUMB SCREW"
    ok(is_reference_only("screen cable", _note), "the cable is shown, not supplied")
    ok(not is_reference_only("thumb screw", _note),
       "and the thumb screw in the very next note still is — a scope marker governs its own "
       "clause, not the page")

    for _marker in ("BY OTHERS", "CUSTOMER SUPPLIED", "NOT SUPPLIED", "FOR REFERENCE ONLY",
                    "SHOWN DOTTED FOR CLARITY", "FREE ISSUE"):
        ok(is_reference_only("monitor arm", f"MONITOR ARM {_marker}"),
           f"'{_marker}' is a scope-of-supply disclaimer")

    # EVERY mention, not any. A part shown for reference on one view and scheduled as
    # supplied elsewhere is supplied — the opposite reading turns a scope note into a silent
    # deletion of a real BOM line, which is the more expensive error and the harder to see.
    ok(not is_reference_only("monitor arm", "MONITOR ARM BY OTHERS. SUPPLY 2x MONITOR ARM"),
       "a part supplied somewhere on the drawing is supplied")
    ok(not is_reference_only("cable tie", "SUPPLY 20x CABLE TIE"), "no marker, no exclusion")
    ok(not is_reference_only("", "anything"), "an empty phrase excludes nothing")
    ok(not is_reference_only("screen cable", ""), "and empty prose excludes nothing")

    # THE CALLER. The helper being right proves nothing if the loop never asks it — which is
    # how a built-not-wired capability has shipped repeatedly in this codebase.
    def _conn():
        raise RuntimeError("fixtures touch no live service")

    def _stub(code, desc, qty):
        return {"part_number": code, "description": desc, "quantity": qty}

    _was = dict(_bir._ELECTRICAL_VOCAB)
    try:
        # Inject the vocabulary rather than the verdict: the real scan, the real guard.
        _bir._ELECTRICAL_VOCAB.update({"screen cable": "cable", "thumb screw": "screw"})
        _got = recognise_bought_in_in_prose(_note, get_connection=_conn, stub_builder=_stub)
    finally:
        _bir._ELECTRICAL_VOCAB.clear()
        _bir._ELECTRICAL_VOCAB.update(_was)

    _codes = sorted(str(i.get("part_number") or "") for i in _got)
    ok("BI-SCREENCABLE" not in _codes,
       f"the reference-only cable must never reach a BOM line, got {_codes}")
    ok("BI-THUMBSCREW" in _codes,
       f"and the genuine bought-in in the same note must survive, got {_codes}")


def test_the_ga_hierarchy_says_what_is_an_assembly():
    """The whole-document extract already returns `assemblies` — parent part numbers with
    their children — and llm_full_job stamps every match is_sub_assembly. pricing_service
    reads that field. Nothing in the estimator did: both suppressions keyed on
    is_assembly_parent, a different name for the same idea.

    So 12120-01-103 SCREEN MOUNTING BRACKET, correctly identified as a sub-assembly from the
    GA tree, still got sheet material, a laser and a fold. Its description carries no
    assembly token, its children (01M, 06M) are siblings rather than prefixed, and it does
    not match the ...-0*-... GA pattern — every parent detector missed it while the answer
    sat on the record under the other name.
    """
    from estimator import estimate_part

    def _run(**kw):
        p = {"part_number": "12120-01-103", "description": "SCREEN MOUNTING BRACKET",
             "quantity": 1, "normalized_material": "MILD STEEL",
             "normalized_thickness_mm": 1.5, "material": "MILD STEEL", "thickness_mm": 1.5,
             "overall_length_mm": 300, "overall_width_mm": 120,
             "normalized_geometry": {"blank_length_mm": 300, "blank_width_mm": 120,
                                     "bend_count": 2, "cut_length_mm": 900, "hole_count": 4},
             "bend_count": 2, "cut_length_mm": 900,
             "manufacturing_interpretation": {
                 "stock_form": "sheet",
                 "operations": ["laser_cutting", "folding", "welding"]},
             "textual_operations": ["laser_cutting", "folding", "welding"]}
        p.update(kw)
        out = estimate_part(p, job_quantity=180)
        _rt = (out.get("process_estimate") or {}).get("run_times_min_per_unit") or {}
        return p, (out.get("material_estimate") or {}), set(_rt)

    # Control: the identical part with no hierarchy claim is a leaf and stays one.
    _leaf, _lm, _lops = _run()
    ok(not _leaf.get("is_assembly_parent"), "an unclaimed part is a leaf")
    ok((_lm.get("cost_per_part_gbp") or 0) > 0, "and a leaf buys its own material")
    ok({"laser_cutting", "folding"} <= _lops, f"and cuts and folds, got {sorted(_lops)}")

    # The GA tree says it is an assembly, and now both axes hear it.
    _asm, _am, _aops = _run(is_sub_assembly=True)
    ok(_asm.get("is_assembly_parent"),
       "a sub-assembly named by the GA hierarchy must reach the fabrication strip")
    eq(_am.get("cost_per_part_gbp"), 0.0, "its material is carried by its children")
    ok(not ({"laser_cutting", "folding"} & _aops),
       f"and it has no blank to cut or fold, got {sorted(_aops)}")
    ok({"welding", "dress_welds"} <= _aops,
       f"while the operations that build it survive, got {sorted(_aops)}")

    # MEASUREMENT OUTRANKS A TRANSCRIBED HIERARCHY. A real flat pattern is evidence the part
    # is a fabricated leaf (dxf 80 > llm_full_extract 40); where they disagree, geometry wins
    # and nothing is silently zeroed on the strength of a text read.
    _both, _bm, _bops = _run(is_sub_assembly=True, geometry_source="dxf_flat_pattern")
    ok(not _both.get("is_assembly_parent"),
       "a part with its own flat pattern stays a leaf whatever the hierarchy claims")
    ok((_bm.get("cost_per_part_gbp") or 0) > 0, "so it keeps its material")
    ok({"laser_cutting", "folding"} <= _bops, "and its cut and fold")


def test_a_tube_is_cut_somewhere_or_it_is_cut_for_free():
    """A tube has no flat blank, so it is not profile-lasered. It is still SAWN TO LENGTH.

    Dropping laser/guillotine on tube stock as "spurious" was half right and expensively so:
    it removed the profile cut that never happened AND the cut-to-length that always does.
    2085's tubes came out of that change with no cutting operation of any kind — the sheet
    was silent on them, which reads as no work rather than as work nobody costed.

    Remapped, not dropped, exactly as fold already was. The module's own predicate says so:
    an operation that has a correct home for this stock form is not spurious.
    """
    from wb_populate import _map_operation, _is_spurious_operation

    for _cut in ("laser", "laser_cutting", "laser_metal", "guillotine"):
        ok(not _is_spurious_operation(_cut, "tube"),
           f"'{_cut}' on a tube is misplaced, not impossible — dropping it makes the cut free")
        eq(_map_operation(_cut, False, "tube"), "Tube",
           f"'{_cut}' on a tube is a tube-department cut")

    # Explicit tube-cutting words land in the same place, whatever the route called it.
    for _cut in ("tube_cut", "tube_cutting"):
        eq(_map_operation(_cut, False, "tube"), "Tube", f"'{_cut}' is a tube cut")
    eq(_map_operation("saw", False, "tube"), "Saw", "a sawn tube books to the saw")

    # PUNCH IS NOT A CUT TO LENGTH. A hole through a tube wall is hole work, which has its
    # own operation; remapping it to a cut would invent a second cut on every punched tube.
    ok(_is_spurious_operation("punch", "tube"), "punching a tube is not cutting it to length")

    # Forming is unchanged — a tube bends on a tube-bender, not a press brake.
    eq(_map_operation("folding", False, "tube"), "Tubebend", "a tube is bent, not folded")

    # SHEET IS UNTOUCHED. The remap is keyed on stock form, so a plate still lasers.
    eq(_map_operation("laser_cutting", False, "sheet"), "Laser (Metal)",
       "a sheet part still profile-lasers")
    eq(_map_operation("guillotine", False, "sheet"), "Guillotine", "and still guillotines")
    ok(not _is_spurious_operation("punch", "sheet"), "and is still punchable")

    # The cut has to SEQUENCE as a cut, or the sheet lists it after the weld it precedes.
    # _SHOP_ORDER is local to populate_workbook, so this reads the table it is declared in.
    _wb = open(__import__("wb_populate").__file__, encoding="utf-8").read()
    _order = _wb.split("_SHOP_ORDER = {")[1].split("}")[0]
    import re as _re
    _rank = {m.group(1): int(m.group(2))
             for m in _re.finditer(r'"([^"]+)":\s*(\d+)', _order)}
    ok(_rank.get("Tube") is not None, "the Tube department must have a place in the route")
    ok(_rank["Tube"] < _rank["Tubebend"], "you cut the tube before you bend it")
    ok(_rank["Tube"] < _rank["Weld (CO2)"], "and before you weld it in")


def test_a_material_stated_once_applies_to_the_whole_drawing():
    """The schema asks for drawing_info.material_general and finish_general, Grok returns
    them, and until now the field name appeared exactly once in the entire codebase: in the
    schema requesting it. Extracted, transported, never read.

    2085 is the case this exists for. Its GA prints "MATERIAL: MILD STEEL" once at assembly
    level and names no material on either tube row, so the tubes reached the sheet with no
    material at all and could not be priced. The drawing does say what they are made of — it
    says it once, in the place nobody looked.

    Strictly a gap-fill: weaker than a part-level reading, so it never overrides one.
    """
    from source_connectors.llm_full_job import apply_full_job_to_pre_estimate

    _job = {"found": True,
            "drawing_info": {"material_general": "MILD STEEL",
                             "finish_general": "POWDER COATED"},
            "parts": [{"part_number": "2085-02", "description": "OUTER TUBE"},
                      {"part_number": "2085-01", "description": "BRACKET PLATE",
                       "material": "STAINLESS STEEL", "finish": "BRUSHED"}]}
    _parts = [{"part_number": "2085-02"}, {"part_number": "2085-01"}]
    _counts = apply_full_job_to_pre_estimate(_parts, _job)
    _tube = {p["part_number"]: p for p in _parts}["2085-02"]
    _plate = {p["part_number"]: p for p in _parts}["2085-01"]

    ok(_tube.get("normalized_material"),
       "a tube whose row states no material still has one — the GA states it for the job")
    eq(_tube.get("normalized_material"), "MILD_STEEL", "and it is the drawing's own material")
    eq(_tube.get("normalized_finish"), "POWDER COATED", "same for the general finish note")
    ok(_counts.get("material_from_general"), "the inheritance is counted, not silent")
    ok(any("GENERAL material note" in str(f) for f in (_tube.get("review_flags") or [])),
       "and flagged on the part: inherited evidence is not the same as a printed row")

    # A ROW THAT SPEAKS FOR ITSELF IS NOT OVERRULED. The general note is the weaker source.
    eq(_plate.get("normalized_material"), "STAINLESS_STEEL",
       "a part naming its own material keeps it — a job-level note never overrides a row")
    eq(_plate.get("normalized_finish"), "BRUSHED", "nor its own finish")

    # NOTHING STATED, NOTHING INVENTED.
    _bare = [{"part_number": "X-1"}]
    apply_full_job_to_pre_estimate(_bare, {"found": True, "drawing_info": {},
                                           "parts": [{"part_number": "X-1"}]})
    ok(not _bare[0].get("normalized_material"),
       "with no general note the part stays blank rather than acquiring a default")


def test_every_field_we_ask_the_model_for_has_a_reader():
    """A field we request, pay for, and never read is indistinguishable from one we never
    asked for — it costs nothing visible and silently under-costs the job.

    drawing_info.material_general was exactly that. 2085's GA prints "MATERIAL: MILD STEEL"
    once at assembly level and names none on either tube row; the field name appeared once
    in the whole codebase, in the schema requesting it, and both tubes reached the sheet
    with no material and no price.

    Every field in the schema must therefore either have a reader or be listed below with a
    reason. The list is the point: adding a field to the schema forces a decision about who
    consumes it, instead of leaving the answer to be discovered from a wrong price.

    Static reads only — a field reached dynamically (dict iteration, a JSON dump) will look
    unread here. That is the safe direction: it can raise a false alarm, answered by adding
    a line below, but it cannot miss a genuinely orphaned field.
    """
    import ast as _ast, os as _os, re as _re, collections as _c

    # ACKNOWLEDGED UNREAD. Each is here because it is genuinely not needed to cost a job,
    # NOT because nobody has looked. Anything that affects a price does not belong here.
    _ACKNOWLEDGED = {
        "drawn_by":          "who drew it — provenance for the report, never a cost input",
        "item_no":           "the BOM table's own line number; parts join on part_number",
        "tolerance_linear":  "tolerance is not yet a cost driver; would gate inspection time",
        "tolerance_angular": "as above",
        "tolerances":        "spec-level duplicate of the two above",
        "timber_note":       "free text about timber grade; material_family already routes it",
        # NOT benign, and deliberately named rather than quietly tolerated: is_fabricated is
        # the per-row make/buy verdict. It stays unread until ONE consumer owns make/buy and
        # everything else reads that verdict — folding a second opinion in blind is how the
        # silent-wrong-decision class gets reintroduced.
        "is_fabricated":     "PENDING: make/buy needs a single owning consumer before wiring",
    }

    _SRC = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), "src")
    _raw = open(_os.path.join(_SRC, "llm_full_extract.py"),
                encoding="utf-8").read().split('_SCHEMA = """')[1].split('"""')[0]
    _fields = sorted(set(_re.findall(r'"([a-z_][a-z0-9_]*)"\s*:', _raw)))
    ok(len(_fields) > 30, f"the schema should carry real content, found {len(_fields)} fields")

    _reads = _c.defaultdict(set)
    for _dp, _dirs, _fs in _os.walk(_SRC):
        for _f in _fs:
            if not _f.endswith(".py") or _f.startswith("_") or ".patch" in _f:
                continue
            _p = _os.path.join(_dp, _f)
            try:
                _t = _ast.parse(open(_p, encoding="utf-8-sig").read())
            except SyntaxError:
                continue
            for _n in _ast.walk(_t):
                if (isinstance(_n, _ast.Call) and isinstance(_n.func, _ast.Attribute)
                        and _n.func.attr in ("get", "pop") and _n.args
                        and isinstance(_n.args[0], _ast.Constant)
                        and isinstance(_n.args[0].value, str)):
                    _reads[_n.args[0].value].add(_f)
                if (isinstance(_n, _ast.Subscript) and isinstance(_n.ctx, _ast.Load)
                        and isinstance(_n.slice, _ast.Constant)
                        and isinstance(_n.slice.value, str)):
                    _reads[_n.slice.value].add(_f)

    _orphans = [f for f in _fields if f not in _reads and f not in _ACKNOWLEDGED]
    ok(not _orphans,
       "asked the model for these and nothing reads them — each is a silent under-cost "
       "waiting for the job that needs it:\n      " + "\n      ".join(_orphans))

    # The list must not rot into a place things are parked. Anything on it that HAS acquired
    # a reader should come off, or the list stops meaning what it says.
    _stale = [f for f in _ACKNOWLEDGED if f in _reads]
    ok(not _stale,
       f"these are listed as acknowledged-unread but now have readers — remove them from "
       f"the list so it keeps meaning something: {_stale}")

    # And the two that caused this: they must stay read.
    for _f in ("material_general", "finish_general"):
        ok(_f in _reads, f"'{_f}' is what this fixture exists for — it must have a reader")


def test_an_operation_already_present_still_gets_its_metadata():
    """Adding an operation's NAME and merging its METADATA are separate acts, and this fold
    used to treat them as one: `if op in ops: continue`, placed above everything that reads
    the route line. Source, sequence, department, SCOPE and qty_per_unit were all skipped.

    The operation survived. The information needed to cost it correctly did not.

    That is how an assembly-level weld loses its scope. If any earlier reader has already
    put 'welding' on 12120-01-02M, the route line saying scope=assembly, participants
    02M/03M/101, qty_per_unit 1, sequence 40 was discarded in full — and the workbook then
    had nothing to tell it the weld happens once rather than once per part it names.
    """
    from source_connectors.llm_full_job import apply_routes_to_parts

    _job = {"found": True, "routes": [
        {"sequence": 40, "operation": "welding", "department": "WELD",
         "part_numbers": ["12120-01-02M", "12120-01-03M"],
         "scope": "assembly", "qty_per_unit": 1, "confidence": "high"}]}

    # 02M already carries the operation from an earlier reader; 03M does not.
    _parts = [{"part_number": "12120-01-02M", "textual_operations": ["welding"]},
              {"part_number": "12120-01-03M", "textual_operations": []}]
    _added = apply_routes_to_parts(_parts, _job)
    _by = {p["part_number"]: p for p in _parts}

    for _pn in ("12120-01-02M", "12120-01-03M"):
        _p = _by[_pn]
        eq((_p.get("operation_scope") or {}).get("welding"), "assembly",
           f"{_pn}: an assembly-scoped weld must carry its scope whether or not the "
           f"operation name was already on the part")
        eq((_p.get("operation_sequence") or {}).get("welding"), 40.0, f"{_pn}: and its sequence")
        eq((_p.get("operation_qty_per_unit") or {}).get("welding"), 1.0, f"{_pn}: and its qty")
        ok((_p.get("operation_sources") or {}).get("welding"), f"{_pn}: and its source")

    eq(_parts[0]["textual_operations"].count("welding"), 1,
       "merging metadata must not duplicate the operation itself")
    eq(_added, 1, "and the counter reports genuinely NEW operations, not metadata merges")

    # GAP-FILL, NEVER OVERWRITE. A stronger earlier source keeps every field it filled.
    _own = [{"part_number": "12120-01-02M", "textual_operations": ["welding"],
             "operation_scope": {"welding": "part"},
             "operation_sequence": {"welding": 10.0}}]
    apply_routes_to_parts(_own, _job)
    eq(_own[0]["operation_scope"]["welding"], "part",
       "a scope already decided by a stronger source is not overwritten by the model")
    eq(_own[0]["operation_sequence"]["welding"], 10.0, "nor its sequence")

    # A RULED-OUT OPERATION IS STILL REFUSED. The metadata merge must not become a way
    # back in for work a measurement already excluded.
    _ruled = [{"part_number": "12120-01-02M", "textual_operations": [],
               "operations_ruled_out": {"welding": "measured: no weld preparation"}}]
    apply_routes_to_parts(_ruled, _job)
    ok("welding" not in (_ruled[0].get("textual_operations") or []),
       "a measurement that ruled the operation out still stands")
    ok(not (_ruled[0].get("operation_scope") or {}).get("welding"),
       "and it acquires no scope either — it is not happening")


def test_costing_a_part_reaches_no_live_service():
    """test_the_rules_suite_touches_no_live_service has guarded exactly one module.

    Everything that costs went straight past it: estimate_part -> estimate_material ->
    _resolve_material_price -> PricingService -> SQL Server, and from there the xAI price
    lookup. Two fixtures added this session call estimate_part four times each, so on a
    machine that can reach those services the suite made live database and paid LLM calls —
    minutes on the SDI network, and pyodbc.connect has no timeout at all, so it could block
    indefinitely. On a machine where nothing is routable it failed instantly and looked fine.

    A guard that only covers the module nobody was worried about is not a guard.
    """
    import os as _os
    import estimator as _est
    import web_ai_price_lookup as _web

    eq(_os.environ.get("SDI_OFFLINE"), "1", "the suite declares itself offline")

    # OBSERVE THE ATTEMPT, NOT THE RETURN VALUE.
    #
    # The first version of this fixture asserted _get_pricing_service() is None. On a
    # machine where SQL is unreachable a REMOVED guard also returns None, so it proved the
    # guard only where the guard does not matter — the exact environment-dependence this
    # exists to remove, and its mutation passed clean.
    import pricing_service as _ps
    _attempts = []

    class _Tripwire:
        def __init__(self, *a, **k):
            _attempts.append("constructed")

    _real_cls, _saved_s, _saved_f = (_ps.PricingService,
                                     _est._PRICING_SERVICE_SINGLETON,
                                     _est._PRICING_SERVICE_FAILED)
    try:
        _ps.PricingService = _Tripwire
        _est._PRICING_SERVICE_SINGLETON = None
        _est._PRICING_SERVICE_FAILED = False
        _got = _est._get_pricing_service()
        eq(_attempts, [], "offline must not even ATTEMPT to construct a PricingService")
        eq(_got, None, "and returns nothing rather than a connection")
    finally:
        _ps.PricingService = _real_cls
        _est._PRICING_SERVICE_SINGLETON = _saved_s
        _est._PRICING_SERVICE_FAILED = _saved_f

    # The paid LLM call, guarded at the primitive rather than at each of its four callers
    # (pricing_service, note_scan, bay_rollup, probe_pipeline) so a fifth inherits it.
    # A key is set for the duration, or the function returns before reaching the network
    # and the tripwire could never fire whether the guard existed or not.
    import urllib.request as _ur
    _calls = []

    def _no_net(*a, **k):
        _calls.append("dialled")
        raise AssertionError("a fixture reached the network")

    _real_open, _had_key = _ur.urlopen, _os.environ.get("XAI_API_KEY")
    try:
        _ur.urlopen = _no_net
        _os.environ["XAI_API_KEY"] = "test-key-not-real"
        eq(_web._call_xai_llm("price a bracket"), None, "no paid model call from a fixture")
        eq(_calls, [], "and no request is made — the guard returns before the network")
    finally:
        _ur.urlopen = _real_open
        if _had_key is None:
            _os.environ.pop("XAI_API_KEY", None)
        else:
            _os.environ["XAI_API_KEY"] = _had_key
    _r = _web.lookup_web_ai_price({"material": "MILD STEEL", "description": "M4 knob"})
    eq(_r.get("found"), False, "and the lookup reports nothing found")
    ok("SDI_OFFLINE" in str(_r.get("review_reason") or ""),
       "saying WHY — an offline run is not evidence that no price exists, and a blank "
       "price column that means 'we did not look' must not read as 'there is none'")

    # THE THIRD AND WIDEST PATH: the legacy connector factory, reached by
    # _resolve_part_system_cost on every part estimate_part prices. It builds a SQL Server
    # connector, an MS Access connector that opens a file on the estimating share, and a web
    # connector. A mounted-but-slow share blocks with no timeout, which is what stopped this
    # suite dead on the estimating machine while it finished in seconds where the network
    # was simply absent.
    import price_sources as _psrc
    eq(_psrc.build_price_connectors(), {},
       "offline builds no connector at all, so nothing can dial SQL, Access or the web")
    _r2 = _psrc.get_best_price(
        _psrc.PriceRequest(kind="part_system_cost", part_code="X", description="bracket"))
    ok(not (_r2.get("selected") or {}).get("unit_price_gbp"),
       "and no price is returned from a source that was never queried")
    ok(_r2.get("audit_trail"),
       "while the audit trail still says WHY — an offline run that found no price must not "
       "read like a part with no price")

    # THE CALLER. A part must still cost end-to-end offline, or the guard has simply
    # broken costing instead of isolating it.
    _p = {"part_number": "GUARD-01", "description": "BRACKET", "quantity": 1,
          "normalized_material": "MILD STEEL", "normalized_thickness_mm": 1.5,
          "material": "MILD STEEL", "thickness_mm": 1.5,
          "normalized_geometry": {"blank_length_mm": 120, "blank_width_mm": 80,
                                  "bend_count": 1, "cut_length_mm": 400},
          "manufacturing_interpretation": {"stock_form": "sheet",
                                           "operations": ["laser_cutting", "folding"]},
          "textual_operations": ["laser_cutting", "folding"]}
    _out = estimate_part_offline_safe(_p)
    ok(_out is not None, "estimate_part still returns offline")
    ok((_out.get("material_estimate") or {}).get("cost_per_part_gbp") is not None,
       "and still prices from the configured rates rather than needing the database")


def estimate_part_offline_safe(part):
    """estimate_part, imported at call time so the module-level guard is already in force."""
    from estimator import estimate_part
    return estimate_part(part, job_quantity=100)


def test_a_reference_only_row_is_excluded_on_the_extract_path_too():
    """The prose recogniser was guarded first, and 12120 still shipped BI-SCREENCABLE with
    no recogniser message at all — the row arrived through the whole-document extract's own
    `bom` list, a different door into the same mistake.

    One predicate, two doors. The marker vocabulary is shared with is_reference_only rather
    than restated, so a phrase added for one path covers both, and nothing here knows the
    word 'cable' — a rule that did would need extending for the next screen, monitor arm or
    customer bracket somebody draws for reference.
    """
    from llm_full_extract import normalize_job
    from bought_in_recogniser import bom_row_is_reference_only

    # Disclaimed on the drawing's own notes, and disclaimed in the row's own text.
    ok(bom_row_is_reference_only("Screen Cable", "", "SCREEN & CABLE SHOWN FOR REFERENCE"),
       "a row the notes disclaim is not supplied")

    # THE TEXT AS THE PDF ACTUALLY RETURNS IT. 12120's disclaimer wraps:
    #     SCREEN & CABLE SHOWN
    #     FOR REFERENCE
    # The first version of this rule split on every newline, so it saw one clause naming the
    # cable with no marker and another carrying the marker with nothing to attach it to —
    # and the cable was supplied and priced. The fixture wrote it on ONE line and passed,
    # which is why the defect reached a real estimate.
    _wrapped = "SCREEN & CABLE SHOWN\nFOR REFERENCE"
    ok(bom_row_is_reference_only("Screen Cable", "", _wrapped),
       "a disclaimer wrapped across lines still disclaims — a PDF line break is layout, "
       "not a sentence")
    _numbered = "1. SCREEN & CABLE SHOWN\nFOR REFERENCE\n2. FIT 4x M4 THUMB SCREW"
    ok(bom_row_is_reference_only("Screen Cable", "", _numbered),
       "wrapped inside a numbered note, still disclaimed")
    ok(not bom_row_is_reference_only("Thumb Screw", "", _numbered),
       "and note 2's hardware is untouched — a NUMBERED note IS a boundary even though a "
       "bare line break is not")
    ok(bom_row_is_reference_only("Monitor Arm", "by others", ""),
       "nor is one whose own row text says who supplies it")
    ok(not bom_row_is_reference_only("Power Cable", "", "SUPPLY 1x POWER CABLE"),
       "and a cable the drawing asks for IS supplied — the rule is about scope, not cables")

    _job = {"drawing_info": {"notes": ["1. SCREEN & CABLE SHOWN FOR REFERENCE",
                                       "2. FIT 4x M4 THUMB SCREW"]},
            "bom": [
                {"part_number": "BI-SCREENCABLE", "description": "Screen Cable", "qty": 1},
                {"part_number": "THUM620", "description": "M4x10mm Mushroom Thumbscrew",
                 "qty": 4},
                {"part_number": "BI-MONARM", "description": "Monitor Arm",
                 "comments": "by others", "qty": 1},
                {"part_number": "BI-PWRCABLE", "description": "Power Cable", "qty": 1},
            ]}
    normalize_job(_job)
    _kept = [p["part_number"] for p in _job["parts"]]

    ok("BI-SCREENCABLE" not in _kept,
       f"the reference-only cable must not become a supplied part, got {_kept}")
    ok("BI-MONARM" not in _kept, f"nor the monitor arm supplied by others, got {_kept}")

    # A GENUINELY SUPPLIED CABLE STAYS. Without this the fixture would pass just as well
    # against a rule that dropped every cable, which is the failure mode to avoid.
    ok("BI-PWRCABLE" in _kept, f"a supplied cable stays in the BOM, got {_kept}")
    ok("THUM620" in _kept,
       f"and the hardware in the very next note stays — a scope marker governs its own "
       f"clause, not the page, got {_kept}")

    # THE EARLY RETURN WAS A BYPASS. A job arriving already in the parts shape — from a
    # cache, a sidecar, or a caller that projected it itself — skipped the filter entirely.
    _pre = {"drawing_info": {"notes": ["SCREEN & CABLE SHOWN", "FOR REFERENCE"]},
            "parts": [{"part_number": "BI-SCREENCABLE", "description": "Screen Cable"},
                      {"part_number": "THUM620", "description": "Thumbscrew"}]}
    normalize_job(_pre)
    _pre_kept = [p["part_number"] for p in _pre["parts"]]
    ok("BI-SCREENCABLE" not in _pre_kept,
       f"an already-projected job is filtered too, got {_pre_kept}")
    ok("THUM620" in _pre_kept, f"and its genuine hardware survives, got {_pre_kept}")

    # Excluded, not vanished: an estimator has to be able to see what was dropped and why.
    ok(any("reference only" in str(w).lower() for w in (_job.get("warnings") or [])),
       "the exclusion is recorded on the job rather than being silent")
    ok(any("BI-SCREENCABLE" in str(w) for w in (_job.get("warnings") or [])),
       "and it names the row it dropped")


def test_an_operation_the_engine_adds_records_where_it_came_from():
    """STEP 0 OF THE ROUTE CUTOVER, and the reason it comes first.

    operation_sources was written in exactly one place — the LLM route fold — so every
    operation the deterministic reader, SolidWorks or the estimator itself concluded carried
    no source at all. rank("") is 0 and llm_full_extract is 40, so ranked arbitration over
    those claims would hand every contested operation to the model by default.

    12120's own shadow measured the scale: thirteen handling decisions, tapping and one
    welding all came out `unknown 0`. A third of the route with no provenance. Fixing that
    after switching arbitration on would make the compiler look wrong when the missing thing
    is underneath it.

    This changes no number. Nothing reads operation_sources for costing — the route compiler
    reads it in shadow, and the arbitration that follows at cutover.
    """
    from estimator import estimate_part, record_operation
    from source_precedence import rank

    _p = {"part_number": "X-01", "description": "BRACKET", "quantity": 1,
          "normalized_material": "MILD STEEL", "normalized_thickness_mm": 1.5,
          "material": "MILD STEEL", "thickness_mm": 1.5,
          "normalized_geometry": {"blank_length_mm": 120, "blank_width_mm": 80,
                                  "bend_count": 2, "cut_length_mm": 400, "hole_count": 3},
          "bend_count": 2, "cut_length_mm": 400,
          "manufacturing_interpretation": {"stock_form": "sheet",
                                           "operations": ["welding"]},
          "textual_operations": ["welding"]}
    estimate_part(_p, job_quantity=100)
    _src = _p.get("operation_sources") or {}

    # HANDLING WAS THE THIRTEEN. 12120's shadow reported thirteen handling decisions at
    # `unknown 0` — because handling was added to the local costing list and never to the
    # record. The commit that cited them did not fix them.
    for _op in ("laser_cutting", "dress_welds", "handling"):
        ok(_src.get(_op), f"the engine added '{_op}' and must record where it came from")
        ok(rank(_src[_op]) > 0,
           f"'{_op}' carries source '{_src.get(_op)}' which ranks 0 — arbitration cannot "
           f"weigh a claim it cannot attribute")

    # The RANKS have to mean something, not merely be non-zero. A weld the engine chained a
    # dressing from is a rule, not a guess, and must outrank a bare inference.
    ok(rank(_src["dress_welds"]) > rank(_src["laser_cutting"]),
       f"a chained rule ({_src['dress_welds']}) outranks an inference "
       f"({_src['laser_cutting']})")

    # Gap-fill, never overwrite: a stronger reader that already attributed an operation
    # keeps its attribution.
    _q = {"operation_sources": {"folding": "dxf"}}
    record_operation(_q, "folding", "inference")
    eq(_q["operation_sources"]["folding"], "dxf",
       "a measured attribution is not replaced by the engine's own guess")

    # RANK, NOT ARRIVAL ORDER. setdefault protected a strong attribution from a weak one and
    # equally blocked a MEASUREMENT from upgrading an earlier guess — the same defect facing
    # the other way. Whoever writes first must not win by writing first.
    _u = {"operation_sources": {"folding": "inference"}}
    record_operation(_u, "folding", "dxf")
    eq(_u["operation_sources"]["folding"], "dxf",
       "a measurement upgrades an earlier inference rather than being refused as a duplicate")

    # NO WRITER MAY BYPASS THE ONE WRITER. Seven call sites appended operations directly;
    # converting six and missing one would leave exactly the hole this exists to close, and
    # nothing would fail.
    import os as _os, re as _re
    _src_text = open(_os.path.join(
        _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
        "src", "estimator.py"), encoding="utf-8").read()
    _bare = _re.findall(r'\["(?:textual|inferred)_operations"\]\s*\.append\(', _src_text)
    ok(not _bare,
       f"{len(_bare)} operation(s) are appended without going through record_operation, so "
       f"they reach the route with no source — use record_operation instead")
def test_route_arbitration_is_ranked_but_metadata_is_gap_filled():
    """Status and metadata are different questions. A strong source can confirm an operation
    while weaker, non-conflicting route evidence supplies its assembly scope and sequence."""
    from route_compiler import (
        REQUIRED, RULED_OUT, UNVERIFIED, arbitrate_event, make_claim,
    )

    _strong = make_claim(
        "welding", REQUIRED, "solidworks_api", "101", "101",
        participants=["02M", "03M"], reason="native feature confirms weld")
    _route = make_claim(
        "welding", REQUIRED, "inference", "101", "101",
        scope="assembly", participants=["02M", "03M"], qty_per_unit=1,
        sequence=40, reason="weld upstand to base", route_id="route:weld-101")
    _decision = arbitrate_event("decision:weld-101", [_strong, _route])

    eq(_decision.status, REQUIRED, "strong measured positive claim decides status")
    eq(_decision.source, "solidworks_api", "the status source remains the strong source")
    eq(_decision.scope, "assembly", "missing scope is filled from compatible route evidence")
    eq(_decision.sequence, 40.0, "missing sequence is filled independently")
    eq(_decision.qty_per_unit, 1.0, "assembly multiplicity comes from the route")
    eq(_decision.participants, ["02M", "03M"],
       "participants describe work without multiplying it")

    _yes = make_claim("folding", REQUIRED, "dxf", "04M", "04M", scope="part")
    _no = make_claim("folding", RULED_OUT, "dxf", "04M", "04M", scope="part",
                     reason="zero measured bend lines")
    _conflict = arbitrate_event("decision:fold-04m", [_yes, _no])
    eq(_conflict.status, UNVERIFIED, "equal-rank disagreement must be explicit")
    eq(_conflict.qty_per_unit, None, "unverified work receives no invented quantity")
    ok(any(c.get("field") == "status" for c in _conflict.conflicts),
       "the conflict remains auditable")

    _weak_yes = make_claim("folding", REQUIRED, "inference", "04M", "04M")
    _measured_no = make_claim("folding", RULED_OUT, "dxf", "04M", "04M",
                              reason="DXF measured zero bend lines")
    _ruled = arbitrate_event("decision:fold-04m", [_weak_yes, _measured_no])
    eq(_ruled.status, RULED_OUT, "measured negative evidence outranks inference")
    eq(_ruled.qty_per_unit, None, "ruled-out work has no quantity")


def test_12120_job_route_is_one_event_per_assembly_not_per_part():
    """One weld targets 101, participant count is not quantity, assembly parents do not
    regain leaf operations, and a DXF ruling survives the LLM route."""
    from route_compiler import (
        NOT_APPLICABLE, REQUIRED, RULED_OUT, compile_job_route,
        project_priced_route,
    )

    _parts = [
        {"part_number": "12120-01-101", "description": "STAND WELD ASSY",
         "is_sub_assembly": True, "is_assembly_parent": True,
         "textual_operations": ["laser_cutting", "folding"],
         "operation_sources": {"laser_cutting": "inference", "folding": "inference"}},
        {"part_number": "12120-01-103", "description": "SCREEN MOUNTING BRACKET",
         "is_sub_assembly": True, "is_assembly_parent": True,
         "textual_operations": ["laser_cutting"],
         "operation_sources": {"laser_cutting": "inference"}},
        {"part_number": "12120-01-01M"},
        {"part_number": "12120-01-02M", "textual_operations": ["welding"]},
        {"part_number": "12120-01-03M", "textual_operations": ["welding"]},
        {"part_number": "12120-01-04M",
         "operations_ruled_out": {"folding": "DXF measured zero bend lines"},
         "operation_ruling_sources": {"folding": "dxf"}},
        {"part_number": "12120-01-08M"},
    ]
    _extract = {
        "top_assembly": {"part_number": "12120-01-GA"},
        "assemblies": [
            {"part_number": "12120-01-GA", "children": [
                {"part_number": "12120-01-101", "qty": 1},
                {"part_number": "12120-01-103", "qty": 1},
                {"part_number": "12120-01-04M", "qty": 1},
                {"part_number": "12120-01-08M", "qty": 1},
            ]},
            {"part_number": "12120-01-101", "children": [
                {"part_number": "12120-01-02M", "qty": 1},
                {"part_number": "12120-01-03M", "qty": 1},
            ]},
            {"part_number": "12120-01-103", "children": [
                {"part_number": "12120-01-01M", "qty": 1},
            ]},
        ],
        "routes": [
            {"operation": "welding", "sequence": 40,
             "part_numbers": ["12120-01-02M", "12120-01-03M"],
             "scope": "assembly", "qty_per_unit": 1, "inferred": True},
            {"operation": "folding", "sequence": 30,
             "part_numbers": ["12120-01-04M"],
             "scope": "part", "qty_per_unit": 1, "inferred": True},
            {"operation": "powder_coating", "sequence": 50,
             "part_numbers": [
                 "12120-01-01M", "12120-01-04M", "12120-01-08M",
                 "12120-01-101", "12120-01-103"],
             "scope": "assembly", "qty_per_unit": 1, "inferred": True},
        ],
    }

    _graph = compile_job_route(_parts, _extract)
    _decisions = _graph["decisions"]
    _welds = [d for d in _decisions
              if d["operation"] == "welding" and d["status"] == REQUIRED]
    eq(len(_welds), 1, "two participating parts compile to one welding event")
    eq(_welds[0]["target_id"], "12120-01-101", "the hierarchy owns the weld")
    eq(_welds[0]["participants"], ["12120-01-02M", "12120-01-03M"],
       "both children remain visible on that event")
    eq(_welds[0]["qty_per_unit"], 1.0,
       "two participants do not become two welding charges")

    _parent_leaf = [
        d for d in _decisions
        if d["target_id"] in ("12120-01-101", "12120-01-103")
        and d["operation"] in ("laser_cutting", "folding")
    ]
    ok(_parent_leaf, "raw parent claims remain auditable")
    ok(all(d["status"] == NOT_APPLICABLE for d in _parent_leaf),
       "hierarchy rules them out instead of deleting or costing them")

    _fold_04 = [d for d in _decisions
                if d["target_id"] == "12120-01-04M"
                and d["operation"] == "folding"]
    eq(len(_fold_04), 1, "inferred and measured fold claims arbitrate once")
    eq(_fold_04[0]["status"], RULED_OUT, "DXF ruling survives the route fold")
    eq(_fold_04[0]["qty_per_unit"], None, "the ruled-out fold has no multiplicity")

    _powder_targets = sorted(
        d["target_id"] for d in _decisions
        if d["operation"] == "powder_coating" and d["status"] == REQUIRED)
    eq(_powder_targets, [
        "12120-01-04M", "12120-01-08M", "12120-01-101", "12120-01-103"],
       "assemblies and standalone leaves are distinct without coating 01M twice")

    _estimates = [
        {"part_number": "12120-01-02M",
         "labour_estimate": {"costs_gbp": {"welding": 0.50}}},
        {"part_number": "12120-01-03M",
         "labour_estimate": {"costs_gbp": {"welding": 0.50}}},
        {"part_number": "12120-01-101",
         "labour_estimate": {"costs_gbp": {"laser_cutting": 2.00}}},
    ]
    _shadow = project_priced_route(_graph, _estimates)
    _weld_rows = [r for r in _shadow["priced_route_rows"]
                  if r["operation"] == "welding"]
    eq(len(_weld_rows), 1, "one decision produces one shadow weld row")
    eq(_weld_rows[0]["qty_per_unit"], 1.0, "projection keeps route multiplicity")
    ok(any(i.get("code") == "assembly_operation_costed_on_multiple_participants"
           for i in _shadow["issues"]),
       "legacy participant charging is exposed before cutover")
    ok(any(i.get("code") == "forbidden_decision_priced"
           for i in _shadow["issues"]),
       "legacy Laser on a parent is identified as resurrection")


def test_estimator_preserves_route_context_without_changing_legacy_fields():
    """The costed record retains route evidence in a nested shadow field."""
    from estimator import estimate_part

    _part = {
        "part_number": "TEST-01M",
        "normalized_material": "MILD_STEEL",
        "normalized_thickness_mm": 1.5,
        "overall_length_mm": 100,
        "overall_width_mm": 50,
        "quantity": 1,
        "textual_operations": ["folding"],
        "inferred_operations": ["laser_cutting"],
        "operation_sources": {"folding": "drawing_deterministic",
                              "laser_cutting": "inference"},
        "operation_scope": {"folding": "part"},
        "operation_sequence": {"folding": 30},
        "operation_qty_per_unit": {"folding": 1},
        "operations_ruled_out": {"punch": "DXF measured no punched features"},
        "operation_ruling_sources": {"punch": "dxf"},
        "manufacturing_features": {"bend_count": 1},
        "geometry_rollup": {"estimated_cut_length_mm": 300.0},
    }
    _estimate = estimate_part(_part, job_quantity=180)
    _context = _estimate.get("route_context") or {}
    eq(_context.get("textual_operations"), _part["textual_operations"],
       "costed record preserves required route names")
    eq((_context.get("operation_scope") or {}).get("folding"), "part",
       "scope reaches the costed record")
    eq((_context.get("operation_sequence") or {}).get("folding"), 30,
       "sequence reaches the costed record")
    eq((_context.get("operations_ruled_out") or {}).get("punch"),
       "DXF measured no punched features", "negative decisions reach the costed record")
    ok("textual_operations" not in _estimate,
       "legacy top-level fields remain untouched during shadow mode")


def test_2085_legacy_route_compiles_to_one_assembly_event():
    """Old cached extracts have no scope field. Hierarchy still makes one weld and one coat,
    Dress inherits that weld event, and sheet-Laser text cannot profile tube stock."""
    from route_compiler import NOT_APPLICABLE, REQUIRED, compile_job_route

    _parts = [
        {"part_number": "2085-01", "description": "BRACKET PLATE",
         "textual_operations": ["powder_coating", "welding", "laser_cutting"],
         "inferred_operations": ["dress_welds"]},
        {"part_number": "2085-02", "description": "OUTER TUBE",
         "section_stock": {"profile_form": "CHS"},
         "textual_operations": [
             "powder_coating", "tube_cut", "welding", "laser_cutting"],
         "inferred_operations": ["dress_welds"]},
        {"part_number": "2085-03", "description": "INNER TUBE",
         "section_stock": {"profile_form": "CHS"},
         "textual_operations": [
             "powder_coating", "tube_cut", "welding", "laser_cutting"],
         "inferred_operations": ["dress_welds"]},
    ]
    _extract = {
        "top_assembly": {"part_number": "2085-GA"},
        "assemblies": [{"part_number": "2085-GA", "children": [
            {"part_number": "2085-01", "qty": 1},
            {"part_number": "2085-02", "qty": 1},
            {"part_number": "2085-03", "qty": 1},
        ]}],
        "routes": [
            {"operation": "powder_coating", "sequence": 10,
             "part_numbers": ["2085-01", "2085-02", "2085-03"],
             "inferred": False},
            {"operation": "tube_cut", "sequence": 10,
             "part_numbers": ["2085-02", "2085-03"], "inferred": True},
            {"operation": "laser_cutting", "sequence": 20,
             "part_numbers": ["2085-01"], "inferred": True},
            {"operation": "welding", "sequence": 30,
             "part_numbers": ["2085-01", "2085-02", "2085-03"],
             "inferred": True},
        ],
    }
    _graph = compile_job_route(_parts, _extract)
    _decisions = _graph["decisions"]

    for _operation in ("welding", "powder_coating", "dress_welds"):
        _events = [d for d in _decisions
                   if d["operation"] == _operation and d["status"] == REQUIRED]
        eq(len(_events), 1, f"{_operation} is one job event, not three part events")
        eq(_events[0]["target_id"], "2085-GA", f"{_operation} belongs to the assembly")
        eq(_events[0]["scope"], "assembly", f"{_operation} retains assembly scope")
        eq(_events[0]["qty_per_unit"], 1.0, f"{_operation} occurs once per product")

    _tube_lasers = [
        d for d in _decisions
        if d["operation"] == "laser_cutting"
        and d["target_id"] in ("2085-02", "2085-03")
    ]
    eq(len(_tube_lasers), 2, "both stale tube-Laser claims remain auditable")
    ok(all(d["status"] == NOT_APPLICABLE for d in _tube_lasers),
       "tube stock routes to tube_cut rather than sheet Laser")
    _tube_cuts = [
        d for d in _decisions
        if d["operation"] == "tube_cut" and d["status"] == REQUIRED
    ]
    eq(sorted(d["target_id"] for d in _tube_cuts), ["2085-02", "2085-03"],
       "each tube retains its real cut event")


def test_estimate_document_stamps_the_shadow_route_without_using_it_for_price():
    """The real estimate_document call site runs the compiler and stamps its audit contract."""
    from estimator import estimate_document

    _part = {
        "part_number": "TEST-02M",
        "normalized_material": "MILD_STEEL",
        "normalized_thickness_mm": 1.5,
        "overall_length_mm": 80,
        "overall_width_mm": 40,
        "quantity": 1,
        "textual_operations": ["laser_cutting"],
        "geometry_rollup": {"estimated_cut_length_mm": 240.0},
    }
    _summary = {
        "assumed_job_quantity": 180,
        "llm_full_extract": {
            "routes": [{
                "operation": "laser_cutting", "part_numbers": ["TEST-02M"],
                "scope": "part", "qty_per_unit": 1, "sequence": 10,
                "inferred": True,
            }],
            "assemblies": [],
        },
    }
    _estimate = estimate_document([_part], summary=_summary)
    _shadow = _estimate.get("canonical_route_shadow") or {}
    eq(_shadow.get("schema"), "priced_route_shadow.v1",
       "estimate_document stamps the canonical shadow contract")
    ok(not _shadow.get("compiler_error"), "the integration executes the real compiler")
    eq(len(_shadow.get("priced_route_rows") or []), 1,
       "one required decision projects to one shadow row")
    ok(_estimate.get("part_estimates"),
       "legacy part estimates remain the source of the live total during shadow mode")


def test_route_shadow_invariant_reports_but_does_not_block_before_cutover():
    """Shadow differences are loud but cannot alter a live quote before the switch."""
    from invariants import check_canonical_route_shadow

    _summary = {"estimate_summary": {"canonical_route_shadow": {
        "schema": "priced_route_shadow.v1",
        "mode": "shadow",
        "decisions": [{
            "decision_id": "decision:x", "operation": "laser_cutting",
            "target_id": "101", "status": "not_applicable", "conflicts": []}],
        "priced_route_rows": [],
        "issues": [{
            "code": "forbidden_decision_priced",
            "decision_id": "decision:x", "operation": "laser_cutting"}],
    }}}
    _violations = check_canonical_route_shadow(_summary)
    ok(_violations, "a resurrection discrepancy must be reported")
    eq({v["severity"] for v in _violations}, {"warning"},
       "shadow diagnostics do not change the firm-price gate before cutover")


def test_hierarchy_quantities_and_pressed_hardware_become_canonical_route_facts():
    from route_compiler import REQUIRED, compile_job_route

    _parts = [
        {"part_number": "ASSY", "description": "WELD ASSY"},
        {"part_number": "PLATE", "description": "PLATE", "quantity": 1},
        {"part_number": "BI-PEMSTUD", "description": "M4 THREADED PEM STUD",
         "page_roles": ["bought_in"], "quantity": 2},
        {"part_number": "BI-CLINCH", "description": "M4 SELF-CLINCH NUT",
         "page_roles": ["bought_in"], "quantity": 4},
    ]
    _extract = {
        "assemblies": [{"part_number": "ASSY", "children": [
            {"part_number": "PLATE", "qty": 1},
            {"part_number": "STD PART", "qty": 2},
            {"part_number": "FIXING", "qty": 4},
        ]}],
        "parts": [
            {"part_number": "STD PART", "description": "M4 THREADED PEM STUD",
             "is_bought_in": True},
            {"part_number": "FIXING", "description": "M4 SELF-CLINCH NUT",
             "is_bought_in": True},
        ],
        "routes": [],
    }
    _graph = compile_job_route(_parts, _extract)
    _nodes = {n["part_number"]: n for n in _graph["nodes"]}
    eq(_nodes["ASSY"]["kind"], "assembly", "the parent is not a sheet part")
    eq(_nodes["STD PART"]["kind"], "bought_in",
       "explicit BOM evidence classifies hardware")
    eq(_nodes["STD PART"]["qty_per_unit"], 2.0,
       "hierarchy quantity beats a flat guess")
    eq(_nodes["FIXING"]["qty_per_unit"], 4.0, "all child multiplicities survive")
    eq(_nodes["STD PART"]["evidence"]["raw_aliases"], ["BI-PEMSTUD"],
       "generated bought-in identities reconcile to the explicit BOM code")
    _insert = [
        d for d in _graph["decisions"]
        if d["operation"] == "hardware_insertion" and d["status"] == REQUIRED
    ]
    eq(len(_insert), 1, "pressed hardware produces one assembly route event")
    eq(_insert[0]["target_id"], "ASSY", "the assembly owns the insertion event")
    eq(_insert[0]["participants"], ["FIXING", "STD PART"],
       "the inserted BOM lines remain auditable")
    eq(_insert[0]["qty_per_unit"], 1.0,
       "six inserts do not become six assembly events")
    ok(_insert[0]["sequence"] < 20,
       "pressed hardware is routed before the folding stage")


def test_canonical_workbook_groups_price_decisions_not_raw_route_words():
    from route_compiler import compile_job_route, project_priced_route
    from wb_populate import canonical_labour_groups

    _parts = [
        {"part_number": "ASSY", "description": "WELD ASSY",
         "textual_operations": ["laser_cutting"]},
        {"part_number": "P1", "description": "PLATE 1",
         "normalized_material": "MILD_STEEL", "normalized_thickness_mm": 1.5},
        {"part_number": "P2", "description": "PLATE 2",
         "normalized_material": "MILD_STEEL", "normalized_thickness_mm": 1.5},
        {"part_number": "PEM", "description": "M4 THREADED PEM STUD",
         "page_roles": ["bought_in"], "quantity": 2},
    ]
    _extract = {
        "top_assembly": {"part_number": "ASSY"},
        "assemblies": [{"part_number": "ASSY", "children": [
            {"part_number": "P1", "qty": 1},
            {"part_number": "P2", "qty": 1},
            {"part_number": "PEM", "qty": 2},
        ]}],
        "routes": [
            {"operation": "laser_cutting", "sequence": 10, "scope": "part",
             "part_numbers": ["P1", "P2"], "inferred": True},
            {"operation": "welding", "sequence": 30, "scope": "assembly",
             "part_numbers": ["P1", "P2"], "inferred": True},
        ],
    }
    _estimates = [
        {"part_number": "P1", "quantity": 1, "normalized_material": "MILD_STEEL",
         "normalized_thickness_mm": 1.5,
         "material_estimate": {"stock_form": "sheet"},
         "labour_estimate": {"batch_hours": {"laser_cutting": 1.0}}},
        {"part_number": "P2", "quantity": 1, "normalized_material": "MILD_STEEL",
         "normalized_thickness_mm": 1.5,
         "material_estimate": {"stock_form": "sheet"},
         "labour_estimate": {"batch_hours": {"laser_cutting": 1.5}}},
        {"part_number": "PEM", "quantity": 2, "page_roles": ["bought_in"],
         "labour_estimate": {}},
    ]
    _graph = compile_job_route(_parts, _extract)
    _shadow = project_priced_route(_graph, _estimates)
    _summary = {
        "parts": _parts,
        "estimate_summary": {
            "part_estimates": _estimates,
            "canonical_route_shadow": _shadow,
        },
    }
    _groups = canonical_labour_groups(_summary, _estimates, 180)
    _laser = [g for g in _groups.values() if g["wb_op"] == "Laser (Metal)"]
    eq(len(_laser), 1, "same-gauge leaf decisions share one tooling setup")
    eq(_laser[0]["qty"], 2.0, "both required leaf events reach that setup")
    eq(len(_laser[0]["decision_ids"]), 2,
       "grouping retains both canonical identities")
    _weld = [g for g in _groups.values() if g["wb_op"] == "Weld (CO2)"]
    eq(len(_weld), 1, "assembly welding is one priced event")
    eq(_weld[0]["qty"], 1.0, "participants do not multiply assembly welding")
    eq(_weld[0]["bh"], 0.0,
       "participant batch hours cannot leak into assembly pricing")
    ok(not any(
        g["wb_op"] == "Laser (Metal)" and "ASSY" in g.get("parts", [])
        for g in _groups.values()
    ), "the assembly parent's raw Laser word is not resurrected")
    _insert = [
        g for g in _groups.values()
        if "hardware_insertion" in g.get("engine_ops", [])
    ]
    eq(len(_insert), 1, "pressed hardware is rendered from its decision")
    eq(_insert[0]["qty"], 1.0, "insertion is one assembly event")
    eq(round(_insert[0]["bh"], 4), 1.5,
       "two inserts x 15 seconds x 180 units becomes 1.5 batch hours")


def test_canonical_workbook_rows_carry_every_source_decision_exactly_once():
    from wb_populate import build_workbook_labour

    _groups = {
        ("setup",): {
            "workbook_row": 95,
            "wb_op": "Laser (Metal)",
            "material": "MILD_STEEL",
            "thickness": 1.5,
            "parts": ["P1", "P2"],
            "group_key": ("setup",),
            "engine_ops": ["laser_cutting"],
            "qty": 2,
            "canonical_route": True,
            "decision_ids": ["decision:p1", "decision:p2"],
        },
    }
    _record = build_workbook_labour(_groups)
    eq(_record["schema"], "workbook_labour_rows.v3",
       "canonical identity has an explicit schema")
    eq(_record["mode"], "canonical", "the renderer records which authority it used")
    eq(_record["rows"][0]["decision_ids"], ["decision:p1", "decision:p2"],
       "shared setup does not erase either route decision")


def test_cutover_invariants_block_missing_unverified_or_resurrected_decisions():
    from invariants import check_canonical_route_shadow

    _summary = {
        "estimate_summary": {"canonical_route_shadow": {
            "schema": "priced_route_shadow.v1",
            "mode": "cutover",
            "decisions": [
                {"decision_id": "required", "operation": "welding",
                 "target_id": "ASSY", "status": "required", "conflicts": []},
                {"decision_id": "uncertain", "operation": "folding",
                 "target_id": "P1", "status": "unverified",
                 "conflicts": [{"field": "status"}]},
                {"decision_id": "no", "operation": "laser_cutting",
                 "target_id": "ASSY", "status": "not_applicable", "conflicts": []},
            ],
            "priced_route_rows": [
                {"decision_id": "required", "operation": "welding"},
            ],
            "issues": [],
        }},
        "workbook_labour": {
            "schema": "workbook_labour_rows.v3",
            "mode": "canonical",
            "rows": [{
                "workbook_row": 95,
                "decision_ids": ["no"],
                "wb_operation": "Laser (Metal)",
            }],
        },
    }
    _violations = check_canonical_route_shadow(_summary)
    eq({v["severity"] for v in _violations}, {"blocking"},
       "cutover discrepancies cannot be advisory")
    _codes = {v["code"] for v in _violations}
    ok("canonical_route_decision_unverified" in _codes,
       "equal-rank uncertainty blocks automatic pricing")
    ok("canonical_non_required_decision_priced" in _codes,
       "a ruled-out event cannot reach the workbook")
    ok("canonical_required_decision_not_rendered" in _codes,
       "required work cannot disappear during rendering")


def test_canonical_cutover_never_falls_back_to_the_legacy_workbook_builder():
    from pathlib import Path
    _root = Path(__file__).resolve().parents[1]
    _main = (_root / "src" / "main.py").read_text(encoding="utf-8")
    ok("NO fallback workbook written: canonical route cutover is" in _main,
       "a compiler failure must fail closed")
    ok("if _canonical_cutover:" in _main,
       "the fallback guard reads the real production switch")
    _wb = (_root / "src" / "wb_populate.py").read_text(encoding="utf-8")
    ok("for pe in ([] if _canonical_cutover else labour_parts):" in _wb,
       "canonical mode cannot execute the raw-route rescue loop")


def test_canonical_bom_identity_quantity_and_missing_lines_reach_the_workbook():
    from wb_populate import canonicalise_part_estimates_for_workbook

    _summary = {"estimate_summary": {"canonical_route_shadow": {
        "nodes": [
            {
                "part_number": "GA", "description": "TOP", "kind": "assembly",
                "qty_per_unit": 1, "parents": [], "children": [
                    {"part_number": "FIXINGTBC", "qty": 2},
                    {"part_number": "THUM620", "qty": 4},
                    {"part_number": "PLATE", "qty": 1},
                    {"part_number": "MISSING-LEAF", "qty": 1},
                ], "evidence": {},
            },
            {
                "part_number": "FIXINGTBC", "description": "M4 KNURLED KNOB",
                "kind": "bought_in", "qty_per_unit": 2, "parents": ["GA"],
                "children": [], "evidence": {"raw_aliases": ["BI-KNURLEDKNOB"]},
            },
            {
                "part_number": "THUM620", "description": "M4 THUMBSCREW",
                "kind": "bought_in", "qty_per_unit": 4, "parents": ["GA"],
                "children": [], "evidence": {},
            },
            {
                "part_number": "PLATE", "description": "PLATE",
                "kind": "leaf", "qty_per_unit": 1, "parents": ["GA"],
                "children": [], "evidence": {},
            },
            {
                "part_number": "MISSING-LEAF", "description": "MISSING PLATE",
                "kind": "leaf", "qty_per_unit": 1, "parents": ["GA"],
                "children": [], "evidence": {},
            },
        ],
        "decisions": [],
        "issues": [],
    }}}
    _estimates = [
        {
            "part_number": "BI-KNURLEDKNOB", "description": "KNOB",
            "quantity": 1, "unit_cost_gbp": 1.25,
            "material_estimate": {"stock_form": "bought_in"},
        },
        {
            "part_number": "PLATE", "description": "PLATE", "quantity": 1,
            "material_estimate": {"stock_form": "sheet"},
        },
    ]
    _normalised = canonicalise_part_estimates_for_workbook(_summary, _estimates)
    _by_id = {item["part_number"]: item for item in _normalised}
    eq(_by_id["FIXINGTBC"]["quantity"], 2,
       "explicit hierarchy quantity replaces the generated alias guess")
    eq(_by_id["FIXINGTBC"]["unit_cost_gbp"], 1.25,
       "identity reconciliation retains the priced evidence")
    eq(_by_id["THUM620"]["quantity"], 4,
       "an explicit unpriced hardware line remains visible")
    ok(_by_id["THUM620"]["_price_explicitly_withheld"],
       "missing hardware pricing is explicit rather than silently absent")
    ok("MISSING-LEAF" not in _by_id,
       "fabricated geometry is never invented to fill a missing estimate")
    _issues = _summary["estimate_summary"]["canonical_route_shadow"]["issues"]
    ok(any(
        issue.get("code") == "bom_leaf_without_estimate"
        and issue.get("part_number") == "MISSING-LEAF"
        for issue in _issues
    ), "a missing fabricated BOM leaf becomes a release blocker")


def test_canonical_powder_grouping_respects_finish_and_empty_routes_keep_v3():
    from wb_populate import build_workbook_labour, canonical_labour_groups

    _decisions = [
        {
            "decision_id": "coat:black", "operation": "powder_coating",
            "status": "required", "target_id": "BLACK", "participants": ["BLACK"],
            "scope": "part", "qty_per_unit": 1, "sequence": 70,
        },
        {
            "decision_id": "coat:white", "operation": "powder_coating",
            "status": "required", "target_id": "WHITE", "participants": ["WHITE"],
            "scope": "part", "qty_per_unit": 1, "sequence": 70,
        },
    ]
    _summary = {
        "parts": [
            {"part_number": "BLACK", "normalized_finish": "RAL 9005 BLACK"},
            {"part_number": "WHITE", "normalized_finish": "RAL 9010 WHITE"},
        ],
        "estimate_summary": {"canonical_route_shadow": {
            "nodes": [
                {"part_number": "BLACK", "kind": "leaf", "qty_per_unit": 1},
                {"part_number": "WHITE", "kind": "leaf", "qty_per_unit": 1},
            ],
            "decisions": _decisions,
            "issues": [],
        }},
    }
    _estimates = [
        {"part_number": "BLACK", "normalized_finish": "RAL 9005 BLACK"},
        {"part_number": "WHITE", "normalized_finish": "RAL 9010 WHITE"},
    ]
    _groups = canonical_labour_groups(_summary, _estimates, 180)
    _powder = [
        group for group in _groups.values()
        if "powder_coating" in group.get("engine_ops", [])
    ]
    eq(len(_powder), 2, "different powder colours require different setup rows")
    _empty = build_workbook_labour({}, canonical_mode=True)
    eq(_empty["schema"], "workbook_labour_rows.v3",
       "a valid no-operation canonical job does not masquerade as legacy")
    eq(_empty["mode"], "canonical", "authority survives an empty route")


def test_shop_sequence_fills_only_required_events_after_arbitration():
    from route_compiler import RULED_OUT, compile_job_route

    _graph = compile_job_route(
        [
            {
                "part_number": "P1",
                "textual_operations": ["tapping"],
                "operations_ruled_out": {"folding": "DXF measured zero bends"},
                "operation_ruling_sources": {"folding": "dxf"},
            },
        ],
        {"routes": [], "assemblies": []},
    )
    _by_operation = {
        decision["operation"]: decision for decision in _graph["decisions"]
    }
    eq(_by_operation["tapping"]["sequence"], 20.0,
       "unsequenced tapping is placed before folding and welding")
    eq(_by_operation["tapping"]["field_provenance"]["sequence"],
       "shop_sequence_rule", "the fallback order remains auditable")
    eq(_by_operation["folding"]["status"], RULED_OUT,
       "shop ordering cannot promote a negative decision")
    eq(_by_operation["folding"]["sequence"], None,
       "ruled-out work receives no invented route position")


def test_child_finish_words_do_not_duplicate_an_owning_assembly_coat():
    from route_compiler import REQUIRED, UNVERIFIED, compile_job_route

    _parts = [
        {
            "part_number": "WELD-ASSY", "description": "WELD ASSY",
            "textual_operations": ["powder_coating"],
            "operation_scope": {"powder_coating": "assembly"},
        },
        {
            "part_number": "CHILD-A", "description": "CHILD A",
            "normalized_finish": "SEE ASSEMBLY DRAWING",
            "textual_operations": ["powder_coating"],
        },
        {
            "part_number": "LOOSE-COVER", "description": "LOOSE COVER",
            "normalized_finish": "SEE ASSEMBLY DRAWING",
            "textual_operations": ["powder_coating"],
        },
    ]
    _extract = {
        "top_assembly": {"part_number": "TOP"},
        "assemblies": [
            {"part_number": "TOP", "children": [
                {"part_number": "WELD-ASSY", "qty": 1},
                {"part_number": "LOOSE-COVER", "qty": 1},
            ]},
            {"part_number": "WELD-ASSY", "children": [
                {"part_number": "CHILD-A", "qty": 1},
            ]},
        ],
        "routes": [{
            "operation": "powder_coating", "scope": "assembly",
            "part_numbers": ["WELD-ASSY"], "sequence": 70,
            "inferred": True,
        }],
    }
    _graph = compile_job_route(_parts, _extract)
    _powder = [
        decision for decision in _graph["decisions"]
        if decision["operation"] == "powder_coating"
    ]
    _required = [d for d in _powder if d["status"] == REQUIRED]
    eq(len(_required), 1,
       "a child finish word corroborates rather than duplicates its assembly coat")
    eq(_required[0]["target_id"], "WELD-ASSY",
       "the explicit assembly remains the owner")
    _uncertain = [d for d in _powder if d["status"] == UNVERIFIED]
    eq(len(_uncertain), 1,
       "a delegated finish with no owning route remains visible for review")
    eq(_uncertain[0]["target_id"], "TOP",
       "the unresolved event points at the assembly it delegates to")


def test_an_unattributed_claim_is_never_promoted_to_a_measurement():
    """An operation nobody attributed must rank 0, whatever its status.

    The temptation is to give a NEGATIVE claim the benefit of the doubt — a "ruled out" with
    no recorded source looks safe to treat as measured, because refusing work is the
    cautious direction. It is not safe. Rank is how this engine decides which evidence wins,
    and handing an unattributed ruling dxf's rank of 80 lets an unsourced exclusion beat a
    real reading. That silently deletes work, which is the more expensive error and the one
    nobody sees on a sheet.

    The compiler already does this correctly. Nothing was checking it: promoting `unknown`
    to `dxf` broke no fixture, so the rule could be undone by a future edit with the suite
    still green. 12120's shadow shows thirteen decisions arriving as `unknown 0` — this is
    the rule that keeps them honest instead of quietly authoritative.
    """
    from route_compiler import make_claim, REQUIRED, RULED_OUT, NOT_APPLICABLE
    from source_precedence import rank

    for _status in (REQUIRED, RULED_OUT, NOT_APPLICABLE):
        for _source in ("", None, "   "):
            _c = make_claim("folding", _status, _source, "P1", "P1")
            eq(_c.source, "unknown",
               f"a {_status} claim with source {_source!r} is unattributed, not measured")
            eq(_c.source_rank, 0,
               f"and it must rank 0 — got {_c.source_rank}, which would let an unsourced "
               f"claim outrank a real reading")

    # A REAL SOURCE STILL RANKS. Without this the fixture would pass against a compiler
    # that ranked everything 0, which would be just as broken in the other direction.
    _m = make_claim("folding", RULED_OUT, "dxf", "P1", "P1")
    eq(_m.source, "dxf", "a measured ruling keeps its source")
    eq(_m.source_rank, rank("dxf"), "and its rank")
    ok(_m.source_rank > 0, "which must actually be greater than an unattributed one")


def test_an_engine_minted_code_yields_to_the_canonical_one():
    """12120 shipped the same PEM stud twice: STD PART qty 2 from the GA BOM, and
    BI-PEMSTUD qty 2 minted by the prose recogniser from the same words. Four studs where
    the drawing calls for two — free while both were unpriced, wrong the moment either got
    a price.

    A "BI-" code is not a part number a draughtsman wrote; it is one this engine invented
    from a phrase. Where a synthesised code describes an item the canonical graph already
    holds, the graph wins. Keyed on that distinction rather than on PEM studs, so any
    recogniser-minted duplicate of a drawing item resolves the same way.
    """
    from wb_populate import canonicalise_part_estimates_for_workbook as _cz

    _summary = {"estimate_summary": {"canonical_route_shadow": {
        "decisions": [{"operation": "handling"}],
        "nodes": [
            {"part_number": "STD PART", "kind": "bought_in", "qty_per_unit": 2,
             "description": "M4 THREADED PEM STUD (LENGTH: 30mm)"},
            {"part_number": "12120-01-01M", "kind": "leaf", "qty_per_unit": 1,
             "description": "MOUNTING BRACKET"},
        ]}}}
    _pes = [
        {"part_number": "STD PART", "description": "M4 THREADED PEM STUD (LENGTH: 30mm)",
         "quantity": 2},
        {"part_number": "BI-PEMSTUD", "description": "M4 THREADED PEM STUD (LENGTH: 30mm)",
         "quantity": 2, "material_estimate": {"unit_material_cost_gbp": 0.44}},
        {"part_number": "12120-01-01M", "description": "MOUNTING BRACKET", "quantity": 1},
        {"part_number": "BI-KNURLEDKNOB", "description": "Knurled Knob", "quantity": 2},
    ]
    _out = _cz(_summary, _pes)
    _codes = [p.get("part_number") for p in _out]

    eq(_codes.count("STD PART"), 1, "the stud appears once")
    ok("BI-PEMSTUD" not in _codes,
       f"the engine-minted duplicate is absorbed, not listed beside it, got {_codes}")
    _stud = [p for p in _out if p["part_number"] == "STD PART"][0]
    eq(_stud.get("quantity"), 2, "at the quantity the drawing states, not doubled")
    ok((_stud.get("material_estimate") or {}).get("unit_material_cost_gbp"),
       "and the costed record survives the merge — identity reconciliation must never "
       "discard the pricing evidence")
    eq(_stud.get("_canonical_source_part_number"), "BI-PEMSTUD",
       "what it was merged from stays recorded")

    # A SYNTHESISED CODE WITH NO CANONICAL TWIN IS LEFT ALONE. Without this the fixture
    # would pass against a rule that simply deleted every BI- line.
    ok("BI-KNURLEDKNOB" in _codes,
       f"a bought-in the graph does not name is kept, got {_codes}")

    # AND A REAL CODE IS NEVER ABSORBED, however well its description matches. Only codes
    # this engine MINTED yield. Dropping the "BI-" test made no fixture fail the first time
    # this was mutated, which would have let a genuine second part with a similar
    # description be silently merged away — a missing BOM line, the error that costs most
    # and shows least.
    _rival = [
        {"part_number": "STD PART", "description": "M4 THREADED PEM STUD (LENGTH: 30mm)",
         "quantity": 2},
        {"part_number": "PEM-M4-30", "description": "M4 THREADED PEM STUD (LENGTH: 30mm)",
         "quantity": 2},
    ]
    _rc = [p.get("part_number") for p in _cz(_summary, _rival)]
    ok("PEM-M4-30" in _rc,
       f"a drawing code is not absorbed by another drawing code, got {_rc}")


def test_the_drill_row_keeps_the_rate_table_title_and_says_what_it_is():
    """DRIL's column-H title is literally "Drill (Acrylic)", and the shop books metal
    drilling and tapping to that same row.

    The TITLE cannot change: it is the lookup key. A title the rate table does not carry
    returns a rate of zero and costs the work at nothing — which is how GRIN silently zeroed
    every deburr on every job. So the title stays and the DESCRIPTION says what the work is,
    which is the half an estimator reads.

    The first version of this checked that a string appeared in wb_populate.py. That proves
    the line was typed, not that it reaches a row — the failure this run has been correcting
    all day. This drives the builder.
    """
    from wb_populate import labour_row_description as _d
    from department_codes import CODE_TITLES

    eq(CODE_TITLES["DRIL"][0], "Drill (Acrylic)",
       "the rate-table title is the lookup key and must not be renamed")

    _metal = _d("Drill (Acrylic)", "MILD STEEL", 1.5, ["12120-01-01M"])
    ok(_metal.startswith("Drill (Acrylic)"),
       f"the row still leads with the exact rate-table title, got {_metal!r}")
    ok("metal drill/tap" in _metal,
       f"and says what the work actually is, got {_metal!r}")

    # AN ACRYLIC PART NEEDS NO NOTE. Without this the fixture would pass against a version
    # that appended the note to every drill row, which would be noise on the parts the
    # department is actually named for.
    _acrylic = _d("Drill (Acrylic)", "ACRYLIC", 5, ["X-1"])
    ok("metal drill/tap" not in _acrylic,
       f"an acrylic part drilled on the acrylic row needs no explanation, got {_acrylic!r}")

    # AND NO OTHER DEPARTMENT IS TOUCHED.
    _fold = _d("Fold", "MILD STEEL", 1.5, ["A", "B"], bends=2)
    ok("metal drill/tap" not in _fold, f"the note is DRIL-only, got {_fold!r}")
    ok(_fold.startswith("Fold —"), f"and other rows read as before, got {_fold!r}")

    # AND THE ROW USES IT. Driving the builder proves the builder; a labour row that
    # constructed its own text would carry none of this and nothing here would notice.
    # This checks the CALL, not that a phrase appears somewhere in the file.
    import os as _os
    _wb = open(_os.path.join(
        _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
        "src", "wb_populate.py"), encoding="utf-8").read()
    ok("_rd = labour_row_description(wb_op," in _wb,
       "the labour row takes its description from the shared builder")


def test_two_names_for_one_department_are_one_operation():
    """OP_NAME_MAP carries synonyms, so Assemble/pack (Metal) inverts to handling, assembly
    AND assemble. The canonical workbook row declares engine_operations ['assembly']; the
    part record carries 'handling'. Same bench work, two names for it.

    Comparing the NAMES reported 12120's handling as work nobody charged for, while the
    three PACM rows on the sheet were charging for exactly that. A false positive here is
    not noise: it invites adding a handling row on top of the assembly/pack that already
    contains it, and double-charging the bench.

    An earlier attempt treated this as supersession — excusing handling wherever a canonical
    assembly decision covered the part. That was a guess at the cause, it was a no-op in the
    case its own fixture tested, and it would have excused a genuinely uncharged handling
    operation on any part an assembly event touched. This fixture uses the row shape the
    live run actually produced, so it cannot pass for the wrong reason.
    """
    from invariants import check_no_unpriced_operations_named as _chk
    from costed_facts import _dept_to_engine_ops

    _sibs = _dept_to_engine_ops().get("assemble/pack (metal)") or []
    ok("handling" in _sibs and "assembly" in _sibs,
       f"handling and assembly must be aliases of one department, got {_sibs}")

    # THE ROW SHAPE THE CANONICAL RUN EMITS, verbatim from 12120's JSON.
    _canonical = {
        "workbook_labour": {"schema": "workbook_labour_rows.v3", "mode": "canonical",
                            "rows": [{"wb_operation": "Assemble/pack (Metal)",
                                      "engine_operations": ["assembly"],
                                      "qty_per_unit": 1.0}]},
        "parts": [{"part_number": "P1", "textual_operations": ["handling"]}],
    }
    eq(_chk(_canonical), [],
       "a part whose handling is charged under the assembly/pack department is not "
       "'work nobody charged for' — the row and the part just name it differently")

    # AND THE CHECK STILL DOES ITS JOB. Without this the fixture would pass against a
    # version that reported nothing at all.
    _with_gap = dict(_canonical, parts=[{"part_number": "P1",
                                         "textual_operations": ["handling", "tapping"]}])
    _v = _chk(_with_gap)
    ok(_v, "an operation no department charged for is still reported")
    eq(sorted((_v[0].get("detail") or {}).get("operations") or {}), ["tapping"],
       "and only that one — the aliased pair must not be dragged back in")


def test_an_ambiguous_description_match_is_not_a_match():
    """The synthesised-code merge kept the FIRST canonical node with a given description.
    Where two canonical bought-ins share a description, the merge target then depended on
    dict order: the quantity would land against whichever was seen first, on the sheet, with
    nothing to say it had been a coin toss.

    An ambiguous match is not a match. The line stays separate and says so, because two
    similar items merged wrongly is a quantity against the wrong part — silent, and worse
    than an extra row an estimator can see and reconcile.
    """
    from wb_populate import canonicalise_part_estimates_for_workbook as _cz

    _amb = {"estimate_summary": {"canonical_route_shadow": {
        "decisions": [{"operation": "handling"}],
        "nodes": [
            {"part_number": "FIXING-A", "kind": "bought_in", "qty_per_unit": 2,
             "description": "M4 THREADED PEM STUD (LENGTH: 30mm)"},
            {"part_number": "FIXING-B", "kind": "bought_in", "qty_per_unit": 2,
             "description": "M4 THREADED PEM STUD (LENGTH: 30mm)"},
        ]}}}
    _pes = [{"part_number": "BI-PEMSTUD",
             "description": "M4 THREADED PEM STUD (LENGTH: 30mm)", "quantity": 2}]
    _codes = [p.get("part_number") for p in _cz(_amb, _pes)]
    ok("BI-PEMSTUD" in _codes,
       f"with two canonical candidates the synthesised line is NOT absorbed into one of "
       f"them at random, got {_codes}")

    # A UNIQUE match still merges — otherwise this fixture would pass against a rule that
    # simply stopped merging anything.
    _uniq = {"estimate_summary": {"canonical_route_shadow": {
        "decisions": [{"operation": "handling"}],
        "nodes": [{"part_number": "STD PART", "kind": "bought_in", "qty_per_unit": 2,
                   "description": "M4 THREADED PEM STUD (LENGTH: 30mm)"}]}}}
    _u = [p.get("part_number") for p in _cz(_uniq, list(_pes))]
    ok("BI-PEMSTUD" not in _u and "STD PART" in _u,
       f"one candidate still absorbs the duplicate, got {_u}")


def test_a_synthesised_code_is_never_absorbed_into_a_part_we_make():
    """The candidate index covered every canonical node, so a recogniser-minted code could
    merge into a fabricated LEAF whose description happened to match — putting a purchased
    item's identity and price onto a part we make.

    A BI- code is a bought-in by construction. Its only legitimate twin is a bought-in.
    """
    from wb_populate import canonicalise_part_estimates_for_workbook as _cz

    _summary = {"estimate_summary": {"canonical_route_shadow": {
        "decisions": [{"operation": "handling"}],
        "nodes": [{"part_number": "12120-01-01M", "kind": "leaf", "qty_per_unit": 1,
                   "description": "MOUNTING BRACKET"}]}}}
    _pes = [
        {"part_number": "12120-01-01M", "description": "MOUNTING BRACKET", "quantity": 1},
        {"part_number": "BI-MOUNTINGBRACKET", "description": "Mounting Bracket",
         "quantity": 1, "material_estimate": {"unit_material_cost_gbp": 8.40}},
    ]
    _codes = [p.get("part_number") for p in _cz(_summary, _pes)]
    ok("BI-MOUNTINGBRACKET" in _codes,
       f"a fabricated leaf must not absorb a bought-in code — that is a make/buy error "
       f"with a price attached, got {_codes}")

    # A CANONICAL BOUGHT-IN STILL ABSORBS IT, or the restriction has simply disabled the
    # dedup this exists to perform.
    _bi = {"estimate_summary": {"canonical_route_shadow": {
        "decisions": [{"operation": "handling"}],
        "nodes": [{"part_number": "STD PART", "kind": "bought_in", "qty_per_unit": 2,
                   "description": "M4 THREADED PEM STUD (LENGTH: 30mm)"}]}}}
    _p2 = [{"part_number": "STD PART", "description": "M4 THREADED PEM STUD (LENGTH: 30mm)",
            "quantity": 2},
           {"part_number": "BI-PEMSTUD",
            "description": "M4 THREADED PEM STUD (LENGTH: 30mm)", "quantity": 2}]
    _c2 = [p.get("part_number") for p in _cz(_bi, _p2)]
    ok("BI-PEMSTUD" not in _c2, f"a canonical bought-in still absorbs it, got {_c2}")


def test_operation_coverage_is_asked_per_part_not_job_wide():
    """Expanding department aliases across the whole job excused `handling` on EVERY part
    the moment any row charged the assembly department — including parts no assembly/pack
    row names.

    A part whose bench time genuinely never reached the sheet would then go unreported: an
    under-charge dressed as a clean check. A row names the parts it covers, so the question
    belongs there — is THIS part's operation charged on a row that includes THIS part.
    """
    from invariants import check_no_unpriced_operations_named as _chk

    _summary = {
        "workbook_labour": {"schema": "workbook_labour_rows.v3", "mode": "canonical",
                            "rows": [{"wb_operation": "Assemble/pack (Metal)",
                                      "engine_operations": ["assembly"],
                                      "part_numbers": ["P1"], "qty_per_unit": 1.0}]},
        "parts": [{"part_number": "P1", "textual_operations": ["handling"]},
                  {"part_number": "P2", "textual_operations": ["handling"]}],
    }
    _v = _chk(_summary)
    ok(_v, "P2's handling is charged on no row that names P2 — it must be reported")
    _ops = (_v[0].get("detail") or {}).get("operations") or {}
    eq(sorted(_ops), ["handling"], f"and named as handling, got {_ops}")
    eq(_ops.get("handling"), ["P2"],
       f"against P2 only — P1 IS covered by the row that names it, got {_ops}")


def test_a_route_group_quantity_is_not_each_targets_quantity():
    """2085's tube_cut names both tubes and states qty_per_unit 2 — two tubes per product.
    Splitting that into a decision per tube while copying the GROUP total onto each gave
    2085-02 x2 and 2085-03 x2: four cuts on two tubes, 18.25 batch hours and GBP 3.24 where
    the honest figure is 9.25 and GBP 1.64. Over half the labour on that job.

    Where a line resolves to ONE target its stated quantity is that target's. Where it
    splits across several, each takes its own multiplicity from the canonical BOM — the only
    place that knows how many of each part the product contains.
    """
    from route_compiler import compile_job_route

    _parts = [{"part_number": "2085-01", "description": "BRACKET PLATE", "quantity": 1},
              {"part_number": "2085-02", "description": "OUTER TUBE", "quantity": 1},
              {"part_number": "2085-03", "description": "INNER TUBE", "quantity": 1}]
    _extract = {
        "found": True,
        "bom": [{"part_number": p["part_number"], "description": p["description"], "qty": 1}
                for p in _parts],
        "assemblies": [{"part_number": "2085-GA",
                        "children": [{"part_number": p["part_number"], "qty": 1}
                                     for p in _parts]}],
        "routes": [
            {"sequence": 10, "operation": "tube_cut", "scope": "part", "qty_per_unit": 2,
             "part_numbers": ["2085-02", "2085-03"]},
            {"sequence": 40, "operation": "welding", "scope": "assembly", "qty_per_unit": 1,
             "part_numbers": ["2085-01", "2085-02", "2085-03"]},
        ]}
    _dec = [d for d in compile_job_route(_parts, _extract).get("decisions") or []
            if d.get("status") == "required"]
    _tubes = [d for d in _dec if d.get("operation") == "tube_cut"]

    eq(len(_tubes), 2, f"one cut decision per tube, got {len(_tubes)}")
    for _t in _tubes:
        eq(_t.get("qty_per_unit"), 1.0,
           f"{_t.get('target_id')} is one tube, not the group's two — got "
           f"{_t.get('qty_per_unit')}")
    eq(sum(d.get("qty_per_unit") or 0 for d in _tubes), 2.0,
       "and the group still totals two cuts, which is what the route said")

    # A SINGLE-TARGET LINE KEEPS ITS STATED QUANTITY. Without this the fixture would pass
    # against a rule that ignored route quantities altogether.
    _weld = [d for d in _dec if d.get("operation") == "welding"]
    eq(len(_weld), 1, "the weld is one event")
    eq(_weld[0].get("qty_per_unit"), 1.0, "and keeps the quantity the route stated")

    # THE SECOND DOOR. The same group total also reaches the compiler through the
    # compatibility adapter, as operation_qty_per_unit stamped onto every participant by the
    # route fold. Fixing only the route split left 2085's tubes doubled by the other path,
    # and the first fixture could not see it because it drove the route path alone.
    _adapter_parts = [
        {"part_number": "2085-01", "description": "BRACKET PLATE", "quantity": 1},
        {"part_number": "2085-02", "description": "OUTER TUBE", "quantity": 1,
         "textual_operations": ["tube_cut"], "operation_qty_per_unit": {"tube_cut": 2},
         "operation_scope": {"tube_cut": "part"}},
        {"part_number": "2085-03", "description": "INNER TUBE", "quantity": 1,
         "textual_operations": ["tube_cut"], "operation_qty_per_unit": {"tube_cut": 2},
         "operation_scope": {"tube_cut": "part"}},
    ]
    _adapter_extract = {
        "found": True,
        "bom": [{"part_number": p["part_number"], "description": p["description"], "qty": 1}
                for p in _adapter_parts],
        "assemblies": [{"part_number": "2085-GA",
                        "children": [{"part_number": p["part_number"], "qty": 1}
                                     for p in _adapter_parts]}],
        "routes": []}
    _ad = [d for d in compile_job_route(_adapter_parts, _adapter_extract).get("decisions") or []
           if d.get("operation") == "tube_cut" and d.get("status") == "required"]
    eq(len(_ad), 2, f"one cut decision per tube through the adapter, got {len(_ad)}")
    for _t in _ad:
        eq(_t.get("qty_per_unit"), 1.0,
           f"{_t.get('target_id')} via the adapter is one tube, not the group's two — got "
           f"{_t.get('qty_per_unit')}")

    # BOTH DOORS AT ONCE — THE PRODUCTION SHAPE, AND THE ONLY ONE THAT FAILS.
    #
    # 2085 has a route line AND part records carrying the same operation, so the event
    # receives two claims. Driving each path alone proves neither: on the live job the route
    # claim said 1 and the corroborating claim said 2, and because operation_sources now
    # attributes both to `inference` they arrive at EQUAL RANK. Equal rank plus disagreeing
    # metadata is a conflict, so arbitration refused to price it and the Tube row vanished
    # from the sheet entirely -- labour fell to GBP 4.11 with the cut missing rather than
    # doubled.
    #
    # Fail-closed did its job. The fix is to make the two claims agree, not to price a
    # conflict; this pins the agreement.
    def _tube_record(_pn, _desc):
        return {"part_number": _pn, "description": _desc, "quantity": 1,
                "textual_operations": ["tube_cut"],
                "operation_qty_per_unit": {"tube_cut": 2},
                "operation_scope": {"tube_cut": "part"},
                "operation_sources": {"tube_cut": "inference"}}

    _both_parts = [{"part_number": "2085-01", "description": "BRACKET PLATE", "quantity": 1},
                   _tube_record("2085-02", "OUTER TUBE"),
                   _tube_record("2085-03", "INNER TUBE")]
    _both_extract = {
        "found": True,
        "bom": [{"part_number": p["part_number"], "description": p["description"], "qty": 1}
                for p in _both_parts],
        "assemblies": [{"part_number": "2085-GA",
                        "children": [{"part_number": p["part_number"], "qty": 1}
                                     for p in _both_parts]}],
        "routes": [{"sequence": 10, "operation": "tube_cut", "scope": "part",
                    "qty_per_unit": 2, "inferred": True,
                    "part_numbers": ["2085-02", "2085-03"]}]}
    _bd = [d for d in compile_job_route(_both_parts, _both_extract).get("decisions") or []
           if d.get("operation") == "tube_cut"]
    eq(len(_bd), 2, f"one decision per tube, got {len(_bd)}")
    for _t in _bd:
        eq(_t.get("status"), "required",
           f"{_t.get('target_id')}: the two claims must AGREE, not conflict — an unverified "
           f"cut is dropped from the sheet and the tube is cut for free")
        eq(_t.get("qty_per_unit"), 1.0, f"{_t.get('target_id')} is one tube")
        _qs = {c.get("qty_per_unit") for c in (_t.get("claims") or [])}
        eq(_qs, {1.0},
           f"{_t.get('target_id')}: every claim on the event agrees on quantity, got {_qs}")
    eq(sum(d.get("qty_per_unit") or 0 for d in _bd), 2.0,
       "and the group still totals the two cuts the route stated")


def test_the_top_assembly_is_packed_whatever_joined_it():
    """Excluding every welded parent from assembly events is right for an INTERMEDIATE
    assembly — welding 12120-01-02M to -03M IS how 101 gets assembled, and a separate
    assemble event would charge that twice.

    It is wrong for the TOP assembly, which is the thing that ships. 2085-GA owns the weld,
    so it received no assembly event at all and the sheet carried no Assemble/pack row: a
    welded bracket nobody handles or packs. The `handling` warning was correct — there was
    genuinely nothing charging for it.
    """
    from route_compiler import compile_job_route

    _parts = [{"part_number": "2085-01", "description": "BRACKET PLATE", "quantity": 1},
              {"part_number": "2085-02", "description": "OUTER TUBE", "quantity": 1}]
    _extract = {
        "found": True,
        "bom": [{"part_number": p["part_number"], "description": p["description"], "qty": 1}
                for p in _parts],
        "assemblies": [{"part_number": "2085-GA",
                        "children": [{"part_number": p["part_number"], "qty": 1}
                                     for p in _parts]}],
        "routes": [{"sequence": 40, "operation": "welding", "scope": "assembly",
                    "qty_per_unit": 1, "part_numbers": ["2085-01", "2085-02"]}]}
    _dec = [d for d in compile_job_route(_parts, _extract).get("decisions") or []
            if d.get("status") == "required"]

    _asm = [d for d in _dec if d.get("operation") == "assembly"]
    eq(len(_asm), 1, f"the welded top assembly still gets ONE pack event, got {len(_asm)}")
    eq(_asm[0].get("target_id"), "2085-GA", "on the thing that ships")
    ok((_asm[0].get("sequence") or 0) > 40,
       f"and after the weld, not before it — got sequence {_asm[0].get('sequence')}")

    # THE WELD IS NOT DUPLICATED BY IT.
    eq(len([d for d in _dec if d.get("operation") == "welding"]), 1,
       "adding the pack event must not create a second weld")

    # AN INTERMEDIATE WELDED ASSEMBLY STILL GETS NOTHING. This is the case the exclusion
    # exists for: welding 12120-01-02M to -03M IS how 101 is assembled, so a pack event on
    # 101 would charge that work twice. Without this the fixture would pass against a rule
    # that packed every welded parent.
    _multi = [{"part_number": "12120-01-02M", "quantity": 1, "description": "UPSTAND"},
              {"part_number": "12120-01-03M", "quantity": 1, "description": "BASE PLATE"},
              {"part_number": "12120-01-05M", "quantity": 1, "description": "COVER PLATE"}]
    _mx = {
        "found": True,
        "bom": [{"part_number": p["part_number"], "description": p["description"], "qty": 1}
                for p in _multi],
        "assemblies": [
            {"part_number": "12120-01-101",
             "children": [{"part_number": "12120-01-02M", "qty": 1},
                          {"part_number": "12120-01-03M", "qty": 1}]},
            {"part_number": "12120-01-SA01",
             "children": [{"part_number": "12120-01-101", "qty": 1},
                          {"part_number": "12120-01-05M", "qty": 1}]},
        ],
        "routes": [{"sequence": 40, "operation": "welding", "scope": "assembly",
                    "qty_per_unit": 1, "target_id": "12120-01-101",
                    "part_numbers": ["12120-01-02M", "12120-01-03M"]}]}
    _md = [d for d in compile_job_route(_multi, _mx).get("decisions") or []
           if d.get("status") == "required" and d.get("operation") == "assembly"]
    _targets = sorted(str(d.get("target_id")) for d in _md)
    ok("12120-01-101" not in _targets,
       f"the welded INTERMEDIATE assembly is not packed as well as welded, got {_targets}")
    ok(any(t.endswith("SA01") for t in _targets),
       f"while the top assembly above it still is, got {_targets}")


def test_an_operation_the_compiler_rejected_is_not_reported_unpriced():
    """2085's tube records still carry laser_cutting — inherited from the shared assembly
    page the plate's route was read off — and the canonical route correctly rules it
    not_applicable, because a tube has no flat blank to profile.

    Reading the raw part field and reporting it as unpriced resurrects the very word the
    compiler rejected, and invites someone to put the laser row back. Decided-against is not
    the same as uncharged: only a required decision, or no decision at all, leaves an
    operation answerable to this check.
    """
    from invariants import check_no_unpriced_operations_named as _chk

    _summary = {
        "workbook_labour": {"schema": "workbook_labour_rows.v3", "mode": "canonical",
                            "rows": [{"wb_operation": "Tube",
                                      "engine_operations": ["tube_cut"],
                                      "part_numbers": ["2085-02"], "qty_per_unit": 1.0}]},
        "estimate_summary": {"canonical_route_shadow": {"decisions": [
            {"operation": "laser_cutting", "status": "not_applicable",
             "target_id": "2085-02", "participants": ["2085-02"]}]}},
        "parts": [{"part_number": "2085-02",
                   "textual_operations": ["tube_cut", "laser_cutting"]},
                  {"part_number": "2085-03", "textual_operations": ["tapping"]}],
    }
    _v = _chk(_summary)
    _ops = (_v[0].get("detail") or {}).get("operations") if _v else {}

    ok("laser_cutting" not in (_ops or {}),
       f"an operation the compiler ruled not_applicable is not 'work nobody charged for' — "
       f"got {_ops}")

    # laser_cutting ALONE proves nothing here. _TUBE_OP_REMAP sends laser to the Tube
    # department for tube stock, so the department inversion already treats the two as
    # siblings and the Tube row excuses it whatever the decision said. Both mutations of
    # the decided-against rule passed against that version of this fixture.
    #
    # `folding` is not a sibling of anything on this row, so it can only be excused by the
    # decision itself.
    _fold = {
        "workbook_labour": {"schema": "workbook_labour_rows.v3", "mode": "canonical",
                            "rows": [{"wb_operation": "Tube",
                                      "engine_operations": ["tube_cut"],
                                      "part_numbers": ["2085-01"], "qty_per_unit": 1.0}]},
        "estimate_summary": {"canonical_route_shadow": {"decisions": [
            {"operation": "folding", "status": "ruled_out", "target_id": "2085-01",
             "participants": ["2085-01"]}]}},
        "parts": [{"part_number": "2085-01", "textual_operations": ["folding"]}],
    }
    eq(_chk(_fold), [],
       "a fold the DXF ruled out is decided, not uncharged — reporting it invites putting "
       "the fold back on a part measured flat")

    # And with no decision it is reported, so the rule cannot simply excuse everything.
    _nodec = dict(_fold, estimate_summary={"canonical_route_shadow": {"decisions": []}})
    ok(_chk(_nodec), "with no decision at all, an uncharged fold is still reported")

    # A REQUIRED DECISION THAT REACHED NO ROW IS THE WHOLE POINT OF THE CHECK. Excusing
    # every decided operation regardless of status would hide work the compiler said must
    # happen and the sheet never charged — the exact silent under-charge this exists to
    # catch, and it broke no fixture until this case was written.
    _req = dict(_fold, estimate_summary={"canonical_route_shadow": {"decisions": [
        {"operation": "folding", "status": "required", "target_id": "2085-01",
         "participants": ["2085-01"]}]}})
    _rv = _chk(_req)
    ok(_rv, "a REQUIRED fold that no row charges must still be reported")
    eq(sorted((_rv[0].get("detail") or {}).get("operations") or {}), ["folding"],
       "naming the operation the compiler required and the sheet omitted")

    # AND THE CHECK STILL REPORTS WHAT NOBODY DECIDED. Without this the fixture would pass
    # against a version that excused everything.
    eq(sorted(_ops or {}), ["tapping"],
       f"an operation with no decision at all is still reported, got {_ops}")
    eq((_ops or {}).get("tapping"), ["2085-03"], "against the part that carries it")


def test_the_quote_header_is_the_logo_not_a_labelled_field():
    """"Prepared for" sat above the embedded customer logo and threw its alignment out on
    any logo whose aspect ratio differed from the one the offset was tuned against.

    A customer's own mark on a quotation needs no caption explaining whose it is. The label
    is gone and the block centres its content, so a tall logo and a wide one both sit level
    with the SDI mark instead of being pushed down by a fixed text offset.
    """
    from client_quote_html import build_quote_html

    _job = {"estimate_summary": {}, "invariants": {"may_quote_firm": True, "violations": []}}
    _html = build_quote_html(_job, job_stem="TEST-HDR")

    ok("Prepared for" not in _html,
       "the rendered quote must not carry the label — it is the thing that misaligns")

    # The block must still EXIST and still centre, or removing the label has simply left a
    # customer mark floating against the top of the header.
    ok('class="cust"' in _html, "the customer block is still rendered")
    _css = _html.split(".head .cust")[1][:200] if ".head .cust" in _html else ""
    ok("align-items:center" in _css.replace(" ", "").replace("align-items:center", "align-items:center"),
       f"and centres its content rather than relying on the removed label's offset, got {_css[:80]!r}")



# ── the two provenance sheets vs the route the sheet charges — job 12120 ─────────────
def _canonical_12120_summary() -> dict:
    """A job in the shape the canonical cutover produces: a compiled hierarchy, accepted
    workbook rows carrying their decision ids, and Excel's own calculated totals.

    `BI-KNOB` is the knurled knob: BOM qty 2 inside sub-assembly `101`, and `101` is used
    twice, so the job needs FOUR per unit and the workbook charges four.
    `12120-01-09M` reached no labour row at all — its drawing text still says powder and
    weld, and the priced route says neither.
    """
    return {
        "manufacturing_writeup": {"parts": [
            {"part_number": "12120-01-04M", "description": "Bracket", "quantity": 1,
             "normalized_material": "MILD_STEEL", "thicknesses_mm": [1.5],
             "geometry_source": "dxf_flat_pattern"},
            {"part_number": "BI-KNOB", "description": "Knurled knob", "quantity": 2,
             "normalized_material": "BOUGHT_IN", "page_roles": ["bought_in"]},
            {"part_number": "12120-01-09M", "description": "Blank plate", "quantity": 1,
             "normalized_material": "MILD_STEEL", "thicknesses_mm": [2.0],
             "textual_operations": ["powder_coating", "welding"]},
        ]},
        "workbook_labour": {
            "schema": "workbook_labour_rows.v3", "mode": "canonical", "rows": [
                {"workbook_row": 41, "wb_operation": "Laser",
                 "engine_operations": ["laser_cutting"], "qty_per_unit": 1.0,
                 "part_numbers": ["12120-01-04M"], "decision_ids": ["d0003-laser"]},
                {"workbook_row": 47, "wb_operation": "Weld (CO2)",
                 "engine_operations": ["welding"], "qty_per_unit": 1.0,
                 "part_numbers": ["12120-GA"], "decision_ids": ["d0011-weld"]},
            ]},
        "final_estimate": {
            "schema": "final_estimate.v2",
            "totals": {"material_gbp": 10.07, "labour_gbp": 31.37, "unit_gbp": 41.44},
            "labour_rows": [
                {"workbook_row": 41, "operation": "Laser", "total_value_gbp": 18.0},
                {"workbook_row": 47, "operation": "Weld (CO2)", "total_value_gbp": 22.0},
            ]},
        "estimate_summary": {
            "part_estimates": [
                {"part_number": "12120-01-04M", "unit_total_cost_gbp": 3.0,
                 "extended_total_cost_gbp": 3.0},
                {"part_number": "BI-KNOB", "unit_total_cost_gbp": 1.2,
                 "extended_total_cost_gbp": 4.8},
                {"part_number": "12120-01-09M", "unit_total_cost_gbp": 0.9,
                 "extended_total_cost_gbp": 0.9},
            ],
            "canonical_route_shadow": {"mode": "cutover", "nodes": [
                {"part_number": "12120-GA", "kind": "assembly", "qty_per_unit": 1.0,
                 "evidence": {}},
                {"part_number": "101", "kind": "assembly", "qty_per_unit": 2.0,
                 "evidence": {}},
                {"part_number": "BI-KNOB", "kind": "bought_in", "qty_per_unit": 4.0,
                 "evidence": {"raw_aliases": ["KNURLED KNOB"]}},
                {"part_number": "12120-01-04M", "kind": "leaf", "qty_per_unit": 1.0,
                 "evidence": {}},
                {"part_number": "12120-01-09M", "kind": "leaf", "qty_per_unit": 1.0,
                 "evidence": {}},
            ], "decisions": []},
        },
    }


def test_canonical_quantity_is_per_unit_not_per_parent():
    """A BOM row says how many the PARENT takes. The knurled knob is qty 2 in sub-assembly
    `101`, and `101` is used twice — the job needs four, and the workbook charges four.

    Every report read the row's own `quantity` and printed 2, so the provenance sheets
    disagreed with the Estimate tab in the same file for every part below the first level.
    wb_populate already rolls the multiplicity through, but on a LOCAL copy that is never
    stamped back, which is why fixing it there fixed nothing on the reports."""
    from costed_facts import canonical_quantity, canonical_identity

    _job = _canonical_12120_summary()
    eq(canonical_quantity(_job, "BI-KNOB"), 4.0,
       "the knob is needed four times per unit, not twice")
    eq(canonical_quantity(_job, "12120-01-04M"), 1.0, "a first-level leaf is unchanged")

    # A record can arrive under a spelling the graph merged away. Looking the number up
    # verbatim misses the node, and a miss is indistinguishable from having no graph.
    eq(canonical_identity(_job, "knurled knob"), "BI-KNOB",
       "a raw alias must resolve to the canonical identity")
    eq(canonical_quantity(_job, "KNURLED KNOB"), 4.0, "and carry the rolled quantity")

    # No graph -> None, so the caller keeps the row quantity rather than being handed a
    # defaulted 1 that looks like an answer.
    eq(canonical_quantity({"manufacturing_writeup": {"parts": []}}, "BI-KNOB"), None,
       "an unknown part must report no canonical quantity, not a default")


def test_decision_ids_survive_the_excel_readback():
    """The audit trail must not end at the workbook.

    `final_estimate` is preferred over `workbook_labour` once Excel has been read back —
    it is the only structure carrying calculated values. Its projection dropped
    `decision_ids`, so the compiler's decisions became unreachable from every downstream
    deliverable on exactly the runs where the route IS canonical."""
    from costed_facts import (decision_ids_for_part, priced_rows_for_part,
                              priced_route_known, _workbook_rows)

    _job = _canonical_12120_summary()
    ok(priced_route_known(_job), "a job with accepted rows has a known priced route")
    ok(any(r.get("_calculated") for r in (_workbook_rows(_job) or [])),
       "the calculated rows are the preferred source once the read-back has run")
    eq(decision_ids_for_part(_job, "12120-01-04M"), ["d0003-laser"],
       "the laser decision must still be reachable after the Excel join")
    eq([r["workbook_row"] for r in priced_rows_for_part(_job, "12120-01-04M")], [41],
       "and the sheet row that charges it")
    eq(decision_ids_for_part(_job, "12120-01-09M"), [],
       "a part on no labour row is priced by no decision")


def test_job_totals_prefer_what_excel_calculated():
    """Two calculators, and the reports were quietly showing the wrong one. The engine's
    per-part sum is a different arithmetic from the Estimate sheet's — on this job £8.70
    against £41.44 — and neither figure means anything unlabelled."""
    from costed_facts import job_totals

    _t = job_totals(_canonical_12120_summary())
    eq(_t["source"], "excel_calculated", "the workbook is authoritative once it has run")
    eq(_t["unit_gbp"], 41.44, "the sheet's unit price")
    eq(round(_t["engine_part_sum_gbp"], 2), 8.70, "and the engine sum, kept to reconcile")

    # Excel errors are carried as null by the read-back. Coercing one to 0.0 here would
    # turn missing data into a figure that reconciles.
    _err = _canonical_12120_summary()
    _err["final_estimate"]["totals"]["labour_gbp"] = None
    eq(job_totals(_err)["labour_gbp"], None, "a #DIV/0! stays missing, it does not become 0")

    # No workbook -> the engine sum is all there is, and it says so.
    _nowb = {"estimate_summary": {"part_estimates": [
        {"part_number": "X", "extended_total_cost_gbp": 5.0}]}}
    eq(job_totals(_nowb)["source"], "engine_part_sum", "with no workbook, the engine sum")


def test_a_part_on_no_labour_row_is_not_described_from_the_drawing():
    """`12120-01-09M` reached no labour row. Its drawing text still says POWDER COATED and
    carries the weld specification legend — the range-wide note that applies to the
    customer's whole product family, not to this job.

    Both provenance sheets fell back to those raw lists whenever the costed lookup came
    back empty, which is precisely the case where a gate had removed the operation. The
    report then narrated the route the workbook had just decided against, inside the same
    file as the Estimate tab that charges neither."""
    from job_decision_report import _ops_explanation
    from estimation_report import build_provenance

    _job = _canonical_12120_summary()
    _blank = _job["manufacturing_writeup"]["parts"][2]

    _txt = _ops_explanation(_blank, None, _job)
    ok("No operation charged" in _txt,
       f"the priced route is known and holds this part in no row, got {_txt!r}")
    ok("powder" not in _txt.lower() and "weld" not in _txt.lower(),
       f"and must not name the operations the gates removed, got {_txt!r}")

    _prov = {p["part_number"]: p for p in build_provenance(_job)}
    eq(_prov["12120-01-09M"]["operations"], "none charged",
       "the provenance sheet must say so too, not list the drawing's words")

    # AND THE FALLBACK MUST STILL WORK WHERE IT IS HONEST. With no workbook nothing has
    # been priced, so the drawing's reading is the best evidence there is — labelled as
    # unpriced rather than passed off as the route. Without this the fixture would pass
    # against a version that simply deleted the fallback.
    _nowb = dict(_job)
    _nowb.pop("workbook_labour", None)
    _nowb.pop("final_estimate", None)
    _txt2 = _ops_explanation(_blank, None, _nowb)
    ok("NOT YET PRICED" in _txt2, f"an unpriced job says so, got {_txt2!r}")
    ok("powder" in _txt2.lower(), f"and still reports what the drawing said, got {_txt2!r}")


def test_the_provenance_sheets_trace_every_part_to_a_sheet_row_and_a_decision():
    """The canonical cutover exists to make the route auditable, and neither sheet showed
    any of it: no decision id, no sheet row, and per-part money columns whose sum does not
    reconcile to the workbook total printed a few rows below them.

    Drives the real writers — a helper returning the right string proves nothing about what
    lands in a cell."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        _fail("openpyxl is required to verify what these sheets actually write; the "
              "workbook pipeline depends on it, so its absence is a broken environment, "
              "not a reason to report this fixture green")
        return
    from openpyxl import Workbook
    from job_decision_report import add_decision_report_sheet
    from estimation_report import add_provenance_sheet

    _job = _canonical_12120_summary()

    wb = Workbook()
    add_decision_report_sheet(wb, _job, {"job_number": "12120"})
    ws = wb["Decision Report"]
    _by_pn = {ws.cell(row=r, column=1).value: r for r in range(6, ws.max_row + 1)}

    eq(ws.cell(row=_by_pn["BI-KNOB"], column=3).value, 4,
       "the Qty column must show the per-unit quantity the sheet charges")

    _trace = str(ws.cell(row=_by_pn["12120-01-04M"], column=12).value or "")
    ok("41" in _trace, f"the Estimate row charging this part, got {_trace!r}")
    ok("d0003-laser" in _trace, f"and the decision behind it, got {_trace!r}")
    eq(str(ws.cell(row=_by_pn["12120-01-09M"], column=12).value or ""),
       "not priced on any labour row", "a part in no row says so plainly")

    # PER-PART MONEY IS MATERIAL. There is no per-part labour figure on a canonical job —
    # labour is a department row's batch value across every part in its setup — so a column
    # headed plainly "Unit £" was reporting an engine-era, labour-inclusive apportionment
    # that reconciles to nothing. On 2085 it read GBP 19.25 against each tube beside a Sell
    # Price of GBP 6.33.
    eq(str(ws.cell(row=5, column=9).value or ""), "Unit £ material",
       "the per-part money column must say what it is")
    eq(str(ws.cell(row=5, column=10).value or ""), "Ext £ material", "and so must its pair")

    _all = " ".join(str(c.value) for r in ws.iter_rows() for c in r if c.value)
    ok("RECONCILIATION" in _all, "what is on the page must be reconciled")
    ok("41.44" in _all, "naming what the Estimate sheet calculated")
    # THE ENGINE PART-SUM IS NOT QUOTED. It is a labour-inclusive figure that reconciles to
    # nothing, and repeating it in the explanation put it back on the page the material
    # basis had just removed it from — on 2085 that was GBP 44.75 beside a GBP 6.33 sheet.
    ok("8.70" not in _all,
       f"the obsolete engine part-sum must not be reintroduced by the explanation")
    ok("charged per department row" in _all,
       "and the reason there is no per-part labour figure must be stated")

    # ── and the same facts on the AI Provenance tab ────────────────────────────
    wb2 = Workbook()
    add_provenance_sheet(wb2, _job, {"job_number": "12120"})
    ws2 = wb2["AI Provenance"]
    _rows2 = {ws2.cell(row=r, column=1).value: r for r in range(6, ws2.max_row + 1)}
    eq(ws2.cell(row=_rows2["BI-KNOB"], column=3).value, 4,
       "the provenance sheet charges the same quantity as the Estimate tab")
    ok("d0011-weld" in " ".join(
        str(ws2.cell(row=r, column=15).value or "") for r in _rows2.values())
       or "d0003-laser" in " ".join(
           str(ws2.cell(row=r, column=15).value or "") for r in _rows2.values()),
       "the provenance sheet carries the decision trail")
    _all2 = " ".join(str(c.value) for r in ws2.iter_rows() for c in r if c.value)
    ok("RECONCILIATION" in _all2 and "41.44" in _all2,
       "and reconciles its own column against what Excel calculated")

# ── stock form vs the canonical route — jobs 1310 (bar) and the acrylic family ───────
def test_a_stock_form_gate_the_canonical_path_never_asks_is_not_a_gate():
    """The physical-impossibility rules lived inside wb_populate's legacy labour loop. The
    canonical cutover replaced that loop wholesale — `for pe in ([] if _canonical_cutover
    else labour_parts)` — so every one of them stopped running the moment the cutover was
    enabled, silently, because a gate nobody asks reports nothing.

    A solid 8mm round bar came back out of the canonical path carrying a Laser (Metal) row:
    the exact 1310-02 STUD defect (diameter read as sheet thickness) the wire rules were
    written for. Acrylic came back punched.

    Ruled out in the COMPILER, where route authority lives, as a not_applicable decision
    carrying its reason — not dropped in the renderer, where a vanished line is
    indistinguishable from one that was never read."""
    from route_compiler import compile_job_route
    import wb_populate as W

    _parts = [
        {"part_number": "1310-02", "description": "STUD 8mm DIA x 65",
         "normalized_material": "MILD_STEEL", "quantity": 1,
         "_bar_recognised": True, "wire_gauge_mm": 8.0, "wire_length_mm": 65.0},
        {"part_number": "ACR-01", "description": "Acrylic face panel",
         "normalized_material": "ACRYLIC", "quantity": 1},
        {"part_number": "MS-01", "description": "Mild steel bracket",
         "normalized_material": "MILD_STEEL", "quantity": 1},
    ]
    _extract = {
        "top_assembly": {"part_number": "GA"},
        "assemblies": [{"part_number": "GA", "children": [
            {"part_number": "1310-02", "qty": 1}, {"part_number": "ACR-01", "qty": 1},
            {"part_number": "MS-01", "qty": 1}]}],
        "routes": [
            {"operation": "laser_cutting", "part_numbers": ["1310-02"], "scope": "part"},
            {"operation": "punch", "part_numbers": ["ACR-01"], "scope": "part"},
            {"operation": "laser_cutting", "part_numbers": ["MS-01"], "scope": "part"},
        ]}
    _graph = compile_job_route(_parts, _extract)
    _status = {(d["operation"], d["target_id"]): d for d in _graph["decisions"]}

    eq(_status[("laser_cutting", "1310-02")]["status"], "not_applicable",
       "a laser cannot profile a solid round bar")
    ok("wire" in (_status[("laser_cutting", "1310-02")].get("reason") or ""),
       "and the decision must carry why, or an estimator cannot tell it from a lost line")
    eq(_status[("punch", "ACR-01")]["status"], "not_applicable",
       "a punch press shatters acrylic")

    # AND THE GATE MUST NOT SWALLOW REAL WORK. Without this the fixture would pass against
    # a version that ruled out every operation it was shown.
    eq(_status[("laser_cutting", "MS-01")]["status"], "required",
       "a mild steel bracket is still lasered")

    # END TO END: the decision has to reach the sheet, not just the graph. Asserting on the
    # compiler alone proved nothing about what gets charged — the whole defect was a rule
    # that was correct and never consulted.
    _summary = {"estimate_summary": {"canonical_route_shadow": {
        "mode": "cutover", "nodes": _graph["nodes"], "decisions": _graph["decisions"]}}}
    _pes = [
        {"part_number": "1310-02", "normalized_material": "MILD_STEEL", "quantity": 1,
         "material_estimate": {"stock_form": "wire", "material": "MILD_STEEL"}},
        {"part_number": "ACR-01", "normalized_material": "ACRYLIC", "quantity": 1,
         "normalized_thickness_mm": 5.0,
         "material_estimate": {"stock_form": "sheet", "material": "ACRYLIC"}},
        {"part_number": "MS-01", "normalized_material": "MILD_STEEL", "quantity": 1,
         "normalized_thickness_mm": 2.0,
         "material_estimate": {"stock_form": "sheet", "material": "MILD_STEEL"}},
    ]
    _rows = W.canonical_labour_groups(_summary, _pes, order_qty=180)
    _charged = {(g["wb_op"], tuple(g["parts"])) for g in _rows.values()}
    ok(not any(pn == ("1310-02",) for _op, pn in _charged),
       f"no row may charge the bar any flat-blank work, got {sorted(_charged)}")
    ok(not any(op == "Punch" for op, _pn in _charged),
       f"and none may punch the acrylic, got {sorted(_charged)}")
    ok(("Laser (Metal)", ("MS-01",)) in _charged,
       f"while the steel bracket is still lasered, got {sorted(_charged)}")


def test_the_impossibility_rules_have_one_home():
    """Two copies of these tables — one in the renderer, one in the compiler — is how one
    of them goes quietly stale while both look maintained. wb_populate's legacy entry point
    must resolve to the same rules the compiler decides from."""
    import wb_populate as W
    from stock_form_rules import (IMPOSSIBLE_OPS_BY_STOCK_FORM,
                                  is_impossible_operation)

    ok(W._SPURIOUS_OPS_BY_STOCK_FORM is IMPOSSIBLE_OPS_BY_STOCK_FORM,
       "the renderer must read the shared table, not a copy of it")
    eq(W._is_spurious_operation("laser_cutting", "wire", "MILD_STEEL"),
       is_impossible_operation("laser_cutting", "wire", "MILD_STEEL"),
       "and reach the same verdict as the compiler")

    # A TUBE IS STILL CUT. The expensive half of this distinction: an over-broad entry here
    # deleted 2085's tube cut entirely and costed both tubes as if they arrived at length.
    ok(not is_impossible_operation("laser_cutting", "tube", "MILD_STEEL"),
       "cutting a tube is redirected to the saw by the operation map, never deleted")
    ok(is_impossible_operation("punch", "tube", "MILD_STEEL"),
       "while a hole through a tube wall is not how it is cut to length")


def test_a_stated_finish_outranks_the_specification_legend():
    """The other half of the gates the cutover switched off, and the expensive half.

    These packs carry a range-wide specification legend — "POWDER COATED STEEL" — that
    applies to the customer's whole product family, not to this job. wb_populate's legacy
    loop dropped powder from a part whose own drawing states a different finish, and dropped
    diamond polish from a part that is powder coated. Under the canonical cutover a
    lacquered timber panel came back with a P.Coat row and a powder-coated steel face came
    back with a Diamond Polish row.

    The part's own title block wins over the legend. Ruled out in the compiler, carrying the
    stated finish as the reason."""
    from route_compiler import compile_job_route
    import wb_populate as W

    _parts = [
        {"part_number": "TMB-01", "description": "Birch ply panel",
         "normalized_material": "MDF", "normalized_finish": "LACQUERED", "quantity": 1},
        {"part_number": "MS-02", "description": "Steel face",
         "normalized_material": "MILD_STEEL",
         "normalized_finish": "POWDER COATED RAL9005", "quantity": 1},
        {"part_number": "ACR-02", "description": "Acrylic edge lit",
         "normalized_material": "ACRYLIC",
         "normalized_finish": "DIAMOND POLISHED EDGES", "quantity": 1},
        {"part_number": "PTR-01", "description": "Bracket",
         "normalized_material": "MILD_STEEL",
         "normalized_finish": "SEE ASSEMBLY", "quantity": 1},
    ]
    _extract = {
        "top_assembly": {"part_number": "GA"},
        "assemblies": [{"part_number": "GA", "children": [
            {"part_number": p["part_number"], "qty": 1} for p in _parts]}],
        "routes": [
            {"operation": "powder_coating", "part_numbers": ["TMB-01"], "scope": "part"},
            {"operation": "diamond_polish", "part_numbers": ["MS-02"], "scope": "part"},
            {"operation": "diamond_polish", "part_numbers": ["ACR-02"], "scope": "part"},
            {"operation": "powder_coating", "part_numbers": ["PTR-01"], "scope": "part"},
        ]}
    _d = {(x["operation"], x["target_id"]): x
          for x in compile_job_route(_parts, _extract)["decisions"]}

    eq(_d[("powder_coating", "TMB-01")]["status"], "not_applicable",
       "a panel the drawing says is LACQUERED does not go through the powder booth")
    ok("LACQUERED" in (_d[("powder_coating", "TMB-01")].get("reason") or ""),
       "and the decision names the finish it read")
    eq(_d[("diamond_polish", "MS-02")]["status"], "not_applicable",
       "a polished edge does not survive a powder coat")

    # THE RULE MUST NOT FIRE WHERE THE DRAWING SUPPORTS THE OPERATION, or it is not a gate,
    # it is a delete. Both of these would fail against an over-broad version.
    eq(_d[("diamond_polish", "ACR-02")]["status"], "required",
       "a part whose stated finish IS a diamond polish keeps it")
    # A finish that defers to another drawing states nothing about THIS part. Reading a
    # pointer as a non-powder finish would strip powder off every part in a pack that
    # specifies it once, on the GA.
    eq(_d[("powder_coating", "PTR-01")]["status"], "required",
       "'SEE ASSEMBLY' is a pointer, not a contradiction — it decides nothing here")

    # END TO END, because the defect was never in the rule — it was that nothing asked it.
    _summary = {"manufacturing_writeup": {"parts": _parts},
                "estimate_summary": {"canonical_route_shadow": {
                    "mode": "cutover",
                    "nodes": [{"part_number": p["part_number"], "kind": "leaf",
                               "qty_per_unit": 1.0, "evidence": {}} for p in _parts],
                    "decisions": list(_d.values())}}}
    _pes = [{"part_number": p["part_number"],
             "normalized_material": p["normalized_material"],
             "normalized_finish": p["normalized_finish"], "quantity": 1,
             "normalized_thickness_mm": 3.0,
             "material_estimate": {"stock_form": "sheet",
                                   "material": p["normalized_material"]}}
            for p in _parts]
    _charged = {(g["wb_op"], tuple(g["parts"]))
                for g in W.canonical_labour_groups(_summary, _pes, order_qty=100).values()}
    # Asserted per OPERATION, not per part: the top assembly's Assemble/pack row names every
    # part in the job and always will. Packing a lacquered panel is not coating it.
    ok(not any(op == "P.Coat" and "TMB-01" in pn for op, pn in _charged),
       f"no coating row may name the lacquered panel, got {sorted(_charged)}")
    ok(not any(op == "Diamond Polish" and "MS-02" in pn for op, pn in _charged),
       f"nor may a polish row name the powder-coated face, got {sorted(_charged)}")
    ok(("Diamond Polish", ("ACR-02",)) in _charged,
       f"while the acrylic keeps its polish, got {sorted(_charged)}")
    ok(("P.Coat", ("PTR-01",)) in _charged,
       f"and the pointer-finish bracket keeps its coat, got {sorted(_charged)}")


def test_a_finish_stated_only_in_surface_finishes_is_still_stated():
    """The legacy diamond-polish gate read normalized_finish alone. A part whose powder
    finish was captured in surface_finishes read as unfinished, so the gate saw no powder
    and left a diamond polish on a coat it will never survive. Both readers now resolve the
    stated finish the same way."""
    import wb_populate as W
    from finish_rules import stated_finish, finish_is_powder, finish_contradiction

    eq(stated_finish({"surface_finishes": ["Powder coated", "RAL 9005"]}),
       "POWDER COATED RAL 9005", "surface_finishes is a statement of finish")
    ok(finish_is_powder(stated_finish({"surface_finishes": ["Powder coated"]})),
       "and must be recognised as powder")
    # normalized_finish still wins when both are present.
    eq(stated_finish({"normalized_finish": "Lacquered", "surface_finishes": ["Powder"]}),
       "LACQUERED", "the normalised reading is preferred over the raw list")
    # A blank finish decides nothing — absence of a reading is not a reading.
    eq(finish_contradiction("powder_coating", ""), None,
       "an unstated finish must not rule anything out")
    ok(W._is_spurious_operation is not None, "the legacy entry point still exists")


# ── one job, one part list — across every deliverable ────────────────────────────────
def _diverging_job() -> dict:
    """A job whose three part lists genuinely disagree, in the three ways they really do.

      STD PART      absorbs the recogniser-minted BI-PEMSTUD (same item, two codes)
      STD PART      is qty 2 on its BOM row and 4 per unit through the hierarchy
      BI-KNOB       is an explicit canonical bought-in with no pricing record at all
    """
    _pes = [
        {"part_number": "12120-01-04M", "description": "Bracket", "quantity": 1,
         "normalized_material": "MILD_STEEL", "unit_total_cost_gbp": 3.0,
         "extended_total_cost_gbp": 3.0,
         # PRESENT AND EMPTY. A costing record carries these slots whether or not anything
         # was read into them, so an overlay that lets the estimate win unconditionally
         # wipes the drawing's real readings and empties every "WHY" column.
         "thicknesses_mm": [], "geometry_source": "", "dxf_source_file": None},
        {"part_number": "BI-PEMSTUD", "description": "PEM stud", "quantity": 2,
         "page_roles": ["bought_in"], "normalized_material": "BOUGHT_IN"},
        {"part_number": "STD PART", "description": "PEM stud", "quantity": 2,
         "page_roles": ["bought_in"], "normalized_material": "BOUGHT_IN"},
    ]
    return {
        "manufacturing_writeup": {"parts": [
            {"part_number": "12120-01-04M", "description": "Bracket", "quantity": 1,
             "normalized_material": "MILD_STEEL", "thicknesses_mm": [1.5],
             "geometry_source": "dxf_flat_pattern"},
            {"part_number": "BI-PEMSTUD", "description": "PEM stud", "quantity": 2,
             "page_roles": ["bought_in"]},
            {"part_number": "STD PART", "description": "PEM stud", "quantity": 2,
             "page_roles": ["bought_in"]},
        ]},
        "estimate_summary": {
            "part_estimates": _pes,
            "canonical_route_shadow": {"mode": "cutover", "decisions": [], "nodes": [
                {"part_number": "12120-01-04M", "kind": "leaf", "qty_per_unit": 1.0,
                 "evidence": {}},
                {"part_number": "STD PART", "kind": "bought_in", "qty_per_unit": 4.0,
                 "description": "PEM stud", "evidence": {"raw_aliases": ["BI-PEMSTUD"]}},
                {"part_number": "BI-KNOB", "kind": "bought_in", "qty_per_unit": 4.0,
                 "description": "Knurled knob", "evidence": {}},
            ]}},
    }


def test_every_deliverable_describes_the_same_part_list():
    """The Estimate sheet, both provenance tabs, the quote and the job report have to be
    describing ONE job. They were reading three different lists, and the sheet's own —
    the canonicalised one — was a local variable in populate_workbook, discarded the moment
    the workbook was saved.

    Canonicalising is not cosmetic. It merges recogniser-minted duplicates into the
    drawing's own code, rolls sub-assembly multiplicity into every quantity, and ADDS
    explicit bought-in BOM lines that never got a pricing record. So a bought-in the sheet
    charges appeared on no other page, a duplicate appeared on every page but the sheet, and
    quantities differed for anything below the first level. Totals were the visible end of
    a mismatch that ran through all of the content."""
    import wb_populate as W
    from costed_facts import job_parts

    _job = _diverging_job()
    _sheet_list = W.canonicalise_part_estimates_for_workbook(
        _job, _job["estimate_summary"]["part_estimates"])
    _sheet = {str(p["part_number"]).upper() for p in _sheet_list}

    # THE DIVERGENCE IS REAL, or this fixture proves nothing. The raw lists must differ
    # from the sheet's before we assert that everything now agrees with it.
    _raw = {str(p["part_number"]).upper()
            for p in _job["manufacturing_writeup"]["parts"]}
    ok(_raw != _sheet,
       f"the drawing-record list must differ from the sheet's, got {sorted(_raw)}")
    ok("BI-KNOB" in _sheet and "BI-KNOB" not in _raw,
       "a canonical bought-in with no pricing record is on the sheet and nowhere else")

    # populate_workbook publishes the sheet's list; from there every consumer reads it.
    _job["estimate_summary"]["canonical_part_estimates"] = _sheet_list

    eq({str(p["part_number"]).upper() for p in job_parts(_job)}, _sheet,
       "the shared accessor must resolve to exactly the sheet's rows")

    # ── every consumer, through its own entry point ────────────────────────────
    from job_report_html import _extract_parts
    from xlsx_output import _part_ests
    for _label, _fn in (("job report HTML", _extract_parts),
                        ("xlsx fallback writer", _part_ests)):
        eq({str(p["part_number"]).upper() for p in _fn(_job)}, _sheet,
           f"{_label} must name the same parts as the Estimate sheet")

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        _fail("openpyxl is required to verify the two provenance tabs")
        return
    from openpyxl import Workbook
    from job_decision_report import add_decision_report_sheet
    from estimation_report import add_provenance_sheet

    wb = Workbook()
    add_decision_report_sheet(wb, _job, {"job_number": "12120"})
    ws = wb["Decision Report"]
    _tab = {str(ws.cell(row=r, column=1).value).upper()
            for r in range(6, ws.max_row + 1)
            if ws.cell(row=r, column=1).value} & (_sheet | _raw)
    eq(_tab, _sheet, "the Decision Report must name the same parts as the sheet")

    wb2 = Workbook()
    add_provenance_sheet(wb2, _job, {"job_number": "12120"})
    ws2 = wb2["AI Provenance"]
    _tab2 = {str(ws2.cell(row=r, column=1).value).upper()
             for r in range(6, ws2.max_row + 1)
             if ws2.cell(row=r, column=1).value} & (_sheet | _raw)
    eq(_tab2, _sheet, "and so must AI Provenance")


def test_the_shared_part_list_keeps_the_evidence_the_tabs_are_built_from():
    """The canonical list is a COSTING record. The provenance columns — thickness source,
    geometry source, DXF filename, material source — are built from the drawing record.
    Switching the tabs to the canonical list without joining the two would have emptied
    every "WHY" column on the sheets whose entire purpose is to answer why."""
    from costed_facts import job_parts
    import wb_populate as W

    _job = _diverging_job()
    _job["estimate_summary"]["canonical_part_estimates"] = \
        W.canonicalise_part_estimates_for_workbook(
            _job, _job["estimate_summary"]["part_estimates"])
    _by_pn = {p["part_number"]: p for p in job_parts(_job)}

    eq(_by_pn["12120-01-04M"].get("thicknesses_mm"), [1.5],
       "the drawing record's evidence must survive the join")
    eq(_by_pn["12120-01-04M"].get("geometry_source"), "dxf_flat_pattern",
       "and so must its geometry provenance")
    # Identity and multiplicity are the canonical answer, whatever the drawing row said.
    eq(_by_pn["STD PART"].get("quantity"), 4.0,
       "quantity is the rolled per-unit figure, not the BOM row's")
    eq(_by_pn["12120-01-04M"].get("unit_total_cost_gbp"), 3.0,
       "and the costed fields come through")

    eq(_by_pn["12120-01-04M"].get("dxf_source_file"), None,
       "an absent reading stays absent — the join invents nothing")

    # A part with no drawing record at all still appears — it is on the sheet.
    ok("BI-KNOB" in _by_pn, "a canonical bought-in with no drawing record is still a row")

    # THE PUBLISHING STEP, checked structurally. Everything above proves the consumers read
    # the canonical list once it is on the summary; only populate_workbook can put it there,
    # and reaching that line needs the estimators' template and a live Excel. So this asserts
    # the link exists rather than leaving the whole chain resting on an unverified write —
    # the list was a discarded local for exactly as long as nothing looked.
    import inspect
    _src = inspect.getsource(W.populate_workbook)
    ok("canonical_part_estimates" in _src,
       "populate_workbook must publish the list it builds the sheet from")
    ok(_src.index("canonicalise_part_estimates_for_workbook")
       < _src.index("canonical_part_estimates"),
       "and publish it AFTER canonicalising, or it publishes the pre-canonical list")

    # WITH NO WORKBOOK there is no canonical list, and the engine's own is all there is.
    # Without this the accessor would return nothing on a job that never reached Excel.
    _nowb = _diverging_job()
    eq({p["part_number"] for p in job_parts(_nowb)},
       {"12120-01-04M", "BI-PEMSTUD", "STD PART"},
       "with no workbook built, the engine's part estimates are the list")


# ── review of 2529cce, and what job 2085 showed on the sheet ─────────────────────────
def test_lacquer_is_applied_by_the_spray_department():
    """The first finish rule asked whether the finish text contained the OPERATION's name —
    "does 'LACQUERED' contain 'spray'?" — and ruled wet_spray out when it did not.

    Lacquer IS applied through the wet-spray booth. That undercosted the exact timber route
    the gate was written to protect: the panel keeps its finish on the drawing and loses the
    department that applies it. A finish phrase and a department name are different
    vocabularies and cannot be compared by substring."""
    from finish_rules import finish_contradiction as fc, finish_families

    for _finish in ("LACQUERED", "PAINTED", "VARNISHED", "WET SPRAY RAL9005"):
        eq(fc("wet_spray", _finish), None,
           f"{_finish} goes through the spray booth — wet_spray must survive it")
        eq(finish_families(_finish), {"wet_spray"}, f"{_finish} is a wet-spray finish")

    # ...and the contradictions it exists for still fire.
    ok(fc("powder_coating", "LACQUERED"), "a lacquered panel is not powder coated")
    ok(fc("diamond_polish", "POWDER COATED RAL9005"),
       "a polished edge does not survive a powder coat")
    ok(fc("wet_spray", "POWDER COATED"), "and powder is not sprayed")
    ok(fc("powder_coating", "RAW"), "an explicitly bare part is not coated")

    # AN UNRECOGNISED PHRASE DECIDES NOTHING. Ruling work out on a finish we could not read
    # is how a gate becomes a delete.
    eq(fc("powder_coating", "AS DRAWING REV C"), None,
       "an unread finish is not a contradiction")
    eq(fc("powder_coating", "SEE ASSEMBLY"), None, "nor is a pointer")
    eq(fc("powder_coating", ""), None, "nor is silence")


def test_a_finish_token_inside_a_longer_word_is_not_that_finish():
    """Two ways to get this wrong, and the first version had one of each.

    Plain substring matching read "AS DRAWING REV C" as a BARE finish — "d-RAW-ing"
    contains RAW — which would strip powder from any part whose finish field points at the
    drawing. Anchoring both ends then broke the other way: drawings write LACQUERED,
    PAINTED, ANODISED, and \bLACQUER\b matches none of them.

    The boundary belongs before the stem and nowhere else."""
    from finish_rules import finish_families

    eq(finish_families("AS DRAWING REV C"), set(),
       "'drawing' does not contain a RAW finish")
    eq(finish_families("LACQUERED"), {"wet_spray"}, "but the -ED form is still lacquer")
    eq(finish_families("ANODISED CLEAR"), {"anodise"}, "and -ISED is still anodising")
    eq(finish_families("POWDER-COATED"), {"powder"}, "a hyphen is a word boundary")
    eq(finish_families("SELF  COLOUR"), {"bare"}, "and so is any run of whitespace")
    eq(finish_families("UNPAINTED"), set(),
       "a stem buried mid-word is not that stem — conservative is the safe direction")


def test_the_gates_read_the_merged_record_not_just_the_raw_one():
    """build_part_graph already reconciles the two extraction paths: `records` is the
    extracted BOM/parts record overlaid with every non-empty raw value. The gates read
    `raw`, so a stock form or finish carried only by the extract bypassed them entirely —
    the rule correct, and simply never shown the evidence."""
    from route_compiler import compile_job_route

    # Nothing on the raw record says bar or lacquer. Both facts arrive via the extract.
    _parts = [{"part_number": "BAR-01", "description": "Stud", "quantity": 1},
              {"part_number": "TMB-02", "description": "Ply panel", "quantity": 1}]
    _extract = {
        "top_assembly": {"part_number": "GA"},
        "assemblies": [{"part_number": "GA", "children": [
            {"part_number": "BAR-01", "qty": 1}, {"part_number": "TMB-02", "qty": 1}]}],
        "parts": [
            {"part_number": "BAR-01", "_bar_recognised": True, "wire_gauge_mm": 8.0},
            {"part_number": "TMB-02", "normalized_finish": "LACQUERED"},
        ],
        "routes": [
            {"operation": "laser_cutting", "part_numbers": ["BAR-01"], "scope": "part"},
            {"operation": "powder_coating", "part_numbers": ["TMB-02"], "scope": "part"},
        ]}
    _d = {(x["operation"], x["target_id"]): x
          for x in compile_job_route(_parts, _extract)["decisions"]}
    eq(_d[("laser_cutting", "BAR-01")]["status"], "not_applicable",
       "a bar identified only by the extract is still a bar")
    eq(_d[("powder_coating", "TMB-02")]["status"], "not_applicable",
       "a finish stated only in the extract still contradicts")


def test_regenerating_a_report_sheet_replaces_it_rather_than_duplicating_it():
    """openpyxl's create_sheet does not overwrite: asked for a name already in the book it
    silently returns "AI Provenance1". The estimators' template can already carry these
    tabs, so a re-run produced BOTH — the stale sheet keeping the name an estimator opens,
    and the fresh one hidden behind a suffix. 2085 shipped with both."""
    try:
        from openpyxl import Workbook
    except ImportError:
        _fail("openpyxl is required to verify what these sheets write")
        return
    from job_decision_report import add_decision_report_sheet
    from estimation_report import add_provenance_sheet

    _job = {"manufacturing_writeup": {"parts": [
                {"part_number": "P1", "description": "Panel", "quantity": 1}]},
            "estimate_summary": {"part_estimates": [
                {"part_number": "P1", "unit_total_cost_gbp": 1.0,
                 "extended_total_cost_gbp": 1.0}]}}

    wb = Workbook()
    wb.create_sheet("AI Provenance")      # already in the template
    wb.create_sheet("Decision Report")
    for _ in range(2):                    # and re-run twice over
        add_decision_report_sheet(wb, _job, {})
        add_provenance_sheet(wb, _job, {})

    eq([n for n in wb.sheetnames if n.startswith("AI Provenance")], ["AI Provenance"],
       f"exactly one provenance tab, got {wb.sheetnames}")
    eq([n for n in wb.sheetnames if n.startswith("Decision Report")], ["Decision Report"],
       f"exactly one decision report, got {wb.sheetnames}")
    # And it is the GENERATED one that survives, not the template's empty placeholder.
    ok(wb["AI Provenance"].max_row > 3, "the surviving provenance tab is the written one")
    ok(wb["Decision Report"].max_row > 3, "and so is the surviving decision report")


def test_the_decision_report_names_every_operation_the_sheet_charges():
    """Job 2085, on the sheet: the Decision Report described the tubes as "powder coating,
    welding" while AI Provenance — two tabs away in the same file, from the same source —
    listed tube_cut, welding, dress_welds, powder_coating and assembly.

    The explanation was built from a whitelist of eight phrases and anything outside it
    vanished. The tube cut is exactly the operation this engine has already lost once, and
    a provenance sheet that omits a charged operation is evidence for a route we did not
    price."""
    from job_decision_report import _ops_explanation

    _job = {
        "workbook_labour": {"schema": "workbook_labour_rows.v3", "mode": "canonical",
                            "rows": [
            {"workbook_row": 97, "wb_operation": "Tube",
             "engine_operations": ["tube_cut"], "part_numbers": ["2085-02"]},
            {"workbook_row": 98, "wb_operation": "Weld (CO2)",
             "engine_operations": ["welding"], "part_numbers": ["2085-02"]},
            {"workbook_row": 99, "wb_operation": "Dress Welds",
             "engine_operations": ["dress_welds"], "part_numbers": ["2085-02"]},
            {"workbook_row": 100, "wb_operation": "P.Coat",
             "engine_operations": ["powder_coating"], "part_numbers": ["2085-02"]},
            {"workbook_row": 101, "wb_operation": "Assemble/pack (Metal)",
             "engine_operations": ["assembly"], "part_numbers": ["2085-02"]},
        ]}}
    _txt = _ops_explanation({"part_number": "2085-02", "description": "OUTER TUBE"},
                            None, _job)
    for _needle, _what in (("tube cutting", "the sawn-to-length cut"),
                           ("weld dressing", "the dressing of the weld"),
                           ("welding", "the weld"),
                           ("powder coating", "the coat"),
                           ("assembly", "the assemble/pack")):
        ok(_needle in _txt.lower(), f"{_what} is charged and must be named, got {_txt!r}")

    # An operation with no written phrase is still named, plainly, rather than dropped.
    _odd = {"workbook_labour": {"schema": "workbook_labour_rows.v3", "mode": "canonical",
                                "rows": [{"workbook_row": 5, "wb_operation": "Wrap",
                                          "engine_operations": ["shrink_wrapping"],
                                          "part_numbers": ["X1"]}]}}
    ok("shrink wrapping" in _ops_explanation({"part_number": "X1"}, None, _odd).lower(),
       "an unrecognised operation is named plainly, never omitted")


def test_the_report_sheets_are_written_after_excel_has_calculated():
    """On 2085 the Ext GBP column summed the engine's per-part figures to GBP 44.75 against
    a Sell Price of GBP 6.33, with the money columns unlabelled and no reconciliation line
    anywhere on the page.

    The code for both was correct and inert: main.py wrote the two tabs BEFORE the Excel
    read-back, so at the moment they asked whether the workbook had calculated, the honest
    answer was no. And the read-back stamps the JSON on disk, not the in-memory summary the
    tabs are written from, so ordering alone was not enough."""
    import inspect
    import main as _main

    _src = inspect.getsource(_main)
    _readback = _src.index("stamp_real_totals_into_json")
    _tabs = _src.index("add_decision_report_sheet(_wb")
    ok(_readback < _tabs,
       "the report sheets must be written after the read-back, or they have nothing to "
       "reconcile against")
    ok(_src.index('summary["final_estimate"] = _stamped') < _tabs,
       "and the calculated totals must be merged into the in-memory summary first — the "
       "read-back stamps the file on disk, not the object the tabs are rendered from")

    # The behaviour that depends on it, driven directly: with the calculated totals present
    # the money columns declare their basis and the page reconciles.
    from costed_facts import job_totals
    _with = job_totals({"estimate_summary": {"part_estimates": [
                            {"part_number": "A", "extended_total_cost_gbp": 44.75}]},
                        "final_estimate": {"totals": {"unit_gbp": 6.33,
                                                      "material_gbp": 0.13,
                                                      "labour_gbp": 5.75}}})
    eq(_with["source"], "excel_calculated", "Excel is authoritative once it has run")
    eq(_with["unit_gbp"], 6.33, "and its figure is the one shown")
    eq(round(_with["engine_part_sum_gbp"], 2), 44.75,
       "with the engine sum kept so the gap is stated rather than left to be noticed")


# ── job 2085 as it actually printed — the reporting layer regression ─────────────────
def _job_2085_as_printed() -> dict:
    """The 2085 sheet's own shape: a plate and two tubes, one Tube row grouping BOTH tubes,
    and engine per-part totals of GBP 19.25 a tube on a job the workbook priced at GBP 6.33.

    The tubes' material genuinely could not be read (native models unread), so their real
    material cost is zero — which is the honest number and the one that flags for review."""
    _rows = [
        {"workbook_row": 96, "wb_operation": "Laser", "engine_operations": ["laser_cutting"],
         "part_numbers": ["2085-01"], "decision_ids": ["d-laser-01"], "qty_per_unit": 1.0},
        {"workbook_row": 97, "wb_operation": "Tube", "engine_operations": ["tube_cut"],
         "part_numbers": ["2085-02", "2085-03"],
         "decision_ids": ["d-tubecut-02", "d-tubecut-03"], "qty_per_unit": 2.0},
        {"workbook_row": 98, "wb_operation": "Weld (CO2)", "engine_operations": ["welding"],
         "part_numbers": ["2085-01", "2085-02", "2085-03"],
         "decision_ids": ["d-weld"], "qty_per_unit": 1.0},
        {"workbook_row": 99, "wb_operation": "Dress Welds",
         "engine_operations": ["dress_welds"],
         "part_numbers": ["2085-01", "2085-02", "2085-03"],
         "decision_ids": ["d-dress"], "qty_per_unit": 1.0},
        {"workbook_row": 100, "wb_operation": "P.Coat",
         "engine_operations": ["powder_coating"],
         "part_numbers": ["2085-01", "2085-02", "2085-03"],
         "decision_ids": ["d-pcoat"], "qty_per_unit": 1.0},
        {"workbook_row": 101, "wb_operation": "Assemble/pack (Metal)",
         "engine_operations": ["assembly"],
         "part_numbers": ["2085-01", "2085-02", "2085-03"],
         "decision_ids": ["d-pack"], "qty_per_unit": 1.0},
    ]
    def _dec(_id, _op, _target, _parts, _scope):
        return {"decision_id": _id, "operation": _op, "status": "required",
                "target_id": _target, "participants": _parts, "scope": _scope,
                "qty_per_unit": 1.0}
    _parts_est = [
        {"part_number": "2085-01", "description": "BRACKET PLATE", "quantity": 1,
         "normalized_material": "MILD_STEEL", "normalized_thickness_mm": 1.2,
         "unit_total_cost_gbp": 6.25, "extended_total_cost_gbp": 6.25,
         "material_estimate": {"stock_form": "sheet", "material": "MILD_STEEL",
                               "unit_material_cost_gbp": 0.13}},
        {"part_number": "2085-02", "description": "OUTER TUBE", "quantity": 1,
         "normalized_material": "MILD_STEEL",
         "unit_total_cost_gbp": 19.25, "extended_total_cost_gbp": 19.25,
         "material_estimate": {"stock_form": "tube", "material": "MILD_STEEL"}},
        {"part_number": "2085-03", "description": "INNER TUBE", "quantity": 1,
         "normalized_material": "MILD_STEEL",
         "unit_total_cost_gbp": 19.25, "extended_total_cost_gbp": 19.25,
         "material_estimate": {"stock_form": "tube", "material": "MILD_STEEL"}},
        {"part_number": "PACKAGING", "description":
         "Packaging (box / pallet — per-unit share, estimator to price)", "quantity": 1,
         "page_roles": ["bought_in"], "material_estimate": {},
         "_price_explicitly_withheld": True},
        {"part_number": "DELIVERY", "description":
         "Delivery (per-unit share of order haulage — estimator to price)", "quantity": 1,
         "page_roles": ["bought_in"], "material_estimate": {},
         "_price_explicitly_withheld": True},
    ]
    return {
        "manufacturing_writeup": {"parts": [
            {"part_number": p["part_number"], "description": p["description"],
             "quantity": 1, "normalized_material": p.get("normalized_material"),
             "page_roles": p.get("page_roles") or []} for p in _parts_est]},
        "workbook_labour": {"schema": "workbook_labour_rows.v3", "mode": "canonical",
                            "rows": _rows},
        "final_estimate": {
            "schema": "final_estimate.v2",
            "totals": {"material_gbp": 0.13, "labour_gbp": 5.75, "unit_gbp": 6.33},
            "labour_rows": [{"workbook_row": r["workbook_row"],
                             "operation": r["wb_operation"], "total_value_gbp": 1.0}
                            for r in _rows]},
        "estimate_summary": {
            "part_estimates": _parts_est,
            "canonical_part_estimates": _parts_est,
            "canonical_route_shadow": {"mode": "cutover", "decisions": [
                _dec("d-laser-01", "laser_cutting", "2085-01", ["2085-01"], "part"),
                _dec("d-tubecut-02", "tube_cut", "2085-02", ["2085-02"], "part"),
                _dec("d-tubecut-03", "tube_cut", "2085-03", ["2085-03"], "part"),
                _dec("d-weld", "welding", "2085-GA",
                     ["2085-01", "2085-02", "2085-03"], "assembly"),
                _dec("d-dress", "dress_welds", "2085-GA",
                     ["2085-01", "2085-02", "2085-03"], "assembly"),
                _dec("d-pcoat", "powder_coating", "2085-GA",
                     ["2085-01", "2085-02", "2085-03"], "assembly"),
                _dec("d-pack", "assembly", "2085-GA",
                     ["2085-01", "2085-02", "2085-03"], "assembly"),
            ], "nodes": [
                {"part_number": p["part_number"], "kind": "leaf", "qty_per_unit": 1.0,
                 "evidence": {}} for p in _parts_est]}},
    }


def test_a_grouped_workbook_row_does_not_give_every_part_every_decision():
    """A workbook row is a tooling SETUP and can group several parts. 2085's two tubes share
    one Tube row carrying both tube-cut decisions, so each tube was told it was cut by the
    other tube's decision as well as its own — the mis-join the canonical route exists to
    prevent, reappearing in the sheet that documents it.

    An ASSEMBLY-scoped decision legitimately belongs to every participant: the weld is one
    event across three parts. So this narrows part events without hiding shared ones."""
    from costed_facts import decision_ids_for_part

    _job = _job_2085_as_printed()
    eq(decision_ids_for_part(_job, "2085-02"),
       ["d-tubecut-02", "d-weld", "d-dress", "d-pcoat", "d-pack"],
       "the outer tube is cut by its OWN decision only")
    eq(decision_ids_for_part(_job, "2085-03"),
       ["d-tubecut-03", "d-weld", "d-dress", "d-pcoat", "d-pack"],
       "and the inner tube by its own")
    ok("d-tubecut-03" not in decision_ids_for_part(_job, "2085-02"),
       "neither tube may carry the other's cut")
    # The shared assembly events must survive, or the fixture passes against a version that
    # simply drops everything it cannot match one-to-one.
    ok("d-weld" in decision_ids_for_part(_job, "2085-01"),
       "one weld event across three parts belongs to all three")


def test_no_labour_inclusive_engine_total_is_shown_as_a_per_part_cost():
    """2085 printed GBP 19.25 against each tube, on a sheet whose whole unit price is
    GBP 6.33, and summed GBP 44.75 into a "cost breakdown by material type" against
    GBP 0.13 of material the workbook actually calculated.

    unit_total_cost_gbp is an engine-era, LABOUR-INCLUSIVE apportionment. There is no
    per-part labour figure on a canonical job — labour is a department row's batch value
    across every part in its setup — so the per-part column is material, which IS costed
    per part and sums to the workbook's own material total."""
    try:
        from openpyxl import Workbook
    except ImportError:
        _fail("openpyxl is required to verify what these sheets write")
        return
    from job_decision_report import add_decision_report_sheet
    from estimation_report import add_provenance_sheet, build_provenance

    _job = _job_2085_as_printed()
    wb = Workbook()
    add_decision_report_sheet(wb, _job, {"job_number": "2085"})
    ws = wb["Decision Report"]
    _all = " ".join(str(c.value) for r in ws.iter_rows() for c in r if c.value)

    ok("19.25" not in _all,
       "the labour-inclusive engine total must not appear anywhere on the sheet")
    ok("44.75" not in _all, "nor its sum in the material breakdown")

    _by_pn = {ws.cell(row=r, column=1).value: r for r in range(6, ws.max_row + 1)}
    eq(ws.cell(row=_by_pn["2085-01"], column=9).value, 0.13,
       "the plate shows the material it was costed at")
    # The tubes' material genuinely could not be read. Zero is the honest number and the
    # one that flags — GBP 19.25 hid a missing reading behind a plausible figure.
    _tube = ws.cell(row=_by_pn["2085-02"], column=9).value
    ok(_tube in (0, 0.0, "—"), f"an unread tube material reads as nothing, got {_tube!r}")

    # The breakdown reconciles to the workbook's material total, and says so.
    ok("MATERIAL COST BREAKDOWN" in _all and "0.13" in _all,
       "the material breakdown must be material, against the sheet's own material total")

    # PACKAGING/DELIVERY are estimator placeholders, not catalogue prices.
    _pack = str(ws.cell(row=_by_pn["PACKAGING"], column=6).value or "")
    ok("NOT YET PRICED" in _pack,
       f"an unpriced placeholder must not claim a catalogue price, got {_pack!r}")
    ok("catalogue/history" not in _pack, "and must not cite a provenance it does not have")

    # AI Provenance shows the same per-part money on the same basis.
    _prov = {p["part_number"]: p for p in build_provenance(_job)}
    eq(_prov["2085-02"]["unit_cost"], 0.0, "the provenance tab agrees about the tube")
    eq(_prov["2085-01"]["unit_cost"], 0.13, "and about the plate")
    wb2 = Workbook()
    add_provenance_sheet(wb2, _job, {"job_number": "2085"})
    ok("19.25" not in " ".join(str(c.value) for r in wb2["AI Provenance"].iter_rows()
                               for c in r if c.value),
       "and never prints the engine total either")


def test_both_2085_tabs_describe_the_same_route():
    """The Decision Report said "powder coating, welding" for the tubes while AI Provenance,
    in the same file, listed tube_cut, welding, dress_welds, powder_coating and assembly.
    Two tabs, one job, two routes. They must agree operation for operation."""
    from estimation_report import build_provenance
    from job_decision_report import _ops_explanation
    from costed_facts import operations_for_part

    _job = _job_2085_as_printed()
    _prov = {p["part_number"]: p for p in build_provenance(_job)}
    for _pn in ("2085-01", "2085-02", "2085-03"):
        _canon = operations_for_part(_job, _pn)
        _dr = _ops_explanation({"part_number": _pn}, None, _job).lower()
        for _op in _canon:
            _words = _op.replace("_", " ")
            ok(_words.split()[0] in _dr,
               f"{_pn}: the Decision Report omits {_op}, which the sheet charges — got {_dr!r}")
            ok(_op in _prov[_pn]["operations"],
               f"{_pn}: AI Provenance omits {_op}")
    ok("tube cutting" in _ops_explanation({"part_number": "2085-02"}, None, _job).lower(),
       "the tube cut in particular — this engine has lost it once already")


# ── the Estimate tab must not look finished when it is not — job 2085 ────────────────
def test_an_unpriced_line_says_what_is_missing_and_what_would_fill_it():
    """2085's Estimate tab showed Unit Cost GBP 6.33 and Sell Price GBP 6.33, and both read
    as a completed price. Neither tube contributed any material, the powder rate is a
    known-wrong assumption, packaging and delivery were zero, and the margin was 0%. Every
    one of those was stated — in a console flag that scrolls past, or as a "GBP -" that
    reads like a real zero.

    A blank that looks like a zero is worse than an error. An error stops someone; a
    plausible number does not."""
    from estimator_inputs import (input_note_for_line, material_input_note,
                                  section_summary, banner_text)

    _outer = {"part_number": "2085-02", "description": "OUTER TUBE",
              "section_stock": {"a": 12.7, "t": 1.2, "profile_form": "CHS",
                                "length_mm": 34.0}}
    _inner = {"part_number": "2085-03", "description": "INNER TUBE",
              "section_stock": {"a": 10.0, "t": 1.2, "profile_form": "CHS"}}

    eq(section_summary(_outer), "12.7 x 1.2 CHS",
       "the section is named as far as we read it, so it can be priced without the drawing")
    _note = material_input_note(_outer)
    ok("MATERIAL UNPRICED" in _note, f"the line must say it is unpriced, got {_note!r}")
    ok("34mm" in _note, f"and name the cut length we read, got {_note!r}")
    ok("section rate" in _note, f"and name the figure that is missing, got {_note!r}")

    # A length we did NOT read must be asked for, never defaulted.
    _note2 = material_input_note(_inner)
    ok("NOT READ" in _note2, f"an unread cut length is asked for, got {_note2!r}")
    ok("34" not in _note2, "and never borrowed from the other tube")

    # A PLACEHOLDER IS A DIFFERENT JOB FROM AN UNPRICED PART. One is a commercial figure to
    # decide, the other a rate to look up, and a checklist that conflates them gets ignored.
    _pack = {"part_number": "PACKAGING", "_price_explicitly_withheld": True,
             "description": "Packaging (box / pallet — per-unit share, estimator to price)"}
    eq(input_note_for_line(_pack)["kind"], "placeholder_unpriced",
       "a withheld placeholder is not an unpriced material line")
    eq(input_note_for_line(_outer)["kind"], "material_unpriced", "and vice versa")

    # The banner counts, because "PROVISIONAL" alone says nothing about how much is missing.
    ok("3 ESTIMATOR INPUTS REQUIRED" in banner_text([{}, {}, {}]), "the count is the point")
    eq(banner_text([]), "", "and a complete sheet carries no banner at all")


def test_the_provisional_banner_never_overwrites_a_total():
    """The banner sits beside Unit Cost and Sell Price — the two figures that read as a
    finished price. It must land PAST the value, not on it: the gap between a label and its
    figure differs per row, and a fixed offset put the banner exactly on the Total Unit Cost
    cell. The write was refused, so the formula was safe either way, but a refused banner is
    no banner and the sheet went out looking finished again."""
    try:
        from openpyxl import Workbook
    except ImportError:
        _fail("openpyxl is required to verify what is written to the Estimate tab")
        return
    import wb_populate as W

    wb = Workbook()
    ws = wb.active
    ws["I200"] = "Total Unit Cost Price"; ws["L200"] = 6.33
    ws["I206"] = "Sell Price";            ws["L206"] = 6.33
    _inputs = [{"kind": "material_unpriced", "part": "2085-02", "where": "BOM row 12",
                "what": "MATERIAL UNPRICED: 12.7 x 1.2 CHS, cut length 34mm"},
               {"kind": "placeholder_unpriced", "part": "PACKAGING", "where": "BOM row 14",
                "what": "NOT YET PRICED: enter the per-unit figure for this line"}]
    _flags = []
    W._write_estimator_inputs(ws, _inputs, _flags)

    eq((ws["L200"].value, ws["L206"].value), (6.33, 6.33),
       "the sheet's own figures must be untouched — a blanked total breaks the read-back "
       "and every deliverable downstream of it")
    for _r, _what in ((200, "Unit Cost"), (206, "Sell Price")):
        _banners = [c.coordinate for c in ws[_r]
                    if c.value and "PROVISIONAL" in str(c.value)]
        eq(len(_banners), 1,
           f"{_what} carries exactly one provisional banner, got {_banners}")
        ok("2 ESTIMATOR INPUTS" in str(ws[_banners[0]].value),
           f"{_what}'s banner names the count")

    # The checklist itself, below the totals, with somewhere to work from.
    _col_c = [str(ws.cell(row=r, column=3).value or "") for r in range(1, ws.max_row + 1)]
    _all = " ".join(_col_c)
    ok("OUTSTANDING ESTIMATOR INPUTS (2)" in _all, f"the checklist is written and counted")
    ok("2085-02" in _all and "PACKAGING" in _all, "and names every outstanding line")
    ok(any("ESTIMATOR INPUTS: 2 outstanding" in f for f in _flags),
       "and the run says where it put them")

    # A COMPLETE SHEET GETS NOTHING. Without this the fixture would pass against a version
    # that banners every job, which would make the banner mean nothing within a week.
    wb2 = Workbook(); ws2 = wb2.active
    ws2["I200"] = "Sell Price"; ws2["L200"] = 6.33
    W._write_estimator_inputs(ws2, [], [])
    ok(not any(c.value and "PROVISIONAL" in str(c.value)
               for r in ws2.iter_rows() for c in r),
       "a sheet with no outstanding inputs carries no banner")


def test_an_input_cell_is_shaded_so_it_reads_as_a_cell_to_type_in():
    """Deliberately a fill, not a comment. Comments are hidden until hovered, and the whole
    failure was that the missing input was discoverable rather than visible."""
    try:
        from openpyxl import Workbook
    except ImportError:
        _fail("openpyxl is required to verify cell formatting")
        return
    import wb_populate as W

    ws = Workbook().active
    W._mark_input_cell(ws, 12, 10)
    _c = ws.cell(row=12, column=10)
    ok(_c.fill is not None and "FFF2CC" in str(_c.fill.fgColor.rgb or ""),
       f"the cell an estimator types into is shaded, got {_c.fill.fgColor.rgb!r}")
    ok(_c.border is not None and _c.border.left is not None
       and _c.border.left.style == "medium", "and edged, so it reads as an input")


def test_the_two_provenance_writers_do_not_fight_over_one_sheet_name():
    """wb_populate writes bought-in PRICE provenance; estimation_report writes the
    whole-estimate provenance report. Both asked for "AI Provenance". openpyxl does not
    overwrite, so 2085 shipped with "AI Provenance" and "AI Provenance1" — the duplicate
    tab, which was never a template problem at all. Two sheets, two purposes, two names."""
    import inspect
    import wb_populate as W

    _src = inspect.getsource(W._append_ai_sheets)
    ok('_add("AI Price Provenance"' in _src,
       "the price-provenance sheet must not claim the report's name")
    ok('_add("AI Provenance"' not in _src,
       "and must not leave the collision behind under a different call")


# ── closing the reporting layer — the last three ─────────────────────────────────────
def test_the_provenance_headline_prefers_the_workbook_not_the_part_column():
    """Its fallback summed the part column and called the result "Est. Total / engine
    part-sum". Once that column became material-only the label was wrong twice over: it is
    not the engine part-sum, and it is not a job total. Driven with a workbook unit cost of
    £6.33 and £0.10 of part material, the sheet printed "Est. Total: £0.10".

    The Decision Report already had the right hierarchy — Sell Price cell, else the
    workbook's calculated unit, and only then anything engine-side. Fixing one and not the
    other is how the two tabs disagreed in the first place."""
    try:
        from openpyxl import Workbook
    except ImportError:
        _fail("openpyxl is required to verify what these sheets write")
        return
    from estimation_report import add_provenance_sheet
    from job_decision_report import add_decision_report_sheet

    _job = _job_2085_as_printed()
    # No "Sell Price" cell to reference — this is the fallback path, and the only one
    # where either sheet has to choose a number for itself.
    for _fn, _tab in ((add_provenance_sheet, "AI Provenance"),
                      (add_decision_report_sheet, "Decision Report")):
        _wb = Workbook()
        _fn(_wb, _job, {"job_number": "2085"})
        _text = " ".join(str(c.value) for r in _wb[_tab].iter_rows()
                         for c in r if c.value)
        ok("6.33" in _text,
           f"{_tab} must fall back to what the Estimate sheet calculated")
        ok("0.13" not in _text.split("RECONCILIATION")[0].split("MATERIAL COST")[0]
           or "Unit cost" in _text,
           f"{_tab} must not present the material column as a job total")
        ok("Est. Total" not in _text,
           f"{_tab} must not label a workbook figure as an engine estimate")

    # AND WITH NO WORKBOOK AT ALL it says so, rather than dressing part material as a total.
    _nowb = _job_2085_as_printed()
    _nowb.pop("final_estimate", None)
    _wb2 = Workbook()
    add_provenance_sheet(_wb2, _nowb, {})
    _t2 = " ".join(str(c.value) for r in _wb2["AI Provenance"].iter_rows()
                   for c in r if c.value)
    ok("no workbook total" in _t2.lower(),
       f"with nothing calculated, the sheet says what the figure actually is")


def test_the_canonical_part_list_reaches_the_saved_json():
    """populate_workbook stamps the canonical list onto the IN-MEMORY summary, and the two
    workbook tabs are written from that — so they were right, and the fix looked done.

    The quote and the job report are generated from the SAVED JSON (generate_report takes a
    path, not an object), which was written before the workbook existed and never received
    it. Both silently fell back to raw part_estimates: merged duplicates back, rolled
    quantities gone, and the bought-in BOM lines the sheet charges missing again — the exact
    divergence the shared list exists to close, live in the two customer-facing files.

    Structural, because reaching that write needs the estimators' template and a live
    Excel; everything downstream of it is covered behaviourally elsewhere."""
    import inspect
    import main as _main

    _src = inspect.getsource(_main)
    ok('["canonical_part_estimates"]) = _canon_pes' in _src,
       "the saved JSON must receive the canonical part list")
    # It has to land in the block that WRITES the file, and before that write.
    _stamp = _src.index('["canonical_part_estimates"]) = _canon_pes')
    _write = _src.index("json.dump(_doc, _fh_w")
    ok(_stamp < _write, "and be stamped before the document is written back")

    # The consumer side, driven for real: with the list absent the accessor falls back, and
    # with it present every deliverable resolves to the sheet's rows.
    from costed_facts import job_parts
    _job = _diverging_job()
    eq({p["part_number"] for p in job_parts(_job)},
       {"12120-01-04M", "BI-PEMSTUD", "STD PART"},
       "without the list, the reports show the raw engine set — the defect")
    import wb_populate as W
    _job["estimate_summary"]["canonical_part_estimates"] = \
        W.canonicalise_part_estimates_for_workbook(
            _job, _job["estimate_summary"]["part_estimates"])
    eq({p["part_number"] for p in job_parts(_job)},
       {"12120-01-04M", "STD PART", "BI-KNOB"},
       "with it, they show the sheet's")


def test_the_material_breakdown_adds_back_to_the_sheets_material_total():
    """The workbook's material total also carries lines belonging to no single part — the
    powder consumable, and the per-line scrap the sheet adds. On 2085 the part column is
    about £0.09 against a workbook material total of £0.1334, so a breakdown of only the
    parts is short, and silently so."""
    try:
        from openpyxl import Workbook
    except ImportError:
        _fail("openpyxl is required to verify what this sheet writes")
        return
    from job_decision_report import add_decision_report_sheet

    _job = _job_2085_as_printed()
    for _bucket in ("canonical_part_estimates", "part_estimates"):
        _job["estimate_summary"][_bucket][0]["material_estimate"][
            "unit_material_cost_gbp"] = 0.09

    _wb = Workbook()
    add_decision_report_sheet(_wb, _job, {"job_number": "2085"})
    _ws = _wb["Decision Report"]
    _rows = {}
    for _r in range(1, _ws.max_row + 1):
        _label = _ws.cell(row=_r, column=1).value
        if _label:
            _rows[str(_label)] = _ws.cell(row=_r, column=9).value

    eq(_rows.get("MILD_STEEL"), 0.09, "the parts account for what they account for")
    eq(_rows.get("Powder / scrap / other workbook material"), 0.04,
       "and the rest is named rather than left as a silent shortfall")
    eq(_rows.get("TOTAL MATERIAL (agrees with the Estimate sheet)"), 0.13,
       "so the section adds back to the workbook's own material total")

    # NO RESIDUAL ROW WHEN THERE IS NO RESIDUAL. Without this the fixture would pass against
    # a version that always invents a balancing line, which would hide a real shortfall by
    # construction.
    _exact = _job_2085_as_printed()
    _wb2 = Workbook()
    add_decision_report_sheet(_wb2, _exact, {"job_number": "2085"})
    _t2 = " ".join(str(c.value) for r in _wb2["Decision Report"].iter_rows()
                   for c in r if c.value)
    ok("Powder / scrap" not in _t2,
       "a breakdown that already reconciles gets no balancing row")


# ── one confidence authority, field by field — job 2085 ──────────────────────────────
def _2085_with_rate_bases() -> dict:
    _job = _job_2085_as_printed()
    for _r in _job["workbook_labour"]["rows"]:
        _r["rate_basis"] = {
            "Laser": "template_calculated",
            "Tube": "unmeasured_default",
            "Dress Welds": "unmeasured_default",
            "P.Coat": "historical_unbanded",
            "Assemble/pack (Metal)": "historical_unbanded",
        }.get(_r["wb_operation"], "historical")
    _job["manufacturing_writeup"]["parts"][0].update({
        "material_source": "llm_full_extract",
        "geometry_source": "dxf_flat_pattern",
        "dxf_source_file": "2085-01 - Bracket Plate_1.2mm MS.DXF",
        "thicknesses_mm": [1.2],
        "_learning_flag": "ZERO_COST_STEEL — review material/thickness",
    })
    return _job


def test_the_weakest_field_decides_the_line_and_nothing_is_averaged():
    """A scalar confidence has to average things that are not commensurable, and averaging
    is how a job hides its own gap. 2085's tubes have NO material at all; a mean of
    "material known, thickness unknown, geometry unknown, route known" lands in the sixties
    and reads as partial knowledge rather than an absent input."""
    from confidence import assess_part, UNKNOWN, ASSUMED, MEASURED, NOT_APPLICABLE

    _job = _2085_with_rate_bases()
    _tube = dict(_job["estimate_summary"]["canonical_part_estimates"][1])
    _a = assess_part(_tube, _job)

    eq(_a["overall"], UNKNOWN,
       "a part with no material price is UNKNOWN, not a percentage")
    _by = {f["field"]: f["status"] for f in _a["fields"]}
    eq(_by["material price"], UNKNOWN, "the tube material was never read")
    eq(_by["thickness"], UNKNOWN, "nor its wall")
    # THE ROUTE IS KNOWN, and must still say so. A single figure could not carry both
    # facts; the point of the field list is that the gap is located, not just totalled.
    eq(_by["route"], MEASURED, "while the route IS known — five operations charged")
    ok("material price" in _a["decided_by"],
       f"and the report names WHICH field decided it, got {_a['decided_by']}")


def test_a_bought_in_placeholder_is_not_a_confident_line():
    """PACKAGING and DELIVERY scored 1.00 for being bought-in, so two lines carrying no
    price at all were counted among the job's HIGH-confidence parts. Being purchased rather
    than made says nothing about whether anyone has priced it.

    But a bought-in must be judged on the fields it HAS: thickness and geometry are
    NOT_APPLICABLE and cannot drag it down, or "no fabrication thickness" becomes a defect
    on a box of packaging."""
    from confidence import assess_part, UNKNOWN, NOT_APPLICABLE, counts_by_status

    _job = _2085_with_rate_bases()
    _pack = dict(_job["estimate_summary"]["canonical_part_estimates"][3])
    _a = assess_part(_pack, _job)

    eq(_a["overall"], UNKNOWN, "an unpriced placeholder is not a confident line")
    _by = {f["field"]: f["status"] for f in _a["fields"]}
    eq(_by["material price"], UNKNOWN, "because it carries no price")
    eq(_by["thickness"], NOT_APPLICABLE, "not because it has no thickness")
    eq(_by["geometry"], NOT_APPLICABLE, "or no geometry")
    ok("thickness" not in _a["decided_by"] and "geometry" not in _a["decided_by"],
       f"a field that does not exist cannot decide the line, got {_a['decided_by']}")

    # And a PRICED bought-in is not dragged down by the same N/A fields.
    _priced_bi = {"part_number": "BI-KNOB", "description": "Knurled knob", "quantity": 4,
                  "page_roles": ["bought_in"],
                  "material_estimate": {"unit_material_cost_gbp": 0.42,
                                        "price_source": {"source_name": "bought_in price book"}}}
    ok(assess_part(_priced_bi, _job)["overall"] != UNKNOWN,
       "a bought-in that HAS a price is not punished for having no thickness")

    _counts = counts_by_status([assess_part(dict(p), _job)
                                for p in _job["estimate_summary"]["canonical_part_estimates"]])
    eq(_counts.get("measured", 0) + _counts.get("confirmed", 0), 0,
       f"no part on 2085 is high-confidence — the sheet said 3, got {_counts}")


def test_material_confidence_is_not_raised_by_the_geometry_source():
    """AI Provenance scored MATERIAL at 95% because "dxf_flat" appeared in
    `geometry_source`, while printing the label from `material_source`. The score and the
    label described different fields, so the 95% was never a statement about the material.
    A DXF flat pattern says nothing about alloy."""
    from confidence import assess_part, ASSUMED, MEASURED, REPORTED

    _base = {"part_number": "X1", "normalized_material": "MILD_STEEL",
             "geometry_source": "dxf_flat_pattern"}
    _mat = {f["field"]: f for f in assess_part(_base, None)["fields"]}["material identity"]
    ok(_mat["status"] in (ASSUMED, REPORTED),
       f"geometry must not raise material confidence, got {_mat['status']}")

    # A material source that IS a material reading does raise it.
    _named = dict(_base, material_source="dxf_filename:PART_1.2mm_MS.DXF")
    _mat2 = {f["field"]: f for f in assess_part(_named, None)["fields"]}["material identity"]
    eq(_mat2["status"], MEASURED, "a material code read from the DXF filename is measured")


def test_both_tabs_report_the_same_status_for_the_same_part():
    """2085-01 came out 0.92 HIGH on the Decision Report and 0.70 MEDIUM on AI Provenance —
    same part, same data, same file, two private confidence ladders. There was a third in
    calibration.py that the reporting pipeline never called."""
    from confidence import assess_part
    from job_decision_report import _conf_info
    from estimation_report import build_provenance

    from costed_facts import job_parts

    _job = _2085_with_rate_bases()
    _prov = {p["part_number"]: p for p in build_provenance(_job)}
    # THE SAME RECORD BOTH TABS ACTUALLY ITERATE. Both read job_parts — the canonical
    # estimate overlaid on the drawing record — and assessing the raw drawing record
    # instead reports a different status for a reason that has nothing to do with the
    # confidence model. Comparing the wrong record is how you "prove" a disagreement that
    # production does not have, and miss one it does.
    for _p in job_parts(_job):
        _pn = _p["part_number"]
        _shared = assess_part(_p, _job)
        _, _label, _, _, _ = _conf_info(_p, 0.0, _job)
        eq(_label, _shared["overall_label"],
           f"{_pn}: the Decision Report must report the shared status")
        eq(_prov[_pn]["overall_status"], _shared["overall"],
           f"{_pn}: and AI Provenance the same one")
    eq(sorted(_prov), sorted(p["part_number"] for p in job_parts(_job)),
       "and both tabs must be describing the same set of lines in the first place")


def test_a_labour_rate_says_how_it_was_arrived_at():
    """Tube 40/hr is an UNMEASURED constant, P.Coat 458/hr a historical observation used
    UNBANDED because no part area was computed, and the laser rate the estimators' OWN
    calculator read off the template. Three very different claims, and the audit tabs
    presented all three identically — the only record of the difference was a "# UNMEASURED"
    code comment."""
    import wb_populate as W
    from confidence import assess_part, ASSUMED, MEASURED, REPORTED

    ok("Tube" in W._THROUGHPUT_UNMEASURED, "the tube rate is not measured, and says so")
    ok("Weld (CO2)" not in W._THROUGHPUT_UNMEASURED,
       "while the weld rate was derived from the corpus")

    _job = _2085_with_rate_bases()
    _tube = dict(_job["estimate_summary"]["canonical_part_estimates"][1])
    _rate = {f["field"]: f
             for f in assess_part(_tube, _job)["fields"]}["labour rate"]
    eq(_rate["status"], ASSUMED,
       "a row priced on an unmeasured constant is an assumption, not a measurement")
    ok("unmeasured" in _rate["source"], f"and names it, got {_rate['source']!r}")

    # A job whose rates are all template-calculated or banded is not downgraded.
    _clean = _2085_with_rate_bases()
    for _r in _clean["workbook_labour"]["rows"]:
        _r["rate_basis"] = "template_calculated"
    _rate2 = {f["field"]: f for f in assess_part(
        dict(_clean["estimate_summary"]["canonical_part_estimates"][1]),
        _clean)["fields"]}["labour rate"]
    eq(_rate2["status"], MEASURED,
       "the estimators' own calculator is the strongest basis on the sheet")


def test_a_pre_cost_flag_the_finished_job_contradicts_is_superseded():
    """ZERO_COST_STEEL fires when a MILD_STEEL part with a DXF still costs nothing. It runs
    BEFORE the workbook and nothing revisited it, so 2085-01 carried "review
    material/thickness" onto the audit tab while the sheet two tabs away charged it £0.13
    of material and a laser row. A flag the finished job contradicts sends an estimator to
    check something already answered."""
    from costed_facts import reconcile_risk_flags

    _job = _2085_with_rate_bases()
    _plate = _job["manufacturing_writeup"]["parts"][0]
    ok("ZERO_COST_STEEL" in str(_plate.get("_learning_flag")), "the flag starts present")

    reconcile_risk_flags(_job)
    ok("ZERO_COST_STEEL" not in str(_plate.get("_learning_flag") or ""),
       "a part the workbook prices no longer carries a zero-cost review flag")
    ok(any(f.get("flag") == "ZERO_COST_STEEL"
           for f in (_plate.get("superseded_risk_flags") or [])),
       "superseded, not deleted — the cue was real when it fired")

    # A PART THAT REALLY IS UNPRICED KEEPS IT. Without this the rule would excuse every
    # zero-cost steel part on the job, which is the flag's whole purpose.
    _job2 = _2085_with_rate_bases()
    _job2["manufacturing_writeup"]["parts"].append({
        "part_number": "2085-99", "description": "UNCOSTED PLATE",
        "normalized_material": "MILD_STEEL", "material_estimate": {},
        "_learning_flag": "ZERO_COST_STEEL — review material/thickness"})
    reconcile_risk_flags(_job2)
    ok("ZERO_COST_STEEL" in str(
        _job2["manufacturing_writeup"]["parts"][-1].get("_learning_flag") or ""),
       "a genuinely uncosted steel part keeps the flag")


def test_there_is_exactly_one_confidence_authority():
    """Three implementations is how the two tabs came to disagree. calibration.py carries a
    fourth scalar that the reporting pipeline never called — dead, and therefore available
    to be revived by anyone who finds it and does not know why it is dead."""
    import inspect
    import calibration, estimation_report, job_decision_report

    ok("DEPRECATED FOR CONFIDENCE" in
       inspect.getdoc(calibration.line_confidence_and_provisos or (lambda: None)) or "",
       "the retired scalar must say so, at the top, where someone about to call it looks")

    # NEITHER REPORT MAY SCORE CONFIDENCE ITSELF. Read the source rather than the behaviour:
    # a private ladder can agree with the shared one today and drift tomorrow, which is
    # exactly what happened before.
    for _mod, _name in ((job_decision_report, "Decision Report"),
                        (estimation_report, "AI Provenance")):
        _src = inspect.getsource(_mod)
        ok("from confidence import" in _src,
           f"{_name} must read the shared authority")
        ok("line_confidence_and_provisos" not in _src,
           f"{_name} must not call the retired scalar")
    ok("mat_conf" not in inspect.getsource(estimation_report),
       "AI Provenance must not keep its own material score")


def test_the_quote_names_the_drawing_and_the_unit_not_the_folder():
    """The headline was the job stem with a leading number stripped off. For the folder
    "2085 - SolidWorks" that leaves "SolidWorks", so a customer's quotation was headed
    "Quotation — SolidWorks". For a combined pack it leaves the whole filename,
    "...-GA2_REV[E]-combined.pdf". Neither names the thing being quoted, and the drawing
    number appeared only as "Job <n>" in small text beneath.

    The drawing states both facts in its own title block and the extract already
    transcribes them."""
    import re as _re
    from client_quote_html import build_quote_html, _drawing_identity

    def _render(stem, extract):
        return build_quote_html(
            {"estimate_summary": {}, "invariants": {"may_quote_firm": True, "violations": []},
             "llm_full_extract": extract}, job_stem=stem)

    # 1. Title block present — it wins over everything.
    _h = _render("2085 - SolidWorks",
                 {"drawing_info": {"drawing_number": "2085", "title": "BRACKET ASSEMBLY"}})
    eq(_re.search(r"<h1>(.*?)</h1>", _h, _re.S).group(1).strip(), "BRACKET ASSEMBLY",
       "the headline names the unit")
    ok("Drawing <b>2085</b>" in _h, "and the drawing number is stated as a drawing number")
    ok("SolidWorks" not in _h, "the folder's own name is not a product")

    # 2. A combined pack: the number and revision come out of the filename soup.
    _n, _rev, _desc = _drawing_identity(
        {"llm_full_extract": {"drawing_info": {
            "drawing_number": "11772-01-09", "revision": "E", "title": "WIDE ARCH"}}},
        "0357299_2 WIDE ARCH_11772-01-09-GA2_REV[E]-combined.pdf")
    eq((_n, _rev, _desc), ("11772-01-09", "Rev E", "WIDE ARCH"),
       "a combined-pack filename must not reach the customer")

    # 3. NOTHING STATED — fall back to the number, never to noise. "SolidWorks",
    # "combined" and "REV[E]" are folder debris, and inventing a product name from them is
    # worse than admitting we only know the drawing number.
    # A JOB NUMBER IS DIGITS, NOT ANY RUN OF ALPHANUMERICS. [0-9A-Za-z_-]* is greedy and
    # every character of a spaceless folder name is in it, so "11350-BootsLadderRackCommsBar"
    # came out whole as the drawing number — the first live job to be run through this.
    _n4, _r4, _d4 = _drawing_identity({}, "11350-BootsLadderRackCommsBar")
    eq(_n4, "11350", "the number stops at the first word")
    eq(_d4, "Boots Ladder Rack Comms Bar",
       "and a spaceless folder name is made readable rather than shipped as typed")
    # Multi-part numbers must still survive whole.
    eq(_drawing_identity({}, "11772-01-09 GA2")[0], "11772-01-09",
       "an all-digit continuation is part of the number")
    eq(_drawing_identity({}, "4471B-WideArch")[0], "4471B",
       "and so is a single letter suffix")

    _n2, _r2, _d2 = _drawing_identity({}, "2085 - SolidWorks")
    eq(_n2, "2085", "the drawing number still comes through")
    eq(_d2, "2085", "and the description falls back to it rather than to 'SolidWorks'")
    _h2 = _render("2085 - SolidWorks", {})
    ok("Manufactured to drawing 2085." in _h2,
       "and the lead does not repeat the number either side of a dash")

    # 4. The canonical top assembly is consulted before the folder name.
    _n3, _r3, _d3 = _drawing_identity(
        {"estimate_summary": {"canonical_route_shadow": {
            "top_assembly": "11772-GA",
            "nodes": [{"part_number": "11772-GA", "description": "Wide arch bay"}]}}},
        "somefolder - combined.pdf")
    eq((_n3, _d3), ("11772-GA", "Wide arch bay"),
       "the drawing's own hierarchy beats the folder every time")


# ── job 11350 (Boots Comms Bar) — a length in the filename is not a gauge ────────────
def test_a_product_dimension_in_a_dxf_filename_is_not_the_sheet_gauge():
    """"Boots Comms Bar - Left Arm 200mm_flat.dxf" returned a thickness of 200.0 — a
    200mm-thick mild steel arm — because the filename carries the PRODUCT LENGTH and the
    parser took the first <n>mm token with no plausibility bound.

    Same class as the 1310-02 STUD, where a bar's diameter was read as a sheet thickness: a
    number with a unit attached is not automatically the number being looked for. A wrong
    gauge is far worse than an absent one — the DXF geometry and the drawing both still
    carry a thickness, and nothing else can catch a plausible-looking 200."""
    from pathlib import Path
    from drawing_job_merge import (thickness_mm_from_dxf_filename as _thk,
                                   SHEET_GAUGE_MIN_MM, SHEET_GAUGE_MAX_MM)

    eq(_thk(Path("Boots Comms Bar - Left Arm 200mm_flat.dxf")), None,
       "a 200mm product length is not a gauge, and no reading beats a wrong one")
    eq(_thk(Path("11350-01-01M_1MM MS.DXF")), 1.0, "a real gauge still reads")

    # THE FIRST PLAUSIBLE TOKEN WINS, not the first token. Both numbers are in the name and
    # the length comes first.
    eq(_thk(Path("Left Arm 200mm_1mm MS.dxf")), 1.0,
       "a length before the gauge must not shadow it")
    eq(_thk(Path("Comms Bar 758mm_1mm MS.dxf")), 1.0, "nor a 758mm bar length")

    # The decimal conventions this parser already handles must keep working.
    eq(_thk(Path("PART_1,2mm.dxf")), 1.2, "European comma decimal")
    eq(_thk(Path("PART_1_5mm.dxf")), 1.5, "underscore decimal")
    eq(_thk(Path("panel 18mm MDF.dxf")), 18.0, "18mm board is a real gauge")
    ok(SHEET_GAUGE_MIN_MM < 1.0 < SHEET_GAUGE_MAX_MM, "the bounds admit ordinary sheet")


def test_solidworks_mirror_naming_is_recognised():
    """SolidWorks names a mirrored part "Mirror<partnumber>" with nothing between the two,
    so the trailing \\b in the pattern never matched: "Mirror11350-01-02M" is one unbroken
    run of word characters. Job 11350 ships exactly that file."""
    import re as _re
    import config as _cfg

    for _name in ("Mirror11350-01-02M", "Mirror-11350-01",
                  "MIRROR 11350-01-02M", "11350-01-02M MIRRORED"):
        ok(_re.search(_cfg.MIRROR_PATTERN, _name, _re.IGNORECASE),
           f"{_name} is a mirrored part")

    # AND IT MUST NOT FIRE INSIDE AN UNRELATED WORD. A lookahead of [\\w-] matched
    # "MIRRORLIKE"; only a digit or a separator can follow the marker.
    for _name in ("MIRRORLIKE-01", "MIRRORED_FINISH_PANEL".replace("MIRRORED", "MIRRORLIKE")):
        ok(not _re.search(_cfg.MIRROR_PATTERN, _name, _re.IGNORECASE),
           f"{_name} is not a mirrored part")


# ── job 11350 — a drawing of a part is not the part's flat pattern ───────────────────
def test_a_drawing_export_is_not_measured_as_a_flat_pattern():
    """Job 11350 shipped five DXFs. Three are drawing exports carrying title blocks,
    dimensions and hundreds of annotation entities — under filenames indistinguishable from
    the two real flats.

    is_flat_part_dxf accepts anything that is a .dxf, is not named as a GA, and has a part
    number in its name, and a SolidWorks drawing export satisfies all three. What that costs
    if it is missed: the title-block border and every dimension line get measured as cut
    path, so the part comes back with a bounding box the size of an A3 sheet, a cut length
    several times the real one, and a hole count that includes the arrowheads — and every
    one of those numbers looks like a measurement, because it was measured."""
    try:
        import ezdxf
    except ImportError:
        _fail("ezdxf is required to verify DXF content classification; the pipeline "
              "depends on it, so its absence is a broken environment, not a green run")
        return
    import tempfile
    from pathlib import Path
    from drawing_job_merge import drawing_export_reason, apply_dxf_geometry_to_part

    _d = Path(tempfile.mkdtemp())

    # A real flat pattern: outline and a bend line, no annotation.
    _doc = ezdxf.new(); _msp = _doc.modelspace()
    _msp.add_lwpolyline([(0, 0), (758, 0), (758, 60), (0, 60)], close=True)
    _msp.add_line((0, 20), (758, 20), dxfattribs={"layer": "BENDLINES"})
    _flat = _d / "11350-01-01M_1MM MS.dxf"; _doc.saveas(_flat)

    # A flat carrying a couple of etch labels. MUST NOT be rejected — over-rejecting costs
    # geometry we really had, and this is the shape that would cause it.
    _doc2 = ezdxf.new(); _msp2 = _doc2.modelspace()
    _msp2.add_lwpolyline([(0, 0), (200, 0), (200, 60), (0, 60)], close=True)
    for _i, _t in enumerate(("ETCH A", "ETCH B")):
        _msp2.add_text(_t).set_placement((10, 10 * _i + 5))
    _etched = _d / "11350-01-02M_1MM MS.dxf"; _doc2.saveas(_etched)

    # A drawing export: a dimension and a title block.
    _doc3 = ezdxf.new(); _msp3 = _doc3.modelspace()
    _msp3.add_lwpolyline([(0, 0), (200, 0), (200, 60), (0, 60)], close=True)
    _msp3.add_lwpolyline([(-50, -50), (450, -50), (450, 300), (-50, 300)], close=True)
    _msp3.add_linear_dim(base=(0, -20), p1=(0, 0), p2=(200, 0)).render()
    for _i, _t in enumerate(("BOOTS COMMS BAR", "REV B", "SCALE 1:2", "MILD STEEL",
                             "DRAWN JG", "11350-01-02", "SHEET 1 OF 1", "DO NOT SCALE",
                             "RAL 9003")):
        _msp3.add_text(_t).set_placement((10, 10 * _i))
    _drawing = _d / "Mirror11350-01-02M_1MM MS.dxf"; _doc3.saveas(_drawing)

    eq(drawing_export_reason(_flat), None, "a real flat pattern is accepted")
    eq(drawing_export_reason(_etched), None,
       "and so is a flat carrying an etch label or two")
    ok(drawing_export_reason(_drawing),
       "a DXF carrying dimensions is a drawing of the part, not the part")
    ok("dimension" in (drawing_export_reason(_drawing) or "").lower(),
       "and the reason names what gave it away")

    # THE FILENAME GATE STILL PASSES IT — which is exactly why the content test exists.
    from drawing_job_merge import is_flat_part_dxf
    ok(is_flat_part_dxf(_drawing),
       "the cheap filename gate cannot tell the difference; that is the point")

    # REFUSED WHERE GEOMETRY IS APPLIED, and recorded rather than silently skipped: a DXF
    # that exists and was rejected is a different fact from no DXF at all.
    _part = {"part_number": "11350-01-02M"}
    apply_dxf_geometry_to_part(_part, _drawing)
    ok(not _part.get("geometry_rollup"),
       "no geometry may be taken from a drawing export")
    eq([r["file"] for r in (_part.get("rejected_dxf_files") or [])],
       ["Mirror11350-01-02M_1MM MS.dxf"], "the rejection is recorded against the part")
    ok(any("rejected" in str(f) for f in (_part.get("review_flags") or [])),
       "and raised as a review flag — the estimator must know a flat is missing")

    # AND A REAL FLAT STILL YIELDS ITS GEOMETRY. Without this the fixture would pass
    # against a version that rejected everything.
    _part2 = {"part_number": "11350-01-01M"}
    apply_dxf_geometry_to_part(_part2, _flat)
    ok((_part2.get("geometry_rollup") or {}).get("estimated_bend_line_count"),
       "the real flat is still measured, bend line and all")


if __name__ == "__main__":
    sys.exit(main())
