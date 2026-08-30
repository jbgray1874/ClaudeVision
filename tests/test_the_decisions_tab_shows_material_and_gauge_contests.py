"""The Decision Report lists the arguments about WHAT THE PART IS, not just about operations.

On 8352 this tab's contest block held two lines, both "powder coating: not_applicable" — while
the same run reported a back panel costed as PLYWOOD that inference read as TIMBER, and three
parts whose gauge two sources put up to 4x apart. Material decides the sheet rate and whether the
part has a rate at all; gauge decides the rate AND steps the cut time, so a part costed on the
wrong one is wrong twice. Those are the decisions worth a person's minute, and the tab that exists
to show decisions did not show them.

The rows are READ from the invariants that already found them, never recomputed — the comparison
rules are fiddly and earned (a material spelling variant is not a disagreement; a gauge must
differ by enough to move the money), and a second copy would let the tab and the review block
disagree about the same part.
"""
from __future__ import annotations

import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from openpyxl import Workbook  # noqa: E402

from job_decision_report import (  # noqa: E402
    add_decision_report_sheet,
    datum_decisions_that_required_resolution,
)


def _material_finding(part="8352-01-03", costed="PLYWOOD", other="TIMBER"):
    return {"code": "two_sources_disagree_about_the_material", "severity": "warning",
            "message": "m", "detail": {"parts": [
                {"part_number": part, "costed_as": costed, "costed_from": "drawing_deterministic",
                 "other": other, "other_from": "inference"}]}}


def _gauge_finding(*rows):
    return {"code": "two_sources_disagree_about_the_gauge", "severity": "warning",
            "message": "g", "detail": {"parts": list(rows)}}


def _gauge_row(part, costed, other, ratio):
    return {"part_number": part, "costed_as": costed, "costed_from": "solidworks_api",
            "other": other, "other_from": "drawing_deterministic", "ratio": ratio}


def _summary(*violations):
    return {"manufacturing_writeup": {"parts": [{"part_number": "8352-01-03", "quantity": 1}]},
            "invariants": {"violations": list(violations)}}


# ── the reader ───────────────────────────────────────────────────────────────────────
def test_material_and_gauge_contests_are_both_listed():
    rows = datum_decisions_that_required_resolution(_summary(
        _material_finding(),
        _gauge_finding(_gauge_row("8352-01-03", 12.0, 3.0, 4.0))))
    assert len(rows) == 2
    assert {r["datum"] for r in rows} == {"material", "gauge"}


def test_the_widest_disagreement_is_listed_first():
    """A gauge out by 4x is read before one out by 1.33x; a categorical material argument, which
    has no measurable distance, sorts after the measured ones."""
    rows = datum_decisions_that_required_resolution(_summary(
        _material_finding(),
        _gauge_finding(_gauge_row("A", 4.0, 3.0, 1.33),
                       _gauge_row("B", 12.0, 3.0, 4.0),
                       _gauge_row("C", 1.2, 3.0, 2.5))))
    assert [r["part_number"] for r in rows] == ["B", "C", "A", "8352-01-03"]


def test_each_row_names_both_readings_and_where_they_came_from():
    row = datum_decisions_that_required_resolution(_summary(_material_finding()))[0]
    assert row["costed_as"] == "PLYWOOD" and row["costed_from"] == "drawing_deterministic"
    assert row["other"] == "TIMBER" and row["other_from"] == "inference"


def test_unrelated_findings_are_ignored():
    rows = datum_decisions_that_required_resolution(_summary(
        {"code": "price_not_reproducible", "severity": "blocking", "message": "x", "detail": {}}))
    assert rows == []


def test_a_job_with_no_invariants_is_safe():
    assert datum_decisions_that_required_resolution({}) == []
    assert datum_decisions_that_required_resolution(None) == []


# ── the rendered sheet ───────────────────────────────────────────────────────────────
def _sheet(summary):
    wb = Workbook()
    add_decision_report_sheet(wb, summary, {})
    return wb["Decision Report"]


def test_the_sheet_renders_the_block_and_its_rows():
    ws = _sheet(_summary(_material_finding(),
                         _gauge_finding(_gauge_row("8352-01-03", 12.0, 3.0, 4.0))))
    text = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any("WHAT THE PART IS" in t for t in text)
    assert "material" in text and "gauge" in text


def test_the_banner_counts_the_material_and_gauge_contests():
    """The banner named only operation contests, so a job whose real arguments were about
    material and gauge announced '2 decisions' with four costlier ones unmentioned."""
    ws = _sheet(_summary(_material_finding(),
                         _gauge_finding(_gauge_row("8352-01-03", 12.0, 3.0, 4.0))))
    banner = ws.cell(row=4, column=1).value or ""
    assert "2 decision(s) required resolution" in banner
    assert "material/gauge" in banner


def test_a_job_with_no_contests_still_says_so():
    ws = _sheet(_summary())
    assert "No decision required resolution" in (ws.cell(row=4, column=1).value or "")
