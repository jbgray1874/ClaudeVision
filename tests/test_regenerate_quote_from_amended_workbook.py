"""The estimator-override path: an amended workbook + three fields -> a fresh client quote.

The estimator edits the AI Estimate sheet, saves it, and later re-enters units, drawing number and
client and asks for a regenerated client quote. This reads the estimator's OWN figure off the sheet
(Sell Price preferred over Unit Cost) and renders the quote through the normal template — no engine
re-run, no drawing read. A sheet whose formulas were never recomputed yields a clear error, not a
blank quote.
"""
from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from openpyxl import Workbook  # noqa: E402

import client_quote_regen as cqr  # noqa: E402


def _amended_sheet(tmp, *, unit_cost=146.0, sell_price=210.5, qty=250, name="10575-02_MANUAL.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"
    ws["E4"] = "Quantity"
    ws["F4"] = qty
    ws["I105"] = "Total Unit Cost Price"
    ws["M105"] = unit_cost
    if sell_price is not None:
        ws["I110"] = "Sell Price"
        ws["M110"] = sell_price
    p = Path(tmp) / name
    wb.save(p)
    return p


def test_it_reads_the_estimators_sell_price_in_preference_to_unit_cost():
    d = tempfile.mkdtemp()
    figs = cqr.read_estimate_figures(_amended_sheet(d, unit_cost=146.0, sell_price=210.5))
    assert figs["price"] == 210.5 and figs["source_label"] == "Sell Price"
    assert figs["unit_cost"] == 146.0


def test_it_falls_back_to_unit_cost_when_no_margin_is_set():
    d = tempfile.mkdtemp()
    figs = cqr.read_estimate_figures(_amended_sheet(d, unit_cost=146.0, sell_price=None))
    assert figs["price"] == 146.0 and "Unit Cost" in figs["source_label"]


def test_a_pound_formatted_cell_still_parses():
    d = tempfile.mkdtemp()
    wb = Workbook(); ws = wb.active; ws.title = "Estimate"
    ws["A1"] = "Sell Price"; ws["B1"] = "£1,234.56"
    p = Path(d) / "s.xlsx"; wb.save(p)
    assert cqr.read_estimate_figures(p)["price"] == 1234.56


def test_the_regenerated_quote_carries_the_amended_price_and_the_three_fields():
    d = tempfile.mkdtemp()
    p = _amended_sheet(d, sell_price=210.5)
    out, figs = cqr.regenerate_quote_from_workbook(
        p, units=250, drawing_number="10575-02", client="Dyson", out_dir=d)
    html = Path(out).read_text(encoding="utf-8")
    assert os.path.exists(out) and out.endswith("_quote.html")
    assert "Dyson" in html                              # client heads the quote
    assert "10575-02" in html                           # drawing number
    assert ("52,625" in html or "52625" in html)        # order value = 250 x 210.50


def test_a_sheet_with_no_computed_figure_errors_rather_than_quoting_a_blank():
    d = tempfile.mkdtemp()
    wb = Workbook(); ws = wb.active; ws.title = "Estimate"
    ws["I105"] = "Total Unit Cost Price"                 # label but no value cached
    p = Path(d) / "blank.xlsx"; wb.save(p)
    try:
        cqr.read_estimate_figures(p)
        assert False, "should have raised"
    except ValueError as e:
        assert "Save" in str(e)                          # tells the estimator what to do


def test_client_and_units_are_required():
    d = tempfile.mkdtemp()
    p = _amended_sheet(d)
    for units, client in ((0, "Dyson"), (10, "")):
        try:
            cqr.regenerate_quote_from_workbook(p, units=units, drawing_number="X",
                                               client=client, out_dir=d)
            assert False, "should have raised"
        except ValueError:
            pass


def test_the_full_override_loop_saves_both_deliverables():
    """run_estimator_override saves the amended sheet as the manual-override record AND the
    regenerated quote, to their (here local) destinations, and returns both paths."""
    src = tempfile.mkdtemp()
    p = _amended_sheet(src, sell_price=195.0, qty=180, name="edited.xlsx")
    q_dir, x_dir = tempfile.mkdtemp(), tempfile.mkdtemp()
    res = cqr.run_estimator_override(p, units=180, drawing_number="10575-02", client="Dyson",
                                     quote_dir=q_dir, override_xlsx_dir=x_dir)
    assert os.path.basename(res["override_xlsx"]) == "10575-02_MANUAL_OVERRIDE.xlsx"
    assert os.path.exists(res["override_xlsx"]) and os.path.exists(res["quote_html"])
    html = Path(res["quote_html"]).read_text(encoding="utf-8")
    assert "Dyson" in html and ("35,100" in html or "35100" in html)   # 180 x 195


def test_the_override_quote_keeps_the_job_it_came_from():
    """THE THIN-QUOTE FIX. Built from the price alone, the override quotation lost everything the
    estimator had not restated: the specification read 'Material: As drawing', the operations
    collapsed to two generic lines and the GA image was dropped. It now starts from the job's own
    saved summary and overlays only the money and the three re-entered facts."""
    import json
    d = tempfile.mkdtemp()
    stem = "8352-010ReuseableBagStand"
    Path(d, f"{stem}.json").write_text(json.dumps({
        "job_output_stem": stem,
        "llm_full_extract": {"drawing_info": {"drawing_number": "8352", "title": "BAG STAND"}},
        "estimate_summary": {
            "part_estimates": [{"part_number": "8352-01-05",
                                "normalized_material": "MILD STEEL",
                                "textual_operations": ["laser_cutting", "welding"]}],
            "workbook_equivalent_pricing": {"m105_total_unit_cost_gbp": 137.48},
            "estimate_workbook_inputs": {"assumed_job_quantity": 400}}}), encoding="utf-8")
    p = _amended_sheet(d, sell_price=195.06, name=f"{stem}_20260818_133037.xlsx")
    out, figs = cqr.regenerate_quote_from_workbook(
        p, units=400, drawing_number="8352", client="M & S", out_dir=d)
    html = Path(out).read_text(encoding="utf-8")
    assert figs["source_summary_found"] is True
    assert "MILD STEEL" in html.upper()          # the real material, not "As drawing"
    assert "195.06" in html                       # and the estimator's own price


def test_the_run_timestamp_is_stripped_to_find_the_job():
    assert cqr._job_stem_from_workbook("8352-010ReuseableBagStand_20260818_133037.xlsx") \
        == "8352-010ReuseableBagStand"
    # a workbook the estimator renamed keeps its own name
    assert cqr._job_stem_from_workbook("my copy.xlsx") == "my copy"


def test_a_job_with_no_saved_summary_still_produces_a_quote():
    """A plain quote beats no quote — but the caller is told which it got."""
    d = tempfile.mkdtemp()
    p = _amended_sheet(d, name="orphan_20260818_133037.xlsx")
    out, figs = cqr.regenerate_quote_from_workbook(
        p, units=10, drawing_number="X1", client="Dyson", out_dir=d)
    assert figs["source_summary_found"] is False
    assert os.path.exists(out)


def test_the_override_validates_before_writing_anything():
    """A bad request must leave no half-made files on the share — validation precedes the copy."""
    src = tempfile.mkdtemp()
    p = _amended_sheet(src, name="edited.xlsx")
    x_dir = tempfile.mkdtemp()
    try:
        cqr.run_estimator_override(p, units=0, drawing_number="X", client="Dyson",
                                   override_xlsx_dir=x_dir)
        assert False, "should have raised"
    except ValueError:
        pass
    assert not any(f.endswith("_MANUAL_OVERRIDE.xlsx") for f in os.listdir(x_dir))
