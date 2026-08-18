"""POST /api/estimate/override: an amended workbook + three fields -> a regenerated client quote.

The backend endpoint runs the engine's regenerator OUT OF PROCESS (config isolation), so this
exercises the whole path: multipart upload -> validation -> subprocess CLI -> JSON response with a
ready-to-open quote URL. The engine-python is pointed at this interpreter and the output root at a
temp dir, so the test needs no share and no separate venv.

Skipped cleanly where the upload stack (python-multipart) or the backend's config is unavailable.
"""
from __future__ import annotations

import os
import sys
import tempfile

import pytest

_REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


@pytest.fixture(scope="module")
def client():
    pytest.importorskip("multipart", reason="python-multipart not installed")
    pytest.importorskip("fastapi", reason="fastapi not installed")
    out = tempfile.mkdtemp()
    os.environ["SDI_ENGINE_PYTHON"] = sys.executable
    os.environ["SDI_ESTIMATE_OUTPUT_ROOT"] = out
    os.environ["SDI_FILE_ROOTS"] = out
    backend_dir = os.path.join(_REPO, "sdi-intelligence-backend")

    # ISOLATE the backend's config from the engine's. Both modules are named `config`, and in the
    # shared test interpreter the engine's src/config is already imported — so the backend's
    # `import config` would return it and miss API_KEY. In production the backend runs in its own
    # process and never sees the engine's config; here we displace the clashing modules, import
    # the backend fresh with its dir first on the path, then restore so later engine tests are
    # unaffected. (The endpoint's own subprocess CLI has a separate interpreter, so no clash there.)
    _clash = ("config", "estimate_routes", "log_filters", "hr_routes")
    saved = {n: sys.modules.pop(n) for n in _clash if n in sys.modules}
    sys.path.insert(0, backend_dir)
    try:
        from fastapi import FastAPI
        from fastapi.testclient import TestClient
        import estimate_routes as _er
        app = FastAPI()
        app.include_router(_er.router)
        tc = TestClient(app)
        # The service may have a real API key configured (from .env). Authenticate as a caller
        # would, so the test exercises the endpoint rather than the gate — and the gate itself
        # stays live: an unauthenticated call is asserted to be refused below.
        _key = getattr(_er.config, "API_KEY", "") or ""
        if _key:
            tc.headers.update({"X-SDI-Key": _key})
        yield tc
    except Exception as exc:                                    # noqa: BLE001
        pytest.skip(f"backend not importable here: {exc}")
    finally:
        try:
            sys.path.remove(backend_dir)
        except ValueError:
            pass
        for n in _clash:
            sys.modules.pop(n, None)
        sys.modules.update(saved)


def _amended_xlsx(sell=195.0):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "Estimate"
    ws["I110"] = "Sell Price"; ws["M110"] = sell
    p = os.path.join(tempfile.mkdtemp(), "edited.xlsx"); wb.save(p)
    return p


def test_a_valid_override_returns_the_quote_and_writes_both_files(client):
    with open(_amended_xlsx(195.0), "rb") as f:
        r = client.post(
            "/api/estimate/override",
            files={"file": ("edited.xlsx", f,
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"units": "180", "drawing": "10575-02", "client": "Dyson"})
    assert r.status_code == 200, r.text
    j = r.json()
    assert j["manual_override"] is True
    assert j["price"] == 195.0 and j["price_source"] == "Sell Price"
    assert j["quote_url"] and j["quote_url"].startswith("/api/file?path=")
    assert os.path.exists(j["quote_html"]) and os.path.exists(j["override_xlsx"])
    assert j["override_xlsx"].endswith("_MANUAL_OVERRIDE.xlsx")


def test_the_endpoint_is_behind_the_api_key_when_one_is_configured(client):
    """This is a WRITE path onto a share — it must not be open. Where a key is configured, a call
    without it is refused; where none is (a dev box), there is no gate to assert."""
    import estimate_routes as _er
    if not getattr(_er.config, "API_KEY", ""):
        pytest.skip("no API key configured in this environment")
    r = client.post("/api/estimate/override",
                    files={"file": ("notes.txt", b"hello", "text/plain")},
                    data={"units": "10", "drawing": "X", "client": "Y"},
                    headers={"X-SDI-Key": "not-the-key"})
    assert r.status_code == 401


def test_a_non_xlsx_upload_is_refused(client):
    r = client.post("/api/estimate/override",
                    files={"file": ("notes.txt", b"hello", "text/plain")},
                    data={"units": "10", "drawing": "X", "client": "Y"})
    assert r.status_code == 415


def test_an_empty_upload_is_refused(client):
    r = client.post(
        "/api/estimate/override",
        files={"file": ("empty.xlsx", b"",
               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"units": "10", "drawing": "X", "client": "Y"})
    assert r.status_code == 400


def test_a_sheet_with_no_computed_price_is_a_400_not_a_500(client):
    """The CLI's plain 'open and Save in Excel first' error must reach the estimator as a 400."""
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "Estimate"
    ws["I105"] = "Total Unit Cost Price"                        # label, no cached value
    p = os.path.join(tempfile.mkdtemp(), "blank.xlsx"); wb.save(p)
    with open(p, "rb") as f:
        r = client.post(
            "/api/estimate/override",
            files={"file": ("blank.xlsx", f,
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"units": "10", "drawing": "X", "client": "Dyson"})
    assert r.status_code == 400
