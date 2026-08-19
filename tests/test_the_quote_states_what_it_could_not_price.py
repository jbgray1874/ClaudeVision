"""Missing drawings are named on every INTERNAL document and on none of the customer's.

An incomplete pack is priced rather than refused, so the total is real for what was supplied and
is not a price for the whole product. Everyone working the job needs to know which parts are
outside it; what the CUSTOMER is told about scope is the estimator's decision and their wording,
not a list this engine generates from its own findings.

So: the Estimate sheet, the AI Provenance tab, the Decision Report and the job report's
drawing-pack section all name the undrawn parts, and the quotation carries none of it — it is
labelled INDICATIVE and says nothing about how the engine reached the figure.

All four read costed_facts.undrawn_bom_lines, so no two documents from one run can name
different parts.
"""
from __future__ import annotations

import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from openpyxl import Workbook  # noqa: E402

from client_quote_html import build_quote_html  # noqa: E402
from costed_facts import undrawn_bom_lines  # noqa: E402
from job_decision_report import add_decision_report_sheet  # noqa: E402

_MISSING = [
    {"part_number": "10575-01-101", "description": "VERSION 1 - BACK WELDED ASSEMBLY"},
    {"part_number": "10575-01-102", "description": "VERSION 1 - BASE WELDED ASSEMBLY"},
    {"part_number": "10575-01-104", "description": "BASE PLATE - WELDED ASSEMBLY"},
    {"part_number": "10575-01-103", "description": "VERTICAL LENS ASSEMBLY"},
]


def _summary(missing=None, qty=1, unit=146.0):
    violations = []
    if missing is not None:
        violations.append({"code": "bom_names_a_drawing_the_pack_does_not_contain",
                           "severity": "blocking", "message": "x",
                           "detail": {"missing": missing}})
    violations.append({"code": "price_not_reproducible", "severity": "blocking",
                       "message": "y", "detail": {}})
    return {"manufacturing_writeup": {"parts": [{"part_number": "10575-02-009", "quantity": 1}]},
            "invariants": {"violations": violations},
            "estimate_summary": {
                "workbook_equivalent_pricing": {"m105_total_unit_cost_gbp": unit},
                "estimate_workbook_inputs": {"assumed_job_quantity": qty}}}


# ── one reader, so no two documents disagree ─────────────────────────────────────────
def test_the_shared_reader_collects_the_undrawn_lines():
    rows = undrawn_bom_lines(_summary(_MISSING))
    assert [r["part_number"] for r in rows] == [m["part_number"] for m in _MISSING]


def test_unrelated_findings_are_ignored():
    assert undrawn_bom_lines(_summary(missing=None)) == []


def test_a_summary_with_nothing_in_it_is_safe():
    assert undrawn_bom_lines({}) == []
    assert undrawn_bom_lines(None) == []


# ── the customer's quotation carries none of it ──────────────────────────────────────
def test_the_quote_does_not_name_the_missing_drawings():
    """What the customer is told about scope is the estimator's call, in their words."""
    html = build_quote_html(_summary(_MISSING), job_stem="10575-02", customer="Dyson")
    for m in _MISSING:
        assert m["part_number"] not in html
    assert "Not included in this price" not in html


def test_the_quote_still_labels_the_price_indicative():
    """Nothing on the page may read as a firm price for a whole product."""
    html = build_quote_html(_summary(_MISSING), job_stem="10575-02", customer="Dyson")
    assert "indicative" in html.lower()


def test_no_engine_language_reaches_the_customer():
    html = build_quote_html(_summary(_MISSING), job_stem="10575-02", customer="Dyson")
    assert "bom_names_a_drawing" not in html
    assert "invariant" not in html.lower()


# ── the internal documents all name them ─────────────────────────────────────────────
def test_the_decision_report_names_them():
    wb = Workbook()
    add_decision_report_sheet(wb, _summary(_MISSING), {})
    text = [c.value for row in wb["Decision Report"].iter_rows()
            for c in row if isinstance(c.value, str)]
    assert any("DRAWINGS MISSING" in t for t in text)
    assert any("10575-01-101" in t for t in text)


def test_a_complete_pack_adds_no_block_anywhere():
    """A job that could price everything must not gain an empty, worrying section."""
    clean = _summary(missing=None, qty=400, unit=100.0)
    wb = Workbook()
    add_decision_report_sheet(wb, clean, {})
    text = [c.value for row in wb["Decision Report"].iter_rows()
            for c in row if isinstance(c.value, str)]
    assert not any("DRAWINGS MISSING" in t for t in text)
