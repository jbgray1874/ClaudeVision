"""Every blank price on the sheet says which kind of nothing it is.

THE READER EXISTED AND THE WRITER DID NOT. price_provenance has carried the unpriced
vocabulary since the day it was written, and invariants.check_every_unpriced_line_says_why
has read `row["unpriced_reason"]` for exactly as long. Nothing anywhere set it. So the check
whose entire purpose is to stop a blank cell reading as free ran on every job and reported on
no line of any of them -- built is not wired, pointing the other way for once, and the
harder direction to notice: the check was green.

WHY THE DISTINCTION IS WORTH THE CODE. 11650's cabinet BOM carries sixteen fabricated lines
at GBP 0.00 whose material is costed in the Sheet Steel block, beside a lock, a mag catch and
a set of feet that nobody has priced at all. On the sheet they are identical -- a blank in
the price column -- and they need opposite actions: one set must NOT be priced, on pain of
doubling the material total, and the other must be. A checklist that cannot tell them apart
is one an estimator stops working, and what gets lost is the half that was real.
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

import price_provenance as pp                                        # noqa: E402
import estimator_inputs as ei                                        # noqa: E402
import wep_readback_from_xlsx as wep                                 # noqa: E402
from invariants import check_every_unpriced_line_says_why as check   # noqa: E402


# ── the categories, decided from the record and never from a name ───────────────────
@pytest.mark.parametrize("record,category,owner", [
    # The majority case, and the one that must never reach the estimator's list.
    ({"_bom_cross_reference": True}, pp.NOT_APPLICABLE, "nobody"),
    ({"_duplicate_of": "FIXING1081"}, pp.NOT_APPLICABLE, "nobody"),
    ({"_canonical_kind": "assembly"}, pp.NOT_APPLICABLE, "nobody"),
    # A figure exists and we decided not to stand behind it. The blank is a policy.
    ({"_ai_indicative_gbp": 86.04}, pp.POLICY_WITHHELD, "estimator"),
    ({"_consumable_qty_unknown": True}, pp.NOT_MEASURED, "estimator"),
    # Nothing we can query holds this item -- the commonest reason on an SDI sheet.
    ({"part_number": "MAG CATCH"}, pp.NO_PRICE_SOURCE, "estimator"),
])
def test_each_kind_of_blank_is_named_from_the_record(record, category, owner):
    reason = ei.unpriced_reason_for_row(record)
    assert reason["category"] == category
    assert reason["owner"] == owner


def test_no_price_source_is_not_dressed_up_as_a_missing_dimension():
    """Both send the line to the estimator, and forcing this one to borrow NOT_MEASURED would
    print "the dimension or quantity it needs was never measured" against a fitting whose
    dimensions are perfectly well known and whose problem is that no catalogue covers it.
    Different action -- ask the supplier, load a price file -- and a checklist that describes
    it wrongly is one people stop trusting."""
    text = pp.describe_unpriced(pp.NO_PRICE_SOURCE)
    assert "never measured" not in text
    assert "catalogue, price file or quote" in text
    assert pp.unpriced_reason(pp.NO_PRICE_SOURCE)["undercharging"] is False


def test_a_cross_reference_is_nobodys_job():
    """Asking someone to price it is asking for the double-count back."""
    reason = ei.unpriced_reason_for_row({"_bom_cross_reference": True})
    assert reason["owner"] == "nobody" and not reason["undercharging"]
    assert "double" in reason["detail"]


# ── the writer, on rows shaped like the ones the sheet gives back ───────────────────
def _es():
    return {"part_estimates": [
        {"part_number": "11650-01-01M", "_bom_cross_reference": True},
        {"part_number": "ESSENTRA FOOT-466122", "_duplicate_of": "FIXING1081"},
        {"part_number": "MAG CATCH"},
    ]}


def _rows():
    return [
        # The fabricated blocks carry NO part code column -- the number is the first word of
        # the description. Reading only the code column leaves every one of them unexplained.
        {"part_code": "", "description": "11650-01-01M  LH UPRIGHT — costed in Sheet Steel below",
         "total_value_gbp": 0},
        {"part_code": "ESSENTRA FOOT-466122", "description": "foot", "total_value_gbp": 0},
        {"part_code": "MAG CATCH", "description": "HAFELE 246.41.745", "total_value_gbp": None},
        {"part_code": "FIXING1081", "description": "foot", "total_value_gbp": 0.46},
    ]


def test_every_unpriced_row_is_stamped_and_priced_rows_are_left_alone():
    rows = _rows()
    assert wep._explain_unpriced_rows(rows, _es()) == 3
    assert [r.get("unpriced_reason", {}).get("category") for r in rows] == [
        pp.NOT_APPLICABLE, pp.NOT_APPLICABLE, pp.NO_PRICE_SOURCE, None]


def test_a_fabricated_row_is_joined_by_the_first_word_of_its_description():
    """The BOM block has a Part code column and the fabricated blocks do not. Joining on the
    code alone explains the bought-ins and silently leaves the sixteen cross-reference rows
    blank -- which is the majority of the sheet and the whole point of the exercise."""
    rows = _rows()
    wep._explain_unpriced_rows(rows, _es())
    assert rows[0]["unpriced_reason"]["category"] == pp.NOT_APPLICABLE


def test_a_row_no_record_matches_is_loudly_unexplained_not_silently_skipped():
    """Falling silent for a row whose part record cannot be found reproduces exactly the
    failure this exists to end. UNEXPLAINED is the honest answer -- we did not price it and
    we cannot say why -- and the invariant reports it."""
    rows = [{"part_code": "GHOST", "description": "?", "total_value_gbp": 0}]
    wep._explain_unpriced_rows(rows, _es())
    assert rows[0]["unpriced_reason"]["category"] == pp.UNEXPLAINED


def test_an_empty_job_does_not_raise():
    assert wep._explain_unpriced_rows([], {}) == 0
    assert wep._explain_unpriced_rows([{"total_value_gbp": 0}], {}) == 1


# ── and the check that has been reporting nothing now reports ───────────────────────
def _job(rows):
    return {"estimate_summary": {"final_estimate": {"material_rows": rows}}}


def test_the_invariant_was_green_on_a_sheet_full_of_unexplained_blanks():
    """The state before this change: every unpriced row carried no reason, so the check found
    nothing to say. It is the shape of failure that survives longest -- a guard reporting
    CLEAR because nobody ever gave it anything to read."""
    silent = [{"part_number": "MAG CATCH", "price_gbp": 0}]
    assert [v["code"] for v in check(_job(silent))] == ["unpriced_line_says_why"]


def test_a_stamped_sheet_passes_and_a_cross_reference_raises_nothing():
    rows = _rows()
    wep._explain_unpriced_rows(rows, _es())
    for r in rows:
        r["price_gbp"] = r.get("total_value_gbp")
    assert check(_job(rows)) == []


def test_an_engine_gap_is_still_reported_as_undercharging():
    """The narrowing must not silence the finding worth a person's time: work that is really
    done, really invoiced, and that nothing on the sheet asks anybody to price."""
    rows = [{"part_number": "VINYL-01", "price_gbp": 0,
             "unpriced_reason": pp.unpriced_reason(pp.NO_VOCABULARY, "REEDED VINYL")}]
    assert [v["code"] for v in check(_job(rows))] == ["unpriced_because_the_engine_cannot"]


# ── the wiring ──────────────────────────────────────────────────────────────────────
def test_the_writer_is_called_on_the_real_readback_path():
    """This module is the one that proved a vocabulary can exist for months with no writer.
    Defining a second one and not calling it would be the same defect wearing this change's
    clothes."""
    src = Path(wep.__file__).read_text(encoding="utf-8")
    body = src[src.index("def stamp_real_totals_into_json"):]
    assert "_explain_unpriced_rows(" in body, \
        "the explainer is defined and never called from the read-back"
    assert body.index("_explain_unpriced_rows(") < body.index("should_stamp_final_estimate"), \
        "rows are stamped into final_estimate before they carry their reasons"


# ── and the report says it, ordered by who has to act ───────────────────────────────
import job_report_html as jrh                                        # noqa: E402


def _report(rows):
    return jrh._unpriced_section(
        {"estimate_summary": {"final_estimate": {"material_rows": rows}}})


def test_the_engine_gap_leads_because_it_is_the_only_one_worth_interrupting_for():
    """An engine gap is work that will be done and invoiced with nothing on the sheet asking
    anyone to price it. No estimator input can fix it. Everything else on this table is
    either somebody's task or correctly nil, so the one that under-charges the job leads."""
    html = _report([
        {"part_number": "NIL-1", "price_gbp": 0,
         "unpriced_reason": pp.unpriced_reason(pp.NOT_APPLICABLE, "costed in Sheet Steel")},
        {"part_number": "ASK-1", "price_gbp": 0,
         "unpriced_reason": pp.unpriced_reason(pp.NO_PRICE_SOURCE, "HAFELE 246.41.745")},
        {"part_number": "GAP-1", "price_gbp": 0,
         "unpriced_reason": pp.unpriced_reason(pp.NO_VOCABULARY, "REEDED VINYL")}])
    assert html.index("GAP-1") < html.index("ASK-1") < html.index("NIL-1")
    assert "under-charged" in html


def test_the_three_owners_are_counted_separately():
    html = _report([
        {"part_number": "A", "price_gbp": 0,
         "unpriced_reason": pp.unpriced_reason(pp.NOT_APPLICABLE)},
        {"part_number": "B", "price_gbp": 0,
         "unpriced_reason": pp.unpriced_reason(pp.NOT_APPLICABLE)},
        {"part_number": "C", "price_gbp": 0,
         "unpriced_reason": pp.unpriced_reason(pp.NO_PRICE_SOURCE)}])
    assert "<b>1</b> waiting on the estimator" in html
    assert "<b>2</b> correctly nil" in html


def test_a_priced_row_is_not_listed_as_a_blank():
    html = _report([{"part_number": "PRICED", "price_gbp": 0.46}])
    assert "PRICED" not in html and "carries a price" in html


def test_blanks_with_no_reasons_are_a_warning_not_an_empty_table():
    """The vocabulary existed for months with no writer and the check stayed green throughout.
    A report that quietly shows an empty table when the stamping did not run would let exactly
    that happen again, one layer up."""
    html = _report([{"part_number": "X", "price_gbp": 0}])
    assert "warn" in html and "no recorded reason" in html


def test_the_section_is_wired_into_the_report():
    src = Path(jrh.__file__).read_text(encoding="utf-8")
    import ast
    body = ast.unparse(next(n for n in ast.walk(ast.parse(src))
                            if isinstance(n, ast.FunctionDef) and n.name == "_render_verdict"))
    assert "_unpriced_section" in body, "the section is defined and never called"


if __name__ == "__main__":                                            # pragma: no cover
    raise SystemExit(pytest.main([__file__, "-q"]))


# ── a check with nothing to look at is not a pass ───────────────────────────────────
# invariants.py states this rule at the top of the file and every reconciliation check obeys
# it. This one, added later, did not: `if not rows: return []`. The Excel COM read-back fails
# for reasons that have nothing to do with the estimate -- an ELEVATED console, a workbook
# that will not open, Excel busy -- and it leaves no final_estimate at all. So on exactly the
# runs where least is known, the newest guard reported a clean sheet.
def test_no_read_back_is_unverified_not_a_pass():
    """A guard that goes green when its input vanishes is worse than no guard, because it
    gets quoted as evidence."""
    out = check({"estimate_summary": {}})
    assert [v["severity"] for v in out] == ["unverified"]
    assert "no material row was read back" in out[0]["message"]


def test_material_rows_of_the_wrong_shape_are_unverified_too():
    out = check({"estimate_summary": {"final_estimate": {"material_rows": "broken"}}})
    assert [v["severity"] for v in out] == ["unverified"]


def test_a_read_back_that_genuinely_found_no_material_rows_is_a_pass():
    """The distinction the fix turns on. An empty list from a sheet that WAS read is a real
    answer; the absence of the sheet is not."""
    assert check({"estimate_summary": {"final_estimate": {"material_rows": []}}}) == []


def test_the_report_says_the_sheet_was_never_read_rather_than_vanishing():
    """The section disappearing tells the same lie the empty table would. The figures on that
    page then come from BEFORE Excel calculated -- a different total -- and nothing else on
    the page says so."""
    html = jrh._unpriced_section({"estimate_summary": {}})
    assert "warn" in html and "never read back" in html
    assert "before Excel calculated" in html


def test_the_check_reads_final_estimate_from_either_shape():
    """Some writers stamp final_estimate on the summary ROOT and some inside estimate_summary,
    which is why invariants keeps a shared resolver -- whose docstring says exactly what a
    private path does: "a check that looks in one place only reports a clean pass on a job it
    never examined." This check reached into estimate_summary directly and was doing that on
    every job of the other shape, silently, from the day it was written. It only surfaced when
    failing closed turned a silent pass into a visible unverified."""
    rows = [{"part_number": "MAG CATCH", "price_gbp": 0}]
    assert [v["code"] for v in check({"final_estimate": {"material_rows": rows}})] \
        == ["unpriced_line_says_why"]
    assert [v["code"] for v in check({"estimate_summary":
                                      {"final_estimate": {"material_rows": rows}}})] \
        == ["unpriced_line_says_why"]


def test_the_report_reads_final_estimate_from_either_shape():
    rows = [{"part_number": "X", "price_gbp": 0,
             "unpriced_reason": pp.unpriced_reason(pp.NO_PRICE_SOURCE)}]
    for job in ({"final_estimate": {"material_rows": rows}},
                {"estimate_summary": {"final_estimate": {"material_rows": rows}}}):
        assert "waiting on the estimator" in jrh._unpriced_section(job)


# ── and on the two surfaces an estimator actually opens ─────────────────────────────
import estimation_report as er                                       # noqa: E402


def _wb_job(parts, final_estimate=None):
    job = {"manufacturing_writeup": {"parts": parts}}
    if final_estimate is not None:
        job["final_estimate"] = final_estimate
    return job


def test_the_provenance_sheet_carries_a_reason_for_every_blank_line():
    """A blank in a money column reads as free on this tab exactly as it does on the Estimate
    tab, and this one is the tab an estimator opens to ask why."""
    rows = er.build_provenance(_wb_job([
        {"part_number": "MAG CATCH", "description": "HAFELE 246.41.745", "quantity": 2},
        {"part_number": "11650-01-01M", "description": "LH UPRIGHT", "quantity": 1,
         "_bom_cross_reference": True}]))
    by = {r["part_number"]: (r.get("unpriced_reason") or {}) for r in rows}
    assert by["MAG CATCH"]["owner"] == "estimator"
    assert by["11650-01-01M"]["owner"] == "nobody"


def test_the_sheet_uses_the_same_classifier_as_the_report_and_the_invariant():
    """A private second opinion here is how two documents describing one job come to disagree
    about which blanks are somebody's job."""
    src = Path(er.__file__).read_text(encoding="utf-8")
    assert "from estimator_inputs import unpriced_reason_for_row" in src


def test_the_sheet_reason_is_computed_from_the_record_not_the_read_back():
    """Joining back through final_estimate.material_rows would explain the blanks only on runs
    where the read-back worked -- and the read-back is precisely what fails on an elevated
    console or a busy Excel. The sheet must explain itself on the runs that needed it most."""
    rows = er.build_provenance(_wb_job(
        [{"part_number": "MAG CATCH", "description": "H", "quantity": 1}]))
    assert (rows[0].get("unpriced_reason") or {}).get("category") == pp.NO_PRICE_SOURCE


def test_a_priced_line_carries_no_reason():
    rows = er.build_provenance({
        "manufacturing_writeup": {"parts": [
            {"part_number": "FIXING1081", "description": "foot", "quantity": 2}]},
        "estimate_summary": {"part_estimates": [
            {"part_number": "FIXING1081", "unit_cost_gbp": 0.22,
             "extended_total_cost_gbp": 0.46}]}})
    assert rows[0].get("unpriced_reason") is None


def _sheet(job):
    import openpyxl
    wb = openpyxl.Workbook()
    er.add_provenance_sheet(wb, job, {"pdf_name": "x", "job_number": "11650"})
    return wb["AI Provenance"]


def test_the_sheet_has_a_column_for_it_and_fills_it():
    ws = _sheet(_wb_job([{"part_number": "MAG CATCH", "description": "H", "quantity": 1}],
                        final_estimate={"material_rows": []}))
    assert ws["P5"].value == "Not priced — why / who"
    assert "ESTIMATOR TO PRICE" in str(ws["P6"].value)


def test_the_sheet_says_when_the_calculated_sheet_was_never_read_back():
    """The money columns on this tab are then the ENGINE's pre-Excel figures rather than what
    the Estimate sheet computes -- two different totals -- and without this the tab looks
    exactly as it does on a run that reconciled perfectly."""
    ws = _sheet(_wb_job([{"part_number": "X", "description": "d", "quantity": 1}]))
    assert "NOT READ BACK" in str(ws["A4"].value)
    # NOT "re-run from a normal PowerShell". That advice was built on a claim about Excel and
    # elevation that turned out to be wrong three times over -- the workbook is written by
    # openpyxl and needs no COM, and the read-back uses DispatchEx, which starts its own
    # instance rather than attaching to one. Telling somebody to change console for a reason
    # that does not exist sends them to fix a machine that is fine.
    assert "Excel busy or absent" in str(ws["A4"].value)
    assert "normal PowerShell" not in str(ws["A4"].value)


def test_the_sheet_is_quiet_about_the_read_back_when_it_ran():
    ws = _sheet(_wb_job([{"part_number": "X", "description": "d", "quantity": 1}],
                        final_estimate={"material_rows": [{"description": "d"}]}))
    assert "NOT READ BACK" not in str(ws["A4"].value or "")


def test_the_html_report_says_it_at_the_top_not_only_in_section_eleven():
    """An estimator reading top-down has formed a view of the number long before section 11.
    Every figure above that point is presented as the workbook's, and on a run with no
    read-back it is not."""
    strip = jrh._provenance_strip({"estimate_summary": {}})
    assert "never read back" in strip and "warn" in strip
    quiet = jrh._provenance_strip(
        {"estimate_summary": {"final_estimate": {"material_rows": [{"a": 1}]}}})
    assert "never read back" not in quiet


# ── the writer and the reader must agree what a price IS ────────────────────────────
# LIVE FAILURE, 12 August. The run stamped "26 unpriced material row(s) explained" and the
# invariant then reported "29 line(s) carry no price and no reason: ?, ?, ?, ?, ?, ?".
#
# The read-back names its columns after the SHEET's headers -- total_value_gbp from "Total
# Value", unit_price_gbp from "Price". The check was written against price_gbp. Neither field
# exists on a real row, so every PRICED line came back as None and read as an unexplained
# blank; the labels printed "?" because the name lookup read part_number on rows that carry
# part_code. Two halves of one mistake: a private guess at the shape of somebody else's
# record, made twice, in code I wrote a day apart.
_SHEET_ROW = {"part_code": "FIXING1081", "description": "foot", "total_value_gbp": 0.46}
_BLANK_ROW = {"part_code": "MAG CATCH", "description": "HAFELE", "total_value_gbp": 0}


def test_a_row_priced_in_the_sheets_own_column_is_not_called_a_blank():
    assert not pp.row_is_unpriced(_SHEET_ROW)
    assert pp.row_is_unpriced(_BLANK_ROW)


def test_the_invariant_and_the_writer_agree_on_the_same_rows():
    """They disagreed on a live job, and the disagreement was invisible until an estimator
    read the two numbers side by side."""
    rows = [dict(_SHEET_ROW), dict(_BLANK_ROW)]
    wep._explain_unpriced_rows(rows, {"part_estimates": []})
    assert check(_job(rows)) and rows[0].get("unpriced_reason") is None
    codes = check(_job(rows))[0]["detail"]["lines"]
    assert codes == ["MAG CATCH"], f"the priced row is still being reported: {codes}"


def test_a_row_is_named_by_whatever_field_carries_its_code():
    """"?" tells a reader nothing they can act on, and six of them tells them the check is
    broken -- which is how a real finding gets ignored."""
    assert pp.row_label(_SHEET_ROW) == "FIXING1081"
    assert pp.row_label({"description": "11650-01-01M  LH UPRIGHT"}) == "11650-01-01M"
    assert pp.row_label({}) == "?"


def test_an_excel_error_is_not_a_price():
    """A cell that failed to calculate is not a zero and not a number. Casting it would stamp
    an error sentinel as money -- which this pipeline has done once already."""
    assert pp.row_price({"total_value_gbp": "#VALUE!"}) is None
    assert pp.row_is_unpriced({"total_value_gbp": "#VALUE!"})


# ── measured everything and still could not cost it ─────────────────────────────────
def test_a_measured_blank_with_no_rate_is_an_engine_gap_not_an_estimator_job():
    """11650's door. The model gave ABS, outranking a drawing and a DXF filename that both
    said POLYCARBONATE; config carries a sheet size and a density for ABS and no rate. A part
    with a measured 1202 x 689 blank went from GBP 35.28 to GBP 0.00 and showed a blank cell.

    Nothing is missing from the drawings, so this is not the estimator's to fill -- filing it
    beside the mag catch would put a phone call to a supplier and a missing number in this
    repository on the same list. And unlike the mag catch, it silently under-charges every
    job that touches that material."""
    r = ei.unpriced_reason_for_row({"normalized_material": "ABS",
                                    "blank_length_mm": 1202, "blank_width_mm": 689})
    assert r["category"] == pp.NO_VOCABULARY
    assert r["owner"] == "engine" and r["undercharging"] is True
    assert "ABS" in r["detail"] and "MATERIAL_PRICE_GBP_PER_KG" in r["detail"]


def test_an_unmeasured_bought_in_is_still_the_estimators():
    """The narrowing matters: a fitting nobody has a price for is a supplier question, and
    calling it an engine gap would bury the one finding a person must act on."""
    r = ei.unpriced_reason_for_row({"part_number": "MAG CATCH"})
    assert r["category"] == pp.NO_PRICE_SOURCE and r["owner"] == "estimator"


def test_a_material_with_no_measured_blank_is_not_claimed_as_an_engine_gap():
    r = ei.unpriced_reason_for_row({"normalized_material": "MILD_STEEL"})
    assert r["category"] == pp.NO_PRICE_SOURCE


# ── the sheet already says why, even when no part record matches ────────────────────
# 11650, 12 August: four rows came back UNEXPLAINED -- STD PART, FIXINGTBC, MAG CATCH,
# YIREE LOCK ASSEMBLY -- because no part record matched them. They are bought-in stubs minted
# late and never reach part_estimates under those codes.
#
# But the row is not silent. wb_populate has already written the reason into its DESCRIPTION,
# from input_note_for_line, and that sentence is the engine's own statement rather than a
# guess about it. Reading it back is the same fact from the only place on the row that still
# carries it -- not inference.
@pytest.mark.parametrize("description,category,owner", [
    ("MAG CATCH  HAFELE 246.41.745  —  NOT YET PRICED: enter the per-unit figure",
     pp.NOT_MEASURED, "estimator"),
    ("STD PART  M4 PEM STUD  —  MATERIAL UNPRICED: enter a unit rate for this item",
     pp.NO_PRICE_SOURCE, "estimator"),
    ("ESSENTRA FOOT-466122 — SAME ARTICLE AS FIXING1081: costed there, not here",
     pp.NOT_APPLICABLE, "nobody"),
    ("11650-01-01M  LH UPRIGHT — costed in Sheet Steel below",
     pp.NOT_APPLICABLE, "nobody"),
])
def test_a_row_with_no_part_record_is_read_from_its_own_description(
        description, category, owner):
    reason = wep._reason_from_the_row_itself({"description": description})
    assert reason["category"] == category and reason["owner"] == owner


def test_a_row_that_really_says_nothing_is_still_unexplained():
    """The fallback must stay honest. Inventing a category for a row that states no reason
    would remove the only signal that the join is failing somewhere it should not."""
    r = wep._reason_from_the_row_itself({"description": "SOME ROW"})
    assert r["category"] == pp.UNEXPLAINED
    assert "says nothing" in r["detail"]


def test_the_four_live_rows_are_no_longer_unexplained():
    """The exact set from the run, through the real entry point."""
    rows = [{"part_code": c, "description": d, "total_value_gbp": 0} for c, d in [
        ("STD PART", "STD PART  M4 PEM STUD  —  MATERIAL UNPRICED: enter a unit rate"),
        ("FIXINGTBC", "FIXINGTBC  M4 KNOB  —  NOT YET PRICED: enter the per-unit figure"),
        ("MAG CATCH", "MAG CATCH  HAFELE  —  NOT YET PRICED: enter the per-unit figure"),
        ("YIREE LOCK ASSEMBLY", "LOCK AND KEY  —  NOT YET PRICED: enter the per-unit figure"),
    ]]
    wep._explain_unpriced_rows(rows, {"part_estimates": []})
    for r in rows:
        r["price_gbp"] = 0
    assert check(_job(rows)) == [], "these four still report as blanks with no reason"
