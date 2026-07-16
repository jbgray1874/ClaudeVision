"""
Discover Estimate-sheet cells from visible labels (no fixed row numbers).

Used by:
- ``estimate_document`` / ``estimate_workbook_inputs`` (qty + manual £/tonne rates via ``read_estimate_workbook_inputs``),
- full parity, template write-back, workbook parity CSV,
- ``extract_workbook_constants`` (preferred over raw D6/L3/L5).

Scans label columns row-by-row, matches configurable regexes, and emits cell_ref -> JSON path maps where needed.
"""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import config


def _col_idx(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        if not ("A" <= ch <= "Z"):
            raise ValueError("bad column %r" % (letters,))
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _cell_text(ws: Any, row: int, col_idx: int) -> str:
    cell = ws.cell(row=row, column=col_idx)
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, (int, float)):
        return str(v)
    return str(v).strip()


def _label_blob(ws: Any, row: int, cols: List[str]) -> str:
    parts: List[str] = []
    for L in cols:
        t = _cell_text(ws, row, _col_idx(L))
        if t:
            parts.append(t)
    return " ".join(parts)


def _split_cell_a1(ref: str) -> Tuple[int, int]:
    """Return (column_index, row_number) for an A1-style reference."""
    m = re.match(r"^([A-Za-z]{1,3})(\d+)$", str(ref).strip())
    if not m:
        raise ValueError("bad cell ref %r" % (ref,))
    return _col_idx(m.group(1)), int(m.group(2))


def _safe_float(value: Any) -> Optional[float]:
    try:
        if value is None:
            return None
        if isinstance(value, str):
            v = value.strip().replace("£", "").replace(",", "")
            if not v:
                return None
            return float(v)
        return float(value)
    except (TypeError, ValueError):
        return None


def discover_output_cells_map(ws: Any) -> Tuple[Dict[str, str], Dict[str, Any]]:
    """
    Return (cell_ref -> summary dotted path, metadata).

    Paths are the same strings used in ``ESTIMATE_TEMPLATE_WRITEBACK.output_cells`` and
    ``_parity_money_specs`` (e.g. estimate_summary.workbook_equivalent_pricing.m59_...).
    """
    meta: Dict[str, Any] = {"rules_hits": [], "warnings": []}
    dsc = getattr(config, "ESTIMATE_SHEET_TOTAL_DISCOVERY", None) or {}

    tb = getattr(config, "ESTIMATE_TEMPLATE_WRITEBACK", {}) or {}
    static_fb: Dict[str, str] = {
        str(k).upper(): str(v) for k, v in (tb.get("output_cells") or {}).items()
    }

    if not dsc.get("enabled", True):
        return dict(static_fb), {**meta, "mode": "discovery_disabled_static_only"}

    r0 = max(1, int(dsc.get("row_min", 1)))
    r1 = max(r0, int(dsc.get("row_max", 300)))
    label_cols = [str(c).strip().upper() for c in (dsc.get("label_columns") or ["I", "J", "K", "L"]) if c]
    if not label_cols:
        label_cols = ["I", "J", "K", "L"]
    policy = str(dsc.get("match_policy") or "last").lower()
    rules: List[Dict[str, Any]] = list(dsc.get("rules") or [])

    blobs: Dict[int, str] = {}
    for r in range(r0, r1 + 1):
        blobs[r] = _label_blob(ws, r, label_cols).lower()

    out: Dict[str, str] = {}
    for i, rule in enumerate(rules):
        rx = rule.get("label_regex")
        path = str(rule.get("summary_path") or "").strip()
        vcols = [str(c).strip().upper() for c in (rule.get("value_columns") or ["M"]) if c]
        if not vcols:
            vcols = ["M"]
        if not rx or not path:
            meta["warnings"].append({"rule_index": i, "detail": "missing_regex_or_path"})
            continue
        try:
            cre = re.compile(str(rx), re.IGNORECASE | re.DOTALL)
        except re.error as e:
            meta["warnings"].append({"rule_index": i, "regex_error": str(e)})
            continue
        hits = [r for r in range(r0, r1 + 1) if cre.search(blobs[r])]
        if not hits:
            meta["warnings"].append({"rule_index": i, "path": path, "detail": "no_row_matched"})
            continue
        row_pick = hits[-1] if policy == "last" else hits[0]
        meta["rules_hits"].append(
            {"rule_index": i, "summary_path": path, "sheet_row": row_pick, "match_policy": policy}
        )
        for col in vcols:
            ref = f"{col}{row_pick}"
            prev = out.get(ref)
            if prev is not None and prev != path:
                meta["warnings"].append({"cell": ref, "conflict_paths": [prev, path]})
            out[ref] = path

    if dsc.get("merge_static_fallback", True) and static_fb:
        discovered_paths = set(out.values())
        by_path: Dict[str, List[str]] = defaultdict(list)
        for cell_ref, path in static_fb.items():
            by_path[str(path)].append(str(cell_ref).upper())
        for path, cells in by_path.items():
            if path in discovered_paths:
                continue
            for cell_ref in cells:
                prev = out.get(cell_ref)
                if prev is not None and prev != path:
                    meta["warnings"].append({"cell": cell_ref, "fallback_skipped_conflict": [prev, path]})
                    continue
                out[cell_ref] = path
                meta["rules_hits"].append(
                    {"summary_path": path, "cell": cell_ref, "source": "static_fallback_missing_path"}
                )

    if not out and dsc.get("use_static_when_empty", True):
        meta["warnings"].append({"detail": "discovery_empty_used_full_static_map"})
        return dict(static_fb), meta

    return out, meta


def discover_quantity_cell_ref(ws: Any) -> Tuple[str, Dict[str, Any]]:
    """
    Find the workbook cell that holds reference order quantity (parity D6 role).

    Returns (A1-style ref, metadata). On failure returns default from config or ``D6``.
    """
    qcfg = getattr(config, "ESTIMATE_QUANTITY_CELL_DISCOVERY", None) or {}
    default_cell = str(qcfg.get("default_cell") or "D6").upper()

    if not qcfg.get("enabled", True):
        return default_cell, {"mode": "disabled", "cell": default_cell}

    r0 = max(1, int(qcfg.get("row_min", 1)))
    r1 = max(r0, int(qcfg.get("row_max", 45)))
    label_cols = [str(c).strip().upper() for c in (qcfg.get("label_columns") or ["A", "B", "C", "D", "E", "F"]) if c]
    if not label_cols:
        label_cols = ["A", "B", "C", "D", "E", "F"]
    try:
        cre = re.compile(str(qcfg.get("label_regex") or r""), re.IGNORECASE | re.DOTALL)
    except re.error:
        return default_cell, {"mode": "regex_error", "cell": default_cell}

    pref_cols = [str(c).strip().upper() for c in (qcfg.get("value_column_preference") or ["D", "E", "G"]) if c]
    if not pref_cols:
        pref_cols = ["D", "E", "G"]

    meta: Dict[str, Any] = {"mode": "scan", "hits": []}
    for r in range(r0, r1 + 1):
        blob = _label_blob(ws, r, label_cols).lower()
        if not cre.search(blob):
            continue
        for col in pref_cols:
            v = _safe_float(ws.cell(row=r, column=_col_idx(col)).value)
            if v is not None and v > 0:
                ref = f"{col}{r}"
                meta["hits"].append({"row": r, "cell": ref, "value": v})
                return ref, meta

    meta["mode"] = "fallback"
    meta["cell"] = default_cell
    return default_cell, meta


def _labour_route_sdi_code_set() -> set:
    """SDI codes from config plus normaliser aliases (MANM, P/C, …)."""
    codes = {
        str(r.get("code", "")).strip().upper()
        for r in (getattr(config, "SDI_OPERATION_CODES", None) or [])
        if r.get("code")
    }
    try:
        from operation_normaliser import AI_TO_SDI

        for code_list in AI_TO_SDI.values():
            for c in code_list:
                s = str(c).strip().upper()
                if s:
                    codes.add(s)
    except ImportError:
        pass
    return codes


def _scan_column_for_sdi_codes(ws: Any, col_letter: str, codes: set, max_r: int) -> List[int]:
    col_i = _col_idx(col_letter)
    hits: List[int] = []
    for r in range(1, max_r + 1):
        v = ws.cell(row=r, column=col_i).value
        if v is None:
            continue
        s = str(v).strip().upper()
        if s in codes:
            hits.append(r)
    return hits


def discover_labour_route_row_span(ws: Any) -> Tuple[int, int, Dict[str, Any]]:
    """
    Find Estimate sheet rows holding SDI operation codes (LASM, FOLD, MANM, …).

    Scans multiple columns when the default column B holds quantity breaks (Route & BOM layouts).
    Returns (start_row, end_row, metadata). Falls back to fixed rows when no codes found.
    """
    cfg_p = getattr(config, "ESTIMATE_FULL_PARITY", None) or {}
    fb0 = int(cfg_p.get("labour_route_row_start") or 117)
    fb1 = int(cfg_p.get("labour_route_row_end") or 148)
    if not cfg_p.get("labour_route_discover", True):
        return fb0, fb1, {"mode": "disabled", "labour_route_row_start": fb0, "labour_route_row_end": fb1}

    codes = _labour_route_sdi_code_set()
    primary = str(cfg_p.get("labour_route_operation_column") or "B").strip().upper() or "B"
    configured = cfg_p.get("labour_route_scan_columns")
    if configured:
        columns = [str(c).strip().upper() for c in configured if c]
    else:
        columns = [primary, "C", "D", "E", "A", "I"]
    seen_cols: set = set()
    col_letters: List[str] = []
    for c in columns:
        if c and c not in seen_cols:
            seen_cols.add(c)
            col_letters.append(c)
    if primary not in seen_cols:
        col_letters.insert(0, primary)

    max_r = min(int(getattr(ws, "max_row", None) or 400), 500)
    best_hits: List[int] = []
    best_col = primary
    hits_per_column: Dict[str, int] = {}
    for col_letter in col_letters:
        hits = _scan_column_for_sdi_codes(ws, col_letter, codes, max_r)
        hits_per_column[col_letter] = len(hits)
        if len(hits) > len(best_hits):
            best_hits = hits
            best_col = col_letter

    if not best_hits:
        return fb0, fb1, {
            "mode": "fallback_no_sdi_codes_in_column",
            "columns_scanned": col_letters,
            "hits_per_column": hits_per_column,
            "labour_route_row_start": fb0,
            "labour_route_row_end": fb1,
        }
    pad = max(0, int(cfg_p.get("labour_route_pad_rows") or 2))
    r0, r1 = min(best_hits), max(best_hits)
    mode = "sdi_multi_column_scan" if len(col_letters) > 1 else "sdi_column_scan"
    return max(1, r0 - pad), r1 + pad, {
        "mode": mode,
        "operation_column": best_col,
        "columns_scanned": col_letters,
        "hits_per_column": hits_per_column,
        "code_rows": best_hits,
        "pad_rows": pad,
    }


def read_estimate_workbook_inputs(
    workbook_path: Path,
    *,
    sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Read Estimate-tab inputs used alongside geometry estimates: qty, manual £/tonne rates, etc.

    Uses the same label regexes as parity discovery. Only ``.xlsx`` / ``.xlsm`` (openpyxl); legacy ``.xls`` is skipped.
    """
    path = workbook_path.expanduser().resolve()
    out: Dict[str, Any] = {
        "workbook": str(path),
        "ok": False,
        "warnings": [],
    }
    if not path.is_file():
        out["warnings"].append("file_not_found")
        return out
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        out["warnings"].append("skipped_requires_xlsx_xlsm")
        return out
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        out["warnings"].append("openpyxl_missing")
        return out

    sn = sheet_name or str((getattr(config, "ESTIMATE_FULL_PARITY", None) or {}).get("estimate_sheet_name") or "Estimate")
    wb = load_workbook(path, data_only=True, rich_text=False)
    try:
        if sn not in wb.sheetnames:
            out["warnings"].append("sheet_not_found:%s" % sn)
            return out
        ws = wb[sn]
        qcell, qmeta = discover_quantity_cell_ref(ws)
        out["quantity_cell_a1"] = qcell
        out["quantity_discovery"] = qmeta
        try:
            qc, qr = _split_cell_a1(qcell)
            qval = _safe_float(ws.cell(row=qr, column=qc).value)
        except ValueError:
            qval = None
        if qval is not None and qval > 0:
            out["assumed_job_quantity"] = int(qval) if abs(qval - round(qval)) < 1e-6 else round(qval, 4)

        rcfg = getattr(config, "ESTIMATE_WORKBOOK_RATE_DISCOVERY", None) or {}
        rate_hits: List[Dict[str, Any]] = []
        if rcfg.get("enabled", True):
            r0 = max(1, int(rcfg.get("row_min", 1)))
            r1 = max(r0, int(rcfg.get("row_max", 35)))
            label_cols = [str(c).strip().upper() for c in (rcfg.get("label_columns") or ["I", "J", "K"]) if c]
            if not label_cols:
                label_cols = ["I", "J", "K"]
            policy = str(rcfg.get("match_policy") or "first").lower()
            for i, rule in enumerate(list(rcfg.get("rates") or [])):
                key = str(rule.get("key") or "").strip()
                rx = rule.get("label_regex")
                vcol = str(rule.get("value_column") or "L").strip().upper() or "L"
                if not key or not rx:
                    continue
                try:
                    cre = re.compile(str(rx), re.IGNORECASE | re.DOTALL)
                except re.error as e:
                    out["warnings"].append("rate_rule_%d_regex:%s" % (i, e))
                    continue
                blobs: Dict[int, str] = {}
                for r in range(r0, r1 + 1):
                    blobs[r] = _label_blob(ws, r, label_cols).lower()
                hits = [r for r in range(r0, r1 + 1) if cre.search(blobs[r])]
                if not hits:
                    out["warnings"].append("rate_rule_%d_no_match:%s" % (i, key))
                    continue
                row_pick = hits[-1] if policy == "last" else hits[0]
                v = _safe_float(ws.cell(row=row_pick, column=_col_idx(vcol)).value)
                rate_hits.append(
                    {"key": key, "sheet_row": row_pick, "value_column": vcol, "value": v, "rule_index": i}
                )
                if v is not None:
                    out[key] = round(float(v), 6) if key.endswith("_gbp") or "cost" in key else float(v)

        out["rate_discovery"] = {"hits": rate_hits, "config": "ESTIMATE_WORKBOOK_RATE_DISCOVERY"}
        out["ok"] = True
        return out
    finally:
        wb.close()


def mapping_to_money_specs(
    mapping: Dict[str, str],
    *,
    quantity_cell: str,
    quantity_path: str,
    extras: Optional[List[Dict[str, str]]] = None,
) -> List[Dict[str, str]]:
    """Build ordered parity/writeback-style specs: one entry per (cell, path)."""
    specs: Dict[Tuple[str, str], Dict[str, str]] = {}
    q_cell = str(quantity_cell).upper()
    specs[(q_cell, quantity_path)] = {
        "cell": q_cell,
        "summary_path": quantity_path,
        "label": "reference_order_qty_workbook",
    }
    for cell_ref, path in mapping.items():
        cu = str(cell_ref).upper()
        p = str(path)
        specs[(cu, p)] = {"cell": cu, "summary_path": p, "label": ""}
    for row in extras or []:
        c = str(row["cell"]).upper()
        specs[(c, str(row["summary_path"]))] = {
            "cell": c,
            "summary_path": str(row["summary_path"]),
            "label": str(row.get("label") or ""),
        }

    def sort_key(spec: Dict[str, str]) -> Tuple[int, int]:
        m = re.match(r"^([A-Za-z]{1,3})(\d+)$", spec["cell"])
        if not m:
            return (0, 0)
        col = _col_idx(m.group(1))
        return (int(m.group(2)), col)

    return sorted(specs.values(), key=sort_key)
