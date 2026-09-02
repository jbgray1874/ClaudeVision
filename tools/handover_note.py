#!/usr/bin/env python3
r"""Every row of an estimate, explained, in one document an estimator can read.

WHY THIS EXISTS. The deliverables already say where each number came from — the report's
"where the bill of materials came from" and "how each operation was decided", the workbook's
AI Provenance tab. What none of them says is WHICH FILE AND WHICH PAGE. The workbook has no
page, sheet or drawing column in any tab, so the one question estimating always asks —
"where did you see that?" — cannot be answered from the pack that was sent, and gets asked
by email instead.

The page numbers are not missing from the engine. Every part record carries `pages` and
`page_roles`, and the run log says so out loud: "39/39 row(s) traced to a sheet; 39 carry
the drawing that owns them". They are simply dropped before anything is written out. So this
joins the workbook (the money, the materials, the routes) to the scan JSON (the pages) on the
part number, and prints one section per question an estimator actually asks.

HAND-BUILT ONCE, THEN GENERATED. This is a tool, not a template to fill in, because a
document typed out by hand is accurate for one run and stale for the next — and the numbers
on this job have moved between runs on identical drawings. What it prints is also the
specification for the report sections that will replace it: file and page on every BOM field
and every route row, and the price source beside every figure.

    python tools/handover_note.py --workbook <estimate.xlsx> --scan-json <12552-00.json>

The scan JSON is optional. Without it every other column still prints and the page column
says so, rather than the tool refusing to run — an estimator with most of the answer is
better off than one with none of it.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
except ImportError:                                              # pragma: no cover
    raise SystemExit("openpyxl is required: pip install openpyxl")


# ── reading the workbook ─────────────────────────────────────────────────────

def _rows(ws, header_row: int = 1) -> List[Dict[str, Any]]:
    """A sheet as dicts keyed by its own header, blank rows dropped."""
    header = [str(ws.cell(header_row, c).value or "").strip()
              for c in range(1, ws.max_column + 1)]
    out: List[Dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(v not in (None, "") for v in values):
            continue
        out.append({h: v for h, v in zip(header, values) if h})
    return out


def _sheet(wb, name: str) -> List[Dict[str, Any]]:
    return _rows(wb[name]) if name in wb.sheetnames else []


# The Sheet Steel block's own columns, by position, from the template.
_STEEL_COLS = {"desc": 3, "qty": 5, "length": 6, "width": 7, "gauge": 8,
               "sheet_l": 9, "sheet_w": 10, "per_sheet": 11, "scrap": 12,
               "cost_per_part": 13, "internal_cut": 20, "rate_per_hour": 23}


def _steel_rows(wb) -> Dict[str, Dict[str, Any]]:
    """The Sheet Steel block, keyed by part number.

    WHERE A FABRICATED PART'S MONEY ACTUALLY IS. A "-M" line shows a dash in the BOM price
    column and says why in its own text — "costed in Sheet Steel below" — because sheet metal
    is priced from blank area, not per piece. An explanation that stops at "priced below"
    ends exactly where the estimator's question begins: how big, off what sheet, how many out
    of it, at what cost. This reads that block so the two halves can be shown together.

    Found by its header text, not a row number: the block moves with the size of the BOM
    above it, and a fixed row would silently read whatever had shifted into it.
    """
    if "Estimate" not in wb.sheetnames:
        return {}
    ws = wb["Estimate"]
    header_row = None
    for r in range(1, min(ws.max_row, 120) + 1):
        joined = " ".join(str(ws.cell(r, c).value or "") for c in range(1, 26)).lower()
        if "part length" in joined and "gauge" in joined and "part description" in joined:
            header_row = r
            break
    if header_row is None:
        return {}
    out: Dict[str, Dict[str, Any]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        text = str(ws.cell(r, _STEEL_COLS["desc"]).value or "").strip()
        if not text:
            # Blank template rows sit between the blocks; stop once we have started and
            # then hit two in a row, rather than running on into "Other Sheet Material".
            if out and not str(ws.cell(r + 1, _STEEL_COLS["desc"]).value or "").strip():
                break
            continue
        code = text.split()[0].strip().upper()
        out[code] = {"row": r, "text": text,
                     **{k: ws.cell(r, c).value for k, c in _STEEL_COLS.items() if k != "desc"}}
    return out


_TOTAL_LABELS = (("material", "total material cost"),
                 ("labour", "total labour cost"),
                 ("unit", "total unit cost"))


def _sheet_totals(wb) -> Dict[str, Optional[float]]:
    """The Estimate sheet's own labelled totals, read from the cells.

    Found by label rather than by cell reference, and the value taken as the first number to
    the right of it — the template's totals sit in column M today and a reference is the one
    thing that cannot survive a template revision. A total whose cell holds a formula with no
    cached result reads as None and the document says the figure was not available, rather
    than reporting a zero that would make every reconciliation look perfect.
    """
    out: Dict[str, Optional[float]] = {k: None for k, _ in _TOTAL_LABELS}
    if "Estimate" not in wb.sheetnames:
        return out
    ws = wb["Estimate"]
    for r in range(1, ws.max_row + 1):
        label = " ".join(str(ws.cell(r, c).value or "") for c in range(1, 8)).strip().lower()
        if not label:
            continue
        for key, needle in _TOTAL_LABELS:
            if out[key] is None and label.startswith(needle):
                for c in range(8, min(ws.max_column, 30) + 1):
                    value = _money(ws.cell(r, c).value)
                    if value is not None:
                        out[key] = round(value, 2)
                        break
    return out


def _estimate_bom(wb) -> List[Dict[str, Any]]:
    """The Bill of Materials block on the Estimate sheet.

    Found by its own header text rather than a row number: the block moves whenever a job
    has more or fewer lines, and a hard-coded row would silently read the wrong band.
    """
    if "Estimate" not in wb.sheetnames:
        return []
    ws = wb["Estimate"]
    header_row = None
    for r in range(1, min(ws.max_row, 60) + 1):
        joined = " ".join(str(ws.cell(r, c).value or "") for c in range(1, 16)).lower()
        if "bill of materials" in joined and "part code" in joined:
            header_row = r
            break
    if header_row is None:
        return []
    out = []
    for r in range(header_row + 1, ws.max_row + 1):
        code = ws.cell(r, 8).value           # "Part code"
        text = ws.cell(r, 3).value           # the description column
        if not code and not text:
            if out:
                break                        # the block has ended
            continue
        out.append({
            "code": str(code or "").strip(),
            "text": str(text or "").strip(),
            "supplier": ws.cell(r, 9).value,
            "price": ws.cell(r, 10).value,
            "qty": ws.cell(r, 11).value,
        })
    return out


# ── reading the scan ─────────────────────────────────────────────────────────

def _load_scan(path: Optional[Path]) -> Any:
    """The run JSON as it sits on disk, or None. Read once and passed around.

    SEPARATED FROM _scan_parts BECAUSE THE PARTS ARE NOT THE ONLY THING IN IT. The same
    document carries `final_estimate` — the rows as Excel calculated them, which is the only
    record that has money on every line and provably sums to the sheet's own totals. Reading
    the file twice to get at both would be two chances to read two different files.
    """
    if not path or not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8", errors="replace"))
    except Exception:                                            # noqa: BLE001
        return None


def _scan_parts(data: Any) -> Dict[str, Dict[str, Any]]:
    """part number -> the scan record, for the pages the workbook does not carry."""
    if data is None:
        return {}
    # THREE SHAPES, BECAUSE THE FULL SCAN IS TOO BIG TO MOVE. 12552's scan JSON is 258,935
    # lines; nobody is going to send that to have three fields read out of it. So a trimmed
    # extract — a bare list of records, or {"parts": [...]} — is accepted alongside the real
    # thing, and only part_number, pages and page_roles are ever read from any of them.
    #
    # This is a workaround for the actual defect, which is that no page number reaches any
    # deliverable. Once the page is written onto the row, this whole path becomes unnecessary.
    if isinstance(data, list):
        parts = data
    elif isinstance(data, dict):
        parts = ((data.get("manufacturing_writeup") or {}).get("parts")
                 or data.get("parts") or [])
    else:
        return {}
    return {str(p.get("part_number") or "").strip().upper(): p
            for p in parts if isinstance(p, dict)}


def _pages_of(record: Dict[str, Any]) -> str:
    pages = record.get("pages") or []
    roles = record.get("page_roles") or []
    if not pages:
        return "no sheet of its own"
    shown = ", ".join(f"p.{p}" for p in pages)
    return f"{shown} ({', '.join(str(r) for r in roles)})" if roles else shown


# ── the rows as Excel calculated them ────────────────────────────────────────
#
# THE ONE RECORD WITH MONEY ON EVERY LINE. `final_estimate` is written by the read-back after
# Excel has recalculated the populated template: every material block (bought-in BOM, tube,
# sheet steel, other sheet) and every labour row, each carrying the value the sheet itself
# computed. It is the only structure that can be summed and held against Total Material Cost
# and Total Labour Cost — which is what "relates to the spreadsheet" has to mean.
#
# Until this was read, this document explained material only, and the labour — a third of the
# unit cost — appeared as a list of operations with no figure against any of them.

def _final_estimate(data: Any) -> Dict[str, Any]:
    """`final_estimate` off the run JSON, wherever the writer put it."""
    if not isinstance(data, dict):
        return {}
    for node in (data.get("final_estimate"),
                 (data.get("estimate_summary") or {}).get("final_estimate")
                 if isinstance(data.get("estimate_summary"), dict) else None):
        if isinstance(node, dict):
            return node
    return {}


def _accepted_labour(data: Any) -> Dict[int, Dict[str, Any]]:
    """workbook_labour rows keyed by sheet row — the PARTS behind each labour line.

    The calculated rows know what a line cost and which department did it; they do not know
    which parts produced it. The accepted rows know exactly that and nothing about cost. They
    join on the sheet row they share, which is the join wb_populate itself documents.
    """
    if not isinstance(data, dict):
        return {}
    node = data.get("workbook_labour")
    if not isinstance(node, dict) and isinstance(data.get("estimate_summary"), dict):
        node = data["estimate_summary"].get("workbook_labour")
    out: Dict[int, Dict[str, Any]] = {}
    for row in ((node or {}).get("rows") or []):
        if isinstance(row, dict) and row.get("workbook_row") is not None:
            try:
                out[int(row["workbook_row"])] = row
            except (TypeError, ValueError):
                continue
    return out


def _money(value: Any) -> Optional[float]:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _sum_money(rows: List[Dict[str, Any]], key: str = "total_value_gbp") -> float:
    return round(sum(_money(r.get(key)) or 0.0 for r in rows), 2)


_BLOCK_NAMES = {"bom": "Bought-in / standard materials", "tube": "Tube and wire",
                "steel": "Sheet steel", "other_sheet": "Other sheet material"}


# ── the money's provenance ───────────────────────────────────────────────────

_INDICATIVE = ("grok", "llm", "xai", "indicative", "market")


def _price_source(bom_row: Dict[str, Any], provenance: Dict[str, Dict[str, Any]],
                  steel_index: Optional[Dict[str, Dict[str, Any]]] = None) -> str:
    """Which book priced this line, in words an estimator can act on.

    THE FABRICATED PARTS ARE NOT UNPRICED. Every "-M" line carries a blank in the BOM's
    price column and says so in its own text — "costed in Sheet Steel below" — because sheet
    metal is priced from blank area on the Sheet Steel block, not per piece here. Reading the
    blank as "no rate" put fifteen made parts on the estimator's to-do list, which is both
    wrong and the fastest way to lose their trust in the rest of the document.

    A BLANK WITH NO EXPLANATION IS THE ONE THAT MATTERS, and it stays loud.
    """
    code = bom_row["code"].upper()
    text = str(bom_row.get("text") or "")
    supplier = str(bom_row.get("supplier") or "").strip()
    prov = provenance.get(code) or {}
    named = str(prov.get("Price Source") or prov.get("price_source") or "").strip()

    if "costed in sheet steel" in text.lower():
        # NAME THE ROW IT IS COSTED ON. "priced below" is true and stops exactly where the
        # question starts; an estimator wants to look at the figure, not be told it exists.
        steel_row = (steel_index or {}).get(code, {}).get("row")
        return (f"by blank area — see `Estimate!{steel_row}` below"
                if steel_row else
                "by blank area on the Sheet Steel block, not per piece")
    if bom_row.get("price") in (None, ""):
        return "**NOT PRICED — needs a rate**"
    if any(token in supplier.lower() for token in _INDICATIVE):
        return f"AI market indication ({supplier}) — NOT A QUOTE, replace it"
    if any(token in named.lower() for token in _INDICATIVE):
        return f"AI market indication ({named}) — NOT A QUOTE, replace it"
    if supplier:
        return f"catalogue — {supplier}"
    if named:
        return named
    # Said plainly rather than guessed at. "config rate card" was asserted here for anything
    # unlabelled, which claimed a source for the bearing's GBP 1.42 that nothing in the
    # workbook actually states.
    return "source not named in the workbook — check AI Provenance"


# ── the document ─────────────────────────────────────────────────────────────

def _description(bom_row: Dict[str, Any]) -> str:
    """What the part IS, with the workbook's warnings stripped off the end.

    The Estimate sheet's description cell carries the part code, then the description, then
    any notice appended to it — "[AI ESTIMATE - INDICATIVE, NOT A QUOTE]", "MATERIAL
    UNPRICED: enter a unit rate for this item". Taking the LAST segment kept the notice and
    threw the description away, so the concrete slab appeared as "[AI ESTIMATE - INDICATIVE,
    NOT A QUOTE]" and the nylon washer as "MATERIAL UNPRICED" — a table naming no parts.
    Those notices are already carried, properly, by the price column beside it.
    """
    text = str(bom_row.get("text") or "")
    code = str(bom_row.get("code") or "")
    if code and text.upper().startswith(code.upper()):
        text = text[len(code):]
    for marker in ("[AI ESTIMATE", "—  MATERIAL UNPRICED", "— MATERIAL UNPRICED",
                   "MATERIAL UNPRICED", "— costed in Sheet Steel", "— costed in sheet steel"):
        cut = text.find(marker)
        if cut > 0:
            text = text[:cut]
    text = text.strip(" —-—\t")
    return (text[:60] or "—")


def _gbp(value: Any) -> str:
    """A number as money, or a plain statement that the figure was not available.

    NEVER A ZERO. A total whose cell holds an uncached formula reads as None, and printing
    that as £0.00 would make a reconciliation against it look exact.
    """
    number = _money(value)
    return f"£{number:,.2f}" if number is not None else "not readable from the sheet"


def _gbp_or(value: Any, dash: str) -> str:
    """Money where there is a number, and the caller's own words where there is not."""
    return _gbp(value) if _money(value) is not None else dash


def _diff(here: float, sheet: Optional[float]) -> str:
    if sheet is None:
        return "cannot be checked"
    gap = round(here - sheet, 2)
    if abs(gap) < 0.01:
        return "**none — it reconciles**"
    # SAID AS A DIRECTION, NOT A SIGNED NUMBER. "£-193.30" makes the reader work out which way
    # round it is, and the two directions mean opposite things: short means lines are missing
    # from this document, over means it is counting something the sheet is not.
    if gap < 0:
        return f"**{_gbp(-gap)} short** — lines this document cannot see"
    return f"**{_gbp(gap)} more** than the sheet — this document is counting something twice"


def _fe_totals(final: Dict[str, Any]) -> Dict[str, Any]:
    node = final.get("totals")
    return node if isinstance(node, dict) else {}


def _measured_sentence(material: Dict[str, Dict[str, Any]]) -> str:
    """How much of the geometry was measured off a model rather than reasoned from a view."""
    sources: Dict[str, int] = {}
    for row in material.values():
        key = str(row.get("Geom source") or "not recorded").strip() or "not recorded"
        sources[key] = sources.get(key, 0) + 1
    if not sources:
        return "The workbook carries no geometry-source column to answer this from."
    ranked = sorted(sources.items(), key=lambda kv: (-kv[1], kv[0]))
    return ("Of " + str(sum(sources.values())) + " part(s) with a material row: "
            + ", ".join(f"{n} {name}" for name, n in ranked) + ".")


def _blocks_sentence(material_rows: List[Dict[str, Any]],
                     labour_rows: List[Dict[str, Any]]) -> str:
    if not material_rows and not labour_rows:
        return "No calculated rows were supplied, so this cannot be answered."
    parts = []
    for block in ("bom", "tube", "steel", "other_sheet"):
        rows = [r for r in material_rows if r.get("block") == block]
        if rows:
            parts.append(f"{_BLOCK_NAMES[block].lower()} {_gbp(_sum_money(rows))}")
    if labour_rows:
        parts.append(f"labour {_gbp(_sum_money(labour_rows))} across "
                     f"{len(labour_rows)} sheet row(s)")
    return "; ".join(parts) + "."


def _reconciles_sentence(material_rows: List[Dict[str, Any]],
                         labour_rows: List[Dict[str, Any]],
                         totals: Dict[str, Optional[float]]) -> str:
    if not material_rows and not labour_rows:
        return ("Not yet — no calculated rows were supplied, so nothing below can be summed "
                "against the sheet.")
    gaps = []
    for label, rows, total in (("material", material_rows, totals.get("material")),
                               ("labour", labour_rows, totals.get("labour"))):
        if total is None:
            gaps.append(f"the sheet's {label} total could not be read")
            continue
        gap = round(_sum_money(rows) - total, 2)
        if abs(gap) >= 0.01:
            gaps.append(f"{label} is {_gbp(abs(gap))} "
                        + ("short" if gap < 0 else "over"))
    if not gaps:
        return ("Yes. Every material and labour line below sums to the Estimate sheet's own "
                "Total Material Cost and Total Labour Cost.")
    return "Not entirely — " + "; ".join(gaps) + ". The table below says where."


def _fmt(value: Any, dash: str = "—") -> str:
    if value in (None, ""):
        return dash
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)


def build(workbook: Path, scan_json: Optional[Path]) -> str:
    wb = openpyxl.load_workbook(workbook, data_only=True)
    scan_doc = _load_scan(scan_json)
    scan = _scan_parts(scan_doc)
    final = _final_estimate(scan_doc)
    accepted = _accepted_labour(scan_doc)
    labour_rows = [r for r in (final.get("labour_rows") or []) if isinstance(r, dict)]
    material_rows = [r for r in (final.get("material_rows") or []) if isinstance(r, dict)]
    totals = _sheet_totals(wb)
    steel = _steel_rows(wb)
    material = {str(r.get("Part") or "").upper(): r for r in _sheet(wb, "AI Material Detail")}
    provenance = {str(r.get("Part") or "").upper(): r for r in _sheet(wb, "AI Price Provenance")}
    routes = _sheet(wb, "Canonical Route")
    bom = _estimate_bom(wb)

    lines: List[str] = []
    add = lines.append

    add(f"# {workbook.stem}")
    add("")
    if not scan:
        add("> **Page numbers are unavailable** — no scan JSON was supplied, so the "
            "\"which sheet\" column reads *not supplied* throughout. Re-run with "
            "`--scan-json output/json/<job>.json` to fill it.")
        add("")

    # ── the questions, answered before the tables ────────────────────────────
    # ASKED AND ANSWERED IN THAT ORDER. A document that opens with four hundred rows makes
    # the reader derive the five facts they came for. These are the five, each computed from
    # the same rows the tables below print — never typed, so they cannot drift from them.
    _unpriced = [r for r in bom
                 if r.get("price") in (None, "")
                 and "costed in sheet steel" not in str(r.get("text") or "").lower()]
    _indicative = [r for r in bom
                   if any(t in f"{r.get('supplier') or ''}".lower() for t in _INDICATIVE)
                   and _money(r.get("price"))]
    _indicative_gbp = round(sum((_money(r.get("price")) or 0) * (_money(r.get("qty")) or 0)
                                for r in _indicative), 2)
    add("## The questions, answered first")
    add("")
    add(f"- **What does a unit cost, and of what?** "
        f"{_gbp(totals['unit'])} — material {_gbp(totals['material'])} + labour "
        f"{_gbp(totals['labour'])}"
        + (f", and the unit cell adds {_gbp(_money(_fe_totals(final).get('other_gbp')))} "
           f"({final.get('unit_price_composition', {}).get('basis') or 'per the sheet'})"
           if _fe_totals(final).get("other_gbp") else "")
        + ".")
    add(f"- **What must be replaced before this is a quote?** "
        + (f"{len(_unpriced)} line(s) carry no price at all — "
           + ", ".join(r["code"] or _description(r) for r in _unpriced[:8])
           + ("…" if len(_unpriced) > 8 else "") + ". "
           if _unpriced else "No line is unpriced. ")
        + (f"{len(_indicative)} line(s) worth {_gbp(_indicative_gbp)} are AI market "
           f"indications, not catalogue prices: "
           + ", ".join(r["code"] or _description(r) for r in _indicative[:8]) + "."
           if _indicative else "No line rests on an AI market indication."))
    add(f"- **How much of this was measured rather than reasoned?** "
        + _measured_sentence(material))
    add(f"- **Where is the money?** " + _blocks_sentence(material_rows, labour_rows))
    add(f"- **Does this document add up to the sheet?** "
        + _reconciles_sentence(material_rows, labour_rows, totals))
    add("")

    # ── does it reconcile ────────────────────────────────────────────────────
    # THE CLAIM THAT MAKES THE REST OF THE DOCUMENT WORTH READING, AND THE ONE THAT CAN BE
    # CHECKED. Every line printed below is summed and held against the Estimate sheet's own
    # labelled total. Where the two differ the difference is stated in pounds — an
    # explanation that quietly covers 81% of a total is worse than one that says which 19% it
    # cannot see, because only the second sends anybody looking.
    if material_rows or labour_rows:
        add("## This document against the sheet")
        add("")
        add("| Block | Lines here | £ here | The sheet's own total | Difference |")
        add("|---|---|---|---|---|")
        for block in ("bom", "tube", "steel", "other_sheet"):
            rows_in = [r for r in material_rows if r.get("block") == block]
            if rows_in:
                add(f"| {_BLOCK_NAMES[block]} | {len(rows_in)} | "
                    f"{_gbp(_sum_money(rows_in))} | — | — |")
        add(f"| **All material** | {len(material_rows)} | {_gbp(_sum_money(material_rows))} "
            f"| {_gbp(totals['material'])} | {_diff(_sum_money(material_rows), totals['material'])} |")
        add(f"| **All labour** | {len(labour_rows)} | {_gbp(_sum_money(labour_rows))} "
            f"| {_gbp(totals['labour'])} | {_diff(_sum_money(labour_rows), totals['labour'])} |")
        add("")
        add("> Every figure in the £ column is the value the Estimate sheet itself computed "
            "for that row, read back after Excel recalculated. Nothing here is re-derived, "
            "so a difference is a row this document cannot see — not a rounding argument.")
        add("")
        for problem in (final.get("adapter_problems") or []):
            if isinstance(problem, dict) and problem.get("message"):
                add(f"> **A block was not read:** {problem['message']}")
                add("")
    elif scan_doc is not None:
        add("## This document against the sheet")
        add("")
        add("> **Not produced.** The run JSON supplied carries no `final_estimate`, which is "
            "the record of the rows as Excel calculated them. Without it the lines below can "
            "be listed but not summed against the sheet's own totals, so this document "
            "cannot claim to be complete. Re-run with the full "
            "`output/json/<job>.json` from a run whose workbook was read back.")
        add("")

    # ── every BOM line ───────────────────────────────────────────────────────
    # TWO ROWS FOR A PART SDI CUTS, because it has two halves on the sheet and an estimator
    # asked for both: the BOM row that shows a dash and points below, and the Sheet Steel row
    # that holds the blank, the gauge and the money. The page is stamped on both — the same
    # drawing owns both halves, and leaving the second one blank is what made the join look
    # like a missing answer.
    add("## Every line on the Bill of Materials")
    add("")
    add("| Part | Description | Qty | Unit £ | Ext £ | Where the £ came from | Material | "
        "Gauge | Blank | Which sheet |")
    add("|---|---|---|---|---|---|---|---|---|---|")
    for row in bom:
        code = row["code"].upper()
        mat = material.get(code, {})
        rec = scan.get(code, {})
        blank_l, blank_w = mat.get("Blank L"), mat.get("Blank W")
        blank = (f"{_fmt(blank_l)} × {_fmt(blank_w)}"
                 if blank_l or blank_w else "not a cut part")
        page = _pages_of(rec) if rec else "not supplied"
        _unit, _qty = _money(row.get("price")), _money(row.get("qty"))
        add(f"| {row['code'] or '—'} "
            f"| {_description(row)} "
            f"| {_fmt(row.get('qty'))} "
            f"| {_fmt(row.get('price'))} "
            f"| {_gbp(round(_unit * _qty, 2)) if _unit and _qty else '—'} "
            f"| {_price_source(row, provenance, steel)} "
            f"| {_fmt(mat.get('Material'))} "
            f"| {_fmt(mat.get('Gauge'), 'none — not sheet')} "
            f"| {blank} "
            f"| {page} |")
        steel_row = steel.get(code)
        if steel_row:
            per_part = _gbp_or(mat.get("Cost/Part"), "not resolved")
            extended = _gbp_or(mat.get("Ext Material"), "not resolved")
            add(f"| ↳ `Estimate!{steel_row['row']}` "
                f"| the same part, on the Sheet Steel block "
                f"| {_fmt(steel_row.get('qty'))} "
                f"| {per_part} "
                f"| {extended} "
                f"| off a {_fmt(steel_row.get('sheet_l'))} × {_fmt(steel_row.get('sheet_w'))} "
                f"sheet, {_fmt(steel_row.get('scrap'))} scrap "
                f"| {_fmt(mat.get('Material'))} "
                f"| {_fmt(steel_row.get('gauge'), 'none — not sheet')} "
                f"| {_fmt(steel_row.get('length'))} × {_fmt(steel_row.get('width'))} "
                f"| {page} |")
    add("")

    # ── where the fabricated money actually is ───────────────────────────────
    if steel:
        add("## The fabricated parts, priced by blank area")
        add("")
        add("Each of these shows a dash in the BOM price column above. That is not a missing "
            "rate — sheet metal is costed from its blank on this block, and pricing it in "
            "both places would double it. This is the other half of those lines.")
        add("")
        add("| Part | Blank L × W | Gauge | Off a sheet | Scrap | Cost/part | Qty | "
            "Extended | Sheet row |")
        add("|---|---|---|---|---|---|---|---|---|")
        for code, row in steel.items():
            # COST FROM THE RESOLVED FIGURE, NOT THE FORMULA CELL. The Sheet Steel block's
            # cost column is an Excel formula, and a workbook written by the engine and never
            # opened has no cached result — so reading that cell alone gives nothing and the
            # table says "computes in Excel" on the one column an estimator came for. The
            # engine's own resolved figure for the same part is on AI Material Detail, which
            # is a value, not a formula. Still read, never recalculated here.
            mat = material.get(code, {})
            add(f"| {code} "
                f"| {_fmt(row.get('length'))} × {_fmt(row.get('width'))} "
                f"| {_fmt(row.get('gauge'))} "
                f"| {_fmt(row.get('sheet_l'))} × {_fmt(row.get('sheet_w'))} "
                f"| {_fmt(row.get('scrap'))} "
                f"| {_fmt(mat.get('Cost/Part'), 'not resolved')} "
                f"| {_fmt(row.get('qty'))} "
                f"| {_fmt(mat.get('Ext Material'), 'not resolved')} "
                f"| `Estimate!{row['row']}` |")
        add("")
        add("> Cost per part and extended are the engine's own resolved figures, read from "
            "the AI Material Detail tab. The Sheet Steel block holds them as Excel formulas, "
            "which have no value until the workbook is opened — nothing here is "
            "recalculated, so this cannot disagree with the sheet.")
        add("")

    # ── every labour line, with the money on it ──────────────────────────────
    # A THIRD OF THE UNIT COST, PREVIOUSLY UNEXPLAINED. This document listed every operation
    # and what decided it, and put no figure against any of them — so an estimator could see
    # that 12552 folds and welds, and could not see that the folding costs £41 until they
    # opened the sheet and found the row themselves. The route section below answers "why is
    # this operation here"; this one answers "what is it charging, and to which parts".
    if labour_rows:
        add("## Every labour line, and what it charges")
        add("")
        add("Sorted by cost. **Batch hours** are for the whole order, not one unit; **£ this "
            "line** is the sheet's own Total Value for the row. Parts come from the accepted "
            "route grouping, joined to the calculated row on the sheet row they share.")
        add("")
        add("| Sheet row | Operation | Dept | Parts | Qty/unit | Set-up (min) | Batch hours | "
            "Rate £/hr | £ this line | Rate basis |")
        add("|---|---|---|---|---|---|---|---|---|---|")
        for row in sorted(labour_rows, key=lambda r: -(_money(r.get("total_value_gbp")) or 0)):
            acc = accepted.get(int(_money(row.get("workbook_row")) or 0), {})
            parts = [str(p) for p in (acc.get("part_numbers") or []) if p]
            add(f"| `Estimate!{_fmt(row.get('workbook_row'))}` "
                f"| {_fmt(row.get('operation'))} "
                f"| {_fmt(row.get('department'))} "
                f"| {', '.join(parts) if parts else 'not recorded on the accepted row'} "
                f"| {_fmt(row.get('qty_per_unit'))} "
                f"| {_fmt(row.get('setup_minutes'))} "
                f"| {_fmt(row.get('batch_hours'))} "
                f"| {_fmt(row.get('dept_rate_gbp_per_hour'))} "
                f"| {_gbp(row.get('total_value_gbp'))} "
                f"| {_fmt(acc.get('rate_basis'), 'not recorded')} |")
        add("")
        add(f"> {len(labour_rows)} row(s), {_gbp(_sum_money(labour_rows))} — "
            f"against the sheet's Total Labour Cost of {_gbp(totals['labour'])}.")
        add("")

    # ── every route line ─────────────────────────────────────────────────────
    add("## Every operation, and who decided it")
    add("")
    add("| Part | Operation | Seq | Scope | Qty | Decided by | On what basis | Which sheet |")
    add("|---|---|---|---|---|---|---|---|")
    for row in routes:
        target = str(row.get("Target") or "").upper()
        rec = scan.get(target, {})
        add(f"| {_fmt(row.get('Target'))} "
            f"| {_fmt(row.get('Operation'))} "
            f"| {_fmt(row.get('Seq'))} "
            f"| {_fmt(row.get('Scope'))} "
            f"| {_fmt(row.get('Qty/unit'))} "
            f"| {_fmt(row.get('Source'))} "
            f"| {str(row.get('Reason') or '—')[:70]} "
            f"| {_pages_of(rec) if rec else 'not supplied'} |")
    add("")

    # ── what each sheet could be read for ────────────────────────────────────
    # A FIELD MISSING FROM THIS EXTRACT IS NOT A FIELD MISSING FROM THE DRAWING.
    #
    # This section reports what each sheet could be read for, and it does that by looking for
    # materials / thicknesses_mm / surface_finishes on the record. The trimmed extract that
    # supplies the page numbers carries only part_number, pages and page_roles — so against
    # one of those every row came out "material **no**, thickness **no**, finish **no**",
    # for a pack whose page 4 plainly reads "MATERIAL: MILD STEEL", "1.5 THK", "POWDER
    # COATED". That is not a weak answer, it is a confident wrong one, and it would have gone
    # to an estimator as a drawing-quality assessment.
    #
    # So the section is built only from an extract that actually carries the fields, and
    # otherwise says what it needs. Refusing to answer is the honest failure mode; the whole
    # document exists to stop people guessing from it.
    _quality_fields = ("materials", "thicknesses_mm", "surface_finishes", "geometry_rollup")
    _has_quality = any(any(k in rec for k in _quality_fields) for rec in scan.values())
    if scan and not _has_quality:
        add("## Drawing quality, sheet by sheet")
        add("")
        add("> **Not produced.** This ran against a trimmed extract carrying only part "
            "numbers and page numbers. Reporting from it would have said every drawing "
            "states no material, no thickness and no finish — which is false: page 4 alone "
            "reads *MATERIAL: MILD STEEL*, *1.5 THK*, *POWDER COATED*. Re-run with the full "
            "`output/json/<job>.json` and this section builds itself.")
        add("")
    if scan and _has_quality:
        add("## Drawing quality, sheet by sheet")
        add("")
        add("| Sheet | Part | Material stated | Thickness stated | Finish stated | "
            "Geometry | What it could not give |")
        add("|---|---|---|---|---|---|---|")
        for code, rec in sorted(scan.items(),
                               key=lambda kv: ((kv[1].get("pages") or [999])[0], kv[0])):
            # ONLY THINGS THAT ARE DRAWINGS. A line with no page is a catalogue item, a
            # commercial line, or a fastener off a BOM table — PACKAGING, DELIVERY, FIXING17.
            # Graded here they came out "could not give: thickness, finish, cut length",
            # which reads as a deficient drawing pack and is nonsense: packaging is not a
            # drawing, and an M8 nyloc nut has no thickness because it is bought, not because
            # somebody forgot to write one. Same false-negative as grading a pack from a
            # trimmed extract, in a different place.
            if not (rec.get("pages") or []):
                continue
            geom = rec.get("geometry_rollup") or {}
            reliability = ((geom.get("confidence") or {}).get("geometry_reliability"))
            # A PURCHASED PART NAMED ON AN ASSEMBLY SHEET IS NOT AN INCOMPLETE DETAIL.
            # The bearing and the rivets appear on p.2 because that is where they are listed,
            # and they have no detail drawing because they do not need one. Saying what they
            # ARE beats listing four fields they were never going to carry.
            _roles = [str(r).lower() for r in (rec.get("page_roles") or [])]
            if "bought_in" in _roles and "detail" not in _roles:
                missing = ["bought in — listed on an assembly sheet, no detail drawing needed"]
            else:
                missing = []
                if not rec.get("materials"):
                    missing.append("material")
                if not rec.get("thicknesses_mm"):
                    missing.append("thickness")
                if not rec.get("surface_finishes"):
                    missing.append("finish")
                if not (geom.get("estimated_cut_length_mm") or 0):
                    missing.append("cut length")
            add(f"| {_pages_of(rec)} | {code or _fmt(rec.get('description'))} "
                f"| {', '.join(str(m) for m in rec.get('materials') or []) or '**no**'} "
                f"| {', '.join(str(t) for t in rec.get('thicknesses_mm') or []) or '**no**'} "
                f"| {', '.join(str(f) for f in rec.get('surface_finishes') or []) or '**no**'} "
                f"| {_fmt(rec.get('geometry_source'))}"
                f"{f' ({reliability:.0%})' if isinstance(reliability, (int, float)) else ''} "
                f"| {', '.join(missing) or 'nothing — complete'} |")
        add("")

    # ── what the pack does not contain, and what it costs ────────────────────
    # ASKED DIRECTLY AND ANSWERED DIRECTLY. "Which drawings are missing, and does it hurt the
    # price?" is a different question from "how good are the drawings we have", and the
    # answer is usually "none of them, for money" — a bought-in bolt has no detail drawing
    # because it does not need one, and saying so stops somebody chasing Design for it.
    #
    # A LINE WITH NO SHEET IS ONLY A PROBLEM IF ITS PRICE DEPENDED ON ONE. So each is
    # classified by what its money actually rests on, not by the absence alone.
    if scan:
        no_sheet = []
        for row in bom:
            code = row["code"].upper()
            if not code:
                continue
            rec = scan.get(code) or {}
            if rec.get("pages"):
                continue
            cut_here = bool(steel.get(code)) or bool((material.get(code) or {}).get("Blank L"))
            unit, qty = _money(row.get("price")), _money(row.get("qty"))
            no_sheet.append({
                "code": row["code"], "desc": _description(row), "cut": cut_here,
                "gbp": round(unit * qty, 2) if unit and qty else None,
                "priced": row.get("price") not in (None, ""),
            })
        add("## Drawings the pack does not contain, and what that costs")
        add("")
        if not no_sheet:
            add("Every costed line on this job is owned by a sheet in the pack. Nothing is "
                "priced off a drawing that was not supplied.")
            add("")
        else:
            _at_risk = [n for n in no_sheet if n["cut"]]
            add(f"{len(no_sheet)} costed line(s) have no sheet of their own in this pack. "
                + (f"**{len(_at_risk)} of them are parts SDI cuts** — those are the ones "
                   f"where a missing drawing moves the price."
                   if _at_risk else
                   "**None of them is a part SDI cuts**, so no missing drawing is holding up "
                   "a fabricated cost: each is a bought item, a fastener off a BOM table, or "
                   "a commercial line, and none of them would have a detail drawing even on "
                   "a complete pack."))
            add("")
            add("| Line | What it is | £ on this job | Does the missing sheet move the £ |")
            add("|---|---|---|---|")
            for item in sorted(no_sheet, key=lambda n: -(n["gbp"] or 0)):
                if item["cut"]:
                    impact = ("**Yes** — this is cut from sheet, so its blank and gauge came "
                              "from somewhere other than a detail drawing. Check them.")
                elif not item["priced"]:
                    impact = ("No — it is unpriced, and a drawing would not price it. It "
                              "needs a rate, not a sheet.")
                else:
                    impact = "No — bought or commercial; priced from a book, not a drawing."
                add(f"| {item['code']} | {item['desc']} "
                    f"| {_gbp(item['gbp']) if item['gbp'] is not None else 'no price'} "
                    f"| {impact} |")
            add("")
        # Pages nothing claimed. Derived from the sheets the parts themselves name — the
        # highest page any record claims is the only page count this document can honestly
        # know, so a gap below is "no part was traced to this sheet", not "the PDF is short".
        claimed = sorted({int(p) for rec in scan.values() for p in (rec.get("pages") or [])
                          if isinstance(p, (int, float))})
        if claimed:
            gaps = [p for p in range(1, max(claimed) + 1) if p not in claimed]
            if gaps:
                add(f"Within p.1–p.{max(claimed)}, **no costed part was traced to "
                    f"{', '.join(f'p.{p}' for p in gaps)}**. Those sheets are usually the "
                    f"cover, the general arrangement and the BOM table, which own no part of "
                    f"their own — but they are also where a part would hide if its drawing "
                    f"were read and never joined to a cost, so they are named rather than "
                    f"assumed harmless.")
            else:
                add(f"Every sheet from p.1 to p.{max(claimed)} is claimed by at least one "
                    f"costed part.")
            add("")

    return "\n".join(lines)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", required=True, type=Path)
    ap.add_argument("--scan-json", type=Path,
                    help="output/json/<job>.json — supplies the page numbers the "
                         "workbook does not carry")
    ap.add_argument("--out", type=Path)
    args = ap.parse_args()

    text = build(args.workbook, args.scan_json)
    if args.out:
        args.out.write_text(text, encoding="utf-8")
        print(f"wrote {args.out}")
    else:
        print(text)


if __name__ == "__main__":
    main()
