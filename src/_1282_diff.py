#!/usr/bin/env python3
r"""
_1282_diff.py — what ACTUALLY changed between two runs?

The workbook's totals are Excel formulas (calc-on-load), so we cannot read those without
opening the file. But every cell the ENGINE writes is a literal — BOM code, supplier, price,
qty, part dimensions, gauge, labour throughput. Those are exactly the inputs we want to
compare, and they are all readable straight off the sheet.

This diffs every literal cell between two workbooks and prints only what differs.

No guessing. If the only line that moved is BI-MAINSCABLE, the £0.51 is the known cached
price-lookup drift and nothing we did last night touched 1282. If ANY other line moved, we
have a real regression and this tells us exactly which one.

Usage:
    python _1282_diff.py "<old.xlsx>" "<new.xlsx>"

To find the candidates:
    Get-ChildItem C:\ClaudeVision\output\estimates\1282*.xlsx |
        Sort-Object LastWriteTime | Select-Object LastWriteTime, Name
"""
from __future__ import annotations
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def literals(path):
    """Every non-formula, non-empty cell on the first sheet."""
    wb = load_workbook(path, data_only=False)
    ws = wb[wb.sheetnames[0]]
    out = {}
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if v is None:
                continue
            if isinstance(v, str) and v.startswith("="):
                continue          # formula — Excel owns it, not us
            out[c.coordinate] = v
    return out


def main():
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    old_p, new_p = sys.argv[1], sys.argv[2]

    old, new = literals(old_p), literals(new_p)
    keys = sorted(set(old) | set(new),
                  key=lambda k: (int("".join(ch for ch in k if ch.isdigit())),
                                 "".join(ch for ch in k if ch.isalpha())))

    diffs = []
    for k in keys:
        a, b = old.get(k), new.get(k)
        if a == b:
            continue
        # ignore pure float noise below a tenth of a penny
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            if abs(float(a) - float(b)) < 0.0005:
                continue
        diffs.append((k, a, b))

    print(f"\nOLD  {old_p}")
    print(f"NEW  {new_p}")
    print(f"\n{len(diffs)} literal cell(s) differ\n")

    if not diffs:
        print("  IDENTICAL INPUTS. Any change in the total is Excel's own roll-up, not ours.")
        return

    print(f"  {'CELL':<8} {'OLD':<42} {'NEW'}")
    print(f"  {'-'*8} {'-'*42} {'-'*42}")
    for k, a, b in diffs:
        sa = ("" if a is None else str(a))[:40]
        sb = ("" if b is None else str(b))[:40]
        print(f"  {k:<8} {sa:<42} {sb}")

    # money columns on this template: J = BOM price, M = BOM total, N/M steel cost, etc.
    money = [(k, a, b) for k, a, b in diffs
             if isinstance(a, (int, float)) and isinstance(b, (int, float))
             and abs(float(b) - float(a)) >= 0.005]
    if money:
        net = sum(float(b) - float(a) for _, a, b in money)
        print(f"\n  {len(money)} numeric cell(s) moved · net {net:+.4f} across all of them")
        print("  (not the unit-cost delta — scrap %, roll-ups and overhead sit on top)")


if __name__ == "__main__":
    main()
