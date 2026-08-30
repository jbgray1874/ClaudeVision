#!/usr/bin/env python3
r"""
_12120_inspect.py   (READ-ONLY - opens the workbook, writes nothing)

Finds the exact cells in the Sheet Steel section that hold Part Length / Part Width / Gauge
for each fabricated part, so a dim-override applier can target the right cells.

Run (from C:\ClaudeVision\src):
    C:\ClaudeVision\.venv\Scripts\python.exe _12120_inspect.py
"""
from __future__ import annotations
import glob, os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def col_letter(idx):
    return get_column_letter(idx) if idx else "?"

OUT_DIR = r"C:\ClaudeVision\output\estimates"
cands = sorted(glob.glob(os.path.join(OUT_DIR, "12120*.xlsx")), key=os.path.getmtime)
if not cands:
    raise SystemExit(f"No 12120 workbook found in {OUT_DIR}")
WB = cands[-1]
print(f"Inspecting: {WB}\n")

wb = load_workbook(WB, data_only=False)
ws = wb.active
print(f"Active sheet: {ws.title}  (dims {ws.dimensions})\n")

steel_hdr_row = None
for r in range(1, ws.max_row + 1):
    for c in range(1, min(ws.max_column, 30) + 1):
        v = ws.cell(r, c).value
        if isinstance(v, str) and v.strip().lower() == "sheet steel":
            steel_hdr_row = r
            break
    if steel_hdr_row:
        break

if not steel_hdr_row:
    print("Could not find 'Sheet Steel' header. First 60 non-empty rows, cols A-N:")
    for r in range(1, min(ws.max_row, 60) + 1):
        cells = [ws.cell(r, c).value for c in range(1, 15)]
        if any(x not in (None, "") for x in cells):
            print(f"  R{r}: {cells}")
    raise SystemExit

print(f"'Sheet Steel' section header at row {steel_hdr_row}.\n")

col_map = {}
hdr_row = None
for r in range(steel_hdr_row, steel_hdr_row + 3):
    labels = {c: str(ws.cell(r, c).value).strip().lower()
              for c in range(1, 20) if isinstance(ws.cell(r, c).value, str)}
    if any("length" in v for v in labels.values()):
        hdr_row = r
        for c, v in labels.items():
            if "part length" in v or v == "length": col_map["L"] = c
            elif "part width" in v or v == "width":  col_map["W"] = c
            elif "gauge" in v:                        col_map["G"] = c
            elif "part description" in v or v == "description": col_map["desc"] = c
            elif "qty per unit" in v:                 col_map["qty"] = c
        break

print(f"Column-header row: {hdr_row}")
print("Column map: " + ", ".join(f"{k}=col{v}({col_letter(v)})" for k, v in col_map.items()) + "\n")

desc_c = col_map.get("desc", 2)
L_c = col_map.get("L"); W_c = col_map.get("W"); G_c = col_map.get("G"); qty_c = col_map.get("qty")

print("Fabricated-part rows (current WRONG PDF-extracted values):")
print("-" * 96)
first_data = (hdr_row or steel_hdr_row) + 1
for r in range(first_data, first_data + 30):
    desc = ws.cell(r, desc_c).value
    if isinstance(desc, str) and desc.strip().lower() in ("other sheet material", "total material cost", "labour"):
        break
    if desc in (None, ""):
        continue
    Lv = ws.cell(r, L_c).value if L_c else "?"
    Wv = ws.cell(r, W_c).value if W_c else "?"
    Gv = ws.cell(r, G_c).value if G_c else "?"
    qv = ws.cell(r, qty_c).value if qty_c else "?"
    print(f"  R{r}: {str(desc)[:40]:40} qty={qv}  L({col_letter(L_c)})={Lv}  W({col_letter(W_c)})={Wv}  G({col_letter(G_c)})={Gv}")
print("-" * 96)
print("\nPaste this back - it gives me the exact cells to write the real dims into.")
