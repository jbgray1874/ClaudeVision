# -*- coding: utf-8 -*-
"""READ-ONLY. Inspect the EXACT formulas in the Other Sheet block cols J (Qty Per Sheet),
L (Cost per sheet), M (Cost Per Part) for the RISER row (51) and an empty row (52), so we know:
  - is M a formula (= L/J style) or a plain writable cell?
  - is L blank (needs a value) or a formula?
  - is J (qty per sheet) a formula computing from sheet vs part dims?
This decides HOW to write the fix (write sheet price to L, OR write per-part cost to M directly).
Run from C:\\ClaudeVision\\src :
  C:\\ClaudeVision\\.venv\\Scripts\\python.exe _probe_othersheet_formula.py
"""
import glob, os
import openpyxl
from openpyxl.utils import get_column_letter
est = sorted(glob.glob(r"C:\ClaudeVision\output\estimates\12532-03RecipeCard_*.xlsx"), key=os.path.getmtime)
path = est[-1]
print(f"=== {os.path.basename(path)} — Other Sheet formulas ===\n")
ws = openpyxl.load_workbook(path, data_only=False)["Estimate"]
for r in (50, 51, 52):
    print(f"--- row {r} ---")
    for c in range(3, 14):  # C..M
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip():
            print(f"    {get_column_letter(c)}{r}: {repr(v)}")
    print()
