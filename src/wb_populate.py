"""
wb_populate.py  —  Populate the real Blank Estimating Workbook template with the
engine's extracted inputs, and let the WB's own formulas do the costing.

WHY THIS EXISTS:
  The previous xlsx_output.py BUILT a workbook from scratch, re-deriving the cost
  model in Python — which diverged from the real WB (the £166 vs £99.99 material
  total). This module instead OPENS the estimators' real template and writes only
  the INPUT cells (dimensions, quantities, operations, BOM prices). The WB's own
  formulas (M59 material, laser rate calc, powder calc, labour lookups, M105 unit
  cost, sell price) then compute everything. One source of truth: the WB.

MODEL (confirmed with JG):
  Engine supplies INPUTS. WB CALCULATES.
    - BOM block:   description, code, PRICE (engine-sourced: UDEF/LLM), qty, scrap
    - Steel block: description, qty, length, width, gauge   -> WB computes cost+laser+powder
    - Other Sheet: description, qty, length, width, thickness
    - Labour:      operation, description, qty              -> WB looks up rate
    - Header:      customer, drawing/job no, ORDER QTY (D6)

LAYOUT CHANGES: every cell address lives in CELL_MAP below. If the estimators hand
  over a WB with a shifted layout, edit CELL_MAP only — not the logic.

Requires openpyxl.
"""
from __future__ import annotations
import os, json, re, shutil, sys
from datetime import datetime

# THE MODULE ITSELF, not just names lifted out of it.
#
# Everything below imports individual settings as `from config import X as _X`, so the bare
# name `config` was never bound here. A later `getattr(config, ...)` therefore raised
# NameError at run time, in the middle of populate_workbook, on every job -- and the whole
# fixture suite passed, because not one fixture CALLS populate_workbook. See
# test_every_module_resolves_the_names_it_uses.
try:
    import config
except Exception:                                  # pragma: no cover - config must import
    config = None                                  # getattr(None, x, default) still returns default

# Powder material rate (£/kg). Single source of truth is config.py — the Excel
# template carries only a static default, which we overwrite on every populate.
try:
    from config import POWDER_COST_PER_KG as _POWDER_COST_PER_KG
except Exception:
    _POWDER_COST_PER_KG = None  # fall back to whatever the template holds
# Coverage: kg of powder per m2 of coated surface. The template's own calculator uses
# 0.1667 (= 6 m2/kg = 100% transfer efficiency, which nothing achieves). See config.py.
try:
    from config import POWDER_KG_PER_M2 as _POWDER_KG_PER_M2
except Exception:
    _POWDER_KG_PER_M2 = 0.20
# Minimum powder booked per coated object. Tim's sheets carry a floor (25-40g) that no
# coverage model explains — you cannot coat a 40mm hook with six grams. ASSUMPTION.
try:
    from config import POWDER_MIN_KG_PER_PIECE as _POWDER_MIN_KG_PER_PIECE
except Exception:
    _POWDER_MIN_KG_PER_PIECE = 0.03
try:
    from config import THROUGHPUT_SIZE_BANDS as _THROUGHPUT_SIZE_BANDS
    from config import THROUGHPUT_AREA_EDGES as _THROUGHPUT_AREA_EDGES
except Exception:
    _THROUGHPUT_SIZE_BANDS = {}
    _THROUGHPUT_AREA_EDGES = (0.05, 0.15, 0.40)
try:
    from config import BOOK_MANM_INSERT_LABOUR as _BOOK_MANM_INSERT_LABOUR
    from config import MANM_INSERT_SECONDS_EACH as _MANM_INSERT_SECONDS_EACH
    from config import MANM_INSERT_PART_TOKENS as _MANM_INSERT_PART_TOKENS
except Exception:
    _BOOK_MANM_INSERT_LABOUR = True
    _MANM_INSERT_SECONDS_EACH = 15.0
    _MANM_INSERT_PART_TOKENS = ["CLINCH", "PEM"]
try:
    # When a fabricated part has no usable blank dims (blank L/W or gauge), the template's
    # per-row material formula errors (#VALUE!/#DIV/0!) and the plain SUM in Total Material
    # Cost propagates that error into Unit and Sell — one missing dim blanks the whole total.
    # With this on, the material total sums with AGGREGATE(9,6,…) which IGNORES errored rows,
    # so the sheet computes a PARTIAL total from the credible lines and self-completes as the
    # estimator fills the flagged dims. Non-regressive: AGGREGATE(9,6,r) == SUM(r) when there
    # are no errors (so 12120/1282 are unchanged). Lever in config for full reversibility.
    from config import MATERIAL_TOTAL_ERROR_TOLERANT as _MATERIAL_TOTAL_ERROR_TOLERANT
except Exception:
    _MATERIAL_TOTAL_ERROR_TOLERANT = True
from typing import Any, Dict, List, Optional, Tuple

# What goes next to a price nobody can reproduce. Short enough for a spreadsheet cell and
# blunt enough that it cannot be read as a supplier quote.
_INDICATIVE_TAG = "[AI ESTIMATE - INDICATIVE, NOT A QUOTE]"

# The engine that answered, named. A provider this map has not met yet still names itself
# from what the lookup recorded, so switching provider needs no edit here.
_LLM_PROVIDER_NAMES = {
    "xai": "xAI Grok LLM",
    "anthropic": "Anthropic Claude LLM",
    "openai": "OpenAI LLM",
}


def _llm_engine_name(block: Dict[str, Any]) -> str:
    raw = str(block.get("llm_provider") or "").strip().lower()
    if not raw or raw == "none":
        return "AI ESTIMATE"
    return _LLM_PROVIDER_NAMES.get(raw, f"{raw} LLM")

# How each kind of source is named in the supplier column, so an estimator reading the sheet
# can see at a glance which lines are firm. Keyed on the source CLASS, never on a part code,
# so a job nobody has seen yet is labelled by the same rule.
_ORIGIN_LABELS = {
    "ai_estimate": "AI ESTIMATE - INDICATIVE",
    "web_catalog": "Web listing - verify",
    "catalogue": "",          # a real catalogue row needs no warning; its supplier name stands
    "config": "",
    "unpriced": "NO PRICE FOUND - estimator to price",
}


def _price_origin(pe: Dict[str, Any]) -> Tuple[str, bool]:
    """(label for the supplier column, is this an unrepeatable guess?) for one BOM line.

    Reads the price stamps the estimator wrote rather than re-deciding anything here: one
    module owns the question of what a price source IS, so the sheet, the report, the quote
    and the invariants cannot disagree about the same line.
    """
    try:
        import price_provenance
    except ImportError:
        return "", False
    best = None
    for _path, block in price_provenance.iter_price_stamps(pe):
        if not price_provenance.stamp_affects_total(block):
            continue
        # The bought-in unit cost is the line's price; a material rate on the same record is
        # not what the BOM row is charging for.
        if "system_cost" in _path:
            best = block
            break
        if best is None:
            best = block
    if best is None:
        return "", False
    cls = price_provenance.stamp_source_class(best)
    if cls == "ai_estimate":
        return f"{_llm_engine_name(best)} - INDICATIVE", True
    label = _ORIGIN_LABELS.get(cls)
    if label is None:
        label = str(best.get("source_name") or "")
    return label, False

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ═══════════════════════════════════════════════════════════════════════════════
# CELL MAP — the ONE place to edit if the WB layout changes.
# Row ranges are the INPUT rows in each block of the 'Estimate' sheet.
# Columns are 1-indexed (A=1). Confirmed against Blank Estimate Sheet WB 2026.
# ═══════════════════════════════════════════════════════════════════════════════
# Refuse to emit an estimate that has silently dropped bought-in parts.
# Set False to revert to the old flag-and-drop behaviour.
STRICT_BOM_OVERFLOW = True

CELL_MAP = {
    # UNC path (not K:) so it resolves in ANY session regardless of drive mapping.
    # K: maps to \\sdi-dc01\shareddata$\Shared — using the UNC directly avoids the
    # drive being unmapped in admin shells / when the engine runs headless via main.py.
    "template_path": r"\\sdi-dc01\shareddata$\Shared\Estimating\Completed\AI Estimating\AISheets\Blank Estimate Sheet  WB 2026.xlsx",
    "output_dir":    r"C:\ClaudeVision\output\estimates",
    "estimate_sheet": "Estimate",
    "structural_sheets": ["Labour", "Material Price Break"],  # NEVER overwrite these

    # Header input cells (row, col) or "A1" style
    "header": {
        "customer":   "C3",
        "drawing_no": "C5",
        "order_qty":  "D6",   # drives $D$6 everywhere — critical
    },

    # BOM / Standard Materials block. Engine writes desc, code, PRICE, qty, scrap.
    # Columns from the dump: C=desc, H=code, I=supplier, J=price, K=qty, L=scrap.
    "bom": {
        "first_row": 11, "last_row": 50,          # 40 slots  (_CELLMAP_WIDENED_BOM40: was 11..25 = 15; widened in Excel 2026-07-13 after 1282 silently dropped its 16th BOM part)
        "col_desc": 3, "col_code": 8, "col_supplier": 9,
        "col_price": 10, "col_qty": 11, "col_scrap": 12,
    },

    # Tube / Wire block. Engine writes desc, qty, gauge, length.
    # Header row 27: C=desc, E=qty, F=gauge, G=length (H..M are formula-driven).
    "tube": {
        "first_row": 53, "last_row": 60,          # 8 slots (+25: BOM widened)
        "col_desc": 3, "col_qty": 5, "col_gauge": 6, "col_length": 7,
    },

    # Sheet Steel block. Engine writes desc, qty, length, width, gauge.
    # Header row 37: C=desc, E=qty, F=len, G=wid, H=gauge, I=sheetL, J=sheetW (K..=formula).
    "steel": {
        "first_row": 63, "last_row": 81,          # 19 slots (+25: BOM widened 2026-07-13)
        "col_desc": 3, "col_qty": 5, "col_length": 6, "col_width": 7, "col_gauge": 8,
        "col_sheet_l": 9, "col_sheet_w": 10,      # optional; WB defaults if blank
        "col_holes": 19, "col_internal_cut": 20,  # S/T: laser-calc inputs (No of holes / Internal Cutting Distance)
    },

    # Other Sheet Material (board/acrylic/HIPS). desc, qty, length, width, thickness.
    # Header row 50: C=desc, D=qty, E=len, F=wid, G=thick, H=sheetL, I=sheetW.
    "other_sheet": {
        "first_row": 84, "last_row": 91,          # 8 slots (+25: BOM widened 2026-07-13)
        "col_desc": 3, "col_qty": 4, "col_length": 5, "col_width": 6, "col_thick": 7,
        "col_sheet_l": 8, "col_sheet_w": 9, "col_cost_per_sheet": 12,
    },

    # Labour block. Engine writes operation-name (C), qty (H), and THROUGHPUT (I).
    # The WB looks up rate/dept/setup from the operation name and computes hours+cost
    # from throughput. Throughput (pieces/hr) = order_qty × qty_per_unit ÷ batch_hours,
    # which makes the WB's own formulas reproduce the engine's run-time cost, and the
    # WB then adds setup from its table. Header row 62: C=operation, H=qty, I=Rate Per Hour
    # (actually throughput/pieces-per-hour, the divisor in the hours & cost formulas).
    "labour": {
        "first_row": 96, "last_row": 167,         # 72 slots (+25: BOM widened 2026-07-13)
        "col_operation": 3, "col_desc": 4, "col_qty": 8, "col_throughput": 9,
    },
}

# Map engine operation names -> WB department operation names (col C of labour block).
# The WB looks up col C against its dept-rate table (rows 115-146). These MUST match
# the WB's EXACT strings or the LOOKUP returns 0. Confirmed from the dept table dump:
#   "Fold", "Laser (Metal)", "P.Coat", "Weld (CO2)", "Punch",
#   "Assemble/pack (Metal)", "Tube", "Saw", "Spotweld", etc.
OP_NAME_MAP = {
    "laser_cutting":  "Laser (Metal)",
    "laser":          "Laser (Metal)",
    "folding":        "Fold",
    "fold":           "Fold",
    "powder_coating": "P.Coat",       # WB table says "P.Coat", not "Powder Coat"
    "powder_coat":    "P.Coat",
    "welding":        "Weld (CO2)",
    "weld":           "Weld (CO2)",
    "dress_welds":    "Dress Welds",   # DRES dept — dress/linish the CO2 weld bead
    "dress":          "Dress Welds",
    "punch":          "Punch",
    "punching":       "Punch",
    "handling":       "Assemble/pack (Metal)",   # engine 'handling' -> WB assemble/pack metal
    "assembly":       "Assemble/pack (Metal)",
    "assemble":       "Assemble/pack (Metal)",
    "robomac":        "Robomac",   # WB dept ROBO £31.45/hr — EXACT string or LOOKUP returns 0
    "wire_forming":   "Robomac",
    "tube_cutting":   "Tube",
    "tube":           "Tube",
    "saw":            "Saw",
    "sawing":         "Saw",
    # Deburr / linish — a quick hand finishing pass. It was unmapped, so the WB fuzzy-matched it
    # to CNCJ (£64.07/hr) and, with no throughput default, billed ~22 hrs. Map it to its real
    # dept so it costs at the deburr rate, not CNC joinery. (Throughput default added above.)
    # GRIN IS NOT IN THE RATE TABLE. These four pointed at "Grinding / Deburr" and the
    # workbook has no such row (H173:K204), so every deburr and every linish on every job
    # LOOKUPed to zero — priced at nothing, and indistinguishable on the sheet from work
    # nobody found. BENC and DRES are the rows that exist.
    "deburr":         "Manual labour (Metal)",
    "linisher":       "Dress Welds",
    # Diamond polish (acrylic edge finish). Unmapped it fell through to a raw op name with no
    # throughput default → ~0.67/hr / £52 on a single diffuser. (Metal DPOL is gated out upstream.)
    "diamond_polish": "Diamond Polish",
    # Timber / joinery operations. These map to the estimators' own route titles
    # (config.SDI_OPERATION_CODES). If a title does not match a row in the WB rate table the
    # WB LOOKUP returns 0 for that line — the run will show which the template actually carries,
    # then we finalise the strings. Rates all exist engine-side (SAW/CNCJ/GLUE/SPRY).
    "cnc_routing":    "CNC Joinery",
    "cnc":            "CNC Joinery",
    "cnc_joinery":    "CNC Joinery",
    "glue":           "Glue",
    "gluing":         "Glue",
    "wet_spray":      "Wet Spray",
    "spray":          "Wet Spray",
    "bench_work":     "Manual labour (Metal)",
    "hardware_insertion": "Manual labour (Metal)",
    "spotweld":       "Spotweld",
    "spot_weld":      "Spotweld",
    "roll":           "Roll",
    "rolling":        "Roll",
    "guillotine":     "Guillotine",
    "linebend":       "Linebend",
    # ── the words the extract prompts ask for, which had no department ──────────────
    # llm_full_extract.ROUTE_OPERATIONS is the vocabulary we TELL the model to use. Five of
    # its words were missing here, so the model returned exactly what it was asked for and
    # the workbook did not know what department it belonged to. tube_cut is the operation
    # M&S 2085's two tubes need: it was on the asking side of the contract and absent from
    # the paying side, so the route could never have been costed however well it was read.
    # A fixture now cross-checks the two lists.
    "tube_cut":       "Tube",              # same department as tube_cutting
    "tube_bending":   "Tubebend",          # SDI bends tube on a tube-bender, not a press
    "tubebend":       "Tubebend",
    "edge_banding":   "Edge Banding",
    # Drilling and tapping share a bench and share the rate table's DRIL row. The previous
    # commit sent them to "Manual labour (Metal)" on my own guess; the rate table
    # (H173:K204) carries DRIL "Drill (Acrylic)", so that is where they go.
    "hole_machining": "Drill (Acrylic)",
    "tapping":        "Drill (Acrylic)",
    "deburring":      "Manual labour (Metal)",
    "linishing":      "Dress Welds",
}

# Acrylic/board parts use DIFFERENT (cheaper) labour operations than metal.
# The WB table has both: "Laser (Acrylic)" £41.21 vs "Laser (Metal)" £68.19,
# "Assemble/pack (Acrylic)" vs "(Metal)", "Manual labour (Acrylic)" vs "(Metal)".
# For an acrylic/board part, these OVERRIDE the metal mapping above. Operations
# with no acrylic variant (fold, powder) fall through to the metal/default name.
# Acrylic/board operations — override the metal defaults where SDI uses different
# equipment. NOTE: 'folding' on acrylic is NOT a press-brake fold — SDI heat-bends
# acrylic on a line-bender (heating element softens it). Map to "Linebend" (LINE).
OP_NAME_MAP_ACRYLIC = {
    "laser_cutting":  "Laser (Acrylic)",
    "laser":          "Laser (Acrylic)",
    "folding":        "Linebend",        # acrylic: heat/line-bend, not press-brake
    "fold":           "Linebend",
    "handling":       "Assemble/pack (Acrylic)",
    "assembly":       "Assemble/pack (Acrylic)",
    "assemble":       "Assemble/pack (Acrylic)",
    "manual":              "Manual labour (Acrylic)",
    "manual_labour":       "Manual labour (Acrylic)",
    "manual_labour_acrylic": "Manual labour (Acrylic)",   # engine emits the _acrylic-suffixed key
    "diamond_polish":      "Diamond Polish",
}

# Tube/section bending: SDI bends RHS tube on a tube-bender, NOT a press-brake.
# The engine correctly retains 'folding' on tubes (SDI does bend them), but the
# WB operation is "Tubebend" (TBEN £32.84 / 45 min setup), not "Fold" (FOLD £40.47).
_TUBE_OP_REMAP = {
    "folding": "Tubebend",
    "fold":    "Tubebend",
    # A TUBE IS STILL CUT. IT IS JUST NOT CUT ON A FLAT-BED LASER.
    #
    # laser/guillotine on a tube used to be DROPPED as spurious, which is half right and
    # expensively so: a tube has no flat blank, so it is not profile-lasered -- but it is
    # sawn to length, and deleting the operation left the cut costing nothing. 2085's tubes
    # came out of that change with no cutting operation of any kind on the sheet.
    #
    # Remapped, not dropped, exactly as fold is. The route said "this part gets cut"; the
    # stock form says the cut happens in the tube department, not on the laser. Punch stays
    # spurious below: putting a hole through a tube wall is not cutting it to length, and
    # hole work has its own operation.
    "laser":         "Tube",
    "laser_cutting": "Tube",
    "laser_metal":   "Tube",
    "guillotine":    "Tube",
    "tube_cut":      "Tube",
    "tube_cutting":  "Tube",
    "saw":           "Saw",
}





def _bom_line_price(_pe: Dict[str, Any]) -> Optional[float]:
    """Best-available unit price for a BOM line (mirrors the per-row logic below).
    Used to sum consolidated overflow value. Withheld/unpriced -> None."""
    if _pe.get("_price_explicitly_withheld"):
        return None
    _me = _pe.get("material_estimate") or {}
    _p = _safe(_pe.get("unit_cost_gbp") or _pe.get("unit_material_cost_gbp")
               or _me.get("unit_material_cost_gbp"))
    if _p is None:
        # THE BOM PRICE COLUMN IS A MATERIAL COLUMN.
        #
        # This fell back to extended_total_cost_gbp, which is the part's WHOLE unit cost
        # -- material AND labour. On a fabricated part with no material price that puts
        # its labour in the material column: M&S 2085's tubes showed GBP 19.25 each,
        # which is GBP 17.80 of powder-coating labour, GBP 0.71 weld, GBP 0.33 dress and
        # GBP 0.42 handling. Not one penny of it is tube.
        #
        # It read as an impossible material price -- 60 metres of stock per bracket --
        # and it is about to become worse than misleading: now that a routed operation
        # reaches the labour block, those same operations get their own rows and the
        # labour is counted twice.
        #
        # Material only. A part whose material nobody could price is UNPRICED, and says
        # so, which is the honest answer and the one an estimator can act on.
        _ext_mat = _safe(_pe.get("extended_material_cost_gbp")
                         or _me.get("extended_material_cost_gbp"))
        _q = int(_safe(_pe.get("quantity"), 1) or 1)
        if _ext_mat is not None and _q > 0:
            _p = round(_ext_mat / _q, 4)
        else:
            # "HAS NO LABOUR" IS NOT THE SAME CONDITION AS "IS A BOUGHT-IN".
            #
            # The first version kept the whole-total fallback only for a part with no labour
            # at all, reasoning that a bought-in has none. But a bought-in DOES carry
            # handling — bought_in_policy deliberately leaves it, because fitting a purchased
            # component is real bench time — so a fixing with a handling cost would have lost
            # its price entirely and shown blank on the sheet.
            #
            # The arithmetic answers it without classifying anything: material is the unit
            # total less the labour on it. That holds for a bought-in and a fabricated part
            # alike, and it is why the tubes come out unpriced — their total IS their labour,
            # so the remainder is nothing.
            _unit_total = _safe(_pe.get("unit_total_cost_gbp"))
            if _unit_total is None:
                _ext = _safe(_pe.get("extended_total_cost_gbp"))
                if _ext is not None and _q > 0:
                    _unit_total = _ext / _q
            _lab = (_pe.get("labour_estimate") or {}).get("costs_gbp") or {}
            _lab_total = sum(_safe(v) or 0.0 for v in _lab.values())
            if _unit_total is not None:
                _material = _unit_total - _lab_total
                # A rounding tail is not a price. At or below a penny, the arithmetic is
                # saying there was no material here.
                _p = round(_material, 4) if _material > 0.01 else None
    return _p

def route_operations_by_part(summary: Dict[str, Any]) -> Dict[str, List[str]]:
    """part number -> the operations on its RAW record.

    THE ROUTE AND THE COST LIVE ON DIFFERENT RECORDS, AND THE SHEET ONLY READS ONE.

    estimator.estimate_part builds a costed part_estimate and never copies
    textual_operations onto it, so estimate_summary.part_estimates — which is what
    wb_populate walks — has no route at all. The route stays on summary["parts"] and
    summary["manufacturing_writeup"]["parts"].

    That is why 2085's tube_cut was visible to the invariant and invisible to the sheet:
    invariants._parts falls through to summary["parts"] and found it, wb_populate read
    part_estimates and found nothing. Both were right about the record they were looking at.
    It is the same wrong-record shape as _parts() returning geometry to a price check, and
    the same shape as the bom/parts bridge — the third time in this job.
    """
    out: Dict[str, List[str]] = {}
    if not isinstance(summary, dict):
        return out
    _sources = [summary.get("parts"),
                (summary.get("manufacturing_writeup") or {}).get("parts")]
    for _src in _sources:
        if not isinstance(_src, list):
            continue
        for _p in _src:
            if not isinstance(_p, dict):
                continue
            _pn = str(_p.get("part_number") or "").strip().upper()
            if not _pn:
                continue
            _bucket = out.setdefault(_pn, [])
            for _k in ("textual_operations", "operations", "inferred_operations"):
                _v = _p.get(_k)
                if not isinstance(_v, list):
                    continue
                for _o in _v:
                    _os = str(_o).strip().lower()
                    if _os and _os not in _bucket:
                        _bucket.append(_os)
            # A ruling by measurement travels with the route, or the sheet would re-add
            # exactly what the measurement removed.
            for _ro in (_p.get("operations_ruled_out") or {}):
                _ros = str(_ro).strip().lower()
                if _ros in _bucket:
                    _bucket.remove(_ros)
    return out



def tube_part_numbers(summary: Dict[str, Any]) -> set:
    """Part numbers that are section/tube stock, on the evidence rather than on a price.

    material_estimate.stock_form is only stamped where the material successfully COSTED, so
    a tube nobody could price does not carry it — which makes every stock-form rule silently
    skip the parts that most need it. The durable evidence is section_stock, written by
    document_builder when it reads a profile off the drawing, and material_family, read off
    the BOM. Both live on the RAW records.

    Deliberately structured evidence only. A description containing "TUBE" would catch
    2085's parts and also a purchased TUBE CLAMP, and this set gates whether an operation is
    dropped — so it stays on facts the readers established, not on a word.
    """
    out = set()
    if not isinstance(summary, dict):
        return out
    _sources = [summary.get("parts"),
                (summary.get("manufacturing_writeup") or {}).get("parts"),
                (summary.get("estimate_summary") or {}).get("part_estimates")]
    for _src in _sources:
        if not isinstance(_src, list):
            continue
        for _p in _src:
            if not isinstance(_p, dict):
                continue
            _pn = str(_p.get("part_number") or "").strip().upper()
            if not _pn:
                continue
            if (_p.get("section_stock")
                    or str(_p.get("material_family") or "").strip().lower() == "tube"
                    or str((_p.get("material_estimate") or {}).get("stock_form") or "").lower() == "tube"):
                out.add(_pn)
    return out



def operation_scope_for(pe: Dict[str, Any], op: str,
                        scope_by_op: Optional[Dict[str, str]] = None) -> Optional[str]:
    """"part" or "assembly" for one operation on one part, or None if nobody said.

    Module-level so it can be driven. The first version lived inline in the group loop and
    its mutation passed clean — the fourth time today a check has verified source text
    instead of behaviour.

    A CHAINED OPERATION INHERITS THE SCOPE OF WHAT SPAWNED IT. estimator.py adds dress_welds
    automatically wherever it finds welding, so that row has no route line of its own and
    therefore no scope. It is the same event: if the weld happens once per assembly, so does
    dressing it. Without this the weld would be charged once and its dressing three times,
    which is worse than either answer applied consistently.
    """
    _o = str(op or "").strip().lower()
    if not _o:
        return None
    _own = (pe or {}).get("operation_scope") or {}
    _job = scope_by_op or {}
    for _key in (_o, _CHAINED_FROM.get(_o)):
        if not _key:
            continue
        _v = _own.get(_key) or _job.get(_key)
        if _v in ("part", "assembly"):
            return _v
    return None


_ACRYLIC_NEVER_POWDER = {"ACRYLIC", "HIGH IMPACT ACRYLIC", "PERSPEX", "PMMA", "POLYCARBONATE"}


def section_coated_area_m2(part: Dict[str, Any]) -> float:
    """Powder-coated surface of one SECTION-STOCK part (tube, box, angle, flat bar), m2.

    The coated-area sum had exactly two contributors: sheet (L x W x 2) and wire (pi x d x L).
    Section stock matched neither. The sheet loop excludes it twice over -- once on stock_form,
    once on a "TUBE" description guard added to stop garbled view geometry inventing 24 m2 of
    phantom blank -- and nothing put the real cylinder surface back. So every powder-coated
    tube, box section and angle on every job has been contributing ZERO coated area, which is
    an under-charge that grows with how much section the job contains.

    The wall is irrelevant: powder lands on the OUTSIDE. So the coated surface is the outer
    perimeter times the cut length.

        round / CHS      pi x D
        everything else  2 x (a + b)

    2(a+b) is right for a closed box and also right for an open angle or channel, where powder
    reaches both faces of both legs: two legs of width a and b, coated top and bottom, is the
    same number. A flat bar is that with b = t. The end faces are ignored -- on a 34 mm tube
    they are under 1% and counting them would imply a precision this does not have.

    Returns 0.0 when the profile or the length is missing, because an unmeasured tube is not a
    zero-area one and inventing a size here would be worse than the gap it fills.
    """
    _p = part or {}
    _ss = _p.get("section_stock") or {}
    if not _ss:
        return 0.0
    _mat = str(_p.get("normalized_material")
               or (_p.get("material_estimate") or {}).get("material") or "").upper().replace("_", " ")
    if _mat in _ACRYLIC_NEVER_POWDER or _p.get("acrylic_no_powder"):
        return 0.0          # acrylic tube is polished, never coated
    _a = _safe(_ss.get("a"))
    _b = _safe(_ss.get("b"))
    if not _a:
        return 0.0
    _me = _p.get("material_estimate") or {}
    _len = _safe((_me.get("stock_estimate") or {}).get("section_length_mm")) or _safe(_ss.get("length_mm"))
    if not _len or _len <= 0:
        return 0.0
    if str(_ss.get("profile_form") or "").strip().upper() in ("CHS", "ROUND", "TUBE_ROUND"):
        _perim_mm = 3.14159265 * _a
    else:
        _perim_mm = 2.0 * (_a + (_b or _a))
    _q = _safe(_p.get("quantity"), 1) or 1
    return (_perim_mm / 1000.0) * (_len / 1000.0) * float(_q)


def template_path() -> str:
    """The estimate template to populate.

    CELL_MAP pins a UNC path on the estimating share, which is right when the engine runs
    on the SDI network and a dead end when it does not: populate_workbook returns None, and
    every job silently falls back to the xlsx_output builder -- which looks identical and
    is not route-grouped.

    SDI_WB_TEMPLATE overrides it with a local copy. The share stays the default, so nothing
    changes for a normal run; an operator off the network gets a way through instead of a
    fallback sheet they have no reason to distrust.
    """
    _env = str(os.environ.get("SDI_WB_TEMPLATE") or "").strip().strip('"')
    return _env or CELL_MAP["template_path"]


def assembly_scoped_qty(group: Dict[str, Any]) -> int:
    """How many times an ASSEMBLY-scoped operation is charged per finished product.

    Normally once -- that is what assembly scope MEANS. But qty_per_unit on the route line
    says how many times the operation happens per unit, and a product containing two welded
    frames is two weldings even though each is assembly-level. The route fold stores it;
    without this it would be a field the schema asks the LLM for, the fold carries, and
    nothing ever reads -- which reads as consumed and is not.

    Never below 1: an assembly-scoped operation that reached this point is on the route, and
    a route line that happens zero times would not be there.
    """
    _q = _safe((group or {}).get("qty_per_unit_by_scope") or None, 1) or 1
    try:
        return max(1, int(_q))
    except (TypeError, ValueError):
        return 1


def routed_operations_without_cost(pe: Dict[str, Any], costs: Any = None,
                                   extra_ops: Any = None) -> List[str]:
    """Fabrication operations on a part's ROUTE that the estimator put no cost against.

    Module-level and pure so it can actually be driven. The first version of this lived
    inline in the middle of a six-hundred-line function, and the fixture written for it
    mirrored the logic instead of calling it — so the mutation that deleted the guard passed.
    That is the same defect shape as the bug it was testing.

    Only FABRICATION operations count. Handling and assembly without a cost are not a
    missing row: bought_in_policy deliberately leaves those on purchased parts, and treating
    them as missing work is how BI-ADHESIVECABLE got an Assemble/pack line it never earned.
    """
    try:
        from bought_in_policy import FABRICATION_OPS as _FAB
    except Exception:
        return []
    if not isinstance(pe, dict):
        return []
    if costs is None:
        costs = (pe.get("labour_estimate") or {}).get("costs_gbp") or {}
    _costed = {str(k).strip().lower() for k in (costs or {})}
    # An operation a MEASUREMENT ruled out is not a missing row — it is an answered
    # question. Belt and braces with the route fold, which already refuses to re-add one:
    # this is the last gate before a labour row, and it is the one that spends money.
    _ruled = {str(k).strip().lower() for k in (pe.get("operations_ruled_out") or {})}
    out: List[str] = []
    _pools = [pe.get(k) for k in ("textual_operations", "operations", "inferred_operations")]
    # extra_ops carries the route off the RAW record, which the costed part_estimate does
    # not have. Without it this function can only ever see an empty list on a real job.
    _pools.append(list(extra_ops or []))
    for vals in _pools:
        if not isinstance(vals, list):
            continue
        for o in vals:
            os_ = str(o).strip().lower()
            if (os_ and os_ not in _costed and os_ not in out and os_ not in _ruled
                    and os_ in _FAB):
                out.append(os_)
    return out


def _map_operation(op: str, is_acrylic: bool, stock_form: str = "") -> Optional[str]:
    """Map an engine operation name to the correct WB department name.

    Priority:
      1. Tube stock → remap fold/folding to Tubebend (tube-bender, not press-brake)
      2. Acrylic/board → use acrylic operation variants (Linebend, Laser (Acrylic), etc.)
      3. Everything else → standard metal mapping
    """
    key = str(op).lower()
    sf = str(stock_form or "").lower()
    # Tube bending: the engine assigns 'folding' but SDI uses a tube-bender
    if sf == "tube" and key in _TUBE_OP_REMAP:
        return _TUBE_OP_REMAP[key]
    # Acrylic/board operations
    if is_acrylic and key in OP_NAME_MAP_ACRYLIC:
        return OP_NAME_MAP_ACRYLIC[key]
    # Standard metal
    _hit = OP_NAME_MAP.get(key)
    if _hit:
        return _hit

    # ── LAST RESORT: THE ALIAS TABLE, NOT A SILENT NONE ──────────────────────────
    # OP_NAME_MAP is keyed on the engine's own operation words. A route now also arrives
    # from a model, and a model writing about manufacturing produces English — "Cut to
    # length", "MIG weld", "Laser cut" — or answers with the department code it was asked
    # for. None of those are engine words, so this returned None and the operation was
    # dropped without a sound.
    #
    # department_codes resolves all of it against the rate table's own closed vocabulary
    # (H173:K204). Returning None still means "nothing recognises this", which is what the
    # caller must flag — but it now means it far less often, and never for a word we
    # ourselves asked the model to use.
    try:
        from department_codes import title_for as _title_for
        return _title_for(op)
    except Exception:
        return None


# Operations that are genuinely impossible / meaningless for a given stock form
# regardless of remapping — these are dropped entirely.
# NOTE: tube-fold and acrylic-fold are NOT in this list — they are REAL operations
# (tube-bending and line-bending) that get REMAPPED above, not dropped.
# Only operations that serve no purpose at all go here.
# Operations the engine ADDS from another operation rather than reading from a route. They
# carry no route line, so they have no scope of their own and inherit their parent's.
_CHAINED_FROM = {
    "dress_welds": "welding",
}

_SPURIOUS_OPS_BY_STOCK_FORM = {
    # A TUBE HAS NO FLAT BLANK, SO IT IS NEITHER PUNCHED NOR PROFILE-LASERED.
    # 2085's tubes carried laser_cutting, inherited from the shared assembly page the
    # plate's route was read off. Nothing measured says either tube is laser-profiled --
    # they are sawn to length and welded in -- and now that a routed operation reaches the
    # labour block, that inherited op would have booked a laser cut on both of them.
    # Dropped, and the drop is flagged by name where it happens, so if SDI ever profiles
    # tube on a tube laser this is one line and the flag says which parts it affected.
    # PUNCH ONLY. The cutting operations moved to _TUBE_OP_REMAP -- a tube is not
    # profile-lasered but it IS cut to length, and dropping the op made the cut free.
    # Putting a hole through a tube wall is not cutting it to length, so punch stays here.
    "tube": {"punch", "punching"},
    # A solid round bar has NO FLAT BLANK. It cannot be lasered, folded, punched,
    # line-bent, guillotined or diamond-polished. It is cut (Robomac / Saw) and welded.
    # 1310-02 STUD (8mm dia x 65) was carrying Laser £4.91 from the original misread
    # that treated its DIAMETER as an 8mm sheet THICKNESS.
    "wire": {
        "laser", "laser_cutting", "laser_metal",
        "fold", "folding",
        "punch", "punching",
        "linebend", "line_bend",
        "guillotine",
        "diamond_polish",
    },
}
# Acrylic doesn't get punched either (laser or CNC cut, not punch press).
_SPURIOUS_OPS_BY_MATERIAL = {
    "acrylic": {"punch", "punching"},
}


def _is_spurious_operation(op: str, stock_form: str, material: str = "") -> bool:
    """True if this operation is physically impossible for this stock form / material.
    Note: fold/folding on tubes and acrylic is NOT spurious — it's remapped to the
    correct WB operation (Tubebend / Linebend) by _map_operation, not dropped here."""
    key = str(op).lower()
    sf = str(stock_form or "").lower()
    if key in _SPURIOUS_OPS_BY_STOCK_FORM.get(sf, set()):
        return True
    mat = str(material or "").lower()
    for mat_key, bad in _SPURIOUS_OPS_BY_MATERIAL.items():
        if mat_key in mat and key in bad:
            return True
    return False


def _safe(v, default=None):
    try:
        if v in (None, "", "None"):
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def _is_sheet_metal(mat: str) -> bool:
    m = (mat or "").upper()
    return any(k in m for k in ("STEEL", "MILD", "MS", "STAINLESS", "ALUM", "GALV", "CR4"))


def _is_board(mat: str) -> bool:
    m = (mat or "").upper()
    return any(k in m for k in ("MDF", "ACRYLIC", "HIPS", "FOAM", "PVC", "POLY", "PERSPEX", "BOARD",
                                # timber families — a glued-and-pinned timber crate is NOT sheet metal
                                "TIMBER", "WOOD", "PINE", "PLYWOOD", "SOFTWOOD", "HARDWOOD", "OAK",
                                "SPRUCE", "BEECH", "BIRCH"))


_TIMBER_TOKENS = ("TIMBER", "WOOD", "PINE", "PLYWOOD", "SOFTWOOD", "HARDWOOD", "OAK",
                  "SPRUCE", "BEECH", "BIRCH", "MDF", "CHIPBOARD", "OSB")


def _is_timber(mat: str) -> bool:
    """Timber/board joinery, as distinct from acrylic and the other plastics.

    _is_board() lumps them together because neither is sheet metal, which is right for
    deciding the cost stream but wrong for naming a DEPARTMENT: a wooden crate is not
    made on the acrylic line. The workbook template has no joinery Assemble/pack, so a
    timber part still has to take the nearest hand-assembly rate — but the estimator is
    told that is what happened rather than reading 'Acrylic' against a timber crate and
    having to work out why."""
    m = (mat or "").upper()
    if "ACRYLIC" in m or "PERSPEX" in m or "PMMA" in m:
        return False          # veneered/laminated acrylic products stay acrylic
    return any(k in m for k in _TIMBER_TOKENS)


class CanonicalRouteUnavailable(RuntimeError):
    """The authoritative route was requested but could not be compiled."""


def labour_row_description(wb_op: Any, material: Any = "", thickness: Any = None,
                           parts: Any = (), bends: Any = 0, holes: Any = 0) -> str:
    """The text an estimator reads on a labour row.

    Module-level so it can be driven. The first check on the DRIL wording asserted that a
    string appeared in this file, which proves the line was typed, not that it reaches a
    row -- the failure this run has been correcting all day.
    """
    _matx = str(material or "").replace("_", " ").strip()
    _spec = []
    if thickness:
        _spec.append(("%g" % float(thickness)) + "mm")
    if _matx:
        _spec.append(_matx)
    _rd = str(wb_op)
    # "DRILL (ACRYLIC)" ON A STEEL PART IS THE RIGHT ROW WITH A MISLEADING NAME.
    #
    # DRIL's column-H title in the rate table is literally "Drill (Acrylic)", and the shop
    # books metal drilling and tapping against that same row. The TITLE cannot change: it is
    # the lookup key, and a title the rate table does not carry returns a rate of zero and
    # costs the work at nothing -- which is how GRIN silently zeroed every deburr on every
    # job. So the title stays exactly as the workbook has it, and the DESCRIPTION says what
    # the work actually is.
    if (str(wb_op) == "Drill (Acrylic)" and _matx
            and "ACRYLIC" not in _matx.upper() and "PERSPEX" not in _matx.upper()):
        _rd += " [metal drill/tap — DRIL is the rate-table row the shop books this to]"
    if _spec:
        _rd += " — " + " ".join(_spec)
    _pl = list(parts or [])
    if _pl:
        _rd += " (" + ", ".join(_pl[:6]) + (", +%d more" % (len(_pl) - 6)
                                            if len(_pl) > 6 else "") + ")"
    if bends:
        _rd += " (%d bend%s)" % (int(bends), "" if int(bends) == 1 else "s")
    elif holes:
        _rd += " (%d hole%s)" % (int(holes), "" if int(holes) == 1 else "s")
    return _rd


def canonical_route_payload(summary: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(summary, dict):
        return {}
    payload = (
        (summary.get("estimate_summary") or {}).get("canonical_route_shadow")
        or summary.get("canonical_route_shadow")
        or {}
    )
    return payload if isinstance(payload, dict) else {}


def canonical_route_cutover_enabled(summary: Dict[str, Any]) -> bool:
    requested = bool(getattr(
        config, "CANONICAL_ROUTE_WORKBOOK_CUTOVER", False))
    if not requested:
        return False
    payload = canonical_route_payload(summary)
    if not payload:
        raise CanonicalRouteUnavailable(
            "canonical route cutover is enabled but no compiled route is present")
    if payload.get("compiler_error"):
        raise CanonicalRouteUnavailable(
            f"canonical route compiler failed: {payload.get('compiler_error')}")
    if not isinstance(payload.get("decisions"), list):
        raise CanonicalRouteUnavailable(
            "canonical route cutover is enabled but decisions are missing")
    return True


def canonical_part_kinds(summary: Dict[str, Any]) -> Dict[str, str]:
    """Part identity -> canonical hierarchy kind."""
    out: Dict[str, str] = {}
    for node in canonical_route_payload(summary).get("nodes") or []:
        if not isinstance(node, dict):
            continue
        part_number = str(node.get("part_number") or "").strip().upper()
        kind = str(node.get("kind") or "").strip().lower()
        if part_number and kind in {"leaf", "assembly", "bought_in"}:
            out[part_number] = kind
    return out


def canonicalise_part_estimates_for_workbook(
    summary: Dict[str, Any],
    part_estimates: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Apply canonical BOM identity and multiplicity without losing cost evidence."""
    nodes = {
        str(node.get("part_number") or "").strip().upper(): node
        for node in canonical_route_payload(summary).get("nodes") or []
        if isinstance(node, dict) and node.get("part_number")
    }
    aliases = {
        str(alias).strip().upper(): identity
        for identity, node in nodes.items()
        for alias in ((node.get("evidence") or {}).get("raw_aliases") or [])
        if str(alias).strip()
    }
    # A SYNTHESISED CODE YIELDS TO A CANONICAL ONE FOR THE SAME ITEM.
    #
    # 12120 shipped the same PEM stud twice: STD PART qty 2 from the GA BOM, and BI-PEMSTUD
    # qty 2 minted by the prose recogniser from the same words. Four studs where the drawing
    # calls for two. Free while both were unpriced; wrong the moment either got a price.
    #
    # A "BI-" code is not a part number the draughtsman wrote -- it is one this engine
    # invented from a phrase ("BI-" + the phrase, uppercased). So where a synthesised code
    # describes the same item as a code the canonical graph already holds, the graph wins.
    # Keyed on that distinction, not on PEM studs: any recogniser-minted duplicate of a
    # drawing item resolves the same way.
    def _desc_key(_t: Any) -> str:
        return re.sub(r"[^A-Z0-9]+", " ", str(_t or "").upper()).strip()

    # ONE CANDIDATE, OR NONE. Keeping the first node with a given description made the
    # merge target depend on dict order: two canonical bought-ins sharing a description
    # would absorb the synthesised line into whichever happened to be seen first, and the
    # sheet would show the quantity against the wrong part with nothing to indicate it.
    # An ambiguous match is not a match.
    _desc_hits: Dict[str, List[str]] = {}
    for _ident, _node in nodes.items():
        _dk = _desc_key(_node.get("description"))
        if _dk:
            _desc_hits.setdefault(_dk, []).append(_ident)
    _canonical_by_desc = {_dk: _ids[0] for _dk, _ids in _desc_hits.items() if len(_ids) == 1}
    _ambiguous = {_dk for _dk, _ids in _desc_hits.items() if len(_ids) > 1}
    _synth_merged: List[str] = []
    for _est in part_estimates or []:
        if not isinstance(_est, dict):
            continue
        _sid = str(_est.get("part_number") or "").strip().upper()
        if not _sid.startswith("BI-") or _sid in nodes or _sid in aliases:
            continue
        _dk = _desc_key(_est.get("description"))
        if _dk in _ambiguous:
            print(f"   [wb_populate] {_sid} matches MORE THAN ONE canonical item by "
                  f"description; left as its own line rather than merged into an arbitrary "
                  f"one — estimator to reconcile", flush=True)
            continue
        _target = _canonical_by_desc.get(_dk)
        if _target and _target != _sid:
            aliases[_sid] = _target
            _synth_merged.append(f"{_sid} -> {_target}")

    normalised: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []
    for estimate in part_estimates or []:
        if not isinstance(estimate, dict):
            continue
        source_id = str(estimate.get("part_number") or "").strip().upper()
        identity = aliases.get(source_id, source_id)
        if not identity:
            continue
        item = dict(estimate)
        node = nodes.get(identity) or {}
        if identity != source_id:
            item["_canonical_source_part_number"] = source_id
            item["part_number"] = identity
        if node.get("qty_per_unit") is not None:
            item["quantity"] = node.get("qty_per_unit")
        if not item.get("description") and node.get("description"):
            item["description"] = node.get("description")
        if identity not in normalised:
            normalised[identity] = item
            order.append(identity)
        else:
            # Prefer the record that actually carries a price/material estimate; aliases
            # exist to reconcile identity, never to discard the costed evidence.
            old = normalised[identity]
            old_score = sum(bool(old.get(key)) for key in (
                "unit_cost_gbp", "unit_total_cost_gbp", "material_estimate"))
            new_score = sum(bool(item.get(key)) for key in (
                "unit_cost_gbp", "unit_total_cost_gbp", "material_estimate"))
            if new_score > old_score:
                normalised[identity] = item

    if _synth_merged:
        print(f"   [wb_populate] canonical identity absorbed {len(_synth_merged)} "
              f"engine-minted duplicate(s): {', '.join(_synth_merged)}", flush=True)

    # An explicit bought-in BOM line must remain visible even when no pricing record was
    # created. It is safer as an unpriced estimator row than absent from the BOM.
    for identity, node in nodes.items():
        if node.get("kind") != "bought_in" or identity in normalised:
            continue
        normalised[identity] = {
            "part_number": identity,
            "description": node.get("description") or "",
            "quantity": node.get("qty_per_unit") or 1,
            "page_roles": ["bought_in"],
            "material_estimate": {},
            "_price_explicitly_withheld": True,
            "review_flag": True,
            "review_flags": [
                "Explicit canonical BOM item has no pricing record; estimator to price."
            ],
        }
        order.append(identity)

    # A fabricated leaf in the explicit hierarchy cannot disappear merely because an
    # upstream adapter failed to produce a cost record. Bought-ins are made visible above;
    # fabricated leaves need geometry/material investigation and therefore block release.
    payload = canonical_route_payload(summary)
    existing_issues = {
        (
            str(issue.get("code") or ""),
            str(issue.get("part_number") or "").strip().upper(),
        )
        for issue in (payload.get("issues") or [])
        if isinstance(issue, dict)
    }
    for identity, node in nodes.items():
        if (
            node.get("kind") == "leaf"
            and node.get("parents")
            and identity not in normalised
            and ("bom_leaf_without_estimate", identity) not in existing_issues
        ):
            payload.setdefault("issues", []).append({
                "code": "bom_leaf_without_estimate",
                "part_number": identity,
                "description": node.get("description") or "",
                "parents": list(node.get("parents") or []),
                "message": (
                    "An explicit fabricated BOM leaf has no estimate record and cannot "
                    "be placed in a material block."
                ),
            })
    return [normalised[identity] for identity in order]


def _canonical_finish_signature(
    identities: List[str],
    estimates: Dict[str, Dict[str, Any]],
    raw: Dict[str, Dict[str, Any]],
) -> str:
    """Stable finish/setup identity; different colours must not share a booth setup."""
    values: List[str] = []
    for identity in identities:
        for record in (estimates.get(identity) or {}, raw.get(identity) or {}):
            for key in (
                "normalized_finish", "finish", "surface_finish",
                "surface_finishes", "finish_general",
            ):
                value = record.get(key)
                if isinstance(value, dict):
                    value = value.get("description") or value.get("name") or value.get("value")
                if isinstance(value, (list, tuple, set)):
                    candidates = value
                else:
                    candidates = [value]
                for candidate in candidates:
                    text = re.sub(r"\s+", " ", str(candidate or "").strip()).upper()
                    if text and text not in values:
                        values.append(text)
    return " | ".join(sorted(values)) or "FINISH-UNSPECIFIED"


def _canonical_record_index(
    summary: Dict[str, Any],
    part_estimates: List[Dict[str, Any]],
) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    estimates = {
        str(item.get("part_number") or "").strip().upper(): item
        for item in (part_estimates or [])
        if isinstance(item, dict) and item.get("part_number")
    }
    raw: Dict[str, Dict[str, Any]] = {}
    for pool in (
        summary.get("parts"),
        (summary.get("manufacturing_writeup") or {}).get("parts"),
    ):
        if not isinstance(pool, list):
            continue
        for item in pool:
            if not isinstance(item, dict):
                continue
            part_number = str(item.get("part_number") or "").strip().upper()
            if part_number:
                raw.setdefault(part_number, item)
    return estimates, raw


def canonical_labour_groups(
    summary: Dict[str, Any],
    part_estimates: List[Dict[str, Any]],
    order_qty: int,
    skipped_part_numbers: Any = (),
    all_fabricated_are_wire: bool = False,
) -> Dict[Any, Dict[str, Any]]:
    """Render required OperationDecisions into workbook pricing groups.

    Route authority stays in the compiler. This function may combine compatible part-level
    decisions into one tooling setup, but it never creates an operation from a raw word or
    a missing legacy cost. Assembly decisions remain one group per target event.
    """
    payload = canonical_route_payload(summary)
    estimates, raw = _canonical_record_index(summary, part_estimates)
    node_qty = {
        str(node.get("part_number") or "").strip().upper():
            (_safe(node.get("qty_per_unit"), 1) or 1)
        for node in (payload.get("nodes") or [])
        if isinstance(node, dict) and node.get("part_number")
    }
    skipped = {
        str(item).strip().upper() for item in (skipped_part_numbers or [])
        if str(item).strip()
    }
    tube_pns = tube_part_numbers(summary)
    groups: Dict[Any, Dict[str, Any]] = {}

    for decision in payload.get("decisions") or []:
        if not isinstance(decision, dict) or decision.get("status") != "required":
            continue
        decision_id = str(decision.get("decision_id") or "").strip()
        operation = str(decision.get("operation") or "").strip().lower()
        target_id = str(decision.get("target_id") or "").strip().upper()
        participants = [
            str(item).strip().upper()
            for item in (decision.get("participants") or [])
            if str(item).strip()
        ]
        participants = list(dict.fromkeys(participants))
        if not decision_id or not operation:
            continue
        if target_id in skipped or any(item in skipped for item in participants):
            continue

        candidate_ids = list(dict.fromkeys(
            [target_id] + participants))
        representative_id = next(
            (item for item in candidate_ids if item in estimates), "")
        pe = estimates.get(representative_id) or {}
        raw_part = raw.get(representative_id) or {}
        material = str(
            pe.get("normalized_material")
            or (pe.get("material_estimate") or {}).get("material")
            or raw_part.get("normalized_material")
            or ""
        )
        thickness = _safe(
            pe.get("normalized_thickness_mm")
            or (pe.get("material_estimate") or {}).get("thickness_mm"),
            0,
        ) or 0
        stock_form = str(
            (pe.get("material_estimate") or {}).get("stock_form") or ""
        ).lower()
        if not stock_form and representative_id in tube_pns:
            stock_form = "tube"
        is_acrylic = _is_board(material)
        wb_op = _map_operation(operation, is_acrylic, stock_form)
        if wb_op is None:
            wb_op = operation
        if (
            wb_op == "Weld (CO2)"
            and stock_form in {"wire", "bar"}
            and all_fabricated_are_wire
        ):
            wb_op = "Spotweld"

        scope = str(decision.get("scope") or "part").lower()
        sequence = _safe(decision.get("sequence"))
        qty = _safe(decision.get("qty_per_unit"), 1) or 1
        if operation == "hardware_insertion":
            key = ("canonical-event", decision_id)
        elif operation == "powder_coating":
            # One colour/booth setup can carry several separately identified coated
            # objects. Different finishes remain separate setups.
            finish_signature = _canonical_finish_signature(
                candidate_ids, estimates, raw)
            key = ("canonical-finish-setup", wb_op, finish_signature)
        elif scope == "assembly":
            key = ("canonical-event", decision_id)
        elif wb_op == "Robomac":
            key = ("canonical-event", decision_id)
        else:
            # Compatible leaf events share one tooling setup. Decision identities remain
            # attached individually, so no event disappears in the grouping.
            key = (
                "canonical-setup", wb_op, material.upper(),
                "%g" % thickness, "%g" % sequence if sequence is not None else "",
            )

        group = groups.setdefault(key, {
            "wb_op": wb_op,
            "material": material,
            "thickness": thickness,
            "group_key": key,
            "qty": 0,
            "bh": 0.0,
            "parts": [],
            "targets": [],
            "bends": 0,
            "holes": 0,
            "engine_ops": [],
            "decision_ids": [],
            "canonical_route": True,
            "assembly_scoped": scope == "assembly",
            "route_sequence": sequence,
        })
        if decision_id not in group["decision_ids"]:
            group["decision_ids"].append(decision_id)
        if operation not in group["engine_ops"]:
            group["engine_ops"].append(operation)
        if target_id and target_id not in group["targets"]:
            group["targets"].append(target_id)
        for part_number in participants or ([target_id] if target_id else []):
            if part_number and part_number not in group["parts"]:
                group["parts"].append(part_number)
        group["qty"] += qty
        if sequence is not None:
            group["route_sequence"] = (
                sequence if group.get("route_sequence") is None
                else min(float(group["route_sequence"]), sequence)
            )

        if operation == "hardware_insertion":
            insert_count = sum(node_qty.get(item, 1) for item in participants)
            group["qty"] = qty
            group["bh"] += (
                float(order_qty) * insert_count
                * float(_MANM_INSERT_SECONDS_EACH) / 3600.0
            )
            group["work_units"] = insert_count
            continue

        # Geometry-derived batch hours are valid for leaf events. Assembly work uses the
        # department throughput because summing participant hours is the old over-count.
        if scope == "part" and representative_id:
            batch_hours = (
                (estimates[representative_id].get("labour_estimate") or {})
                .get("batch_hours") or {}
            )
            normalised_hours = {
                str(name).strip().lower(): value
                for name, value in batch_hours.items()
            }
            hours = _safe(normalised_hours.get(operation))
            if hours and hours > 0:
                group["bh"] += hours

            geometry = estimates[representative_id].get("normalized_geometry") or {}
            if operation == "folding":
                group["bends"] += int(_safe(
                    geometry.get("estimated_bend_line_count"), 0) or 0) * int(qty)
            elif operation in {"hole_machining", "drilling", "punch"}:
                group["holes"] += int(_safe(
                    geometry.get("estimated_hole_count"), 0) or 0) * int(qty)

    for group in groups.values():
        group["decision_ids"] = sorted(set(group.get("decision_ids") or []))
        group["decision_id"] = (
            group["decision_ids"][0]
            if len(group["decision_ids"]) == 1 else None
        )
    return groups


def _is_tube(pe: Dict[str, Any]) -> bool:
    """A part is tube if its description/geometry indicates section/tube stock
    (no flat-pattern DXF; priced by length). Heuristic: 'tube' in desc, or the
    part has a length but no width (a bar), or explicit geometry_source hints."""
    desc = f"{pe.get('part_number','')} {pe.get('description','')}".upper()
    if "TUBE" in desc or "SECTION" in desc or "BOX SECTION" in desc:
        return True
    return False



def _verify_template_matches_cellmap(ws, cm, flags=None):
    """The template is the single source of truth for its own layout.

    M92 (Total Material Cost) sums every material block:
        =(SUM(M11:M50)+SUM(M53:M60)+SUM(M63:M81)+SUM(M84:M91)+AF83)
    Read it back and check CELL_MAP agrees. If the estimators re-shape the sheet again,
    we find out HERE — not by writing into merged header cells, and not by silently
    truncating a block.

    Non-fatal if the formula cannot be found or parsed: we fall back to the constants.
    """
    import re as _re
    try:
        f = ws["M92"].value
        if not isinstance(f, str) or "SUM(" not in f:
            return
        spans = _re.findall(r"SUM\(M(\d+):M(\d+)\)", f)
        if not spans:
            return
        found = {int(a): int(b) for a, b in spans}
        for key in ("bom", "tube", "steel", "other_sheet"):  # CELL_MAP names it "tube"
            blk = cm.get(key)
            if not blk:
                continue
            fr, lr = blk["first_row"], blk["last_row"]
            if fr in found and found[fr] != lr:
                raise RuntimeError(
                    f"[wb_populate] TEMPLATE/CELL_MAP MISMATCH in '{key}': "
                    f"CELL_MAP says rows {fr}..{lr}, but the template's M92 formula sums "
                    f"M{fr}:M{found[fr]}. The template layout changed. "
                    f"Update CELL_MAP — do NOT run with a stale map."
                )
    except RuntimeError:
        raise
    except Exception:
        return  # never let the safety check itself break a run


def _make_material_total_error_tolerant(ws, flags=None):
    """Rewrite the Total Material Cost formula (M92) so an errored material row does not
    blank the whole total.

    The template's M92 is a plain SUM across the material blocks:
        =(SUM(M11:M50)+SUM(M53:M60)+SUM(M63:M81)+SUM(M84:M91)+AF83)
    A single #VALUE!/#DIV/0! in any summed cell (a steel part with no blank dims) makes the
    whole SUM error, which cascades into Unit Cost and Sell Price. We replace each SUM(range)
    with AGGREGATE(9,6,range) — 9 = SUM, 6 = ignore errors — so the total is the sum of the
    CREDIBLE rows and the errored rows contribute nothing until the estimator dimensions them.

    NON-REGRESSIVE: AGGREGATE(9,6,range) is arithmetically identical to SUM(range) whenever the
    range has no errors, so fully-dimensioned jobs (12120/1282) are unchanged. Bare cell terms
    (e.g. +AF83, the powder total) are left as-is — they are not the cascade source. Best-effort:
    if M92 is not the expected shape, leave it untouched and flag.
    """
    import re as _re
    try:
        cell = ws["M92"]
        f = cell.value
        if not isinstance(f, str) or "SUM(" not in f:
            if flags is not None:
                _flag("material-total error-tolerance: M92 is not the expected SUM formula — "
                      "left as-is (a missing dim will still #VALUE! the total).", flags)
            return False
        # SUM(  ->  AGGREGATE(9,6,   for every SUM( in the formula. Ranges/bare refs unchanged.
        # AGGREGATE is an Excel 2010+ "future function": openpyxl MUST store it with the
        # _xlfn. prefix or Excel shows #NAME? (it reads bare AGGREGATE as an unknown name).
        # Excel displays/evaluates _xlfn.AGGREGATE(...) as AGGREGATE(...).
        new_f = _re.sub(r"\bSUM\(", "_xlfn.AGGREGATE(9,6,", f)
        if new_f == f:
            return False
        cell.value = new_f
        if flags is not None:
            _flag("material total made error-tolerant (SUM->AGGREGATE ignore-errors): a part "
                  "with no blank dims no longer blanks Total Material/Unit/Sell — it totals the "
                  "credible lines and self-completes when the flagged dims are filled.", flags)
        return True
    except Exception:
        return False  # never let this optional hardening break a run


def _clean_error_cells(ws, flags=None):
    """Wrap the formula-driven output cells in the Sheet-Steel / Other-Sheet rows (Qty Per Sheet,
    Cost Per Part, and the row-aligned Laser/CNC/Powder rate-calculator cells) in IFERROR(...,"")
    so a part with no blank dims shows BLANK cells instead of #DIV/0!/#VALUE! scattered across the
    sheet. NON-REGRESSIVE: IFERROR(valid_formula,"") == valid_formula when there is no error, so a
    fully dimensioned row is untouched — and the row SELF-HEALS: the instant the estimator fills
    the dims, the formula evaluates and the value appears (and the error-tolerant material total
    picks it up). The 'DIMS REQUIRED' description marker still names which rows to complete.
    """
    wrapped = 0
    try:
        blocks = [CELL_MAP.get("steel"), CELL_MAP.get("other_sheet")]
        for blk in blocks:
            if not blk:
                continue
            fr, lr = blk["first_row"], blk["last_row"]
            for r in range(fr, lr + 1):
                # K(11)..AH(34): Qty Per Sheet, Cost Per Part, and the row's rate-calculator cells
                for c in range(11, 35):
                    cell = ws.cell(row=r, column=c)
                    v = cell.value
                    if isinstance(v, str) and v.startswith("=") and "IFERROR(" not in v.upper():
                        cell.value = f'=IFERROR({v[1:]},"")'
                        wrapped += 1
    except Exception:
        return 0  # never let cosmetic cleanup break a run
    if flags is not None and wrapped:
        _flag(f"cleaned {wrapped} error-prone formula cell(s) with IFERROR — not-yet-dimensioned "
              f"rows show blank, not #DIV/0!/#VALUE! (self-heal when dims are filled).", flags)
    return wrapped


def _flag(msg: str, flags: List[str]):
    flags.append(msg)
    print(f"   [wb_populate] ⚠ {msg}")


def route_group_id(wb_op: Any, material: Any, thickness: Any, part_numbers: Any = (),
                   group_key: Any = None) -> str:
    """A stable identity for a route group, independent of where it lands on the sheet.

    The sheet row is the join key WITHIN a run, and it is the right one — it is what Excel
    calculated against. But it is not stable BETWEEN runs: insert a row in the template, or
    let a group drop out because a part became bought-in, and every row below shifts. A
    baseline compared on row number then reports a change on every line and the real change
    is invisible.

    The identity is what the group IS — the operation, the material, the gauge, and the parts
    it covers. Same group, same id, whatever row it lands on and whatever job it belongs to.
    """
    import hashlib
    parts = sorted({str(p).strip().upper() for p in (part_numbers or []) if str(p).strip()})
    if group_key is not None:
        # THE GROUPING KEY IS THE IDENTITY. Deriving the id from the group's stored material
        # and thickness is wrong for the one-row-per-job departments, whose key is
        # (operation, "", ""): those fields hold whatever the FIRST part to reach the group
        # happened to carry, so reordering mixed-gauge P.Coat or Weld parts changed an id
        # that is supposed to be stable. The key is what actually decided the grouping.
        basis = "|".join(str(k or "").strip().lower() for k in
                         (group_key if isinstance(group_key, (list, tuple)) else [group_key]))
    else:
        _thk = _num_or_none(thickness)
        basis = "|".join((
            str(wb_op or "").strip().lower(),
            str(material or "").strip().lower(),
            f"{_thk:.3f}" if _thk is not None else "",
        ))
    return "rg_" + hashlib.sha1((basis + "||" + ",".join(parts)).encode("utf-8")).hexdigest()[:12]


def _num_or_none(v: Any) -> Optional[float]:
    try:
        f = float(str(v).strip())
    except (TypeError, ValueError):
        return None
    return f if f == f else None


def build_workbook_labour(
    groups: Any,
    skipped_part_numbers: Any = (),
    canonical_mode: bool = False,
) -> Dict[str, Any]:
    """The CANONICAL ROUTE RECORD: the labour rows the workbook accepted, each keyed to the
    sheet row it was written to.

    Must be called AFTER the write loop, because `workbook_row` is stamped onto each group
    there. An earlier version built this record BEFORE the loop and then back-filled row
    numbers by matching on department name in insertion order — which is wrong the moment
    one department holds more than one group. On 12120 that swapped the 1.2mm and 1.5mm
    Fold rows, and the Laser rows with them, attaching each group's calculated cost to the
    other's parts. Reading the row straight off the group cannot do that.

    These rows are post-filter: after every spurious-op, finish and material gate, after
    department mapping, and including injected operations. Engine-side
    labour_estimate.costs_gbp is PRE-filter and must not be used to describe the job.

    Identity only. Hours, rates and values come from Excel via `final_estimate`, which joins
    to these rows on `workbook_row`.
    """
    _rows = [g for g in (groups.values() if isinstance(groups, dict) else (groups or []))
             if isinstance(g, dict) and g.get("workbook_row")]
    _canonical = bool(canonical_mode) or (
        bool(_rows) and all(g.get("canonical_route") for g in _rows)
    )
    return {
        "schema": "workbook_labour_rows.v3" if _canonical else "workbook_labour_rows.v2",
        "mode": "canonical" if _canonical else "legacy",
        "identity": ("route_group_id is stable across runs and template revisions; "
                     "workbook_row is the join key within THIS run's sheet."),
        "note": ("Labour rows as ACCEPTED by wb_populate, keyed to the sheet row each was "
                 "written to. Identity only — hours, rates and values come from Excel via "
                 "final_estimate, which joins to these on workbook_row."),
        "rows": [
            {
                "workbook_row": g.get("workbook_row"),
                "route_group_id": route_group_id(g.get("wb_op"), g.get("material"),
                                                 g.get("thickness"), g.get("parts"),
                                                 group_key=g.get("group_key")),
                "wb_operation": g.get("wb_op"),
                # Real engine operations, recorded when the group was formed. NOT the group
                # key: that holds the mapped department name.
                "engine_operations": list(g.get("engine_ops") or []),
                "engine_operation": (list(g.get("engine_ops") or []) or [None])[0],
                "decision_id": g.get("decision_id"),
                "decision_ids": list(g.get("decision_ids") or []),
                "material": g.get("material"),
                "thickness_mm": g.get("thickness"),
                "qty_per_unit": g.get("qty"),
                "part_numbers": list(g.get("parts") or []),
            }
            for g in sorted(_rows, key=lambda g: g["workbook_row"])
        ],
        "skipped_part_numbers": sorted(skipped_part_numbers or []),
    }


def populate_workbook(summary: Dict[str, Any], job_folder_name: str) -> Optional[str]:
    """Open the template, populate inputs from `summary`, save-as to output dir.
    `job_folder_name` is the drawing-folder basename, used for the output filename.
    Returns the output path, or None on failure."""
    if openpyxl is None:
        print("   [wb_populate] openpyxl not installed — cannot populate template.")
        return None

    cm = CELL_MAP
    tpl = template_path()
    if not os.path.exists(tpl):
        print(f"   [wb_populate] TEMPLATE NOT FOUND: {tpl}")
        print(f"   [wb_populate] the estimate will fall back to the xlsx_output builder, "
              f"whose labour rows are NOT route-grouped and therefore OVERSTATE assembly "
              f"operations. If the share is unreachable, point SDI_WB_TEMPLATE at a local "
              f"copy of the template and re-run.")
        return None

    flags: List[str] = []

    # part estimates live under estimate_summary.part_estimates (or top-level parts)
    pes = (
        (summary.get("estimate_summary") or {}).get("part_estimates")
        or summary.get("part_estimates")
        or summary.get("parts")
        or []
    )
    _canonical_cutover = canonical_route_cutover_enabled(summary)
    if _canonical_cutover:
        pes = canonicalise_part_estimates_for_workbook(summary, list(pes))
    _canonical_kinds = canonical_part_kinds(summary) if _canonical_cutover else {}
    if _canonical_cutover:
        canonical_route_payload(summary)["mode"] = "cutover"
        _flag(
            "CANONICAL BOM/ROUTE CUTOVER is active: hierarchy controls material placement "
            "and required OperationDecisions are the only source of labour rows.",
            flags,
        )

    wb = openpyxl.load_workbook(tpl, data_only=False, keep_vba=False)
    if cm["estimate_sheet"] not in wb.sheetnames:
        print(f"   [wb_populate] no '{cm['estimate_sheet']}' sheet in template.")
        return None
    ws = wb[cm["estimate_sheet"]]
    _verify_template_matches_cellmap(ws, cm, flags)

    # Powder £/kg — write the code-controlled rate into the sheet (cell AF82),
    # overwriting the template's static default. AF83 (=AD82*AF82) then computes
    # powder material cost at the correct rate. Source: config.POWDER_COST_PER_KG.
    if _POWDER_COST_PER_KG is not None:
        try:
            ws["AF82"] = float(_POWDER_COST_PER_KG)
        except Exception:
            pass

    # ── Header ──────────────────────────────────────────────────────────────
    hdr = cm["header"]
    order_qty = int(_safe(summary.get("assumed_job_quantity")
                          or summary.get("order_quantity")
                          or (summary.get("estimate_summary") or {}).get("assumed_job_quantity")
                          or 180) or 180)
    ws[hdr["customer"]]   = summary.get("customer") or summary.get("client") or job_folder_name
    ws[hdr["drawing_no"]] = re.match(r"\s*(\d+)", job_folder_name).group(1) if re.match(r"\s*(\d+)", job_folder_name) else job_folder_name
    ws[hdr["order_qty"]]  = order_qty

    # ── Classify parts into blocks — from the FULL part audit ──────────────
    # Rules, in order, based on the engine's own fields (confirmed by _classify_audit):
    #   1. PACKAGING / DELIVERY placeholders (£0.00)     -> DROP (estimator adds manually)
    #   2. weldment/assembly parent (suppressed flag, OR assembly-name + no material)
    #                                                    -> exclude from material (labour only)
    #   3. stock_form in (sheet, stated_weight) w/ geom  -> Steel block (WB computes L×W×gauge)
    #   4. stock_form == tube                            -> BOM (catalogue-priced section)
    #   5. page_roles 'bought_in'                        -> BOM (price in unit_cost_gbp)
    #   6. board material                                -> Other Sheet block
    #   7. has blank geometry (any sheet metal)          -> Steel block
    #   8. no role, no geom, has ext cost (BI- item)     -> BOM (price = ext / qty)
    STEEL_STOCK_FORMS = {"sheet", "stated_weight"}
    DROP_CODES = {"PACKAGING", "DELIVERY"}
    bom_parts, steel_parts, board_parts, weldment_parts, excluded = [], [], [], [], []
    wire_parts = []
    for pe in pes:
        pn = str(pe.get("part_number") or "").upper()
        canonical_kind = _canonical_kinds.get(pn)
        me = pe.get("material_estimate") or {}
        stock_form = str(me.get("stock_form") or "").lower()
        rflags = [str(f).lower() for f in (me.get("reliability_flags") or [])]
        roles = [str(r).lower() for r in (pe.get("page_roles") or [])]
        mat = str(pe.get("normalized_material") or me.get("material") or "").upper()
        blank_l = _safe(me.get("blank_length_mm"))
        unit_price = _safe(pe.get("unit_cost_gbp") or pe.get("unit_material_cost_gbp"))
        ext_total = _safe(pe.get("extended_total_cost_gbp"))
        cost_method = str(me.get("cost_method") or "").lower()
        name_is_assembly = bool(re.search(r"-GA\b|ASSEMBL|WELDMENT", pn)
                                or re.search(r"ASSEMBL|WELDMENT", str(pe.get("description") or "").upper())
                                or "weldment" in cost_method
                                or "children" in cost_method)

        # 1. commercial placeholders (PACKAGING / DELIVERY): write as blank-price
        #    (£0) BOM rows for the estimator to fill in, rather than dropping.
        #    Saves manual typing. Price is left blank on purpose — delivery/packaging
        #    cost is order-level (pallets, haulage share) and NOT derivable from the
        #    drawing, so the engine must not invent it; the BOM writer flags "no price".
        #    (Reverses the earlier "estimator adds manually" DROP, per estimator request.)
        if pn in DROP_CODES:
            bom_parts.append(pe)
            continue

        # 2. weldment/assembly parent — material carried by children, labour only
        if canonical_kind == "assembly":
            excluded.append(pe)
            print(f"   [wb_populate] excluded canonical assembly parent: {pn} "
                  "(material belongs to leaf children)")
        elif "weldment_parent_material_suppressed" in rflags:
            weldment_parts.append(pe)
        elif name_is_assembly and stock_form not in STEEL_STOCK_FORMS and "bought_in" not in roles:
            # assembly rollup (e.g. 1453-GA-C kick plate ASSEMBLY, 1455-C-101 weldment):
            # exclude from material to avoid double-counting its children.
            excluded.append(pe)
            print(f"   [wb_populate] excluded assembly rollup: {pn} "
                  f"(£{ext_total} carried by children)")
        # 2b. Stated-weight part with NO nesting geometry (e.g. a timber panel costed by its
        #     printed weight × £/kg). The area-based steel/board blocks can't display it — they
        #     read L×W and would show £0. Write it as a DIRECT-priced BOM line (the same proven
        #     path the tubes use) so its real per-part material cost reaches the sheet. Steel
        #     stated-weight parts that DO carry a blank still route to the Sheet Steel block below.
        elif stock_form == "stated_weight" and blank_l is None and _safe(me.get("cost_per_part_gbp")):
            bom_parts.append(pe)
        # 3. steel (sheet or stated-weight, with real geometry) — but NOT a board/timber part.
        #    A timber panel costed by stated weight has stock_form 'stated_weight' (in
        #    STEEL_STOCK_FORMS); without the board guard it would be dumped in Sheet Steel and
        #    metal-lasered. Board/timber materials fall through to the board block below.
        elif stock_form in STEEL_STOCK_FORMS and not _is_board(mat):
            steel_parts.append(pe)
        # 3b. wire / round bar -> Wire block (WB prices it from gauge + length)
        elif stock_form == "wire":
            wire_parts.append(pe)
        # 4. tube — catalogue-priced section -> BOM
        elif stock_form == "tube":
            bom_parts.append(pe)
        # 5. priced bought-in
        elif "bought_in" in roles:
            bom_parts.append(pe)
        # 6. board / other sheet
        elif _is_board(mat):
            board_parts.append(pe)
        # 7. sheet metal with geometry but no stock_form set
        elif _is_sheet_metal(mat) and blank_l is not None:
            steel_parts.append(pe)
        # 8. no role, no geometry, has a cost -> a bought-in costed via the full path
        elif unit_price is None and ext_total is not None and blank_l is None:
            bom_parts.append(pe)
        else:
            _flag(f"part {pe.get('part_number')} unclassifiable "
                  f"(stock_form={stock_form!r}, role={roles}, unit={unit_price}, "
                  f"ext={ext_total}, blankL={blank_l}) — skipped.", flags)

    # ── BOM block: desc, code, PRICE (engine), qty, scrap ──────────────────
    # Holds true bought-ins (fixings, vinyl, electricals) AND tube sections
    # (catalogue-priced £/EA). All have an engine-sourced price we write directly.
    # ── Powder on WIRE / BAR — the area the workbook cannot see ─────────────────
    # The Powder Qty Calculator sums SHEET area over the Sheet Steel block. Wire and bar
    # live in a different block with no length x width, so they have ALWAYS contributed
    # zero powder. It only surfaced when a job (7670) was entirely wire and the calculator
    # returned a clean 0 against Tim's £0.40.
    #
    # A wire is a CYLINDER: its whole surface is coated, so area = pi * d * L. (No x2 —
    # that is a flat sheet, which has two faces.) Gauge and length are already on the
    # sheet; we read them off the PDF.
    #
    # Read from `summary`, NOT from wire_parts: that list may not be built yet at this
    # point in the file, and an empty loop would silently produce 0.0 and drop the powder
    # line without a word. `summary` is in scope everywhere.
    #
    # Sheet powder stays in the workbook's own AF82/AF83. This adds only what that
    # calculation cannot see, so nothing is double-counted.
    _wire_powder_area_m2 = 0.0
    _wire_powder_diag = []
    _all_pes_pw = ((summary.get("estimate_summary") or {}).get("part_estimates")
                   or summary.get("parts") or [])
    for _wp in _all_pes_pw:
        _wme = _wp.get("material_estimate") or {}
        if str(_wme.get("stock_form") or "").lower() not in ("wire", "bar"):
            continue
        _wg = _safe(_wme.get("wire_gauge_mm") or _wp.get("wire_gauge_mm"))
        _wl = _safe(_wme.get("wire_length_mm") or _wp.get("wire_length_mm"))
        _wq = _safe(_wp.get("quantity"), 1) or 1
        _wire_powder_diag.append(f"{_wp.get('part_number')}(g={_wg},l={_wl},q={_wq})")
        if _wg and _wl:
            _wire_powder_area_m2 += 3.14159265 * (_wg / 1000.0) * (_wl / 1000.0) * float(_wq)
    # ── SHEET AREA TOO ──────────────────────────────────────────────────────────
    # The template's calculator sees only the Sheet Steel block. Now that powder is a BOM
    # LINE and no longer comes from that calculator, we must compute the sheet area here
    # as well — otherwise moving it to the BOM would silently drop it.
    #
    # Two faces, because a sheet has two sides. For 1310-01 this gives
    #     167.04 x 113 x 2 / 1e6 = 0.03775 m2
    # which matches the template's own AB63 to five decimal places. The area maths agrees
    # with theirs; only the coverage RATE is in dispute.
    _sheet_powder_area_m2 = 0.0
    # acrylic_excluded_from_powder (2026-07-15): acrylic / perspex / PMMA / polycarbonate are
    # diamond polished, NEVER powder coated. An acrylic part is a sheet form, so without this
    # guard its area was summed into the coated total and a phantom POWDER BOM row was written
    # (12439). Exclude acrylic/plastic from the coated-area sum. Steel parts are unaffected.
    _ACR_NO_POWDER = {"ACRYLIC", "HIGH IMPACT ACRYLIC", "PERSPEX", "PMMA", "POLYCARBONATE"}
    def _is_acrylic_pw(_p):
        _m = str(_p.get("normalized_material")
                 or (_p.get("material_estimate") or {}).get("material") or "").upper().replace("_", " ")
        return _m in _ACR_NO_POWDER or bool(_p.get("acrylic_no_powder"))
    for _sp in _all_pes_pw:
        _sme = _sp.get("material_estimate") or {}
        if str(_sme.get("stock_form") or "").lower() not in ("sheet", "plate", "stated_weight", ""):  # include stated_weight: it is coated steel routed by weight, must not drop from the powder sum (aligns with STEEL_STOCK_FORMS routing filter). Powder basis stays GROSS L x W; this only keeps steel parts in the sum when a valid blank area flips them onto the weight path.
            continue
        if _is_acrylic_pw(_sp):
            continue   # acrylic is not powder coated — contributes zero coated area
        # A part named TUBE is a hollow section, NOT a flat coated sheet: its blank (if any) is
        # garbled view geometry, and its true coated area is a thin cylinder surface, not L×W×2.
        # Without this guard a tube the LLM missed (e.g. 10M read as a 2431×2431 blank) injects
        # ~24 m² of phantom coated area and dominates the powder line. Skip it from the sheet sum.
        _sdesc = str(_sp.get("part_description") or _sp.get("description")
                     or _sme.get("description") or "").upper()
        if "TUBE" in _sdesc:
            continue
        _sng = _sp.get("normalized_geometry") or {}
        _sl = _safe(_sme.get("blank_length_mm") or _sng.get("blank_length_mm"))
        _sw = _safe(_sme.get("blank_width_mm") or _sng.get("blank_width_mm"))
        _sq = _safe(_sp.get("quantity"), 1) or 1
        # Sanity guard: a single fabricated part on a retail display is never a >3.5 m² flat blank
        # (2431×2431 = 5.9 m²). An area that large is garbled PDF view geometry — exclude it from
        # the coated sum rather than let one bad blank invent the whole powder cost.
        if _sl and _sw and ((_sl / 1000.0) * (_sw / 1000.0)) > 3.5:
            continue
        if _sl and _sw:
            _sheet_powder_area_m2 += (_sl / 1000.0) * (_sw / 1000.0) * 2.0 * float(_sq)

    # ── SECTION AREA TOO ────────────────────────────────────────────────────────
    # Sheet and wire were the only two contributors. A powder-coated tube, box section or
    # angle matched neither and contributed nothing at all — see section_coated_area_m2.
    _section_powder_area_m2 = 0.0
    _section_powder_diag = []
    for _cp in _all_pes_pw:
        _ca = section_coated_area_m2(_cp)
        if _ca > 0:
            _section_powder_area_m2 += _ca
            _section_powder_diag.append(f"{_cp.get('part_number')}({_ca:.5f} m2)")

    _powder_area_m2 = _sheet_powder_area_m2 + _wire_powder_area_m2 + _section_powder_area_m2
    if _section_powder_area_m2 > 0:
        _flag(f"POWDER: section stock adds {_section_powder_area_m2:.5f} m2 of coated surface "
              f"({', '.join(_section_powder_diag)}) — outer perimeter x cut length. Until now "
              f"section contributed ZERO to the coated total, since the sum saw only flat "
              f"blanks and wire.", flags)
    _powder_by_area_kg = _powder_area_m2 * float(_POWDER_KG_PER_M2)

    # ── A MINIMUM PER PIECE, NOT JUST A COVERAGE RATE ───────────────────────────
    # A "piece" is one object on the booth line. If the job welds, the components become
    # ONE object — the rule the estimators gave us on 1310 this morning. Otherwise each
    # fabricated part hangs on its own.
    _mw_pw2 = (summary.get("manufacturing_writeup") or {}).get("parts") or []
    _job_welds_pw = any(
        "weld" in str(_o).lower()
        for _m in _mw_pw2
        for _o in (_m.get("textual_operations") or _m.get("operations") or [])
    )
    _fab_pieces = 0
    for _fp in _all_pes_pw:
        _fsf = str((_fp.get("material_estimate") or {}).get("stock_form") or "").lower()
        if _is_acrylic_pw(_fp):
            continue   # acrylic is not coated — must not count toward the per-piece powder floor
        if _fsf in ("sheet", "plate", "wire", "bar", "board"):
            _fab_pieces += int(_safe(_fp.get("quantity"), 1) or 1)
    # powder_floor_zero_when_no_coated (2026-07-15): a job with ZERO coatable pieces gets
    # ZERO powder, not a forced minimum of one. The old max(1, _fab_pieces) guaranteed at
    # least one coated object even when nothing in the job is coatable (e.g. a pure-acrylic
    # job — acrylic is excluded from _fab_pieces upstream), so the per-piece floor booked
    # 0.03 kg of phantom powder. Now: no coatable parts -> 0 pieces -> floor 0 -> no powder.
    # Steel jobs are unaffected (they always have _fab_pieces >= 1).
    _coated_pieces = (0 if _fab_pieces == 0
                      else (1 if _job_welds_pw else _fab_pieces))
    _powder_by_floor_kg = _coated_pieces * float(_POWDER_MIN_KG_PER_PIECE)

    _powder_kg_total = round(max(_powder_by_area_kg, _powder_by_floor_kg), 5)

    # Largest fabricated part by area (m2) - the size proxy for throughput banding. Pack and
    # coat speed track the biggest part in the job, not the average. Reuses the same geometry
    # the powder area calc reads; known when the labour block runs, unlike the unit-cost cell.
    _max_part_area_m2 = 0.0
    for _ap in _all_pes_pw:
        _ame = _ap.get("material_estimate") or {}
        _asf = str(_ame.get("stock_form") or "").lower()
        _ang = _ap.get("normalized_geometry") or {}
        _aq = _safe(_ap.get("quantity"), 1) or 1
        _area = 0.0
        if _asf in ("sheet", "plate", "board", ""):
            _al = _safe(_ame.get("blank_length_mm") or _ang.get("blank_length_mm"))
            _aw = _safe(_ame.get("blank_width_mm") or _ang.get("blank_width_mm"))
            if _al and _aw:
                _area = (_al / 1000.0) * (_aw / 1000.0)
        elif _asf in ("wire", "bar"):
            _ag = _safe(_ame.get("gauge_mm") or _ame.get("diameter_mm"))
            _aln = _safe(_ame.get("length_mm") or _ame.get("cut_length_mm"))
            if _ag and _aln:
                _area = 3.14159265 * (_ag / 1000.0) * (_aln / 1000.0)
        if _area > _max_part_area_m2:
            _max_part_area_m2 = _area
    _powder_basis = ("MINIMUM PER PIECE" if _powder_by_floor_kg >= _powder_by_area_kg
                     else "COATED AREA")
    _wire_powder_kg = _powder_kg_total          # the BOM branch below reads this name

    if _powder_kg_total > 0:
        _flag(f"POWDER QUANTITY IS AN ASSUMPTION — estimator to confirm. "
              f"{_powder_area_m2:.4f} m2 coated x {_POWDER_KG_PER_M2} kg/m2 "
              f"= {_powder_by_area_kg:.4f} kg; floor of {_POWDER_MIN_KG_PER_PIECE} kg x "
              f"{_coated_pieces} coated object(s) = {_powder_by_floor_kg:.4f} kg. "
              f"BOOKED {_powder_kg_total} kg (the {_powder_basis}). "
              f"The template assumes 0.1667 kg/m2 = 100% transfer efficiency, which nothing "
              f"achieves; we use 0.20 (70um film, ~50% efficiency). The floor exists because "
              f"Tim's sheets show 25-40g on parts far too small to explain by area — you "
              f"cannot coat a 40mm hook with six grams. BOTH numbers are levers in config.py "
              f"(POWDER_KG_PER_M2, POWDER_MIN_KG_PER_PIECE) — tell us the real rule and it is "
              f"a one-line change.", flags)
    if _wire_powder_diag and _wire_powder_kg <= 0:
        # Never fail silently. If wire parts exist but the area is zero, say what they held.
        _flag(f"powder: found {len(_wire_powder_diag)} wire/bar part(s) but computed ZERO "
              f"coated area — {'; '.join(_wire_powder_diag)}. Powder NOT costed on the wire; "
              f"gauge/length missing from the pricing record.", flags)

    # ── POWDER IS A LINE ON THE BILL OF MATERIALS ───────────────────────────────
    # Tim writes it as one:   Powder | GBP 9.73 | 0.03 kg | 4% | GBP 0.30
    # We had it hidden in the Powder Qty Calculator and bolted onto M92 via AF83, so an
    # estimator reading the BOM saw nothing. That is Dave's "no powder allowed for".
    #
    # If the drawing NAMES a powder (7670's TLP-J125-T RYOBI GREEN) that row already exists
    # and gets the quantity. If the drawing names none (1310) the part is still coated, so
    # add a generic row rather than let the cost vanish.
    def _is_powder_row(_p):
        return ("POWDER" in str(_p.get("part_number") or "").upper()
                or "POWDER" in str(_p.get("description") or "").upper()
                or bool(_p.get("_consumable_qty_unknown")))

    if _powder_kg_total > 0 and not any(_is_powder_row(_p) for _p in bom_parts):
        bom_parts = list(bom_parts) + [{
            "part_number": "POWDER",
            "description": "Powder — computed from coated surface area "
                           f"({_powder_area_m2:.4f} m2)",
            "quantity": 1,
            "_price_explicitly_withheld": True,   # routes through the consumable branch,
            "_consumable_qty_unknown": True,      # which sets qty = kg and price = GBP/kg
            "_catalogue_rate_gbp": float(_POWDER_COST_PER_KG or 9.73),
        }]

    # Kill the template's own powder term. M92 adds AF83 (= total kg x GBP/kg) on top of the
    # material blocks. Powder is now a BOM row inside SUM(M11:M50), so leaving AF83 alive
    # would charge it TWICE.
    try:
        ws["AF83"] = 0
    except Exception:
        pass

    b = cm["bom"]


    # ── EVERY FABRICATED PART APPEARS IN THE BILL OF MATERIALS ─────────────────────
    #
    # The template costs a nested sheet part in the Sheet Steel block and a board panel in
    # Other Sheet, so neither has ever appeared in the "Bill of Materials (Per Unit)" list.
    # That is arithmetically correct and practically wrong: on M&S 2085 an estimator opens
    # the sheet, reads the bill of materials, and the main part of the job — the bracket
    # plate everything else welds into — is not in it. You have to know the template's
    # internals to work out that it is costed twenty rows further down.
    #
    # So they are listed, and listed FIRST, with the block carrying their cost named on the
    # line. The price is deliberately ZERO: Total Material Cost (M92) sums BOM + Wire +
    # Sheet Steel + Other Sheet, so a priced duplicate here would double that part's
    # material and the sheet would be wrong in a way nobody would spot.
    _xref_rows: List[Dict[str, Any]] = []
    for _blk_name, _blk in (("Sheet Steel", steel_parts), ("Other Sheet Material", board_parts),
                            ("Wire", wire_parts)):
        for _xp in _blk:
            if not isinstance(_xp, dict) or not _xp.get("part_number"):
                continue
            _xref_rows.append({
                "part_number": _xp.get("part_number"),
                "description": f"{_xp.get('description') or ''} — costed in {_blk_name} below",
                "quantity": int(_safe(_xp.get("quantity"), 1) or 1),
                "unit_cost_gbp": 0.0,
                "_bom_cross_reference": True,
            })
    if _xref_rows:
        bom_parts = _xref_rows + list(bom_parts)
        _flag(f"BOM: {len(_xref_rows)} fabricated part(s) listed in the Bill of Materials for "
              f"completeness at GBP 0.00 — their material is costed in the Sheet Steel / Other "
              f"Sheet / Wire block, and pricing them here too would double it.", flags)

    # ── BOM overflow: spill-in-code, no template surgery, no lines dropped ──────────
    # The template's BOM block holds a fixed number of rows. A big job (e.g. Cocktails,
    # 48 bought-in/tube lines) exceeds it. Rather than fail the whole sheet (dropping to
    # the legacy builder) or silently drop lines, write the first (_n_rows - 1) items
    # individually and turn the FINAL BOM row into a consolidated 'overflow' line that
    # carries the summed value of the remainder — so the material total stays correct —
    # while every spilled line is itemised in full on a dedicated 'BOM Overflow' sheet.
    _n_rows = b["last_row"] - b["first_row"] + 1
    _bom_overflow_parts: List[Dict[str, Any]] = []
    if len(bom_parts) > _n_rows:
        _bom_overflow_parts = bom_parts[_n_rows - 1:]
        _ov_total = 0.0
        for _op in _bom_overflow_parts:
            _pp = _bom_line_price(_op)
            if _pp is not None:
                _ov_total += _pp * int(_safe(_op.get("quantity"), 1) or 1)
        _consolidated = {
            "part_number": "BOM-OVERFLOW",
            "description": (f"+{len(_bom_overflow_parts)} more bought-in items "
                            f"(itemised on 'BOM Overflow' sheet)"),
            "quantity": 1,
            "unit_cost_gbp": round(_ov_total, 2),
            "_bom_overflow_consolidated": True,
        }
        bom_parts = bom_parts[:_n_rows - 1] + [_consolidated]
        _flag(f"BOM overflow: {len(_bom_overflow_parts) + _n_rows - 1} BOM/tube lines > "
              f"{_n_rows} template rows. {len(_bom_overflow_parts)} spilled to the "
              f"'BOM Overflow' sheet and consolidated (£{_ov_total:.2f}) on the last BOM "
              f"row — no lines dropped.", flags)

    row = b["first_row"]
    for pe in bom_parts:
        if row > b["last_row"]:
            break
        me = pe.get("material_estimate") or {}
        se = me.get("stock_estimate") or {}
        # description: prefer THIS part's OWN identity (number + description). The catalogue
        # section-match can borrow a *different job's* tube row (e.g. 11406-02-02M / TUBE0173) when
        # this tube's own dims weren't extracted — showing another job's part number is misleading.
        # Keep the catalogue's size text only as a parenthetical REFERENCE, never as the identity.
        _own_pn = pe.get("part_number")
        _own_desc = pe.get("description")
        _cat_desc = se.get("catalogue_description")
        if _own_pn or _own_desc:
            desc = f"{_own_pn or ''}  {_own_desc or ''}".strip()
            # append catalogue size ref (helpful for tubes) without borrowing the foreign identity
            if _cat_desc and se.get("catalogue_part_code"):
                # strip any leading "ITEM n - <foreign pn> - " so only the size/length remains
                import re as _re
                _size = _re.sub(r"^\s*ITEM\s*\d+\s*-\s*[\w-]+\s*-\s*", "", str(_cat_desc)).strip()
                _size = _re.sub(r"\s*-\s*LASER TUBE\s*$", "", _size, flags=_re.I).strip()
                if _size:
                    desc = f"{desc} (cat ref: {_size})"
        else:
            desc = _cat_desc or _own_pn
        # code: THIS part's own number; fall back to catalogue code only if the part has none
        code = _own_pn or se.get("catalogue_part_code")
        # supplier: engine may put it at top level OR in material_estimate
        supplier = pe.get("supplier") or me.get("supplier") or ""
        # WHERE THIS PRICE CAME FROM, ON THE SHEET ITSELF. The supplier column was blank on
        # every BOM line, so a catalogue price and an AI market estimate looked identical to
        # the estimator reading the workbook — and on this job the guessed lines were 96% of
        # material, with one knob quoted at GBP 1.25, GBP 11.52 and GBP 1.77 on three runs of
        # the same inputs. A figure nobody can reproduce must not sit in a price column
        # looking like a quote.
        _origin, _indicative = _price_origin(pe)
        if _origin and not supplier:
            supplier = _origin
        if _indicative:
            # Also on the description, because the supplier column is narrow and this is the
            # one thing a reader must not miss. The cell is truncated to 120 characters on
            # write, and the tag is appended LAST — so on a long description the warning is
            # the first thing that would be cut. Trim the description instead: a shortened
            # part name is a small loss, a silently dropped warning is the whole point gone.
            _room = 120 - len(_INDICATIVE_TAG) - 2
            desc = f"{str(desc)[:_room].rstrip()}  {_INDICATIVE_TAG}"
            _flag(f"BOM {code or desc}: priced by an AI market estimate, not a catalogue. "
                  f"The figure changes every run and is INDICATIVE ONLY — confirm before "
                  f"quoting.", flags)
        # price: bought-in unit_cost_gbp, or tube's unit_material_cost_gbp (catalogue £/EA),
        # or — for BI-/fuller-costed items — derive unit from extended_total_cost_gbp / qty
        qty = int(_safe(pe.get("quantity"), 1))
        # ── "Not priced" is a DECISION, not a missing value ──────────────────────────
        # The chain below is built to KEEP LOOKING until it finds a number: four candidate
        # fields, then a division as a last resort. That is right for a part whose price is
        # merely recorded somewhere unexpected. It is WRONG for a part the estimator layer
        # has DELIBERATELY refused to price.
        #
        # 7670: the engine correctly refused to invent a powder quantity (a consumable is
        # sold by weight — "assume 1" means 1kg, and 1kg would coat 6 m2 of a 0.023 m2 wire
        # frame). It cleared unit_cost_gbp. The chain moved to unit_material_cost_gbp, found
        # £7.72 still sitting there, and put the £8.03 straight back on the sheet.
        #
        # Honour the marker and short-circuit. Whatever stale numbers survive elsewhere on
        # the record, unpriced means unpriced.
        if pe.get("_price_explicitly_withheld"):
            _cat_rate = _safe(pe.get("_catalogue_rate_gbp"))
            _is_consumable_line = bool(pe.get("_consumable_qty_unknown")) or \
                                  "POWDER" in str(pe.get("part_number") or "").upper()
            if _is_consumable_line and _cat_rate and _wire_powder_kg > 0:
                # We withheld because we could not know the QUANTITY. We now can: the wire's
                # coated area is real geometry (pi x d x L) that the workbook's sheet-only
                # calculator cannot see. So cost it — and say exactly how, and how wrong the
                # rate is.
                price = _cat_rate
                qty = _powder_kg_total
                _flag(f"POWDER computed from WIRE geometry: {_wire_powder_area_m2:.5f} m2 of "
                      f"coated surface (pi x dia x length) x {_POWDER_KG_PER_M2} kg/m2 = "
                      f"{_wire_powder_kg} kg @ £{_cat_rate:.2f}/kg. "
                      f"COVERAGE RATE IS THE TEMPLATE'S {_POWDER_KG_PER_M2} kg/m2 = 100% "
                      f"TRANSFER EFFICIENCY, which nothing achieves. Tim's sheet for THIS job "
                      f"implies 1.70 kg/m2 (an open wire frame lets most of the cloud "
                      f"through) — about 10x this. His sheets imply 2.7x-4.9x even on FLAT "
                      f"parts. THIS LINE UNDER-READS until the rate is measured "
                      f"(config.POWDER_KG_PER_M2). Estimator to check.", flags)
            else:
                price = None
                _flag(f"BOM {pe.get('part_number') or (str(desc)[:30])}: price WITHHELD by the "
                      f"engine (quantity not on the drawing and cannot be guessed). Row is on "
                      f"the sheet with its code and supplier — ESTIMATOR TO PRICE. Not an "
                      f"error.", flags)
        else:
            # ONE price chain, not two. This was a hand-copy of _bom_line_price with the
            # same whole-total fallback, so fixing the helper alone would have left the
            # SHEET unchanged and only corrected the overflow sum nobody sees. That is the
            # test-the-caller-not-the-helper trap, in the code rather than in a fixture.
            price = _bom_line_price(pe)
        ws.cell(row=row, column=b["col_desc"],     value=str(desc)[:120])
        ws.cell(row=row, column=b["col_code"],     value=code)
        ws.cell(row=row, column=b["col_supplier"], value=supplier)
        ws.cell(row=row, column=b["col_price"],    value=price if price is not None else None)
        ws.cell(row=row, column=b["col_qty"],      value=qty)
        ws.cell(row=row, column=b["col_scrap"],    value=0.04)  # 4% default; WB applies
        if price is None:
            _flag(f"BOM item {pe.get('part_number')} has no price — line will be £0.", flags)
        row += 1

    # Itemise every spilled BOM line on a dedicated sheet so nothing is hidden behind the
    # consolidated overflow row. Failure-isolated: a sheet-write error never breaks the run.
    if _bom_overflow_parts:
        try:
            _ov_ws = wb.create_sheet("BOM Overflow")
            _ov_ws.append(["Part code", "Description", "Supplier", "Unit Price (GBP)", "Qty"])
            for _op in _bom_overflow_parts:
                _ome = _op.get("material_estimate") or {}
                _ov_ws.append([
                    _op.get("part_number") or "",
                    str(_op.get("description") or "")[:120],
                    _op.get("supplier") or _ome.get("supplier") or "",
                    _bom_line_price(_op),
                    int(_safe(_op.get("quantity"), 1) or 1),
                ])
        except Exception:
            pass

    # ── Wire block: desc, qty, gauge, length ───────────────────────────────
    # Round bar / stud / wire. The WB prices it itself:
    #   gauge -> Metres Per Tonne (lookup) -> Price Per Metre = L3 / m-per-tonne
    #   cost  = (length_mm / 1000) * price_per_m * qty * (1 + scrap)
    # 8mm dia: 2534 m/t, L3 £1600/t -> £0.6314/m. 65mm -> £0.041. Tim's sheet: £0.04.
    # This block existed in the template from day one and had NEVER been populated.
    #
    # (Tube sections stay in the BOM block above — they are catalogue-priced £/EA, not
    #  costed from gauge+length like bar stock.)
    if wire_parts:
        w = cm["tube"]
        row = w["first_row"]
        for pe in wire_parts:
            if row > w["last_row"]:
                _flag(f"Wire overflow: {len(wire_parts)} wire/bar parts but only "
                      f"{w['last_row']-w['first_row']+1} rows — extras DROPPED.", flags)
                break
            me = pe.get("material_estimate") or {}
            gauge = _safe(me.get("wire_gauge_mm") or pe.get("wire_gauge_mm"))
            length = _safe(me.get("wire_length_mm") or pe.get("wire_length_mm"))
            qty = int(_safe(pe.get("quantity"), 1))
            desc = f"{pe.get('part_number') or ''}  {pe.get('description') or ''}".strip()
            if gauge is None or length is None:
                _flag(f"wire part {pe.get('part_number')} missing gauge/length "
                      f"(gauge={gauge}, length={length}) — line will be £0.", flags)
            ws.cell(row=row, column=w["col_desc"],   value=str(desc)[:120])
            ws.cell(row=row, column=w["col_qty"],    value=qty)
            ws.cell(row=row, column=w["col_gauge"],  value=gauge)
            ws.cell(row=row, column=w["col_length"], value=length)
            row += 1

    # part_number -> the Sheet Steel row it was written to. Used below so the Laser
    # labour row can reference the template's OWN rate calculator instead of our model.
    _steel_row_by_pn = {}
    # Parts excluded from fabrication (empty detail/callout artefacts): keep them
    # out of the Sheet Steel block (blank gauge -> #DIV/0!) AND the labour groups.
    _skip_pns = set()

    # ── Steel block: desc, qty, length, width, gauge ───────────────────────
    s = cm["steel"]
    row = s["first_row"]
    for _si, pe in enumerate(steel_parts):
        if row > s["last_row"]:
            # Cannot widen: the steel rows are wired 1:1 into hidden laser/CNC rate calculators
            # with row-locked absolute refs + per-row merged cells (a prior widen was reverted for
            # 'failed MergedCell'). So make the overflow LOUD ON THE SHEET instead of silently
            # dropping parts: overwrite the last steel row's description with a marker naming the
            # count and the dropped part numbers, so the estimator sees the BOM is incomplete.
            _dropped_pns = [str(x.get("part_number") or x.get("description") or "?")
                            for x in steel_parts[_si:]]
            _n = len(_dropped_pns)
            ws.cell(row=s["last_row"], column=s["col_desc"],
                    value=f"⚠ +{_n} STEEL PART(S) NOT SHOWN (block full): "
                          f"{', '.join(_dropped_pns)} — see Decision Report")
            _flag(f"Steel overflow: {len(steel_parts)} steel parts but only "
                  f"{s['last_row']-s['first_row']+1} rows — {_n} shown on sheet as LOUD marker "
                  f"in last row (not silently dropped): {', '.join(_dropped_pns)}", flags)
            break
        me = pe.get("material_estimate") or {}
        ng = pe.get("normalized_geometry") or {}
        # Guard: an empty detail/callout artefact (no gauge, no geometry, no DXF)
        # is not a fabricatable steel part. It must not reach the Sheet Steel cost
        # row (blank gauge -> #DIV/0!). Skip it here and record it so the labour
        # loop drops its phantom ops too. Catches e.g. 'D-M4' regardless of how it
        # entered the part-estimate flow (upstream false-part filter missed it).
        _pn_g = str(pe.get("part_number") or "")
        _len_g = _safe(me.get("blank_length_mm") or ng.get("blank_length_mm"))
        _wid_g = _safe(me.get("blank_width_mm")  or ng.get("blank_width_mm"))
        _gau_g = _safe(pe.get("normalized_thickness_mm") or me.get("thickness_mm"))
        _dxf_g = pe.get("geometry_source") in ("dxf_flat_pattern", "solidworks_flat_pattern")
        if (not _gau_g) and (not _len_g) and (not _wid_g) and (not _dxf_g):
            _skip_pns.add(_pn_g)
            _flag("excluded non-fabricatable part '" + _pn_g + "' from Sheet Steel "
                  "(no gauge, no geometry, no DXF) - detail/callout artefact, "
                  "estimator to verify", flags)
            continue
        length = _safe(me.get("blank_length_mm") or ng.get("blank_length_mm"))
        width  = _safe(me.get("blank_width_mm")  or ng.get("blank_width_mm"))
        gauge  = _safe(pe.get("normalized_thickness_mm") or me.get("thickness_mm"))
        ws.cell(row=row, column=s["col_desc"],   value=f"{pe.get('part_number','')}  {pe.get('description','')}")
        ws.cell(row=row, column=s["col_qty"],    value=int(_safe(pe.get("quantity"), 1)))
        ws.cell(row=row, column=s["col_length"], value=length)
        ws.cell(row=row, column=s["col_width"],  value=width)
        ws.cell(row=row, column=s["col_gauge"],  value=gauge)
        # Which row did this part land on? The template's own Laser Rate Calculator
        # computes a throughput on THIS row (col W = 3600/V). The labour block should
        # READ that, not substitute our own model — ours is ~4x slow on small parts.
        _steel_row_by_pn[str(pe.get("part_number") or "")] = row
        # Laser-calc inputs S (No of holes) and T (Internal Cutting Distance).
        # Drawing-derived; feeds the sheet's laser calculator display. Honest gaps:
        # blank (not 0) when not read, so a genuine no-hole part is not a false claim.
        # Canonical geometry lives in pe["geometry_rollup"] (NOT "geometry").
        _geom = pe.get("geometry_rollup") or pe.get("normalized_geometry") or {}
        _holes = _geom.get("estimated_hole_count")
        if _holes is None:
            _holes = _geom.get("hole_count")
        if isinstance(_holes, (int, float)) and int(_holes) > 0:
            ws.cell(row=row, column=s["col_holes"], value=int(_holes))
        _cutlen = _safe(_geom.get("estimated_cut_length_mm") or _geom.get("cut_length_mm"))
        if length and width and _cutlen:
            _internal = round(max(0.0, float(_cutlen) - 2.0 * (float(length) + float(width))), 1)
            ws.cell(row=row, column=s["col_internal_cut"], value=_internal)
            if _internal > 0:
                _flag(f"steel {pe.get('part_number')}: internal-cut T={_internal}mm is DERIVED "
                      f"(cut_len {_cutlen} - bounding perim); overshoots on complex profiles.", flags)
        elif _holes:
            _flag(f"steel {pe.get('part_number')}: {int(_holes or 0)} holes read but internal-cut T "
                  f"not derivable (missing L/W/cut_len) — left blank, not 0.", flags)
        if not (length and width and gauge):
            _flag(f"steel {pe.get('part_number')} missing dim(s) "
                  f"(L={length} W={width} G={gauge}) — WB cost will be 0/wrong.", flags)
            # Loud, on-sheet marker so the estimator sees exactly which rows to complete —
            # the row keeps its formula, so filling L/W/G recomputes it and the (error-tolerant)
            # material total picks it up automatically. Which dim is missing is named.
            _miss = ", ".join(
                _n for _n, _v in (("L", length), ("W", width), ("gauge", gauge)) if not _v
            )
            _dcell = ws.cell(row=row, column=s["col_desc"])
            _dcur = str(_dcell.value or "")
            if "DIMS REQUIRED" not in _dcur:
                _dcell.value = f"{_dcur}  ⚠ DIMS REQUIRED ({_miss}) — not costed"
        row += 1

    # ── Other Sheet block: desc, qty, length, width, thickness ─────────────
    o = cm["other_sheet"]
    row = o["first_row"]
    for pe in board_parts:
        if row > o["last_row"]:
            _flag(f"Other-sheet overflow: {len(board_parts)} board parts, block full — extras DROPPED.", flags)
            break
        me = pe.get("material_estimate") or {}
        ng = pe.get("normalized_geometry") or {}
        length = _safe(me.get("blank_length_mm") or ng.get("blank_length_mm"))
        width  = _safe(me.get("blank_width_mm")  or ng.get("blank_width_mm"))
        thick  = _safe(pe.get("normalized_thickness_mm") or me.get("thickness_mm"))
        ws.cell(row=row, column=o["col_desc"],   value=f"{pe.get('part_number','')}  {pe.get('description','')}")
        ws.cell(row=row, column=o["col_qty"],    value=int(_safe(pe.get("quantity"), 1)))
        ws.cell(row=row, column=o["col_length"], value=length)
        ws.cell(row=row, column=o["col_width"],  value=width)
        ws.cell(row=row, column=o["col_thick"],  value=thick)
        # sheet size: WB nesting divides by sheet L/W — MUST be non-blank or it
        # produces #VALUE! that propagates into M59 and poisons the whole total.
        # Board stock default 2440×1220 (standard sheet); use engine's if present.
        se = me.get("stock_estimate") or {}
        sh = se.get("candidate_sheet_size_mm") or [2440, 1220]
        ws.cell(row=row, column=o["col_sheet_l"], value=_safe(sh[0]) if len(sh) > 0 else 2440)
        ws.cell(row=row, column=o["col_sheet_w"], value=_safe(sh[1]) if len(sh) > 1 else 1220)
        # Cost per sheet (col L): the WB formula M=(L/J)*(1+K)*D needs this input, else Cost Per
        # Part = 0. Prefer the material result's raw pre-scrap sheet price; fall back to
        # reconstructing it from the per-part cost × parts-per-sheet ÷ (1+scrap). Board parts with
        # neither are left blank + flagged (honest £0, not a silent guess).
        _sheet_price = _safe(me.get("sheet_price_gbp"))
        if not _sheet_price:
            _cpp = _safe(me.get("cost_per_part_gbp") or me.get("unit_material_cost_gbp"))
            _pps = _safe(me.get("parts_per_sheet"))
            _scrap_frac = _safe(me.get("scrap_pct")) or 0.04
            if _cpp and _pps:
                _sheet_price = round(float(_cpp) * float(_pps) / (1.0 + float(_scrap_frac)), 2)
        if _sheet_price:
            ws.cell(row=row, column=o["col_cost_per_sheet"], value=_sheet_price)
        else:
            _flag(f"Other-sheet {pe.get('part_number')} has no sheet price — Cost Per Part will be 0.", flags)
        row += 1

    # ── Labour block: one row per (part, operation) ────────────────────────
    # Operations live in labour_estimate.costs_gbp (its KEYS are the operation names).
    # We write THREE inputs per row and let the WB compute everything else:
    #   C = operation name  -> WB looks up rate/dept/setup from its own table
    #   H = qty per unit
    #   I = throughput (pieces/hour) = order_qty × qty_per_unit ÷ batch_hours
    # The WB's hours formula J=(60/I)*H/60*D6 and cost formula M=rate/I*H+setup then
    # compute using the WB's OWN rate table — reproducing the engine's run-time cost
    # and adding the WB's setup allowance on top. Every number is WB-calculated.
    #
    # THROUGHPUT DEFAULTS (Option B): when the engine's batch_hours is near-zero
    # (e.g. bend_count=None means fold time ≈ 0 → throughput = 10,810/hr), we fall
    # back to realistic per-operation shop rates sourced from Tim's manual estimates.
    # The ceiling is 5× the default; if derived throughput exceeds the ceiling we use
    # the default instead. Geometry-driven ops (laser) stay engine-derived where valid.
    # MEASURED from 1,982 historical jobs (dbo.historical_quote_labour_line, 2026-07-13).
    # Previously these were eyeballed medians. Nearly all were TOO LOW — and a low
    # throughput means MORE hours, so we were OVER-charging labour on every single job.
    # Format: op: throughput_per_hr,   # n lines in corpus | was
    _THROUGHPUT_DEFAULTS = {
        "Robomac":                  709,    # 34 lines  | WAS MISSING ENTIRELY — bar cutting
                                            #             cost £0 (1310 stud vs Tim's £0.17)
        "P.Coat":                   458,    # 316 lines | was 424 — close; my "2-3x too slow"
                                            #             claim was wrong, this one was fine
        "Laser (Metal)":            269,    # 305 lines | was 180
        "Laser (Acrylic)":          252,    # 13 lines  | was 120
        "Manual labour (Acrylic)":  122,    # 13 lines  | was 40
        "Linebend":                 118,    # 18 lines  | was 40
        "Punch":                    116,    # 126 lines | was 100
        "Salvagnini":               110,    # 26 lines  | was 60
        "Saw":                      105,    # 10 lines  | was 60
        "Roll":                     100,    # 12 lines  | was 120
        "Assemble/pack (Acrylic)":   99,    # 15 lines  | was 35
        "Fold":                      93,    # 329 lines | was 50
        "Manual labour (Metal)":     79,    # 23 lines  | was 40
        "Assemble/pack (Metal)":     58,    # 166 lines | was 40
        "Spotweld":                  51,    # 41 lines  | was 23
        "Weld (CO2)":                29,    # 110 lines | was 42 — the ONLY one where we were
                                            #             too FAST, i.e. UNDER-charging weld.
                                            #             Correcting it makes weld dearer.
        # Not present in the corpus sample — left at the previous values, and FLAGGED as
        # unmeasured rather than quietly presented as if they were derived like the rest.
        "Tubebend":                  30,    # UNMEASURED
        "Tube":                      40,    # UNMEASURED
        "Guillotine":                80,    # UNMEASURED
        "Drill (Acrylic)":           30,    # UNMEASURED
        # Deburr/linish is a quick hand pass. Without a default the engine's time model gave a
        # garbage ~1/hr, so a single grouped deburr line billed 22 hrs / £365 at the CNCJ rate —
        # the biggest single error on a metal job. A sane default lets the floor cap it.
        "Grinding / Deburr":        120,    # UNMEASURED — quick hand deburr, config-tunable
        # Joinery / timber-route ops. Without defaults the engine's time model derived a
        # garbage ~0.17/hr on no-DXF timber parts, so grouped CNC/Glue/Spray lines billed
        # 5-6 hrs/part (£1,050 CNC + £610 spray + £472 glue on Cocktails). Sane defaults let
        # the floor cap them back to ~a minute or two per part. UNMEASURED, config-tunable.
        "CNC Joinery":   30,    # UNMEASURED — CNC router pass
        "Glue":          40,    # UNMEASURED — manual glue-up
        "Wet Spray":         60,    # UNMEASURED — wet spray booth
        # Dress Welds (DRES): linish/grind the CO2 weld bead — a quick hand pass, like deburr.
        # Had no default, so the engine's garbage ~0.88/hr stood: a grouped 16-part line billed
        # 18.6 hrs / £533 — the single largest labour line on Cocktails and the whole gap to the
        # engine's £976. A sane default lets the floor cap it.
        "Dress Welds":               60,    # UNMEASURED — hand weld dressing, config-tunable
        # Diamond Polish (DPOL): acrylic edge polish. Unmapped + no default gave ~0.67/hr / £52
        # on a single diffuser. (Metal DPOL is already gated out upstream.)
        "Diamond Polish":            90,    # UNMEASURED — acrylic edge polish, config-tunable
    }
    _THROUGHPUT_CEILING_MULTIPLIER = 5   # derived > default × 5 → use default
    # The ceiling above only catches derived throughputs that are too FAST. A derived
    # throughput that is too SLOW sails through — and slow means MORE HOURS, which
    # INFLATES labour. On 1310 the stud's weld derived at 14.85/hr against a default of
    # 42 (Tim's sheet implies ~50): 2.8x too slow, unguarded, £3.23 vs Tim's £1.25.
    # Symmetric, and deliberately as conservative as the ceiling: only implausible
    # outliers are substituted, and every substitution is FLAGGED with both numbers.
    _THROUGHPUT_FLOOR_DIVISOR = 5        # derived < default ÷ 5 → use default
    lb = cm["labour"]
    row = lb["first_row"]
    labour_overflow = False
    # Labour applies to FABRICATED parts (steel, board, weldment) plus tubes (which get
    # powder/handling as real finishing). Bought-in BOM items (fixings, BI-*, NOTE-*,
    # electricals) are purchased finished goods — they must NOT generate fabrication
    # labour rows (e.g. BI-ADHESIVECABLE was wrongly getting an 'Assemble/pack' line).
    labour_parts = list(steel_parts) + list(board_parts) + list(weldment_parts) + list(wire_parts)

    def _pe_material(p):
        return str(p.get("normalized_material") or (p.get("material_estimate") or {}).get("material") or "")

    labour_parts += [p for p in bom_parts
                     if str((p.get("material_estimate") or {}).get("stock_form") or "").lower() == "tube"
                     # Timber/board panels are direct-priced in the BOM block for MATERIAL, but they
                     # still need their sawing/routing/gluing/lacquer LABOUR — pull them into the
                     # labour path (their stated-weight stock_form + board material is the signature;
                     # this excludes the PVC sticker and metal fasteners, which are not stated_weight).
                     or (str((p.get("material_estimate") or {}).get("stock_form") or "").lower() == "stated_weight"
                         and _is_board(_pe_material(p)))]

    # ── IF WE KNOW THE ROUTE, WE COST THE ROUTE ──────────────────────────────────
    #
    # Every rule above decides by CLASSIFICATION: is this steel, board, wire, a tube, a
    # board panel. A part the engine could not classify gets no labour row no matter what it
    # is doing — so an operation we read off the drawing, or concluded from it, is simply
    # dropped on the floor at the last step before the sheet.
    #
    # M&S 2085 spent four runs proving it. Its tubes are round; the section detector only
    # knew rectangular; no section meant no `tube` stock form; and no stock form meant that
    # even a saw and a weld sitting on the part in black and white would never have reached
    # the labour block. Fixing the detector fixes THIS job. This fixes the shape of it: the
    # route we extracted is the route we cost, and a classification gap can no longer silently
    # delete work.
    #
    # Gated on FABRICATION operations specifically, never on handling/assembly, because
    # bought_in_policy already strips fabrication ops from anything purchased — so a part
    # still carrying one is something we make. That is what kept BI-ADHESIVECABLE from
    # getting an Assemble/pack line, and it stays kept.
    try:
        from bought_in_policy import FABRICATION_OPS as _FAB_OPS
    except Exception:
        _FAB_OPS = frozenset()
    # The route lives on the RAW part records, not on the costed ones this loop walks.
    _route_by_pn = route_operations_by_part(summary)
    _tube_pns = tube_part_numbers(summary)
    # Scope travels on the extract's routes, not on the part records, so read it from there
    # too — a part-level record only knows the ops it carries, not how often they happen.
    _scope_by_op: Dict[str, str] = {}
    for _r in ((summary.get("llm_full_extract") or {}).get("routes") or []):
        if isinstance(_r, dict):
            _o = str(_r.get("operation") or "").strip().lower()
            _s2 = str(_r.get("scope") or "").strip().lower()
            if _o and _s2 in ("part", "assembly"):
                _scope_by_op[_o] = _s2
    _already = {id(p) for p in labour_parts}
    _routed_in = []
    for p in bom_parts:
        if id(p) in _already or not isinstance(p, dict):
            continue
        # A CROSS-REFERENCE ROW IS A DISPLAY LINE, NOT A PART.
        #
        # The BOM list carries a GBP 0.00 row for each fabricated part so the bill of
        # materials reads as the parts list it claims to be. That row is a stub with the
        # same part number as the real record — so this gate matched its route and put it in
        # the labour block as a SECOND copy of a part already there. 2085 booked Weld,
        # Dress Welds and P.Coat at qty 4 across three parts, and grew a second bare
        # "Laser (Metal) (2085-01)" row beside the real one. The cross-reference exists to
        # stop a double count in the material column; it must not create one in labour.
        if p.get("_bom_cross_reference") or p.get("_bom_overflow_consolidated"):
            continue
        _ops = set(_route_by_pn.get(str(p.get("part_number") or "").strip().upper(), []))
        for _k in ("textual_operations", "operations", "inferred_operations"):
            _v = p.get(_k)
            if isinstance(_v, list):
                _ops |= {str(o).strip().lower() for o in _v if str(o).strip()}
        if _ops & _FAB_OPS:
            labour_parts.append(p)
            _routed_in.append(f"{p.get('part_number')} ({', '.join(sorted(_ops & _FAB_OPS))})")
    if _routed_in:
        _flag(f"labour: {len(_routed_in)} part(s) pulled into the labour block because they "
              f"carry a fabrication route, not because the engine classified their stock: "
              f"{'; '.join(_routed_in[:6])}. A route we read is a route we cost.", flags)
    # Powder gate: costs_gbp carries powder_coating blind to finish (part_estimates
    # has finish=None). The reliable drawing-derived signal is textual_operations on
    # the fuller manufacturing_writeup.parts record. Build {part_number -> is_powder}.
    # POWDER GATE — reads the DRAWING FINISH, and FOLLOWS POINTERS. (2026-07-13)
    #
    # "SEE ASSEMBLY DRAWING" is a POINTER, not an absence. It was being read as "no finish",
    # so powder was dropped on every part carrying it — including both parts of 1310, where
    # Tim charges P.Coat £2.00, and four parts of the 1282 Milwaukee bay, which the assembly
    # drawing states is POWDER COATED - SEMI-GLOSS, RAL3020 TRAFFIC RED.
    #
    #   1. finish contains POWDER              -> coat it
    #   2. finish is a POINTER                 -> resolve against the assembly page's finish
    #   3. finish is RAW / SCRAPED EDGES / ... -> do NOT coat (explicit; honour it)
    #   4. no finish text at all               -> fall back to textual_operations (old signal)
    #
    # If a part points at an assembly and NO assembly page states a finish, we cost NOTHING
    # and raise a DRAWING DEFECT flag. An unanswerable drawing goes to Design; it is not
    # guessed at.
    #
    # normalized_finish is empty on some parts (e.g. 1455-C-101 HEADER WELDMENT, which does
    # carry "POWDER COATED - SEMI-GLOSS" in surface_finishes), so surface_finishes is the
    # fallback. That alone fixes the one part that unambiguously says "coat me" and wasn't.
    _POWDER_POINTER_HINTS = ("SEE ASSEMBLY", "SEE GA", "AS ASSEMBLY",
                             "PER ASSEMBLY", "REFER TO ASSEMBLY")

    def _pg_role(_pg):
        _r = _pg.get("page_role")
        if isinstance(_r, dict):
            return str(_r.get("primary_role") or "")
        return str(_r or "")

    def _pg_text(_pg):
        _t = ""
        for _k in ("pdfplumber_text", "normalized_text", "pypdf_text", "text_preview"):
            if _pg.get(_k):
                _t += " " + str(_pg[_k])
        return _t.upper()

    _assembly_is_powder = False
    _assembly_finish_seen = False
    for _pg in (summary.get("pages") or []):
        if not _pg_role(_pg).lower().startswith("assembl"):
            continue
        _at = _pg_text(_pg)
        if "SURFACE FINISH" in _at or "POWDER" in _at:
            _assembly_finish_seen = True
        if "POWDER" in _at:
            _assembly_is_powder = True
            break

    _mw_parts = (summary.get("manufacturing_writeup") or {}).get("parts") or []
    _powder_ok = {}
    # Resolve each part's finish ONCE, here, and let every later reader use this. The last
    # attempt at the assembly-level rule re-derived the finish from normalized_finish alone
    # and missed the surface_finishes fallback below — so it silently found nothing and did
    # nothing. One value, one home.
    _fin_by_pn = {}
    for _mp in _mw_parts:
        _pn = str(_mp.get("part_number") or "")
        # ── A TIN OF PAINT IS NOT A PAINTED OBJECT ───────────────────────────────────
        # The bought-in powder line (TLP-J125-T RYOBI GREEN) carries
        # surface_finishes = ["POWDER COATED - FINE TEXTURE"] — of course it does, IT IS
        # POWDER — and its part_number is None. It was landing in the gate as
        # _powder_ok[""] = True, so the job LOOKED like it already had a coated part. That
        # silently suppressed the assembly-level finish rule (and both of its diagnostics),
        # and cost 7670 its entire £1.92 of P.Coat.
        #
        # A bought-in consumable is not a fabricated part and cannot be the thing that goes
        # through the booth. Any job carrying a powder CODE in its BOM had this bug.
        _pg_roles_boughtin = [str(_r).lower() for _r in (_mp.get("page_roles") or [])]
        if (not _pn) or ("bought_in" in _pg_roles_boughtin):
            continue
        _fin = str(_mp.get("normalized_finish") or "").strip()
        if not _fin:
            _fin = " ".join(str(x) for x in (_mp.get("surface_finishes") or [])).strip()
        _fin_u = _fin.upper()
        _fin_by_pn[_pn] = _fin_u
        _tops = _mp.get("textual_operations") or []
        if isinstance(_tops, str):
            _tops = [_tops]

        if "POWDER" in _fin_u:
            _powder_ok[_pn] = True
        elif any(_h in _fin_u for _h in _POWDER_POINTER_HINTS):
            _powder_ok[_pn] = _assembly_is_powder
            if _assembly_is_powder:
                _flag(f"{_pn}: finish is a POINTER ('{_fin[:38]}') — resolved to POWDER "
                      f"from the assembly drawing.", flags)
            else:
                _flag(f"{_pn}: finish points to the assembly drawing, but NO assembly page "
                      f"states a finish. DRAWING DEFECT — not coated, and not invented. "
                      f"Raise with Design.", flags)
        elif _fin_u:
            _powder_ok[_pn] = False      # RAW / SCRAPED EDGES / etc — explicit
        else:
            _powder_ok[_pn] = any("powder" in str(_o).lower() for _o in _tops)

    # ── ASSEMBLY-LEVEL FINISH ────────────────────────────────────────────────────
    # A finish belongs to the OBJECT THAT GOES THROUGH THE BOOTH, and that object is often
    # an ASSEMBLY, not a part. On 7670 you form raw wire, weld the frame, THEN coat it:
    #
    #     details  (pages 2-4): SURFACE FINISH: RAW            <- correct
    #     assembly (page 1)   : POWDER COATED - FINE TEXTURE   <- the thing that gets coated
    #
    # The drawings are right; the per-part gate is not. It saw three RAW parts, dropped
    # powder on all three, and lost Tim's £1.92 of P.Coat.
    #
    # THE TRAP: 1282 has the same shape — 1455-C-001..004 are RAW and are welded into
    # 1455-C-101, which IS powder coated. But 1282's P.Coat group already contains nine
    # parts INCLUDING that weldment. Flipping its four RAW children to "coated" would hang
    # ONE OBJECT FIVE TIMES.
    #
    # So: apply this ONLY when NOTHING ELSE in the job qualifies for powder.
    #     7670  nothing coated -> the RAW parts ARE what goes through the booth. Coat them.
    #     1282  nine coated    -> the weldment already represents its children. Leave them.
    _assembly_level_powder = False
    if not any(_powder_ok.values()):
        if not _assembly_is_powder:
            _flag("nothing in this job carries a POWDER finish, and no assembly page states "
                  "one either. If the product IS coated, the drawing does not say so — raise "
                  "with Design. NOT coating anything, and not guessing.", flags)
        else:
            # Read the finish from the ONE place it was resolved. (The previous attempt
            # re-derived it from normalized_finish only, missed the surface_finishes
            # fallback, found nothing, and failed silently.)
            _raw_components = [
                _pn2 for _pn2, _ok in _powder_ok.items()
                if (not _ok) and _pn2 and "RAW" in _fin_by_pn.get(_pn2, "")
            ]
            if _raw_components:
                _assembly_level_powder = True
                for _pn2 in _raw_components:
                    _powder_ok[_pn2] = True
                _flag(f"ASSEMBLY-LEVEL FINISH: every detail says RAW, the assembly drawing "
                      f"says POWDER. The components are formed raw, welded, and the ASSEMBLY "
                      f"is coated ({', '.join(_raw_components)}). P.Coat applied ONCE, to one "
                      f"object — not once per component.", flags)
            else:
                # Loud, specific, and it prints what it actually saw. No third blind attempt.
                _flag(f"assembly drawing says POWDER and nothing else in the job qualifies, "
                      f"but no part's finish reads RAW. Finishes seen: "
                      f"{ {k: (v[:24] or '<empty>') for k, v in _fin_by_pn.items()} }. "
                      f"NOT coating anything. Check the drawing finish fields.", flags)

    # ── A WELDMENT IS ONE OBJECT (Dave, 2026-07-14) ──────────────────────────────
    # The branch above only fires when NOTHING in the job qualifies for powder. That was a
    # hack to protect 1282, and it asks the wrong question.
    #
    # On 1310 both parts read 'SEE ASSEMBLY DRAWING'. The pointer resolver correctly turns
    # that into POWDER — so _powder_ok is True, the guard fails, the block is skipped, and
    # P.Coat stays at qty 2. We hang ONE OBJECT TWICE. Tim books qty 1, £2.00; we book
    # qty 2, £3.33.
    #
    # The right question is the one the estimator asked: IS THIS ONE WELDED OBJECT?
    #
    #     no part carries its OWN finish (all point at the assembly)   AND   the job welds
    #        -> they are joined into one thing, and one thing hangs once.
    #
    # 1282 is untouched: its peg panels carry POWDER on their own drawings, so the finishes
    # are MIXED, not all pointers — they are coated individually and then bolted. Correct
    # at qty 16, and this rule stands down.
    _weldment_is_one_object = False
    if not _assembly_level_powder:
        _fab_pns = [str(_p.get("part_number") or "") for _p in labour_parts
                    if _p.get("part_number")]
        _all_point_at_assembly = bool(_fab_pns) and all(
            "SEE ASS" in str(_fin_by_pn.get(_pn, "")).upper() for _pn in _fab_pns
        )
        _job_welds = any(
            "weld" in str(_o).lower()
            for _mp in _mw_parts
            for _o in (_mp.get("textual_operations") or _mp.get("operations") or [])
        )
        if _all_point_at_assembly and _job_welds:
            _weldment_is_one_object = True
            _assembly_level_powder = True          # the P.Coat qty already reads this flag
            _flag(f"WELDMENT IS ONE OBJECT: no part carries its own finish — every one points "
                  f"at the assembly ({', '.join(_fab_pns)}) — and the job welds. They are "
                  f"joined into a single object, and a single object hangs on the booth line "
                  f"ONCE. P.Coat qty 1, not one per component.", flags)
        elif _all_point_at_assembly and not _job_welds:
            # Do NOT guess. Pointing at the assembly does not by itself mean welded — the
            # parts could be coated separately and then bolted. Say why we did not fire.
            _flag(f"every part points at the assembly for its finish "
                  f"({', '.join(_fab_pns)}) but NOTHING WELDS on this job. They may be one "
                  f"object, or coated separately and bolted. NOT collapsing P.Coat to qty 1 "
                  f"— charging one coat per part. Estimator to check.", flags)

    # Diamond-polish gate: diamond_polish is spurious on powder-coated parts
    # (mutually exclusive finishes; boilerplate misfire). Build {pn -> finish is
    # powder} from the fuller record's normalized_finish.
    _finish_is_powder = {}
    for _mp in _mw_parts:
        _pn = str(_mp.get("part_number") or "")
        _fin = str(_mp.get("normalized_finish") or "").upper()
        _finish_is_powder[_pn] = "POWDER" in _fin

    # ── LABOUR — GROUPED BY SETUP, NOT BY PART ──────────────────────────────
    # The WB books SETUP on every row (Fold 30min = £20.24, P.Coat 15min = £88.86,
    # Weld 30min = £20.89). The engine wrote ONE ROW PER PART, so a ten-part job invented
    # ten press-brake setups that never happen on the floor.
    #
    # Measured across 1,982 historical jobs, on a 6-10 part job the estimators write:
    #     Assemble/pack 1.21 rows | Weld 1.23 | Fold 2.03 | Laser 1.92 | P.Coat 1.70
    # Not one per part. But not exactly one per job either — and that tells us the rule:
    #
    #     SETUP BELONGS TO A TOOLING CHANGE, NOT TO A PART.
    #
    # You set the brake for 1.2mm, run every 1.2mm part, then change for 1.0mm. Two gauges,
    # two setups — which is exactly why Tim writes ~2 fold lines on a ten-part job.
    #
    #     Fold / Laser / Punch / P.Coat  -> group by (operation, material, gauge)
    #     Assemble/pack                  -> ONE row per job (pack the product once)
    #     Weld / Spotweld                -> ONE row per job (weld the assembly)
    #
    # Every part number in a group is named in the description cell — nothing is lost.
    # P.Coat belongs here, NOT grouped by gauge. The powder booth does not care how
    # thick the metal is: one colour, one line, one oven run, ONE setup. On 1310 the stud
    # is welded to the hook plate and the assembly goes through powder as a single object —
    # grouping by gauge charged two coating setups to coat one thing (£5.90 vs Tim's £2.00).
    # Fold/Laser/Punch stay grouped by gauge, where the gauge genuinely IS the tooling.
    _ONE_ROW_PER_JOB = {"Assemble/pack (Metal)", "Assemble/pack (Acrylic)",
                        "Weld (CO2)", "Spotweld", "Dress Welds",
                        "P.Coat"}
    _PACK_OPS = {"Assemble/pack (Metal)", "Assemble/pack (Acrylic)"}

    # ── YOU CANNOT SPOT-WELD A BAR TO A PLATE ────────────────────────────────────
    # Spot welding squeezes two thin OVERLAPPING SHEETS between electrodes. 7670 is three
    # 4mm wire forms welded to each other — Tim spotwelds it (buttweld 150/hr + spotweld
    # 45/hr). 1310-02 is an 8mm ROUND STUD welded to a 2mm HOOK PLATE — that is MIG/stud
    # weld, not spot weld.
    #
    # I generalised from 7670 to "any part with stock_form=wire" and it caught 1310's stud:
    # the number moved closer to Tim's £1.25 while the PHYSICS got worse. A lucky match
    # hiding a wrong model is the one thing worse than an honest gap.
    #
    # The engine cannot see WHICH part is welded to which — the joint list is not in the
    # geometry. So the honest rule is the conservative one: Spotweld only when EVERYTHING
    # fabricated in the job is wire/bar. The moment there is sheet present, we cannot tell
    # what the wire is being joined to, and CO2 is the safe assumption.
    _fab_forms = [
        str((_p.get("material_estimate") or {}).get("stock_form") or "").lower()
        for _p in labour_parts
    ]
    _all_fabricated_are_wire = bool(_fab_forms) and all(
        _f in ("wire", "bar") for _f in _fab_forms
    )
    if (not _all_fabricated_are_wire) and any(_f in ("wire", "bar") for _f in _fab_forms):
        _flag("welding: this job mixes wire/bar with sheet, and the drawing does not say "
              "which parts are joined to which. Assuming Weld (CO2) — a wire-to-wire joint "
              "would be Spotweld and cheaper. Estimator to check.", flags)
    # A solid round bar is cut on the Robomac. This is a MANUFACTURING ROUTE (material form
    # -> machine), the same class of rule as "sheet steel gets lasered" — which the engine
    # already infers without anyone writing it on a drawing. document_builder does add
    # 'robomac' to the part's operations, but wb_populate reads ops from the PRICING record
    # (part_estimate.labour_estimate.costs_gbp), and the op was written to the WRITE-UP
    # record. Rather than plumb the pricing layer — which has no bar time model to offer
    # anyway — inject the row here from stock_form.
    _ROBOMAC_STOCK_FORMS = {"wire"}
    # Robomac is ONE ROW PER WIRE FORM, not one per job. Tim's 7670 sheet:
    #     Robomac  main frame    qty 1  100/hr  setup 15  £0.47
    #     Robomac  back wire     qty 2  450/hr  setup 15  £0.30
    #     Robomac  bottom frame  qty 1  300/hr  setup 15  £0.26
    # Three rows, three setups, throughput swinging 100 -> 450. Each wire form is a
    # different bend program on the machine, so each is a genuine separate setup.
    # Grouping these into one row UNDER-charges — the failure mode we cannot see.
    _PER_PART_OPS = {"Robomac"}

    _groups = (
        canonical_labour_groups(
            summary,
            list(pes),
            order_qty,
            skipped_part_numbers=_skip_pns,
            all_fabricated_are_wire=_all_fabricated_are_wire,
        )
        if _canonical_cutover else {}
    )
    if _canonical_cutover:
        _flag(
            f"canonical route rendered {len(_groups)} labour pricing group(s); "
            "the raw-route rescue and legacy per-part cost union were not consulted.",
            flags,
        )
    for pe in ([] if _canonical_cutover else labour_parts):
        if str(pe.get("part_number") or "") in _skip_pns:
            continue   # phantom excluded from Sheet Steel — drop its ops too
        le = pe.get("labour_estimate") or {}
        costs = le.get("costs_gbp") or {}
        batch_hours = le.get("batch_hours") or {}
        _pn = str(pe.get("part_number") or "")
        ops = list(costs.keys())

        # ── A ROUTED OPERATION WITH NO COST MODEL IS STILL WORK ──────────────────────
        #
        # This list came only from labour_estimate.costs_gbp — the operations the estimator
        # managed to put a NUMBER against. An operation that is on the part's route but has
        # no time model behind it was not merely uncosted, it was invisible: no row, no
        # flag, nothing on the sheet to say the job includes it.
        #
        # M&S 2085 is that exactly. Both tubes carry tube_cut and tube_bending, the plate
        # carries folding, all three map to real departments (Tube, Tubebend, Fold) — and
        # the invariant reported them as "named but not priced" while the sheet showed five
        # labour rows on one part. The route was right the whole way down and stopped here.
        #
        # So the union is taken. An operation the estimator costed keeps its number; one it
        # did not gets a row anyway, timed from the department's own throughput default (the
        # same default the sheet already uses when it cannot size-band a part) and flagged
        # by name. Every filter below still runs — spurious-by-stock-form, the powder gate,
        # the diamond-polish gate — so this widens what reaches them, not what survives them.
        _routed_extra = routed_operations_without_cost(
            pe, costs, _route_by_pn.get(_pn.strip().upper()))
        if _routed_extra:
            ops = ops + _routed_extra
            _flag(f"labour {_pn or '?'}: {len(_routed_extra)} routed operation(s) carry no "
                  f"cost model ({', '.join(_routed_extra)}) — costed from the department's "
                  f"throughput default so the work appears on the sheet. Confirm the times.",
                  flags)

        if not ops:
            continue
        _qty_pu = int(_safe(pe.get("quantity"), 1))
        _is_acr = _is_board(str(pe.get("normalized_material") or ""))
        _sf = (pe.get("material_estimate") or {}).get("stock_form")
        # STOCK FORM IS ONLY STAMPED WHEN THE MATERIAL COSTED SUCCESSFULLY.
        #
        # estimator sets stock_form="tube" inside the two branches that PRODUCE A PRICE — the
        # catalogue match and the mass calculation. A tube whose material could not be priced
        # at all (2085's, which have a section but no printed cut length) therefore arrives
        # with the field blank, and every rule keyed on it silently does nothing. The laser
        # suppression written for exactly these parts was inert on exactly these parts.
        #
        # Identity is not a by-product of pricing. And the evidence for it — section_stock —
        # is on the RAW record, not this costed one, so it comes through the same bridge the
        # route does. Reading it off `pe` looked right and resolved to None on the real data:
        # the fourth time in this job that a fix has been aimed at the wrong record.
        if not _sf and _pn.strip().upper() in _tube_pns:
            _sf = "tube"
        _mat = pe.get("normalized_material") or ""
        _me2 = pe.get("material_estimate") or {}
        _ng2 = pe.get("normalized_geometry") or {}
        _thk = _safe(pe.get("normalized_thickness_mm") or _me2.get("thickness_mm"), 0)

        for op in ops:
            if _is_spurious_operation(op, _sf, _mat):
                _flag(f"dropped spurious op '{op}' on {_pn} "
                      f"(stock_form={_sf}, material={_mat})", flags)
                continue
            if "powder" in str(op).lower():
                if _pn in _powder_ok and not _powder_ok[_pn]:
                    _flag(f"dropped powder on {_pn} — drawing finish is not powder "
                          f"(RAW/assembly/weldment); costs_gbp over-applied it.", flags)
                    continue
            if "diamond" in str(op).lower() or ("polish" in str(op).lower()
                                                and "edge" not in str(op).lower()):
                if _finish_is_powder.get(_pn):
                    _flag(f"dropped diamond_polish on {_pn} — part is POWDER COATED "
                          f"(diamond-polish is spurious/boilerplate on a powder finish).", flags)
                    continue

            wb_op = _map_operation(op, _is_acr, _sf or "")
            # A timber part routed onto an ACRYLIC department: the template has no joinery
            # equivalent, so the nearest hand rate is used. Say so rather than leaving
            # "Assemble/pack (Acrylic)" sitting against a wooden crate unexplained — the
            # rate is a substitution, and only an estimator can confirm it is the right one.
            if wb_op and "(Acrylic)" in str(wb_op) and _is_timber(_mat):
                _flag(f"'{wb_op}' booked on {_pn} ({_mat}) — the template has no joinery "
                      f"equivalent of this department, so the ACRYLIC hand rate is used as "
                      f"the nearest available. Confirm the rate is right for timber.", flags)
            if wb_op is None:
                _flag(f"labour op '{op}' ({_pn}) not in OP_NAME_MAP — WB rate lookup will "
                      f"return 0 for it. Add mapping.", flags)
                wb_op = str(op)

            # Wire frames go on the SPOT WELDER — but ONLY when the whole job is wire.
            # 7670 (3 wire forms welded to each other): Spotweld, £1.61 on Tim's sheet.
            # 1310 (8mm stud welded to a 2mm plate):    CO2 — you cannot spot-weld a bar
            #                                           to a plate, whatever the number says.
            if (wb_op == "Weld (CO2)"
                    and str(_sf or "").lower() in ("wire", "bar")
                    and _all_fabricated_are_wire):
                wb_op = "Spotweld"

            if wb_op in _PER_PART_OPS:
                key = (wb_op, _pn, "")         # one setup PER PART (per wire form)
            elif wb_op in _ONE_ROW_PER_JOB:
                key = (wb_op, "", "")          # one setup for the whole job
            else:
                key = (wb_op, str(_mat), "%g" % (_thk or 0))   # one setup per tooling change

            g = _groups.setdefault(key, {
                "wb_op": wb_op, "material": _mat, "thickness": _thk,
                # The tuple that DECIDED this grouping. material/thickness above are only
                # those of the first part to land here, which for a one-row-per-job
                # department is an accident of ordering.
                "group_key": key,
                "qty": 0, "bh": 0.0, "parts": [], "bends": 0, "holes": 0,
                # The ENGINE operation(s) behind this row. The group key carries the mapped
                # DEPARTMENT name, so it cannot answer "what operation is this" — reading it
                # as one turned wet_spray into 'Spray / Wet Paint' and left every consumer
                # matching on a string the engine never emits.
                "engine_ops": [],
            })
            if op and op not in g.setdefault("engine_ops", []):
                g["engine_ops"].append(str(op))
            # The sequence the extract read off THIS drawing, if it gave one. Lowest wins:
            # a group can gather several parts, and the row belongs where the first of them
            # is done. None stays None, and the shop's own order applies at sort time.
            # SCOPE — how often this operation happens per product, which is not the same
            # as how many parts the route line names. Recorded on the group so the emit loop
            # can flag it (or honour it, once the rates are confirmed).
            _op_l = str(op or "").strip().lower()
            _sc = operation_scope_for(pe, _op_l, _scope_by_op)
            if _sc == "assembly":
                g["assembly_scoped"] = True
                _qpu_part = (pe.get("operation_qty_per_unit") or {}).get(_op_l)
                if _qpu_part:
                    _prev = _safe(g.get("qty_per_unit_by_scope"), 0) or 0
                    g["qty_per_unit_by_scope"] = max(_prev, _safe(_qpu_part, 1) or 1)
            _rs = (pe.get("operation_sequence") or {}).get(str(op or "").strip().lower())
            if _rs is not None:
                try:
                    _rs = float(_rs)
                    g["route_sequence"] = (_rs if g.get("route_sequence") is None
                                           else min(float(g["route_sequence"]), _rs))
                except (TypeError, ValueError):
                    pass
            g["qty"] += _qty_pu
            _bh = _safe(batch_hours.get(op))
            if _bh and _bh > 0:
                g["bh"] += float(_bh)
            if _pn and _pn not in g["parts"]:
                g["parts"].append(_pn)

        # Robomac: a solid bar has to be CUT, and no upstream record delivers that op to the
        # pricing layer. Inject it from the stock form — same manufacturing-route reasoning
        # that already gives sheet steel its laser. Throughput 709/hr is the corpus median
        # (34 lines), NOT back-solved from Tim's £0.17.
        # Inject Robomac ONLY if the pricing record did not already carry the op.
        # 1310's stud had no robomac op at all, so the row was missing and I injected it
        # unconditionally. On 7670 the op IS present — and the unconditional injection
        # produced TWO Robomac rows for the same work (one keyed ("Robomac","",""), one
        # keyed ("Robomac","MILD_STEEL","4")). Check before injecting.
        if str(_sf or "").lower() in _ROBOMAC_STOCK_FORMS:
            _has_robo = any(
                _map_operation(_o, _is_acr, _sf or "") == "Robomac" for _o in ops
            )
            if not _has_robo:
                _rg = _groups.setdefault(("Robomac", _pn, ""), {
                    "wb_op": "Robomac", "material": _mat, "thickness": 0,
                    "qty": 0, "bh": 0.0, "parts": [], "bends": 0, "holes": 0,
                })
                _rg["qty"] += _qty_pu
                if _pn and _pn not in _rg["parts"]:
                    _rg["parts"].append(_pn)
            _ol = str(op).lower()
            if _ol == "folding":
                g["bends"] += int(_safe((_ng2 or {}).get("estimated_bend_line_count"), 0)) * _qty_pu
            elif _ol in ("hole_machining", "drilling", "punch"):
                g["holes"] += int(_safe((_ng2 or {}).get("estimated_hole_count"), 0)) * _qty_pu

    # ── MANM: insert labour for pressed fasteners (self-clinch nuts, PEM studs) ──
    # Tim books the press/insert time as Manual labour (Metal) (MANM). His 12120 sheet
    # gives the rule directly and consistently: Clinch x4 @60/hr and Pem x2 @120/hr both
    # = 15 s/insert (config.MANM_INSERT_SECONDS_EACH). Knurled knobs and thumbscrews are
    # hand-assembled (Assemble/pack), NOT pressed, so only clinch/PEM parts count. Counts
    # come from bom_parts, which by now carry the reconciled quantities (self-clinch 1->4,
    # PEM added), so this books labour on the true insert count. Injected here — like the
    # Robomac row above — because the insert count is a BOM fact, not a per-part route op.
    if (not _canonical_cutover
            and _BOOK_MANM_INSERT_LABOUR and _MANM_INSERT_SECONDS_EACH > 0):
        _ins_tokens = [str(t).upper() for t in (_MANM_INSERT_PART_TOKENS or [])]
        _insert_count = 0
        _insert_parts: List[str] = []
        for _bp in bom_parts:
            _blob = (str(_bp.get("description") or "") + " "
                     + str(_bp.get("part_number") or "")).upper()
            if _ins_tokens and any(_tok in _blob for _tok in _ins_tokens):
                _q = int(_safe(_bp.get("quantity"), 0))
                if _q > 0:
                    _insert_count += _q
                    _pn_ins = str(_bp.get("part_number") or _bp.get("description") or "insert")
                    if _pn_ins not in _insert_parts:
                        _insert_parts.append(_pn_ins)
        if _insert_count > 0:
            # bh chosen so the derived throughput below resolves to exactly
            # 3600/(inserts * seconds_each): one MANM row, qty 1.
            _manm_bh = order_qty * _insert_count * float(_MANM_INSERT_SECONDS_EACH) / 3600.0
            _mg = _groups.setdefault(("Manual labour (Metal)", "insert", ""), {
                "wb_op": "Manual labour (Metal)", "material": "", "thickness": 0,
                "qty": 1, "bh": 0.0, "parts": [], "bends": 0, "holes": 0,
            })
            _mg["qty"] = 1
            _mg["bh"] += _manm_bh
            _mg["parts"] = _insert_parts
            _manm_tp = 3600.0 / (_insert_count * float(_MANM_INSERT_SECONDS_EACH))
            _flag(f"MANM insert labour: {_insert_count} pressed insert(s) "
                  f"(self-clinch/PEM, counted from the reconciled BOM: "
                  f"{', '.join(_insert_parts)}) x {_MANM_INSERT_SECONDS_EACH:.0f}s each "
                  f"-> Manual labour (Metal) at {_manm_tp:.0f}/hr. Rule from Tim's 12120 "
                  f"sheet (clinch x4 @60/hr, pem x2 @120/hr both = 15s/insert). Knurled "
                  f"knob & thumbscrew are hand-assembled (Assemble/pack), not counted.", flags)

    # ── LABOUR ROWS COME OUT IN MANUFACTURING ORDER, NOT ALPHABETICAL ORDER ─────────
    #
    # This was sorted(_groups.keys()), and the keys start with the department name — so
    # 2085's sheet read Assemble/pack, Laser, P.Coat. Pack before cut. That is not a route,
    # it is a word list, and an estimator reading down it cannot sanity-check a sequence
    # that is not in sequence.
    #
    # The extract already returns `sequence` per route (10, 20, 30...) and it was thrown
    # away. Where it is present it wins, because it is the model reading THIS drawing's
    # order of work. Where it is absent — a job with no routes extracted, or an operation
    # the engine inferred itself — the shop's own order applies: you cut before you form,
    # form before you weld, weld before you coat, and you pack last. That is a fact about
    # the shop, not about a job, so it is inheritable by every job that follows.
    _SHOP_ORDER = {
        "Laser (Metal)": 10, "Laser (Acrylic)": 10, "Punch": 10, "Guillotine": 10,
        "Saw": 15, "Tube": 15, "CNC Joinery": 15,
        "Fold": 20, "Linebend": 20, "Tubebend": 20, "Roll": 20, "Robomac": 20,
        "Manual labour (Metal)": 25,
        "Edge Banding": 28, "Glue": 28,
        "Weld (CO2)": 30, "Spotweld": 30,
        "Dress Welds": 35, "Grinding / Deburr": 38,
        "P.Coat": 60, "Wet Spray": 60, "Diamond Polish": 62,
        "Assemble/pack (Metal)": 90, "Assemble/pack (Acrylic)": 90,
    }
    _DEFAULT_ORDER = 50          # something we do not recognise sits mid-route, not last

    def _group_order(_k):
        _g = _groups[_k]
        _read = _g.get("route_sequence")
        if _read is not None:
            return (0, float(_read), str(_g.get("wb_op") or ""))
        return (0, float(_SHOP_ORDER.get(str(_g.get("wb_op") or ""), _DEFAULT_ORDER)),
                str(_g.get("wb_op") or ""))

    _charge_once = bool(getattr(config, "ASSEMBLY_SCOPED_OPS_CHARGE_ONCE", False))
    for _key in sorted(_groups.keys(), key=_group_order):
        g = _groups[_key]
        # An assembly-level operation is done once per product however many parts it joins.
        # Today it is charged once per part. Whether that is wrong turns on what the
        # throughput rate MEANS -- assemblies per hour or parts per hour -- which is the
        # estimators' ruling about their own table, so the default only SAYS SO.
        if (not g.get("canonical_route")
                and g.get("assembly_scoped")
                and int(_safe(g.get("qty"), 1) or 1) > 1):
            _was = int(_safe(g.get("qty"), 1) or 1)
            if _charge_once:
                g["qty"] = assembly_scoped_qty(g)
                _flag(f"labour '{g.get('wb_op')}': ASSEMBLY-scoped operation charged ONCE "
                      f"per product (was qty {_was}, one per part). "
                      f"config.ASSEMBLY_SCOPED_OPS_CHARGE_ONCE is on.", flags)
            else:
                _flag(f"labour '{g.get('wb_op')}': the route marks this an ASSEMBLY-level "
                      f"operation -- done once per product -- but it is charged at qty "
                      f"{_was}, one per part it names. If the {g.get('wb_op')} rate is "
                      f"ASSEMBLIES per hour this line is about {_was}x too high; if it is "
                      f"PARTS per hour it is correct. Nothing changed: set "
                      f"config.ASSEMBLY_SCOPED_OPS_CHARGE_ONCE once the rates are confirmed.",
                      flags)
        if row > lb["last_row"]:
            labour_overflow = True
            break
        wb_op = g["wb_op"]

        # Assemble/pack is PER PRODUCT: you pack the finished product once, not once per
        # part. Tim books qty 1 (1298: "Poly bag & bulk pack", qty 1, 90/hr).
        # Assemble/pack is PER PRODUCT: you pack the finished product once, not once per part.
        # P.Coat is the same WHEN THE COAT HAPPENS AT ASSEMBLY LEVEL: one welded frame goes
        # on the hook, not three loose components. Tim books exactly that — P.Coat qty 1.
        # (When the parts themselves carry POWDER, they are coated individually before
        #  assembly and the per-part count is right — so this only applies to the
        #  assembly-level case.)
        if g.get("canonical_route"):
            _qty = int(_safe(g.get("qty"), 1) or 1)
        else:
            _qty = 1 if (wb_op in _PACK_OPS
                         or (wb_op == "P.Coat" and _assembly_level_powder)) else int(g["qty"] or 1)

        _matx = str(g["material"] or "").replace("_", " ").strip()
        _rd = labour_row_description(wb_op, g["material"], g["thickness"],
                                     g["parts"], g["bends"], g["holes"])

        ws.cell(row=row, column=lb["col_operation"], value=wb_op)
        ws.cell(row=row, column=lb["col_desc"],      value=_rd[:200])
        ws.cell(row=row, column=lb["col_qty"],       value=_qty)

        default_tp = _THROUGHPUT_DEFAULTS.get(wb_op or "")

        # ── SIZE-BANDED DEFAULT, keyed on the job's largest part AREA ────────────────
        # For Assemble/pack and P.Coat one number cannot be right - a small part is packed
        # and coated far faster than a big one. Pick the band from _max_part_area_m2, which
        # is known here (unlike unit cost, a workbook formula). Fold/Laser are not in the
        # band table (derived); Robomac/Weld are not (not size-driven) - all fall through.
        _bands = _THROUGHPUT_SIZE_BANDS.get(wb_op or "")
        if _bands and _max_part_area_m2 > 0:
            _e1, _e2, _e3 = _THROUGHPUT_AREA_EDGES
            _a = _max_part_area_m2
            _band = "A" if _a < _e1 else "B" if _a < _e2 else "C" if _a < _e3 else "D"
            _banded = _bands.get(_band)
            if _banded:
                _flag(f"throughput for '{wb_op}' size-banded on part area: largest part "
                      f"{_max_part_area_m2:.4f} m2 -> band {_band} -> {_banded}/hr "
                      f"(was default {default_tp}/hr). MEASURED from your own history by product "
                      f"size - a small part runs faster than a big one, and one median cannot "
                      f"say that. Retune in config.THROUGHPUT_SIZE_BANDS.", flags)
                default_tp = _banded
        elif _bands and _max_part_area_m2 <= 0:
            _flag(f"throughput for '{wb_op}': wanted to size-band it but no fabricated part "
                  f"area was computed - using the un-banded default {default_tp}/hr. Not "
                  f"guessing a band.", flags)

        # Assembly, packing and welding time is NOT in the DXF. There is no geometry from
        # which to derive "how long does it take to pack this" — the engine's derived value
        # for those ops is fiction dressed as measurement (1310's weld derived at 14.85/hr
        # against a corpus average of 29). Use the MEASURED default and say so.
        # For laser (cut path) and fold (bend count) the geometry genuinely does carry the
        # information, so the derived value is kept.
        # ── THE TEMPLATE ALREADY COMPUTES THE LASER RATE. READ IT. ──────────────────
        # Sheet Steel block, per row:  V = total seconds,  W = 3600/V = pieces per hour.
        # Every input to it (blank L/W, gauge, hole count, internal cut) is written by
        # THIS module on every run. For 1310-01 it computes 311.1/hr; Tim books 300.
        # We were writing 80 — our own second, worse time model — and over-charging the
        # laser by 3-4x on small parts.
        #
        # For a group of parts sharing one setup, the correct combined rate is
        #     3600 * SUM(qty) / SUM(qty * seconds)
        # which for a single part collapses to 3600/V, i.e. exactly the W column.
        #
        # Written as a FORMULA, not a value, so it tracks any change the estimators make
        # to their own cutting speeds — and so it is visibly THEIR number, not ours.
        _laser_formula = None
        if wb_op == "Laser (Metal)":
            _rws = [_steel_row_by_pn.get(str(_p)) for _p in (g.get("parts") or [])]
            _rws = [_r for _r in _rws if _r]
            if _rws:
                _qs = "+".join("E%d" % _r for _r in _rws)
                _ts = "+".join("E%d*V%d" % (_r, _r) for _r in _rws)
                _fb = float(default_tp or 269)
                _laser_formula = "=IFERROR(3600*(%s)/(%s),%s)" % (_qs, _ts, _fb)

        if _laser_formula:
            ws.cell(row=row, column=lb["col_throughput"], value=_laser_formula)
            _flag(f"laser throughput now READS THE TEMPLATE'S OWN Laser Rate Calculator "
                  f"(rows {_rws}) instead of the engine's time model. The calculator uses "
                  f"the estimators' cutting speeds, the blank size, the hole count and the "
                  f"internal cut distance — all of which we already write into it. On 1310 "
                  f"it computes 311/hr where our model said 80 (Tim books 300).", flags)
        elif (wb_op in _ONE_ROW_PER_JOB or wb_op in _PER_PART_OPS) and default_tp:
            ws.cell(row=row, column=lb["col_throughput"], value=float(default_tp))
        else:
            bh = g["bh"]
            if bh and bh > 0:
                _total_pieces = order_qty * _qty
                _derived = _total_pieces / bh
                throughput = _derived
                if default_tp:
                    _ceiling = default_tp * _THROUGHPUT_CEILING_MULTIPLIER
                    _floor = default_tp / _THROUGHPUT_FLOOR_DIVISOR
                    if _derived > _ceiling:
                        throughput = float(default_tp)
                        _flag(f"throughput CEILING hit on '{wb_op}': derived {_derived:.2f}/hr "
                              f"is {_derived/default_tp:.1f}x the default {default_tp}/hr "
                              f"— using default (was UNDER-charging).", flags)
                    elif _derived < _floor:
                        throughput = float(default_tp)
                        _flag(f"throughput FLOOR hit on '{wb_op}': derived {_derived:.2f}/hr "
                              f"is {default_tp/_derived:.1f}x SLOWER than the default "
                              f"{default_tp}/hr — using default (was OVER-charging).", flags)
                ws.cell(row=row, column=lb["col_throughput"], value=round(throughput, 4))
            elif default_tp:
                ws.cell(row=row, column=lb["col_throughput"], value=float(default_tp))
            else:
                _flag(f"labour op '{wb_op}' has no batch_hours and no default throughput — "
                      f"WB hours/cost will be #DIV/0! for this row.", flags)
        # Remember WHICH sheet row this group became. The calculated read-back
        # returns rows keyed only by their position, and without this join key it
        # cannot recover which engine operations and which parts a row represents —
        # it then falls back to inverting the department name, which expands every
        # alias and duplicates operations on the client quote.
        g["workbook_row"] = row
        row += 1

    # The canonical route record — built here, after the write loop, so every row carries
    # the sheet row it actually landed on. See build_workbook_labour for why that matters.
    summary["workbook_labour"] = build_workbook_labour(
        _groups, _skip_pns, canonical_mode=_canonical_cutover)

    _flag(f"labour: {len(_groups)} grouped row(s) — setup is booked once per tooling group, "
          f"not once per part.", flags)
    if labour_overflow:
        _flag(f"Labour overflow: more operations than {lb['last_row']-lb['first_row']+1} rows — extras DROPPED.", flags)

    # ── Make Total Material Cost tolerate not-yet-dimensioned rows ─────────
    # One steel part with no blank L/W (or blank gauge) errors its per-row cost and, via the
    # plain SUM in M92, blanks the WHOLE total (Material -> Unit -> Sell). Rewrite M92 to sum
    # ignoring errors so the sheet shows a PARTIAL total from the credible lines + the flagged
    # gaps, and self-completes as the estimator fills the dims. Non-regressive on clean jobs.
    if _MATERIAL_TOTAL_ERROR_TOLERANT:
        _make_material_total_error_tolerant(ws, flags)
        # Also clean the scattered per-row #DIV/0!/#VALUE! so the sheet is presentable to
        # estimating: wrap the error-prone row formulas in IFERROR (blank on error, real value
        # once dims are filled). Same 'don't blank the credible work' intent, at cell level.
        _clean_error_cells(ws, flags)

    # ── Append AI supplementary sheets (renamed to avoid clashing with WB) ──
    _append_ai_sheets(wb, summary, flags)

    # ── Force Excel to recalc on open ──────────────────────────────────────
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        try:
            from openpyxl.workbook.properties import CalcProperties
            wb.calculation = CalcProperties(fullCalcOnLoad=True)
        except Exception:
            _flag("could not set fullCalcOnLoad — estimator may need to press F9.", flags)

    # ── Save-As to output dir with folder-name + timestamp ─────────────────
    os.makedirs(cm["output_dir"], exist_ok=True)
    safe_name = re.sub(r'[<>:"/\\|?*]', "_", job_folder_name).strip()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(cm["output_dir"], f"{safe_name}_{stamp}.xlsx")
    wb.save(out_path)

    print(f"   [wb_populate] Populated template saved: {out_path}")
    print(f"   [wb_populate] Parts: {len(bom_parts)} BOM (incl. tube sections), "
          f"{len(wire_parts)} wire/bar, "
          f"{len(steel_parts)} steel, {len(board_parts)} other-sheet, "
          f"{len(weldment_parts)} weldment (labour-only)")
    if flags:
        print(f"   [wb_populate] {len(flags)} flag(s) raised — see above.")
    print(f"   [wb_populate] NOTE: totals compute when opened in Excel (calc-on-load).")
    return out_path


def _flag_to_text(rf):
    """review_flags is meant to be a list of strings, but something upstream sometimes
    appends a dict — which killed populate_workbook on 7670 (TypeError in str.join) and
    sent the whole job down the xlsx_output fallback path SILENTLY, producing the wrong
    workbook with no loud failure.

    Coerce to text without losing the content. A dict is rendered compactly rather than
    dumped as a raw Python repr into the estimator's review column.

    The upstream producer writing structured data into a list-of-strings field is a
    separate defect and still needs fixing at source.
    """
    if isinstance(rf, str):
        return rf
    if isinstance(rf, dict):
        for _k in ("message", "msg", "text", "reason", "flag", "detail", "description"):
            if rf.get(_k):
                _code = rf.get("code") or rf.get("type") or rf.get("name")
                return f"{_code}: {rf[_k]}" if _code else str(rf[_k])
        return "; ".join(f"{k}={v}" for k, v in rf.items() if v not in (None, "", []))
    if isinstance(rf, (list, tuple)):
        return " / ".join(_flag_to_text(x) for x in rf)
    return str(rf)

def _append_ai_sheets(wb, summary: Dict[str, Any], flags: List[str]):
    """Append the engine's own detail/provenance sheets under NON-colliding names,
    so the WB's structural 'Labour' and 'Material Price Break' sheets are untouched."""
    # AI Labour Detail — the engine's own labour breakdown (informational)
    pes = (summary.get("estimate_summary") or {}).get("part_estimates") or summary.get("parts") or []

    def _add(title: str, header: List[str], rows: List[List[Any]]):
        # ensure name doesn't clash with structural sheets
        if title in CELL_MAP["structural_sheets"] or title in wb.sheetnames:
            title = "AI " + title
        ws = wb.create_sheet(title=title[:31])  # Excel 31-char sheet name limit
        ws.append(header)
        for r in rows:
            ws.append(r)

    # AI Material Detail
    mat_rows = []
    for pe in pes:
        me = pe.get("material_estimate") or {}
        mat_rows.append([
            pe.get("part_number"), pe.get("description"),
            pe.get("normalized_material"),
            (me.get("blank_length_mm")), (me.get("blank_width_mm")),
            pe.get("normalized_thickness_mm"),
            me.get("cost_per_part_gbp"), me.get("extended_material_cost_gbp"),
            (pe.get("geometry") or {}).get("estimated_cut_length_mm"),
            # Name the real source. A modelled flat pattern is measured geometry, but the
            # estimator must be able to tell a SolidWorks cut list from a DXF from a PDF.
            {"dxf_flat_pattern": "dxf",
             "solidworks_flat_pattern": "solidworks"}.get(
                str(pe.get("geometry_source") or ""), "pdf"),
        ])
    _add("AI Material Detail",
         ["Part", "Desc", "Material", "Blank L", "Blank W", "Gauge",
          "Cost/Part", "Ext Material", "Cut len (mm)", "Geom source"],
         mat_rows)

    # AI Provenance — where each bought-in price came from
    prov_rows = []
    for pe in pes:
        if "bought_in" in [str(r).lower() for r in (pe.get("page_roles") or [])]:
            prov_rows.append([
                pe.get("part_number"), pe.get("description"),
                pe.get("unit_cost_gbp"),
                pe.get("cost_source") or pe.get("source"),
                "verified" if pe.get("price_verified") else "UNVERIFIED",
                pe.get("supplier") or "",
                " | ".join(_flag_to_text(_rf) for _rf in (pe.get("review_flags") or [])),
            ])
    _add("AI Provenance",
         ["Part", "Desc", "Unit £", "Price Source", "Verified", "Supplier", "Review Flags"],
         prov_rows)

    # The costing sheet shows only accepted required rows. These two review sheets preserve
    # the full hierarchy and every route decision, including ruled-out and unverified work.
    canonical = canonical_route_payload(summary)
    if canonical:
        bom_rows = []
        for node in canonical.get("nodes") or []:
            if not isinstance(node, dict):
                continue
            bom_rows.append([
                node.get("part_number"),
                node.get("description"),
                node.get("kind"),
                node.get("qty_per_unit"),
                ", ".join(node.get("parents") or []),
                ", ".join(
                    f"{edge.get('part_number')} x{edge.get('qty')}"
                    for edge in (node.get("children") or [])
                    if isinstance(edge, dict)
                ),
            ])
        _add(
            "Canonical BOM",
            ["Part", "Description", "Kind", "Qty/unit", "Parent(s)", "Children"],
            bom_rows,
        )

        route_rows = []
        for decision in canonical.get("decisions") or []:
            if not isinstance(decision, dict):
                continue
            route_rows.append([
                decision.get("sequence"),
                decision.get("operation"),
                decision.get("status"),
                decision.get("target_id"),
                decision.get("scope"),
                decision.get("qty_per_unit"),
                ", ".join(decision.get("participants") or []),
                decision.get("source"),
                decision.get("reason"),
                decision.get("decision_id"),
            ])
        _add(
            "Canonical Route",
            [
                "Seq", "Operation", "Status", "Target", "Scope", "Qty/unit",
                "Participants", "Source", "Reason", "Decision ID",
            ],
            route_rows,
        )


# ── standalone test entry ──────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python wb_populate.py <run_json_path> <job_folder_name>")
        print('Example: python wb_populate.py "C:\\ClaudeVision\\output\\json\\1282 - Milwaukee Wall Bay.json" "1282 - Milwaukee Wall Bay"')
        sys.exit(1)
    with open(sys.argv[1], encoding="utf-8") as fh:
        summary = json.load(fh)
    populate_workbook(summary, sys.argv[2])
