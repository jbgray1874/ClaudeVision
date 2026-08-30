#!/usr/bin/env python3
r"""
_probe_powder_wireup.py  —  READ-ONLY.

Before wiring the orphan powder-material calc into the total (M67), answer 3 things
from the widened template so we don't guess or double-count:

  1. Does 'Total Powder Per Unit' (near Z57) actually have a FORMULA summing the AD
     'Powder Qty Per Part' column, or is it just an orphan label? Dump Z57..AE57.
  2. Is there a POWDER PRICE (£/kg) anywhere in the sheet to multiply the kg by?
     Scan for 'powder' + '£'/'price'/'kg' and any price-looking constants near the
     powder calc. (If none exists, wiring kg->£ needs a price input we don't have.)
  3. What EXACTLY is in M67 now (the material total), and is there any existing
     powder term in it? Show M67 + the cells around 'Total Material Cost' so we can
     see how a powder term would be added without double-counting the P.Coat labour.

READ-ONLY.

Usage:
  C:\ClaudeVision\.venv\Scripts\python.exe _probe_powder_wireup.py
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as L

TEMPLATE = r'\\sdi-dc01\shareddata$\Shared\Estimating\Completed\AI Estimating\AISheets\Blank Estimate Sheet  WB 2026.xlsx'


def main():
    wb = load_workbook(TEMPLATE)
    ws = wb['Estimate']

    print("=" * 90)
    print("POWDER WIRE-UP FEASIBILITY PROBE (read-only)")
    print("=" * 90)

    # 1. Total Powder Per Unit — formula or orphan?
    print("\n[1] 'Total Powder Per Unit' area (rows 55-58, cols X(24)..AF(32)):")
    for r in range(55, 59):
        cells = []
        for c in range(24, 33):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                cells.append(f"{L(c)}{r}={v!r}")
        if cells:
            print(f"    row {r}: " + " | ".join(cells))

    # 1b. dump the AD column (Powder Qty Per Part) for steel rows so we see the values summed
    print("\n[1b] AD 'Powder Qty Per Part' formulas, steel rows 38-56:")
    for r in range(38, 57):
        v = ws.cell(row=r, column=30).value  # AD = 30
        if v not in (None, ""):
            print(f"    AD{r}={v!r}")

    # 2. Powder price anywhere?
    print("\n[2] Any POWDER PRICE (£/kg) — scan whole sheet for 'powder' + price/kg/£ nearby:")
    found_price = False
    for r in range(1, 200):
        for c in range(1, 40):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "powder" in v.lower():
                # look at neighbours for a number or £
                nbrs = []
                for dr in (0, 1):
                    for dc in range(-2, 5):
                        nv = ws.cell(row=r+dr, column=c+dc).value
                        if isinstance(nv, (int, float)) or (isinstance(nv, str) and ("£" in nv or "price" in nv.lower() or "kg" in nv.lower())):
                            nbrs.append(f"{L(c+dc)}{r+dr}={nv!r}")
                print(f"    {L(c)}{r}={v!r}  neighbours: {nbrs}")
                if nbrs:
                    found_price = True
    if not found_price:
        print("    (no obvious powder price / £-per-kg found near powder labels)")

    # 3. M67 total + surroundings
    print("\n[3] Material total (M67) and 'Total Material Cost' row:")
    for r in range(66, 69):
        cells = []
        for c in range(3, 14):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                cells.append(f"{L(c)}{r}={v!r}")
        if cells:
            print(f"    row {r}: " + " | ".join(cells))
    print(f"\n    M67 formula = {ws['M67'].value!r}")

    print("\n" + "=" * 90)
    print("Decisions this answers:")
    print("  - [1] Is Total Powder a real sum (kg) we can use, or an orphan label?")
    print("  - [2] Is there a £/kg to convert kg->cost? (No price -> can't cost material honestly)")
    print("  - [3] Can we add a powder term to M67 without double-counting P.Coat labour?")


if __name__ == "__main__":
    main()
