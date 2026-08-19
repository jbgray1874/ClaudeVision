"""
SDI Intelligence — Job Decision Report
========================================
Generates a detailed per-job report showing exactly:
  - Every part estimated
  - What material was used and WHY
  - Where the thickness came from
  - What operations were detected and how
  - Cost breakdown per part
  - Confidence level with explanation
  - What's certain vs what needs review
Added as "Decision Report" sheet to every estimate xlsx.
Also generates a standalone per-job summary.
Called from estimator.py:
    from job_decision_report import add_decision_report_sheet
    add_decision_report_sheet(wb, summary, scan_meta)
"""
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional

from geometry_inference import _has_geometry as _has_real_dxf_geometry

# WHERE A DECISION WAS TAKEN, named by the module that owns the ranks. This report exists
# to explain the costing; a source it cannot name is a decision it cannot explain.
try:
    from source_precedence import display_name as _display_source, was_measured as _was_measured
except Exception:                                                   # pragma: no cover
    def _display_source(s):                                         # type: ignore[misc]
        return str(s or "").replace("_", " ")

    def _was_measured(s):                                           # type: ignore[misc]
        return bool(s) and str(s) not in ("llm_extract", "llm_full_extract",
                                          "inference", "geometry_inference")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OK = True
except ImportError:
    _OK = False
# ── Colours ────────────────────────────────────────────────────────────────────
C_NAVY      = "1F3864"
C_BLUE      = "2F5496"
C_WHITE     = "FFFFFF"
C_HIGH      = "C6EFCE"
C_HIGH_TXT  = "276221"
C_MED       = "FFEB9C"
C_MED_TXT   = "7D6608"
C_LOW       = "FFC7CE"
C_LOW_TXT   = "9C0006"
C_BOUGHT    = "EDEDED"
C_LIGHT     = "EBF3FB"
C_ALT       = "F5F5F5"
C_BORDER    = "BDD7EE"
C_SECTION   = "D6E4F0"


def _is_bought_in(part: Dict) -> bool:
    """True when a part is a bought-in / catalogue component (not fabricated).

    Bought-in items have no fabrication material, so they must be kept OUT of the
    material-inference paths (DXF-filename tokens, part-number suffix heuristics).
    A code like BI-50CMLOOM ends in 'M' and BI-LEDLINKLIGHT ends in 'T'; without this
    guard the suffix heuristic mislabels them "-M → Mild Steel" / "-T → MDF/Timber".
    Detected by: normalized_material BOUGHT_IN, a 'bought_in' page role, the layer-2
    recogniser source, or the BI-/FIXING/VINYL code families.
    """
    # THE CANONICAL GRAPH ALREADY DECIDED THIS, and asking it beats re-deriving it. The
    # compiler classifies every node as assembly / bought_in / leaf and writes that back
    # onto the part as canonical_kind; this function was inferring the same fact from
    # materials, roles and code families, and disagreeing. 12422-24's 79814P is a wood screw
    # the graph calls bought_in, and this report gave it MFC and a 16mm fabrication
    # thickness because none of the heuristics below recognise a supplier's own part number.
    _kind = str(part.get("canonical_kind") or "").lower()
    if _kind == "bought_in":
        return True
    if _kind in ("assembly", "leaf"):
        return False
    if part.get("is_bought_in") is True:
        return True
    mat = str(part.get("normalized_material") or part.get("material") or "").upper()
    if mat == "BOUGHT_IN":
        return True
    roles = part.get("page_roles") or []
    if "bought_in" in [str(r).lower() for r in roles]:
        return True
    src = str(part.get("source") or "").lower()
    if "recogniser" in src or "bought_in" in src or "note_scan" in src:
        return True
    pn = str(part.get("part_number") or "").upper()
    if pn.startswith(("BI-", "FIXING", "VINYL", "PACKAGING", "DELIVERY")):
        return True
    return False


def _c(ws, row, col, value="", bold=False, bg=None, fg="000000",
       align="left", wrap=False, size=10, italic=False,
       num_fmt=None, border=False):
    cell = _writable(ws, row, col)
    if cell is None:
        return None
    cell.value = value
    cell.font = Font(name="Arial", bold=bold, color=fg,
                     size=size, italic=italic)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    if num_fmt:
        cell.number_format = num_fmt
    if border:
        s = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=s, right=s, top=s, bottom=s)
    return cell



def _writable(ws, row: int, col: int):
    """The cell at (row, col), or its merged-range ANCHOR.

    openpyxl exposes every cell of a merged range but only the top-left one accepts a
    value; the rest raise "'MergedCell' object attribute 'value' is read-only". wb_populate
    learned this on 11350, where it killed an entire estimator-input block after the
    heading. This tab appends below a totals row whose footer the template merges, so the
    same trap sits directly under the new block. Returns None when the range cannot be
    resolved, so the caller skips rather than raising: a report that dies explains nothing.
    """
    try:
        cell = ws.cell(row=row, column=col)
        if cell.__class__.__name__ != "MergedCell":
            return cell
        for rng in getattr(ws, "merged_cells", []).ranges:
            if (rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col):
                return ws.cell(row=rng.min_row, column=rng.min_col)
        return None
    except Exception:
        return None


def replace_generated_sheet(wb, title: str):
    """Create `title`, replacing any existing sheet of that name.

    openpyxl's create_sheet does NOT overwrite — asked for a name already in the book it
    silently returns "AI Provenance1". The estimators' template can already contain these
    tabs, and re-running a job against a workbook that has them produced BOTH: the stale
    sheet keeping the name an estimator looks for, and the fresh one hidden behind a
    suffix. 2085 shipped with both.

    These sheets are generated wholesale from the summary every run, so replacing is the
    correct semantic — there is nothing in the old one worth merging, and leaving it is
    worse than deleting it because it still reads as this job's provenance."""
    if title in wb.sheetnames:
        del wb[title]
    # Anything openpyxl previously suffixed is also a stale copy of this sheet.
    for _name in [n for n in wb.sheetnames
                  if n.startswith(title) and n[len(title):].isdigit()]:
        del wb[_name]
    return wb.create_sheet(title)


def _find_wb_sell_price_ref(wb) -> Optional[str]:
    """Locate the WB's Sell Price VALUE cell by scanning for its LABEL, and return a
    cross-sheet formula reference string like "='Estimate'!M143".

    WHY a formula, not a value: the WB computes Sell Price with its own Excel formulas
    that only evaluate when the file is opened (calc-on-load). At report-build time the
    value is still 0 in memory, so we cannot read it. Instead we point the report's total
    cell at the WB's cell with a live formula — Excel then shows the same authoritative
    number on both sheets.

    WHY scan for the label, not hardcode M143: the estimators' template layout shifts when
    the BOM block grows, moving the Sell Price down. Anchoring to the "Sell Price" label
    (which moves with its value) survives that; a hardcoded row would silently go stale.
    Returns None if the sheet/label is not found, so the caller falls back to the engine sum.
    """
    try:
        # The populated estimate lives on the "Estimate" sheet (CELL_MAP estimate_sheet).
        ws = None
        for name in ("Estimate", "estimate"):
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            return None
        # Scan for a cell whose text is (or contains) "Sell Price"; the £ value sits to its
        # right on the same row (from the template: label in col ~I-L, value in col M).
        for r in ws.iter_rows():
            for cell in r:
                v = cell.value
                if isinstance(v, str) and "sell" in v.lower() and "price" in v.lower():
                    # Find the first numeric/blank value cell to the RIGHT on this row.
                    # Prefer column M (the template's value column) if it is to the right.
                    label_col = cell.column  # 1-indexed
                    row_idx = cell.row
                    # Look rightward up to 8 columns for the value cell; take the furthest
                    # populated/known money column. Template value column is M (13).
                    # A VALUE IS A NUMBER OR A FORMULA — NEVER PROSE.
                    #
                    # This took the first non-empty cell to the right. An Excel formula reads
                    # as None through openpyxl, so the real value cell looks empty and the
                    # walk continues — and once the estimator-input banner was added to the
                    # right of the totals, it became the first populated cell on that row.
                    # The report then pointed its SELL PRICE at a sentence, and the audit
                    # tabs displayed a malformed number instead of the price.
                    #
                    # Text on this row is a caption, not a figure, so it is skipped and the
                    # column-M fallback below does its job as intended.
                    target_col = None
                    for c in range(label_col + 1, label_col + 9):
                        cc = ws.cell(row=row_idx, column=c)
                        _v = cc.value
                        if _v is None or _v == "":
                            continue
                        if isinstance(_v, str) and not _v.lstrip().startswith("="):
                            continue          # a caption, not the value
                        target_col = c
                        break
                    # If nothing populated found (value is a not-yet-calculated formula that
                    # openpyxl read as 0/None), fall back to column M on the label's row.
                    if target_col is None:
                        target_col = 13  # M
                    col_letter = get_column_letter(target_col)
                    return f"='{ws.title}'!{col_letter}{row_idx}"
        return None
    except Exception:
        return None


def _conf_info(part: Dict, unit_cost: float = 0.0, summary: Any = None) -> tuple:
    """(score, label, bg, fg, explanation) — from the ONE shared confidence authority.

    This function used to carry its own ladder, and estimation_report carried a different
    one, and calibration.py a third. The two report tabs disagreed about the same part in
    the same file: 2085-01 came out 0.92 HIGH here and 0.70 MEDIUM there. Its ladder also
    claimed "material from DXF filename" without ever reading material_source, so the
    confidence REASON contradicted the Material Source column on its own row.

    The score survives only because callers threshold on it. It is derived FROM the status
    the shared assessment returned — never averaged, because averaging is how a part with
    no material price at all lands in the sixties and reads as partial knowledge.
    """
    from confidence import (assess_part, STATUS_FILL, UNKNOWN, ASSUMED,
                            REPORTED, MEASURED, CONFIRMED)
    _a = assess_part(part, summary)
    _status = _a["overall"]
    _bg, _fg = STATUS_FILL.get(_status, ("EDEDED", "555555"))
    _score = {UNKNOWN: 0.20, ASSUMED: 0.45, REPORTED: 0.70,
              MEASURED: 0.90, CONFIRMED: 0.99}.get(_status, 0.20)
    _why = _a.get("reason") or "; ".join(_a.get("decided_by") or [])
    return _score, _a["overall_label"], _bg, _fg, _why


def _mat_source_explanation(part: Dict) -> str:
    """Plain English explanation of why this material was chosen."""
    # Bought-in components have NO fabrication material — never run them through the
    # DXF-token / part-number-suffix material heuristics (BI-...T would misread as
    # "-T → MDF/Timber", BI-...M as "-M → Mild Steel"). Return an honest bought-in note.
    if _is_bought_in(part):
        # "Priced from catalogue/history" states a provenance that does not exist on a
        # line carrying no price. PACKAGING and DELIVERY are per-unit shares an estimator
        # fills in, and they are the two lines most likely to be forgotten — so they say so.
        from costed_facts import is_placeholder_price as _placeholder
        if _placeholder(part):
            return ("Bought-in / no fabrication material — NOT YET PRICED, estimator to "
                    "enter a per-unit figure")
        return "Bought-in / catalogue component — no fabrication material (priced from catalogue/history)"
    mat = str(part.get("normalized_material") or part.get("material") or "").upper()
    src = str(part.get("material_source") or "")
    geo = str(part.get("geometry_source") or "")
    dxf = str(part.get("dxf_source_file") or "")
    pn  = str(part.get("part_number") or "")
    if "knowledge_base" in src:
        return f"✅ SDI Knowledge Base — previously confirmed by estimator"
    if "_MS_" in dxf.upper() or "MS_" in dxf.upper():
        return f"✅ DXF filename contains '_MS_' → Mild Steel"
    if "PETG" in dxf.upper():
        return f"✅ DXF filename contains 'PETG' → Acrylic"
    if "JOINERY" in dxf.upper():
        return f"✅ DXF filename contains 'JOINERY' → MDF/Timber"
    if pn and pn.endswith("M"):
        return f"✅ Part number suffix '-M' → Mild Steel (SDI naming convention)"
    if pn and pn.endswith("A"):
        return f"✅ Part number suffix '-A' → Acrylic (SDI naming convention)"
    if pn and pn.endswith("T"):
        return f"✅ Part number suffix '-T' → MDF/Timber (SDI naming convention)"
    if mat == "MILD_STEEL" and "pdf" in geo:
        return "⚡ PDF drawing text — 'MILD STEEL' found in title block"
    if mat == "MDF":
        return "⚡ PDF drawing text — 'MDF' or 'FSC ACCREDITED' found"
    if mat == "BOUGHT_IN":
        return "✅ Description contains bought-in keyword (hinge/magnet/fixing)"
    if not mat or mat in ("UNKNOWN","LED","CARD"):
        return "⚠ Material unresolved — check drawing title block"
    return f"⚡ AI inference from drawing context"


def _thk_source_explanation(part: Dict) -> str:
    """Plain English explanation of thickness source.
    DXF filename checked FIRST (most reliable, avoids real 2mm/3mm acrylic
    being discarded as tolerance-table values)."""
    # Bought-in components have no fabrication thickness — don't imply one.
    if _is_bought_in(part):
        return "— bought-in component (no fabrication thickness)"

    # THE THICKNESS THE JOB WAS COSTED AT, BEFORE ANY RE-DERIVATION.
    #
    # This function read `thicknesses_mm` — every thickness-looking number found on the
    # drawing — and took the first. On 12422-24 that reported 01J as 16mm while the sheet
    # costed a 28mm MFC panel, and 02M as 16mm against 1.5mm steel. The report was not
    # describing the estimate; it was performing its own extraction and publishing the
    # result beside the estimate's numbers as though they agreed.
    #
    # normalized_thickness_mm is the arbitrated datum, written through the resolver with a
    # recorded source. Asking it is what makes this a REPORT of the costing rather than a
    # second opinion about the drawing. The derivation below is kept for records that have
    # no arbitrated value — it is a fallback now, not the first answer.
    _costed = part.get("normalized_thickness_mm")
    try:
        _costed = float(_costed) if _costed not in (None, "") else None
    except (TypeError, ValueError):
        _costed = None
    if _costed and _costed > 0:
        _src = ""
        try:
            from source_precedence import source_of
            _src = str(source_of(part, "normalized_thickness_mm") or "")
        except Exception:
            _src = ""
        # NAMED BY THE MODULE THAT OWNS THE RANKS, not by a copy kept here. This was a
        # private eight-entry table, so a source the waterfall knows about but the table
        # did not — mirror_of_measured, pdf_overall_dims, override_rule — rendered as a
        # bare internal key in the one document written to explain the decision.
        _pretty = _display_source(_src)
        _mark = "✅" if _was_measured(_src) else "⚡"
        return (f"{_mark} {_costed:g}mm — the thickness this job was COSTED at"
                + (f", from {_pretty}" if _pretty else " (source not recorded)"))

    import re
    dxf  = str(part.get("dxf_source_file") or "")
    geo  = str(part.get("geometry_source") or "")
    tol  = {0.5, 1.0, 1.5, 2.0, 3.0}
    # 1. DXF filename thickness — most reliable
    m = re.search(r'[_\-\s](\d+\.?\d*)\s*mm', dxf, re.IGNORECASE)
    if m:
        tv = float(m.group(1))
        if 0.3 <= tv <= 25.0:
            return f"✅ {tv}mm — from DXF filename: {dxf}"
    # 2. thicknesses_mm — only strip tolerance values if the FULL sequence is
    #    present (a standalone 2.0/3.0 is a real thickness, not table noise).
    thks = part.get("thicknesses_mm") or []
    thk_set = {round(float(t), 1) for t in thks if t}
    if tol.issubset(thk_set):
        real_thks = [t for t in thks if t and round(float(t), 1) not in tol]
    else:
        real_thks = [t for t in thks if t]
    if real_thks:
        thk = real_thks[0]
        if "dxf" in geo:
            return f"✅ {thk}mm — from DXF geometry / drawing dimensions"
        return f"⚡ {thk}mm — extracted from PDF drawing text"
    if thks and tol.issubset({round(float(t),1) for t in thks if t}):
        return "⚠ Tolerance table values only — real thickness not extracted"
    return "⚠ No thickness found — assembly-only page or missing dimension"


def _ops_explanation(part: Dict, est: Optional[Dict] = None,
                     summary: Optional[Dict] = None) -> str:
    """Explain how operations were determined.

    Driven by what we actually COSTED where the part estimate is available, not by the raw
    textual/inferred op lists. Those lists are the drawing's interpretation, and on these
    packs the shared specification legend puts processes on parts that never carry them —
    which is how "laser cutting" and "powder coating" ended up described against timber
    panels that the Estimate sheet charges only saw, glue, CNC and spray for. A provenance
    sheet that describes operations the estimate does not contain is worse than no
    provenance: it reads as evidence for a route we did not price. Same rule already
    applied to the client quote.
    """
    # Canonical where the workbook has run (the route the Estimate sheet two tabs away
    # actually charges); the part's own PRE-FILTER costed fields only as a fallback.
    from costed_facts import operations_for_part, priced_route_known
    ops: List[str] = operations_for_part(summary, part.get("part_number"), est)
    if not ops and priced_route_known(summary):
        # THE PRICED ROUTE IS KNOWN AND THIS PART IS IN NONE OF IT.
        #
        # That is an answer, not a gap, and the honest one is to say so. The old fallback
        # reached for the drawing's textual + inferred lists here, which is where the
        # suppressed route came back: every part a gate removed — powder on a timber panel,
        # weld/dress on an artefact record — lost its costed evidence and was then described
        # from the specification legend instead. The report ended up narrating exactly the
        # operations the workbook had just decided against.
        return ("No operation charged on this job — the priced route contains this part "
                "in no labour row")
    _priced = True
    if not ops:
        # No workbook yet (a report generated from a JSON alone). Nothing has been priced,
        # so the drawing's own reading is the best available evidence — and it is labelled
        # as such below rather than presented as what we charged.
        ops = list(part.get("textual_operations") or []) + list(part.get("inferred_operations") or [])
        _priced = False
    mat = str(part.get("normalized_material") or part.get("material") or "").upper()
    geo = str(part.get("geometry_source") or "")
    if not ops:
        return "No operations detected — assembly-only record"
    # EVERY OPERATION GETS NAMED.
    #
    # This was a whitelist of eight phrases, and any operation outside it vanished from the
    # sentence. On 2085 the tubes read "powder coating, welding" here while AI Provenance —
    # two tabs away in the same file, from the same source — listed tube_cut, welding,
    # dress_welds, powder_coating and assembly. Same job, two different routes on the page,
    # because a renderer that can only describe what it recognises silently under-reports
    # everything else, and the tube cut is exactly the operation this engine has already
    # lost once.
    #
    # An operation with no written explanation is still an operation. Naming it plainly
    # beats omitting it — the sheet charges it either way.
    _PHRASES = {
        "cnc_routing":        "CNC routing (timber/MDF material)",
        "folding":            "folding (bend lines in DXF)",
        "powder_coating":     "powder coating (finish specified in drawing)",
        "wet_spray":          "wet spray (finish specified in drawing)",
        "edge_banding":       "edge banding (MDF/timber with 'EDGED' finish)",
        "welding":            "welding (assembly drawing indicates welds)",
        "handling":           "handling / assembly (bench time)",
        "assembly":           "handling / assembly (bench time)",
        "dress_welds":        "weld dressing (linishing the welded joint)",
        "tube_cut":           "tube cutting (sawn to length — section stock has no flat blank)",
        "tube_bending":       "tube bending (tube bender, not the press brake)",
        "linebend":           "line bending (acrylic heat-bent, not press-braked)",
        "line_bending":       "line bending (acrylic heat-bent, not press-braked)",
        "diamond_polish":     "diamond polishing (acrylic edge finish)",
        "hardware_insertion": "hardware insertion (PEM / clinch studs pressed in)",
        "saw":                "sawing to length",
        "guillotine":         "guillotine shearing",
        "punch":              "punching",
        "hole_machining":     "hole machining / drilling",
        "drilling":           "hole machining / drilling",
        "tapping":            "tapping",
        "deburring":          "deburring",
        "linishing":          "linishing",
        "robomac":            "Robomac cutting (wire / bar stock)",
    }
    sources: List[str] = []
    _seen = set()
    for _op in ops:
        _key = str(_op).strip().lower()
        if not _key or _key in _seen:
            continue
        _seen.add(_key)
        if _key == "laser_cutting":
            _phrase = ("laser cutting (flat DXF detected)" if "dxf" in geo
                       else "laser cutting (inferred from material/geometry)")
        else:
            _phrase = _PHRASES.get(_key, _key.replace("_", " "))
        if _phrase not in sources:
            sources.append(_phrase)
    # "Read from the drawing" is not "charged". Saying which one this is costs a word and
    # stops an unpriced reading being mistaken for the route the sheet contains.
    _lead = "Operations" if _priced else "Read from drawing — NOT YET PRICED"
    return f"{_lead}: " + ", ".join(sources) if sources else \
           f"{_lead}: {', '.join(ops)}"



def decisions_that_required_resolution(summary: Dict[str, Any]) -> List[Dict[str, Any]]:
    """The operations where two equally-ranked sources disagreed and the arbiter chose.

    THE FIRST THING AN ESTIMATOR SHOULD READ, and the reason it gets its own block at the
    top of the sheet rather than a column halfway down a hundred rows. Every other line on
    this tab describes a decision that made itself: one strongest source, nothing at that
    rank contradicting it. These are the lines where the engine had to pick, and they are
    where fine-tuning feedback is worth most.

    Ordered weakest-resolution first: a contest settled by the claim-id backstop had no
    distinguishing evidence at all and is a coin flip made reproducible, which deserves a
    person's attention before one settled by the drawing's own words.
    """
    payload = ((summary.get("estimate_summary") or {}).get("canonical_route_shadow")
               or summary.get("canonical_route_shadow") or {})
    out = []
    for d in (payload.get("decisions") or []):
        if isinstance(d, dict) and d.get("contested"):
            out.append(d)
    def _worst_first(d):
        key = str(d.get("settled_by_key") or "")
        # later in RESOLUTION_KEYS == weaker ground for the choice
        try:
            from route_compiler import RESOLUTION_KEYS
            return (-RESOLUTION_KEYS.index(key) if key in RESOLUTION_KEYS else 1,
                    str(d.get("target_id") or ""))
        except Exception:
            return (0, str(d.get("target_id") or ""))
    out.sort(key=_worst_first)
    return out


# The DATA contests, by the invariant code that already finds them. Read rather than
# recomputed: the comparison rules are fiddly and earned (a material spelling variant is not a
# disagreement; a gauge must differ by enough to move the money), and a second copy here would
# be the two-copies-of-one-rule defect this codebase keeps digging out — with the tab and the
# review block quietly disagreeing about the same part.
_DATUM_CONTEST_CODES = {
    "two_sources_disagree_about_the_material": "material",
    "two_sources_disagree_about_the_gauge": "gauge",
}


def datum_decisions_that_required_resolution(summary: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Contests about the DATA the price is computed from — what the part is made of, and how
    thick — alongside the operation contests already listed.

    THE OPERATION CONTESTS WERE THE ONLY ONES SHOWN, AND THEY ARE NOT THE EXPENSIVE ONES. On
    8352 this block held two lines, both 'powder coating: not_applicable', while the same run
    reported a back panel costed as PLYWOOD that inference read as TIMBER, and three parts whose
    gauge two sources put 4x apart — 12mm against 3mm on the panel that sets the sheet rate.
    Material decides the rate and whether the part has a rate at all; gauge decides the rate and
    steps the cut time, so a part costed on the wrong one is wrong twice. Those are the decisions
    worth a person's minute, and the tab that exists to show decisions did not show them.

    Ordered worst-first by how far the two readings are apart, so the 4x argument is read before
    the 1.1x one.
    """
    if not isinstance(summary, dict):
        return []
    out: List[Dict[str, Any]] = []
    for violation in ((summary.get("invariants") or {}).get("violations") or []):
        if not isinstance(violation, dict):
            continue
        datum = _DATUM_CONTEST_CODES.get(str(violation.get("code") or ""))
        if not datum:
            continue
        for d in ((violation.get("detail") or {}).get("parts") or []):
            if not isinstance(d, dict):
                continue
            out.append({
                "part_number": d.get("part_number"),
                "datum": datum,
                "costed_as": d.get("costed_as"),
                "costed_from": d.get("costed_from"),
                "other": d.get("other"),
                "other_from": d.get("other_from"),
                # Only the gauge check measures how far apart the readings are; a material
                # disagreement is categorical, so it sorts after the measured ones.
                "ratio": d.get("ratio"),
            })
    out.sort(key=lambda d: (-(float(d.get("ratio") or 0)), str(d.get("part_number") or "")))
    return out


def powder_authority(summary: Dict[str, Any]) -> str:
    """Who decided powder on this job, in one sentence, naming source and rank.

    Never a silent zero and never a competing gate. Powder has twice this month produced a
    figure on a sheet that no reader could trace to a decision -- once as phantom mass from
    a misclassified plastic, once from a geometry sum that had never consulted the route.
    A cell that says who decided makes both failures visible the moment they recur.
    """
    payload = ((summary.get("estimate_summary") or {}).get("canonical_route_shadow")
               or summary.get("canonical_route_shadow") or {})
    decisions = [d for d in (payload.get("decisions") or []) if isinstance(d, dict)]
    powder = [d for d in decisions
              if str(d.get("operation") or "").strip().lower() in
              ("powder_coating", "powder_coat", "powder", "p_coat", "pcoat")]
    if not decisions:
        return ("Powder: decided by the LEGACY FINISH GATE - no compiled route on this job. "
                "Not arbitrated, and not traceable to a ranked source.")
    if not powder:
        return ("Powder: NO DECISION on this route. Any powder mass came from geometry, "
                "not from an arbitrated decision - treat it as unverified.")
    required = [d for d in powder if str(d.get("status") or "").lower() == "required"]
    if not required:
        return ("Powder: decided by the route compiler - NOTHING COATED on this job. "
                "A powder figure on this sheet would contradict the route.")
    best = max(required, key=lambda d: int(d.get("source_rank") or 0))
    return ("Powder: decided by the route compiler ({} part(s) coated; strongest source "
            "{}, rank {}).".format(len(required),
                                   best.get("decided_by") or best.get("source") or "unrecorded",
                                   best.get("source_rank") or "?"))


def add_decision_report_sheet(wb, summary: Dict[str, Any],
                               scan_meta: Dict[str, Any] = None) -> None:
    """Add a 'Decision Report' sheet to an existing workbook."""
    if not _OK:
        return

    # THE SAME ROWS THE ESTIMATE SHEET HAS. job_parts is the canonical list the sheet was
    # built from, each record overlaid on its manufacturing_writeup entry so the provenance
    # columns below still have the geometry, thickness and material-source fields they read.
    # Iterating manufacturing_writeup directly put this tab on a different set of rows from
    # the sheet two tabs away: a bought-in the sheet charges appearing nowhere here, a
    # recogniser-minted duplicate appearing here and not there.
    from costed_facts import job_parts
    parts = job_parts(summary) or (
        (summary.get("manufacturing_writeup") or {}).get("parts") or [])
    if not parts:
        return

    # SDI Intelligence — cost lives in estimate_summary.part_estimates,
    # keyed by part_number. Build a lookup so the report shows real costs.
    #
    # PRE-FILTER, and the columns fed from it say so. These are the engine's own per-part
    # numbers; the Estimate sheet's totals are calculated by Excel from the accepted labour
    # and material rows, and the two are different calculators. Per-part cost is not
    # recoverable from the sheet — a labour row is a department's batch value across every
    # part in the group — so the honest presentation is the engine figure, labelled as the
    # engine figure, reconciled against the workbook below.
    _est_lookup = {}
    for _pe in (summary.get("estimate_summary") or {}).get("part_estimates", []):
        _pn = _pe.get("part_number")
        if _pn:
            _est_lookup[_pn] = _pe

    from costed_facts import (canonical_quantity, decision_ids_for_part,
                              is_placeholder_price, job_totals, part_material_cost,
                              priced_route_known, priced_rows_for_part)
    _totals = job_totals(summary)
    _canonical = priced_route_known(summary)

    scan_meta = scan_meta or {}
    ws = replace_generated_sheet(wb, "Decision Report")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_BLUE
    # ── Column widths ──────────────────────────────────────────────────────────
    col_widths = [16, 30, 6, 14, 10, 34, 34, 34, 10, 10, 14, 26]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    _c(ws, 1, 1, "SDI Intelligence — Estimate Decision Report",
       bold=True, bg=C_NAVY, fg=C_WHITE, align="center", size=14)
    ws.row_dimensions[1].height = 30
    pdf_name = scan_meta.get("pdf_name") or summary.get("source_file") or "—"
    job_no   = scan_meta.get("job_number") or "—"
    # THE HEADLINE FIGURE IS THE WORKBOOK'S. Summing the engine's labour-inclusive per-part
    # totals put "Total Estimate: £44.75" at the top of a 2085 sheet the workbook priced at
    # £6.33 — and repeated it in the total row whenever the Sell Price cell could not be
    # located. The engine sum is only a fallback for a job that never reached Excel.
    _wb_unit = _totals.get("unit_gbp")
    total = float(_wb_unit) if _wb_unit is not None else sum(
        float(_est_lookup.get(p.get("part_number"), {}).get("extended_total_cost_gbp") or 0)
        for p in parts)

    # The AUTHORITATIVE total is the WB's Sell Price (computed by the WB's own formulas).
    # Find its cell by label so the report can reference it live; falls back to the engine
    # part-sum (`total`) if the WB sheet/label is not present.
    _sell_ref = _find_wb_sell_price_ref(wb)

    ds = (summary.get("estimate_summary") or {}).get("data_sufficiency") or {}
    if ds.get("status") == "insufficient_data":
        # Do NOT print the engine part-sum (ds.document_total_provisional_gbp) as "the
        # provisional total" — it is a different calculator from the workbook Sell Price and
        # can differ materially (Horti Crate: engine £102.07 vs workbook £46.53), which made
        # this header contradict the SELL PRICE total row below and the report/quote HTML.
        # Warn here; the authoritative provisional figure is the Sell Price shown below.
        total_line = (
            f"⚠ INSUFFICIENT DATA — PROVISIONAL, NOT for quoting "
            f"(credible {float(ds.get('credible_cost_ratio') or 0) * 100:.0f}% · "
            f"DXF {float(ds.get('dxf_part_ratio') or 0) * 100:.0f}% of parts) — see total below"
        )
    else:
        # Prefer the WB Sell Price; the header text still shows the engine sum as a
        # reference, but the authoritative figure is the WB's (see the TOTAL row below).
        if _sell_ref:
            total_line = "Total (Sell Price): see Estimate sheet — mirrored below"
        elif _wb_unit is not None:
            total_line = f"Unit cost (calculated by the Estimate sheet): £{total:,.2f}"
        else:
            total_line = f"Total Estimate (engine part-sum — no workbook): £{total:,.2f}"
    ws.merge_cells("A2:L2")
    _c(ws, 2, 1,
       f"Drawing: {pdf_name}   |   Job: {job_no}   |   "
       f"{total_line}   |   "
       f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   "
       f"SDI Intelligence — wearesdi.com",
       bg=C_BLUE, fg=C_WHITE, align="center", size=10)
    ws.row_dimensions[2].height = 16
    # ── Key ────────────────────────────────────────────────────────────────────
    ws.merge_cells("A3:L3")
    _c(ws, 3, 1,
       "CONFIDENCE:   ✅ HIGH — DXF matched or knowledge base confirmed   "
       "⚡ MEDIUM — PDF extraction or AI inference   "
       "⚠ REVIEW — zero cost, unknown material or assembly-only record",
       bg="F0F0F0", size=9, italic=True)
    ws.row_dimensions[3].height = 14
    ws.row_dimensions[4].height = 6  # spacer
    # ── Column headers ─────────────────────────────────────────────────────────
    # The money columns are the ENGINE's per-part figures, and once a workbook exists they
    # are not what the sheet charges. Naming the basis in the header is the difference
    # between a working figure and a price the reader will quote from.
    # Material, not "cost". See the per-part note below: labour has no per-part figure.
    _money_basis = " material" if _canonical else ""
    headers = [
        "Part Number", "Description", "Qty",
        "Material", "Thickness",
        "Material Source — WHY",
        "Thickness Source — WHY",
        "Operations — HOW DETECTED",
        f"Unit £{_money_basis}", f"Ext £{_money_basis}", "Confidence",
        "Priced by — sheet row / decision",
    ]
    # ── ROW 4: WHAT AN ESTIMATOR SHOULD READ BEFORE THE TABLE ──────────────────
    # Two facts that decide how much of the rest to trust, on the sheet itself rather than
    # in a log nobody opens: who owns powder, and how many decisions the engine had to
    # settle rather than simply read. Powder is here because it has twice this month put a
    # figure on a sheet that no reader could trace to a decision.
    _contested = decisions_that_required_resolution(summary)
    _datum_contested = datum_decisions_that_required_resolution(summary)
    _banner = powder_authority(summary)
    # BOTH KINDS COUNTED. The operation contests were the only ones named here, and on a job
    # whose real arguments were about material and gauge the banner read "2 decisions" while
    # four costlier ones sat unmentioned two tabs away in the review list.
    _n = len(_contested) + len(_datum_contested)
    if _n:
        _bits = []
        if _contested:
            _bits.append(f"{len(_contested)} operation")
        if _datum_contested:
            _bits.append(f"{len(_datum_contested)} material/gauge")
        _banner += (f"   |   {_n} decision(s) required resolution "
                    f"({', '.join(_bits)}) — see the blocks below the table.")
    else:
        _banner += "   |   No decision required resolution: no two equal sources disagreed."
    _c(ws, 4, 1, _banner, bold=True, size=9,
       bg=(C_LOW if _contested else C_LIGHT),
       fg=(C_LOW_TXT if _contested else "000000"), wrap=False)

    for ci, hdr in enumerate(headers, 1):
        _c(ws, 5, ci, hdr, bold=True, bg=C_NAVY, fg=C_WHITE,
           align="center", size=9, border=True)
    ws.row_dimensions[5].height = 20
    # ── Part rows ──────────────────────────────────────────────────────────────
    row = 6
    review_parts = []
    for i, part in enumerate(parts):
        pn   = str(part.get("part_number") or "—")
        desc = str(part.get("description") or "—")
        # QUANTITY PER TOP-LEVEL UNIT, not per parent.
        #
        # A BOM row states how many the PARENT takes. For anything reached through a
        # sub-assembly that is not the quantity the job needs: a knob at qty 2 inside a
        # sub-assembly used twice is 4 per unit. The compiled hierarchy rolls the
        # multiplicity through, and the workbook already charges on the rolled figure —
        # so the raw row quantity here made the report disagree with its own Estimate sheet
        # for every part below the first level. Falls back to the row when the graph does
        # not know the part.
        _cq = canonical_quantity(summary, pn)
        qty  = int(_cq) if _cq is not None and float(_cq).is_integer() else (
            _cq if _cq is not None else int(part.get("quantity") or 1))
        # Bought-in components carry no fabrication material — show a clean label
        # instead of a defaulted/mis-inferred one (e.g. "MILD_STEEL" on a foam tape).
        if _is_bought_in(part):
            mat = "Bought-in"
        else:
            mat = str(part.get("normalized_material") or part.get("material") or "—")
        # PER-PART MONEY IS MATERIAL ONLY.
        #
        # unit_total_cost_gbp is an engine-era, LABOUR-INCLUSIVE apportionment. On 2085 it
        # put GBP 19.25 against each tube on a sheet whose entire unit price is GBP 6.33 —
        # a number in a column headed like a cost that reconciles to nothing, on the two
        # parts whose material could not be read at all. Labour is charged per DEPARTMENT
        # ROW as a batch value across every part in that setup, so there is no per-part
        # labour figure to show; material is costed per part and sums to the workbook's own
        # material total, so that is what this column is and what it now says.
        if _canonical:
            unit, ext = part_material_cost(part)
        else:
            unit = float(_est_lookup.get(pn, {}).get("unit_total_cost_gbp") or 0)
            ext  = float(_est_lookup.get(pn, {}).get("extended_total_cost_gbp") or 0)
        # Thickness — real value only
        # Thickness column — DXF filename first, then non-tolerance values
        import re as _re_t
        _dfn_t = str(part.get("dxf_source_file") or "")
        _m_t = _re_t.search(r'[_\-\s](\d+\.?\d*)\s*mm', _dfn_t, _re_t.IGNORECASE)
        thks     = part.get("thicknesses_mm") or []
        tol      = {0.5, 1.0, 1.5, 2.0, 3.0}
        if _is_bought_in(part):
            real_thk = "—"
        elif _m_t and 0.3 <= float(_m_t.group(1)) <= 25.0:
            real_thk = f"{float(_m_t.group(1)):.1f}mm"
        else:
            _thk_set = {round(float(t),1) for t in thks if t}
            if tol.issubset(_thk_set):
                real_thk = next((f"{float(t):.1f}mm" for t in thks
                                 if t and round(float(t), 1) not in tol), "—")
            else:
                real_thk = next((f"{float(t):.1f}mm" for t in thks if t), "—")
        conf, conf_label, conf_bg, conf_fg, conf_expl = _conf_info(part, unit, summary)
        mat_why  = _mat_source_explanation(part)
        thk_why  = _thk_source_explanation(part)
        ops_why  = _ops_explanation(part, _est_lookup.get(pn), summary)
        bg = C_ALT if i % 2 == 0 else C_WHITE
        if _is_bought_in(part):
            bg = C_BOUGHT
        _c(ws, row, 1,  pn,         bg=bg, bold=True, size=9, border=True)
        _c(ws, row, 2,  desc,       bg=bg, size=9,    border=True, wrap=True)
        _c(ws, row, 3,  qty,        bg=bg, align="center", size=9, border=True)
        _c(ws, row, 4,  mat,        bg=conf_bg, bold=True, fg=conf_fg,
           size=9, border=True)
        _c(ws, row, 5,  real_thk,   bg=bg, align="center", size=9, border=True)
        _c(ws, row, 6,  mat_why,    bg=bg, size=8, border=True, wrap=True)
        _c(ws, row, 7,  thk_why,    bg=bg, size=8, border=True, wrap=True)
        _c(ws, row, 8,  ops_why,    bg=bg, size=8, border=True, wrap=True)
        _c(ws, row, 9,  unit if unit > 0 else "—",
           bg=bg, align="right", bold=True, size=9,
           num_fmt="£#,##0.00", border=True)
        _c(ws, row, 10, ext if ext > 0 else "—",
           bg=bg, align="right", bold=True, size=9,
           num_fmt="£#,##0.00", border=True)
        _c(ws, row, 11, conf_label, bg=conf_bg, fg=conf_fg,
           align="center", bold=True, size=8, border=True)
        # ── Traceability: part -> the sheet rows charging it -> the decisions behind them.
        # Without this the Decision Report asserts a route and offers nothing to check it
        # against; with it every line on the page can be walked back to the Estimate tab and
        # forward to the compiler decision that put it there.
        _prows = priced_rows_for_part(summary, pn) if _canonical else []
        if _prows:
            _wbrows = sorted({int(float(r["workbook_row"])) for r in _prows
                              if r.get("workbook_row")})
            _dids = decision_ids_for_part(summary, pn)
            _trace = ("Estimate row " + ", ".join(str(r) for r in _wbrows)) if _wbrows else ""
            if _dids:
                _trace = (_trace + "\n" if _trace else "") + " · ".join(_dids)
        elif _canonical:
            _trace = "not priced on any labour row"
        else:
            _trace = "— no workbook built"
        _c(ws, row, 12, _trace, bg=bg, size=8, border=True, wrap=True)
        ws.row_dimensions[row].height = 36
        row += 1
        if (conf < 0.5 or unit == 0) and not _is_bought_in(part):
            review_parts.append((pn, desc, conf_expl))
    # ── Total row ──────────────────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    # Label reflects which number is shown: WB Sell Price (authoritative) if we found it,
    # otherwise the engine part-sum.
    _total_label = ("SELL PRICE (from Estimate sheet)" if _sell_ref
                    else "UNIT COST (calculated by the Estimate sheet)" if _wb_unit is not None
                    else "TOTAL ESTIMATE (engine part-sum — no workbook)")
    _resolution_row = row + 2                  # written after the total, see below
    _c(ws, row, 1, _total_label, bold=True, bg=C_NAVY, fg=C_WHITE,
       align="right", size=12)
    _c(ws, row, 9, "", bg=C_NAVY)
    if _sell_ref:
        # Live cross-sheet formula — Excel computes it on open, so the report total always
        # equals the WB's authoritative Sell Price (no duplicated maths, no drift).
        _tc = _c(ws, row, 10, _sell_ref, bold=True, bg=C_NAVY, fg=C_WHITE,
                 align="right", size=13, num_fmt="£#,##0.00")
    else:
        _c(ws, row, 10, total, bold=True, bg=C_NAVY, fg=C_WHITE,
           align="right", size=13, num_fmt="£#,##0.00")
    _c(ws, row, 11, "", bg=C_NAVY)
    _c(ws, row, 12, "", bg=C_NAVY)
    ws.row_dimensions[row].height = 24
    # ── Reconciliation: the engine part-sum against what Excel calculated ───────
    # Two calculators, both on this page: the Ext £ column sums the engine's per-part
    # figures, the total row shows the workbook's. They are not the same arithmetic and on
    # real jobs they differ materially. Leaving the reader to notice — and to guess which to
    # believe — is what made this sheet read as the engine contradicting itself.
    if _totals["source"] == "excel_calculated":
        row += 1
        ws.merge_cells(f"A{row}:L{row}")
        # RECONCILE WHAT IS ON THE PAGE. This used to quote the engine part-sum — GBP 44.75
        # on 2085 — as the figure to reconcile against. Now that the per-part columns are
        # material, that number is not on this sheet and does not belong in its arithmetic
        # either: it is an obsolete labour-inclusive total, and repeating it in the
        # explanation put it back on the page the fix had just removed it from.
        _unit = float(_totals.get("unit_gbp") or 0.0)
        _mat = _totals.get("material_gbp")
        _lab = _totals.get("labour_gbp")
        _col_mat = sum(part_material_cost(p)[1] for p in parts) if _canonical else 0.0
        _mat_txt = (f"The material column above sums to £{_col_mat:,.2f} against the "
                    f"sheet's £{float(_mat):,.2f}. " if _mat is not None else "")
        _lab_txt = (f"Labour is £{float(_lab):,.2f} and is charged per department row — a "
                    f"batch value across every part in that setup, with no per-part figure "
                    f"to show. The 'Priced by' column names the rows and the decisions "
                    f"behind each part's share. " if _lab is not None else "")
        _c(ws, row, 1,
           f"RECONCILIATION — the Estimate sheet calculated £{_unit:,.2f} per unit. "
           f"{_mat_txt}{_lab_txt}The workbook is authoritative.",
           bg=C_LIGHT, size=9, wrap=True, italic=True)
        ws.row_dimensions[row].height = 34
    # ── Parts requiring review ─────────────────────────────────────────────────
    if review_parts:
        row += 2
        ws.merge_cells(f"A{row}:L{row}")
        _c(ws, row, 1, f"⚠  PARTS REQUIRING REVIEW ({len(review_parts)} items)",
           bold=True, bg="FFC7CE", fg="9C0006", size=11)
        ws.row_dimensions[row].height = 20
        row += 1
        for pn, desc, reason in review_parts:
            ws.merge_cells(f"B{row}:L{row}")
            _c(ws, row, 1, "⚠", bg="FFC7CE", align="center", size=9)
            _c(ws, row, 2, f"{pn}  —  {desc}  |  {reason}",
               bg="FFC7CE", fg="9C0006", size=9, wrap=True)
            ws.row_dimensions[row].height = 18
            row += 1
    # ── Insufficient data / unreliable-cost section ────────────────────────────
    if ds.get("status") == "insufficient_data":
        row += 2
        ws.merge_cells(f"A{row}:L{row}")
        _c(ws, row, 1, "⚠  INSUFFICIENT DATA — DO NOT QUOTE FROM THIS TOTAL",
           bold=True, bg="FFC7CE", fg="9C0006", size=11)
        ws.row_dimensions[row].height = 20
        row += 1
        ws.merge_cells(f"A{row}:L{row}")
        # Do NOT cite a second, static "provisional total" here. The authoritative total is
        # the workbook Sell Price shown in the SELL PRICE row directly above (a live cross-
        # sheet formula). The engine part-sum (ds.document_total_provisional_gbp) is a
        # different calculator and can differ materially — e.g. on the Horti Crate the engine
        # part-sum was £102.07 while the workbook Sell Price was £46.53. Printing that figure
        # here made the Decision Report contradict its own total row (and the report/quote
        # HTML). The banner's job is to WARN; the number is already above it.
        _c(ws, row, 1,
           f"Most of this estimate is not DXF-backed. The total shown above is "
           f"PROVISIONAL and must not be quoted — request part DXFs first. "
           f"Credible share: {float(ds.get('credible_cost_ratio') or 0) * 100:.0f}% · "
           f"Part DXFs: {int(ds.get('parts_with_dxf') or 0)}/"
           f"{int(ds.get('fabricated_part_count') or 0)} fabricated parts.",
           bg="FFC7CE", fg="9C0006", size=9, wrap=True)
        ws.row_dimensions[row].height = 28
        row += 1
        for up in (ds.get("unreliable_parts") or [])[:12]:
            ws.merge_cells(f"B{row}:L{row}")
            _c(ws, row, 1, "✗", bg="FFC7CE", align="center", size=9)
            _c(ws, row, 2,
               f"{up.get('part_number')}  —  {up.get('description')}  |  "
               f"£{float(up.get('extended_cost_gbp') or 0):,.2f}  —  "
               f"{', '.join(up.get('reasons') or [])}",
               bg="FFC7CE", fg="9C0006", size=9, wrap=True)
            ws.row_dimensions[row].height = 18
            row += 1
    # ── Missing-DXF / inferred-geometry section ────────────────────────────────
    _parts_all = (summary.get("manufacturing_writeup") or {}).get("parts") or []
    _inferred = [p for p in _parts_all if p.get("geometry_inferred") and not p.get("dxf_augmented")]
    _no_dxf   = [p for p in _parts_all
                 if (p.get("source") == "sdi_bom_row_no_geometry"
                     and not p.get("geometry_inferred")
                     and not _has_real_dxf_geometry(p))]
    if _inferred or _no_dxf:
        row += 2
        ws.merge_cells(f"A{row}:L{row}")
        _c(ws, row, 1, "⚠  DRAWINGS OUTSTANDING — PROVISIONAL / MISSING COSTS",
           bold=True, bg="FFE699", fg="7F6000", size=11)
        ws.row_dimensions[row].height = 20
        row += 1
        ws.merge_cells(f"A{row}:L{row}")
        _c(ws, row, 1, "These parts have no flat DXF. Request the DXF from the "
           "drawing office; figures below are AI-inferred and provisional.",
           bg="FFF2CC", fg="7F6000", size=9, wrap=True)
        ws.row_dimensions[row].height = 16
        row += 1
        for p in _inferred:
            gi = p.get("geometry_inference") or {}
            basis = gi.get("basis", "inferred")
            _bn = {"historical_sdilive": "from SDILive history",
                   "sibling_borrow": "borrowed from similar part",
                   "category_default": "typical size for type"}.get(basis, basis)
            ws.merge_cells(f"B{row}:L{row}")
            _c(ws, row, 1, "✎", bg="FFF2CC", align="center", size=9)
            _c(ws, row, 2, f"{p.get('part_number')}  —  {p.get('description')}  |  "
               f"INFERRED ({_bn}): {gi.get('blank_length_mm')}×{gi.get('blank_width_mm')}mm "
               f"— VERIFY before quoting", bg="FFF2CC", fg="7F6000", size=9, wrap=True)
            ws.row_dimensions[row].height = 18
            row += 1
        for p in _no_dxf:
            ws.merge_cells(f"B{row}:L{row}")
            _c(ws, row, 1, "✗", bg="FFC7CE", align="center", size=9)
            _c(ws, row, 2, f"{p.get('part_number')}  —  {p.get('description')}  |  "
               f"NO DXF + could not infer — PRICE MANUALLY (currently £0)",
               bg="FFC7CE", fg="9C0006", size=9, wrap=True)
            ws.row_dimensions[row].height = 18
            row += 1
    # ── Cost breakdown by material ─────────────────────────────────────────────
    row += 2
    ws.merge_cells(f"A{row}:L{row}")
    # Named base. These are engine per-part figures and the percentages are shares of their
    # own sum — not of the Sell Price in the total row above, which is a different number.
    # An unlabelled "%" next to an unlabelled "£" invited exactly that reading.
    # Named base AND named total. An unlabelled "%" beside an unlabelled "£" invited the
    # reading that these were shares of the Sell Price, which they never were.
    _mat_total = _totals.get("material_gbp")
    _c(ws, row, 1,
       "MATERIAL COST BREAKDOWN BY TYPE"
       + (f"  —  per-part material; the Estimate sheet calculated "
          f"£{float(_mat_total):,.2f} of material in total"
          if _mat_total is not None else
          "  —  per-part material only; labour is charged per department row"),
       bold=True, bg=C_BLUE, fg=C_WHITE, size=11)
    ws.row_dimensions[row].height = 20
    row += 1
    mat_totals: Dict[str, float] = {}
    for part in parts:
        # Group bought-ins together rather than under a mis-inferred "MILD_STEEL".
        if _is_bought_in(part):
            mat = "Bought-in"
        else:
            mat = str(part.get("normalized_material") or part.get("material") or "Unknown")
        # Same basis as the per-part columns. Built from the engine's labour-inclusive
        # totals this section reported GBP 44.75 of "material" on a job whose sheet
        # calculated GBP 0.13 of it — a breakdown by material type that was not a breakdown
        # of material.
        _e = part_material_cost(part)[1] if _canonical else float(
            _est_lookup.get(part.get("part_number"), {}).get("extended_total_cost_gbp") or 0)
        mat_totals[mat] = mat_totals.get(mat, 0) + _e
    # WHAT THE PARTS DO NOT ACCOUNT FOR.
    #
    # The sheet's material total also carries lines that belong to no single part — the
    # powder consumable, and the per-line scrap the workbook adds. On 2085 the part column
    # is about £0.09 against a workbook material total of £0.1334, so a breakdown of only
    # the parts is short and silently so. Named as its own row, from the difference, so the
    # section adds back to the authoritative figure instead of merely being near it.
    if _mat_total is not None:
        _residual = round(float(_mat_total) - sum(mat_totals.values()), 4)
        if abs(_residual) >= 0.005:
            mat_totals["Powder / scrap / other workbook material"] = _residual
    for mi, (mat, cost) in enumerate(sorted(
            mat_totals.items(), key=lambda x: x[1], reverse=True)):
        bg = C_ALT if mi % 2 == 0 else C_WHITE
        _base = sum(mat_totals.values())
        pct = (cost / _base * 100) if _base > 0 else 0
        ws.merge_cells(f"A{row}:H{row}")
        _c(ws, row, 1, mat, bg=bg, bold=True, size=10, border=True)
        _c(ws, row, 9, cost, bg=bg, align="right", bold=True,
           num_fmt="£#,##0.00", size=10, border=True)
        _c(ws, row, 10, "", bg=bg, border=True)
        _c(ws, row, 11, f"{pct:.1f}%", bg=bg, align="center",
           size=10, border=True)
        _c(ws, row, 12, "", bg=bg, border=True)
        ws.row_dimensions[row].height = 18
        row += 1
    # The section must be checkable against the sheet, not just internally consistent.
    if _mat_total is not None:
        ws.merge_cells(f"A{row}:H{row}")
        _c(ws, row, 1, "TOTAL MATERIAL (agrees with the Estimate sheet)",
           bold=True, bg=C_SECTION, size=10, border=True)
        _c(ws, row, 9, sum(mat_totals.values()), bg=C_SECTION, align="right", bold=True,
           num_fmt="£#,##0.00", size=10, border=True)
        _c(ws, row, 10, "", bg=C_SECTION, border=True)
        _c(ws, row, 11, "100.0%", bg=C_SECTION, align="center", size=10, border=True)
        _c(ws, row, 12, "", bg=C_SECTION, border=True)
        ws.row_dimensions[row].height = 18
        row += 1
    # ── Footer ─────────────────────────────────────────────────────────────────
    row += 2
    ws.merge_cells(f"A{row}:L{row}")
    _c(ws, row, 1,
       f"Generated by SDI Intelligence  |  wearesdi.com  |  "
       f"{datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
       f"Estimates based on SolidWorks drawings + SDILive knowledge base  |  "
       f"Subject to estimator review before quoting",
       size=8, italic=True, fg="888888", align="center")
    # ── DECISIONS THAT REQUIRED RESOLUTION ─────────────────────────────────────
    # Every other line on this tab describes a decision that made itself: one strongest
    # source, nothing at that rank contradicting it. These are the lines where the engine
    # had to choose between sources the waterfall calls equal, and they are where an
    # estimator's feedback is worth most. Weakest resolution first -- a contest settled by
    # the claim-id backstop had no distinguishing evidence at all.
    if _contested:
        # BELOW EVERYTHING ALREADY ON THE SHEET. A fixed offset from the
        # totals row lands inside the template's merged footer.
        _r = max(_resolution_row, ws.max_row + 2)
        _c(ws, _r, 1, "DECISIONS THAT REQUIRED RESOLUTION", bold=True,
           bg=C_NAVY, fg=C_WHITE, size=10)
        _r += 1
        for _ci, _h in enumerate(("Part", "Operation", "Settled as", "Decided by",
                                  "Rank", "Other source claimed", "Settled by"), 1):
            _c(ws, _r, _ci, _h, bold=True, bg=C_SECTION, size=9, border=True)
        _r += 1
        for _d in _contested:
            _c(ws, _r, 1, str(_d.get("target_id") or ""), size=9, border=True)
            _c(ws, _r, 2, str(_d.get("operation") or "").replace("_", " "),
               size=9, border=True)
            _c(ws, _r, 3, str(_d.get("status") or ""), size=9, border=True)
            _c(ws, _r, 4, str(_d.get("decided_by") or _d.get("source") or "unrecorded"),
               size=9, border=True)
            _c(ws, _r, 5, _d.get("source_rank") or "", align="center", size=9, border=True)
            _c(ws, _r, 6, ", ".join(str(x) for x in (_d.get("losing_statuses") or [])),
               size=9, border=True)
            _c(ws, _r, 7, str(_d.get("settled_by_key") or "rank"), size=9, border=True,
               wrap=True)
            _r += 1

    # ── WHAT THE PART IS, WHERE TWO SOURCES DISAGREED ──────────────────────────
    # The block above covers OPERATIONS, and those are not the expensive arguments. On 8352
    # it held two lines, both "powder coating: not_applicable", while the same run had a back
    # panel costed as PLYWOOD that inference read as TIMBER and three parts whose gauge two
    # sources put 4x apart. Material decides the rate and whether the part has a rate at all;
    # gauge decides the rate AND steps the cut time, so a part costed on the wrong one is
    # wrong twice. A decisions tab that omits them is not describing the decisions that matter.
    if _datum_contested:
        _r = ws.max_row + 2
        _c(ws, _r, 1, "WHAT THE PART IS — WHERE TWO SOURCES DISAGREED", bold=True,
           bg=C_NAVY, fg=C_WHITE, size=10)
        _r += 1
        _c(ws, _r, 1,
           "The higher-ranked source was used and the figure stands. Confirm these before "
           "quoting firm — material and gauge both move the money.",
           size=8, italic=True, fg="666666")
        _r += 1
        for _ci, _h in enumerate(("Part", "Datum", "Costed as", "Read from",
                                  "Other source said", "Which source", "Apart by"), 1):
            _c(ws, _r, _ci, _h, bold=True, bg=C_SECTION, size=9, border=True)
        _r += 1
        for _d in _datum_contested:
            _ratio = _d.get("ratio")
            _c(ws, _r, 1, str(_d.get("part_number") or ""), size=9, border=True)
            _c(ws, _r, 2, str(_d.get("datum") or ""), size=9, border=True)
            _c(ws, _r, 3, str(_d.get("costed_as") or ""), size=9, border=True)
            _c(ws, _r, 4, str(_d.get("costed_from") or ""), size=9, border=True, wrap=True)
            _c(ws, _r, 5, str(_d.get("other") or ""), size=9, border=True)
            _c(ws, _r, 6, str(_d.get("other_from") or ""), size=9, border=True, wrap=True)
            # Only the gauge contest measures a distance; a material argument is categorical.
            _c(ws, _r, 7, (f"{float(_ratio):g}x" if _ratio else "—"),
               align="center", size=9, border=True)
            _r += 1

    # ── THE DECISIONS NOBODY COULD MAKE, BECAUSE THE DRAWING NEVER ARRIVED ─────
    # Every other block on this tab describes a decision made from evidence. These parts had
    # none: no drawing, so nothing read them, nothing costed them and no decision about them
    # exists to report. Naming them keeps this tab an account of the whole BOM rather than of
    # the parts that happened to have drawings.
    try:
        from costed_facts import undrawn_bom_lines as _undrawn
        _missing = _undrawn(summary)
    except Exception:                                            # noqa: BLE001
        _missing = []
    if _missing:
        _r = ws.max_row + 2
        _c(ws, _r, 1, "DRAWINGS MISSING FROM THIS PACK — NOTHING DECIDED, NOTHING COSTED",
           bold=True, bg=C_LOW, fg=C_LOW_TXT, size=10)
        _r += 1
        _c(ws, _r, 1,
           "No detail drawing was supplied for these BOM lines, so the engine never read them. "
           "They carry no material, no route and no cost — the total is real for what was "
           "supplied and is not a price for the whole product.",
           size=8, italic=True, fg="666666")
        _r += 1
        for _ci, _h in enumerate(("Part", "Description", "Status"), 1):
            _c(ws, _r, _ci, _h, bold=True, bg=C_SECTION, size=9, border=True)
        _r += 1
        for _m in _missing:
            _c(ws, _r, 1, str(_m.get("part_number") or ""), size=9, border=True)
            _c(ws, _r, 2, str(_m.get("description") or ""), size=9, border=True, wrap=True)
            _c(ws, _r, 3, "no drawing supplied", size=9, border=True)
            _r += 1

    ws.freeze_panes = "A6"


if __name__ == "__main__":
    print("SDI Intelligence — Job Decision Report module ready.")
    print()
    print("Add to estimator.py before wb.save():")
    print("  from job_decision_report import add_decision_report_sheet")
    print("  add_decision_report_sheet(wb, summary, scan_meta)")
