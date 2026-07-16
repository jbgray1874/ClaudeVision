#!/usr/bin/env python3
r"""
_diff_1282_runs.py  —  READ-ONLY.

WHY: 1282 (regression anchor) was £273.55 at the old £4/kg powder rate. With the powder
fix it should be ~£278.76. It came out at £250.59 — about £28 BELOW expectation, after
the phantom-bought-in guards landed. An anchor that moves without explanation is worse
than no anchor, so we find the £28 before touching anything else.

Compares the last TWO 1282 workbooks cell-for-cell in the three blocks that matter:
    - Bill of Materials   (bought-in lines + prices)
    - Sheet Steel         (fabricated parts)
    - Labour              (operations)

Locates each block by its HEADER TEXT, not by hard-coded row numbers (the template has
been widened; hard-coded offsets are how the previous quick-diff came back empty).

Usage (from C:\ClaudeVision\src):
    C:\ClaudeVision\.venv\Scripts\python.exe _diff_1282_runs.py
"""
from __future__ import annotations
import glob, os, sys
from openpyxl import load_workbook

PAT = r"C:\ClaudeVision\output\estimates\1282*.xlsx"
SHEET = "Estimate"
MAXROW = 200
MAXCOL = 20


def cells(ws, r):
    return [ws.cell(r, c).value for c in range(1, MAXCOL + 1)]


def rowtext(ws, r):
    return " ".join(str(v) for v in cells(ws, r) if v not in (None, ""))


def find_row(ws, needle, start=1):
    n = needle.upper()
    for r in range(start, MAXROW):
        if n in rowtext(ws, r).upper():
            return r
    return None


def block(ws, header, stop_headers):
    """Rows between a header and the next section header."""
    h = find_row(ws, header)
    if h is None:
        return []
    stops = [find_row(ws, s, h + 1) for s in stop_headers]
    stops = [s for s in stops if s]
    end = min(stops) if stops else MAXROW
    out = []
    for r in range(h + 1, end):
        vals = [v for v in cells(ws, r) if v not in (None, "")]
        if not vals:
            continue
        # skip pure-zero template filler rows
        if all(str(v).strip() in ("0", "0.0", "4%", "£-", "2500", "1250", "50", "0.00")
               for v in vals):
            continue
        out.append((r, vals))
    return out


def summarise(path):
    wb = load_workbook(path, data_only=False)
    ws = wb[SHEET]
    print("\n" + "=" * 96)
    print(os.path.basename(path))
    print("  AF57 (powder £/kg):", ws["AF57"].value, "   D6 (qty):", ws["D6"].value)
    print("=" * 96)

    bom = block(ws, "Bill of Materials", ["Wire", "Sheet Steel"])
    print(f"\n-- BILL OF MATERIALS ({len(bom)} line(s)) --")
    for r, vals in bom:
        print(f"   r{r:<4}", " | ".join(str(v) for v in vals[:8]))

    steel = block(ws, "Sheet Steel", ["Other Sheet Material", "Total Material"])
    print(f"\n-- SHEET STEEL ({len(steel)} part(s)) --")
    for r, vals in steel:
        print(f"   r{r:<4}", " | ".join(str(v) for v in vals[:6]))

    lab = block(ws, "Operation", ["Total Labour"])
    print(f"\n-- LABOUR ({len(lab)} row(s)) --")
    for r, vals in lab:
        print(f"   r{r:<4}", " | ".join(str(v) for v in vals[:6]))

    return {
        "bom": {str(v[1][0]): v for v in bom},
        "steel": {str(v[1][0]): v for v in steel},
        "labour_n": len(lab),
        "bom_n": len(bom),
        "steel_n": len(steel),
    }


def main():
    fs = sorted(glob.glob(PAT), key=os.path.getmtime)
    if len(fs) < 2:
        sys.exit("need at least two 1282 workbooks")
    prev, curr = fs[-2], fs[-1]

    print("PREVIOUS:", os.path.basename(prev))
    print("CURRENT :", os.path.basename(curr))

    a = summarise(prev)
    b = summarise(curr)

    print("\n" + "=" * 96)
    print("DIFF")
    print("=" * 96)
    print(f"  BOM lines   : {a['bom_n']:>3}  ->  {b['bom_n']:>3}")
    print(f"  Steel parts : {a['steel_n']:>3}  ->  {b['steel_n']:>3}")
    print(f"  Labour rows : {a['labour_n']:>3}  ->  {b['labour_n']:>3}")

    lost = set(a["bom"]) - set(b["bom"])
    gained = set(b["bom"]) - set(a["bom"])
    if lost:
        print("\n  !! BOM LINES LOST:")
        for k in sorted(lost):
            print("     -", a["bom"][k][1][:6])
    if gained:
        print("\n  ++ BOM LINES GAINED:")
        for k in sorted(gained):
            print("     +", b["bom"][k][1][:6])

    print("\n  -- BOM price changes on lines present in BOTH --")
    for k in sorted(set(a["bom"]) & set(b["bom"])):
        va, vb = a["bom"][k][1], b["bom"][k][1]
        if va != vb:
            print(f"     ~ {k}")
            print(f"         was: {va[:8]}")
            print(f"         now: {vb[:8]}")

    lost_s = set(a["steel"]) - set(b["steel"])
    if lost_s:
        print("\n  !! STEEL PARTS LOST:")
        for k in sorted(lost_s):
            print("     -", a["steel"][k][1][:6])

    print("""
READ: a ~£28 drop with the SAME line count means PRICE drift (history lookup returning a
different historical row per run — the known non-determinism). A drop in LINE COUNT means
the guards removed something they shouldn't have — that is a real regression, revert.
""")


if __name__ == "__main__":
    main()
