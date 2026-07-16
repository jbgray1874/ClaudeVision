"""
PROOF-OF-CONCEPT — does NOT touch production. Tests the single riskiest assumption:
can openpyxl open the heavily-formula'd Blank WB, populate a few INPUT cells, set
auto-recalc, and save a copy WITHOUT corrupting the WB's formulas / structural sheets?

It fills ONLY:
  - D6 (order quantity)
  - C3 (customer), C5 (drawing/job no)
  - the BOM block (rows 11-15) with a few sample bought-in lines

Then saves to the output dir as a _PROOF file. You open it in EXCEL and check:
  1. Does M59 (Total Material Cost) show a real number, not 0?  -> calc-on-open works
  2. Did the BOM data land in the right cells (C=desc, K=qty, L=scrap)?
  3. Are the Labour and Material Price Break sheets still intact with formulas?

Requires openpyxl. If not installed:
  C:\ClaudeVision\.venv\Scripts\python.exe -m pip install openpyxl

Run:
  C:\ClaudeVision\.venv\Scripts\python.exe _wb_populate_proof.py
"""
import shutil, os, sys

TEMPLATE = r"K:\Estimating\Completed\AI Estimating\AISheets\Blank Estimate Sheet  WB 2026.xlsx"
OUTDIR   = r"C:\ClaudeVision\output\estimates"
OUTNAME  = "1282 - Milwaukee Wall Bay_PROOF.xlsx"

def main():
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed. Run:")
        print(r"  C:\ClaudeVision\.venv\Scripts\python.exe -m pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(TEMPLATE):
        print(f"TEMPLATE NOT FOUND:\n  {TEMPLATE}")
        print("Check the exact filename (note the double-space in 'Sheet  WB').")
        sys.exit(1)

    os.makedirs(OUTDIR, exist_ok=True)
    out_path = os.path.join(OUTDIR, OUTNAME)

    # Load the template. keep_vba False, data_only False (we want formulas preserved).
    print("Opening template (this is the risky step — does openpyxl preserve the formulas?)...")
    wb = openpyxl.load_workbook(TEMPLATE, data_only=False, keep_vba=False)

    print("Sheets found in template:", wb.sheetnames)
    if "Estimate" not in wb.sheetnames:
        print("!! No 'Estimate' sheet — aborting."); sys.exit(1)

    ws = wb["Estimate"]

    # --- populate a FEW input cells only (the proof) ---
    # Header
    ws["C3"] = "TTI / Milwaukee (PROOF)"
    ws["C5"] = "1282"
    ws["D6"] = 180                      # order quantity — drives $D$6 references everywhere

    # BOM block: rows 11-15. Columns from the dump:
    #   C=description, H=part code, I=supplier, J=price(LOOKUP-driven, leave blank),
    #   K=qty per unit, L=scrap %
    # We fill description / code / qty / scrap and LEAVE J (price) to the WB's LOOKUP.
    sample_bom = [
        # (row, description,                    code,        qty, scrap)
        (11, "POP RIVET 4.0 x 10mm",            "FIXING5",   2,   0.04),
        (12, "M8 FLANGED NUTSERT",              "FIXING236", 2,   0.04),
        (13, "M8x38mm DIA GLIDE",               "FIXING125", 4,   0.04),
        (14, "Milwaukee Base Shelf Vinyl",      "VINYL76",   1,   0.04),
        (15, "Junction box",                    "",          1,   0.04),
    ]
    for row, desc, code, qty, scrap in sample_bom:
        ws[f"C{row}"] = desc
        ws[f"H{row}"] = code
        ws[f"K{row}"] = qty
        ws[f"L{row}"] = scrap

    # --- force Excel to recalc on open so the cached 0s get refreshed ---
    # openpyxl exposes this via calc properties.
    try:
        wb.calculation.fullCalcOnLoad = True
        print("Set fullCalcOnLoad = True (Excel will recalc on open).")
    except Exception as e:
        # older openpyxl: set via the workbook's calcPr
        try:
            from openpyxl.workbook.properties import CalcProperties
            wb.calculation = CalcProperties(fullCalcOnLoad=True)
            print("Set fullCalcOnLoad via CalcProperties.")
        except Exception as e2:
            print(f"WARN: could not set fullCalcOnLoad ({e2}). You may need to press F9 in Excel.")

    print(f"Saving proof to:\n  {out_path}")
    wb.save(out_path)
    print("\nSAVED. Now OPEN THAT FILE IN EXCEL and check:")
    print("  1. Does M59 (Estimate!M59, 'Total Material Cost') show a real number, not 0?")
    print("  2. Did rows 11-15 populate in columns C/H/K/L correctly?")
    print("  3. Open the 'Labour' and 'Material Price Break' tabs — formulas intact?")
    print("  4. Did J11-J15 (price) auto-fill from the Material Price Break LOOKUP?")
    print("\nIf all yes -> openpyxl round-trips this template safely; we build the full version.")
    print("If M59 stays 0 or formulas are broken -> we pivot to COM automation instead.")

if __name__ == "__main__":
    main()
