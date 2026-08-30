# -*- coding: utf-8 -*-
"""READ-ONLY. Before widening the Sheet Steel block we must SEE the live template layout: which
rows hold block headers, data rows, totals, and which cells carry formulas that reference block
rows (the #REF! risk if we shift things). Dumps the Estimate sheet structure so we can plan the
widen: how many rows to add, which config first_row/last_row values shift, and which formulas
need re-checking.

Reads the BLANK template (what wb_populate writes into), NOT an output file.

Run from C:\\ClaudeVision\\src :
  C:\\ClaudeVision\\.venv\\Scripts\\python.exe _probe_template_layout.py
"""
from pathlib import Path
import openpyxl

# find the blank template wb_populate uses
CANDIDATES = [
    r"C:\ClaudeVision\input\spreadsheets\EmptyEstimating\Blank Estimate Sheet 2026.xlsx",
    r"C:\ClaudeVision\input\spreadsheets\Blank Estimate Sheet 2026.xlsx",
]
tpl = None
for c in CANDIDATES:
    if Path(c).exists():
        tpl = Path(c); break
if tpl is None:
    # search for it
    base = Path(r"C:\ClaudeVision\input")
    hits = list(base.rglob("*Blank*Estimat*.xlsx")) + list(base.rglob("*Estimate*2026*.xlsx"))
    print("Template not at expected path. Candidates found:")
    for h in hits: print("   ", h)
    if hits: tpl = hits[0]
if tpl is None:
    print("No template found — paste the path wb_populate loads."); raise SystemExit(1)

print(f"Template: {tpl}\n")
wb = openpyxl.load_workbook(tpl, data_only=False)
ws = wb["Estimate"] if "Estimate" in wb.sheetnames else wb.active
print(f"Sheet: {ws.title}   dims: {ws.dimensions}   max_row: {ws.max_row}\n")

# dump rows 1..140 col A-C text + any formulas, so we see block headers/totals/formula rows
print("=== rows 30..135: col A/B/C content + formula cells (F,G,H,I,J) ===")
for r in range(30, 136):
    a = ws.cell(r, 1).value
    label = ""
    for c in range(1, 6):
        v = ws.cell(r, c).value
        if isinstance(v, str) and v.strip():
            label = v.strip()[:40]; break
    # find any formula in this row (cols A..M)
    formulas = []
    for c in range(1, 14):
        v = ws.cell(r, c).value
        if isinstance(v, str) and v.startswith("="):
            formulas.append(f"{openpyxl.utils.get_column_letter(c)}{r}={v[:30]}")
    if label or formulas:
        line = f"  r{r:3}: {label:40}"
        if formulas:
            line += "  FORMULAS: " + " | ".join(formulas[:3])
        print(line)

print("\n=== key: locate Sheet Steel header, its data rows, Other Sheet header, Total Material row ===")
for r in range(30, 136):
    for c in range(1, 6):
        v = ws.cell(r, c).value
        if isinstance(v, str) and any(k in v.upper() for k in
            ("SHEET STEEL","OTHER SHEET","TOTAL MATERIAL","BILL OF MATERIAL","WIRE","LABOUR","TOTAL LABOUR")):
            print(f"   r{r} col{c}: {v.strip()[:50]}")
            break
