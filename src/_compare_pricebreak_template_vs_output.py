# -*- coding: utf-8 -*-
"""READ-ONLY. Answers: does the Material Price Break sheet have formulas in the BASE TEMPLATE,
in the ENGINE OUTPUT, in both, or neither? This decides whether adding price-break formulas
is a RESTORE (template lost them) or a NEW FEATURE (never had them).

Reads the 'Material Price Break' sheet in:
  1. the base template  (config: EmptyEstimating\\Blank Estimate Sheet 2026.xlsx)
  2. the latest 1298 engine output
and reports formula vs value cells in each, side by side. Touches nothing.

Run from C:\\ClaudeVision\\src :
  C:\\ClaudeVision\\.venv\\Scripts\\python.exe _compare_pricebreak_template_vs_output.py
"""
from pathlib import Path
import openpyxl, glob, os

# resolve template path from config
import importlib.util
def find_template():
    # try config
    try:
        cfgpath = Path(r"C:\ClaudeVision\src\config.py")
        txt = cfgpath.read_text(encoding="utf-8", errors="ignore")
        # look for SPREADSHEETS_DIR
        import re
        m = re.search(r'SPREADSHEETS_DIR\s*=\s*(.+)', txt)
        # fall back to common locations
    except Exception:
        pass
    # brute-force search for the file
    for base in [r"C:\ClaudeVision", r"C:\ClaudeVision\src", r"C:\ClaudeVision\data"]:
        for p in glob.glob(os.path.join(base, "**", "Blank Estimate Sheet 2026.xlsx"), recursive=True):
            return p
    return None

def dump_pricebreak(path, label):
    print(f"\n===== {label} =====\n  {path}")
    if not path or not os.path.exists(path):
        print("  (FILE NOT FOUND)")
        return
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:
        print("  load error:", e); return
    # find the price-break sheet
    target = None
    for n in wb.sheetnames:
        if "price break" in n.lower() or "material price" in n.lower():
            target = n; break
    if not target:
        print("  sheets:", wb.sheetnames)
        print("  (no 'Material Price Break' sheet)")
        return
    ws = wb[target]
    print(f"  sheet '{target}', dims {ws.dimensions}")
    formula_cells = 0; value_cells = 0
    samples = []
    for row in ws.iter_rows():
        for c in row:
            if c.value in (None, ""):
                continue
            if isinstance(c.value, str) and c.value.startswith("="):
                formula_cells += 1
                if len(samples) < 12:
                    samples.append(f"{c.coordinate}={c.value[:34]}")
            else:
                value_cells += 1
    print(f"  FORMULA cells: {formula_cells}   VALUE cells: {value_cells}")
    if samples:
        print("  sample formulas:")
        for s in samples:
            print("    ", s)
    else:
        print("  (no formulas — sheet is values/blank only)")

tmpl = find_template()
dump_pricebreak(tmpl, "BASE TEMPLATE (Blank Estimate Sheet 2026)")

est_dir = Path(r"C:\ClaudeVision\output\estimates")
outs = sorted(est_dir.glob("1298*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
dump_pricebreak(str(outs[0]) if outs else None, "ENGINE OUTPUT (latest 1298)")

print("\n=== VERDICT LOGIC ===")
print("  TEMPLATE has formulas + OUTPUT has formulas -> working as designed; ladder computes on open.")
print("  TEMPLATE has formulas + OUTPUT missing them -> engine STRIPS them when populating -> RESTORE (bug).")
print("  TEMPLATE missing + OUTPUT missing          -> never had them -> adding = NEW FEATURE (design needed).")
print("  Compare against Tim's MANUAL sheet too (does his have the formulas?) to know the real baseline.")
