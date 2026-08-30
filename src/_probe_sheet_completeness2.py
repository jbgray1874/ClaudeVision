# -*- coding: utf-8 -*-
"""READ-ONLY v2 — corrected columns (v1 mis-read them). From the real sheet dump the layout is:
  BOM block (rows 11-25): C=desc, then Part code / Supplier / Price / Qty / Scrap in the columns
     after a gap. We'll print ALL non-empty cells per row so we don't guess column indices.
  Steel (38-48): C=desc, F=len, G=wid, H=gauge (confirmed good in v1).
  Other Sheet / acrylic (51-58): find the RISER row, print all non-empty cells.
  Labour routes (63-135): A=operation, B=part-desc.
Prints each row's full non-empty cell map so columns can't be mis-guessed.

Run from C:\\ClaudeVision\\src :
  C:\\ClaudeVision\\.venv\\Scripts\\python.exe _probe_sheet_completeness2.py
"""
import glob, os
try:
    import openpyxl
except ImportError:
    print("run with venv python"); raise SystemExit(1)

est = sorted(glob.glob(r"C:\ClaudeVision\output\estimates\12532-03RecipeCard_*.xlsx"), key=os.path.getmtime)
path = est[-1]
print(f"=== Sheet: {os.path.basename(path)} ===\n")
wb = openpyxl.load_workbook(path, data_only=False)
ws = wb["Estimate"] if "Estimate" in wb.sheetnames else wb.active
from openpyxl.utils import get_column_letter

def rowmap(r, cmax=15):
    cells = []
    for c in range(1, cmax+1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip():
            cells.append(f"{get_column_letter(c)}={str(v)[:38]}")
    return cells

print("=== (A) BOM BLOCK rows 11-25 (all non-empty cells) ===")
for r in range(11, 26):
    cm = rowmap(r)
    if cm: print(f"  r{r}: " + " | ".join(cm))

print("\n=== OTHER SHEET / ACRYLIC rows 50-59 (all non-empty cells) ===")
for r in range(50, 60):
    cm = rowmap(r)
    if cm: print(f"  r{r}: " + " | ".join(cm))

print("\n=== (B) LABOUR ROUTES rows 63-135 (op + part) ===")
seen = {}
for r in range(63, 136):
    op = ws.cell(row=r, column=1).value
    pd = ws.cell(row=r, column=2).value
    if op and str(op).strip():
        pn = str(pd).split("—")[-1].strip() if pd and "—" in str(pd) else str(pd or "")
        seen.setdefault(pn, []).append(str(op).strip())
for pn, ops in seen.items():
    print(f"  {pn[:42]:42} -> {', '.join(ops)}")

print(f"\n  (total labour rows populated: {sum(len(v) for v in seen.values())}, parts: {len(seen)})")

print("\n=== DISPLAY BOARDS on the sheet (how many of 3 rendered?) ===")
found = 0
for r in range(11, 26):
    for c in range(1, 16):
        v = ws.cell(row=r, column=c).value
        if v and "DISPLAY BOARD" in str(v):
            print(f"  r{r}: {v}")
            found += 1
            break
print(f"  -> {found} of 3 display boards on the sheet")
