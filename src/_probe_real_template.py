# -*- coding: utf-8 -*-
"""READ-ONLY. The REAL template (from _labour_formulas.py / _merged_cell_diag.py) is on the share:
  \\\\sdi-dc01\\shareddata$\\Shared\\Estimating\\Completed\\AI Estimating\\AISheets\\Blank Estimate Sheet  WB 2026.xlsx
(note two spaces before WB). My earlier probe hit a stale LOCAL copy with no formulas — wrong file.
Also: _revert_bom_lastrow.py shows a PRIOR widen attempt was REVERTED due to 'failed MergedCell'.
So merged cells are the real hazard here.

Dump from the CORRECT template: block headers, formulas (full width), AND merged-cell ranges in
the steel/other/total region — because merged cells are what broke the last widen.

Run from C:\\ClaudeVision\\src :
  C:\\ClaudeVision\\.venv\\Scripts\\python.exe _probe_real_template.py
"""
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

TPL = r"\\sdi-dc01\shareddata$\Shared\Estimating\Completed\AI Estimating\AISheets\Blank Estimate Sheet  WB 2026.xlsx"
if not Path(TPL).exists():
    print(f"NOT FOUND at: {TPL}")
    print("Paste the exact template path wb_populate loads (check wb_populate.py for TEMPLATE=).")
    raise SystemExit(1)

wb = openpyxl.load_workbook(TPL, data_only=False)
ws = wb["Estimate"] if "Estimate" in wb.sheetnames else wb.active
print(f"Template: {TPL}")
print(f"Sheet: {ws.title}  dims: {ws.dimensions}  max_row: {ws.max_row}\n")

# block headers
print("=== block headers (rows 30-135) ===")
for r in range(30, 136):
    for c in range(1, 6):
        v = ws.cell(r, c).value
        if isinstance(v, str) and any(k in v.upper() for k in
            ("SHEET STEEL","OTHER SHEET","TOTAL MATERIAL","BILL OF MATERIAL","WIRE","LABOUR")):
            print(f"   r{r} col{c}: {v.strip()[:45]}")

# formulas in steel/other/total region (full width)
print("\n=== formulas in rows 36-65 (full width to col 30) ===")
for r in range(36, 66):
    for c in range(1, 31):
        v = ws.cell(r, c).value
        if isinstance(v, str) and v.startswith("="):
            print(f"   {get_column_letter(c)}{r} = {v[:60]}")

# MERGED CELLS in the region (the hazard)
print("\n=== MERGED CELL ranges intersecting rows 36-60 (the widen hazard) ===")
for mr in ws.merged_cells.ranges:
    if mr.min_row <= 60 and mr.max_row >= 36:
        print(f"   {mr}")

print("\nVERDICT: merged cells across the steel/other blocks are what broke the last widen. If the")
print("steel data rows are individually merged (e.g. desc spans cols), inserting rows needs the")
print("merges replicated too — that's what 'failed MergedCell' was. This tells us feasibility.")
