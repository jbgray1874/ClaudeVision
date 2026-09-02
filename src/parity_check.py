#!/usr/bin/env python3
r"""
parity_check.py — SDI estimating-engine parity harness
======================================================

Compare one engine run against the estimator's manual workbook for the same job, print a
report, and append one row to a ledger so patterns across the week surface.

    python parity_check.py <engine.json> <manual.xls> [--scan-json <job>.json]
                           [--ledger parity_ledger.csv] [--job NAME]

WHAT THE PREVIOUS VERSION MEASURED, AND WHY IT HAD TO BE REBUILT
----------------------------------------------------------------
It ran once, in July, and produced one ledger row. Read again in September against everything
the engine had learned since, it was measuring the wrong things in five ways at once:

  * IT READ A DIFFERENT RECORD OF MATERIAL COST THAN THE SHEET IT COMPARED TO. The engine side
    came from `estimate_summary.part_estimates` — blank-area figures computed BEFORE the
    workbook exists — while the manual side read a cell. On 12552 those two records of one
    fact are £49.76 and £136.32. The £86.56 between them would have been booked as a lane gap
    and sent somebody hunting a supplier line that does not exist.
  * IT ENUMERATED ONE BLOCK OF FIVE. Manual line items came from the Bill of Materials only,
    stopping at the first "Wire"/"Sheet Steel" header — so sheet steel, other sheet material,
    tube and every labour row were compared as grand totals with no line to point at. On
    12552 the Sheet Steel block WAS the whole material gap.
  * ITS PART-NUMBER RULE TRUNCATED AT THREE SEGMENTS. `\d{4,5}-\d{2}-[A-Z0-9.]+` turns
    12349-02-69-01A into 12349-02-69 — the identical defect fixed in the DXF filename parser,
    in a second regex that disagreed with the first. Every four-segment fabrication keyed to
    its parent, missed the engine index, and fell into "engine-missing" at full value.
  * IT ONLY LOOKED ONE WAY. The report was headed "engine-missing / under-captured", and a
    50% threshold declared any engine figure above half the manual one to be a match. Over-
    pricing — a bounding box costed as a blank, the single worst defect on 12349 — was
    invisible, and a 40% under-read contributed zero.
  * NEITHER SIDE WAS RECONCILED AGAINST ITSELF. The parsed lines were never summed and held
    against the sheet's own Total Material Cost, so a parse that dropped rows produced a
    confident lane analysis of its own omissions.

WHAT THIS ONE DOES
------------------
ONE BLOCK READER, BOTH SIDES. The manual workbook and the engine's are the same house
template, so the block definitions come from `wep_readback_from_xlsx` — the module that
already reads one of them — rather than a third private copy. A block is a block whichever
file it is in.

RECONCILE BEFORE COMPARING. Each side's lines are summed and held against that side's own
labelled totals, and the report says so before it says anything else. A parity built on an
unbalanced parse is measuring the parser.

SIGNED, AND BOTH WAYS. Over and under are the same question with a different sign. Every line
whose difference clears a floor is reported; nothing is absorbed by a threshold.

THE INPUTS, NOT JUST THE MONEY. Two wrong inputs that cancel read as perfect parity. Blank,
gauge, nest and quantity are compared per part, because that is where the fixable engineering
is — a £ difference tells you there is a problem and an input difference tells you what it is.

AND WHO DECIDED IT. With a scan JSON, every difference carries the reader that supplied the
engine's side, so after ten jobs the ledger answers the question it exists for: which lane of
readers is costing us money.
"""
from __future__ import annotations

import argparse
import csv
import datetime
import json
import os
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

sys.path.insert(0, str(Path(__file__).resolve().parent))

# ── ONE DEFINITION OF WHAT A BLOCK IS ────────────────────────────────────────────────
#
# Imported, not restated. wep_readback_from_xlsx reads the engine's copy of this template and
# its header maps are the tested description of it; a private second copy here is how the two
# come to disagree about which column holds a cost, silently, on the one tool whose job is to
# find disagreements.
try:
    from wep_readback_from_xlsx import (                    # type: ignore[attr-defined]
        _LABOUR_HEADER_KEYS as LABOUR_KEYS,
        _MATERIAL_HEADER_KEYS as BOM_KEYS,
        _SHEET_HEADER_KEYS as SHEET_KEYS,
        _TUBE_HEADER_KEYS as TUBE_KEYS,
    )
except Exception:                                           # pragma: no cover - import guard
    LABOUR_KEYS = BOM_KEYS = SHEET_KEYS = TUBE_KEYS = {}

# label that opens the block  ->  (block name, its header key map)
BLOCKS: List[Tuple[str, str, Dict[str, str]]] = [
    ("bill of materials", "bom", BOM_KEYS),
    ("wire", "tube", TUBE_KEYS),
    ("sheet steel", "steel", SHEET_KEYS),
    ("other sheet material", "other_sheet", SHEET_KEYS),
]
BLOCK_ORDER = ["bom", "tube", "steel", "other_sheet"]

TOTAL_LABELS = (("material", "total material cost"),
                ("labour", "total labour cost"),
                ("unit", "total unit cost"))

# A line whose difference is smaller than this is noise: rounding, a penny of scrap, a rate
# card a month apart. Reported in the block subtotal, not as a line for somebody to chase.
LINE_FLOOR_GBP = 0.50

# A BLOCK ENDS AT ITS TOTAL. Running each block to the next block's header let the last one
# swallow "Total Material Cost" and count it as a line — the reconciliation then read
# material lines of £577.85 against a stated £288.92, which is the stated figure plus itself.
# The labour table did the same to everything below it and came out four times over.
_TOTAL_ROW = re.compile(r"^\s*(sub\s*)?total\b", re.IGNORECASE)


def _is_total_row(text: Any) -> bool:
    return bool(_TOTAL_ROW.match(str(text or "")))


def _f(v: Any) -> Optional[float]:
    try:
        f = float(str(v).replace(",", "").replace("£", "").strip())
    except (TypeError, ValueError):
        return None
    return f if f == f else None


def _key(code: Any) -> str:
    """THE ONE PART-NUMBER RULE. Not a regex of this module's own — the engine stores parts
    under `normalize_part_code`, and a comparison that normalises differently from the thing
    it compares is measuring its own normaliser."""
    try:
        from part_identity import normalize_part_code
        return normalize_part_code(code).upper()
    except Exception:                                       # pragma: no cover - import guard
        return str(code or "").strip().upper()


def _lead_code(text: Any) -> str:
    """The part code out of a description cell. The fabricated blocks carry no code column;
    the engine writes '<part number>  <description>' and the estimator writes the same."""
    return _key(str(text or "").strip().split("  ")[0].split(" ")[0])


# ── reading a house workbook, whichever side wrote it ────────────────────────────────

def _open_rows(path: Path) -> List[List[Any]]:
    """The Estimate sheet as a plain grid, from .xls or .xlsx alike.

    The estimators' files are .xls and the engine writes .xlsx. Two readers, one grid, so
    everything downstream is written once.
    """
    name = str(path).lower()
    if name.endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["Estimate"] if "Estimate" in wb.sheetnames else wb[wb.sheetnames[0]]
        return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                for r in range(1, ws.max_row + 1)]
    import xlrd
    book = xlrd.open_workbook(path)
    try:
        sheet = next((book.sheet_by_name(n) for n in book.sheet_names()
                      if n.strip().upper() == "ESTIMATE"), book.sheet_by_index(0))
        return [[sheet.cell_value(r, c) for c in range(sheet.ncols)]
                for r in range(sheet.nrows)]
    finally:
        book.release_resources()


def _label(row: List[Any]) -> str:
    return " ".join(str(v or "") for v in row[:8]).strip().lower()


def _header_positions(row: List[Any], keys: Dict[str, str]) -> Dict[str, int]:
    """Which column holds which field, found by header TEXT.

    The template shifts columns as a block grows, and a hardcoded index goes quietly stale
    rather than failing — the same reasoning the read-back gives for doing it this way.
    """
    out: Dict[str, int] = {}
    for i, cell in enumerate(row):
        text = str(cell or "").strip().lower()
        if not text:
            continue
        for needle, field in keys.items():
            if field not in out and text.startswith(needle):
                out[field] = i
    return out


def read_house_workbook(path: Path) -> Dict[str, Any]:
    """Totals, every material block and every labour row, out of an SDI house workbook."""
    grid = _open_rows(Path(path))
    out: Dict[str, Any] = {
        "path": str(path), "headline": {}, "totals": {k: None for k, _ in TOTAL_LABELS},
        "blocks": {b: [] for b in BLOCK_ORDER}, "labour_rows": [], "problems": [],
    }

    # Where each block starts, and where the labour table starts.
    starts: List[Tuple[int, str, Dict[str, str]]] = []
    labour_start: Optional[int] = None
    for r, row in enumerate(grid):
        lab = _label(row)
        if not lab:
            continue
        for needle, name, keys in BLOCKS:
            if lab.startswith(needle) and not any(s[1] == name for s in starts):
                starts.append((r, name, keys))
        if labour_start is None and "operation" in lab and "dept" in lab:
            labour_start = r
        for field, needle in TOTAL_LABELS:
            if out["totals"][field] is None and lab.startswith(needle):
                for c in range(8, min(len(row), 30)):
                    v = _f(row[c])
                    if v is not None:
                        out["totals"][field] = round(v, 2)
                        break
        if lab.startswith("quantity"):
            out["headline"].setdefault("quantity", _f(row[3]))
        elif lab.startswith("drawing no"):
            out["headline"].setdefault("drawing", row[3] if len(row) > 3 else None)

    starts.sort()
    for i, (r, name, keys) in enumerate(starts):
        end = starts[i + 1][0] if i + 1 < len(starts) else (labour_start or len(grid))
        cols = _header_positions(grid[r], keys)
        if "total_value_gbp" not in cols:
            # The header row is usually the block label's own row on the BOM and the NEXT row
            # on the fabricated blocks. Try it before giving up, and say so if it fails.
            if r + 1 < len(grid):
                cols = _header_positions(grid[r + 1], keys)
        if "total_value_gbp" not in cols:
            out["problems"].append(f"{name}: could not find a value column from its header")
            continue
        for rr in range(r + 1, end):
            row = grid[rr]
            if _is_total_row(_label(row)):
                break
            desc = str(row[cols.get("description", 2)] or "").strip() if row else ""
            total = _f(row[cols["total_value_gbp"]]) if len(row) > cols["total_value_gbp"] else None
            if not desc and total is None:
                continue
            if desc.lower().startswith(tuple(n for n, _, _ in BLOCKS)) or \
                    desc.lower().startswith(("part description", "part code")):
                continue
            line = {"block": name, "description": desc,
                    "code": _key(row[cols["part_code"]]) if "part_code" in cols
                            and len(row) > cols["part_code"] else _lead_code(desc)}
            for field, c in cols.items():
                if field in ("description", "part_code") or len(row) <= c:
                    continue
                line[field] = _f(row[c]) if field != "supplier" else row[c]
            out["blocks"][name].append(line)
        # A COLUMN OF FORMULAS NOBODY CALCULATED IS NOT A COLUMN OF ZEROS. openpyxl returns
        # None for a formula with no cached result, and reporting a block of eleven rows at
        # £0.00 is a confident wrong answer about somebody's estimate. Say which file it is,
        # because the answer differs: an engine .xlsx wants its run's final_estimate, and a
        # manual .xls wants opening and saving in Excel once.
        _rows = out["blocks"][name]
        if _rows and all(_f(l.get("total_value_gbp")) is None for l in _rows):
            out["problems"].append(
                f"{name}: {len(_rows)} row(s) read, and every value cell is an uncalculated "
                f"formula — this workbook has never been recalculated and saved, so its "
                f"figures cannot be compared with anything")

    if labour_start is not None:
        cols = _header_positions(grid[labour_start], LABOUR_KEYS)
        if "total_value_gbp" in cols:
            for rr in range(labour_start + 1, len(grid)):
                row = grid[rr]
                if len(row) <= cols["total_value_gbp"]:
                    continue
                if _is_total_row(_label(row)):
                    break
                total = _f(row[cols["total_value_gbp"]])
                op = str(row[cols.get("operation", 2)] or "").strip()
                if total is None and not op:
                    continue
                if _is_total_row(op) or op.lower().startswith("operation"):
                    continue
                line = {"operation": op}
                for field, c in cols.items():
                    if len(row) > c:
                        line[field] = row[c] if field in ("operation", "description",
                                                          "department") else _f(row[c])
                out["labour_rows"].append(line)
        else:
            out["problems"].append("labour: could not find a value column from its header")
    return out


# ── the engine's side ────────────────────────────────────────────────────────────────

def read_engine_json(path: Path) -> Dict[str, Any]:
    """The engine's rows, preferring the record that provably sums to its own sheet.

    `final_estimate` is written by the read-back AFTER Excel recalculates the populated
    template: every block and every labour row, carrying the value the sheet itself computed.
    `part_estimates` is what the sheet was BUILT from, which is a different number — so a run
    without a read-back is compared on the older basis and the report SAYS so, rather than
    quietly comparing two unlike things.
    """
    with open(path, encoding="utf-8") as fh:
        doc = json.load(fh)
    fe = doc.get("final_estimate")
    if not isinstance(fe, dict):
        fe = ((doc.get("estimate_summary") or {}).get("final_estimate")
              if isinstance(doc.get("estimate_summary"), dict) else None)
    out: Dict[str, Any] = {"path": str(path), "basis": None, "problems": [],
                           "blocks": {b: [] for b in BLOCK_ORDER}, "labour_rows": [],
                           "totals": {}}
    if isinstance(fe, dict) and (fe.get("material_rows") or fe.get("labour_rows")):
        out["basis"] = "final_estimate — the rows as Excel calculated them"
        totals = fe.get("totals") or {}
        out["totals"] = {"material": _f(totals.get("material_gbp")),
                         "labour": _f(totals.get("labour_gbp")),
                         "unit": _f(totals.get("unit_gbp"))}
        for row in fe.get("material_rows") or []:
            if not isinstance(row, dict):
                continue
            block = str(row.get("block") or "bom")
            line = dict(row)
            line["code"] = _key(row.get("part_code")) or _lead_code(row.get("description"))
            out["blocks"].setdefault(block, []).append(line)
        out["labour_rows"] = [r for r in (fe.get("labour_rows") or []) if isinstance(r, dict)]
        for p in fe.get("adapter_problems") or []:
            out["problems"].append(f"read-back: {(p or {}).get('message')}")
        return out

    es = doc.get("estimate_summary") or {}
    out["basis"] = ("estimate_summary.part_estimates — the engine's own arithmetic, NOT the "
                    "workbook's. This run has no read-back")
    out["problems"].append(
        "no final_estimate on this run: material is compared on the engine's blank-area "
        "basis against a manual sheet's nest-derived cells, which are different records of "
        "the same fact. Re-run so the read-back stamps final_estimate.")
    for pe in es.get("part_estimates") or []:
        me = pe.get("material_estimate") or {}
        out["blocks"]["bom"].append({
            "block": "bom", "code": _key(pe.get("part_number")),
            "description": pe.get("description"),
            "qty_per_unit": _f(pe.get("quantity")),
            "total_value_gbp": _f(me.get("extended_material_cost_gbp")),
        })
    out["totals"] = {"material": _f(es.get("total_material_cost_gbp")),
                     "labour": _f(es.get("total_labour_cost_gbp")),
                     "unit": _f(es.get("document_total_estimated_cost_gbp"))}
    return out


# ── does each side add up to itself ──────────────────────────────────────────────────

def reconcile(side: Dict[str, Any]) -> Dict[str, Any]:
    """Do this side's own lines sum to this side's own labelled totals?

    Asked of BOTH sides and printed BEFORE any comparison. A parity built on a parse that
    dropped rows is an analysis of its own omissions, and it reads exactly like a finding.
    """
    material = round(sum(_f(l.get("total_value_gbp")) or 0.0
                         for b in BLOCK_ORDER for l in side["blocks"].get(b, [])), 2)
    labour = round(sum(_f(l.get("total_value_gbp")) or 0.0
                       for l in side.get("labour_rows") or []), 2)
    stated_m, stated_l = side["totals"].get("material"), side["totals"].get("labour")
    return {
        "material_lines": material, "material_stated": stated_m,
        "material_gap": None if stated_m is None else round(stated_m - material, 2),
        "labour_lines": labour, "labour_stated": stated_l,
        "labour_gap": None if stated_l is None else round(stated_l - labour, 2),
        "balances": (stated_m is not None and abs(stated_m - material) < 0.01
                     and stated_l is not None and abs(stated_l - labour) < 0.01),
    }


# ── the comparison ───────────────────────────────────────────────────────────────────

def _index(lines: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for line in lines:
        code = line.get("code") or _lead_code(line.get("description"))
        if not code:
            continue
        if code in out:      # two rows for one part on one side: sum them, note it
            out[code]["total_value_gbp"] = ((_f(out[code].get("total_value_gbp")) or 0)
                                            + (_f(line.get("total_value_gbp")) or 0))
            out[code].setdefault("_rows", 1)
            out[code]["_rows"] += 1
            continue
        out[code] = dict(line)
    return out


# The inputs a £ difference is usually MADE of. Compared per part, because a money delta says
# there is a problem and an input delta says what it is.
INPUT_FIELDS = [("length_mm", "blank L", 1.0), ("width_mm", "blank W", 1.0),
                ("gauge", "gauge", 0.01), ("thickness_mm", "thickness", 0.01),
                ("qty_per_unit", "qty", 0.01), ("qty_per_sheet", "nest", 0.01)]


def compare_blocks(manual: Dict[str, Any], engine: Dict[str, Any],
                   floor: float = LINE_FLOOR_GBP) -> Dict[str, Any]:
    """Every block, line by line, signed both ways, with the inputs behind each difference."""
    blocks: Dict[str, Any] = {}
    for name in BLOCK_ORDER:
        m_idx, e_idx = _index(manual["blocks"].get(name, [])), _index(engine["blocks"].get(name, []))
        rows: List[Dict[str, Any]] = []
        for code in sorted(set(m_idx) | set(e_idx)):
            m, e = m_idx.get(code), e_idx.get(code)
            m_gbp = _f((m or {}).get("total_value_gbp")) or 0.0
            e_gbp = _f((e or {}).get("total_value_gbp")) or 0.0
            delta = round(e_gbp - m_gbp, 2)
            inputs = []
            if m and e:
                for field, label, tol in INPUT_FIELDS:
                    mv, ev = _f(m.get(field)), _f(e.get(field))
                    if mv is None or ev is None or abs(mv - ev) <= tol:
                        continue
                    inputs.append({"field": label, "manual": mv, "engine": ev})
            if abs(delta) < floor and not inputs:
                continue
            rows.append({
                "code": code, "manual_gbp": round(m_gbp, 2), "engine_gbp": round(e_gbp, 2),
                "delta_gbp": delta, "inputs": inputs,
                "only_on": None if (m and e) else ("manual" if m else "engine"),
                "description": str((e or m or {}).get("description") or "")[:48],
            })
        blocks[name] = {
            "manual_total": round(sum(_f(l.get("total_value_gbp")) or 0
                                      for l in manual["blocks"].get(name, [])), 2),
            "engine_total": round(sum(_f(l.get("total_value_gbp")) or 0
                                      for l in engine["blocks"].get(name, [])), 2),
            "rows": sorted(rows, key=lambda r: -abs(r["delta_gbp"])),
        }
        blocks[name]["delta"] = round(blocks[name]["engine_total"]
                                      - blocks[name]["manual_total"], 2)
    return blocks


def compare_labour(manual: Dict[str, Any], engine: Dict[str, Any]) -> Dict[str, Any]:
    """By department, and split into set-up and run.

    One total against one total says nothing about a quantity: set-up does not move with the
    order and run time does, so two jobs whose labour agrees at 100 off can be far apart at
    10. The sheet records set-up in MINUTES against a department rate; both sides are
    converted the same way or neither is.
    """
    def _by_dept(side: Dict[str, Any]) -> Dict[str, Dict[str, float]]:
        out: Dict[str, Dict[str, float]] = {}
        qty = max(1, int(_f((side.get("headline") or {}).get("quantity")) or 1))
        for row in side.get("labour_rows") or []:
            dept = str(row.get("department") or "?").strip().upper() or "?"
            total = _f(row.get("total_value_gbp")) or 0.0
            mins, rate = _f(row.get("setup_minutes")), _f(row.get("dept_rate_gbp_per_hour"))
            setup = (rate / 60.0) * mins / qty if (mins is not None and rate is not None) else 0.0
            node = out.setdefault(dept, {"total": 0.0, "setup": 0.0})
            node["total"] += total
            node["setup"] += setup
        return {d: {"total": round(v["total"], 2), "setup": round(v["setup"], 2),
                    "run": round(v["total"] - v["setup"], 2)} for d, v in out.items()}

    m, e = _by_dept(manual), _by_dept(engine)
    rows = []
    for dept in sorted(set(m) | set(e)):
        mv, ev = m.get(dept, {}), e.get(dept, {})
        rows.append({
            "department": dept,
            "manual": mv.get("total", 0.0), "engine": ev.get("total", 0.0),
            "delta": round(ev.get("total", 0.0) - mv.get("total", 0.0), 2),
            "manual_setup": mv.get("setup", 0.0), "engine_setup": ev.get("setup", 0.0),
            "manual_run": mv.get("run", 0.0), "engine_run": ev.get("run", 0.0),
        })
    return {"rows": sorted(rows, key=lambda r: -abs(r["delta"])),
            "manual_setup": round(sum(v["setup"] for v in m.values()), 2),
            "engine_setup": round(sum(v["setup"] for v in e.values()), 2),
            "manual_run": round(sum(v["run"] for v in m.values()), 2),
            "engine_run": round(sum(v["run"] for v in e.values()), 2)}


def readers_for(scan_json: Optional[Path]) -> Dict[str, Dict[str, Any]]:
    """part code -> which reader supplied its geometry, material and gauge."""
    if not scan_json:
        return {}
    try:
        with open(scan_json, encoding="utf-8") as fh:
            doc = json.load(fh)
    except Exception:                                        # noqa: BLE001
        return {}
    parts = ((doc.get("manufacturing_writeup") or {}).get("parts")
             or doc.get("parts") or []) if isinstance(doc, dict) else doc
    out: Dict[str, Dict[str, Any]] = {}
    for p in parts or []:
        if isinstance(p, dict) and p.get("part_number"):
            out[_key(p["part_number"])] = {
                "geometry": p.get("geometry_source"), "material": p.get("material_source"),
                "thickness": p.get("thickness_source"),
            }
    return out


# ── the report ───────────────────────────────────────────────────────────────────────

def _money(v: Any) -> str:
    f = _f(v)
    return "—" if f is None else f"£{f:,.2f}"


def _signed(v: Any) -> str:
    f = _f(v)
    return "—" if f is None else f"{'+' if f >= 0 else '−'}£{abs(f):,.2f}"


def build_report(manual: Dict[str, Any], engine: Dict[str, Any], job: str,
                 scan_json: Optional[Path] = None) -> Tuple[str, Dict[str, Any]]:
    L: List[str] = [f"PARITY — {job}", "=" * 72, ""]
    readers = readers_for(scan_json)

    L.append(f"  Engine basis: {engine.get('basis')}")
    for p in (manual.get("problems") or []) + (engine.get("problems") or []):
        L.append(f"  ! {p}")
    L.append("")

    # ── 1. does each side add up to itself ──────────────────────────────────────────
    L.append("  DOES EACH SIDE ADD UP TO ITSELF")
    for name, side in (("manual", manual), ("engine", engine)):
        r = reconcile(side)
        for what in ("material", "labour"):
            lines, stated = r[f"{what}_lines"], r[f"{what}_stated"]
            gap = r[f"{what}_gap"]
            L.append(f"      {name:6} {what:8} lines {_money(lines):>11}   "
                     f"sheet {_money(stated):>11}   "
                     + ("agree" if gap is not None and abs(gap) < 0.01
                        else f"OFF BY {_money(gap)}" if gap is not None
                        else "no total on the sheet"))
    L.append("")
    if not (reconcile(manual)["balances"] and reconcile(engine)["balances"]):
        L.append("  ! One side does not sum to its own totals. Every difference below is "
                 "suspect until that is settled — a parse that dropped rows produces an "
                 "analysis of its own omissions, and it reads exactly like a finding.")
        L.append("")

    # ── 2. block by block, then line by line ────────────────────────────────────────
    blocks = compare_blocks(manual, engine)
    L.append("  MATERIAL, BLOCK BY BLOCK        manual      engine       delta")
    for name in BLOCK_ORDER:
        b = blocks[name]
        if not (b["manual_total"] or b["engine_total"]):
            continue
        L.append(f"      {name:14} {_money(b['manual_total']):>12} "
                 f"{_money(b['engine_total']):>12} {_signed(b['delta']):>12}")
    L.append("")
    for name in BLOCK_ORDER:
        rows = blocks[name]["rows"]
        if not rows:
            continue
        L.append(f"  {name.upper()} — every line that differs by £{LINE_FLOOR_GBP:.2f} or "
                 f"more, or whose inputs disagree")
        for r in rows:
            where = ("engine only" if r["only_on"] == "engine" else
                     "manual only" if r["only_on"] == "manual" else "")
            L.append(f"      {r['code'] or r['description']:<24} "
                     f"manual {_money(r['manual_gbp']):>10}  engine {_money(r['engine_gbp']):>10}"
                     f"  {_signed(r['delta_gbp']):>11}  {where}")
            for i in r["inputs"]:
                L.append(f"          {i['field']}: manual {i['manual']:g} vs engine "
                         f"{i['engine']:g}")
            rd = readers.get(r["code"])
            if rd and (r["inputs"] or abs(r["delta_gbp"]) >= LINE_FLOOR_GBP):
                L.append(f"          engine read it from: geometry {rd.get('geometry') or '—'},"
                         f" material {rd.get('material') or '—'},"
                         f" gauge {rd.get('thickness') or '—'}")
        L.append("")

    # ── 3. labour, by department and by set-up vs run ───────────────────────────────
    lab = compare_labour(manual, engine)
    L.append("  LABOUR, BY DEPARTMENT           manual      engine       delta")
    for r in lab["rows"]:
        L.append(f"      {r['department']:14} {_money(r['manual']):>12} "
                 f"{_money(r['engine']):>12} {_signed(r['delta']):>12}")
    L.append(f"      {'set-up':14} {_money(lab['manual_setup']):>12} "
             f"{_money(lab['engine_setup']):>12} "
             f"{_signed(round(lab['engine_setup'] - lab['manual_setup'], 2)):>12}")
    L.append(f"      {'run':14} {_money(lab['manual_run']):>12} "
             f"{_money(lab['engine_run']):>12} "
             f"{_signed(round(lab['engine_run'] - lab['manual_run'], 2)):>12}")
    L.append("")

    # ── 4. the headline ─────────────────────────────────────────────────────────────
    mt, et = manual["totals"], engine["totals"]
    for what in ("material", "labour", "unit"):
        m, e = _f(mt.get(what)), _f(et.get(what))
        if m is None and e is None:
            continue
        L.append(f"  {what.upper():9} manual {_money(m):>11}   engine {_money(e):>11}   "
                 + (_signed(round((e or 0) - (m or 0), 2)) if m is not None and e is not None
                    else "one side did not state it"))
    return "\n".join(L), {"blocks": blocks, "labour": lab,
                          "manual_reconciles": reconcile(manual)["balances"],
                          "engine_reconciles": reconcile(engine)["balances"]}


# ── the ledger ───────────────────────────────────────────────────────────────────────

def ledger_row(job: str, manual: Dict[str, Any], engine: Dict[str, Any],
               detail: Dict[str, Any], run_date: str = "") -> Dict[str, Any]:
    """One row per JOB, not per run of this tool.

    The old row stamped `datetime.date.today()` — the date somebody happened to run parity,
    not the date of the estimate — and appended unconditionally, so re-running a parity
    double-counted it in any weekly aggregate. The key is the job.
    """
    row = {
        "job": job,
        "job_date": run_date or datetime.date.fromtimestamp(
            os.path.getmtime(engine["path"])).isoformat(),
        "quantity": (manual.get("headline") or {}).get("quantity"),
        "drawing": (manual.get("headline") or {}).get("drawing"),
        "engine_basis": "final_estimate" if "final_estimate" in str(engine.get("basis"))
                        else "part_estimates",
        "manual_reconciles": detail["manual_reconciles"],
        "engine_reconciles": detail["engine_reconciles"],
    }
    for what in ("material", "labour", "unit"):
        row[f"manual_{what}"] = manual["totals"].get(what)
        row[f"engine_{what}"] = engine["totals"].get(what)
    for name in BLOCK_ORDER:
        row[f"delta_{name}"] = detail["blocks"][name]["delta"]
    row["delta_setup"] = round(detail["labour"]["engine_setup"]
                               - detail["labour"]["manual_setup"], 2)
    row["delta_run"] = round(detail["labour"]["engine_run"]
                             - detail["labour"]["manual_run"], 2)
    row["lines_differing"] = sum(len(detail["blocks"][n]["rows"]) for n in BLOCK_ORDER)
    row["lines_with_input_disagreement"] = sum(
        1 for n in BLOCK_ORDER for r in detail["blocks"][n]["rows"] if r["inputs"])
    return row


def append_ledger(path: str, row: Dict[str, Any]) -> str:
    """Write this job's row, REPLACING any earlier row for the same job.

    Appending unconditionally meant a job parity'd twice appeared twice, and every aggregate
    over the ledger counted it twice — in the file that exists to be aggregated.
    """
    rows: List[Dict[str, Any]] = []
    if os.path.exists(path):
        with open(path, newline="", encoding="utf-8") as fh:
            rows = [r for r in csv.DictReader(fh) if r.get("job") != row["job"]]
    fields = list(row.keys())
    for existing in rows:
        for k in existing:
            if k not in fields:
                fields.append(k)
    rows.append({k: row.get(k) for k in fields})
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k) for k in fields})
    return "replaced" if len(rows) > 1 or os.path.exists(path) else "created"


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("engine_json")
    ap.add_argument("manual_xls")
    ap.add_argument("--scan-json", type=Path, default=None,
                    help="the run's output/json/<job>.json — names the reader behind each "
                         "engine figure, so the ledger can say which lane costs money")
    ap.add_argument("--ledger", default="parity_ledger.csv")
    ap.add_argument("--job", default=None)
    a = ap.parse_args()

    job = a.job or os.path.splitext(os.path.basename(a.engine_json))[0]
    manual = read_house_workbook(Path(a.manual_xls))
    engine = read_engine_json(Path(a.engine_json))
    report, detail = build_report(manual, engine, job, a.scan_json)
    print(report)
    what = append_ledger(a.ledger, ledger_row(job, manual, engine, detail))
    print(f"\n  -> ledger row {what} in {a.ledger}")


if __name__ == "__main__":
    main()
