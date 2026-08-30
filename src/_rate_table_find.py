#!/usr/bin/env python3
r"""
_rate_table_find.py  -- read-only. Finds the rate table instead of guessing where it is.

My last attempt looked at rows 108-155 and found nothing. That range came from an OLD
historical sheet's LOOKUP formula ($H$118:$H$149). The BOM was widened by 25 rows on
2026-07-13 and everything below it shifted. The range is stale.

So: do not guess. The template tells us exactly where its own rate table is, in the
LOOKUP formula sitting in the labour block's K column:

    K96: =IF(H96=0,"0",(LOOKUP(C96, Estimate!$H$nnn:$H$mmm, Estimate!$I$nnn:$I$mmm)))
                                              ^^^^^^^^^^^^ this is the answer

Read the formula, extract the range, dump the range. Then belt-and-braces: scan the
WHOLE sheet for any cell containing P.Coat / POWDER, so we cannot miss it even if the
formula is shaped differently than expected.

THE QUESTION: Tim's 1310 books P.Coat at 331.42, dept POWDER. Ours books 355.43, dept P/C.
    Two powder rows in the table -> the LOOKUP is matching the wrong one. ENGINE BUG.
    One row at 355.43           -> Tim is on a different template version. SDI's call, not ours.
"""
from __future__ import annotations
from openpyxl import load_workbook
import re, sys

TEMPLATE = (r"\\sdi-dc01\shareddata$\Shared\Estimating\Completed"
            r"\AI Estimating\AISheets\Blank Estimate Sheet  WB 2026.xlsx")


def main():
    print(f"\n  {TEMPLATE}\n")
    try:
        ws = load_workbook(TEMPLATE, data_only=False)["Estimate"]
    except Exception as e:
        sys.exit(f"could not open: {e}")

    # 1. Read the LOOKUP formula out of the labour block. CELL_MAP says labour starts at 96.
    print("  LOOKUP FORMULAS IN THE LABOUR BLOCK (row 96)")
    print("  " + "-" * 68)
    lo = hi = None
    for col, name in ((7, "G dept"), (11, "K rate"), (12, "L setup")):
        f = ws.cell(96, col).value
        print(f"  {name:<10} {f}")
        if isinstance(f, str) and "LOOKUP" in f.upper():
            m = re.search(r"\$H\$(\d+)\s*:\s*\$H\$(\d+)", f)
            if m:
                lo, hi = int(m.group(1)), int(m.group(2))
    # also try the neighbouring columns in case the map is off by one
    if lo is None:
        for col in range(3, 16):
            f = ws.cell(96, col).value
            if isinstance(f, str) and "LOOKUP" in f.upper():
                m = re.search(r"\$H\$(\d+)\s*:\s*\$H\$(\d+)", f)
                if m:
                    lo, hi = int(m.group(1)), int(m.group(2))
                    print(f"  found LOOKUP in column {col}: {f}")
                    break

    if lo:
        print(f"\n  ==> THE RATE TABLE IS AT ROWS {lo}-{hi}\n")
        rng = range(lo - 2, hi + 3)
    else:
        print("\n  no LOOKUP formula found at row 96 - falling back to a full-sheet scan\n")
        rng = range(1, ws.max_row + 1)

    # 2. Dump it.
    print(f"  {'row':>4}  {'H  operation':<32} {'I  rate':>10}  {'J  dept':<10} {'K  setup'}")
    print(f"  {'-'*4}  {'-'*32} {'-'*10}  {'-'*10} {'-'*8}")
    powder = []
    for r in rng:
        op = ws.cell(r, 8).value
        if op in (None, ""):
            continue
        rate, dept, setup = ws.cell(r, 9).value, ws.cell(r, 10).value, ws.cell(r, 11).value
        hit = (any(t in str(op).upper() for t in ("COAT", "POWDER"))
               or "P/C" in str(dept or "").upper()
               or "POWDER" in str(dept or "").upper())
        if hit:
            powder.append((r, op, rate, dept))
        print(f"  {r:>4}  {str(op):<32} {str(rate):>10}  {str(dept or ''):<10} "
              f"{setup}{'   <== POWDER' if hit else ''}")

    # 3. Belt and braces - scan the whole sheet for the string, wherever it lives.
    print("\n  FULL-SHEET SCAN for 'P.Coat' / 'POWDER' (so we cannot miss it):")
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str) and ("P.COAT" in v.upper() or "POWDER" in v.upper()):
                print(f"      {c.coordinate:<8} {v!r}")

    print("\n" + "=" * 72)
    if len(powder) > 1:
        print("  TWO OR MORE POWDER ROWS -> the LOOKUP may well be picking the wrong one.")
        print("  That is an ENGINE-SIDE fix and we do it today.")
        for r, op, rate, dept in powder:
            print(f"        row {r}: {op!r}  rate={rate}  dept={dept}")
    elif len(powder) == 1:
        r, op, rate, dept = powder[0]
        print(f"  ONE powder row: row {r}  {op!r}  rate={rate}  dept={dept}")
        print("  If that is 355.43 / P/C, the engine is reading the LIVE template correctly")
        print("  and Tim's 331.42 / POWDER is a different template VERSION. The rate is then")
        print("  SDI's to set - take it to Tim and Tony. Do not change the rate card from here.")
    else:
        print("  STILL no powder row. Read the full-sheet scan above for where it actually is.")
    print("=" * 72 + "\n")


if __name__ == "__main__":
    main()
