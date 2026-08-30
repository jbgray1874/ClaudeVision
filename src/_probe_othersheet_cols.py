# -*- coding: utf-8 -*-
"""READ-ONLY. Dump the 'other_sheet' block column-map definition (wb_populate.py ~80-100) so we
know the exact column key for the 'Cost per sheet' cell (col L on the sheet) to write the price into.
Also dump the sheet's row 50 headers to confirm which column index 'Cost per sheet' is.
Run from C:\\ClaudeVision\\src :
  C:\\ClaudeVision\\.venv\\Scripts\\python.exe _probe_othersheet_cols.py
"""
from pathlib import Path
WB = Path(r"C:\ClaudeVision\src\wb_populate.py")
lines = WB.read_text(encoding="utf-8", errors="replace").splitlines()
print("=== other_sheet block column map (wb_populate.py ~78-110) ===")
for i in range(77, min(112, len(lines))):
    print(f"  {i+1}: {lines[i]}")

# Also confirm from a populated sheet which column header is 'Cost per sheet'
print("\n=== Confirm 'Cost per sheet' column index from the populated sheet row 50 ===")
import glob, os
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    est = sorted(glob.glob(r"C:\ClaudeVision\output\estimates\12532-03RecipeCard_*.xlsx"), key=os.path.getmtime)
    ws = openpyxl.load_workbook(est[-1], data_only=False)["Estimate"]
    for c in range(1, 16):
        v = ws.cell(row=50, column=c).value
        if v and str(v).strip():
            print(f"  col {get_column_letter(c)} ({c}): {v}")
except Exception as e:
    print(f"  (sheet read skipped: {e})")
