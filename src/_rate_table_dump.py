#!/usr/bin/env python3
r"""
_rate_table_dump.py  -- read-only.

WHERE DOES 355.43 COME FROM?

Not from us. wb_populate writes only FOUR cells per labour row:
    C = operation   D = description   H = qty   I = throughput (pieces/hr)

Everything else is an Excel LOOKUP inside the estimators' OWN template. From a real
historical sheet's raw JSON:

    K (the £/hr): =IF(H96=0,"0",(LOOKUP(C96,Estimate!$H$118:$H$149,Estimate!$I$118:$I$149)))
    L (setup):    =IF(C96=0,"0",(LOOKUP(C96,Estimate!$H$118:$H$149,Estimate!$K$118:$K$149)))
    G (dept):     =IF(C96=0,"", (LOOKUP(C96,Estimate!$H$118:$H$149,Estimate!$J$118:$J$149)))

So the rate table lives in the template at roughly H..K, rows 115-149:
    H = operation name    I = £/hr    J = dept code    K = setup mins

THE QUESTION THIS ANSWERS

Tim's 1310 sheet books P.Coat at 331.42, dept POWDER.
Our populated sheet books P.Coat at 355.43, dept P/C.

Same operation name. Different rate. Different dept code. So either:

  (a) there are TWO powder rows in this table and the LOOKUP is matching the wrong
      one -> that is an ENGINE BUG and we fix it today; or
  (b) there is ONE row at 355.43 and Tim's sheet is a different template version
      -> then the rate is SDI's to set, not ours, and it goes back to Tim and Tony.

Do not change SDI's rate card off a single sheet. The corpus has already shown
355.43 (2,141 lines) AND 304.13 (419 lines) coexisting, so powder rates have moved
over time and 331.42 could easily be the OLDER one. Read the table first.

Note the template path has TWO SPACES before "WB".
"""
from __future__ import annotations
from openpyxl import load_workbook
import sys

TEMPLATE = (r"\\sdi-dc01\shareddata$\Shared\Estimating\Completed"
            r"\AI Estimating\AISheets\Blank Estimate Sheet  WB 2026.xlsx")


def main():
    print(f"\n  {TEMPLATE}\n")
    try:
        ws = load_workbook(TEMPLATE, data_only=False)["Estimate"]
    except Exception as e:
        sys.exit(f"could not open: {e}")

    print(f"  {'row':>4}  {'H  operation':<30} {'I  £/hr':>10}  {'J  dept':<10} {'K  setup'}")
    print(f"  {'-'*4}  {'-'*30} {'-'*10}  {'-'*10} {'-'*8}")

    powder = []
    for r in range(108, 156):
        op = ws.cell(r, 8).value          # H
        if op in (None, ""):
            continue
        rate  = ws.cell(r, 9).value       # I
        dept  = ws.cell(r, 10).value      # J
        setup = ws.cell(r, 11).value      # K
        mark = ""
        if any(t in str(op).upper() for t in ("COAT", "POWDER", "P/C")) \
           or any(t in str(dept or "").upper() for t in ("POWDER", "P/C")):
            mark = "   <== POWDER"
            powder.append((r, op, rate, dept))
        print(f"  {r:>4}  {str(op):<30} {str(rate):>10}  {str(dept or ''):<10} {setup}{mark}")

    print("\n" + "=" * 72)
    if len(powder) > 1:
        print("  TWO OR MORE POWDER ROWS. Excel's LOOKUP takes the LAST match on a sorted")
        print("  range and misbehaves badly on an unsorted one — so which row wins is not")
        print("  obvious, and we may well be picking the wrong one. THIS IS AN ENGINE-SIDE")
        print("  FIX: match the operation name to the row the estimators actually use.")
        for r, op, rate, dept in powder:
            print(f"        row {r}: {op!r}  rate={rate}  dept={dept}")
    elif len(powder) == 1:
        r, op, rate, dept = powder[0]
        print(f"  ONE powder row: row {r}  {op!r}  rate={rate}  dept={dept}")
        print("  If that rate is 355.43 and dept is P/C, then the engine is reading this")
        print("  template CORRECTLY, and Tim's 331.42 / POWDER is a DIFFERENT template")
        print("  version. In that case the rate is SDI's to set — take it to Tim and Tony,")
        print("  do not change it from here.")
    else:
        print("  NO powder row found in rows 108-155. The rate table is somewhere else —")
        print("  widen the range, or read the LOOKUP range straight off a populated K cell.")
    print("=" * 72)

    print("\n  KEEP THE SIZES IN VIEW:")
    print("      rate        355.43 -> 331.42    =   7%  on the P.Coat line")
    print("      throughput     458 -> 957       = 109%  on the SAME line")
    print("  The throughput is FIFTEEN TIMES the error. Fix the rate if it is genuinely")
    print("  wrong — but it will not move parity.\n")


if __name__ == "__main__":
    main()
