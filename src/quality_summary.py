"""
quality_summary.py  —  rebuilds the "ESTIMATE QUALITY SUMMARY" workbook tab.

This sheet had been removed from the codebase; this restores it, corrected:
  * manufactured-vs-bought-in is classified by NATURE (real fab material + real
    part number) — not by whether the line happened to be priced via Tim's
    bought-in JSON, which is what made the old tab call almost everything
    "Bought-in".
  * geometry-source counts read each part's real geometry_source
    (dxf_flat_pattern -> DXF-exact, pdf -> PDF-inferred, else none).
  * shortcoming flags are rendered in plain English from a glossary.

Public entry point:
    add_quality_summary_sheet(wb, summary)  -> openpyxl worksheet (or None on no data)

Defensive throughout: every field is read with .get and several fallback names,
so a missing/renamed field degrades to a blank cell rather than an exception.
"""
from __future__ import annotations
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime

try:
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - openpyxl always present in the engine
    Font = Alignment = PatternFill = Border = Side = None  # type: ignore
    def get_column_letter(i): return chr(64 + i)

# ---------------------------------------------------------------- glossary ----
# Plain-English meaning for each flag code, keyed by the exact code the engine
# emits. Kept in sync with SDI_estimate_flag_glossary.md / _RISK_PLAIN.
_FLAG_PLAIN: Dict[str, str] = {
    "missing_material_spec":      "no material identified",
    "missing_material_thickness": "thickness not found",
    "missing_material_price":     "material + thickness known but no price on file",
    "many_bends":                 "3+ bends — fold time/tooling risk",
    "large_flat":                 "large flat — sheet size / nesting check",
    "hanging_holes":              "hanging holes — check fit/spacing",
    "weld_required":              "welding/assembly labour applies",
    "section_or_wire_stock_pricing_review": "tube/angle/wire — priced as section, verify",
    "zero_cost_steel":            "steel part returned £0 material — must not happen",
    "web_ai_indicative_system_cost": "price is a web/AI estimate, not a real quote",
    "no material":                "no material identified",
    "no thickness":               "thickness not found",
    "no price found":             "no price could be resolved",
}

_FAB_NON_MATERIALS = {"BOUGHT_IN", "PAPER", "PRINTED_PAPER", "CARD", "UNKNOWN", ""}
_ASSEMBLY_MARKERS = ("-GA", "-WA", "GA-", "WA0", "WELDMENT", "WELDED")


# ---------------------------------------------------------- field helpers ----
def _first(d: Dict[str, Any], *names, default=None):
    for n in names:
        v = d.get(n)
        if v not in (None, "", [], {}):
            return v
    return default

def _flag_list(est: Dict[str, Any]) -> List[str]:
    out: List[str] = []
    for key in ("risk_flags", "review_flags", "flags", "shortcomings"):
        v = est.get(key)
        if isinstance(v, list):
            for x in v:
                if isinstance(x, dict):
                    # review_flags are dicts: {'severity':..,'field':..,'reason':..}
                    # Use 'flag' key if present, else 'reason', else 'field'
                    txt = x.get("flag") or x.get("reason") or x.get("field") or ""
                    if txt:
                        out.append(str(txt))
                elif x:
                    out.append(str(x))
        elif isinstance(v, str) and v:
            out += [s.strip() for s in v.replace(";", ",").split(",") if s.strip()]
    seen, uniq = set(), []
    for f in out:
        if f not in seen:
            seen.add(f); uniq.append(f)
    return uniq

def _plain_flags(flags: List[str]) -> str:
    if not flags:
        return "\u2014"  # em dash
    return "; ".join(_FLAG_PLAIN.get(f, f) for f in flags)

def _material(est: Dict[str, Any]) -> str:
    m = _first(est, "normalized_material", "material")
    if not m:
        ms = est.get("materials")
        if isinstance(ms, list) and ms:
            m = ms[0]
    return str(m or "").upper().replace(" ", "_")

def _thickness(est: Dict[str, Any]) -> str:
    t = _first(est, "normalized_thickness_mm", "thickness_mm", "thickness")
    if t in (None, ""):
        ts = est.get("thicknesses")
        if isinstance(ts, list) and ts:
            t = ts[0]
    if t in (None, ""):
        # check inside material_estimate
        me = est.get("material_estimate") or {}
        t = me.get("thickness_mm") or me.get("normalized_thickness_mm")
    if t in (None, "", 0):
        return ""
    try:
        f = float(t)
        return f"{f:g}" if f > 0 else ""
    except (TypeError, ValueError):
        return str(t)

def _geometry_source(est: Dict[str, Any]) -> str:
    """-> 'dxf' | 'pdf' | 'none' (normalised bucket)."""
    # geometry_source may live at top level or inside normalized_geometry dict
    g = str(_first(est, "geometry_source", "geometry_origin", default="") or "").lower()
    if not g:
        ng = est.get("normalized_geometry") or {}
        g = str(ng.get("geometry_source") or ng.get("source") or "").lower()
    if not g:
        # also check material_estimate path
        me = est.get("material_estimate") or {}
        g = str(me.get("geometry_source") or "").lower()
    if "dxf" in g or "flat_pattern" in g:
        return "dxf"
    if "pdf" in g:
        return "pdf"
    # Final check: if any geometry values are non-zero, call it pdf
    ng = est.get("normalized_geometry") or {}
    if ng.get("estimated_cut_length_mm") or ng.get("estimated_hole_count"):
        return "pdf"
    return "none"

def _geom_label(bucket: str) -> str:
    return {"dxf": "DXF-exact", "pdf": "PDF", "none": "\u2014"}.get(bucket, "\u2014")

def _ext_cost(est: Dict[str, Any]) -> float:
    try:
        return float(_first(est, "extended_total_cost_gbp", "extended_cost_gbp",
                            "line_cost_gbp", default=0.0) or 0.0)
    except (TypeError, ValueError):
        return 0.0

def _unit_cost(est: Dict[str, Any]) -> float:
    try:
        return float(_first(est, "unit_estimate", "unit_cost_gbp", "unit_total_cost_gbp",
                            default=0.0) or 0.0)
    except (TypeError, ValueError):
        return 0.0

def _qty(est: Dict[str, Any]) -> int:
    try:
        return int(float(_first(est, "quantity", "qty", "qty_per_unit", default=1) or 1))
    except (TypeError, ValueError):
        return 1

def _is_indicative(est: Dict[str, Any], flags: List[str]) -> bool:
    if any("indicative" in f or "web_ai" in f or "ai_estimate" in f for f in flags):
        return True
    src = str(_first(est, "price_source", "pricing_basis", "provenance", default="")).lower()
    return any(k in src for k in ("web", "indicative", "ai_estimate", "historical"))

def _classify(est: Dict[str, Any]) -> str:
    """Manufactured | Bought-in | Assembly — by nature, not by pricing route."""
    pn = str(_first(est, "part_number", "part_no", default="")).upper()
    desc = str(est.get("description") or "").upper()
    if any(m in pn for m in _ASSEMBLY_MARKERS) or "WELDMENT" in desc:
        return "Assembly"
    mat = _material(est)
    if est.get("bought_in") is True or mat == "BOUGHT_IN":
        return "Bought-in"
    # fixings / electrics / loom are bought-in regardless of material blank
    if any(k in pn or k in desc for k in ("FIXING", "ELECTRIC", "LOOM", "RIVET", "NUTSERT")):
        return "Bought-in"
    if mat and mat not in _FAB_NON_MATERIALS and pn not in ("", "NONE", "?"):
        return "Manufactured"
    return "Bought-in"


# -------------------------------------------------------------- the sheet ----
def add_quality_summary_sheet(wb, summary: Dict[str, Any]):
    if Font is None:
        return None
    es = (summary.get("estimate_summary") or {})
    parts: List[Dict[str, Any]] = es.get("part_estimates") or []
    if not parts:
        return None

    ds = es.get("data_sufficiency") or {}
    pages = summary.get("pages") or []
    job = (summary.get("source_file", "").split("\\")[-1] or "estimate")

    # ----- classify + tabulate every part -----
    rows = []
    counts = {"Manufactured": 0, "Bought-in": 0, "Assembly": 0}
    geom = {"dxf": 0, "pdf": 0, "none": 0}
    missing_mat, missing_thk, mfg_no_geom = [], [], []
    all_flags: List[str] = []
    bi_verified = bi_indicative = bi_unpriced = 0
    indic_lines: List[str] = []
    priced_count = 0
    parts_cost = 0.0

    for est in parts:
        pn = str(_first(est, "part_number", "part_no", default="?")) or "?"
        desc = str(est.get("description") or "") or "\u2014"
        qty = _qty(est)
        typ = _classify(est)
        gbucket = _geometry_source(est)
        mat = _material(est) or "\u2014"
        thk = _thickness(est)
        flags = _flag_list(est)
        ext = _ext_cost(est)
        unit = _unit_cost(est)
        indicative = _is_indicative(est, flags)
        has_price = (ext > 0) or (unit > 0)

        counts[typ] = counts.get(typ, 0) + 1
        geom[gbucket] = geom.get(gbucket, 0) + 1
        all_flags += flags
        parts_cost += ext if ext > 0 else unit

        if mat in ("", "\u2014") or "missing_material_spec" in flags or "no material" in flags:
            missing_mat.append(pn)
        if not thk:
            missing_thk.append(pn)
        if typ == "Manufactured" and gbucket == "none":
            mfg_no_geom.append(pn)

        if typ == "Bought-in":
            if not has_price:
                bi_unpriced += 1
            elif indicative:
                bi_indicative += 1
            else:
                bi_verified += 1

        if has_price:
            priced_count += 1
        if indicative:
            indic_lines.append(f"{pn} [indicative]")

        priced_lbl = "UNPRICED" if not has_price else ("Indicative" if indicative else "Priced")
        rows.append([pn, desc, qty, typ, _geom_label(gbucket), mat, thk, priced_lbl, _plain_flags(flags)])

    # Supplement bought-in counts from bay_estimate (Tim's JSON lines, tube, etc.)
    # These live outside part_estimates so the loop above misses them.
    be = summary.get("bay_estimate") or {}
    for bl in (be.get("costed_lines") or be.get("bought_in_lines") or []):
        cost = float(bl.get("cost_gbp") or bl.get("total_gbp") or 0.0)
        src = str(bl.get("cost_source") or bl.get("source") or "").lower()
        typ_bl = _classify({"part_number": bl.get("code",""), "description": bl.get("description","")})
        if typ_bl == "Bought-in" or bl.get("kind") in ("catalogue","bought_in"):
            counts["Bought-in"] = counts.get("Bought-in", 0) + 1
            if cost <= 0:
                bi_unpriced += 1
            elif any(k in src for k in ("indicative","web","ai")):
                bi_indicative += 1
            else:
                bi_verified += 1
            parts_cost += cost

    # pages
    p_assembly = sum(1 for p in pages if ((p.get("page_role") or {}).get("primary_role") or "") == "assembly")
    p_detail = sum(1 for p in pages if ((p.get("page_role") or {}).get("primary_role") or "") == "detail")

    distinct_flags = []
    for f in all_flags:
        if f not in distinct_flags:
            distinct_flags.append(f)

    # ----- styling -----
    H = Font(bold=True, size=11)
    TITLE = Font(bold=True, size=13)
    SUB = Font(italic=True, size=9, color="666666")
    HEADFILL = PatternFill("solid", fgColor="1F3864")
    HEADFONT = Font(bold=True, color="FFFFFF", size=10)
    SECTIONFILL = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    WRAP = Alignment(wrap_text=True, vertical="top")

    ws = wb.create_sheet("Estimate Quality Summary")
    ws.sheet_view.showGridLines = False
    r = 1

    def put(row, col, val, font=None, fill=None, align=None, border=False):
        c = ws.cell(row=row, column=col, value=val)
        if font: c.font = font
        if fill: c.fill = fill
        if align: c.alignment = align
        if border: c.border = BORDER
        return c

    # ----- header -----
    put(r, 1, f"ESTIMATE QUALITY SUMMARY  \u2014  {job}", TITLE); r += 1
    qty_job = (
        ((es.get("estimate_workbook_inputs") or {}).get("assumed_job_quantity"))
        or summary.get("assumed_job_quantity")
        or summary.get("quantity")
        or ""
    )
    rev = ds.get("rev") or summary.get("revision") or "\u2014"
    put(r, 1, f"Qty {qty_job}    Rev {rev}    {datetime.now().strftime('%d/%m/%Y')}    |   "
              f"AI estimate \u2014 review before quoting", SUB); r += 2

    # ----- section 1: drawing / input quality -----
    put(r, 1, "1.  DRAWING / INPUT QUALITY", H, SECTIONFILL)
    for col in range(2, 10): put(r, col, "", fill=SECTIONFILL)
    r += 1
    def kv(label, value):
        nonlocal r
        put(r, 1, label, Font(bold=True, size=10))
        put(r, 3, value)
        r += 1
    total_parts = len(parts)
    kv("Parts identified",
       f"{total_parts}   (manufactured {counts.get('Manufactured',0)}, "
       f"bought-in {counts.get('Bought-in',0)}, assembly {counts.get('Assembly',0)})")
    kv("Geometry source",
       f"DXF flat-pattern (exact): {geom['dxf']}     PDF-inferred: {geom['pdf']}     none: {geom['none']}")
    kv("Pages scanned", f"{len(pages)}   (assembly: {p_assembly}, detail: {p_detail})")
    kv("Missing material", f"{len(missing_mat)}" + (f": {', '.join(missing_mat[:8])}" if missing_mat else ""))
    kv("Missing thickness", f"{len(missing_thk)}" + (f": {', '.join(missing_thk[:8])}" if missing_thk else ""))
    kv("Manufactured parts w/o geometry",
       f"{len(mfg_no_geom)} priced without DXF/PDF geometry"
       + (f": {', '.join(mfg_no_geom[:8])}" if mfg_no_geom else ""))
    r += 1

    # ----- per-part table -----
    headers = ["Part No", "Description", "Qty", "Type", "Geometry", "Material", "Thk mm", "Priced", "Shortcomings / flags"]
    for i, h in enumerate(headers, start=1):
        put(r, i, h, HEADFONT, HEADFILL, border=True)
    r += 1
    for row in rows:
        for i, val in enumerate(row, start=1):
            put(r, i, val, border=True, align=WRAP if i in (2, 9) else None)
        r += 1
    r += 1

    # ----- section 2: pricing quality -----
    put(r, 1, "2.  PRICING QUALITY", H, SECTIONFILL)
    for col in range(2, 10): put(r, col, "", fill=SECTIONFILL)
    r += 1
    kv("Parts cost (sum of lines, ex packaging & margin)", f"\u00a3{parts_cost:,.2f}")
    kv("Data sufficiency", str(ds.get("status", "\u2014")))
    bi_total = counts.get("Bought-in", 0)
    kv("Bought-in pricing",
       f"{bi_total} item(s):   VERIFIED {bi_verified},   indicative {bi_indicative},   UNPRICED {bi_unpriced}")
    kv("Indicative-priced lines", "; ".join(indic_lines) if indic_lines else "none")
    kv("Distinct issues flagged",
       f"{len(distinct_flags)}  -  " + "; ".join(_FLAG_PLAIN.get(f, f) for f in distinct_flags[:6])
       + (" ..." if len(distinct_flags) > 6 else ""))
    r += 1

    # ----- what went well -----
    put(r, 1, "What went well", H); r += 1
    well = []
    if priced_count:
        well.append(f"{priced_count} of {total_parts} line(s) resolved a price.")
    if bi_verified:
        well.append(f"{bi_verified} bought-in item(s) priced from a VERIFIED catalogue / supplier rate.")
    if (summary.get("bay_estimate") or {}).get("assembly_pack_labour"):
        well.append("Assembly / pack labour applied from the estimator-confirmed rate.")
    if geom["dxf"]:
        well.append(f"{geom['dxf']} part(s) costed from exact DXF flat-pattern geometry.")
    for w in well:
        put(r, 1, f"+   {w}"); r += 1
    r += 1

    # ----- shortcomings / actions -----
    put(r, 1, "Shortcomings / actions before quoting", H); r += 1
    short = []

    # LLM/web fallbacks — state WHY for each part
    for est in parts:
        pn = str(_first(est, "part_number", "part_no", default="?")) or "?"
        flags = _flag_list(est)
        if not _is_indicative(est, flags):
            continue
        cause = []
        gs = _geometry_source(est)
        mat = _material(est) or ""
        if not mat or mat in ("\u2014", ""):
            cause.append("no material identified on drawing")
        if gs == "none":
            cause.append("no DXF or flat-pattern geometry \u2014 detail drawing missing or not issued")
        elif gs == "pdf":
            cause.append("geometry inferred from PDF only, not a flat DXF")
        if "missing_material_spec" in flags:
            cause.append("material spec absent from title block")
        if not cause:
            cause.append("no catalogue or RAG match for this part code")
        short.append(f"-   {pn} priced by web/AI: {'; '.join(cause)} \u2014 verify against a real quote.")

    if missing_mat:
        short.append(f"-   {len(missing_mat)} part(s) missing material spec ({', '.join(missing_mat[:5])}) \u2014 "
                     f"confirm from the drawing title block.")
    if mfg_no_geom:
        short.append(f"-   {len(mfg_no_geom)} manufactured part(s) have no geometry ({', '.join(mfg_no_geom[:5])}) \u2014 "
                     f"check whether a detail DXF exists; if not, request it from design.")
    if bi_unpriced:
        short.append(f"-   {bi_unpriced} bought-in line(s) UNPRICED \u2014 source a price before quoting.")

    ds_status = str(ds.get("status", ""))
    if ds_status == "insufficient_data":
        dxf_n = geom.get("dxf", 0)
        fab_n = counts.get("Manufactured", 0)
        bi_n  = counts.get("Bought-in", 0)
        short.append(
            f"-   Data sufficiency INSUFFICIENT \u2014 headline total suppressed. "
            f"DXF geometry on {dxf_n} of {fab_n} fabricated part(s). "
            + (f"{bi_n} bought-in/commodity line(s) have no DXF (expected \u2014 they don\u2019t need one). "
               if bi_n else "")
            + "Fabricated parts need flat DXFs from design for a credible auto-estimate."
        )

    short.append("-   Material prices may carry generic provenance / placeholder dates \u2014 confirm stockholder and price freshness.")
    for s_line in short:
        put(r, 1, s_line if s_line.startswith("-") else f"-   {s_line}"); r += 1

    # ----- widths -----
    widths = [14, 34, 6, 13, 11, 14, 8, 11, 46]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws
