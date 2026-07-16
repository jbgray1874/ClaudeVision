"""
Read key cells from the Blank Estimate workbook after template changes.

Usage:
  python src/extract_workbook_constants.py --workbook "input/spreadsheets/.../Blank Estimate Sheet 2026.xlsx"

Pairs with config.WORKBOOK_SOURCE_MAP and WORKBOOK_INPUT_DEFAULTS — paste values into config
or set env vars WORKBOOK_WIRE_COST_PER_TONNE_GBP / WORKBOOK_SHEET_STEEL_COST_PER_TONNE_GBP.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional


def extract_constants(workbook_path: Path) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc

    path = workbook_path.expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(path)

    wb = load_workbook(path, data_only=True, read_only=True)
    out: Dict[str, Any] = {"workbook": str(path)}

    estimate_sheet = None
    for name in wb.sheetnames:
        if name.strip().lower() == "estimate":
            estimate_sheet = wb[name]
            break

    if estimate_sheet is not None:
        try:
            from estimate_sheet_discovery import read_estimate_workbook_inputs

            scan = read_estimate_workbook_inputs(path, sheet_name=estimate_sheet.title)
            out["estimate_sheet_scan"] = scan
            if scan.get("ok"):
                out["estimate_default_job_quantity"] = scan.get("assumed_job_quantity")
                out["estimate_quantity_cell_a1"] = scan.get("quantity_cell_a1")
                if scan.get("wire_cost_per_tonne_gbp") is not None:
                    out["estimate_wire_cost_per_tonne_gbp"] = scan["wire_cost_per_tonne_gbp"]
                if scan.get("sheet_steel_cost_per_tonne_gbp") is not None:
                    out["estimate_sheet_steel_cost_per_tonne_gbp"] = scan["sheet_steel_cost_per_tonne_gbp"]
            if scan.get("assumed_job_quantity") is None:
                out["estimate_D6_default_job_quantity"] = estimate_sheet["D6"].value
            if scan.get("wire_cost_per_tonne_gbp") is None:
                out["estimate_L3_wire_cost_per_tonne_gbp"] = estimate_sheet["L3"].value
            if scan.get("sheet_steel_cost_per_tonne_gbp") is None:
                out["estimate_L5_sheet_steel_cost_per_tonne_gbp"] = estimate_sheet["L5"].value
        except Exception as exc:
            out["estimate_sheet_scan_error"] = str(exc)
            out["estimate_D6_default_job_quantity"] = estimate_sheet["D6"].value
            out["estimate_L3_wire_cost_per_tonne_gbp"] = estimate_sheet["L3"].value
            out["estimate_L5_sheet_steel_cost_per_tonne_gbp"] = estimate_sheet["L5"].value

    mp_sheet = None
    for name in wb.sheetnames:
        if "material price break" in name.lower().replace("_", " "):
            mp_sheet = wb[name]
            break

    headers: List[str] = []
    if mp_sheet is not None:
        out["material_price_break_sheet"] = mp_sheet.title
        # Columns D through N => Excel columns 4..14
        for col in range(4, 15):
            cell = mp_sheet.cell(row=4, column=col)
            headers.append(str(cell.value).strip() if cell.value is not None else "")
        out["material_price_break_row4_D_to_N"] = headers

    wb.close()
    return out


def _print_hints(data: Dict[str, Any]) -> None:
    print("\n# Suggested updates (merge into config.py or environment):\n")
    d6 = data.get("estimate_default_job_quantity") or data.get("estimate_D6_default_job_quantity")
    if d6 is not None:
        print(f"# DEFAULT_JOB_QUANTITY / WORKBOOK_INPUT_DEFAULTS default_job_quantity ≈ {d6}")
    l3 = data.get("estimate_wire_cost_per_tonne_gbp") or data.get("estimate_L3_wire_cost_per_tonne_gbp")
    if l3 is not None:
        print(f"# export WORKBOOK_WIRE_COST_PER_TONNE_GBP={l3}")
    l5 = data.get("estimate_sheet_steel_cost_per_tonne_gbp") or data.get("estimate_L5_sheet_steel_cost_per_tonne_gbp")
    if l5 is not None:
        print(f"# export WORKBOOK_SHEET_STEEL_COST_PER_TONNE_GBP={l5}")
    hdrs = data.get("material_price_break_row4_D_to_N")
    if hdrs:
        print("# Verify MATERIAL_PRICE_BREAK_HEADERS in config.py still matches row 4 above.")


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Extract Estimate / Material Price Break constants from .xlsx")
    parser.add_argument("--workbook", type=Path, required=True, help="Path to Blank Estimate .xlsx")
    parser.add_argument("--hints", action="store_true", help="Print suggested env/config lines after JSON")
    args = parser.parse_args(argv)

    try:
        data = extract_constants(args.workbook)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(json.dumps(data, indent=2, default=str))
    if args.hints:
        _print_hints(data)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
