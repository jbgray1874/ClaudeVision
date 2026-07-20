r"""READ-ONLY. Get the ENGINE's REAL section subtotals from the populated workbook (same cells
Tim has), so parity compares section-to-section correctly — not derived buckets. Also resolve
the £214.11 vs £187.43 unit-cost discrepancy by reading the populated workbook's own total cell.
Reads the latest populated 1282 xlsx. No edits."""
import os, glob, re

est_dir=r"C:\ClaudeVision\output\estimates"
cands=sorted(glob.glob(os.path.join(est_dir,"1282*Milwaukee*Wall*Bay*.xlsx")), key=os.path.getmtime)
if not cands:
    print("no populated 1282 xlsx found in", est_dir); raise SystemExit
wb_path=cands[-1]
print("populated workbook:", os.path.basename(wb_path))

try:
    import openpyxl
    wb=openpyxl.load_workbook(wb_path, data_only=True)  # cached computed values
    ws=wb["Estimate"] if "Estimate" in wb.sheetnames else wb[wb.sheetnames[0]]
    print("sheet:", ws.title, " dims:", ws.max_row,"x",ws.max_column, "\n")

    # Find labelled subtotal rows exactly as on Tim's sheet
    wants=["total material cost","total labour cost","total unit cost","sell price","rebate",
           "sheet steel","other sheet material","wire","standard materials"]
    def rowtext(r):
        return " ".join(str(ws.cell(r,c).value) for c in range(1,min(ws.max_column,16)+1) if ws.cell(r,c).value not in (None,""))
    print("--- labelled rows + last numeric on row ---")
    for r in range(1, ws.max_row+1):
        txt=rowtext(r).lower()
        for w in wants:
            if w in txt:
                nums=[ws.cell(r,c).value for c in range(1,ws.max_column+1)
                      if isinstance(ws.cell(r,c).value,(int,float)) and ws.cell(r,c).value not in (0,None)]
                val=nums[-1] if nums else None
                print(f"  r{r:<3} '{w}': {val}")
                break

    # subtotal cells the engine names in the bundle: M59 material, M103 labour, M105 total
    print("\n--- specific cells (M/L col) ---")
    for ref in ["M59","M103","M105","L105","M107","M109","L111","M115","M120"]:
        try: print(f"  {ref} = {ws[ref].value}")
        except Exception as e: print(f"  {ref} = <err {e}>")

    # section subtotals: look for the block-total cells near each section header
    print("\n--- scan col M (13) for nonzero subtotals by row band ---")
    for r in range(1, ws.max_row+1):
        v=ws.cell(r,13).value
        if isinstance(v,(int,float)) and v not in (0,None):
            lbl=rowtext(r)[:46]
            print(f"  M{r} = {v:<10} | {lbl}")
except Exception as e:
    print("openpyxl read failed (formulas may be uncached):", repr(e))
    print("If cached values are zero, the file needs opening+saving in Excel to cache formula results.")
