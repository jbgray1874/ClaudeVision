# -*- coding: utf-8 -*-
"""READ-ONLY. Locate the LABOUR ROUTES on the sheet. v2 probe found 0 in cols A/B rows 63-135, but
the earlier full dump showed a populated labour block. Scan the whole sheet for the 'Labour' header
and dump the actual rows/columns where operations live, so we can confirm routes are present + sensible.

Run from C:\\ClaudeVision\\src :
  C:\\ClaudeVision\\.venv\\Scripts\\python.exe _probe_routes.py
"""
import glob, os
try:
    import openpyxl
except ImportError:
    print("run with venv python"); raise SystemExit(1)
from openpyxl.utils import get_column_letter

est = sorted(glob.glob(r"C:\ClaudeVision\output\estimates\12532-03RecipeCard_*.xlsx"), key=os.path.getmtime)
path = est[-1]
wb = openpyxl.load_workbook(path, data_only=False)
print(f"=== Sheet: {os.path.basename(path)} — worksheets: {wb.sheetnames} ===\n")
ws = wb["Estimate"] if "Estimate" in wb.sheetnames else wb.active

# 1) Find the 'Labour' header and the 'Operation' column header anywhere on the sheet
print("=== Locate Labour section + Operation/Dept headers ===")
labour_hdr_row = None
for r in range(1, ws.max_row+1):
    for c in range(1, 20):
        v = ws.cell(row=r, column=c).value
        if v and str(v).strip().lower() in ("labour", "operation", "dept.", "rate per hour", "total hours", "labour cost"):
            print(f"  r{r} {get_column_letter(c)}: {v}")
            if str(v).strip().lower() == "labour":
                labour_hdr_row = r

# 2) Dump the rows just after the Operation header — find real op names
print("\n=== First 20 populated rows in the labour region (any op-like text in cols A-C) ===")
start = (labour_hdr_row or 60)
count = 0
for r in range(start, min(start+80, ws.max_row+1)):
    a = ws.cell(row=r, column=1).value
    b = ws.cell(row=r, column=2).value
    c = ws.cell(row=r, column=3).value
    line = " | ".join(f"{get_column_letter(i+1)}={str(x)[:40]}" for i,x in enumerate([a,b,c]) if x and str(x).strip())
    if line:
        print(f"  r{r}: {line}")
        count += 1
    if count >= 25:
        break

print(f"\n  max_row on sheet: {ws.max_row}")
