from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

import config

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover
    load_workbook = None


def _safe_float(value: Any) -> Optional[float]:
    try:
        if value is None:
            return None
        if isinstance(value, str):
            v = value.strip().replace("£", "").replace(",", "")
            if not v:
                return None
            return float(v)
        return float(value)
    except (TypeError, ValueError):
        return None


def _extract_by_path(payload: Dict[str, Any], dotted_path: str) -> Any:
    cur: Any = payload
    for key in dotted_path.split("."):
        if not isinstance(cur, dict):
            return None
        cur = cur.get(key)
    return cur


def _percent_diff(expected: Optional[float], actual: Optional[float]) -> Optional[float]:
    if expected is None or actual is None:
        return None
    if abs(expected) < 1e-9:
        return None
    return round(abs(actual - expected) / abs(expected) * 100.0, 4)


def _status_from_pct(pct: Optional[float], match: float, warn: float) -> str:
    if pct is None:
        return "review"
    if pct <= match:
        return "match"
    if pct <= warn:
        return "warning"
    return "fail"


def build_report_rows(summary: Dict[str, Any], workbook_path: Path, sheet_name: str = "Estimate") -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for workbook parity report")
    wb = load_workbook(workbook_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' not found in workbook")
    ws = wb[sheet_name]

    thresholds = (config.WORKBOOK_EQUIVALENT_PRICING or {}).get("variance_thresholds_pct") or {}
    match_pct = float(thresholds.get("match", 3.0))
    warn_pct = float(thresholds.get("warning", 10.0))

    rows: List[Dict[str, Any]] = []
    mapping = (config.ESTIMATE_TEMPLATE_WRITEBACK or {}).get("output_cells") or {}
    for cell_ref, path in mapping.items():
        expected_raw = _extract_by_path(summary, str(path))
        actual_raw = ws[cell_ref].value
        expected = _safe_float(expected_raw)
        actual = _safe_float(actual_raw)
        pct = _percent_diff(expected, actual)
        status = _status_from_pct(pct, match_pct, warn_pct)
        rows.append(
            {
                "cell": cell_ref,
                "summary_path": path,
                "expected_value": expected,
                "actual_value": actual,
                "actual_cell_raw": actual_raw,
                "pct_variance": pct,
                "status": status,
            }
        )
    return rows


def write_reports(rows: List[Dict[str, Any]], csv_path: Path, json_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "cell",
        "summary_path",
        "expected_value",
        "actual_value",
        "actual_cell_raw",
        "pct_variance",
        "status",
    ]
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    json_path.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate cell-level parity report between estimate_summary and workbook cells.")
    parser.add_argument("--summary-json", required=True, help="Path to scan summary JSON containing estimate_summary")
    parser.add_argument("--workbook", required=True, help="Workbook to validate")
    parser.add_argument("--sheet", default="Estimate", help="Sheet name to validate")
    parser.add_argument("--out-csv", default=str(config.CSV_DIR / "estimate_workbook_parity_report.csv"), help="CSV report path")
    parser.add_argument("--out-json", default=str(config.CSV_DIR / "estimate_workbook_parity_report.json"), help="JSON report path")
    args = parser.parse_args()

    summary_path = Path(args.summary_json).resolve()
    workbook_path = Path(args.workbook).resolve()
    summary = json.loads(summary_path.read_text(encoding="utf-8"))

    rows = build_report_rows(summary, workbook_path, sheet_name=args.sheet)
    write_reports(rows, Path(args.out_csv).resolve(), Path(args.out_json).resolve())

    total = len(rows)
    fail = sum(1 for r in rows if r["status"] == "fail")
    warn = sum(1 for r in rows if r["status"] == "warning")
    match = sum(1 for r in rows if r["status"] == "match")
    print(f"Parity rows={total} match={match} warning={warn} fail={fail}")
    print(f"CSV: {Path(args.out_csv).resolve()}")
    print(f"JSON: {Path(args.out_json).resolve()}")


if __name__ == "__main__":
    main()
