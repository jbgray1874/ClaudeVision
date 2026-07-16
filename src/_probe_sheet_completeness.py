# -*- coding: utf-8 -*-
"""READ-ONLY. Answer 3 questions for handoff-readiness:
  (A) Are all parts IDENTIFIED + VISIBLE on the sheet? (BOM block + steel block + other-sheet)
  (B) Are the ROUTES (labour operations per part) sensible? (right ops for each part type)
  (C) What's still £0 / None?
Reads the latest populated xlsx directly.

Run from C:\\ClaudeVision\\src :
  C:\\ClaudeVision\\.venv\\Scripts\\python.exe _probe_sheet_completeness.py
"""
import glob, os
from pathlib import Path
try:
    import openpyxl
except ImportError:
    print("openpyxl not in this python — run with the venv python"); raise SystemExit(1)

# newest estimate xlsx
est = sorted(glob.glob(r"C:\ClaudeVision\output\estimates\12532-03RecipeCard_*.xlsx"), key=os.path.getmtime)
if not est:
    print("no estimate xlsx found"); raise SystemExit(1)
path = est[-1]
print(f"=== Sheet: {os.path.basename(path)} ===\n")
wb = openpyxl.load_workbook(path, data_only=False)
ws = wb["Estimate"] if "Estimate" in wb.sheetnames else wb.active

def val(r, c):
    v = ws.cell(row=r, column=c).value
    return "" if v is None else str(v)

print("=== (A) BOM BLOCK (bought-in + tube) rows 11-25, col C=desc, H=price, I=qty ===")
for r in range(11, 26):
    d = val(r, 3)
    if d.strip():
        print(f"  row {r}: {d[:60]:60}  price={val(r,8):>10}  qty={val(r,9)}")

print("\n=== (B) STEEL BLOCK rows 38-48, col C=desc, F=len, G=wid, H=gauge ===")
for r in range(38, 49):
    d = val(r, 3)
    if d.strip():
        print(f"  row {r}: {d[:55]:55}  L={val(r,6):>8} W={val(r,7):>8} g={val(r,8)}")

print("\n=== OTHER SHEET (acrylic) rows 51-58, col B=desc ===")
for r in range(51, 59):
    d = val(r, 2)
    if d.strip():
        print(f"  row {r}: {d[:55]:55}  L={val(r,3):>8} W={val(r,4):>8} t={val(r,5)}  costsheet={val(r,10)}")

print("\n=== (B) ROUTES — Labour block rows 63-135, col A=operation, B=part-desc ===")
ops_by_part = {}
for r in range(63, 136):
    op = val(r, 1); pd = val(r, 2)
    if op.strip():
        # extract part name after the emdash
        pn = pd.split("—")[-1].strip() if "—" in pd else pd
        ops_by_part.setdefault(pn, []).append(op)
for pn, ops in ops_by_part.items():
    print(f"  {pn[:45]:45} -> {', '.join(ops)}")

print("\n=== (C) WHAT'S STILL BLANK/ZERO on the sheet ===")
# BOM prices that are blank
for r in range(11, 26):
    d = val(r,3)
    if d.strip() and not val(r,8).strip():
        print(f"  BOM row {r} NO PRICE: {d[:50]}")
