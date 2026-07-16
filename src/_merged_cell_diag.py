"""READ-ONLY. Diagnoses the wb_populate 'MergedCell ... read-only' crash on the
widened template WITHOUT trial-and-error.

It:
  1. Opens the widened template, lists EVERY merged cell range on the 'Estimate' sheet.
  2. Reads wb_populate.py's cell-map (which columns/rows it writes to).
  3. Flags any merged range that overlaps a cell wb_populate tries to write —
     that overlap IS the crash. Tells you exactly which cells to unmerge.

No writes, no changes. Just tells us precisely what to unmerge to keep the widened sheet.
"""
import re
from pathlib import Path
import openpyxl

TEMPLATE = r"\\sdi-dc01\shareddata$\Shared\Estimating\Completed\AI Estimating\AISheets\Blank Estimate Sheet  WB 2026.xlsx"
SRC = Path(r"C:\ClaudeVision\src")

print("=" * 78)
print("1. ALL merged cell ranges on the 'Estimate' sheet")
print("=" * 78)
try:
    wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
except Exception as e:
    print(f"  could not open template: {e}")
    raise SystemExit(1)

# find the estimate sheet (name may vary)
ws = None
for name in wb.sheetnames:
    if name.strip().lower() == "estimate":
        ws = wb[name]; break
if ws is None:
    ws = wb[wb.sheetnames[0]]
    print(f"  (no 'Estimate' tab; using first sheet '{ws.title}')")
else:
    print(f"  sheet: '{ws.title}'")

merged = list(ws.merged_cells.ranges)
print(f"  {len(merged)} merged range(s):")
for m in sorted(merged, key=lambda r: (r.min_row, r.min_col)):
    print(f"     {m}   (rows {m.min_row}-{m.max_row}, cols {m.min_col}-{m.max_col})")

print("\n" + "=" * 78)
print("2. Cell-map from wb_populate.py (which cells the engine writes)")
print("=" * 78)
wp = (SRC / "wb_populate.py").read_text(encoding="utf-8", errors="replace")
# pull the CELL_MAP / block dicts — first_row/last_row/col_* entries
for i, l in enumerate(wp.splitlines(), 1):
    if re.search(r'first_row|last_row|col_\w+|"bom"|"steel"|"labour"|"other|"header"|quantity|D6|first_col', l):
        s = l.strip()
        if s and not s.startswith("#"):
            print(f"    {i:5}: {s[:100]}")

print("\n" + "=" * 78)
print("3. OVERLAP CHECK — which write-columns fall inside a merged range")
print("=" * 78)
# Heuristic: the BOM block writes cols 3,8,9,10,11,12 (C,H,I,J,K,L) rows 11-31 per the map.
# Flag any merged range intersecting those. Also check labour/steel areas broadly.
write_targets = []
# BOM: rows 11-31, cols C,H,I,J,K,L
for row in range(11, 32):
    for col in (3, 8, 9, 10, 11, 12):
        write_targets.append((row, col))
conflicts = []
for (r, c) in write_targets:
    for m in merged:
        if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
            # is (r,c) the top-left (writable) or a non-top-left (read-only)?
            is_topleft = (r == m.min_row and c == m.min_col)
            if not is_topleft:
                conflicts.append((r, c, str(m)))
if conflicts:
    print("  *** CONFLICTS — these write-target cells are NON-top-left in a merge (CRASH source): ***")
    seen = set()
    for r, c, m in conflicts:
        col_letter = openpyxl.utils.get_column_letter(c)
        key = m
        if key not in seen:
            print(f"     merged range {m}  -> unmerge this")
            seen.add(key)
    print("\n  FIX: unmerge the ranges listed above (they sit in the BOM write columns).")
else:
    print("  No BOM-column conflicts found. The crash may be in a DIFFERENT block")
    print("  (steel/labour/header). Check the full merged-range list in section 1 against")
    print("  the cell-map in section 2 — look for merges the engine writes into that are")
    print("  NOT top-left cells.")

wb.close()
