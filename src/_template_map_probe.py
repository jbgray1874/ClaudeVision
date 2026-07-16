#!/usr/bin/env python3
r"""
_template_map_probe.py  —  READ-ONLY.

WHY: the BOM block holds only 15 rows. Job 1282 produces 16 bought-in parts, so one is
silently thrown away:

    [wb_populate] BOM overflow: 16 BOM/tube parts but only 15 rows — extra parts DROPPED.

That cost 1282 its £26 LED Downlights (~£27 with scrap) — the whole of the "£28 mystery".
Any job with >15 bought-ins is quietly under-costed on the deliverable. This is the
highest-severity defect open: it does not produce an obviously-wrong number, it produces a
plausible one that is short.

We are widening the BOM block IN EXCEL (Excel re-points every formula and LOOKUP range
automatically; openpyxl does NOT, which would silently break the dept-rate lookups at
$H$155:$H$186 and the wire-gauge lookups at $H$191:$H$199).

RUN THIS BEFORE AND AFTER THE WIDENING. It reports where every block actually starts, so
wb_populate's CELL_MAP is rewritten from MEASURED positions, not assumed ones.

Usage:
    C:\ClaudeVision\.venv\Scripts\python.exe _template_map_probe.py
    C:\ClaudeVision\.venv\Scripts\python.exe _template_map_probe.py "C:\path\to\template.xlsx"
"""
from __future__ import annotations
import sys, re
from openpyxl import load_workbook

DEFAULT_TPL = None   # falls back to wb_populate's configured template
SHEET = "Estimate"
MAXROW = 260
MAXCOL = 40

MARKERS = [
    ("BOM header",            "Bill of Materials"),
    ("Wire header",           "Wire"),
    ("Sheet Steel header",    "Sheet Steel"),
    ("Other Sheet header",    "Other Sheet Material"),
    ("Total Material Cost",   "Total Material Cost"),
    ("Labour header",         "Operation"),
    ("Total Labour Cost",     "Total Labour Cost"),
    ("Total Unit Cost",       "Total Unit Cost"),
    ("Powder £/kg label",     "Powder £/kg"),
    ("Total Powder Per Unit", "Total Powder Per Unit"),
    ("Rebate Calculator",     "Rebate Calculator"),
]


def resolve_template(argv):
    if len(argv) > 1:
        return argv[1]
    try:
        import wb_populate  # noqa
        for attr in ("TEMPLATE_PATH", "template_path", "TPL", "TEMPLATE"):
            v = getattr(wb_populate, attr, None)
            if isinstance(v, str) and v.lower().endswith((".xlsx", ".xlsm")):
                return v
    except Exception:
        pass
    sys.exit("Could not resolve the template path — pass it as an argument:\n"
             '   python _template_map_probe.py "C:\\path\\to\\Blank Estimate Sheet  WB 2026.xlsx"')


def rowtext(ws, r):
    return " ".join(
        str(ws.cell(r, c).value)
        for c in range(1, MAXCOL + 1)
        if ws.cell(r, c).value not in (None, "")
    )


def main():
    tpl = resolve_template(sys.argv)
    print("TEMPLATE:", tpl)
    wb = load_workbook(tpl, data_only=False)
    if SHEET not in wb.sheetnames:
        sys.exit(f"No '{SHEET}' sheet. Sheets: {wb.sheetnames}")
    ws = wb[SHEET]

    print(f"\ndimensions: {ws.max_row} rows x {ws.max_column} cols")
    print("\n" + "=" * 80)
    print("BLOCK MARKERS (first row whose text contains the marker)")
    print("=" * 80)

    found = {}
    for label, needle in MARKERS:
        hit = None
        for r in range(1, MAXROW):
            if needle.upper() in rowtext(ws, r).upper():
                hit = r
                break
        found[label] = hit
        print(f"  {label:<24} row {hit if hit else '-- NOT FOUND --'}")

    print("\n" + "=" * 80)
    print("DERIVED DATA-ROW RANGES  (header row + 1  ->  next block header - 1)")
    print("=" * 80)
    order = ["BOM header", "Wire header", "Sheet Steel header",
             "Other Sheet header", "Total Material Cost", "Labour header",
             "Total Labour Cost"]
    rows = [(l, found[l]) for l in order if found[l]]
    for i, (label, r) in enumerate(rows[:-1]):
        nxt = rows[i + 1][1]
        # data starts after the header row and its column-title row
        start = r + 2 if label in ("BOM header", "Wire header", "Sheet Steel header",
                                   "Other Sheet header", "Labour header") else r + 1
        end = nxt - 1
        n = end - start + 1
        print(f"  {label:<24} data rows {start:>4} .. {end:<4}  ({n} row(s))")

    print("\n" + "=" * 80)
    print("KEY CELLS")
    print("=" * 80)
    for cell in ("D6", "L3", "L5", "AD57", "AE57", "AF57", "AF58", "M67", "M145"):
        try:
            print(f"  {cell:<6} {ws[cell].value!r}")
        except Exception as e:
            print(f"  {cell:<6} <error {e}>")

    print("\n" + "=" * 80)
    print("LOOKUP RANGES REFERENCED IN FORMULAS (these MUST move with the rows)")
    print("=" * 80)
    refs = set()
    pat = re.compile(r"\$[A-Z]{1,2}\$(\d+):\$[A-Z]{1,2}\$(\d+)")
    for r in range(1, MAXROW):
        for c in range(1, MAXCOL + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith("="):
                for m in pat.finditer(v):
                    refs.add(m.group(0))
    for ref in sorted(refs):
        print("  ", ref)

    print("""
AFTER WIDENING: re-run this. Every block start should shift by the number of rows
inserted, and the lookup ranges above should have moved with them (Excel does this
automatically — that is why we are NOT doing this in openpyxl).
Paste the output and the CELL_MAP patch will be written from these MEASURED rows.
""")


if __name__ == "__main__":
    main()
