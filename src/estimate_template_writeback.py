from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict

import config

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover
    load_workbook = None


def _extract_by_path(payload: Dict[str, Any], dotted_path: str) -> Any:
    cur: Any = payload
    for key in dotted_path.split("."):
        if not isinstance(cur, dict):
            return None
        cur = cur.get(key)
    return cur


def write_estimate_template_from_summary(
    summary: Dict[str, Any],
    template_workbook: Path,
    output_workbook: Path,
    sheet_name: str = "Estimate",
) -> Path:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for estimate template write-back")
    if template_workbook.suffix.lower() == ".xls":
        raise RuntimeError("Write-back requires .xlsx template (openpyxl cannot save legacy .xls).")

    wb = load_workbook(template_workbook)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' not found in template workbook")
    ws = wb[sheet_name]

    mapping = (config.ESTIMATE_TEMPLATE_WRITEBACK or {}).get("output_cells") or {}
    changed = {}
    for cell_ref, dotted_path in mapping.items():
        value = _extract_by_path(summary, str(dotted_path))
        if value is None:
            continue
        ws[cell_ref] = value
        changed[cell_ref] = value

    output_workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_workbook)

    audit_path = output_workbook.with_suffix(".writeback.audit.json")
    audit_path.write_text(
        json.dumps(
            {
                "template_workbook": str(template_workbook),
                "output_workbook": str(output_workbook),
                "sheet_name": sheet_name,
                "cells_written": changed,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return output_workbook

