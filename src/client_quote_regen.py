"""Regenerate a client quote from an estimator's AMENDED workbook — the manual-override path.

THE SPREADSHEET IS THE SOURCE OF TRUTH HERE, NOT THE ENGINE. The estimator opens the AI Estimate
workbook, edits it (fills a missing rate, corrects a price, sets a margin), saves it, and — hours
or days later — uploads that amended sheet and re-enters three facts a saved sheet cannot be
trusted to still carry: the number of units, the drawing number, and the client. This reads the
estimator's OWN figure straight off that sheet (Excel has already recomputed it) and re-renders
ONLY the client quote, through the same we.are.sdi template the engine uses. The job report and the
provenance tab are deliberately NOT touched — this is the estimator's number now, not the engine's.

Any drawings uploaded alongside are ignored: this path never re-reads a drawing or re-runs the
engine. It turns an amended sheet + three fields into a fresh, correctly-headed client quote.
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, Optional, Tuple

# The labels on the Estimate sheet whose row carries the figure we want, weakest-preferred last.
# Read by LABEL, not by cell address, so a template that shifts a row down does not silently read
# the wrong number. "Sell Price" is what the customer pays (post-margin); "Total Unit Cost Price"
# is the fallback when no margin has been set and the two are equal.
_PRICE_LABELS = ("Sell Price", "Total Unit Cost Price", "Unit Cost")
_QTY_LABELS = ("Quantity",)


def _num(value: Any) -> Optional[float]:
    """A number from a cell that may be a float, an int, or '£146.00' text."""
    if isinstance(value, (int, float)):
        return float(value) if value == value else None      # reject NaN
    if value is None:
        return None
    s = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(s) if s not in ("", "-", ".") else None
    except ValueError:
        return None


def _find_value_in_row(ws, row_idx: int, label_col: int) -> Optional[float]:
    """The first credible number to the RIGHT of a label cell, on the same row."""
    for col in range(label_col + 1, ws.max_column + 1):
        v = _num(ws.cell(row=row_idx, column=col).value)
        if v is not None:
            return v
    return None


def read_estimate_figures(workbook_path: str | Path) -> Dict[str, Any]:
    """The amended sheet's own numbers: {sell_price, unit_cost, quantity, price, source_label}.

    Reads the CACHED values Excel wrote when the estimator saved (openpyxl data_only), so it sees
    the recomputed figure, not the formula. If the sheet was never opened-and-saved in Excel the
    cache is empty — that is reported as a clear error, not a silent zero, because a quote built on
    a blank is worse than no quote."""
    from openpyxl import load_workbook

    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"amended workbook not found: {path}")
    wb = load_workbook(str(path), data_only=True, read_only=True)
    try:
        ws = wb["Estimate"] if "Estimate" in wb.sheetnames else wb[wb.sheetnames[0]]
        found: Dict[str, float] = {}
        qty: Optional[float] = None
        wanted = {lbl.lower(): lbl for lbl in _PRICE_LABELS}
        qty_wanted = {lbl.lower() for lbl in _QTY_LABELS}
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                text = cell.value.strip().lower()
                if text in wanted and wanted[text] not in found:
                    v = _find_value_in_row(ws, cell.row, cell.column)
                    if v is not None:
                        found[wanted[text]] = v
                elif text in qty_wanted and qty is None:
                    qty = _find_value_in_row(ws, cell.row, cell.column)
    finally:
        wb.close()

    # The customer price, weakest-preferred last — Sell Price first.
    price = None
    source_label = ""
    for lbl in _PRICE_LABELS:
        if lbl in found and found[lbl] > 0:
            price, source_label = found[lbl], lbl
            break
    if price is None:
        raise ValueError(
            f"no Unit Cost / Sell Price figure could be read from {path.name}. Open the workbook "
            f"in Excel, let it recalculate, and Save before uploading — the figures are formulas "
            f"and are only stored once Excel has computed them.")
    return {
        "sell_price": found.get("Sell Price"),
        "unit_cost": found.get("Total Unit Cost Price") or found.get("Unit Cost"),
        "quantity": int(qty) if qty else None,
        "price": price,
        "source_label": source_label,
    }


def _job_stem_from_workbook(workbook_path: Path) -> str:
    """The job's own stem from an estimate workbook's filename, with the run timestamp removed.

    The engine writes '<stem>_20260818_133037.xlsx'; its summary JSON is '<stem>.json'. Stripping
    the timestamp is what lets an override find the job it came from days later."""
    return re.sub(r"_\d{8}_\d{6}$", "", Path(workbook_path).stem)


def find_original_summary(workbook_path: str | Path) -> Optional[Dict[str, Any]]:
    """The engine's saved summary for the job this workbook came from, or None.

    THE OVERRIDE QUOTE WAS THINNER THAN THE ENGINE'S, AND THAT IS WHAT A CUSTOMER SAW. Built from
    price alone, it had no parts — so the specification read 'Material: As drawing', the operations
    list collapsed to two generic lines, and the GA image was dropped because nothing told it which
    drawing to render. The estimator's number was right and everything around it had been lost.

    The job's own summary is on disk from the original run. Starting from it and overlaying only
    what the estimator changed keeps the whole quotation — GA, materials, finish, the real
    operations — and changes just the money and the three re-entered facts. Missing (a job run on
    another machine, or output cleared) falls back to the minimal summary: a plain quote beats no
    quote, and the caller says which happened."""
    stem = _job_stem_from_workbook(workbook_path)
    candidates = []
    try:
        import config as _cfg
        candidates.append(Path(getattr(_cfg, "JSON_DIR")) / f"{stem}.json")
    except Exception:                                            # noqa: BLE001
        pass
    # Beside the workbook too, for a pack whose outputs were filed with it.
    candidates.append(Path(workbook_path).parent / f"{stem}.json")
    for path in candidates:
        try:
            if path.is_file():
                data = json.loads(path.read_text(encoding="utf-8"))
                if isinstance(data, dict):
                    return data
        except Exception:                                        # noqa: BLE001
            continue
    return None


def _summary_from_figures(figures: Dict[str, Any], *, units: int, drawing_number: str,
                          client: str, job_stem: str,
                          base: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """The summary the quote renderer reads: the ORIGINAL job where one is on disk, with the
    estimator's price and the three re-entered facts written over it — else a minimal stand-in.

    Only the four things the estimator restated are overwritten. Everything else the quotation
    shows (the GA, the material, the finish, the operation list) is the engine's own reading of
    the drawings, which the estimator did not change by editing a price."""
    summary: Dict[str, Any] = dict(base) if isinstance(base, dict) else {}
    summary["job_output_stem"] = job_stem
    summary["manual_estimator_override"] = True

    # The drawing number the estimator typed wins over the one read off the title block: they are
    # looking at the pack now, and a regen is when a misread number gets corrected.
    _extract = dict(summary.get("llm_full_extract") or {})
    _info = dict(_extract.get("drawing_info") or {})
    if str(drawing_number or "").strip():
        _info["drawing_number"] = str(drawing_number).strip()
    _extract["drawing_info"] = _info
    summary["llm_full_extract"] = _extract

    _es = dict(summary.get("estimate_summary") or {})
    _wep = dict(_es.get("workbook_equivalent_pricing") or {})
    _wep["m105_total_unit_cost_gbp"] = figures["price"]
    _es["workbook_equivalent_pricing"] = _wep
    _inputs = dict(_es.get("estimate_workbook_inputs") or {})
    _inputs["assumed_job_quantity"] = int(units)
    _es["estimate_workbook_inputs"] = _inputs
    summary["estimate_summary"] = _es
    return summary


def regenerate_quote_from_workbook(workbook_path: str | Path, *, units: int, drawing_number: str,
                                   client: str, out_dir: Optional[str | Path] = None,
                                   job_stem: Optional[str] = None) -> Tuple[str, Dict[str, Any]]:
    """Read an amended workbook + three re-entered fields and write a fresh client quote HTML.

    Returns (quote_html_path, figures). Reuses build_quote_html so the override quote is
    pixel-identical to an engine quote — same branding, layout and provisional labelling — it
    just carries the estimator's number. Nothing else is regenerated."""
    from client_quote_html import build_quote_html

    path = Path(workbook_path)
    stem = job_stem or re.sub(r"[^\w\- ]", "", str(drawing_number or path.stem)).strip() or "quote"
    figures = read_estimate_figures(path)
    if not str(client or "").strip():
        raise ValueError("client is required — it heads the quotation and picks the logo")
    if not int(units or 0) > 0:
        raise ValueError("units must be a positive whole number")

    # START FROM THE JOB, NOT FROM THE PRICE. Without this the quotation loses everything the
    # estimator did not restate — the GA image, the material, the finish, the operation list.
    _base = find_original_summary(path)
    figures["source_summary_found"] = _base is not None
    if _base is None:
        print(f"   [regen] no saved summary found for '{stem}' — the quote will carry the price "
              f"and the three fields you entered, but not the GA image, materials or operations "
              f"from the original run.", flush=True)
    summary = _summary_from_figures(figures, units=int(units), drawing_number=str(drawing_number),
                                    client=str(client), job_stem=stem, base=_base)
    html = build_quote_html(summary, job_stem=stem, manual_workbook=str(path), customer=str(client))

    out_dir_p = Path(out_dir) if out_dir else path.parent
    out_dir_p.mkdir(parents=True, exist_ok=True)
    out_path = out_dir_p / f"{stem}_quote.html"
    out_path.write_text(html, encoding="utf-8")
    return str(out_path), figures


def run_estimator_override(uploaded_workbook: str | Path, *, units: int, drawing_number: str,
                           client: str, quote_dir: Optional[str | Path] = None,
                           override_xlsx_dir: Optional[str | Path] = None,
                           job_stem: Optional[str] = None) -> Dict[str, Any]:
    """The full override loop: save the estimator's amended workbook as the manual-override record,
    regenerate the client quote to the AISheets share, and return both paths plus the figures read.

    Destinations default to config (AISheets share for the quote; same for the override sheet unless
    SDI_OVERRIDE_XLSX_DIR points elsewhere). Nothing else is regenerated — the job report and the
    provenance tab are the engine's and are left as they were.
    """
    import shutil
    try:
        import config as _cfg
        _default_quote = getattr(_cfg, "MANUAL_OVERRIDE_QUOTE_DIR", None)
        _default_xlsx = getattr(_cfg, "MANUAL_OVERRIDE_XLSX_DIR", None) or _default_quote
    except Exception:                                            # noqa: BLE001
        _default_quote = _default_xlsx = None

    src = Path(uploaded_workbook)
    if not src.exists():
        raise FileNotFoundError(f"uploaded workbook not found: {src}")
    stem = job_stem or re.sub(r"[^\w\- ]", "", str(drawing_number or src.stem)).strip() or "quote"

    # Validate BEFORE writing anything, so a bad request leaves no half-made files on the share.
    figures = read_estimate_figures(src)
    if not str(client or "").strip():
        raise ValueError("client is required — it heads the quotation and picks the logo")
    if not int(units or 0) > 0:
        raise ValueError("units must be a positive whole number")

    # 1. The amended workbook, saved as the manual-override record.
    xlsx_dir = Path(override_xlsx_dir or _default_xlsx or src.parent)
    xlsx_dir.mkdir(parents=True, exist_ok=True)
    override_xlsx = xlsx_dir / f"{stem}_MANUAL_OVERRIDE.xlsx"
    shutil.copy2(src, override_xlsx)

    # 2. The client quote, regenerated from that record, to the AISheets share.
    q_dir = Path(quote_dir or _default_quote or xlsx_dir)
    quote_path, _ = regenerate_quote_from_workbook(
        override_xlsx, units=int(units), drawing_number=str(drawing_number),
        client=str(client), out_dir=q_dir, job_stem=stem)

    return {
        "override_xlsx": str(override_xlsx),
        "quote_html": quote_path,
        "figures": figures,
        "job_stem": stem,
    }


def main() -> None:
    """Run the override loop from the command line, so the feature is usable before the portal
    button exists: an estimator (or a PowerShell wrapper) points it at their amended sheet and
    the three fields, and it writes the manual-override record and the regenerated client quote.

        python src/client_quote_regen.py --workbook edited.xlsx --units 180 \\
               --drawing 10575-02 --client Dyson
    """
    import argparse
    ap = argparse.ArgumentParser(
        description="Regenerate a client quote from an estimator's amended workbook.")
    ap.add_argument("--workbook", required=True, help="Path to the amended (estimator-edited) xlsx")
    ap.add_argument("--units", required=True, type=int, help="Number of units for the order")
    ap.add_argument("--drawing", required=True, help="Drawing number for the quote header")
    ap.add_argument("--client", required=True, help="Client name (heads the quote, picks the logo)")
    ap.add_argument("--quote-dir", help="Override the quote destination (default: AISheets share)")
    ap.add_argument("--override-xlsx-dir", help="Override where the _MANUAL_OVERRIDE sheet is saved")
    ap.add_argument("--json", action="store_true",
                    help="Emit the result as a single JSON line on stdout (for the backend endpoint)")
    a = ap.parse_args()
    res = run_estimator_override(
        a.workbook, units=a.units, drawing_number=a.drawing, client=a.client,
        quote_dir=a.quote_dir, override_xlsx_dir=a.override_xlsx_dir)
    if a.json:
        import json as _json
        print(_json.dumps({
            "manual_override": True,
            "override_xlsx": res["override_xlsx"],
            "quote_html": res["quote_html"],
            "job_stem": res["job_stem"],
            "price": res["figures"]["price"],
            "price_source": res["figures"]["source_label"],
            "sell_price": res["figures"].get("sell_price"),
            "unit_cost": res["figures"].get("unit_cost"),
        }))
    else:
        print(f"  price read:    £{res['figures']['price']:.2f} "
              f"({res['figures']['source_label']})")
        print(f"  override sheet: {res['override_xlsx']}")
        print(f"  client quote:   {res['quote_html']}")


if __name__ == "__main__":
    main()
