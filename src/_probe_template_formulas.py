# -*- coding: utf-8 -*-
"""READ-ONLY. The layout probe found block headers but NO formulas in cols A-M — the cost columns
run further right (sheet goes to AD/col 30). Before widening the steel block we MUST see the
formulas: the per-part cost formulas in the steel rows, and especially the 'Total Material Cost'
(row 59) sum range — if we add rows, that range must expand or new parts won't count.

Scans the FULL column range (1..30) for formulas in the key rows, and specifically dumps:
  - steel data rows 38-48 (per-part cost formulas — the template to copy into new rows)
  - Other Sheet 51-58
  - Total Material Cost row 59 (the sum range that must expand)
  - anything referencing rows 38-59 from elsewhere (labour, totals, qty ladder)

Run from C:\\ClaudeVision\\src :
  C:\\ClaudeVision\\.venv\\Scripts\\python.exe _probe_template_formulas.py
"""
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

CANDIDATES = [
    r"C:\ClaudeVision\input\spreadsheets\EmptyEstimating\Blank Estimate Sheet 2026.xlsx",
]
tpl = None
for c in CANDIDATES:
    if Path(c).exists(): tpl = Path(c); break
if tpl is None:
    hits = list(Path(r"C:\ClaudeVision\input").rglob("*Blank*Estimat*.xlsx"))
    if hits: tpl = hits[0]
print(f"Template: {tpl}\n")
wb = openpyxl.load_workbook(tpl, data_only=False)
ws = wb["Estimate"]

def dump_row_formulas(r, tag):
    cells = []
    for c in range(1, 31):
        v = ws.cell(r, c).value
        if isinstance(v, str) and v.startswith("="):
            cells.append(f"{get_column_letter(c)}{r}={v}")
    if cells:
        print(f"  [{tag}] r{r}:")
        for x in cells: print(f"        {x}")
    else:
        print(f"  [{tag}] r{r}: (no formulas)")

print("=== STEEL data rows 38-48 (per-part cost formulas to replicate) ===")
for r in range(38, 49): dump_row_formulas(r, "steel")

print("\n=== OTHER SHEET rows 51-58 ===")
for r in range(51, 59): dump_row_formulas(r, "other")

print("\n=== TOTAL MATERIAL COST row 59 (the sum range that must expand) ===")
for r in (59, 60): dump_row_formulas(r, "totalmat")

print("\n=== any formula ANYWHERE (rows 1-159) that references steel rows 38-48 or 59 ===")
import re
for r in range(1, 160):
    for c in range(1, 31):
        v = ws.cell(r, c).value
        if isinstance(v, str) and v.startswith("="):
            # does it reference rows 38..48 or 59?
            refs = re.findall(r"[A-Z]{1,3}(\d{1,3})", v)
            if any(38 <= int(x) <= 48 or int(x)==59 for x in refs):
                print(f"   {get_column_letter(c)}{r} = {v[:70]}")

print("\nVERDICT: the Total Material sum range + the per-steel-row cost formula pattern tell us exactly")
print("what to replicate/expand when we insert rows. Then widen = insert rows 49+, copy steel formula")
print("pattern down, expand the material sum, shift config for Other Sheet/Labour/rollup blocks.")
