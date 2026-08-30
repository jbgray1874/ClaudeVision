r"""READ-ONLY. Parity must read Excel-computed cells from the populated .xlsx, but openpyxl
reads formulas as None until something calculates them. Determine what calc engine is available:
  1) Excel via win32com (COM) — best, uses real Excel
  2) LibreOffice headless (soffice) — cross-platform recompute
  3) Does the populated .xlsx ALREADY have cached values? (maybe wb_populate triggers calc)
  4) Does estimate_full_parity_report already support read_via_excel? (we saw a flag)
Report what exists so the fix uses the right mechanism. No edits."""
import os, glob, subprocess, shutil

print("="*66); print("1 — Excel COM (win32com) available?"); print("="*66)
try:
    import win32com.client  # noqa
    print("  win32com IMPORTABLE — Excel COM likely usable")
    # is Excel actually installed / registerable?
    try:
        import win32com.client as w
        xl = w.Dispatch("Excel.Application")
        print("  Excel.Application Dispatch OK — Excel IS installed and COM-callable")
        xl.Quit()
    except Exception as e:
        print(f"  Excel Dispatch failed: {type(e).__name__}: {str(e)[:60]}")
except Exception as e:
    print(f"  win32com NOT importable: {type(e).__name__}")

print("\n"+"="*66); print("2 — LibreOffice headless (soffice) available?"); print("="*66)
for cand in [r"C:\Program Files\LibreOffice\program\soffice.exe",
             r"C:\Program Files (x86)\LibreOffice\program\soffice.exe", "soffice", "libreoffice"]:
    found = shutil.which(cand) if not os.path.isabs(cand) else (cand if os.path.exists(cand) else None)
    if found:
        print(f"  FOUND: {found}")
        break
else:
    print("  LibreOffice not found in usual locations")

print("\n"+"="*66); print("3 — does the fresh populated .xlsx have cached values already?"); print("="*66)
xlsx=r"C:\ClaudeVision\output\estimates\1282 - Milwaukee Wall Bay_20260720_122217.xlsx"
if os.path.exists(xlsx):
    try:
        import openpyxl
        wb=openpyxl.load_workbook(xlsx, data_only=True)
        ws=wb["Estimate"] if "Estimate" in wb.sheetnames else wb.active
        # check the key cells — M92 material, unit cost, labour
        # find 'Total Material Cost' / 'Total Unit Cost Price' rows and read the value to their right
        def scan(label):
            for r in range(1, ws.max_row+1):
                for c in range(1, min(ws.max_column,16)+1):
                    v=ws.cell(r,c).value
                    if isinstance(v,str) and label.lower() in v.lower():
                        # value = rightmost numeric on this row
                        nums=[ws.cell(r,cc).value for cc in range(1,ws.max_column+1) if isinstance(ws.cell(r,cc).value,(int,float))]
                        return (r, nums[-1] if nums else None)
            return (None,None)
        for lab in ["Total Material Cost","Total Unit Cost","Total Labour Cost"]:
            r,v=scan(lab)
            print(f"  '{lab}': row {r}, cached value = {v}")
        print("\n  If cached values are None -> file NOT calculated, need a calc engine.")
        print("  If they show 133.45/189.01/42.33 -> file IS calculated, openpyxl can read directly!")
    except Exception as e:
        print(f"  openpyxl read failed: {e}")

print("\n"+"="*66); print("4 — does estimate_full_parity_report support read_via_excel?"); print("="*66)
p=r"C:\ClaudeVision\src\estimate_full_parity_report.py"
src=open(p,encoding="utf-8",errors="replace").read()
import re
for i,ln in enumerate(src.splitlines()):
    if re.search(r"read_via_excel|win32com|xlwings|soffice|libreoffice|calculate|Dispatch|recalc", ln, re.I):
        print(f"  {i+1}: {ln.strip()[:96]}")
