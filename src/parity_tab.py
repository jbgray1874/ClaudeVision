"""
parity_tab.py  —  adds a "Parity" sheet to the estimate workbook.

This is the JR-style comparison: AI estimate vs manual workbook, side by side,
with structural checks (missing lines, wrong part numbers, description/qty
mismatches) as well as financial variances.  It replaces the HTML parity report
as the primary comparison tool — everything in the workbook, nothing separate.

Public entry point:
    add_parity_sheet(wb, summary, manual_xlsx_path=None)
        -> worksheet or None

If no manual_xlsx_path is supplied, the tab still appears but the manual columns
show "— awaiting manual comparison —".  The AI columns are always populated.
"""
from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import re

try:
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
    _HAS_OPX = True
except Exception:
    _HAS_OPX = False

# ----------------------------------------------------------------- colours ---
_C = {
    "header_dark":  "1F3864",
    "header_mid":   "2E5FA3",
    "section":      "D9E1F2",
    "match":        "E2EFDA",   # green
    "warn":         "FFEB9C",   # amber
    "fail":         "FFC7CE",   # red
    "missing":      "F4CCCC",   # red-ish (line in manual, absent in AI)
    "extra":        "FCE5CD",   # orange (line in AI, absent in manual)
    "white":        "FFFFFF",
    "light_grey":   "F2F2F2",
}

def _fill(hex_):
    return PatternFill("solid", fgColor=hex_)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)

def _border():
    t = Side(style="thin", color="BFBFBF")
    return Border(left=t, right=t, top=t, bottom=t)

_WRAP = Alignment(wrap_text=True, vertical="top")
_MID  = Alignment(horizontal="center", vertical="top")
_PCT  = '0.0"%"'
_GBP  = '£#,##0.00'


# --------------------------------------------------------------- helpers ----
def _safe_float(v) -> Optional[float]:
    try:
        f = float(v)
        return f if f == f else None   # NaN guard
    except (TypeError, ValueError):
        return None

def _norm_code(s) -> str:
    return re.sub(r"\s+", "", str(s or "")).upper()

def _pct(ai, manual) -> Optional[float]:
    if ai is None or manual is None or manual == 0:
        return None
    return round(100.0 * (ai - manual) / manual, 1)

def _rag(pct) -> str:
    if pct is None:
        return "no_data"
    if abs(pct) <= 3.0:
        return "match"
    if abs(pct) <= 10.0:
        return "warn"
    return "fail"


# ------------------------------------------------------- manual xlsx read ---
def _load_manual_parts(manual_xlsx_path: str) -> Dict[str, Dict[str, Any]]:
    """
    Read the manual estimate workbook and extract per-part lines.
    Returns {normalised_part_number: {description, qty, material_gbp, labour_gbp, total_gbp}}.
    Tolerant — returns {} if anything goes wrong.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(manual_xlsx_path, data_only=True)
        # Try the first sheet named 'Estimate', 'Sheet Steel', 'Labour', or just sheet 1
        ws = None
        for name in wb.sheetnames:
            if any(k in name.upper() for k in ("ESTIMATE", "SHEET STEEL", "LABOUR", "BOM")):
                ws = wb[name]; break
        if ws is None:
            ws = wb.worksheets[0]

        parts: Dict[str, Dict[str, Any]] = {}
        # Scan every row for a pattern: col contains a plausible part number
        pn_re = re.compile(r"^\d{3,}[-A-Z0-9]*$", re.I)
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            # Find the first cell that looks like a part number
            for i, c in enumerate(cells):
                if pn_re.match(c) and len(c) >= 4:
                    code = _norm_code(c)
                    # pull numeric values from remaining cells
                    nums = []
                    for cv in cells[i+1:]:
                        f = _safe_float(cv)
                        if f is not None and f > 0:
                            nums.append(f)
                    desc = ""
                    for cv in cells[i+1:]:
                        if cv and not cv.replace(".","").replace(",","").isdigit():
                            desc = cv[:80]; break
                    parts[code] = {
                        "description": desc,
                        "total_gbp": nums[0] if nums else None,
                        "source_col": i,
                    }
                    break
        return parts
    except Exception:
        return {}


# ------------------------------------------------------ build the sheet ----
def add_parity_sheet(
    wb,
    summary: Dict[str, Any],
    manual_xlsx_path: Optional[str] = None,
) -> Any:
    if not _HAS_OPX:
        return None

    est       = summary.get("estimate_summary") or {}
    parts_est = est.get("part_estimates") or []
    if not parts_est:
        return None

    # Load manual comparison if available
    manual_parts: Dict[str, Dict[str, Any]] = {}
    manual_loaded = False
    if manual_xlsx_path and Path(manual_xlsx_path).exists():
        manual_parts = _load_manual_parts(manual_xlsx_path)
        manual_loaded = bool(manual_parts)

    ws = wb.create_sheet("Parity")
    ws.sheet_view.showGridLines = False
    r = 1

    # ---- helpers ----
    def cell(row, col, val="", bold=False, fill_hex=None, fmt=None,
             align=None, color="000000", size=10):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color=color, size=size)
        if fill_hex:
            c.fill = _fill(fill_hex)
        if fmt:
            c.number_format = fmt
        if align:
            c.alignment = align
        c.border = _border()
        return c

    def hrow(row, texts, fill_hex=_C["header_dark"], color="FFFFFF"):
        for i, t in enumerate(texts, 1):
            cell(row, i, t, bold=True, fill_hex=fill_hex, color=color,
                 align=_MID, size=10)

    # ---- title ----
    job = summary.get("source_file", "estimate")
    ws.cell(row=r, column=1, value=f"PARITY REPORT  —  {Path(job).stem}").font = Font(bold=True, size=13)
    r += 1
    status_note = (f"Manual comparison: {Path(manual_xlsx_path).name}"
                   if manual_loaded else "Manual comparison: NOT LOADED — AI columns only")
    ws.cell(row=r, column=1, value=status_note).font = Font(italic=True, size=9, color="666666")
    r += 2

    # ================================================================
    # SECTION 1: STRUCTURAL CHECK
    # Which GA BOM lines appear in the AI output, and which are missing?
    # ================================================================
    ws.cell(row=r, column=1, value="1.  STRUCTURAL CHECK — part presence").font = Font(bold=True, size=11)
    ws.cell(row=r, column=1).fill = _fill(_C["section"])
    r += 1

    hrow(r, ["Part No", "Description", "GA Qty", "In AI output?",
             "Manual match?", "Description match?", "Qty match?", "Notes"])
    r += 1

    bom_rows = (summary.get("document_analysis") or {}).get("bom_rows") or []
    ai_codes = {_norm_code(p.get("part_number")): p for p in parts_est}
    manual_codes = set(manual_parts.keys())

    struct_issues: List[str] = []
    for brow in bom_rows:
        code   = _norm_code(brow.get("part_number"))
        desc   = str(brow.get("description") or "")
        qty    = brow.get("quantity") or "?"
        in_ai  = code in ai_codes
        in_man = code in manual_codes if manual_loaded else None

        ai_part   = ai_codes.get(code) or {}
        ai_desc   = str(ai_part.get("description") or "")
        ai_qty    = ai_part.get("quantity")
        desc_ok   = (desc[:30].lower() in ai_desc.lower() or ai_desc[:30].lower() in desc.lower()) if (in_ai and desc and ai_desc) else None
        qty_ok    = (str(ai_qty) == str(qty)) if in_ai else None

        ai_flag   = "✓" if in_ai  else "✗ MISSING"
        man_flag  = ("✓" if in_man else "✗ MISSING") if manual_loaded else "—"
        desc_flag = ("✓" if desc_ok else "⚠ check") if desc_ok is not None else "—"
        qty_flag  = ("✓" if qty_ok  else f"⚠ AI={ai_qty}") if qty_ok is not None else "—"

        row_fill = (_C["fail"] if not in_ai else
                    (_C["warn"] if desc_ok is False or qty_ok is False else _C["match"]))
        for col, val in enumerate([code, desc, qty, ai_flag, man_flag, desc_flag, qty_flag,
                                    "" if in_ai else "Part absent from AI output — check extraction"], 1):
            cell(r, col, val, fill_hex=row_fill)
        if not in_ai:
            struct_issues.append(f"{code} absent from AI output")
        r += 1

    if struct_issues:
        ws.cell(row=r, column=1, value=f"⚠  {len(struct_issues)} part(s) in GA BOM not found in AI output: "
                                        + ", ".join(struct_issues[:6])).font = Font(italic=True, color="CC0000", size=9)
        r += 1
    r += 1

    # ================================================================
    # SECTION 2: FINANCIAL COMPARISON
    # AI unit cost vs manual unit cost, per part, with RAG
    # ================================================================
    ws.cell(row=r, column=1, value="2.  FINANCIAL COMPARISON — unit cost per part").font = Font(bold=True, size=11)
    ws.cell(row=r, column=1).fill = _fill(_C["section"])
    r += 1

    hrow(r, ["Part No", "Description", "Qty",
             "AI Mat £", "AI Lab £", "AI Unit £", "AI Ext £",
             "Manual Unit £", "Variance %", "RAG", "Price source", "Flags"])
    r += 1

    ai_total = man_total = 0.0
    warn_lines: List[str] = []
    fail_lines: List[str] = []

    for pe in parts_est:
        pn      = str(pe.get("part_number") or "")
        code    = _norm_code(pn)
        desc    = str(pe.get("description") or "")[:60]
        qty     = pe.get("quantity") or 1
        cb      = pe.get("cost_breakdown") or {}
        mat_e   = pe.get("material_estimate") or {}
        lab_e   = pe.get("labour_estimate") or {}

        ai_mat  = _safe_float(pe.get("material_cost_gbp") or mat_e.get("extended_material_cost_gbp")) or 0.0
        ai_lab  = _safe_float(pe.get("labour_cost_gbp")  or lab_e.get("total_labour_cost_gbp"))       or 0.0
        ai_unit = _safe_float(pe.get("unit_total_cost_gbp") or cb.get("unit_total_cost_gbp") or
                               pe.get("unit_estimate")) or (ai_mat + ai_lab)
        ai_ext  = _safe_float(cb.get("extended_total_cost_gbp") or pe.get("extended_estimate")) or (ai_unit * int(qty))
        ai_total += ai_ext

        man_row  = manual_parts.get(code) or {}
        man_unit = _safe_float(man_row.get("total_gbp"))
        if man_unit:
            man_total += man_unit * int(qty)

        var_pct = _pct(ai_unit, man_unit)
        rag     = _rag(var_pct)
        row_fill = {"match": _C["match"], "warn": _C["warn"], "fail": _C["fail"],
                    "no_data": _C["white"]}[rag]

        ps      = pe.get("price_source") or {}
        ps_lbl  = str(ps.get("source") or ps.get("source_type") or "—")[:30]
        flags   = "; ".join(str(f) for f in (pe.get("risk_flags") or [])[:3])[:60]

        vals = [pn, desc, qty,
                ai_mat, ai_lab, ai_unit, ai_ext,
                man_unit if man_unit is not None else ("— await" if manual_loaded else "—"),
                f"{var_pct:+.1f}%" if var_pct is not None else "—",
                {"match":"✓ Match","warn":"⚠ Review","fail":"✗ Fail","no_data":"—"}[rag],
                ps_lbl, flags]
        fmts = [None, None, None, _GBP, _GBP, _GBP, _GBP, _GBP, None, None, None, None]
        for col, (val, fmt) in enumerate(zip(vals, fmts), 1):
            c = cell(r, col, val, fill_hex=row_fill, fmt=fmt)
        if rag == "warn":
            warn_lines.append(pn)
        elif rag == "fail":
            fail_lines.append(pn)
        r += 1

    # totals row
    for col, val in enumerate(
        ["TOTAL", "", "", "", "", "", ai_total,
         man_total if man_total else None, "", "", "", ""], 1):
        cell(r, col, val, bold=True, fill_hex=_C["section"],
             fmt=_GBP if col in (7, 8) else None)
    if man_total and ai_total:
        var_tot = _pct(ai_total, man_total)
        cell(r, 9, f"{var_tot:+.1f}%" if var_tot else "—", bold=True, fill_hex=_C["section"])
    r += 2

    # ================================================================
    # SECTION 3: SUMMARY + ACTIONS
    # ================================================================
    ws.cell(row=r, column=1, value="3.  SUMMARY & ACTIONS").font = Font(bold=True, size=11)
    ws.cell(row=r, column=1).fill = _fill(_C["section"])
    r += 1

    def note(text, color="000000"):
        ws.cell(row=r, column=1, value=text).font = Font(size=10, color=color)

    note(f"Structural: {len(struct_issues)} GA BOM line(s) absent from AI output" if struct_issues
         else "Structural: all GA BOM lines present in AI output  ✓", "CC0000" if struct_issues else "006600")
    r += 1
    if fail_lines:
        note(f"Financial FAIL (>10% variance): {', '.join(fail_lines[:8])}", "CC0000"); r += 1
    if warn_lines:
        note(f"Financial WARN (3-10% variance): {', '.join(warn_lines[:8])}", "996600"); r += 1
    if not manual_loaded:
        note("Manual comparison not loaded — drop a manual estimate xlsx into the job folder "
             "and re-run to populate the Manual Unit £ and Variance columns.", "666666"); r += 1
    r += 1
    note("How to read this tab:", "444444"); r += 1
    note("  Section 1: structural — is every GA BOM line present in the AI output, with the right description and quantity?", "444444"); r += 1
    note("  Section 2: financial — AI unit cost vs manual unit cost per part, RAG-rated.  ✓ = within 3%  ⚠ = 3-10%  ✗ = >10%.", "444444"); r += 1
    note("  Price source column shows where each AI price came from (catalogue / RAG / web / AI).", "444444"); r += 1

    # column widths
    for col, w in enumerate([14, 36, 6, 10, 10, 10, 10, 12, 10, 10, 24, 40], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return ws
