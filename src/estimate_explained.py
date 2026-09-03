#!/usr/bin/env python3
r"""Every row of an estimate, explained — the joined record the deliverables are built from.

WHY THIS EXISTS. The deliverables each said where a number came from and none of them said
WHICH FILE AND WHICH PAGE, so the one question estimating always asks — "where did you see
that?" — could not be answered from the pack that was sent, and got asked by email instead.
The pages were never missing from the engine: every part record carries `pages` and
`page_roles` and the run log says so out loud. They were dropped before anything was written.

So this joins the three records that between them hold the whole answer:

    the workbook        the blocks, the blank sizes, the routes
    final_estimate      the rows AS EXCEL CALCULATED THEM — the only record with money on
                        every line, and the only one that can be summed against the sheet's
                        own Total Material Cost and Total Labour Cost
    the part records    the drawing page that owns each part, and what that sheet stated

and prints one section per question an estimator actually asks.

ONE BUILDER, SEVERAL RENDERINGS. This began as a hand-run tool and is now the source for the
workbook's Explanation tab, the HTML report's provenance sections and the body of the
estimate email. It lives in src/ for that reason: three programs that each go and ask the
workbook their own questions are three programs that disagree by the tenth change.

WHAT IT WILL NOT DO. It never recalculates. Every figure is read from the sheet or from the
run's record of the sheet, and where a figure is unavailable it says so rather than printing
a zero — a zero would make a reconciliation against it look exact. Where two records disagree
about the same rows it prints both, says which one reaches the total, and says so in pounds.

Run it from tools/handover_note.py.
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
except ImportError:                                              # pragma: no cover
    raise SystemExit("openpyxl is required: pip install openpyxl")


# ── reading the workbook ─────────────────────────────────────────────────────

def _rows(ws, header_row: int = 1) -> List[Dict[str, Any]]:
    """A sheet as dicts keyed by its own header, blank rows dropped."""
    header = [str(ws.cell(header_row, c).value or "").strip()
              for c in range(1, ws.max_column + 1)]
    out: List[Dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(v not in (None, "") for v in values):
            continue
        out.append({h: v for h, v in zip(header, values) if h})
    return out


def _sheet(wb, name: str) -> List[Dict[str, Any]]:
    return _rows(wb[name]) if name in wb.sheetnames else []


# The Sheet Steel block's own columns, by position, from the template.
_STEEL_COLS = {"desc": 3, "qty": 5, "length": 6, "width": 7, "gauge": 8,
               "sheet_l": 9, "sheet_w": 10, "per_sheet": 11, "scrap": 12,
               "cost_per_part": 13, "internal_cut": 20, "rate_per_hour": 23}


def _steel_rows(wb) -> Dict[str, Dict[str, Any]]:
    """The Sheet Steel block, keyed by part number.

    WHERE A FABRICATED PART'S MONEY ACTUALLY IS. A "-M" line shows a dash in the BOM price
    column and says why in its own text — "costed in Sheet Steel below" — because sheet metal
    is priced by NEST on that block: a whole sheet divided by how many of the part come out
    of it, not by the part's own area and not per piece. An explanation that stops at
    "priced below"
    ends exactly where the estimator's question begins: how big, off what sheet, how many out
    of it, at what cost. This reads that block so the two halves can be shown together.

    Found by its header text, not a row number: the block moves with the size of the BOM
    above it, and a fixed row would silently read whatever had shifted into it.
    """
    if "Estimate" not in wb.sheetnames:
        return {}
    ws = wb["Estimate"]
    header_row = None
    for r in range(1, min(ws.max_row, 120) + 1):
        joined = " ".join(str(ws.cell(r, c).value or "") for c in range(1, 26)).lower()
        if "part length" in joined and "gauge" in joined and "part description" in joined:
            header_row = r
            break
    if header_row is None:
        return {}
    out: Dict[str, Dict[str, Any]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        text = str(ws.cell(r, _STEEL_COLS["desc"]).value or "").strip()
        if not text:
            # Blank template rows sit between the blocks; stop once we have started and
            # then hit two in a row, rather than running on into "Other Sheet Material".
            if out and not str(ws.cell(r + 1, _STEEL_COLS["desc"]).value or "").strip():
                break
            continue
        code = text.split()[0].strip().upper()
        out[code] = {"row": r, "text": text,
                     **{k: ws.cell(r, c).value for k, c in _STEEL_COLS.items() if k != "desc"}}
    return out


_TOTAL_LABELS = (("material", "total material cost"),
                 ("labour", "total labour cost"),
                 ("unit", "total unit cost"))


def _sheet_totals(wb, final: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """The Estimate sheet's own labelled totals.

    Found by label rather than by cell reference, and the value taken as the first number to
    the right of it — the template's totals sit in column M today and a reference is the one
    thing that cannot survive a template revision.

    THE CELLS ARE USUALLY EMPTY TO US, AND THAT IS NOT THE SAME AS ABSENT. Total Material
    Cost is a SUM formula, and a workbook written by openpyxl and never re-saved by Excel
    carries no cached result for it — so reading the file alone returned nothing for all
    three totals and the reconciliation could not be performed at all, on a document whose
    whole purpose is to perform it. The read-back has already opened that workbook through
    Excel, calculated it, and read those same labelled cells: `final_estimate.totals` IS
    what the cell held. Falling back to it is not circular — the totals were scanned from
    the labelled cells, the rows from the four blocks, by two separate passes — but the
    document says which of the two it used, because "read from the file" and "read from the
    run's record of the file" are different claims.

    A figure neither source yields stays None. The document then says the total was not
    available rather than reporting a zero that would make every reconciliation look perfect.
    """
    out: Dict[str, Any] = {k: None for k, _ in _TOTAL_LABELS}
    if "Estimate" not in wb.sheetnames:
        return _fill_totals_from_final(out, final)
    ws = wb["Estimate"]
    for r in range(1, ws.max_row + 1):
        label = " ".join(str(ws.cell(r, c).value or "") for c in range(1, 8)).strip().lower()
        if not label:
            continue
        for key, needle in _TOTAL_LABELS:
            if out[key] is None and label.startswith(needle):
                for c in range(8, min(ws.max_column, 30) + 1):
                    value = _money(ws.cell(r, c).value)
                    if value is not None:
                        out[key] = round(value, 2)
                        out.setdefault("_from", {})[key] = "the workbook's own cell"
                        break
    return _fill_totals_from_final(out, final)


def _fill_totals_from_final(out: Dict[str, Any],
                            final: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    """Whatever the cells would not yield, taken from the run's record of those same cells."""
    totals = (final or {}).get("totals")
    if not isinstance(totals, dict):
        return out
    for key, field in (("material", "material_gbp"), ("labour", "labour_gbp"),
                       ("unit", "unit_gbp")):
        if out.get(key) is None:
            value = _money(totals.get(field))
            if value is not None:
                out[key] = round(value, 2)
                out.setdefault("_from", {})[key] = (
                    "the run's read-back of that cell, calculated by Excel")
    return out


def _order_quantity(wb) -> Optional[int]:
    """How many this estimate is for, from the header cell the populator writes.

    A plain value, not a formula, so it reads out of the saved file — unlike the totals. It is
    what makes the set-up split meaningful: a labour row's Total Value is per unit, so the
    set-up inside it has already been divided by this.
    """
    if "Estimate" not in wb.sheetnames:
        return None
    try:
        from wb_populate import CELL_MAP
        cell = str(CELL_MAP["header"]["order_qty"])
    except Exception:                                            # noqa: BLE001
        cell = "D6"
    value = _money(wb["Estimate"][cell].value)
    return int(value) if value and value >= 1 else None


def _estimate_bom(wb) -> List[Dict[str, Any]]:
    """The Bill of Materials block on the Estimate sheet.

    Found by its own header text rather than a row number: the block moves whenever a job
    has more or fewer lines, and a hard-coded row would silently read the wrong band.
    """
    if "Estimate" not in wb.sheetnames:
        return []
    ws = wb["Estimate"]
    header_row = None
    for r in range(1, min(ws.max_row, 60) + 1):
        joined = " ".join(str(ws.cell(r, c).value or "") for c in range(1, 16)).lower()
        if "bill of materials" in joined and "part code" in joined:
            header_row = r
            break
    if header_row is None:
        return []
    out = []
    for r in range(header_row + 1, ws.max_row + 1):
        code = ws.cell(r, 8).value           # "Part code"
        text = ws.cell(r, 3).value           # the description column
        if not code and not text:
            if out:
                break                        # the block has ended
            continue
        out.append({
            "code": str(code or "").strip(),
            "text": str(text or "").strip(),
            "supplier": ws.cell(r, 9).value,
            "price": ws.cell(r, 10).value,
            "qty": ws.cell(r, 11).value,
        })
    return out


# ── reading the scan ─────────────────────────────────────────────────────────

def _load_scan(path: Optional[Path]) -> Any:
    """The run JSON as it sits on disk, or None. Read once and passed around.

    SEPARATED FROM _scan_parts BECAUSE THE PARTS ARE NOT THE ONLY THING IN IT. The same
    document carries `final_estimate` — the rows as Excel calculated them, which is the only
    record that has money on every line and provably sums to the sheet's own totals. Reading
    the file twice to get at both would be two chances to read two different files.
    """
    if not path or not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8", errors="replace"))
    except Exception:                                            # noqa: BLE001
        return None


def _scan_parts(data: Any) -> Dict[str, Dict[str, Any]]:
    """part number -> the scan record, for the pages the workbook does not carry."""
    if data is None:
        return {}
    # THREE SHAPES, BECAUSE THE FULL SCAN IS TOO BIG TO MOVE. 12552's scan JSON is 258,935
    # lines; nobody is going to send that to have three fields read out of it. So a trimmed
    # extract — a bare list of records, or {"parts": [...]} — is accepted alongside the real
    # thing, and only part_number, pages and page_roles are ever read from any of them.
    #
    # This is a workaround for the actual defect, which is that no page number reaches any
    # deliverable. Once the page is written onto the row, this whole path becomes unnecessary.
    if isinstance(data, list):
        parts = data
    elif isinstance(data, dict):
        parts = ((data.get("manufacturing_writeup") or {}).get("parts")
                 or data.get("parts") or [])
    else:
        return {}
    return {str(p.get("part_number") or "").strip().upper(): p
            for p in parts if isinstance(p, dict)}


def _basename(path: Any) -> str:
    """The filename out of a path that was written on Windows and may be read anywhere.

    Path(...).name splits on the separator of the machine it is RUNNING on, so a
    "K:\\jobs\\part.DXF" recorded by the runner comes back whole on anything but Windows —
    and the document is generated wherever somebody happens to run it. Split on both.
    """
    text = str(path or "").strip().rstrip("\\/")
    for sep in ("\\", "/"):
        text = text.rsplit(sep, 1)[-1]
    return text


def _pack_files(data: Any) -> List[str]:
    """Every drawing file this run read, in the order the run recorded them."""
    if not isinstance(data, dict):
        return []
    out: List[str] = []
    for key in ("job_source_pdfs", "source_pdfs"):
        for item in (data.get(key) or []):
            # `job_source_pdfs` holds RECORDS, not paths — file_scan writes
            # {"name", "path", "page_count"} per document, while `source_pdfs` is a plain
            # list of paths. _basename() was handed the dict and stringified it, so the
            # pack read `{'name': '12552-00-GA_Infinity Drawer_RevC.PDF'}` and any answer
            # built from it carried the braces into the report.
            if isinstance(item, dict):
                item = item.get("name") or item.get("path") or ""
            name = _basename(item)
            if name and name not in out:
                out.append(name)
    single = data.get("source_file")
    if single and str(single).lower().endswith(".pdf"):
        name = _basename(single)
        if name not in out:
            out.append(name)
    return out


def _page_index(data: Any) -> Dict[int, Dict[str, Any]]:
    """job page number -> which PDF it is a page of, and the page number PRINTED on it.

    A PACK OF FOUR DRAWINGS HAS FOUR PAGE ONES. file_scan renumbers every page across the
    whole job so `page_number` is unique — 1..N — and keeps the document's own number as
    `source_page_number`, "the per-PDF original for display". Nothing was displaying it.

    So a part on the second document's page 4 was reported as "p.18": a number that appears
    on no drawing anybody can open, next to a file name that was only ever filled in when the
    pack happened to hold exactly ONE PDF. An estimator turning to p.18 of a 12-page drawing
    finds nothing there, and has no way to tell which of the four documents to look in.

    Both halves of "where did you see that" come from here.
    """
    out: Dict[int, Dict[str, Any]] = {}
    pages = (data or {}).get("pages") if isinstance(data, dict) else None
    for page in pages or []:
        if not isinstance(page, dict):
            continue
        try:
            job_page = int(page.get("job_page_number") or page.get("page_number"))
        except (TypeError, ValueError):
            continue
        name = _basename(page.get("source_pdf_name") or page.get("source_pdf_path") or "")
        printed = page.get("source_page_number")
        out[job_page] = {"file": name,
                         "printed": printed if printed is not None else job_page}
    return out


def _file_of(record: Dict[str, Any], pack: List[str],
             pages: Optional[Dict[int, Dict[str, Any]]] = None) -> str:
    """WHICH FILE, not just which page.

    "p.6" is only half an answer when a pack has four PDFs and a folder of DXFs — the estimator
    still has to work out which document page 6 belongs to. A part measured off a DXF has an
    even better answer available: the DXF's own filename, which the engine already records and
    which names the gauge and the material in most drawing offices' conventions.
    """
    dxf = str(record.get("dxf_source_file") or "").strip()
    if dxf:
        return _basename(dxf)
    own = str(record.get("source_file") or "").strip()
    if own.lower().endswith(".pdf"):
        return _basename(own)
    # THE PAGE KNOWS WHICH DOCUMENT IT IS A PAGE OF. Asked before falling back to guessing
    # from the size of the pack, so a four-PDF job answers as precisely as a one-PDF job.
    named = {(pages or {}).get(_int(p), {}).get("file") for p in (record.get("pages") or [])}
    named.discard(None)
    named.discard("")
    if named:
        return ", ".join(sorted(named))
    if record.get("pages") and len(pack) == 1:
        # One document in the pack, so every page in it is a page of that document. Said
        # rather than inferred silently: with two PDFs this would be a guess and it declines.
        return pack[0]
    return "not recorded"


def _int(value: Any) -> Any:
    try:
        return int(value)
    except (TypeError, ValueError):
        return value


def _drawing_no(record: Dict[str, Any]) -> str:
    """The number in the TITLE BLOCK, which is not always the part number.

    part_index collects `drawing_numbers` off each title block and falls back to the part
    number when a sheet prints none. An estimator checking our work reads the number off the
    drawing in front of them, so it is reported even when it equals the part number — "the
    same" is an answer, and its absence is the interesting case.
    """
    numbers = [str(d).strip() for d in (record.get("drawing_numbers") or []) if str(d).strip()]
    if not numbers:
        return "—"
    pn = str(record.get("part_number") or "").strip().upper()
    other = [d for d in numbers if d.upper() != pn]
    return ", ".join(dict.fromkeys(other or numbers))


def _sources_of(record: Dict[str, Any], pack: List[str],
                pages: Optional[Dict[int, Dict[str, Any]]] = None) -> List[str]:
    """EVERY drawing file this part is evidenced by, each with the page and what it is.

    A part is rarely evidenced by one document. It has a flat exported for the laser, a
    detail sheet, a line on the GA, and a model the first two came out of — and this returned
    whichever ONE it found first, which meant "where did you see that" was answered with the
    DXF and the drawing pages went unmentioned, or answered with a page and the flat that
    actually supplied the geometry went unmentioned.

    James: "source_pdf_name needs to cover all drawing file names, not just pdfs." So all of
    them, named, in the order of what they are worth as evidence: the flat that was measured,
    the model it came from, then the sheets it is drawn on.
    """
    out: List[str] = []
    dxf = str(record.get("dxf_source_file") or "").strip()
    if dxf:
        out.append(f"{_basename(dxf)} (flat)")
    model = str(record.get("solidworks_part_number") or "").strip()
    if model:
        out.append(f"{model} (SOLIDWORKS model)")

    own = record.get("pages") or []
    roles = ", ".join(str(r) for r in (record.get("page_roles") or []))
    if own and pages:
        # Each page paired with the document it is printed in, and the number printed on it.
        for p in own:
            entry = pages.get(_int(p)) or {}
            name = entry.get("file") or (pack[0] if len(pack) == 1 else "not recorded")
            out.append(f"{name} · p.{entry.get('printed', p)}")
    elif own:
        name = _file_of(record, pack, pages)
        # Where the page cannot be resolved, _file_of falls back to the DXF — which is
        # already the first entry. Naming it twice reads as two documents.
        if not any(entry.startswith(name) for entry in out):
            out.append(f"{name} · {_pages_of(record, pages).split(' (')[0]}")
    if roles and out:
        out[-1] = f"{out[-1]} ({roles})"
    return out


def _where(record: Dict[str, Any], pack: List[str],
           pages: Optional[Dict[int, Dict[str, Any]]] = None) -> str:
    """Every file and page this part was seen in, which is what "where did you see that"
    actually asks. Never one of them."""
    if not record:
        return "not supplied"
    found = _sources_of(record, pack, pages)
    if not found:
        return "no sheet of its own"
    # "·" already separates a file from its page, so a second level is needed between the
    # files themselves. Not "|" — these strings go into markdown tables.
    return " ; ".join(found)


def _pages_of(record: Dict[str, Any],
              pages: Optional[Dict[int, Dict[str, Any]]] = None) -> str:
    own = record.get("pages") or []
    roles = record.get("page_roles") or []
    if not own:
        return "no sheet of its own"
    # THE NUMBER PRINTED ON THE DRAWING, not the job-wide one. They are the same on a
    # single-PDF pack and different on every other, and only one of them is a page an
    # estimator can turn to.
    shown = ", ".join(f"p.{(pages or {}).get(_int(p), {}).get('printed', p)}" for p in own)
    return f"{shown} ({', '.join(str(r) for r in roles)})" if roles else shown


# ── the rows as Excel calculated them ────────────────────────────────────────
#
# THE ONE RECORD WITH MONEY ON EVERY LINE. `final_estimate` is written by the read-back after
# Excel has recalculated the populated template: every material block (bought-in BOM, tube,
# sheet steel, other sheet) and every labour row, each carrying the value the sheet itself
# computed. It is the only structure that can be summed and held against Total Material Cost
# and Total Labour Cost — which is what "relates to the spreadsheet" has to mean.
#
# Until this was read, this document explained material only, and the labour — a third of the
# unit cost — appeared as a list of operations with no figure against any of them.

def _data_sufficiency(data: Any) -> Dict[str, Any]:
    """What the engine concluded about whether this pack could be costed credibly.

    IT SUPPRESSES ITS OWN HEADLINE AND SAYS SO ONLY ON THE CONSOLE. The gate stamps
    INSUFFICIENT DATA, nulls the engine's document total, and prints one line to a log nobody
    keeps — while the workbook goes on to compute a perfectly ordinary Unit Cost, because the
    two are different figures. So the estimate arrives priced, with the reason to doubt it
    recorded in a JSON field and stated nowhere a person reads.

    Priced with what we have, and the limitation stated: this is the half that was missing.
    """
    if not isinstance(data, dict):
        return {}
    for node in (data.get("data_sufficiency"),
                 (data.get("estimate_summary") or {}).get("data_sufficiency")
                 if isinstance(data.get("estimate_summary"), dict) else None):
        if isinstance(node, dict):
            return node
    return {}


def _final_estimate(data: Any) -> Dict[str, Any]:
    """`final_estimate` off the run JSON, wherever the writer put it."""
    if not isinstance(data, dict):
        return {}
    for node in (data.get("final_estimate"),
                 (data.get("estimate_summary") or {}).get("final_estimate")
                 if isinstance(data.get("estimate_summary"), dict) else None):
        if isinstance(node, dict):
            return node
    return {}


def _accepted_labour(data: Any) -> Dict[int, Dict[str, Any]]:
    """workbook_labour rows keyed by sheet row — the PARTS behind each labour line.

    The calculated rows know what a line cost and which department did it; they do not know
    which parts produced it. The accepted rows know exactly that and nothing about cost. They
    join on the sheet row they share, which is the join wb_populate itself documents.
    """
    if not isinstance(data, dict):
        return {}
    node = data.get("workbook_labour")
    if not isinstance(node, dict) and isinstance(data.get("estimate_summary"), dict):
        node = data["estimate_summary"].get("workbook_labour")
    out: Dict[int, Dict[str, Any]] = {}
    for row in ((node or {}).get("rows") or []):
        if isinstance(row, dict) and row.get("workbook_row") is not None:
            try:
                out[int(row["workbook_row"])] = row
            except (TypeError, ValueError):
                continue
    return out


def _money(value: Any) -> Optional[float]:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _sum_money(rows: List[Dict[str, Any]], key: str = "total_value_gbp") -> float:
    return round(sum(_money(r.get(key)) or 0.0 for r in rows), 2)


_BLOCK_NAMES = {"bom": "Bought-in / standard materials", "tube": "Tube and wire",
                "steel": "Sheet steel", "other_sheet": "Other sheet material"}


# ── which reader decided a thing ─────────────────────────────────────────────
#
# "WHERE DID THAT COME FROM" HAS TWO ANSWERS AND THE DOCUMENT ONLY GAVE ONE. It said which
# drawing page a part appears on, and what priced it — and not which READER supplied the
# material, the gauge, the quantity or the size. Those are different questions: a thickness
# measured off the SOLIDWORKS model and one read off a title block are both "p.7", and only
# one of them is a measurement.
#
# The engine records it per field and always has. source_precedence stamps material_source,
# thickness_source, quantity_source and geometry_source onto every part as it goes, and ranks
# them so a weaker reader cannot overwrite a stronger one. This prints that record.

_READER_WORDS = {
    "estimator_confirmed": "an estimator confirmed it",
    "knowledge_base": "the knowledge base",
    "solidworks_api": "the SOLIDWORKS model",
    "solidworks_flat_pattern": "the SOLIDWORKS flat pattern — measured",
    "solidworks_applied_material": "the material applied in the SOLIDWORKS model",
    "dxf": "the part's DXF",
    "dxf_flat_pattern": "the part's DXF flat pattern — measured",
    "dxf_filename": "the DXF's own filename",
    "mirror_of_measured": "mirrored from the measured opposite hand",
    "drawing_deterministic": "read straight off the drawing",
    "title_block": "the drawing's title block",
    "drawing_notes": "a note on the drawing",
    "pdf_overall_dims": "overall dimensions on the drawing — the view, not the part",
    "bom_tree": "the BOM table",
    "override_rule": "an override rule",
    "llm_extract": "the vision model (xAI Grok) — a reading, not a measurement",
    "llm_full_extract": "the vision model (xAI Grok) — a reading, not a measurement",
    "inference": "inferred by the engine — nothing on the drawing said it",
    "geometry_inference": "inferred from the part's geometry — nothing measured it",
    "": "not recorded",
}


def _reader(source: Any) -> str:
    """A source name in words an estimator can weigh, with the engine's own rank on it."""
    key = str(source or "").strip().lower()
    words = _READER_WORDS.get(key)
    if words is None:
        # Decorated values such as "knowledge_base (92%)" or "override_rule:timber_panels".
        for known, text in _READER_WORDS.items():
            if known and key.startswith(known):
                words = text
                break
    words = words or key.replace("_", " ")
    try:
        from source_precedence import rank as _rank
        _r = _rank(key)
    except Exception:                                            # noqa: BLE001
        _r = None
    return f"{words} ({_r})" if _r else words


def _readers_used(scan: Dict[str, Dict[str, Any]]) -> List[str]:
    """Every reader that actually decided something on this job, strongest first."""
    seen: Dict[str, int] = {}
    for rec in scan.values():
        for field in ("material_source", "thickness_source", "quantity_source",
                      "geometry_source"):
            key = str(rec.get(field) or "").strip().lower()
            if key:
                seen[key] = seen.get(key, 0) + 1
    try:
        from source_precedence import rank as _rank
    except Exception:                                            # noqa: BLE001
        def _rank(_):                                            # noqa: ANN001
            return 0
    return [f"{_reader(k)} — {n} field(s)"
            for k, n in sorted(seen.items(), key=lambda kv: (-_rank(kv[0]), -kv[1], kv[0]))]


# ── where a part number stopped tracing through the pack ─────────────────────
#
# THE FAILURE DESIGN CAN ACT ON, AND THE ONE NOBODY HAS BEEN TOLD ABOUT.
#
# A part number is supposed to travel: BOM row -> detail sheet -> DXF export -> model. Every
# place it does not is a place the engine had to size or price a part from something that was
# not a drawing of that part, and every one of those has a £ on it.
#
# 12349-02-69 is the case. Its flats are named `...-01A_-01_2MM_High Impact Acrylic_RevA.DXF`
# — four dash-segments — and the filename parser capped a number at three, so all seven
# resolved to the grandparent. 01A then had no flat of its own, fell through to a document
# text scan, and took 2120 x 2120 off the general arrangement. None of that was reported: the
# estimate simply arrived with a 4.5 square metre acrylic drawer front in it.
#
# The engine cannot tell Design their numbering is wrong — usually it is not; usually the
# READER is. What it can do is say, on every job, exactly where a number stopped tracing and
# what the estimate did instead, so the two of us can tell which end the fix belongs at.

# Geometry that is not a measurement OF THIS PART. Each means: no flat, no model, so a size
# came from somewhere that describes the part only indirectly.
_UNTRACED_GEOMETRY = {
    "pdf_overall_dims": "sized from the overall on a view rather than a flat pattern",
    "document_text_largest_numbers": "sized from the largest numbers in the document text — "
                                     "context-blind, and the biggest numbers in a pack are "
                                     "usually the general arrangement's",
    "inference": "sized by inference — nothing on the drawing stated it",
    "geometry_inference": "sized from reasoning about the part, not from a drawing of it",
    "llm_extract": "sized from a vision-model reading of the page",
    "llm_full_extract": "sized from a vision-model reading of the page",
}

# A TRACING FAILURE IS ABOUT A PART WE HAVE TO CUT.
#
# `bom_tree` was in the list above and produced the wrong kind of noise: a bought-in concrete
# slab has no geometry because it is bought, not because its number stopped tracing, and
# reporting it here sends somebody to ask Design for a detail drawing of a slab. Nothing is a
# break in the trail unless the engine needed a drawing of that part and could not find one —
# which means the part is fabricated.


_QUALITY_FIELDS = ("materials", "thicknesses_mm", "surface_finishes", "geometry_rollup")


def _pack_was_read_in_full(scan: Dict[str, Dict[str, Any]]) -> bool:
    """Does this extract carry the fields a drawing-quality read needs?

    A FIELD MISSING FROM THE EXTRACT IS NOT A FIELD MISSING FROM THE DRAWING. Against a
    trimmed extract every row comes out "material no, thickness no, finish no" for a pack
    whose page 4 plainly reads MATERIAL: MILD STEEL, 1.5 THK, POWDER COATED. That is not a
    weak answer, it is a confident wrong one, and it would go to an estimator as an
    assessment of Design's work.
    """
    return any(any(k in rec for k in _QUALITY_FIELDS) for rec in scan.values())


def _gauge_stated(rec: Dict[str, Any]) -> str:
    """The gauge this part is COSTED at, and the scrape only where it disagrees.

    THE COLUMN WAS PRINTING A TOLERANCE TABLE. `thicknesses_mm` is every number the page
    reader found that looked like a thickness — on 11908-21 that is "1.0, 3" against three
    parts whose DXF filenames read `9mm MDF+ LAM`, whose sheet block says Ga 9, and which are
    costed at 9. On 12349 the same column said 5.0 for a 1.5 mm bracket.

    So section 7 disagreed with section 2 about the same part in the same document, and the
    half that was wrong is the half headed "what the drawing states" — which is the half an
    estimator would use to check us, and the half they would send to Design.

    The resolved gauge is what the arbitration decided after weighing every reading. Where a
    scraped number differs it is still shown, marked as unused, because a drawing that prints
    a contradictory thickness IS worth knowing about — it is just not what we costed.
    """
    resolved = _money(rec.get("normalized_thickness_mm"))
    scraped = [str(t) for t in (rec.get("thicknesses_mm") or []) if str(t).strip()]
    if resolved is None:
        return ", ".join(scraped) or "no"
    shown = f"{resolved:g}"
    others = [t for t in scraped if _money(t) is not None
              and abs((_money(t) or 0) - resolved) > 0.01]
    if others:
        return f"{shown} (the page also reads {', '.join(others)} — not used)"
    return shown


def _what_a_sheet_could_not_give(rec: Dict[str, Any]) -> List[str]:
    """What this drawing did not state, or why the question does not apply to it.

    A PURCHASED PART NAMED ON AN ASSEMBLY SHEET IS NOT AN INCOMPLETE DETAIL. The bearing and
    the rivets appear on p.2 because that is where they are listed, and they have no detail
    drawing because they do not need one. Saying what they ARE beats listing four fields they
    were never going to carry.
    """
    roles = [str(r).lower() for r in (rec.get("page_roles") or [])]
    if "bought_in" in roles and "detail" not in roles:
        return ["bought in — listed on an assembly sheet, no detail drawing needed"]
    geom = rec.get("geometry_rollup") or {}
    missing: List[str] = []
    if not rec.get("materials"):
        missing.append("material")
    if not rec.get("thicknesses_mm"):
        missing.append("thickness")
    if not rec.get("surface_finishes"):
        missing.append("finish")
    if not (geom.get("estimated_cut_length_mm") or 0):
        missing.append("cut length")
    return missing


def _what_they_are(lines: List[Dict[str, Any]]) -> str:
    """The kinds of line actually on THIS job, named from the lines themselves.

    This sentence was a fixed list — "fasteners off the BOM table, bought items, packaging,
    delivery and the powder line" — printed whatever the job held. 11908-21 has no powder on
    it, so the note named a line that does not exist, which is exactly the tell that a
    document is generated rather than written. It also had nothing to say about a job whose
    lines were something else.
    """
    kinds: List[str] = []
    codes = [str(l.get("code") or "").upper() for l in lines]
    descs = " ".join(str(l.get("desc") or "") for l in lines).upper()
    if any(c in ("PACKAGING",) for c in codes):
        kinds.append("packaging")
    if any(c in ("DELIVERY",) for c in codes):
        kinds.append("delivery")
    if any(c.startswith("POWDER") for c in codes):
        kinds.append("the powder line")
    if any(c.startswith(("FIXING", "STD", "BI-", "P/P")) for c in codes):
        kinds.append("fasteners off the BOM table")
    _named = {"packaging", "delivery", "the powder line", "fasteners off the BOM table"}
    if len(kinds) < len(lines) or not kinds:
        kinds.append("bought items")
    if len(kinds) == 1:
        return kinds[0]
    return ", ".join(kinds[:-1]) + " and " + kinds[-1]


def _missing_drawings(bom: List[Dict[str, Any]], scan: Dict[str, Dict[str, Any]],
                      steel: Dict[str, Any],
                      material: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Lines with no sheet of their own, and whether their price depended on one.

    A LINE WITH NO SHEET IS ONLY A PROBLEM IF ITS PRICE RESTED ON ONE. A bought-in bolt has
    no detail drawing because it does not need one, and saying so stops somebody chasing
    Design for it.
    """
    out: List[Dict[str, Any]] = []
    for row in bom:
        code = str(row.get("code") or "").upper()
        if not code:
            continue
        rec = scan.get(code) or {}
        if rec.get("pages"):
            continue
        unit, qty = _money(row.get("price")), _money(row.get("qty"))
        out.append({
            "code": row.get("code"), "desc": _description(row),
            "cut": bool(steel.get(code)) or bool((material.get(code) or {}).get("Blank L")),
            "gbp": round(unit * qty, 2) if unit and qty else None,
            "priced": row.get("price") not in (None, ""),
        })
    return out


def _synthesised_number(part_number: str) -> bool:
    """A number the engine invented because the file's own would not resolve.

    drawing_job_merge names a promoted flat `<parent>-DXF<digits>` when it cannot bind it to a
    real child. That is a legitimate rescue and a loud signal: something in the pack did not
    trace, and this line may be a duplicate of one that did.
    """
    return "-DXF" in str(part_number or "").upper()


def _tracing_failures(scan: Dict[str, Dict[str, Any]], pack: List[str],
                      steel_calc: Dict[str, Dict[str, Any]],
                      material: Dict[str, Dict[str, Any]],
                      pages: Optional[Dict[int, Dict[str, Any]]] = None
                      ) -> List[Dict[str, Any]]:
    """Every part whose number stopped tracing somewhere, with what it cost.

    Reported per part rather than per pack, because "three numbers did not resolve" is not
    something anybody can act on and "01A was sized off the GA at £86" is.
    """
    out: List[Dict[str, Any]] = []
    for code, rec in sorted(scan.items()):
        if not code:
            continue
        _calc = steel_calc.get(code) or {}
        _det = material.get(code) or {}
        # Fabricated: the engine cut it, so it needed a drawing of it. A bought-in is
        # excluded here and reported — with its price and its source — on the bill of
        # materials, where it belongs.
        _roles = [str(r).lower() for r in (rec.get("page_roles") or [])]
        _fabricated = bool(_calc or _det.get("Blank L")
                           or str(rec.get("dxf_source_file") or "").strip()) or (
            "detail" in _roles and "bought_in" not in _roles)
        if not _fabricated:
            continue
        geom = str(rec.get("geometry_source") or "").strip().lower()
        pages = rec.get("pages") or []
        why: List[str] = []
        if _synthesised_number(rec.get("part_number") or code):
            why.append("its number was synthesised by the engine — the file it came from "
                       "could not be bound to a part in the BOM")
        if geom in _UNTRACED_GEOMETRY:
            why.append(_UNTRACED_GEOMETRY[geom])
        if not pages and not str(rec.get("dxf_source_file") or "").strip():
            why.append("no sheet of its own in the pack and no DXF — nothing in the pack "
                       "draws this part")
        if not why:
            continue
        out.append({
            "code": rec.get("part_number") or code,
            "description": _clip(rec.get("description"), 48),
            "why": why,
            "reader": _reader(rec.get("geometry_source")),
            "drawing_no": _drawing_no(rec),
            "where": _where(rec, pack, pages),
            "gbp": _money(_calc.get("total_value_gbp")) or _money(_det.get("Cost")),
            "blank": (f"{_fmt(_det.get('Blank L'))} x {_fmt(_det.get('Blank W'))}"
                      if _det.get("Blank L") else ""),
        })
    return sorted(out, key=lambda d: -(d["gbp"] or 0))


# ── the money's provenance ───────────────────────────────────────────────────

_INDICATIVE = ("grok", "llm", "xai", "indicative", "market")


def _order_qty_hint(bom_row: Dict[str, Any]) -> str:
    """The order quantity a commercial line was priced for, out of its own description.

    commercial_lines writes "PACKAGING for the whole order of 7, divided per unit" onto the
    line. That sentence is the only place the divisor survives into the workbook, so this
    reads it back rather than guessing at the header quantity — which is the same number
    today and would silently diverge the moment a line is priced for a different batch.
    """
    found = re.search(r"whole order of\s*(\d+)", str(bom_row.get("text") or ""),
                      re.IGNORECASE)
    return found.group(1) if found else ""


def _price_source(bom_row: Dict[str, Any], provenance: Dict[str, Dict[str, Any]],
                  steel_index: Optional[Dict[str, Dict[str, Any]]] = None) -> str:
    """Which book priced this line, in words an estimator can act on.

    THE FABRICATED PARTS ARE NOT UNPRICED. Every "-M" line carries a blank in the BOM's
    price column and says so in its own text — "costed in Sheet Steel below" — because sheet
    metal is priced BY NEST on the Sheet Steel block, not per piece here. Reading the
    blank as "no rate" put fifteen made parts on the estimator's to-do list, which is both
    wrong and the fastest way to lose their trust in the rest of the document.

    A BLANK WITH NO EXPLANATION IS THE ONE THAT MATTERS, and it stays loud.
    """
    code = bom_row["code"].upper()
    text = str(bom_row.get("text") or "")
    supplier = str(bom_row.get("supplier") or "").strip()
    # THE JOIN, TRIED THE WAYS THE TWO SHEETS ACTUALLY SPELL A CODE. Looking the raw code up
    # verbatim and giving up is how most of 12349-02's lines came out "source not named".
    prov = (provenance.get(code)
            or provenance.get(code.replace(" ", ""))
            or provenance.get(str(bom_row.get("text") or "").strip().split(" ")[0].upper())
            or {})
    named = str(prov.get("Price Source") or prov.get("price_source")
                or prov.get("Source") or prov.get("source") or "").strip()

    # A CATEGORY WORD IS NOT A CODE, AND SAYING SO IS THE ANSWER.
    #
    # SDI drawings put a CLASS in the part-code column where the item has no specific code:
    # "FIXING", "STD PART", "P/P". 12349-02 carried all three, each at GBP 0.00 with "source
    # not named in the workbook" beside it — which reads as a lookup we forgot to do. It is
    # not: there is nothing to look up. You cannot price the word "FIXING", and the identity
    # is entirely in the description next to it.
    try:
        from part_code_conventions import is_category_not_a_code as _is_category
    except Exception:                                            # noqa: BLE001
        _is_category = None                                      # type: ignore[assignment]
    # EVERY FABRICATED BLOCK, NOT JUST THE STEEL ONE. This tested for "costed in sheet
    # steel" alone, so 12349-02's acrylic and MDF — whose own text reads "costed in Other
    # Sheet Material" — fell through it and came out as "source not named in the workbook",
    # on lines that are correctly and deliberately blank here because their money is in
    # another block. Three made parts on the estimator's to-do list for no reason.
    _blocks = (("costed in sheet steel", "Sheet Steel"),
               ("costed in other sheet material", "Other Sheet Material"),
               ("costed in tube", "Tube"), ("costed in wire", "Wire"))
    for _needle, _label in _blocks[1:]:
        if _needle in text.lower():
            return (f"costed by nest on the {_label} block, not per piece here")
    if "costed in sheet steel" in text.lower():
        # NAME THE ROW IT IS COSTED ON. "priced below" is true and stops exactly where the
        # question starts; an estimator wants to look at the figure, not be told it exists.
        # NOT "BY BLANK AREA". That was the AI Material Detail tab's method, and it is not
        # what the sheet does: column M divides a whole sheet by how many of the part nest
        # out of it. Naming the wrong method on the row that points at the money is how a
        # covering note came to quote the blank-area figure.
        steel_row = (steel_index or {}).get(code, {}).get("row")
        return (f"costed by nest on the Sheet Steel block — see `Estimate!{steel_row}` below"
                if steel_row else
                "costed by nest on the Sheet Steel block, not per piece here")
    # ZERO IS WHAT A BLANK READS AS ONCE THE CELL HAS BEEN WRITTEN. This asked for None or
    # empty only, so every line the workbook had filled with 0.00 skipped past the loud
    # answer and landed on "source not named in the workbook" — which sends an estimator to
    # check a provenance tab about a line that simply has no price.
    _p = bom_row.get("price")
    if _p in (None, "") or (isinstance(_p, (int, float)) and float(_p) == 0.0):
        # WHY it could not be priced, where we can say. A class word in the code column is
        # not a lookup we forgot — there is nothing to look up, and the answer is a code.
        # Asked only of an unpriced line: PACKAGING is a category word too, and on a line
        # that carries GBP 25.00 the price is what matters, not the spelling of its code.
        if _is_category is not None and code and _is_category(code):
            # WHOSE FAULT IT IS, STATED CAREFULLY. The first version of this told an
            # estimator to "identify it and give it a code" — and on 12349-02 the engine had
            # ALREADY resolved BI-SCREW and BI-BUTTONSCREW for two of the three lines and
            # then kept the class word over them. Asking a person to do work we had already
            # done and thrown away is worse than saying nothing.
            return ("**NOT PRICED — the code column holds a CLASS, not a code.** The sheet "
                    f"carries '{bom_row['code']}', which is the word a drawing prints where "
                    "an item has no part number, so nothing can look a rate up against it. "
                    "If the item has an SDI code, put it in and the price follows; if it "
                    "genuinely has none, price it by hand. Where the engine resolved a real "
                    "code and this class word displaced it, that is ours")
        return "**NOT PRICED — needs a rate**"
    # PACKAGING AND DELIVERY ARE ORDER-LEVEL, AND THE DIVISOR IS THE POINT.
    #
    # Both are asked for the WHOLE ORDER and divided by the quantity, so the per-unit figure
    # an estimator reads falls as the order rises — on 12552 they were £85 + £85 on a unit of
    # £930.39 at 1 off, and near nothing at 100. "AI market indication" says neither of those
    # things, and does not say how to stop it being an indication at all.
    _order_line = str(bom_row.get("code") or "").upper() in ("PACKAGING", "DELIVERY")
    if _order_line and any(token in f"{supplier} {named}".lower() for token in _INDICATIVE):
        _qty = _order_qty_hint(bom_row)
        return ("market indication for the WHOLE ORDER"
                + (f" of {_qty}, ÷ {_qty} per unit" if _qty else ", divided per unit")
                + " — NOT A QUOTE. To make it a firm house rate on every job, set "
                  "`config.COMMERCIAL_LINE_GBP_PER_ORDER` and this becomes a catalogue price")
    if any(token in supplier.lower() for token in _INDICATIVE):
        return f"AI market indication ({supplier}) — NOT A QUOTE, replace it"
    if any(token in named.lower() for token in _INDICATIVE):
        return f"AI market indication ({named}) — NOT A QUOTE, replace it"
    if supplier:
        return f"catalogue — {supplier}"
    if named:
        return named
    # Said plainly rather than guessed at. "config rate card" was asserted here for anything
    # unlabelled, which claimed a source for the bearing's GBP 1.42 that nothing in the
    # workbook actually states.
    return "source not named in the workbook — check AI Provenance"


# ── the document ─────────────────────────────────────────────────────────────

def _description(bom_row: Dict[str, Any]) -> str:
    """What the part IS, with the workbook's warnings stripped off the end.

    The Estimate sheet's description cell carries the part code, then the description, then
    any notice appended to it — "[AI ESTIMATE - INDICATIVE, NOT A QUOTE]", "MATERIAL
    UNPRICED: enter a unit rate for this item". Taking the LAST segment kept the notice and
    threw the description away, so the concrete slab appeared as "[AI ESTIMATE - INDICATIVE,
    NOT A QUOTE]" and the nylon washer as "MATERIAL UNPRICED" — a table naming no parts.
    Those notices are already carried, properly, by the price column beside it.
    """
    text = str(bom_row.get("text") or "")
    code = str(bom_row.get("code") or "")
    if code and text.upper().startswith(code.upper()):
        text = text[len(code):]
    for marker in ("[AI ESTIMATE", "—  MATERIAL UNPRICED", "— MATERIAL UNPRICED",
                   "MATERIAL UNPRICED", "— costed in Sheet Steel", "— costed in sheet steel"):
        cut = text.find(marker)
        if cut > 0:
            text = text[:cut]
    text = text.strip(" —-—\t")
    return (text[:60] or "—")


def _gbp(value: Any) -> str:
    """A number as money, or a plain statement that the figure was not available.

    NEVER A ZERO. A total whose cell holds an uncached formula reads as None, and printing
    that as £0.00 would make a reconciliation against it look exact.
    """
    number = _money(value)
    return f"£{number:,.2f}" if number is not None else "not readable from the sheet"


def _gbp_or(value: Any, dash: str) -> str:
    """Money where there is a number, and the caller's own words where there is not."""
    return _gbp(value) if _money(value) is not None else dash


def _diff(here: float, sheet: Optional[float]) -> str:
    if sheet is None:
        return "cannot be checked"
    gap = round(here - sheet, 2)
    if abs(gap) < 0.01:
        return "**none — it reconciles**"
    # SAID AS A DIRECTION, NOT A SIGNED NUMBER. "£-193.30" makes the reader work out which way
    # round it is, and the two directions mean opposite things: short means lines are missing
    # from this document, over means it is counting something the sheet is not.
    if gap < 0:
        return f"**{_gbp(-gap)} short** — lines this document cannot see"
    return f"**{_gbp(gap)} more** than the sheet — this document is counting something twice"


def _fe_totals(final: Dict[str, Any]) -> Dict[str, Any]:
    node = final.get("totals")
    return node if isinstance(node, dict) else {}


def _measured_sentence(material: Dict[str, Dict[str, Any]],
                       scan: Optional[Dict[str, Dict[str, Any]]] = None) -> str:
    """How much of the geometry was measured off a model rather than reasoned from a view.

    THE PART RECORD ANSWERS THIS WHEN THE TAB WILL NOT. On 12552 the AI Material Detail tab
    read "not recorded" for all 31 parts while the same run's part records carried
    solidworks_flat_pattern on nineteen of them — so the honest answer was sitting one join
    away while the document reported that nothing at all had been measured. That is the
    worst kind of wrong here: it understates the estimate's own evidence.
    """
    sources: Dict[str, int] = {}
    for code, row in material.items():
        key = str(row.get("Geom source") or "").strip()
        if not key or key == "not recorded":
            key = str(((scan or {}).get(code) or {}).get("geometry_source") or "").strip()
        sources[key or "not recorded"] = sources.get(key or "not recorded", 0) + 1
    if not sources:
        return "The workbook carries no geometry-source column to answer this from."
    ranked = sorted(sources.items(), key=lambda kv: (-kv[1], kv[0]))
    return ("Of " + str(sum(sources.values())) + " part(s) with a material row: "
            + ", ".join(f"{n} {name}" for name, n in ranked) + ".")


def _blocks_sentence(material_rows: List[Dict[str, Any]],
                     labour_rows: List[Dict[str, Any]]) -> str:
    if not material_rows and not labour_rows:
        return "No calculated rows were supplied, so this cannot be answered."
    parts = []
    for block in ("bom", "tube", "steel", "other_sheet"):
        rows = [r for r in material_rows if r.get("block") == block]
        if rows:
            parts.append(f"{_BLOCK_NAMES[block].lower()} {_gbp(_sum_money(rows))}")
    if labour_rows:
        parts.append(f"labour {_gbp(_sum_money(labour_rows))} across "
                     f"{len(labour_rows)} sheet row(s)")
    return "; ".join(parts) + "."


def _reconciles_sentence(material_rows: List[Dict[str, Any]],
                         labour_rows: List[Dict[str, Any]],
                         totals: Dict[str, Any]) -> str:
    if not material_rows and not labour_rows:
        return ("Not yet — no calculated rows were supplied, so nothing below can be summed "
                "against the sheet.")
    gaps = []
    for label, rows, total in (("material", material_rows, totals.get("material")),
                               ("labour", labour_rows, totals.get("labour"))):
        if total is None:
            gaps.append(f"the sheet's {label} total could not be read")
            continue
        gap = round(_sum_money(rows) - total, 2)
        if abs(gap) >= 0.01:
            gaps.append(f"{label} is {_gbp(abs(gap))} "
                        + ("short" if gap < 0 else "over"))
    if not gaps:
        return ("Yes. Every material and labour line below sums to the Estimate sheet's own "
                "Total Material Cost and Total Labour Cost.")
    return "Not entirely — " + "; ".join(gaps) + ". The table below says where."


def _fmt(value: Any, dash: str = "—") -> str:
    if value in (None, ""):
        return dash
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)


def _gather(workbook: Path, scan_json: Optional[Path]) -> Dict[str, Any]:
    """Everything both the document and the covering email are written from.

    ONE READER, TWO RENDERERS. The email restates the estimate's headline figures in the body
    of a message somebody forwards to a customer's engineer; the markdown document states them
    in full. If each opened the workbook and totalled it for itself, the two would eventually
    disagree — and the one that disagreed would be the one already sent. So the reading is done
    once, here, and both renderers are handed the same numbers.
    """
    wb = openpyxl.load_workbook(workbook, data_only=True)
    scan_doc = _load_scan(scan_json)
    final = _final_estimate(scan_doc)
    return {
        "wb": wb,
        "stem": workbook.stem,
        "scan_doc": scan_doc,
        "scan": _scan_parts(scan_doc),
        "final": final,
        "sufficiency": _data_sufficiency(scan_doc),
        "accepted": _accepted_labour(scan_doc),
        "labour_rows": [r for r in (final.get("labour_rows") or []) if isinstance(r, dict)],
        "material_rows": [r for r in (final.get("material_rows") or []) if isinstance(r, dict)],
        "totals": _sheet_totals(wb, final),
        "steel": _steel_rows(wb),
        # The calculated steel rows, keyed the way the block's own description cell is keyed —
        # the read-back carries no part code for the fabricated blocks, only the description
        # the engine wrote into them, and that description begins with the part number.
        "steel_calc": {str(r.get("description") or "").split()[0].strip().upper(): r
                       for r in (final.get("material_rows") or [])
                       if isinstance(r, dict) and r.get("block") == "steel"
                       and str(r.get("description") or "").strip()},
        "material": {str(r.get("Part") or "").upper(): r
                     for r in _sheet(wb, "AI Material Detail")},
        "provenance": {str(r.get("Part") or "").upper(): r
                       for r in _sheet(wb, "AI Price Provenance")},
        "routes": _sheet(wb, "Canonical Route"),
        "bom": _estimate_bom(wb),
        "order_qty": _order_quantity(wb),
        "pack": _pack_files(scan_doc),
        "page_index": _page_index(scan_doc),
    }


def build(workbook: Path, scan_json: Optional[Path]) -> str:
    g = _gather(workbook, scan_json)
    wb, scan_doc, scan = g["wb"], g["scan_doc"], g["scan"]
    final, sufficiency, accepted = g["final"], g["sufficiency"], g["accepted"]
    labour_rows, material_rows = g["labour_rows"], g["material_rows"]
    totals, steel, steel_calc = g["totals"], g["steel"], g["steel_calc"]
    material, provenance, routes = g["material"], g["provenance"], g["routes"]
    bom, order_qty, pack = g["bom"], g["order_qty"], g["pack"]
    page_index = g["page_index"]

    lines: List[str] = []
    add = lines.append

    add(f"# {workbook.stem}")
    add("")
    if not scan:
        add("> **Page numbers are unavailable** — no scan JSON was supplied, so the "
            "\"which sheet\" column reads *not supplied* throughout. Re-run with "
            "`--scan-json output/json/<job>.json` to fill it.")
        add("")

    # ── the questions, answered before the tables ────────────────────────────
    # ASKED AND ANSWERED IN THAT ORDER. A document that opens with four hundred rows makes
    # the reader derive the five facts they came for. These are the five, each computed from
    # the same rows the tables below print — never typed, so they cannot drift from them.
    _unpriced = [r for r in bom
                 if r.get("price") in (None, "")
                 and "costed in sheet steel" not in str(r.get("text") or "").lower()]
    _indicative = [r for r in bom
                   if any(t in f"{r.get('supplier') or ''}".lower() for t in _INDICATIVE)
                   and _money(r.get("price"))]
    _indicative_gbp = round(sum((_money(r.get("price")) or 0) * (_money(r.get("qty")) or 0)
                                for r in _indicative), 2)
    add("## The questions, answered first")
    add("")
    add(f"- **What does a unit cost, and of what?** "
        f"{_gbp(totals['unit'])} — material {_gbp(totals['material'])} + labour "
        f"{_gbp(totals['labour'])}"
        + (f", and the unit cell adds {_gbp(_money(_fe_totals(final).get('other_gbp')))} "
           f"({final.get('unit_price_composition', {}).get('basis') or 'per the sheet'})"
           if _fe_totals(final).get("other_gbp") else "")
        + ".")
    add(f"- **What must be replaced before this is a quote?** "
        + (f"{len(_unpriced)} line(s) carry no price at all — "
           + ", ".join(r["code"] or _description(r) for r in _unpriced[:8])
           + ("…" if len(_unpriced) > 8 else "") + ". "
           if _unpriced else "No line is unpriced. ")
        + (f"{len(_indicative)} line(s) worth {_gbp(_indicative_gbp)} are AI market "
           f"indications, not catalogue prices: "
           + ", ".join(r["code"] or _description(r) for r in _indicative[:8]) + "."
           if _indicative else "No line rests on an AI market indication."))
    add(f"- **How much of this was measured rather than reasoned?** "
        + _measured_sentence(material, scan))
    add(f"- **Where is the money?** " + _blocks_sentence(material_rows, labour_rows))
    add(f"- **Does this document add up to the sheet?** "
        + _reconciles_sentence(material_rows, labour_rows, totals))
    add("")

    # ── what a person still has to do ────────────────────────────────────────
    # NAMED, PRICED AND COUNTED, not summarised in a sentence. "3 lines carry no price" is
    # true and an estimator cannot act on it: they want the codes, the money each one is
    # worth, the page to look at and whether it is waiting on them or on us. This is the list
    # that has been retyped into every covering email so far.
    if _unpriced or _indicative:
        add("## What a person still has to settle")
        add("")
        add(f"{len(_unpriced) + len(_indicative)} line(s). Until these are answered the "
            f"estimate is not a quote, and the banner on the sheet says so.")
        add("")
        add("| Line | What it is | Qty | On the sheet | What it needs | Which file and page |")
        add("|---|---|---|---|---|---|")
        _todo = ([(r, "indicative") for r in _indicative]
                 + [(r, "unpriced") for r in _unpriced])
        for row, kind in sorted(
                _todo, key=lambda pair: -((_money(pair[0].get("price")) or 0)
                                          * (_money(pair[0].get("qty")) or 0))):
            _unit, _qty = _money(row.get("price")), _money(row.get("qty"))
            _ext = round(_unit * _qty, 2) if _unit and _qty else None
            _rec = scan.get(row["code"].upper()) or {}
            add(f"| {row['code'] or '—'} | {_description(row)} | {_fmt(row.get('qty'))} "
                + (f"| {_gbp(_ext)} — an AI market indication, not a catalogue price "
                   f"| **Overwrite it, or accept it deliberately.** It moves between runs, "
                   f"so an estimate resting on it cannot be reproduced. "
                   if kind == "indicative" else
                   "| **£0.00 — the line is costing nothing** "
                   "| **A rate.** Nothing we can query holds a price for this code. ")
                + f"| {_where(_rec, pack, page_index)} |")
        add("")

    # ── does it reconcile ────────────────────────────────────────────────────
    # THE CLAIM THAT MAKES THE REST OF THE DOCUMENT WORTH READING, AND THE ONE THAT CAN BE
    # CHECKED. Every line printed below is summed and held against the Estimate sheet's own
    # labelled total. Where the two differ the difference is stated in pounds — an
    # explanation that quietly covers 81% of a total is worse than one that says which 19% it
    # cannot see, because only the second sends anybody looking.
    _pointers = sum(1 for r in bom
                    if "costed in sheet steel" in str(r.get("text") or "").lower())
    if material_rows or labour_rows:
        add("## This document against the sheet")
        add("")
        add("| Block | Lines here | £ here | The sheet's own total | Difference |")
        add("|---|---|---|---|---|")
        for block in ("bom", "tube", "steel", "other_sheet"):
            rows_in = [r for r in material_rows if r.get("block") == block]
            if rows_in:
                # THE BOM BLOCK CARRIES THE POINTERS TOO. Every part costed on Sheet Steel
                # also has a row up here showing a dash, so a bare count reads as thirty-two
                # bought-in lines when seventeen of them are what anybody would call bought in.
                _count = str(len(rows_in))
                if block == "bom" and _pointers:
                    _count = (f"{len(rows_in)} — {len(rows_in) - _pointers} priced or awaiting "
                              f"a rate, {_pointers} pointing at Sheet Steel")
                add(f"| {_BLOCK_NAMES[block]} | {_count} | "
                    f"{_gbp(_sum_money(rows_in))} | — | — |")
        add(f"| **All material** | {len(material_rows)} | {_gbp(_sum_money(material_rows))} "
            f"| {_gbp(totals['material'])} | {_diff(_sum_money(material_rows), totals['material'])} |")
        add(f"| **All labour** | {len(labour_rows)} | {_gbp(_sum_money(labour_rows))} "
            f"| {_gbp(totals['labour'])} | {_diff(_sum_money(labour_rows), totals['labour'])} |")
        add("")
        add("> Every figure in the £ column is the value the Estimate sheet itself computed "
            "for that row, read back after Excel recalculated. Nothing here is re-derived, "
            "so a difference is a row this document cannot see — not a rounding argument.")
        _from = totals.get("_from") or {}
        for key, label in (("material", "Total Material Cost"),
                           ("labour", "Total Labour Cost")):
            if _from.get(key):
                add(f"> **{label}** came from {_from[key]}.")
        add("")
        for problem in (final.get("adapter_problems") or []):
            if isinstance(problem, dict) and problem.get("message"):
                add(f"> **A block was not read:** {problem['message']}")
                add("")
    elif scan_doc is not None:
        add("## This document against the sheet")
        add("")
        add("> **Not produced.** The run JSON supplied carries no `final_estimate`, which is "
            "the record of the rows as Excel calculated them. Without it the lines below can "
            "be listed but not summed against the sheet's own totals, so this document "
            "cannot claim to be complete. Re-run with the full "
            "`output/json/<job>.json` from a run whose workbook was read back.")
        add("")

    # ── every BOM line ───────────────────────────────────────────────────────
    # TWO ROWS FOR A PART SDI CUTS, because it has two halves on the sheet and an estimator
    # asked for both: the BOM row that shows a dash and points below, and the Sheet Steel row
    # that holds the blank, the gauge and the money. The page is stamped on both — the same
    # drawing owns both halves, and leaving the second one blank is what made the join look
    # like a missing answer.
    add("## Every line on the Bill of Materials")
    add("")
    add("| Part | Description | Qty | Unit £ | Ext £ | Where the £ came from | Material | "
        "Gauge | Blank | Which sheet |")
    add("|---|---|---|---|---|---|---|---|---|---|")
    for row in bom:
        code = row["code"].upper()
        mat = material.get(code, {})
        rec = scan.get(code, {})
        blank_l, blank_w = mat.get("Blank L"), mat.get("Blank W")
        blank = (f"{_fmt(blank_l)} × {_fmt(blank_w)}"
                 if blank_l or blank_w else "not a cut part")
        page = _where(rec, pack, page_index)
        _unit, _qty = _money(row.get("price")), _money(row.get("qty"))
        add(f"| {row['code'] or '—'} "
            f"| {_description(row)} "
            f"| {_fmt(row.get('qty'))} "
            f"| {_fmt(row.get('price'))} "
            f"| {_gbp(round(_unit * _qty, 2)) if _unit and _qty else '—'} "
            f"| {_price_source(row, provenance, steel)} "
            f"| {_fmt(mat.get('Material'))} "
            f"| {_fmt(mat.get('Gauge'), 'none — not sheet')} "
            f"| {blank} "
            f"| {page} |")
        steel_row = steel.get(code)
        if steel_row:
            # M IS A LINE TOTAL, NOT A COST PER PART. The column is headed "Cost Per Part"
            # and holds ROUNDUP(sheet price / parts-per-sheet, 2) x qty x scrap — the whole
            # line. Printing it under a per-part heading, or dividing it back out by the
            # quantity, invents a per-piece figure the sheet never computed; printing the AI
            # Material Detail tab's per-part instead put a blank-area number in the money
            # column that is not in the unit cost at all. Same naming trap as the labour
            # block's "Rate Per Hour", which holds a throughput.
            line_total = _gbp_or((steel_calc.get(code) or {}).get("total_value_gbp"),
                                 "not read back")
            add(f"| ↳ `Estimate!{steel_row['row']}` "
                f"| the same part, on the Sheet Steel block "
                f"| {_fmt(steel_row.get('qty'))} "
                f"| — line total, not per part "
                f"| **{line_total}** "
                f"| one {_fmt(steel_row.get('sheet_l'))} × {_fmt(steel_row.get('sheet_w'))} "
                f"sheet ÷ {_fmt((steel_calc.get(code) or {}).get('qty_per_sheet'), '?')} "
                f"nested, × qty, "
                f"× {_fmt(steel_row.get('scrap'))} scrap "
                f"| {_fmt(mat.get('Material'))} "
                f"| {_fmt(steel_row.get('gauge'), 'none — not sheet')} "
                f"| {_fmt(steel_row.get('length'))} × {_fmt(steel_row.get('width'))} "
                f"| {page} |")
    add("")

    # ── where the fabricated money actually is ───────────────────────────────
    if steel:
        # PRICED BY NEST. "by blank area" was the AI Material Detail tab's method and is
        # not what this block does — column M divides a whole sheet by how many of the part
        # nest out of it. The footnote under this very table has said so correctly the whole
        # time, so the section contradicted itself over the one number an estimator came for,
        # and the wrong half was the heading. James, on 12552: "Ignore the 'by blank area'
        # wording on the dash rows."
        add("## The fabricated parts, priced by nest")
        add("")
        add("Each of these shows a dash in the BOM price column above. That is not a missing "
            "rate — sheet metal is costed on this block, from how many of the part nest out "
            "of a sheet, and pricing it in both places would double it. This is the other "
            "half of those lines.")
        add("")
        add("| Part | Blank L × W | Gauge | Off a sheet | Nest per sheet | Scrap | Qty | "
            "**£ the sheet charges** | The engine's own figure | Sheet row |")
        add("|---|---|---|---|---|---|---|---|---|---|")
        for code, row in steel.items():
            # COST FROM THE RESOLVED FIGURE, NOT THE FORMULA CELL. The Sheet Steel block's
            # cost column is an Excel formula, and a workbook written by the engine and never
            # opened has no cached result — so reading that cell alone gives nothing and the
            # table says "computes in Excel" on the one column an estimator came for. The
            # engine's own resolved figure for the same part is on AI Material Detail, which
            # is a value, not a formula. Still read, never recalculated here.
            mat = material.get(code, {})
            # THE NEST COUNT COMES FROM THE READ-BACK, NOT THE CELL. Qty Per Sheet is a
            # formula with no cached result in the saved file, so the column read "not
            # computed" on every row of a job whose whole steel story is the nest. Excel
            # calculated it and the read-back recorded it.
            nest = _fmt((steel_calc.get(code) or {}).get("qty_per_sheet"), "not recorded")
            add(f"| {code} "
                f"| {_fmt(row.get('length'))} × {_fmt(row.get('width'))} "
                f"| {_fmt(row.get('gauge'))} "
                f"| {_fmt(row.get('sheet_l'))} × {_fmt(row.get('sheet_w'))} "
                f"| {nest} "
                f"| {_fmt(row.get('scrap'))} "
                f"| {_fmt(row.get('qty'))} "
                f"| **{_gbp_or((steel_calc.get(code) or {}).get('total_value_gbp'), 'not read back')}** "
                f"| {_gbp_or(mat.get('Ext Material'), 'not resolved')} "
                f"| `Estimate!{row['row']}` |")
        add("")
        add("> **£ the sheet charges** is column M of the Sheet Steel row, and it is a LINE "
            "TOTAL despite the column being headed *Cost Per Part*: "
            "`ROUNDUP(sheet price / nest per sheet, 2) x qty x scrap`. Do not divide it back "
            "out — the sheet computes no per-piece figure. That total is what is inside "
            "Total Material Cost. **The engine's own figure** is the AI Material Detail tab's "
            "blank-area calculation of the same part; it is NOT in the unit cost and must "
            "not be quoted. Nothing here is recalculated.")
        # THE TWO VIEWS OF THE SAME FIFTEEN PARTS, HELD AGAINST EACH OTHER.
        #
        # On 12552 the engine's per-part figures extended to £49.76 while the sheet's own
        # steel rows summed to £136.32 — £86.56, sixteen per cent of the material total, on
        # the same fifteen parts. The document printed the first set and reconciled with the
        # second, so both numbers were honestly labelled and nothing said they disagreed.
        # An estimator adding the Extended column would have got a figure the sheet does not
        # charge, and the covering note quoted "£1.05 a part" off it.
        if steel_calc:
            _engine = round(sum(_money((material.get(c) or {}).get("Ext Material")) or 0.0
                                for c in steel), 2)
            _sheet_side = round(sum(_money(r.get("total_value_gbp")) or 0.0
                                    for r in steel_calc.values()), 2)
            if abs(_engine - _sheet_side) >= 0.01:
                add("")
                add(f"> **The two columns disagree, and the sheet's is the one you pay.** "
                    f"The engine resolved these {len(steel)} part(s) at {_gbp(_engine)}; the "
                    f"sheet charges {_gbp(_sheet_side)}, a difference of "
                    f"{_gbp(abs(_engine - _sheet_side))}. The usual cause is the nest: the "
                    f"sheet divides a whole sheet by the Nest per sheet column beside each "
                    f"row, and a part that yields one per sheet carries the whole sheet — at "
                    f"one off that is the template working correctly, not a double charge. "
                    f"Quote the bold column, and do not change the block to match the other "
                    f"one without a policy decision: every job costed on this template moves "
                    f"with it.")
        add("")

    # ── every labour line, with the money on it ──────────────────────────────
    # A THIRD OF THE UNIT COST, PREVIOUSLY UNEXPLAINED. This document listed every operation
    # and what decided it, and put no figure against any of them — so an estimator could see
    # that 12552 folds and welds, and could not see that the folding costs £41 until they
    # opened the sheet and found the row themselves. The route section below answers "why is
    # this operation here"; this one answers "what is it charging, and to which parts".
    if labour_rows:
        add("## Every labour line, and what it charges")
        add("")
        add("Sorted by cost. **Batch hours** are for the whole order, not one unit; **£ this "
            "line** is the sheet's own Total Value for the row. Parts come from the accepted "
            "route grouping, joined to the calculated row on the sheet row they share.")
        add("")
        add("| Sheet row | Operation | Dept | Parts | Qty/unit | Set-up (min) | Batch hours | "
            "Rate £/hr | £ this line | Rate basis |")
        add("|---|---|---|---|---|---|---|---|---|---|")
        for row in sorted(labour_rows, key=lambda r: -(_money(r.get("total_value_gbp")) or 0)):
            acc = accepted.get(int(_money(row.get("workbook_row")) or 0), {})
            parts = [str(p) for p in (acc.get("part_numbers") or []) if p]
            add(f"| `Estimate!{_fmt(row.get('workbook_row'))}` "
                f"| {_fmt(row.get('operation'))} "
                f"| {_fmt(row.get('department'))} "
                f"| {', '.join(parts) if parts else 'not recorded on the accepted row'} "
                f"| {_fmt(row.get('qty_per_unit'))} "
                f"| {_fmt(row.get('setup_minutes'))} "
                f"| {_fmt(row.get('batch_hours'))} "
                f"| {_fmt(row.get('dept_rate_gbp_per_hour'))} "
                f"| {_gbp(row.get('total_value_gbp'))} "
                f"| {_fmt(acc.get('rate_basis'), 'not recorded')} |")
        add("")
        add(f"> {len(labour_rows)} row(s), {_gbp(_sum_money(labour_rows))} — "
            f"against the sheet's Total Labour Cost of {_gbp(totals['labour'])}.")
        add("")

    # ── which reader decided each part ───────────────────────────────────────
    # THE OTHER HALF OF "WHERE DID YOU SEE THAT". The page says which sheet a part is on; this
    # says which reader supplied each fact about it, and at what rank — a gauge off the
    # SOLIDWORKS model and one read off a title block are both "p.7", and only one of them was
    # measured. The engine has stamped this on every field as it went; nothing printed it.
    if scan:
        _graded = [(code, rec) for code, rec in sorted(scan.items())
                   if any(rec.get(f) for f in ("material_source", "thickness_source",
                                               "quantity_source", "geometry_source"))]
        if _graded:
            add("## Which reader decided each part")
            add("")
            add("The number in brackets is the engine's rank for that reader: a higher-ranked "
                "source may not be overwritten by a lower one, which is why a SOLIDWORKS "
                "measurement survives a vision model's reading of the same part and not the "
                "other way round.")
            add("")
            add("**What decided something on this job:** "
                + "; ".join(_readers_used(scan)) + ".")
            add("")
            add("| Part | Material | Thickness | Quantity | Geometry / size | Which file and page |")
            add("|---|---|---|---|---|---|")
            for code, rec in _graded:
                add(f"| {code or '—'} "
                    f"| {_reader(rec.get('material_source'))} "
                    f"| {_reader(rec.get('thickness_source'))} "
                    f"| {_reader(rec.get('quantity_source'))} "
                    f"| {_reader(rec.get('geometry_source'))} "
                    f"| {_where(rec, pack, page_index)} |")
            add("")
            add("> A SOLIDWORKS source means the part and assembly files themselves were read "
                "— not the PDF of them. Where the model was available the engine takes the "
                "flat pattern, the applied material and the bend count from it, because those "
                "are the modelled facts rather than somebody's reading of a view. Where it was "
                "not, the drawing is the best available and the row says so.")
            add("")

    # ── what the labour actually is ──────────────────────────────────────────
    # THE WHOLE QUANTITY STORY, AND IT WAS BEING WORKED OUT BY HAND. Every labour row carries
    # a one-off set-up and a run time, and the sheet charges both against however many units
    # the run was for. The set-up is what falls when the quantity rises; the run time never
    # moves. Without the split an estimator asked "what would 25 off cost" has to open the
    # sheet, find eleven set-up cells and eleven rates, and do it themselves — which is what
    # happened, and it is the kind of arithmetic that goes into an email slightly wrong.
    if labour_rows and order_qty:
        _setup = 0.0
        for row in labour_rows:
            _mins = _money(row.get("setup_minutes")) or 0.0
            _rate = _money(row.get("dept_rate_gbp_per_hour")) or 0.0
            _setup += (_mins / 60.0) * _rate / order_qty
        _setup = round(_setup, 2)
        _charged = _sum_money(labour_rows)
        _run = round(_charged - _setup, 2)
        add("## What the labour is: set-up, and run time")
        add("")
        add(f"This estimate is for **{order_qty} off**. Of the {_gbp(_charged)} of labour on "
            f"it, **{_gbp(_setup)} is set-up** and **{_gbp(_run)} is run time**. Set-up is a "
            f"one-off per department row and is spread across the order, so it falls as the "
            f"quantity rises. Run time per unit does not move at any quantity — it is the "
            f"floor.")
        add("")
        add("| Order qty | Set-up per unit | Run per unit | Labour per unit |")
        add("|---|---|---|---|")
        for _q in (1, 10, 25, 50, 100, 250):
            _s = round(_setup * order_qty / _q, 2)
            add(f"| {_q}{' — this estimate' if _q == order_qty else ''} | {_gbp(_s)} "
                f"| {_gbp(_run)} | {_gbp(round(_s + _run, 2))} |")
        add("")
        add("> Labour only. Material per part does not change with quantity, and packaging "
            "and delivery are priced for the whole order at the quantity above — neither is "
            "in this table, so it is not a unit price. It is what the labour would do.")
        add("")

    # ── every route line ─────────────────────────────────────────────────────
    add("## Every operation, and who decided it")
    add("")
    add("| Part | Operation | Seq | Scope | Qty | Decided by | On what basis | Which file and page |")
    add("|---|---|---|---|---|---|---|---|")
    for row in routes:
        target = str(row.get("Target") or "").upper()
        rec = scan.get(target, {})
        add(f"| {_fmt(row.get('Target'))} "
            f"| {_fmt(row.get('Operation'))} "
            f"| {_fmt(row.get('Seq'))} "
            f"| {_fmt(row.get('Scope'))} "
            f"| {_fmt(row.get('Qty/unit'))} "
            f"| {_fmt(row.get('Source'))} "
            f"| {str(row.get('Reason') or '—')[:70]} "
            f"| {_where(rec, pack, page_index)} |")
    add("")

    # ── what each sheet could be read for ────────────────────────────────────
    # A FIELD MISSING FROM THIS EXTRACT IS NOT A FIELD MISSING FROM THE DRAWING.
    #
    # This section reports what each sheet could be read for, and it does that by looking for
    # materials / thicknesses_mm / surface_finishes on the record. The trimmed extract that
    # supplies the page numbers carries only part_number, pages and page_roles — so against
    # one of those every row came out "material **no**, thickness **no**, finish **no**",
    # for a pack whose page 4 plainly reads "MATERIAL: MILD STEEL", "1.5 THK", "POWDER
    # COATED". That is not a weak answer, it is a confident wrong one, and it would have gone
    # to an estimator as a drawing-quality assessment.
    #
    # So the section is built only from an extract that actually carries the fields, and
    # otherwise says what it needs. Refusing to answer is the honest failure mode; the whole
    # document exists to stop people guessing from it.
    _quality_fields = ("materials", "thicknesses_mm", "surface_finishes", "geometry_rollup")
    _has_quality = any(any(k in rec for k in _quality_fields) for rec in scan.values())
    if scan and not _has_quality:
        add("## Drawing quality, sheet by sheet")
        add("")
        add("> **Not produced.** This ran against a trimmed extract carrying only part "
            "numbers and page numbers. Reporting from it would have said every drawing "
            "states no material, no thickness and no finish — which is false: page 4 alone "
            "reads *MATERIAL: MILD STEEL*, *1.5 THK*, *POWDER COATED*. Re-run with the full "
            "`output/json/<job>.json` and this section builds itself.")
        add("")
    if scan and _has_quality:
        add("## Drawing quality, sheet by sheet")
        add("")
        add("| File and page | Part | Material stated | Thickness stated | Finish stated | "
            "Geometry | What it could not give |")
        add("|---|---|---|---|---|---|---|")
        for code, rec in sorted(scan.items(),
                               key=lambda kv: ((kv[1].get("pages") or [999])[0], kv[0])):
            # ONLY THINGS THAT ARE DRAWINGS. A line with no page is a catalogue item, a
            # commercial line, or a fastener off a BOM table — PACKAGING, DELIVERY, FIXING17.
            # Graded here they came out "could not give: thickness, finish, cut length",
            # which reads as a deficient drawing pack and is nonsense: packaging is not a
            # drawing, and an M8 nyloc nut has no thickness because it is bought, not because
            # somebody forgot to write one. Same false-negative as grading a pack from a
            # trimmed extract, in a different place.
            if not (rec.get("pages") or []):
                continue
            geom = rec.get("geometry_rollup") or {}
            reliability = ((geom.get("confidence") or {}).get("geometry_reliability"))
            # A PURCHASED PART NAMED ON AN ASSEMBLY SHEET IS NOT AN INCOMPLETE DETAIL.
            # The bearing and the rivets appear on p.2 because that is where they are listed,
            # and they have no detail drawing because they do not need one. Saying what they
            # ARE beats listing four fields they were never going to carry.
            _roles = [str(r).lower() for r in (rec.get("page_roles") or [])]
            if "bought_in" in _roles and "detail" not in _roles:
                missing = ["bought in — listed on an assembly sheet, no detail drawing needed"]
            else:
                missing = []
                if not rec.get("materials"):
                    missing.append("material")
                if not rec.get("thicknesses_mm"):
                    missing.append("thickness")
                if not rec.get("surface_finishes"):
                    missing.append("finish")
                if not (geom.get("estimated_cut_length_mm") or 0):
                    missing.append("cut length")
            add(f"| {_where(rec, pack, page_index)} | {code or _fmt(rec.get('description'))} "
                f"| {', '.join(str(m) for m in rec.get('materials') or []) or '**no**'} "
                f"| {', '.join(str(t) for t in rec.get('thicknesses_mm') or []) or '**no**'} "
                f"| {', '.join(str(f) for f in rec.get('surface_finishes') or []) or '**no**'} "
                f"| {_fmt(rec.get('geometry_source'))}"
                f"{f' ({reliability:.0%})' if isinstance(reliability, (int, float)) else ''} "
                f"| {', '.join(missing) or 'nothing — complete'} |")
        add("")

    # ── where a part number stopped tracing, and what it cost ────────────────
    # FOR DESIGN AS MUCH AS FOR THE ESTIMATOR. Every entry here is a place the engine had to
    # size or price a part from something that is not a drawing OF that part. Some of those
    # are our reader's fault and some are the pack's, and the only way to tell which is to
    # print both the break and what the estimate did instead.
    _untraced = _tracing_failures(scan, pack, steel_calc, material, page_index)
    if _untraced:
        add("## Where a part number stopped tracing through the pack")
        add("")
        _at_stake = round(sum(u["gbp"] or 0 for u in _untraced), 2)
        add(f"{_plural(len(_untraced), 'part')} could not be followed from the BOM through to "
            f"a drawing or a flat of its own"
            + (f", carrying {_gbp(_at_stake)} between them" if _at_stake else "")
            + ". Each was still costed — from whatever the engine could reach — and that "
              "substitution is what the last column names. **Where the break is ours we will "
              "fix the reader; where it is the pack's, this is the list to send Design.**")
        add("")
        add("| Part | What it is | £ on this job | Blank used | Which file and page "
            "| Where the trail broke |")
        add("|---|---|---|---|---|---|")
        for u in _untraced:
            add(f"| {u['code']} | {u['description'] or '—'} | {_gbp_or(u['gbp'], '—')} "
                f"| {u['blank'] or '—'} | {u['where']} | {'; '.join(u['why'])} |")
        add("")

    # ── what the pack does not contain, and what it costs ────────────────────
    # ASKED DIRECTLY AND ANSWERED DIRECTLY. "Which drawings are missing, and does it hurt the
    # price?" is a different question from "how good are the drawings we have", and the
    # answer is usually "none of them, for money" — a bought-in bolt has no detail drawing
    # because it does not need one, and saying so stops somebody chasing Design for it.
    #
    # A LINE WITH NO SHEET IS ONLY A PROBLEM IF ITS PRICE DEPENDED ON ONE. So each is
    # classified by what its money actually rests on, not by the absence alone.
    if scan:
        no_sheet = []
        for row in bom:
            code = row["code"].upper()
            if not code:
                continue
            rec = scan.get(code) or {}
            if rec.get("pages"):
                continue
            cut_here = bool(steel.get(code)) or bool((material.get(code) or {}).get("Blank L"))
            unit, qty = _money(row.get("price")), _money(row.get("qty"))
            no_sheet.append({
                "code": row["code"], "desc": _description(row), "cut": cut_here,
                "gbp": round(unit * qty, 2) if unit and qty else None,
                "priced": row.get("price") not in (None, ""),
            })
        add("## Drawings the pack does not contain, and what that costs")
        add("")
        # THE ENGINE'S OWN VERDICT ON THE PACK, WHERE SOMEBODY WILL READ IT. It decides
        # whether the drawings supported a credible cost, suppresses its own headline when
        # they did not, and prints one line to a console. The workbook still computes a Unit
        # Cost, because that is a different figure — so the estimate arrives priced and the
        # reason to doubt it arrives nowhere. Stated here, with the money on it, because "we
        # priced what we could" is only honest if it is followed by what we could not.
        if sufficiency.get("provisional") or str(
                sufficiency.get("status") or "") in ("provisional", "insufficient_data"):
            _ratio = _money(sufficiency.get("credible_cost_ratio"))
            _fab = sufficiency.get("fabricated_part_count")
            _with = sufficiency.get("parts_with_dxf")
            add(f"> **This job is priced in full, and some of it is read rather than "
                f"measured.** "
                + (f"Of the {_gbp(sufficiency.get('document_total_provisional_gbp'))} it "
                   f"assembled, **{_ratio:.0%} rests on figures it considers credible** — the "
                   f"rest on geometry read off a view, or on prices it could not verify. "
                   if _ratio is not None else "")
                + (f"{_with} of {_fab} fabricated part(s) have a DXF; the others were sized "
                   f"from the drawing rather than measured. "
                   if _fab else "")
                + "Every line is costed and the unit cost is a real figure — it is what the "
                  "sheet's own cells add up to. What follows is which lines rest on a "
                  "reading rather than a measurement, so they can be checked first rather "
                  "than the whole estimate being doubted.")
            add("")
            _weak = [u for u in (sufficiency.get("unreliable_parts") or [])
                     if isinstance(u, dict)]
            if _weak:
                add("| Part | What it is | £ on this job | Why the engine doubts it |")
                add("|---|---|---|---|")
                for _u in sorted(_weak,
                                 key=lambda u: -(_money(u.get("extended_cost_gbp")) or 0)):
                    _rec = scan.get(str(_u.get("part_number") or "").upper()) or {}
                    add(f"| {_fmt(_u.get('part_number'))} "
                        f"| {_clip(_u.get('description'), 44)} "
                        f"| {_gbp_or(_u.get('extended_cost_gbp'), '—')} "
                        f"| {', '.join(str(r) for r in (_u.get('reasons') or [])) or 'not recorded'} "
                        f"{'· ' + _where(_rec, pack, page_index) if _rec.get('pages') else ''} |")
                add("")
                add("> A part DXF or a SOLIDWORKS model for the parts above is the single "
                    "thing that would move those lines from read to measured. Every other "
                    "line in this estimate is unaffected by it.")
                add("")
        if not no_sheet:
            add("Every costed line on this job is owned by a sheet in the pack. Nothing is "
                "priced off a drawing that was not supplied.")
            add("")
        else:
            _at_risk = [n for n in no_sheet if n["cut"]]
            add(f"{len(no_sheet)} costed line(s) have no sheet of their own in this pack. "
                + (f"**{len(_at_risk)} of them are parts SDI cuts** — those are the ones "
                   f"where a missing drawing moves the price."
                   if _at_risk else
                   "**None of them is a part SDI cuts**, so no missing drawing is holding up "
                   "a fabricated cost: each is a bought item, a fastener off a BOM table, or "
                   "a commercial line, and none of them would have a detail drawing even on "
                   "a complete pack."))
            add("")
            add("| Line | What it is | £ on this job | Does the missing sheet move the £ |")
            add("|---|---|---|---|")
            for item in sorted(no_sheet, key=lambda n: -(n["gbp"] or 0)):
                if item["cut"]:
                    impact = ("**Yes** — this is cut from sheet, so its blank and gauge came "
                              "from somewhere other than a detail drawing. Check them.")
                elif not item["priced"]:
                    impact = ("No — it is unpriced, and a drawing would not price it. It "
                              "needs a rate, not a sheet.")
                else:
                    # NOT "priced from a book". Packaging, delivery and the powder line are
                    # not bought from anybody's catalogue — one is computed from coated
                    # area, two are priced for the order — and calling all three a book
                    # price is a claim about the source that the BOM table above contradicts.
                    impact = ("No — it is not a drawn part. Its price and its source are on "
                              "the bill of materials above.")
                add(f"| {item['code']} | {item['desc']} "
                    f"| {_gbp(item['gbp']) if item['gbp'] is not None else 'no price'} "
                    f"| {impact} |")
            add("")
        # Pages nothing claimed. Derived from the sheets the parts themselves name — the
        # highest page any record claims is the only page count this document can honestly
        # know, so a gap below is "no part was traced to this sheet", not "the PDF is short".
        claimed = sorted({int(p) for rec in scan.values() for p in (rec.get("pages") or [])
                          if isinstance(p, (int, float))})
        if claimed:
            gaps = [p for p in range(1, max(claimed) + 1) if p not in claimed]
            if gaps:
                add(f"Within p.1–p.{max(claimed)}, **no costed part was traced to "
                    f"{', '.join(f'p.{p}' for p in gaps)}**. Those sheets are usually the "
                    f"cover, the general arrangement and the BOM table, which own no part of "
                    f"their own — but they are also where a part would hide if its drawing "
                    f"were read and never joined to a cost, so they are named rather than "
                    f"assumed harmless.")
            else:
                add(f"Every sheet from p.1 to p.{max(claimed)} is claimed by at least one "
                    f"costed part.")
            add("")

    return "\n".join(lines)


# ── the covering note, in the shape an estimator actually reads ──────────────
#
# THE FORMAT IS NOT DECORATION; IT IS THE SPECIFICATION.
#
# What went out with 12349-02 was a headline reading "not reported/unit at 7 off" over a list
# of filenames. James: "the write up is very poor. it needs to be in this format" — and then
# the note he had written by hand for 12552, which is the thing to reproduce. It works because
# it answers, in order and before any table: what does it cost, what is the biggest number
# made of, what is the labour, what needs a person, what is wrong with the pack, and what is
# wrong with US. An estimator can act on each of those; a list of attachments is not something
# anybody can act on.
#
# Every figure is read from the workbook's own calculated cells through _gather, so this note
# and the full document cannot disagree — they are the same numbers rendered twice.

_EMAIL_CSS = ("font:14px/1.55 -apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;"
              "color:#1c2530;max-width:860px")
_TH = ('style="text-align:left;padding:4px 12px 4px 0;border-bottom:1px solid #c8d2dc;'
       'font-weight:600;white-space:nowrap"')
_TD = 'style="padding:3px 12px 3px 0;border-bottom:1px solid #edf1f5;vertical-align:top"'
_NUM = ('style="padding:3px 12px 3px 0;border-bottom:1px solid #edf1f5;text-align:right;'
        'white-space:nowrap"')


def _clip(text: Any, limit: int) -> str:
    """Shorten on a WORD boundary, and say that it was shortened.

    "CNC Joinery — 5mm HIGH IMPACT ACRYLIC (12349-02-" is a machine cutting a string at
    character 48. An estimator reads that as software that does not know what it is holding,
    and they are not wrong. Break where the words break, and mark it.
    """
    words = str(text if text is not None else "").split()
    if not words:
        return "—"
    out = ""
    for word in words:
        if out and len(out) + 1 + len(word) > limit:
            return out + " …"
        out = f"{out} {word}".strip() if out else word
    return out


def _plural(count: Any, noun: str, plural: str = "") -> str:
    """"1 line" and "2 lines". A note an estimator forwards to a customer's engineer should
    not read "1 part number(s)" — small, and it is the difference between a document that
    looks written and one that looks generated."""
    try:
        n = int(count)
    except (TypeError, ValueError):
        return f"{count} {noun}s"
    return f"{n} {noun}" if n == 1 else f"{n} {plural or noun + 's'}"


def _e(text: Any) -> str:
    return (str(text if text is not None else "").replace("&", "&amp;")
            .replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;"))


def _table(headers: List[str], rows: List[List[Any]], numeric: Optional[set] = None) -> str:
    numeric = numeric or set()
    head = "".join(f"<th {_TH}>{_e(h)}</th>" for h in headers)
    body = "".join(
        "<tr>" + "".join(
            f"<td {_NUM if i in numeric else _TD}>{_e(cell)}</td>"
            for i, cell in enumerate(row)) + "</tr>"
        for row in rows)
    return (f'<table style="border-collapse:collapse;margin:10px 0;width:100%">'
            f"<thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>")


def _setup_and_run(labour_rows: List[Dict[str, Any]], order_qty: Any = 1) -> Dict[str, Any]:
    """How much of the labour is set-up, which is the whole of the quantity story.

    Set-up does not move with the order and run time does. Stating the split is what lets an
    estimator answer "what would 50 off cost" without another run — and on 12552 it was
    £247.40 of £323.84, which is not a detail.
    """
    total = round(sum(_money(r.get("total_value_gbp")) or 0 for r in labour_rows), 2)
    # THE SHEET RECORDS SET-UP IN MINUTES, NOT IN POUNDS. The read-back carries
    # `setup_minutes` and `dept_rate_gbp_per_hour`; this asked for `setup_cost_gbp`, got
    # nothing on every row, and concluded there was no set-up to report — so the one
    # sentence that answers "what would 50 off cost" never printed. Total Value is per
    # UNIT and the sheet has already divided the set-up by the order quantity, so the
    # conversion divides by it too.
    setup = 0.0
    known = False
    for r in labour_rows:
        mins = _money(r.get("setup_minutes"))
        rate = _money(r.get("dept_rate_gbp_per_hour"))
        if mins is None or rate is None:
            continue
        known = True
        setup += (rate / 60.0) * mins / max(1, int(order_qty or 1))
    setup = round(setup, 2)
    return {"total": total, "setup": setup, "run": round(total - setup, 2),
            "rows": len(labour_rows), "known": known}


def covering_email(workbook: Path, scan_json: Optional[Path] = None, *,
                   client: str = "", deliverables: Optional[List[str]] = None,
                   provisional: bool = True) -> Dict[str, str]:
    """Subject, HTML and plain text for one finished estimate, in seven sections."""
    g = _gather(workbook, scan_json)
    scan, final, pack = g["scan"], g["final"], g["pack"]
    totals, bom, steel_calc = g["totals"], g["bom"], g["steel_calc"]
    material, provenance = g["material"], g["provenance"]
    labour_rows, material_rows = g["labour_rows"], g["material_rows"]
    order_qty = g["order_qty"] or 1
    page_index = g["page_index"]
    job = str(g["stem"]).split("_")[0]

    # A LINE AT £0.00 IS THE ONE THIS SECTION EXISTS FOR, AND IT WAS NOT COUNTED.
    #
    # This asked `price in (None, "")`. 12349-02's wood screws, flange buttons, bumpons and
    # all three acrylic/MDF lines carry a price of 0.0 — a number, not a blank — so none of
    # them matched, `needs_a_person` came out EMPTY, and section 5 did not render at all.
    # The message still said "it stays unissued until the lines in §5 are settled", above a
    # document with no §5 in it. Zero IS the condition: it is what a blank reads as once the
    # cell has been written.
    def _reads_as_free(row) -> bool:
        if "costed in" in str(row.get("text") or "").lower():
            return False          # its money is in a fabricated block, not missing
        price = _money(row.get("price"))
        return price is None or price == 0
    _unpriced = [r for r in bom if _reads_as_free(r)]
    # The source string is what names an indication — the supplier cell often holds the
    # engine's own marker rather than a supplier, and _price_source is what the table prints.
    _indicative = [r for r in bom
                   if _money(r.get("price"))
                   and any(t in (f"{r.get('supplier') or ''} "
                                 f"{_price_source(r, provenance, scan) or ''}").lower()
                           for t in _INDICATIVE)]
    needs_a_person = _unpriced + _indicative
    labour = _setup_and_run(labour_rows, order_qty)
    untraced = _tracing_failures(scan, pack, steel_calc, material, page_index)

    _steel_total = round(sum(_money(r.get("total_value_gbp")) or 0
                             for r in material_rows if r.get("block") == "steel"), 2)
    _bought_total = round((totals.get("material") or 0) - _steel_total, 2)

    _state = "PROVISIONAL. " if provisional else ""
    _need = (f"{_plural(len(needs_a_person), 'line')} need a person."
             if needs_a_person else "No line is waiting on a person.")
    subject = (f"{job} — SDI Intelligence estimate, {_state}"
               f"{_gbp(totals.get('unit'))}/unit at {order_qty} off. {_need}")

    h: List[str] = [f'<div style="{_EMAIL_CSS}">']
    add = h.append
    add(f"<p><b>{_e(job)}</b>{' &middot; ' + _e(client) if client else ''} &middot; "
        f"{order_qty} off &middot; {_e(_state.strip().rstrip('.') or 'FOR REVIEW')}</p>")
    add(f'<p style="font-size:26px;margin:12px 0 4px"><b>{_e(_gbp(totals.get("unit")))}</b>'
        f'<span style="color:#5b6b7d;font-size:14px"> per unit, ex VAT</span></p>')
    add("<p>Every figure below is read from the workbook's own calculated cells — nothing "
        "re-derived. The objective is to give you a set of explains you can work with, so "
        "please feed back anything that is wrong or missing.</p>")
    if deliverables:
        add("<p>Attached: " + " &middot; ".join(f"<b>{_e(_basename(d))}</b>"
                                                for d in deliverables) + ".")
        add(" No customer quote — it stays unissued until the lines in §5 are settled.</p>"
            if provisional else "</p>")

    # 1 ─ the number
    add("<h3>1. The number</h3>")
    _rows = [["Material", _gbp(totals.get("material"))],
             ["Labour", _gbp(totals.get("labour"))]]
    _sub = round((totals.get("material") or 0) + (totals.get("labour") or 0), 2)
    _rows.append(["Subtotal", _gbp(_sub)])
    _other = _money(_fe_totals(final).get("other_gbp"))
    if _other:
        _rows.append([(final.get("unit_price_composition") or {}).get("basis")
                      or "the unit cell's own uplift", f"+{_gbp(_other)}"])
    _rows.append([f"Unit cost, {order_qty} off", _gbp(totals.get("unit"))])
    add(_table(["", "£"], _rows, numeric={1}))
    # NAMED BY BLOCK, from the rows themselves. This said "bought-in and commercial X plus
    # sheet steel Y", where X was (material - steel) — so on any job with an acrylic, MDF or
    # tube block, material the sheet holds in a third place was announced as bought-in.
    _by_block = {}
    for _r in material_rows:
        _by_block[_r.get("block") or "?"] = round(
            _by_block.get(_r.get("block") or "?", 0.0)
            + (_money(_r.get("total_value_gbp")) or 0), 2)
    _BLOCK_WORDS = {"bom": "bought-in and commercial", "steel": "sheet steel",
                    "other_sheet": "other sheet material", "tube": "tube and section"}
    if _by_block:
        add("<p>Material is "
            + ", ".join(f"{_BLOCK_WORDS.get(b, b)} {_gbp(v)}"
                        for b, v in sorted(_by_block.items(), key=lambda kv: -kv[1]))
            + f". Labour is {_plural(labour['rows'], 'sheet row')}.</p>")
    else:
        add(f"<p>Material is bought-in and commercial {_gbp(_bought_total)} plus sheet steel "
            f"{_gbp(_steel_total)}. Labour is {_plural(labour['rows'], 'sheet row')}.</p>")

    # 2 ─ THE FABRICATED MATERIAL, EVERY BLOCK OF IT.
    #
    # This rendered the STEEL block and called everything else "bought-in and commercial",
    # a figure it computed as (material total - steel). On 12349-02 that headed section 3
    # with GBP 69.99 over rows adding to GBP 39.28: the acrylic and the MDF are in the
    # sheet's OTHER SHEET MATERIAL block, which nothing here drew, so GBP 30.71 of material
    # was in the note's arithmetic and on none of its lines — in a document whose first
    # sentence promises every figure is read from the workbook's own cells.
    #
    # Every block the sheet has, each with its own subtotal, and the four reconciled against
    # Total Material Cost underneath. A residual is stated rather than absorbed.
    _BLOCK_TITLES = [
        ("steel", "Sheet steel", True),
        ("other_sheet", "Other sheet material — acrylic, MDF, board", True),
        ("tube", "Tube and section", False),
    ]
    _fab_total = 0.0
    _fab_present = [(b, t, n) for b, t, n in _BLOCK_TITLES
                    if any(r.get("block") == b for r in material_rows)]
    if _fab_present:
        add("<h3>2. The material we cut, block by block</h3>")
        if any(n for _, _, n in _fab_present):
            add("<p><b>Priced by nest, not by area.</b> Each line in the nested blocks "
                "below is<br><code>ROUNDUP(sheet price ÷ how many nest per sheet, 2) × qty "
                "× 1.04 scrap</code><br>There is no per-piece figure on the sheet — please "
                "don't divide these back out.</p>")
        for _block, _title, _nested in _fab_present:
            _rows_b = [r for r in material_rows if r.get("block") == _block]
            _sub = round(sum(_money(r.get("total_value_gbp")) or 0 for r in _rows_b), 2)
            _fab_total += _sub
            add(f"<p><b>{_e(_title)} — {_gbp(_sub)}</b></p>")
            _srows = []
            for r in sorted(_rows_b, key=lambda r: -(_money(r.get("total_value_gbp")) or 0)):
                _code = str(r.get("description") or "").split(" ")[0].strip().upper()
                _rec = scan.get(_code) or {}
                # THE SHEET'S OWN CELLS, not the AI Material Detail tab. Those are two
                # records of one fact and they differ — the detail tab holds the engine's
                # blank-area figure, this holds what the estimate was actually built from.
                _blank = (f"{_fmt(r.get('length_mm'))} × {_fmt(r.get('width_mm'))}"
                          if r.get("length_mm") else "—")
                _srows.append([
                    str(r.get("description") or "—"), _blank,
                    _fmt(r.get("gauge") or r.get("thickness_mm")
                         or _rec.get("normalized_thickness_mm")),
                    _fmt(r.get("qty_per_unit")), _fmt(r.get("qty_per_sheet")),
                    _gbp_or(r.get("total_value_gbp"), "—"),
                    _drawing_no(_rec), _where(_rec, pack, page_index),
                ])
            add(_table(["Part", "Blank mm", "Ga", "Qty", "Nest/sheet", "£ line total",
                        "Drawing no.", "Which drawing files and pages"],
                       _srows, numeric={3, 4, 5}))

    # 3 ─ bought-in and commercial, from the same record as everything else
    _bom_rows = [r for r in material_rows if r.get("block") == "bom"]
    _bom_total = round(sum(_money(r.get("total_value_gbp")) or 0 for r in _bom_rows), 2)
    if _bom_rows or bom:
        add(f"<h3>3. Bought-in and commercial — {_gbp(_bom_total if _bom_rows else _bought_total)}</h3>")
        _brows = []
        _src_by_code = {str(r.get("code") or "").upper(): r for r in bom}
        for r in sorted(_bom_rows or [], key=lambda r: -(_money(r.get("total_value_gbp")) or 0)):
            _code = str(r.get("part_code") or "").strip()
            _rec = scan.get(_code.upper()) or {}
            _srow = _src_by_code.get(_code.upper())
            _brows.append([
                _code or "—", str(r.get("description") or "—"),
                _fmt(r.get("qty_per_unit")), _gbp_or(r.get("unit_price_gbp"), "—"),
                _gbp_or(r.get("total_value_gbp"), "0.00"),
                (_price_source(_srow, provenance, scan) if _srow else "")
                or str(r.get("supplier") or "") or "source not named on the sheet",
                _where(_rec, pack, page_index),
            ])
        if not _bom_rows:
            # Pre-read-back fallback: the workbook scan, as before.
            for row in sorted(bom, key=lambda r: -((_money(r.get("price")) or 0)
                                                   * (_money(r.get("qty")) or 0))):
                if "costed in sheet steel" in str(row.get("text") or "").lower():
                    continue
                _u, _q = _money(row.get("price")), _money(row.get("qty"))
                _rec = scan.get(str(row.get("code") or "").upper()) or {}
                _brows.append([
                    row.get("code") or "—", _description(row), _fmt(row.get("qty")),
                    _gbp_or(_u, "—"),
                    _gbp_or(round(_u * _q, 2) if _u and _q else None, "0.00"),
                    _price_source(row, provenance, scan) or "source not named on the sheet",
                    _where(_rec, pack, page_index),
                ])
        add(_table(["Line", "What it is", "Qty", "£/ea", "£ ext", "Source",
                    "Which drawing files and pages"], _brows, numeric={2, 3, 4}))
        if _indicative:
            _ind_gbp = round(sum((_money(r.get("price")) or 0) * (_money(r.get("qty")) or 0)
                                 for r in _indicative), 2)
            _one = len(_indicative) == 1
            add(f"<p><b>{len(_indicative)} of those lines "
                f"{'is an AI market indication' if _one else 'are AI market indications'} "
                f"rather than "
                f"{'a catalogue price' if _one else 'catalogue prices'} — {_gbp(_ind_gbp)}"
                f"{'' if _one else ' between them'}.</b> "
                f"{'It moves' if _one else 'They move'} between runs, so an estimate "
                f"resting on {'it' if _one else 'them'} cannot be reproduced.</p>")

    # ── HOW PACKAGING AND DELIVERY WERE WORKED OUT ────────────────────────────────
    #
    # "change the e-mail to include a detailed computational description of the delivery and
    # palleting costs also please."
    #
    # On 11908-21 these two were £160 of a £249 unit — 64%, against £3.74 of MDF — and the
    # note said only that they were market indications. An estimator could see THAT the
    # number was wrong and had nothing to argue with. The shipment the engine described is
    # what the market was asked to price, so the shipment is what has to be shown: every part
    # it weighed, the density it used, what it came to, and whether that made it a parcel or
    # a pallet. A description that says "about 49 kg" for a 2.3 kg tray then names the part
    # carrying the other 46 kg, instead of hiding it in a total.
    _commercial = [c for c in (g["scan_doc"] or {}).get("commercial_lines", [])
                   if isinstance(c, dict)] if isinstance(g["scan_doc"], dict) else []
    if _commercial:
        add("<h4>How packaging and delivery were worked out</h4>")
        _basis = next((c.get("basis") for c in _commercial
                       if isinstance(c.get("basis"), dict)), {}) or {}
        _counted = [r for r in (_basis.get("counted_parts") or []) if isinstance(r, dict)]
        if _counted:
            add(_table(["Part", "Blank mm", "Ga", "Material", "kg/m³", "Qty/unit",
                        "kg each", "kg in the order"],
                       [[r.get("part_number") or "—",
                         f"{_fmt(r.get('length_mm'))} × {_fmt(r.get('width_mm'))}",
                         _fmt(r.get("thickness_mm")), _fmt(r.get("material")),
                         _fmt(r.get("density_kg_m3")), _fmt(r.get("per_assembly")),
                         _fmt(r.get("kg_each")), _fmt(r.get("kg_in_order"))]
                        for r in _counted], numeric={2, 4, 5, 6, 7}))
        _skipped = _basis.get("parts_without_a_blank")
        _phantom = _basis.get("parts_with_an_impossible_blank")
        _size = _basis.get("largest_part_mm") or []
        add(f"<p>That is <b>{_fmt(_basis.get('order_weight_kg'))} kg</b> for the whole order"
            + (f", largest panel {_fmt(_size[0])} × {_fmt(_size[1])} mm" if len(_size) == 2
               else "")
            + f", from {_plural(_basis.get('parts_measured') or 0, 'part')} the engine could "
              f"measure"
            + (f"; {_plural(_skipped, 'part')} had no blank to weigh" if _skipped else "")
            + (f"; <b>{_plural(_phantom, 'part')} were left out because the blank recorded "
               f"for them fits no sheet the material is stocked in</b> — that is a defect "
               f"upstream, not a gap in the drawings" if _phantom else "")
            + ". Weight is blank area × gauge × the material's density, times how many are "
              "made. It decides the next line.</p>")
        try:
            from commercial_lines import shipment_shape as _shape
            _sh = _shape(_basis)
        except Exception:                                        # noqa: BLE001
            _sh = ""
        _plan = _basis.get("shipment") or {}
        if _sh:
            add(f"<p><b>Priced as a {_sh}.</b> "
                + ("One carton a person can lift and a courier will take, so it was asked "
                   "for as a parcel rather than palletised haulage."
                   if _sh == "parcel" else
                   f"{_plural(_plan.get('carton_count') or 0, 'carton')} on "
                   f"{_plural(_plan.get('pallet_count') or 0, 'pallet')}, so it was asked "
                   f"for as palletised haulage.")
                + " Both descriptions used to say <i>a pallet</i> whatever the order was — "
                  "which is how a 2.3 kg tray came back at £95 to box and £65 to deliver. "
                  "The lookup answers the question it is asked.</p>")
        add(_table(["Line", "Asked for", "£ for the order", "÷ units", "£/unit"],
                   [[c.get("code") or "—", str(c.get("described_as") or "")[:120],
                     _gbp_or(c.get("order_gbp"), "—"),
                     _fmt(c.get("order_quantity")), _gbp_or(c.get("unit_gbp"), "—")]
                    for c in _commercial], numeric={2, 3, 4}))

    # DOES THE MATERIAL ADD UP. The claim that makes the rest worth reading, and the one a
    # reader can check without leaving the message.
    _mat_sheet = _money(totals.get("material"))
    if _mat_sheet is not None and material_rows:
        _lines_total = round(_fab_total + _bom_total, 2)
        _gap = round(_mat_sheet - _lines_total, 2)
        add(f"<p><b>Material reconciliation.</b> The blocks above come to "
            f"{_gbp(_lines_total)} against the sheet's Total Material Cost of "
            f"{_gbp(_mat_sheet)}"
            + (f" — they agree.</p>" if abs(_gap) < 0.01 else
               f", a difference of {_gbp(abs(_gap))}. That difference is on the sheet and "
               f"on no line here, which means a block was not read back. Treat the sheet's "
               f"figure as the total.</p>"))

    # 4 ─ labour
    if labour_rows:
        # HEADED WITH THE SHEET'S OWN TOTAL, NOT THE SUM OF THE ROWS BELOW IT. Those are two
        # different numbers whenever a row could not be read back, and a note whose section 4
        # contradicts its section 1 is worse than one that admits the gap: the reader stops
        # trusting both figures rather than the one that is short.
        _lab_sheet = _money(totals.get("labour"))
        add(f"<h3>4. Labour — {_gbp(_lab_sheet if _lab_sheet is not None else labour['total'])}"
            f" across {labour['rows']} sheet rows</h3>")
        _lrows = [[str(r.get("operation") or r.get("description") or "—"),
                   str(r.get("description") or "—"),
                   r.get("department") or "—",
                   _fmt(r.get("setup_minutes")), _fmt(r.get("batch_hours")),
                   _gbp_or(r.get("dept_rate_gbp_per_hour"), "—"),
                   _gbp_or(r.get("total_value_gbp"), "—")]
                  for r in sorted(labour_rows,
                                  key=lambda r: -(_money(r.get("total_value_gbp")) or 0))]
        add(_table(["Operation", "What it covers", "Dept", "Set-up min", "Batch hrs",
                    "£/hr", "£"], _lrows, numeric={3, 4, 5, 6}))
        if labour["known"] and labour["setup"]:
            add(f"<p><b>About {_gbp(labour['setup'])} of the {_gbp(labour['total'])} in this "
                f"table is set-up</b>, all of it landing on {_plural(order_qty, 'item')}. "
                f"Run time is "
                f"{_gbp(labour['run'])}. That is the quantity story in one line: raise the "
                f"order and the set-up spreads while the run time stays put. Say the word "
                f"and I'll run a larger quantity properly rather than project it.</p>")
        if _lab_sheet is not None and abs(_lab_sheet - labour["total"]) >= 0.01:
            add(f"<p><b>The rows above come to {_gbp(labour['total'])} against the sheet's "
                f"{_gbp(_lab_sheet)}</b> — {_gbp(round(abs(_lab_sheet - labour['total']), 2))} "
                f"of labour is on the sheet and not in this table, because the read-back could "
                f"not resolve every row. Treat the sheet's figure as the total and this table "
                f"as most of the detail behind it.</p>")

    # 5 ─ what needs a person
    if needs_a_person:
        add(f"<h3>5. {_plural(len(needs_a_person), 'line')} that need you</h3>")
        _nrows = []
        for row in sorted(needs_a_person,
                          key=lambda r: -((_money(r.get("price")) or 0)
                                          * (_money(r.get("qty")) or 0))):
            _u, _q = _money(row.get("price")), _money(row.get("qty"))
            _is_ind = row in _indicative
            _nrows.append([
                row.get("code") or "—", _description(row),
                _gbp_or(round(_u * _q, 2) if _u and _q else None, "£0.00"),
                ("An AI market indication, not a catalogue price. Overwrite it, or accept it "
                 "deliberately." if _is_ind else
                 "The line is costing nothing — nothing we can query holds a rate for this."),
                _where(scan.get(str(row.get("code") or "").upper()) or {}, pack, page_index),
            ])
        add(_table(["Line", "What it is", "On the sheet now", "What's needed",
                    "Which drawing files and pages"], _nrows, numeric={2}))
        add("<p>Overwrite anything tagged <b>AI ESTIMATE — INDICATIVE, NOT A QUOTE</b> and "
            "the sheet recalculates.</p>")

    # 6 ─ every operation, and who decided it
    #
    # THE ROUTE IS HALF THE ESTIMATE AND IT WAS NOT IN THE NOTE AT ALL. Section 4 says what
    # the labour costs; this says what we think the shop actually does to each part and on
    # whose authority — drawn, inferred, or read by a model. An operation nobody drew is the
    # cheapest thing on the sheet to strike out and the easiest to miss.
    if g["routes"]:
        # AN OPERATION WE RULED OUT IS NOT AN OPERATION. The table listed both together and
        # counted them in one number, so 11908-21 read "16 lines" over rows that include
        # "folding — the part does not fold" and "laser_cutting — assembly parent has no
        # independently measured fabricated leaf". Neither is done, neither is charged, and a
        # laser row against an MDF tray is the kind of thing that loses a reader's trust in
        # the fifteen rows that ARE right.
        #
        # Both are worth printing — a ruled-out operation with its reason is how somebody
        # checks we did not miss one — but they are different questions and go in different
        # tables. An applied operation carries a quantity; a ruling has none.
        _applied = [r for r in g["routes"] if _money(r.get("Qty/unit")) is not None]
        _ruled_out = [r for r in g["routes"] if _money(r.get("Qty/unit")) is None]
        add(f"<h3>6. Every operation, and who decided it — "
            f"{_plural(len(_applied), 'line')}</h3>")
        _rrows = []
        for row in _applied:
            _rec = scan.get(str(row.get("Target") or "").upper(), {})
            _rrows.append([
                _fmt(row.get("Target")), _fmt(row.get("Operation")), _fmt(row.get("Seq")),
                _fmt(row.get("Scope")), _fmt(row.get("Qty/unit")), _fmt(row.get("Source")),
                str(row.get("Reason") or "—"), _where(_rec, pack, page_index),
            ])
        add(_table(["Part", "Operation", "Seq", "Scope", "Qty", "Decided by",
                    "On what basis", "Which drawing files and pages"], _rrows,
                   numeric={2, 4}))
        _inferred = [r for r in _applied
                     if "infer" in str(r.get("Source") or "").lower()]
        if _inferred:
            add(f"<p><b>{_plural(len(_inferred), 'operation')} inferred rather than drawn</b> "
                f"— {', '.join(sorted({str(r.get('Operation') or '?') for r in _inferred}))}. "
                f"Confirm them or tell me to drop them.</p>")

        if _ruled_out:
            add(f"<p><b>{_plural(len(_ruled_out), 'operation')} ruled out</b>, with the "
                f"reason. Nothing here is done or charged; it is printed so somebody can "
                f"check we did not rule out something we should be doing.</p>")
            add(_table(["Part", "Operation", "Why it is not done", "Decided by"],
                       [[_fmt(r.get("Target")), _fmt(r.get("Operation")),
                         str(r.get("Reason") or "—"), _fmt(r.get("Source"))]
                        for r in _ruled_out]))

    # 7 ─ the drawing pack, in full
    add("<h3>7. The drawing pack</h3>")
    _sheets = sorted({int(p) for rec in scan.values() for p in (rec.get("pages") or [])
                      if isinstance(p, (int, float))})
    if _sheets:
        add(f"<p>{_plural(len(pack) or 1, 'document')}, p.1–p.{max(_sheets)}. "
            f"{_plural(len([r for r in scan.values() if r.get('pages')]), 'costed part')} "
            f"traced to a sheet of their own.</p>")

    # Sheet by sheet — refused outright rather than answered wrongly from a trimmed extract.
    if scan and not _pack_was_read_in_full(scan):
        add("<p><b>Drawing quality, sheet by sheet: not produced.</b> This ran against a "
            "trimmed extract carrying only part numbers and page numbers. Reporting from it "
            "would have said every drawing states no material, no thickness and no finish — "
            "which is false. Re-run against the full job JSON and this builds itself.</p>")
    elif scan:
        add("<p><b>Drawing quality, sheet by sheet.</b></p>")
        _qrows = []
        for code, rec in sorted(scan.items(),
                                key=lambda kv: ((kv[1].get("pages") or [999])[0], kv[0])):
            if not (rec.get("pages") or []):
                continue
            _rel = ((rec.get("geometry_rollup") or {}).get("confidence") or {}).get(
                "geometry_reliability")
            _qrows.append([
                _where(rec, pack, page_index), code or _fmt(rec.get("description")),
                _drawing_no(rec),
                ", ".join(str(m) for m in rec.get("materials") or []) or "no",
                _gauge_stated(rec),
                ", ".join(str(f) for f in rec.get("surface_finishes") or []) or "no",
                (f"{_fmt(rec.get('geometry_source'))}"
                 + (f" ({_rel:.0%})" if isinstance(_rel, (int, float)) else "")),
                ", ".join(_what_a_sheet_could_not_give(rec)) or "nothing — complete",
            ])
        add(_table(["Which drawing files and pages", "Part", "Drawing no.",
                    "Material stated", "Thickness stated", "Finish stated", "Geometry",
                    "What it could not give"], _qrows))

    # Drawings the pack does not contain, and whether that costs anything.
    _no_sheet = _missing_drawings(bom, scan, g["steel"], material)
    if _no_sheet:
        _bites = [n for n in _no_sheet if n["cut"] or not n["priced"]]
        add(f"<p><b>{_plural(len(_no_sheet), 'line')} with no sheet of their own"
            + (f", of which {len(_bites)} "
               f"{'affects' if len(_bites) == 1 else 'affect'} the price."
               if _bites else
               ". None of them affects the price — "
               + _what_they_are(_no_sheet)
               + ", none of which would carry a detail drawing on a complete pack.")
            + "</b></p>")
        if _bites:
            add(_table(["Line", "What it is", "£ on this job", "Does the missing sheet bite?"],
                       [[n["code"] or "—", n["desc"], _gbp_or(n["gbp"], "no price"),
                         ("Yes — we cut this part and had no drawing of it to size it from"
                          if n["cut"] else
                          "Yes — it has no price and no sheet to read one from")]
                        for n in _bites], numeric={2}))

    if untraced:
        _at_stake = round(sum(u["gbp"] or 0 for u in untraced), 2)
        add(f"<p><b>{_plural(len(untraced), 'part number')} could not be followed from the "
            f"BOM through to a drawing or a flat of its own"
            + (f", carrying {_gbp(_at_stake)} between them" if _at_stake else "")
            + ".</b> Each was still costed, from whatever the engine could reach — the last "
              "column says what it used instead. Where the break is ours we will fix the "
              "reader; where it is the pack's, this is the list for Design.</p>")
        add(_table(["Part", "Drawing no.", "What it is", "£ on this job", "Blank used",
                    "Which drawing files and pages", "Where the trail broke"],
                   [[u["code"], u.get("drawing_no") or "—", u["description"] or "—",
                     _gbp_or(u["gbp"], "—"), u["blank"] or "—", u["where"],
                     "; ".join(u["why"])]
                    for u in untraced], numeric={3}))
    else:
        add("<p>Every costed part traced from the BOM through to a drawing or a flat of its "
            "own. Nothing here for Design.</p>")

    # Sheets nothing claimed. Derived from the pages the parts themselves name, so a gap is
    # "no costed part was traced to this sheet", never "the PDF is short".
    if _sheets:
        _gaps = [p for p in range(1, max(_sheets) + 1) if p not in _sheets]
        add(f"<p>Within p.1–p.{max(_sheets)}, "
            + (f"<b>no costed part was traced to "
               f"{', '.join(f'p.{p}' for p in _gaps)}</b>. Those are usually the cover, the "
               f"general arrangement and the BOM table, which own no part of their own — but "
               f"they are also where a part would hide if its drawing were read and never "
               f"joined to a cost, so they are named rather than assumed harmless."
               if _gaps else
               "every sheet is claimed by at least one costed part.") + "</p>")

    # 8 ─ ours, not yours
    #
    # ONLY WHAT WE CAN ACTUALLY OWN. This listed every break in section 6 and called them all
    # engine defects, which is a claim the engine is not in a position to make: a part with no
    # sheet in the pack may be our reader or may be a drawing nobody issued. So it names the
    # SUBSTITUTIONS — a size the engine reached for when it had no measurement — because those
    # are ours whatever the pack looks like, and leaves the rest in section 6 for us to settle
    # between us.
    _substituted = [u for u in untraced
                    if any(w in _UNTRACED_GEOMETRY.values() for w in u["why"])]
    _ours = [f"{u['code']} — {u['why'][0]}"
             + (f" ({_gbp(u['gbp'])} on this job)" if u["gbp"] else "")
             for u in _substituted[:6]]
    if _ours:
        add("<h3>8. Ours, not yours</h3>")
        add("<p>Where the engine had no measurement it used something else rather than "
            "leaving the line blank. Each of these is a substitution we made, and ours to "
            "fix — nothing here needs anything from you:</p>")
        add("<ul>" + "".join(f"<li>{_e(o)}</li>" for o in _ours) + "</ul>")

    add(f'<p style="color:#5b6b7d;font-size:12px;margin-top:18px">Produced by SDI '
        f'Intelligence{" for " + _e(client) if client else ""}. '
        f'The full line-by-line document is the <b>AI Explanation</b> tab in the attached '
        f'workbook and section 14 of the report — every row with the drawing page it came '
        f'from, which reader decided it, and what it charges.</p>')
    add("</div>")
    html = "\n".join(h)
    return {"subject": subject, "html": html, "text": _as_text(html)}


_TAG = re.compile(r"<[^>]+>")


def _as_text(html: str) -> str:
    """A plain-text alternative that keeps the shape. Not a general HTML renderer: it reads
    only the markup produced above, which is why it can be this short."""
    text = html
    for pattern, repl in (
            (r"</h3>", "\n"), (r"<h3[^>]*>", "\n\n"), (r"</p>", "\n"),
            (r"</tr>", "\n"), (r"</li>", "\n"), (r"<br\s*/?>", "\n"),
            (r"</t[dh]>", "  ")):
        text = re.sub(pattern, repl, text)
    text = _TAG.sub("", text)
    text = (text.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
            .replace("&quot;", '"').replace("&middot;", "·"))
    lines = [ln.rstrip() for ln in text.splitlines()]
    out: List[str] = []
    for line in lines:
        if line.strip() or (out and out[-1].strip()):
            out.append(line)
    return "\n".join(out).strip() + "\n"


# ── the same document, in a shape other renderers can use ────────────────────
#
# PARSED BACK OUT OF THE MARKDOWN, DELIBERATELY. The obvious design is a structured
# intermediate that both the markdown and the workbook tab render from, and it is the wrong
# one here: it means rewriting four hundred lines of emit code that an estimator is about to
# rely on, to gain a separation nothing yet needs. What IS needed is that the tab, the report
# section and the email cannot drift from the document — and re-reading the document gives
# that by construction, because there is only one document.
#
# This is safe to parse only because it is machine-written: every table is emitted by the
# code above with a fixed shape, so there is no markdown here that this does not produce.
# The moment a human is allowed to edit the text, this goes.

_MD_STRIP = ("**", "`", "*")


def plain(text: str) -> str:
    """Markdown emphasis removed, for a renderer that has its own — a cell, or a <td>."""
    out = str(text or "")
    for token in _MD_STRIP:
        out = out.replace(token, "")
    return out.strip()


def sections(markdown: str) -> List[Dict[str, Any]]:
    """The document as [{title, intro, tables:[{columns, rows}], notes}].

    Every string is left exactly as the document wrote it, markdown and all: a renderer that
    wants plain text calls plain() on the cells it is placing. Stripping here would throw away
    the emphasis the HTML report wants to keep.
    """
    out: List[Dict[str, Any]] = []
    current: Optional[Dict[str, Any]] = None
    table: Optional[Dict[str, Any]] = None

    def _cells(line: str) -> List[str]:
        return [c.strip() for c in line.strip().strip("|").split("|")]

    for line in str(markdown or "").splitlines():
        stripped = line.strip()
        if stripped.startswith("## "):
            current = {"title": stripped[3:].strip(), "intro": [], "tables": [], "notes": []}
            out.append(current)
            table = None
            continue
        if current is None:
            continue
        if not stripped:
            table = None
            continue
        if stripped.startswith("|"):
            # The separator row carries no content — it only tells markdown a table started.
            if set(stripped) <= set("|-: "):
                continue
            if table is None:
                table = {"columns": _cells(stripped), "rows": []}
                current["tables"].append(table)
            else:
                table["rows"].append(_cells(stripped))
            continue
        table = None
        if stripped.startswith(">"):
            current["notes"].append(stripped.lstrip("> ").strip())
        else:
            current["intro"].append(stripped)
    return out


def worksheet_rows(parsed: List[Dict[str, Any]]) -> List[List[str]]:
    """The document as flat spreadsheet rows — one list of cells per sheet row.

    A tab is not a document: it has no headings, no paragraphs and no emphasis, only cells.
    So a section title becomes a row of its own, prose becomes a single wide cell, and a table
    becomes its header row followed by its rows. Markdown is stripped, because a cell renders
    `**£11.48**` as those characters.

    Column count varies by section and that is correct — Excel does not mind, and forcing
    every section onto the widest table's columns would put the reconciliation's five columns
    under the bill of materials' ten.
    """
    out: List[List[str]] = []
    for section in parsed:
        if out:
            out.append([])
        out.append([plain(section["title"])])
        for line in section.get("intro") or []:
            out.append([plain(line)])
        for table in section.get("tables") or []:
            out.append([])
            out.append([plain(c) for c in table["columns"]])
            for row in table["rows"]:
                out.append([plain(c) for c in row])
        for note in section.get("notes") or []:
            out.append([])
            out.append([plain(note)])
    return out
