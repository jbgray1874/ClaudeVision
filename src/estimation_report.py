"""
SDI Intelligence — Estimation Provenance Report
===========================================
Generates a detailed "how we got there" report alongside
every estimate xlsx. Shows exactly where each material,
thickness, cost and operation came from.
Critical for:
  - MD/FD confidence in AI estimates
  - Estimator trust and verification
  - Learning system audit trail
  - Identifying where AI is uncertain
Output: Added as extra sheet "AI Provenance" in estimate xlsx
        AND as a standalone PDF report (optional)
Called from estimator.py after xlsx is written.
"""
import json
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional
try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter
    _XLSX_OK = True
except ImportError:
    _XLSX_OK = False
try:
    import corrections_db as db
    _DB_OK = True
except ImportError:
    _DB_OK = False

# Reuse the SAME bought-in detection + WB Sell Price finder proven on the Decision
# Report — one source of truth, so both supplementary sheets behave identically.
# Fallbacks keep this module importable if job_decision_report is unavailable.
try:
    from job_decision_report import (_is_bought_in, _find_wb_sell_price_ref,
                                     replace_generated_sheet)
except Exception:  # pragma: no cover - defensive
    def _is_bought_in(part: Dict) -> bool:
        mat = str(part.get("normalized_material") or part.get("material") or "").upper()
        if mat == "BOUGHT_IN":
            return True
        roles = part.get("page_roles") or []
        if "bought_in" in [str(r).lower() for r in roles]:
            return True
        src = str(part.get("source") or "").lower()
        if "recogniser" in src or "bought_in" in src or "note_scan" in src:
            return True
        pn = str(part.get("part_number") or "").upper()
        if pn.startswith(("BI-", "FIXING", "VINYL", "PACKAGING", "DELIVERY")):
            return True
        return False

    def replace_generated_sheet(wb, title):
        if title in wb.sheetnames:
            del wb[title]
        for _n in [n for n in wb.sheetnames
                   if n.startswith(title) and n[len(title):].isdigit()]:
            del wb[_n]
        return wb.create_sheet(title)

    def _find_wb_sell_price_ref(wb):
        try:
            ws = None
            for name in ("Estimate", "estimate"):
                if name in wb.sheetnames:
                    ws = wb[name]
                    break
            if ws is None:
                return None
            for r in ws.iter_rows():
                for c in r:
                    v = c.value
                    if isinstance(v, str) and "sell" in v.lower() and "price" in v.lower():
                        target_col = None
                        for cc in range(c.column + 1, c.column + 9):
                            if ws.cell(row=c.row, column=cc).value not in (None, ""):
                                target_col = cc
                                break
                        if target_col is None:
                            target_col = 13
                        return f"='{ws.title}'!{get_column_letter(target_col)}{c.row}"
            return None
        except Exception:
            return None
# ── Colours ────────────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # SDI dark navy
C_HEADER_FG   = "FFFFFF"
C_HIGH        = "C6EFCE"   # green  — high confidence
C_MEDIUM      = "FFEB9C"   # amber  — medium confidence
C_LOW         = "FFC7CE"   # red    — low confidence / needs review
C_KB          = "DDEEFF"   # blue   — from knowledge base
C_HIST        = "E8D5FF"   # purple — from historical data
C_AI          = "FFF2CC"   # yellow — AI inference
C_RULE        = "D9EAD3"   # light green — override rule fired
C_BOUGHT      = "EDEDED"   # grey   — bought-in / catalogue component
C_SECTION     = "2F5496"   # section header blue
C_ALT_ROW     = "F5F5F5"   # alternating row
def confidence_colour(confidence: float) -> str:
    if confidence >= 0.85:
        return C_HIGH
    elif confidence >= 0.60:
        return C_MEDIUM
    else:
        return C_LOW
def confidence_label(confidence: float) -> str:
    if confidence >= 0.85:
        return "HIGH ✓"
    elif confidence >= 0.60:
        return "MEDIUM"
    else:
        return "LOW — REVIEW"


# The PHYSICAL FACTS the engine read off the drawing and model — what the part is made of, how
# thick, and its geometry. Kept separate from whether a rate has been typed (pricing) and from
# what is DONE to the part (route is a decision, reported on the Decisions tab, not a read), so a
# strong read is never hidden behind a pending price or a routine BOM-sourced route.
_READING_FIELDS = frozenset({"material identity", "thickness", "geometry"})


def reading_and_pricing_counts(provenance):
    """Split a job's parts into a READING readout (how well the engine pulled material,
    thickness, geometry and route off the drawing/model) and a PRICING readout (how many lines
    are priced versus waiting on the estimator's rate).

    Reading is scored on its own fields only — never on the price. An estimate is EXPECTED to
    arrive with prices pending; folding that into one weakest-link band is what made a fully-read
    job report '0 HIGH / 22 LOW' and read as an engine failure. Returns a dict of six counts."""
    from confidence import (STATUS_ORDER, MEASURED, CONFIRMED, REPORTED, UNKNOWN)

    def _weakest(statuses):
        real = [s for s in statuses if s in STATUS_ORDER]
        return min(real, key=STATUS_ORDER.index) if real else None

    out = {"read_high": 0, "read_med": 0, "read_low": 0, "priced": 0, "pending": 0}
    for _p in (provenance or []):
        _fields = _p.get("fields") or []
        _r = _weakest([f.get("status") for f in _fields
                       if f.get("field") in _READING_FIELDS])
        if _r in (MEASURED, CONFIRMED):
            out["read_high"] += 1
        elif _r == REPORTED:
            out["read_med"] += 1
        elif _r is not None:
            out["read_low"] += 1
        _mp = next((f.get("status") for f in _fields
                    if f.get("field") == "material price"), None)
        if _mp == UNKNOWN:
            out["pending"] += 1
        else:
            out["priced"] += 1
    return out
def source_label(source: str) -> str:
    """Human-readable explanation of where a value came from."""
    s = str(source or "").lower()
    if "knowledge_base" in s:
        return "SDI Knowledge Base (previously confirmed)"
    elif "override_rule" in s:
        name = re.search(r'override_rule:(.+)', s)
        rule = name.group(1) if name else "learning rule"
        return f"Learning Rule fired: {rule}"
    elif "dxf_flat_pattern" in s:
        return "DXF flat pattern file (exact geometry)"
    elif "dxf_filename" in s:
        return "DXF filename material code (e.g. _MS_, PETG)"
    elif "historical" in s:
        return "Historical SDI estimate match"
    elif "pn_suffix" in s:
        return "Part number suffix (-M=Steel, -A=Acrylic, -T=MDF)"
    elif "pdf" in s:
        return "PDF drawing text extraction"
    elif "ai" in s or "inference" in s:
        return "AI inference (Claude)"
    else:
        return str(source or "AI inference")
def _price_basis_label(price_source: Dict[str, Any], material: str = "") -> str:
    """One-line 'where did this price come from?' for the provenance sheet, built
    from the engine's own per-part price_source metadata (pricing_service /
    _build_price_source_metadata). Surfaces the source the engine actually used —
    DB price book, historical quote, ERP/UDEF, config rate card — or a loud flag
    when no price source was found. No new pricing logic; display only."""
    if not price_source:
        return "—"
    src   = str(price_source.get("source_name") or "").strip()
    stype = str(price_source.get("source_type") or "").strip()
    supp  = str(price_source.get("supplier_source") or "").strip()
    pdate = str(price_source.get("price_date") or "").strip()
    low   = src.lower()
    _supp = f" — {supp}" if supp and supp.lower() not in low else ""
    if low == "fallback" or "no price" in low or "no_price" in low:
        return "⚠ No price source — add to price book"
    if "bought_in" in low:
        return f"Price book (bought-in){_supp}" + (f", {pdate}" if pdate else "")
    if "historical_quote" in low:
        return f"Historical quote{_supp}"
    if "udef" in low:
        return "UDEF parts table"
    if "pma" in low or "erp" in low:
        return "ERP parts master"
    if stype == "web_ai_fallback" or "web_ai" in low:
        return "⚠ Web/AI fallback — verify"
    if stype == "web_catalog" or "catalog" in low:
        return f"Supplier catalogue{_supp}"
    if "config" in low or stype == "config":
        return "Config rate card"
    if stype == "external":
        return (supp or src) + (f", {pdate}" if pdate else "")
    return src or "—"
def build_provenance(summary: Dict[str, Any]) -> List[Dict]:
    """
    Extract provenance data from a scan summary.
    Returns list of part provenance records.
    """
    # The canonical job part list — the same rows the Estimate sheet was built from, each
    # overlaid on its manufacturing_writeup entry so the provenance fields below survive.
    from costed_facts import job_parts as _job_parts
    parts = _job_parts(summary) or (
        (summary.get("manufacturing_writeup") or {}).get("parts") or [])
    provenance = []
    # SDI Intelligence — cost lives in estimate_summary.part_estimates, keyed by
    # part_number. Build a lookup so the provenance report shows real costs.
    # PRE-FILTER engine figures; the sheet's totals are Excel's. See the reconciliation
    # note the provenance sheet writes under its total.
    _est_lookup = {}
    for _pe in (summary.get("estimate_summary") or {}).get("part_estimates", []):
        _pn = _pe.get("part_number")
        if _pn:
            _est_lookup[_pn] = _pe
    from costed_facts import (canonical_quantity, decision_ids_for_part,
                              is_placeholder_price, operations_for_part,
                              part_material_cost, priced_route_known,
                              priced_rows_for_part)
    _canonical = priced_route_known(summary)
    for part in parts:
        pn   = part.get("part_number") or "—"
        desc = part.get("description") or "—"
        # Bought-in components carry no fabrication material — surface them honestly
        # rather than defaulting/mis-inferring MILD_STEEL (foam tape, loom, cable…).
        _bought = _is_bought_in(part)
        if _bought:
            mat = "Bought-in"
        else:
            mat = part.get("normalized_material") or part.get("material") or "Unknown"
        # Quantity PER TOP-LEVEL UNIT from the compiled hierarchy, not the BOM row's
        # per-parent figure. A part reached through a sub-assembly needs the parent's
        # multiplicity rolled in, which is what the workbook charges on; without it this
        # sheet under-states every item below the first level.
        _cq = canonical_quantity(summary, pn)
        qty  = _cq if _cq is not None else part.get("quantity", 1)
        if isinstance(qty, float) and qty.is_integer():
            qty = int(qty)
        _pe = _est_lookup.get(pn, {})
        # Material only — see costed_facts.part_material_cost. The engine's
        # unit_total_cost_gbp is labour-inclusive and reconciles to nothing on a canonical
        # job; labour lives on the department rows, not on the part.
        if _canonical:
            unit, ext = part_material_cost(part)
        else:
            unit = float(_pe.get("unit_total_cost_gbp") or 0)
            ext  = float(_pe.get("extended_total_cost_gbp") or 0)
        geo  = str(part.get("geometry_source") or "pdf")
        # Operations as the workbook ACCEPTED them, from the one shared post-costing source.
        # This tab sits inside the same .xlsx as the Estimate, so narrating the raw textual
        # + inferred lists here made one workbook describe two different routes: laser,
        # powder and weld against timber panels the Estimate sheet charges saw, glue, CNC
        # and spray for. Falls back to the raw lists only when no workbook rows exist.
        ops = operations_for_part(summary, pn, _pe)
        _ops_priced = True
        if not ops and not _canonical:
            # No workbook — nothing is priced yet, so the drawing's own reading is the best
            # available evidence. Labelled as unpriced below rather than passed off as the
            # route. Once the workbook HAS run, a part in no labour row carries no charged
            # operation, and reaching for the raw lists there is how the gated-off route —
            # powder on timber, weld/dress on artefact records — came back onto the page.
            ops = (list(part.get("textual_operations") or [])
                   + list(part.get("inferred_operations") or []))
            _ops_priced = False
        # ── Material provenance ────────────────────────────────────────────────
        # ONE SHARED ASSESSMENT, FIELD BY FIELD.
        #
        # This block scored MATERIAL confidence from `geometry_source` — "dxf_flat" in geo
        # raised it to 95% — while printing the label from `material_source`. The score and
        # the label described different fields, so the 95% was never a statement about the
        # material at all. And a bought-in scored 1.0 for being bought-in, which is how two
        # unpriced placeholders were counted among the job's HIGH-confidence parts.
        from confidence import assess_part as _assess
        _assessment = _assess(part, summary)
        _by_field = {f["field"]: f for f in _assessment["fields"]}
        _mat_field = _by_field.get("material identity") or {}
        if _bought:
            mat_source_str = "Bought-in / catalogue component — no fabrication material"
        else:
            mat_source   = part.get("material_source") or geo
            mat_source_str = source_label(mat_source)
        # ── Thickness provenance ───────────────────────────────────────────────
        # DXF filename FIRST — most reliable, and avoids real 2mm/3mm acrylic
        # being wrongly stripped as tolerance-table values.
        import re as _re
        thk_val = None
        thk_source = "Not extracted (tolerance table only)"
        if _bought:
            thk_source = "— bought-in component (no fabrication thickness)"
        else:
            # THE ARBITRATED DATUM FIRST, exactly as the Decision Report asks for it.
            #
            # This re-derived the source by parsing the DXF filename, and on 10575-02 it
            # published "DXF filename (10575-01-001_MS_1.2mm_Rev D.DXF)" against the Decision
            # Report's "from the SolidWorks model" — one part, one thickness, two tabs of one
            # workbook naming different origins for it. Both said 1.2mm; only one was
            # describing the estimate.
            #
            # The model is rank 90 and the filename rank 70, so with both reading 1.2 the model
            # wins the arbitration and the Decision Report was right. This tab simply was not
            # asking. That is the same defect 7060e27 fixed for MATERIAL in the other file —
            # a document whose purpose is provenance must READ the provenance, because two
            # readers that each guess will eventually disagree in front of an estimator
            # deciding whether to trust a number.
            _costed = part.get("normalized_thickness_mm")
            try:
                _costed = float(_costed) if _costed not in (None, "") else None
            except (TypeError, ValueError):
                _costed = None
            if _costed and _costed > 0:
                thk_val = _costed
                _tsrc = ""
                try:
                    from source_precedence import source_of, display_name as _disp
                    _tsrc = str(source_of(part, "normalized_thickness_mm") or "")
                except Exception:
                    _tsrc = ""
                thk_source = _disp(_tsrc) if _tsrc else "costed value (source not recorded)"

            _dfn = str(part.get("dxf_source_file") or "")
            if thk_val is None:
                # FALLBACK ONLY, for records with no arbitrated value — a report of an older
                # JSON, or a part the resolver never touched.
                _tm = _re.search(r"[_\-\s](\d+\.?\d*)\s*mm", _dfn, _re.IGNORECASE)
                if _tm:
                    _tv = float(_tm.group(1))
                    if 0.3 <= _tv <= 25.0:
                        thk_val = _tv
                        thk_source = f"DXF filename ({_dfn})"
            if thk_val is None:
                thicknesses = part.get("thicknesses_mm") or []
                tol_set     = {0.5, 1.0, 1.5, 2.0, 3.0}
                _t_set      = {round(float(t),1) for t in thicknesses if t}
                # Only strip tolerance values if the FULL sequence is present;
                # a standalone 2.0/3.0 is a real thickness.
                if tol_set.issubset(_t_set):
                    thk_clean = [t for t in thicknesses
                                 if t and round(float(t), 1) not in tol_set]
                else:
                    thk_clean = [t for t in thicknesses if t]
                if thk_clean:
                    thk_val = thk_clean[0]
                    thk_source = "DXF geometry" if "dxf" in geo else "PDF text"
        _thk_field = _by_field.get("thickness") or {}
        # ── Geometry provenance ────────────────────────────────────────────────
        geo_data     = part.get("geometry_rollup") or {}
        cut_len      = float(geo_data.get("estimated_cut_length_mm") or 0)
        n_holes      = int(geo_data.get("estimated_hole_count") or 0)
        n_bends      = int(geo_data.get("estimated_bend_line_count") or 0)
        geo_conf_raw = float((geo_data.get("confidence") or {})
                             .get("estimated_cut_length_mm") or
                             geo_data.get("geometry_reliability") or 0)
        if _bought:
            geo_source = "— bought-in component (no fabrication geometry)"
        else:
            geo_source = "DXF flat pattern (exact)" if "dxf" in geo else \
                         f"PDF vector extraction (reliability {geo_conf_raw:.0%})"
        # ── Cost provenance ────────────────────────────────────────────────────
        hist_match   = None
        if _DB_OK and pn and pn != "—":
            hist_match = db.get_historical_cost(part_number=pn)
        # ── Flags ──────────────────────────────────────────────────────────────
        flags = []
        learning_flag = part.get("_learning_flag") or ""
        if learning_flag:
            flags.extend(learning_flag.split(" | "))
        # Bought-in components are catalogue-priced with no fabrication geometry, so
        # the zero-cost / unresolved-material / no-thickness review flags do not apply.
        if not _bought:
            if unit == 0.0 and mat not in ("BOUGHT_IN",):
                flags.append("Zero cost — thickness or geometry missing")
            if mat in ("Unknown", "UNKNOWN", "LED", "CARD"):
                flags.append(f"Material unresolved: {mat!r}")
            if not thk_val and mat in ("MILD_STEEL", "ACRYLIC", "MDF"):
                flags.append("No thickness extracted — manual review needed")
        # ── Override rules that fired ──────────────────────────────────────────
        # Skip for bought-ins — the DXF-token material codes never apply to them.
        overrides_fired = []
        if not _bought:
            if "override_rule:" in str(part.get("material_source") or ""):
                rule_name = str(part.get("material_source")).split("override_rule:")[-1]
                overrides_fired.append(f"Material rule: {rule_name}")
            if part.get("dxf_source_file"):
                dxf = part["dxf_source_file"].upper()
                if "_MS_" in dxf:
                    overrides_fired.append("DXF filename _MS_ → MILD_STEEL")
                elif "PETG" in dxf:
                    overrides_fired.append("DXF filename PETG → ACRYLIC")
                elif "JOINERY" in dxf:
                    overrides_fired.append("DXF filename JOINERY → MDF")
        # ── Overall confidence ─────────────────────────────────────────────────
        # The WEAKEST REQUIRED field, never a mean — and a bought-in is judged on the
        # fields it actually has, so "no fabrication thickness" cannot drag it down and
        # "it is bought-in" cannot prop it up.
        _status = _assessment["overall"]
        # ── Rate / price source ────────────────────────────────────────────────
        # The engine tags every price with a source; it lives on the material
        # estimate (part_estimate.material_estimate.price_source), not top-level.
        _ps = ((_pe.get("material_estimate") or {}).get("price_source")
               or _pe.get("price_source") or {})
        rate_basis = _price_basis_label(_ps, mat)
        # ── Route text, and the audit trail behind it ──────────────────────────
        _ops_text = (", ".join(ops) if ops
                     else ("none charged" if _canonical else "—"))
        if not _ops_priced:
            _ops_text += "  (read from drawing — not yet priced)"
        if not _canonical:
            _priced_by = "— no workbook built"
        else:
            _rows_for_part = priced_rows_for_part(summary, pn)
            _wb_rows = sorted({int(float(r["workbook_row"])) for r in _rows_for_part
                               if r.get("workbook_row")})
            _dids = decision_ids_for_part(summary, pn)
            if _wb_rows or _dids:
                _priced_by = "\n".join(filter(None, [
                    ("Estimate row " + ", ".join(str(r) for r in _wb_rows)) if _wb_rows else "",
                    " · ".join(_dids) if _dids else "",
                ]))
            else:
                _priced_by = "not priced on any labour row"
        provenance.append({
            "part_number":       pn,
            "rate_basis":        rate_basis,
            "description":       desc,
            "quantity":          qty,
            "material":          mat,
            "material_source":   mat_source_str,
            "material_status":   (_mat_field.get("status") or "unknown"),
            "material_reason":   (_mat_field.get("reason") or ""),
            "thickness_mm":      thk_val,
            "thickness_source":  thk_source,
            "thickness_status":  (_thk_field.get("status") or "unknown"),
            "cut_length_mm":     cut_len,
            "n_holes":           n_holes,
            "n_bends":           n_bends,
            "geometry_source":   geo_source,
            "geometry_conf":     geo_conf_raw,
            "dxf_file":          part.get("dxf_source_file") or "—",
            "operations":        _ops_text,
            # Part -> the Estimate rows charging it -> the compiler decisions behind them.
            "priced_by":         _priced_by,
            "unit_cost":         unit,
            "extended_cost":     ext,
            "overall_status":    _status,
            "overall_label":     _assessment["overall_label"],
            "overall_reason":    _assessment.get("reason") or "",
            "decided_by":        list(_assessment.get("decided_by") or []),
            "fields":            _assessment["fields"],
            "is_bought_in":      _bought,
            "flags":             flags,
            "overrides_fired":   overrides_fired,
            "historical_match":  hist_match,
            # WHY THIS LINE CARRIES NO MONEY. Computed from the part record by the SAME
            # classifier the workbook read-back and the HTML report use, rather than joined
            # back through final_estimate.material_rows — because the read-back is exactly
            # what fails when Excel is busy or a workbook will not open, and a sheet that explains
            # its blanks only when everything worked explains nothing on the runs that
            # needed it. A private second opinion here is also how two documents describing
            # one job come to disagree about which blanks are somebody's job.
            "unpriced_reason":   (_unpriced_reason_for(part)
                                  if not (unit or ext) else None),
        })
    return provenance


def _unpriced_reason_for(part: Dict[str, Any]) -> Dict[str, Any]:
    try:
        from estimator_inputs import unpriced_reason_for_row
        return unpriced_reason_for_row(part)
    except Exception:                                    # pragma: no cover
        return {}
def add_provenance_sheet(wb, summary: Dict[str, Any],
                          scan_meta: Dict[str, Any] = None) -> None:
    """
    Add an 'AI Provenance' sheet to an existing openpyxl workbook.
    Call this after the main estimate sheet is written.
    """
    if not _XLSX_OK:
        return
    provenance = build_provenance(summary)
    if not provenance:
        return
    ws = replace_generated_sheet(wb, "AI Provenance")
    scan_meta = scan_meta or {}

    # Authoritative total = the WB's Sell Price, found by label so it survives layout
    # shifts. If present we write a LIVE cross-sheet formula (Excel computes on open),
    # so this sheet agrees with the WB and the Decision Report. Else fall back below.
    _sell_ref = _find_wb_sell_price_ref(wb)
    from costed_facts import job_totals
    _totals = job_totals(summary)

    # THE HEADLINE FIGURE IS THE WORKBOOK'S.
    #
    # This summed the part column and called the result "Est. Total / engine part-sum".
    # Once that column became material-only the label was wrong twice over: it is not the
    # engine part-sum, and it is not a job total — driven with a workbook unit cost of
    # £6.33 and £0.10 of part material, this sheet printed "Est. Total: £0.10". The
    # Decision Report already prefers the workbook; this is the same hierarchy.
    _part_material_total = sum(p["extended_cost"] for p in provenance)
    _wb_unit = _totals.get("unit_gbp")
    _engine_total = float(_wb_unit) if _wb_unit is not None else _part_material_total

    def cell(row, col, value="", bold=False, bg=None, fg="000000",
             align="left", wrap=False, size=10, border=False, num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Calibri", bold=bold, color=fg, size=size)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center",
                                 wrap_text=wrap)
        if num_fmt:
            c.number_format = num_fmt
        if border:
            thin = Side(style="thin", color="BBBBBB")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return c
    # ── Title block ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:P1")
    cell(1, 1, "SDI Intelligence — Estimate Provenance Report",
         bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, align="center", size=13)
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:P2")
    pdf_name = scan_meta.get("pdf_name") or summary.get("source_file") or "—"
    job_no   = scan_meta.get("job_number") or "—"
    scan_dt  = scan_meta.get("scan_date") or datetime.now().strftime("%d/%m/%Y %H:%M")
    # Header total text: prefer the WB Sell Price label when we can reference it.
    _tot_txt = ("Sell Price: see Estimate sheet (mirrored below)"
                if _sell_ref else
                f"Unit cost (calculated by the Estimate sheet): £{_engine_total:.2f}"
                if _wb_unit is not None else
                f"Part material only — no workbook total: £{_engine_total:.2f}")
    cell(2, 1,
         f"Drawing: {pdf_name}   |   Job: {job_no}   |   Scanned: {scan_dt}   |   "
         f"Parts: {len(provenance)}   |   {_tot_txt}",
         bg="2F5496", fg=C_HEADER_FG, align="center", size=10)
    ws.row_dimensions[2].height = 18
    # ── Legend ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A3:P3")
    cell(3, 1,
         "STATUS — the WEAKEST field decides the line, never an average:   "
         "CONFIRMED/MEASURED — read from a model, a DXF or the estimators' own calculator   "
         "REPORTED — read from the drawing; reproducible, not verified   "
         "ASSUMED — a default or an inference is standing in   "
         "UNKNOWN — a required field has no reading   "
         "N/A — that field does not exist for this line",
         bg="F0F0F0", align="left", size=9)
    ws.row_dimensions[3].height = 16
    # ── THE SHEET SAYS WHEN IT COULD NOT BE CHECKED ────────────────────────────
    # The Excel read-back fails for reasons nothing to do with the estimate — an elevated
    # console, a workbook that will not open, Excel busy — and when it does, the figures on
    # this tab are the engine's PRE-Excel numbers rather than the ones the Estimate sheet
    # calculated. Those are different totals. Silence there is the worst case: the tab looks
    # exactly as it does on a run that reconciled perfectly.
    _fe = summary.get("final_estimate")
    if not isinstance(_fe, dict):
        _fe = (summary.get("estimate_summary") or {}).get("final_estimate")
    if not isinstance(_fe, dict) or not _fe:
        ws.merge_cells("A4:P4")
        cell(4, 1,
             "THE CALCULATED SHEET WAS NOT READ BACK — Excel did not return this workbook's "
             "computed totals (Excel busy or absent, or a workbook that would not open). "
             "The money columns below are the ENGINE's figures, not what the Estimate sheet "
             "calculates. Re-run once Excel can be driven before using these numbers.",
             bg=C_LOW, align="left", size=9, wrap=True)
        ws.row_dimensions[4].height = 30
    else:
        ws.row_dimensions[4].height = 6  # spacer
    # ── Column headers ─────────────────────────────────────────────────────────
    # The money columns are the ENGINE's per-part figures. Once Excel has calculated the
    # sheet they are not what the job is charged, and a column headed plainly "Unit £" next
    # to a Sell Price that disagrees is the report contradicting itself.
    from costed_facts import priced_route_known as _prk
    _money_basis = " material" if _prk(summary) else ""
    headers = [
        ("Part Number",       15), ("Description",     28), ("Qty", 5),
        ("Material",          14), ("Mat. Source",      32), ("Conf.",  8),
        ("Thickness",          9), ("Thk. Source",      22), ("Geometry Source", 26),
        ("Cut (mm)",          10), ("Ops",              22),
        (f"Unit £{_money_basis}",  11), (f"Ext £{_money_basis}", 11),
        ("Rate / source",     34), ("Priced by — sheet row / decision", 26),
        # A BLANK IN A MONEY COLUMN READS AS FREE, on this sheet as much as on the Estimate
        # tab. The three kinds of nothing need different people: one must NOT be priced
        # (its material is costed in another block), one is waiting on the estimator, and
        # one is work this engine cannot charge for at all.
        ("Not priced — why / who", 46),
    ]
    for ci, (hdr, width) in enumerate(headers, 1):
        c = cell(5, ci, hdr, bold=True, bg=C_SECTION, fg=C_HEADER_FG,
                 align="center", size=10)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[5].height = 20
    # ── Part rows ──────────────────────────────────────────────────────────────
    row = 6
    for i, p in enumerate(provenance):
        bg = C_ALT_ROW if i % 2 == 0 else "FFFFFF"
        # Bought-ins get a neutral grey; fabricated parts keep the confidence colour.
        # Shaded by STATUS, from the shared table, so both tabs colour a status the same
        # way. A bought-in is no longer given a neutral grey that reads as "fine": an
        # unpriced placeholder shades UNKNOWN like anything else missing a required field.
        from confidence import STATUS_FILL as _SF
        conf_bg = _SF.get(p.get("overall_status"), ("EDEDED", "555555"))[0]
        cell(row, 1,  p["part_number"],        bg=bg,       border=True)
        cell(row, 2,  p["description"],        bg=bg,       border=True, wrap=True)
        cell(row, 3,  p["quantity"],            bg=bg,       align="center", border=True)
        cell(row, 4,  p["material"],            bg=conf_bg,  bold=True, border=True)
        cell(row, 5,  p["material_source"],     bg=conf_bg,  border=True, wrap=True, size=9)
        cell(row, 6,  {"n/a": "N/A"}.get(p.get("material_status"),
                                          str(p.get("material_status") or "").upper()),
                                                bg=conf_bg,  align="center", border=True)
        cell(row, 7,  f"{p['thickness_mm']}mm" if p["thickness_mm"] else "—",
                                                bg=bg,       align="center", border=True)
        cell(row, 8,  p["thickness_source"],    bg=bg,       border=True, size=9, wrap=True)
        cell(row, 9,  p["geometry_source"],     bg=bg,       border=True, size=9, wrap=True)
        cell(row, 10, f"{p['cut_length_mm']:.0f}" if p["cut_length_mm"] else "—",
                                                bg=bg,       align="right", border=True)
        cell(row, 11, p["operations"],          bg=bg,       border=True, size=9, wrap=True)
        cell(row, 12, f"£{p['unit_cost']:.2f}", bg=bg,       align="right",
             bold=True, border=True)
        cell(row, 13, f"£{p['extended_cost']:.2f}", bg=bg,   align="right",
             bold=True, border=True)
        _rb = p.get("rate_basis") or "—"
        cell(row, 14, _rb, bg=(C_LOW if _rb.startswith("⚠") else bg),
             border=True, size=9, wrap=True)
        cell(row, 15, p.get("priced_by") or "—", bg=bg, border=True, size=8, wrap=True)
        # WHOSE BLANK THIS IS. Coloured by owner, not by severity of the number: an engine
        # gap is work that will be done and invoiced with nothing on the sheet asking anyone
        # to price it, so it is the one an estimator cannot fix and the one that gets the
        # warning fill. A line that is correctly nil is left plain — it needs no action, and
        # colouring it would teach people to ignore the colour.
        _ur = p.get("unpriced_reason") or {}
        if _ur:
            _owner = {"estimator": "ESTIMATOR TO PRICE",
                      "engine": "ENGINE GAP — THIS JOB IS UNDER-CHARGED",
                      "nobody": "nothing to charge here"}.get(_ur.get("owner"), "")
            _txt = f"{_ur.get('why')}" + (f" — {_ur['detail']}" if _ur.get("detail") else "")
            cell(row, 16, f"{_owner}: {_txt}",
                 bg=(C_LOW if _ur.get("undercharging") else bg),
                 border=True, size=8, wrap=True)
        else:
            cell(row, 16, "—", bg=bg, border=True, size=8, align="center")
        ws.row_dimensions[row].height = 28
        row += 1
        # ── Flags / warnings ───────────────────────────────────────────────────
        if p["flags"]:
            for flag in p["flags"]:
                ws.merge_cells(f"B{row}:P{row}")
                cell(row, 1, "⚠",              bg=C_LOW, align="center", size=9)
                cell(row, 2, f"REVIEW: {flag}", bg=C_LOW, size=9, wrap=True)
                ws.row_dimensions[row].height = 16
                row += 1
        # ── Override rules that fired ──────────────────────────────────────────
        if p["overrides_fired"]:
            ws.merge_cells(f"B{row}:P{row}")
            cell(row, 1, "🧠",                  bg=C_RULE, align="center", size=9)
            cell(row, 2, "Learning: " + " | ".join(p["overrides_fired"]),
                 bg=C_RULE, size=9)
            ws.row_dimensions[row].height = 14
            row += 1
        # ── Historical matches ─────────────────────────────────────────────────
        if p["historical_match"]:
            for hm in (p["historical_match"] or [])[:2]:
                avg  = hm.get("AvgCost") or hm.get("avg_cost") or 0
                cnt  = hm.get("SampleCount") or hm.get("sample_count") or 0
                hmat = hm.get("Material") or hm.get("material") or "?"
                ws.merge_cells(f"B{row}:P{row}")
                cell(row, 1, "📚",              bg=C_HIST, align="center", size=9)
                cell(row, 2,
                     f"Historical: {cnt} SDI estimate(s) for this part as "
                     f"{hmat} — avg £{avg:.2f}",
                     bg=C_HIST, size=9)
                ws.row_dimensions[row].height = 14
                row += 1
    # ── Summary footer ─────────────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:L{row}")
    # Authoritative total: WB Sell Price (live formula) when found, else engine sum.
    _total_label = ("SELL PRICE (from Estimate sheet)" if _sell_ref
                    else "UNIT COST (calculated by the Estimate sheet)" if _wb_unit is not None
                    else "PART MATERIAL ONLY — no workbook total")
    cell(row, 1,  _total_label, bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG,
         align="right", size=11)
    if _sell_ref:
        cell(row, 13, _sell_ref, bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG,
             align="right", size=12, num_fmt="£#,##0.00")
    else:
        cell(row, 13, f"£{_engine_total:.2f}",  bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG,
             align="right", size=12)
    ws.row_dimensions[row].height = 22
    # Two calculators on one page. The Ext £ column sums the engine's per-part figures; the
    # total row shows what Excel computed from the accepted labour and material rows. They
    # differ, and saying which is authoritative is not optional on a sheet whose whole
    # purpose is to be checkable.
    if _totals["source"] == "excel_calculated":
        row += 1
        ws.merge_cells(f"A{row}:P{row}")
        # Same basis as the Decision Report: reconcile the MATERIAL column against the
        # sheet's material total, and state labour as what it is — a department-row charge
        # with no per-part figure. The engine part-sum is an obsolete labour-inclusive
        # number and is deliberately not quoted here.
        _mat, _lab = _totals.get("material_gbp"), _totals.get("labour_gbp")
        _col_mat = sum(float(p.get("extended_cost") or 0) for p in provenance)
        # NAME THE GAP — IN THE WORDS THE OTHER TAB ALREADY USES.
        #
        # This printed both figures and stopped, leaving the reader to decide which was wrong.
        # Neither is. The first attempt at explaining it then got the CONTENTS wrong, and that
        # was worse than saying nothing: it named "purchased items on the Bill of Materials,
        # packaging and delivery", every one of which is already IN this column — packaging at
        # £28, the pallet, FIXING2104, the screws, the TESA tape. A reader who went looking for
        # packaging in the gap would have found it in the column instead and concluded the tab
        # could not account for itself, which is exactly the distrust the sentence exists to
        # prevent.
        #
        # The Decision Report has computed this residual all along and labels it
        # "Powder / scrap / other workbook material" — the powder consumable and the per-line
        # scrap uplift, lines that belong to no single part and so cannot appear in a per-part
        # column. Same arithmetic, same words, and the reader is sent there to see the figure
        # broken out rather than asked to take this sentence on trust.
        if _mat is None:
            _mat_txt = ""
        else:
            _gap = float(_mat) - _col_mat
            _gap_txt = ""
            if abs(_gap) >= 0.01:
                _gap_txt = (f"The £{abs(_gap):,.2f} difference is the sheet's "
                            f"POWDER / SCRAP / OTHER WORKBOOK MATERIAL — the powder consumable "
                            f"and the per-line scrap uplift, which belong to no single part and "
                            f"so cannot appear in a per-part column. The Decision Report's "
                            f"material breakdown shows it as its own row. Neither figure is "
                            f"wrong — this column is per-part provenance, the sheet is the "
                            f"money. ")
            _mat_txt = (f"The material column above sums to £{_col_mat:,.2f} against the "
                        f"sheet's £{float(_mat):,.2f}. {_gap_txt}")
        _lab_txt = (f"Labour is £{float(_lab):,.2f}, charged per department row across "
                    f"every part in that setup — see 'Priced by' for the rows and "
                    f"decisions behind each part. " if _lab is not None else "")
        cell(row, 1,
             f"RECONCILIATION — the Estimate sheet calculated "
             f"£{float(_totals.get('unit_gbp') or 0):,.2f} per unit. "
             f"{_mat_txt}{_lab_txt}The workbook is authoritative.",
             bg=C_KB, size=9, wrap=True)
        ws.row_dimensions[row].height = 32

    # WHAT THIS TAB COULD NOT DESCRIBE, BECAUSE NOTHING READ IT.
    #
    # A part with no drawing has no row worth reading here: no material source, no geometry, no
    # price — it simply is not in the job the engine saw. Leaving it out silently makes this tab
    # look like a complete account of the product when it is an account of the drawings that
    # arrived. Named here so the reader knows which parts this tab is NOT about.
    try:
        from costed_facts import undrawn_bom_lines as _undrawn
        _missing = _undrawn(summary)
    except Exception:                                            # noqa: BLE001
        _missing = []
    if _missing:
        row += 2
        ws.merge_cells(f"A{row}:P{row}")
        _names = "; ".join(
            f"{m['part_number']}" + (f" ({m['description']})" if m.get("description") else "")
            for m in _missing[:8])
        _more = f" …and {len(_missing) - 8} more" if len(_missing) > 8 else ""
        cell(row, 1,
             f"DRAWINGS MISSING FROM THIS PACK ({len(_missing)}) — no detail drawing was "
             f"supplied for these, so nothing read them and nothing costed them. They are not "
             f"in the figures above: {_names}{_more}",
             bg=C_LOW, size=9, wrap=True, bold=True)
        ws.row_dimensions[row].height = 30

    row += 2
    ws.merge_cells(f"A{row}:P{row}")
    # READING SEPARATED FROM PRICING, because they are two different questions and merging
    # them destroyed the tab's trust. The old summary took each part's WEAKEST field — and on
    # an estimate that field is almost always the price, which is "NOT YET PRICED, estimator to
    # enter a figure": a normal, expected state, not a bad read. So a part whose material,
    # thickness and geometry were ALL measured off the SolidWorks model was counted LOW because
    # nobody had typed its rate yet, and a whole job read "0 HIGH / 22 LOW" — which says the
    # engine failed when it did not. Now the engine's READING (what it pulled off the drawing and
    # the model) is scored on its own, and PRICING is reported as what it is: how many lines are
    # priced versus waiting on the estimator's rate.
    _rp = reading_and_pricing_counts(provenance)
    read_hi, read_md, read_lo = _rp["read_high"], _rp["read_med"], _rp["read_low"]
    priced, pending = _rp["priced"], _rp["pending"]
    cell(row, 1,
         f"Engine read the drawing & model:  "
         f"🟢 {read_hi} measured/confirmed   "
         f"🟡 {read_md} reported   "
         f"🔴 {read_lo} needs a look      |      "
         f"Pricing:  💷 {priced} priced   "
         f"⏳ {pending} awaiting your rate      |      "
         f"Generated by SDI Intelligence  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}",
         bg="F0F0F0", size=9, align="left")
    # Freeze panes below header + column A
    ws.freeze_panes = "B6"
    # Tab colour
    ws.sheet_properties.tabColor = "1F3864"
def generate_standalone_report(summary: Dict[str, Any],
                                output_path: str,
                                scan_meta: Dict[str, Any] = None) -> str:
    """
    Generate a standalone provenance xlsx report.
    Returns path to created file.
    """
    if not _XLSX_OK:
        print("[provenance] openpyxl not available")
        return ""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cover"
    # Simple cover page
    ws["A1"] = "SDI Intelligence Estimation Provenance Report"
    ws["A2"] = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"] = f"Drawing: {(scan_meta or {}).get('pdf_name', '—')}"
    ws["A4"] = f"Job: {(scan_meta or {}).get('job_number', '—')}"
    add_provenance_sheet(wb, summary, scan_meta)
    wb.save(output_path)
    print(f"[provenance] Report saved: {output_path}")
    return output_path
if __name__ == "__main__":
    print("SDI Intelligence Provenance Report module ready.")
    print()
    print("Integration into estimator.py:")
    print("  from estimation_report import add_provenance_sheet")
    print("  # After writing main estimate sheet:")
    print("  add_provenance_sheet(wb, summary, scan_meta)")
    print("  wb.save(xlsx_path)")
