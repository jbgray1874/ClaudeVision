"""
SDI Intelligence — Tim assemble/pack labour ingester (E2 learn-from-Tim).

The engine's E2 method takes a *median* assemble/pack time across all history,
which over/under-states any single job (1282: median => £29.13/bay vs Tim £13.44).
This ingester reads the actual assemble/pack labour from one SDI manual estimate
workbook and stores the per-bay total per drawing, so the engine uses Tim's real
figure for known jobs and only falls back to the E2 median for novel ones.

It keys off the labour table's **Dept. column** (PACM/PACP = assemble/pack), located
by header name, NOT free-text description words — so "KICK PLATE ASSEMBLY",
"UPPER LEG ASSEMBLY", "MAKE BOXES" etc. on weld/laser/P-Coat rows are NOT mistaken
for assembly. Columns are resolved from the header row ("Operation / Dept. /
Total Value"), so it generalises across SDI estimate layouts.

Input must be .xlsx/.xlsm (openpyxl); convert a true .xls first.

Outputs:
  job_assembly_labour.json   {by_drawing: {DRAWING: {lines:[...], total_gbp, source}}}
  job_assembly_labour.sql    MERGE into AIEstimating.JobAssemblyLabour
"""
from __future__ import annotations
import argparse, json, datetime
from pathlib import Path
import openpyxl

# Assemble/pack departments (config-comparable). PACM = metal assemble/pack, PACP = acrylic.
ASM_DEPTS = {"PACM", "PACP"}


def parse(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Estimate"] if "Estimate" in wb.sheetnames else wb.worksheets[0]

    drawing = None
    for r in range(1, 20):
        if str(ws.cell(r, 3).value or "").strip().lower().startswith("drawing"):
            drawing = str(ws.cell(r, 4).value or "").strip()
            break

    # Locate the labour table header: a row carrying both a "Dept" cell and a
    # "Total Value" cell. Resolve column indices from it so layout can vary.
    hdr, cols = None, {}
    for r in range(20, ws.max_row + 1):
        rowmap = {}
        for c in range(1, 16):
            t = str(ws.cell(r, c).value or "").strip().lower()
            if t:
                rowmap[t] = c
        if any("dept" in k for k in rowmap) and any("total value" in k for k in rowmap):
            hdr = r
            cols["op"]   = next((v for k, v in rowmap.items() if k.startswith("operation")), 3)
            cols["part"] = next((v for k, v in rowmap.items() if "description" in k), 4)
            cols["dept"] = next((v for k, v in rowmap.items() if "dept" in k), 7)
            cols["val"]  = next((v for k, v in rowmap.items() if "total value" in k), 13)
            break
    if hdr is None:
        return drawing, [], 0.0

    lines, total = [], 0.0
    for r in range(hdr + 1, ws.max_row + 1):
        op = str(ws.cell(r, cols["op"]).value or "").strip()
        if "total labour" in op.lower():
            break
        dept = str(ws.cell(r, cols["dept"]).value or "").strip().upper()
        val = ws.cell(r, cols["val"]).value
        if dept in ASM_DEPTS and isinstance(val, (int, float)) and val > 0:
            part = str(ws.cell(r, cols["part"]).value or "").strip()
            lines.append({
                "operation": op,
                "part_description": part,
                "dept": dept,
                "description": f"{op} — {part}".strip(" —"),
                "row": r,
                "cost_per_bay_gbp": round(float(val), 4),
            })
            total += float(val)
    return drawing, lines, round(total, 4)


def emit_sql(drawing, lines, total, source):
    out = [
        "IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='AIEstimating') EXEC('CREATE SCHEMA AIEstimating');",
        "IF OBJECT_ID('AIEstimating.JobAssemblyLabour','U') IS NULL",
        "CREATE TABLE AIEstimating.JobAssemblyLabour (",
        "    id int IDENTITY PRIMARY KEY, drawing_number varchar(40) NOT NULL,",
        "    dept varchar(10) NULL, description varchar(260) NULL,",
        "    cost_per_bay_gbp decimal(12,4) NOT NULL, is_total bit NOT NULL DEFAULT 0,",
        "    source_workbook varchar(260) NULL, ingested_utc datetime2 NOT NULL DEFAULT SYSUTCDATETIME());",
        f"DELETE FROM AIEstimating.JobAssemblyLabour WHERE drawing_number='{drawing}';",
    ]

    def q(v):
        return "NULL" if v is None else "'" + str(v).replace("'", "''") + "'"

    for ln in lines:
        out.append(
            "INSERT INTO AIEstimating.JobAssemblyLabour(drawing_number,dept,description,cost_per_bay_gbp,is_total,source_workbook) "
            f"VALUES('{drawing}',{q(ln['dept'])},{q(ln['description'])},{ln['cost_per_bay_gbp']},0,{q(source)});"
        )
    out.append(
        "INSERT INTO AIEstimating.JobAssemblyLabour(drawing_number,dept,description,cost_per_bay_gbp,is_total,source_workbook) "
        f"VALUES('{drawing}','PACM','TOTAL assemble/pack per bay',{total},1,{q(source)});"
    )
    return "\n".join(out) + "\n"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--out-dir", default=".")
    ap.add_argument("--write-db", action="store_true")
    a = ap.parse_args()
    drawing, lines, total = parse(a.workbook)
    src = Path(a.workbook).name
    outd = Path(a.out_dir)

    jpath = outd / "job_assembly_labour.json"
    payload = {"by_drawing": {}}
    if jpath.exists():
        try:
            payload = json.loads(jpath.read_text())
        except Exception:
            payload = {"by_drawing": {}}
    payload.setdefault("by_drawing", {})[drawing] = {
        "lines": lines, "total_gbp": total, "source": src,
        "ingested": datetime.datetime.now(datetime.timezone.utc).isoformat(),
    }
    jpath.write_text(json.dumps(payload, indent=2))
    (outd / "job_assembly_labour.sql").write_text(emit_sql(drawing, lines, total, src))
    print(f"Drawing {drawing}: {len(lines)} assemble/pack (PACM/PACP) line(s), total £{total:.2f}/bay")
    for ln in lines:
        print(f"    row {ln['row']:>3}  {ln['dept']:<4}  £{ln['cost_per_bay_gbp']:>7.2f}  {ln['description']}")
    print(f"  -> {jpath}\n  -> {outd/'job_assembly_labour.sql'}")
    if a.write_db:
        try:
            import pyodbc, config  # noqa: F401
            cn = config.get_connection(); cur = cn.cursor()
            cur.execute(open(outd / "job_assembly_labour.sql").read()); cn.commit()
            print("  -> written to AIEstimating.JobAssemblyLabour")
        except Exception as exc:
            print(f"  (DB write skipped: {exc})")


if __name__ == "__main__":
    main()
