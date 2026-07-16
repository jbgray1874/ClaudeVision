#!/usr/bin/env python3
r"""
_probe_powder_calc_formulas.py  —  READ-ONLY.

Goal: understand how the WB turns a steel part's POWDER AREA into COST, so we can
make the 4 real P.Coat labour lines SIZE-SCALE (like Tim's 424/hr varying by part)
instead of the flat 369/184 throughput -> flat £10.81.

The sheet already computes, per steel row (right-hand calculators):
    Powder Qty Calculator cols (approx AB-AE): m2 Per Part, Cutting Speed, Qty Per Kilo,
    Powder Qty Per Part ; plus a "Total Powder Per Unit" cell.

But the P.Coat LABOUR cost (Labour block col M) is driven by the operation name's
dept rate (P/C £355.43/hr) and the THROUGHPUT we write (flat 369/184). The labour
cost does NOT read the powder-area calc. This probe dumps BOTH so we can see the
relationship and decide the correct throughput to write.

Dumps, from the widened template:
  1. Header row 37 labels for the right-hand calculator columns (P..AE) so we know
     which column is which by name (not guessed).
  2. Rows 38-40 formulas for those columns (the powder area/qty per part).
  3. The "Total Powder Per Unit" cell + any powder cost cell.
  4. A sample P.Coat labour row's cost formula (col M in the labour block) so we see
     how throughput (col I) drives cost — the number we'd change.

READ-ONLY. No writes.

Usage:
  C:\ClaudeVision\.venv\Scripts\python.exe _probe_powder_calc_formulas.py
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as L

TEMPLATE = r'\\sdi-dc01\shareddata$\Shared\Estimating\Completed\AI Estimating\AISheets\Blank Estimate Sheet  WB 2026.xlsx'


def main():
    wb = load_workbook(TEMPLATE)
    ws = wb['Estimate']

    print("=" * 90)
    print("POWDER CALCULATOR + P.COAT COST PATH PROBE (read-only)")
    print("=" * 90)

    # 1. Header labels for the right-hand calculator columns (P=16 .. AE=31)
    print("\n[1] Calculator column headers (row 37), cols P(16)..AE(31):")
    for c in range(16, 32):
        v = ws.cell(row=37, column=c).value
        if v not in (None, ""):
            print(f"    {L(c)}37 = {v!r}")

    # 2. Powder calc formulas for first 3 steel rows (38-40), cols AB(28)..AE(31)
    print("\n[2] Powder-calc formulas, steel rows 38-40, cols Z(26)..AF(32):")
    for r in (38, 39, 40):
        cells = []
        for c in range(26, 33):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                cells.append(f"{L(c)}{r}={v!r}")
        print(f"    row {r}: " + " | ".join(cells))

    # 3. Search for a 'Total Powder' label and any nearby cost cell
    print("\n[3] 'Total Powder' / powder-cost cells (scan rows 36-60, all cols):")
    for r in range(36, 61):
        for c in range(1, 40):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "powder" in v.lower():
                # print the label and the two cells to its right
                right = [f"{L(c+i)}{r}={ws.cell(row=r,column=c+i).value!r}" for i in range(1, 4)]
                print(f"    {L(c)}{r} = {v!r}  ->  " + " | ".join(right))

    # 4. A P.Coat labour row cost formula: find first labour row whose col C maps P/C.
    #    Labour block is 71-142 (widened). Show col C (op), col I (throughput), col M (cost).
    print("\n[4] Labour block P.Coat row cost path (cols C op, H qty, I throughput, K rate, M cost):")
    printed = 0
    for r in range(71, 143):
        opv = ws.cell(row=r, column=3).value  # C
        # in a blank template these are formulas/empty; show the FORMULA structure from row 71
        if r == 71 or (isinstance(opv, str) and "coat" in str(opv).lower()):
            cvals = {L(c): ws.cell(row=r, column=c).value for c in (3, 8, 9, 11, 13)}
            print(f"    row {r}: " + " | ".join(f"{k}={v!r}" for k, v in cvals.items()))
            printed += 1
            if printed >= 3:
                break

    print("\n" + "=" * 90)
    print("From [2]/[3]: how powder AREA -> powder QTY/cost is computed.")
    print("From [4]: how the P.Coat labour COST uses throughput (col I) — the number we'd change.")
    print("Decision: can we write a size-scaled throughput (from area) instead of flat 369/184?")


if __name__ == "__main__":
    main()
