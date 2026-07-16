#!/usr/bin/env python3
r"""
_probe_laser_path.py  —  READ-ONLY diagnostic. Writes nothing.

Answers three questions for 1300-01 (or any single-steel-part job):
  1. Does the Estimate-sheet Laser Rate Calculator (cells P..W) feed the Labour
     tab's laser cost, or are they computed on separate paths? (the "decoupling")
  2. What is the No. of holes (already in the JSON) and where should it go? (cell S)
  3. What is a drawing-derived Internal Cutting Distance? (cell T)
        internal_cut ~= total_cut_length_mm - bounding_perimeter
        bounding_perimeter = 2*(PartLength + PartWidth)

USAGE (run AFTER opening the xlsx in Excel once so formulas have cached values,
       or it will note that values are blank):

  C:\ClaudeVision\.venv\Scripts\python.exe _probe_laser_path.py ^
      "C:\ClaudeVision\output\estimates\1300-01FlatShelf_20260709_094613.xlsx" ^
      "C:\ClaudeVision\output\json\1300-01FlatShelf.json"

Nothing is written. It only reads the two files.
"""
import sys, json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def find_steel_row(ws):
    """First row under a 'Part Length'/'Cost Per Part' header that has a numeric Part Length."""
    # locate header row + the Part Length / cols by label
    cols = {}
    hdr_row = None
    for r in range(1, min(ws.max_row, 80) + 1):
        labels = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                labels[v.strip()] = c
        if "Part Length" in labels and ("Cost Per Part" in labels or "Gauge" in labels):
            hdr_row = r
            cols = labels
            break
    if not hdr_row:
        return None, {}, None
    # laser calc + input columns by header label (row may differ slightly)
    want = ["Part Length", "Part Width", "Gauge", "No of holes", "Intenal Cutting Distance",
            "Internal Cutting Distance", "Cutting Speed (mm per sec)", "Profile Cutting (secs)",
            "Non Profile Cutting (secs)", "Total Time", "Rate Per Hour", "Cost Per Part",
            "Load /  Unload"]
    colmap = {}
    for r in (hdr_row, hdr_row - 1, hdr_row + 1):
        if r < 1:
            continue
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() in want:
                colmap.setdefault(v.strip(), c)
    # first data row with numeric Part Length
    plc = colmap.get("Part Length")
    data_row = None
    for r in range(hdr_row + 1, hdr_row + 15):
        if plc and _num(ws.cell(row=r, column=plc).value):
            data_row = r
            break
    return data_row, colmap, hdr_row


def find_labour_laser(wb):
    for name in wb.sheetnames:
        ws = wb[name]
        # header map
        hdr = {}
        hdr_row = None
        for r in range(1, min(ws.max_row, 40) + 1):
            row_labels = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    row_labels[v.strip()] = c
            if "Rate Per Hour" in row_labels and "Total Value" in row_labels:
                hdr = row_labels
                hdr_row = r
                break
        if not hdr_row:
            continue
        # find the laser line
        for r in range(hdr_row + 1, ws.max_row + 1):
            rowvals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            joined = " ".join(str(x) for x in rowvals if x is not None).upper()
            if "LASM" in joined or "LASER (METAL)" in joined:
                return name, r, hdr
    return None, None, {}


def main(xlsx, jpath):
    print("=" * 78)
    print("LASER PATH PROBE  (read-only)")
    print("=" * 78)

    # ---- JSON side ----
    with open(jpath, "r", encoding="utf-8") as fh:
        s = json.load(fh)
    parts = (s.get("estimate_summary") or {}).get("part_estimates") or s.get("part_estimates") or []
    steel = next((p for p in parts if (p.get("material") or "").upper().replace(" ", "_") == "MILD_STEEL"
                  and (p.get("geometry") or {}).get("estimated_cut_length_mm")), parts[0] if parts else {})
    g = steel.get("geometry") or {}
    cut = _num(g.get("estimated_cut_length_mm"))
    holes = g.get("estimated_hole_count")
    pierces = g.get("estimated_pierce_count")
    print("\n[JSON]  part:", steel.get("part_number") or steel.get("name"))
    print(f"        estimated_cut_length_mm : {cut}")
    print(f"        estimated_hole_count    : {holes}   (pierces: {pierces})")

    # ---- xlsx side ----
    wb = load_workbook(xlsx, data_only=True)
    est = wb["Estimate"] if "Estimate" in wb.sheetnames else wb[wb.sheetnames[0]]
    row, cmap, hdr = find_steel_row(est)
    print(f"\n[XLSX]  Estimate steel data row: {row}")
    def cell(label):
        c = cmap.get(label)
        return est.cell(row=row, column=c).value if (row and c) else None

    F = _num(cell("Part Length")); G = _num(cell("Part Width")); H = _num(cell("Gauge"))
    S = cell("No of holes")
    T = cell("Intenal Cutting Distance") or cell("Internal Cutting Distance")
    R = _num(cell("Profile Cutting (secs)")); U = _num(cell("Non Profile Cutting (secs)"))
    V = _num(cell("Total Time")); W = _num(cell("Rate Per Hour"))
    print(f"        Part Length (F)   : {F}")
    print(f"        Part Width  (G)   : {G}")
    print(f"        Gauge       (H)   : {H}")
    print(f"        No of holes (S)   : {S!r}   <-- INPUT cell")
    print(f"        Internal Cut(T)   : {T!r}   <-- INPUT cell")
    print(f"        Profile secs(R)   : {R}")
    print(f"        NonProfile  (U)   : {U}")
    print(f"        Total Time  (V)   : {V}")
    print(f"        Calc Rate/hr(W)   : {W}   (parts/hr the CALCULATOR produces)")

    # ---- Labour side ----
    lname, lrow, lhdr = find_labour_laser(wb)
    print(f"\n[XLSX]  Labour laser line: sheet={lname} row={lrow}")
    lab = {}
    if lrow:
        lws = wb[lname]
        for lbl in ("Rate Per Hour", "Total Hours", "Labour Cost", "Set Up (Mins)", "Total Value"):
            c = lhdr.get(lbl)
            lab[lbl] = lws.cell(row=lrow, column=c).value if c else None
        for k, v in lab.items():
            print(f"        {k:<16}: {v}")

    # ---- derive internal distance + verdict ----
    print("\n" + "-" * 78)
    if F and G and cut:
        perim = 2 * (F + G)
        internal = round(cut - perim, 1)
        print(f"DERIVED internal cut distance = cut_length - 2*(F+G)")
        print(f"        = {cut} - 2*({F}+{G}) = {cut} - {round(perim,1)} = {internal} mm")
        print(f"        -> candidate value for cell T (holes candidate for S = {holes})")

    print("-" * 78)
    calc_rate = W
    lab_rate = _num(lab.get("Rate Per Hour"))
    if calc_rate and lab_rate:
        ratio = round(calc_rate / lab_rate, 2) if lab_rate else None
        print(f"DECOUPLING CHECK:")
        print(f"   Estimate calculator parts/hr (W) : {calc_rate}")
        print(f"   Labour tab Rate/Hour             : {lab_rate}")
        if abs(calc_rate - lab_rate) > max(2.0, 0.1 * calc_rate):
            print(f"   => DECOUPLED (differ by {ratio}x). The Labour laser cost is NOT read")
            print(f"      from the Estimate laser calculator. Writing S/T into the calc will")
            print(f"      change the DISPLAY (V,W) but may NOT change the charged Labour cost.")
            print(f"      Next: find where wb_populate computes the LASM Total Value.")
        else:
            print(f"   => COUPLED. The calculator feeds the Labour cost; writing S/T WILL move £.")
    else:
        print("DECOUPLING CHECK: could not read both rates (open xlsx in Excel once so formulas cache).")
    print("=" * 78)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("usage: python _probe_laser_path.py <estimate.xlsx> <summary.json>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
