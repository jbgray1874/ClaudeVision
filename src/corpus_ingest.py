"""
corpus_ingest.py — turn historical SDI estimate workbooks into clean, normalised,
comparable records for the vector / RAG corpus.

This is a STANDALONE, READ-ONLY tool. It does not import or touch the live estimating
engine; it only reads .xls/.xlsx estimate workbooks and writes record JSONL. Safe to
run alongside everything else.

Why "normalised / comparable":
  RAG comparables only help if they compare like-for-like. Each record therefore keeps
    * raw_manufacturing_cost_gbp = material + labour  (BEFORE rebate / overhead)
    * the quantity the estimate was priced at  (setup amortises differently at qty 40 vs 1000)
    * the date / year  (a 2019 price is not a 2026 price)
    * the rebate fraction and a DERIVED overhead divisor (each estimator's choice, kept explicit)
  so retrieval can age- and quantity-normalise rather than trusting a raw unit cost.

Grains emitted:
  * one "job" record  (top-down sanity band: similar whole jobs)
  * one "part" record per detail part (component analogy: similar parts/ops)

Each record carries an `embedding_text` (what gets vectorised) and `source` provenance
(workbook + sheet) so any retrieved comparable can be traced back to its sheet.

Usage:
    python corpus_ingest.py --glob "K:/Estimating/.../**/*.xls" --out corpus.jsonl
    python corpus_ingest.py --paths a.xls b.xlsx --out corpus.jsonl --pretty
"""

from __future__ import annotations

import argparse
import datetime as _dt
import glob as _glob
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


# --------------------------------------------------------------------------------------
# Workbook loading (xls via xlrd, xlsx via openpyxl) -> uniform 2D grid of values
# --------------------------------------------------------------------------------------
def _load_sheet(path: Path) -> Tuple[List[List[Any]], int]:
    """Return (rows, datemode). datemode: 0 = 1900 system, 1 = 1904 system."""
    suffix = path.suffix.lower()
    if suffix == ".xls":
        import xlrd

        wb = xlrd.open_workbook(str(path))
        sh = wb.sheet_by_index(0)
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
        return rows, wb.datemode
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(str(path), data_only=True, read_only=True)
        sh = wb[wb.sheetnames[0]]
        rows = [[c if c is not None else "" for c in row] for row in sh.iter_rows(values_only=True)]
        return rows, 0
    raise ValueError(f"Unsupported workbook type: {path.suffix}")


def _cell(rows: List[List[Any]], r: int, c: int) -> Any:
    if 0 <= r < len(rows) and 0 <= c < len(rows[r]):
        return rows[r][c]
    return ""


def _col(row: List[Any], c: int) -> Any:
    return row[c] if 0 <= c < len(row) else ""


def _f(v: Any) -> Optional[float]:
    try:
        if v is None or v == "":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _s(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _to_year(v: Any, datemode: int) -> Optional[int]:
    """Excel serial or datetime/string -> calendar year."""
    if isinstance(v, _dt.datetime):
        return v.year
    fv = _f(v)
    if fv and fv > 20000:  # plausible Excel serial date
        try:
            import xlrd

            y, *_ = xlrd.xldate_as_tuple(fv, datemode)
            return int(y)
        except Exception:
            pass
    m = re.search(r"(20\d{2})", _s(v))
    return int(m.group(1)) if m else None


# --------------------------------------------------------------------------------------
# Section / label location (robust to row drift across template versions)
# --------------------------------------------------------------------------------------
def _find_row(rows: List[List[Any]], label: str, *, col: Optional[int] = None) -> Optional[int]:
    lab = label.upper().strip()
    # Pass 1: exact title-cell match. Section titles ("Sheet Steel", "Wire", "Labour")
    # are standalone cells, so this avoids shadowing by meta cells that merely CONTAIN
    # the words (e.g. "Sheet Steel Cost Per Tonne" near the header).
    for r, row in enumerate(rows):
        cells = range(len(row)) if col is None else [col]
        for c in cells:
            if c < len(row) and _s(row[c]).upper().rstrip(" :").strip() == lab:
                return r
    # Pass 2: fall back to a contains match.
    for r, row in enumerate(rows):
        cells = range(len(row)) if col is None else [col]
        for c in cells:
            if c < len(row) and lab in _s(row[c]).upper():
                return r
    return None


def _meta_value_right_of(rows: List[List[Any]], label: str) -> Any:
    """Find a label cell and return the first non-empty cell to its right on the same row."""
    lab = label.upper()
    for row in rows:
        for c, cell in enumerate(row):
            if lab in _s(cell).upper():
                for cc in range(c + 1, len(row)):
                    if _s(row[cc]):
                        return row[cc]
    return ""


# --------------------------------------------------------------------------------------
# Block readers — each finds its header row by label, then reads data rows until the
# next section header or a run of blanks. Column layout is detected from the header row.
# --------------------------------------------------------------------------------------
_STOP_TITLES = {
    "STANDARD MATERIALS", "BILL OF MATERIALS", "WIRE", "SHEET STEEL",
    "OTHER SHEET MATERIAL", "LABOUR", "DEPARTMENT",
}


def _is_section_header(row: List[Any]) -> bool:
    """A row starts a new section only if an early cell is a section TITLE (exact),
    or a 'Total ... Cost' line. Substring matching is avoided so operation names like
    'Manual labour (Acrylic)' do not get mistaken for the 'Labour' section header."""
    for c in row[:4]:
        v = _s(c).upper().rstrip(" :").strip()
        if v in _STOP_TITLES:
            return True
        if v.startswith("TOTAL ") and "COST" in v:
            return True
    return False


def _read_block(rows: List[List[Any]], start_label: str) -> Tuple[Optional[int], List[List[Any]]]:
    """Return (header_row_index, data_rows) for the section beginning at start_label.
    The header row is the row at/after the label that names columns (contains
    'Part Description' / 'Operation' / 'Price'). Data rows follow until next section."""
    hdr = _find_row(rows, start_label)
    if hdr is None:
        return None, []
    # locate the column-header row (label row may itself be the title only)
    col_hdr = hdr
    for r in range(hdr, min(hdr + 3, len(rows))):
        joined = " ".join(_s(c).upper() for c in rows[r])
        if any(k in joined for k in ("PART DESCRIPTION", "OPERATION", "PRICE", "PART CODE")):
            col_hdr = r
            break
    data: List[List[Any]] = []
    for r in range(col_hdr + 1, len(rows)):
        row = rows[r]
        if _is_section_header(row) and r != col_hdr:
            break
        data.append(row)
    return col_hdr, data


def _parse_steel_parts(rows: List[List[Any]]) -> List[Dict[str, Any]]:
    """Sheet Steel + Other Sheet Material blocks -> material part records."""
    out: List[Dict[str, Any]] = []
    for label, mat_class in (("SHEET STEEL", "mild_steel"), ("OTHER SHEET MATERIAL", "sheet_other")):
        _, data = _read_block(rows, label)
        for row in data:
            pn = _s(_col(row, 0))
            desc = _s(_col(row, 2))
            if not pn and not desc:
                continue
            # Other Sheet Material shifts dims one column left (no gauge col split)
            if label == "OTHER SHEET MATERIAL":
                rec = dict(part_number=pn, description=desc, material_class=mat_class,
                           qty_per_unit=_f(_col(row, 3)), length_mm=_f(_col(row, 4)),
                           width_mm=_f(_col(row, 5)), thickness_mm=_f(_col(row, 6)),
                           cost_per_sheet_gbp=_f(_col(row, 11)),
                           material_cost_per_part_gbp=_f(_col(row, 12)),
                           scrap_pct=_f(_col(row, 10)))
            else:
                rec = dict(part_number=pn, description=desc, material_class=mat_class,
                           qty_per_unit=_f(_col(row, 4)), length_mm=_f(_col(row, 5)),
                           width_mm=_f(_col(row, 6)), thickness_mm=_f(_col(row, 7)),
                           material_cost_per_part_gbp=_f(_col(row, 12)),
                           scrap_pct=_f(_col(row, 11)))
            if rec["material_cost_per_part_gbp"] or rec["length_mm"]:
                out.append(rec)
    return out


def _parse_bought_in(rows: List[List[Any]]) -> List[Dict[str, Any]]:
    _, data = _read_block(rows, "STANDARD MATERIALS")
    out: List[Dict[str, Any]] = []
    for row in data:
        desc = _s(_col(row, 2))
        price = _f(_col(row, 9))
        value = _f(_col(row, 12))
        if not desc or (value in (None, 0.0) and price in (None, 0.0)):
            continue
        out.append(dict(description=desc, supplier=_s(_col(row, 8)),
                        unit_price_gbp=price, qty_per_unit=_f(_col(row, 10)),
                        scrap_pct=_f(_col(row, 11)), extended_value_gbp=value))
    return out


def _parse_labour(rows: List[List[Any]]) -> List[Dict[str, Any]]:
    _, data = _read_block(rows, "LABOUR")
    out: List[Dict[str, Any]] = []
    for row in data:
        op = _s(_col(row, 2))
        dept = _s(_col(row, 6))
        if not op and not dept:
            continue
        if "TOTAL" in op.upper():
            break
        out.append(dict(part_number=_s(_col(row, 0)), operation=op,
                        part_description=_s(_col(row, 3)), dept=dept,
                        qty_per_unit=_f(_col(row, 7)), rate_per_hour=_f(_col(row, 8)),
                        total_hours=_f(_col(row, 9)), labour_rate_gbp_per_hr=_f(_col(row, 10)),
                        setup_mins=_f(_col(row, 11)), value_per_unit_gbp=_f(_col(row, 12))))
    return out


# --------------------------------------------------------------------------------------
# Embedding text
# --------------------------------------------------------------------------------------
def _job_embedding_text(job: Dict[str, Any], materials: List[str], depts: List[str]) -> str:
    bits = [job.get("description", ""), job.get("customer", ""),
            " ".join(sorted(set(materials))), " ".join(sorted(set(depts)))]
    return " | ".join(b for b in bits if b)


def _part_embedding_text(p: Dict[str, Any], ops: List[str]) -> str:
    dims = ""
    if p.get("length_mm") and p.get("width_mm"):
        dims = f"{p['length_mm']:.0f}x{p['width_mm']:.0f}"
    mat = p.get("material_class", "")
    th = f"{p['thickness_mm']}mm" if p.get("thickness_mm") else ""
    bits = [p.get("description", ""), f"{mat} {th}".strip(), dims, " ".join(ops)]
    return " | ".join(b for b in bits if b)


# --------------------------------------------------------------------------------------
# Bought-in categorisation + pack labour (the "commercial wrap" the comparables need)
# --------------------------------------------------------------------------------------
_BOUGHT_IN_CATEGORIES = [
    ("print",     ("print", "litho", "digi", "poster", "header", "pos ", "graphic",
                   "vinyl", "decal", "label", "artwork")),
    ("packaging", ("box", "carton", "case", "mailer", "wrap", "foam", "insert",
                   "void", "bag", "sleeve")),
    ("pallet",    ("pallet", "skid")),
    ("delivery",  ("deliver", "carriage", "freight", "transport", "courier",
                   "haulage", "postage", "shipping")),
    ("fixing",    ("screw", "bolt", "nut", "washer", "rivet", "fixing", "fastener",
                   "stud", "clip", "magnet", "standoff", "stand-off", "bumpon")),
]


def _categorise_bought_in(desc: str) -> str:
    d = (desc or "").lower()
    for cat, kws in _BOUGHT_IN_CATEGORIES:
        if any(k in d for k in kws):
            return cat
    return "other"


_PACK_DEPTS = ("ASSEMBLY", "ASSY", "PACK", "PACKING", "PACKAGING")


def _pack_labour_total(labour: List[Dict[str, Any]]) -> Optional[float]:
    tot, hit = 0.0, False
    for o in labour:
        dept = _s(o.get("dept")).upper()
        if any(p in dept for p in _PACK_DEPTS):
            tot += _f(o.get("value_per_unit_gbp")) or 0.0
            hit = True
    return round(tot, 4) if hit else None


def _bought_in_rollup(bought_in: List[Dict[str, Any]]) -> Tuple[Optional[float], Dict[str, float]]:
    breakdown: Dict[str, float] = {}
    total = 0.0
    for b in bought_in:
        v = _f(b.get("extended_value_gbp")) or 0.0
        cat = _categorise_bought_in(b.get("description"))
        breakdown[cat] = round(breakdown.get(cat, 0.0) + v, 4)
        total += v
    return (round(total, 4) if bought_in else None), breakdown


# --------------------------------------------------------------------------------------
# Main parse
# --------------------------------------------------------------------------------------
def parse_sdi_estimate_workbook(path: Path) -> Dict[str, Any]:
    rows, datemode = _load_sheet(path)
    warnings: List[str] = []

    job_no = _s(_meta_value_right_of(rows, "Drawing No"))
    job_no = re.sub(r"-GA.*$", "", job_no) or job_no  # 10897-01-GA -> 10897-01
    job_no = re.sub(r"\.0$", "", job_no)               # numeric 1282.0 -> 1282
    rev = _s(_meta_value_right_of(rows, "Rev"))
    desc = _s(_meta_value_right_of(rows, "Description"))
    customer = _s(_meta_value_right_of(rows, "Customer"))
    qty = _f(_meta_value_right_of(rows, "Quantity"))
    prepared_by = _s(_meta_value_right_of(rows, "Prepared By"))
    year = _to_year(_meta_value_right_of(rows, "Date"), datemode)
    unit_cost = _f(_meta_value_right_of(rows, "Unit Cost"))

    mat_total = _f(_meta_value_right_of(rows, "Total Material Cost"))
    lab_total = _f(_meta_value_right_of(rows, "Total Labour Cost"))
    rebate = _f(_meta_value_right_of(rows, "Rebate Calculator"))
    sell = _f(_meta_value_right_of(rows, "Sell Price"))

    steel_parts = _parse_steel_parts(rows)
    bought_in = _parse_bought_in(rows)
    labour = _parse_labour(rows)

    # ---- normalisation -------------------------------------------------------------
    raw_cost = None
    if mat_total is not None and lab_total is not None:
        raw_cost = round(mat_total + lab_total, 4)
    overhead_div = None
    if raw_cost and unit_cost and rebate is not None and unit_cost > 0:
        # unit_cost = (mat+lab)/(1-rebate)/overhead_div  ->  derive overhead_div
        overhead_div = round(raw_cost / ((1.0 - rebate) * unit_cost), 4)
        if not (0.5 <= overhead_div <= 1.2):
            warnings.append(f"derived overhead divisor {overhead_div} out of sane range "
                            f"(0.5-1.2) — possible data-entry error in unit cost")

    # group labour ops by part — ONLY by a real part number. Older templates omit the
    # part-number column on labour rows; attributing those to parts would smear every
    # op onto every part, so they go to a job-level 'unattributed_ops' bucket instead.
    ops_by_part: Dict[str, List[Dict[str, Any]]] = {}
    unattributed_ops: List[Dict[str, Any]] = []
    for op in labour:
        opn = _s(op.get("part_number"))
        if opn:
            ops_by_part.setdefault(opn, []).append(op)
        else:
            unattributed_ops.append(op)

    parts_with_pn = sum(1 for p in steel_parts if _s(p.get("part_number")))
    if steel_parts and parts_with_pn / max(1, len(steel_parts)) < 0.5:
        warnings.append("older template — most parts lack a part-number column; "
                        "part/op attribution limited (job-level record is unaffected)")
    if unattributed_ops and not ops_by_part:
        warnings.append("labour rows carry no part numbers — operations recorded at "
                        "job level only (per-part op attribution unavailable)")

    materials_used = sorted({p["material_class"] for p in steel_parts})
    depts_used = sorted({_s(o["dept"]) for o in labour if _s(o["dept"])})
    _bi_total, _bi_breakdown = _bought_in_rollup(bought_in)

    job_record = {
        "record_type": "job",
        "job_no": job_no,
        "revision": rev,
        "description": desc,
        "customer": customer,
        "quantity": qty,
        "year": year,
        "prepared_by": prepared_by,
        "material_cost_gbp": mat_total,
        "labour_cost_gbp": lab_total,
        "raw_manufacturing_cost_gbp": raw_cost,   # pre rebate/overhead — the comparable basis
        "rebate_fraction": rebate,
        "overhead_divisor_derived": overhead_div,
        "unit_cost_gbp": unit_cost,
        "sell_price_gbp": sell,
        "materials_used": materials_used,
        "departments_used": depts_used,
        "part_count": len(steel_parts),
        "bought_in_count": len(bought_in),
        "bought_in_total_gbp": _bi_total,
        "bought_in_breakdown_gbp": _bi_breakdown or None,
        "pack_labour_gbp": _pack_labour_total(labour),
        "unattributed_op_count": len(unattributed_ops),
        "unattributed_op_value_gbp": round(
            sum(_f(o.get("value_per_unit_gbp")) or 0.0 for o in unattributed_ops), 4
        ) or None,
        "embedding_text": _job_embedding_text(
            {"description": desc, "customer": customer}, materials_used, depts_used),
        "source": {"workbook": path.name, "sheet": 0},
        "warnings": warnings,
    }

    part_records: List[Dict[str, Any]] = []
    for p in steel_parts:
        pn = _s(p.get("part_number"))
        ops = ops_by_part.get(pn, [])
        lab_cost = round(sum(_f(o.get("value_per_unit_gbp")) or 0.0 for o in ops), 4) if ops else None
        rec = dict(p)
        rec.update({
            "record_type": "part",
            "job_no": job_no,
            "year": year,
            "operations": [
                {"operation": o["operation"], "dept": o["dept"],
                 "rate_per_hour": o["rate_per_hour"], "total_hours": o["total_hours"],
                 "labour_rate_gbp_per_hr": o["labour_rate_gbp_per_hr"],
                 "setup_mins": o["setup_mins"], "value_per_unit_gbp": o["value_per_unit_gbp"]}
                for o in ops
            ],
            "labour_cost_per_part_gbp": lab_cost,
            "embedding_text": _part_embedding_text(p, [o["operation"] for o in ops]),
            "source": {"workbook": path.name, "sheet": 0},
        })
        part_records.append(rec)

    return {"job": job_record, "parts": part_records,
            "bought_in": bought_in, "warnings": warnings}


def ingest_paths(paths: List[Path]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    for path in paths:
        try:
            parsed = parse_sdi_estimate_workbook(path)
        except Exception as e:  # one bad workbook must not sink the batch
            records.append({"record_type": "error", "source": {"workbook": path.name},
                            "error": f"{type(e).__name__}: {e}"})
            continue
        records.append(parsed["job"])
        records.extend(parsed["parts"])
        _jn = parsed["job"].get("job_no")
        _yr = parsed["job"].get("year")
        for b in parsed.get("bought_in", []):
            rec = dict(b)
            rec.update({"record_type": "bought_in", "job_no": _jn, "year": _yr,
                        "category": _categorise_bought_in(b.get("description")),
                        "source": {"workbook": path.name, "sheet": 0}})
            records.append(rec)
    return records


def main() -> None:
    ap = argparse.ArgumentParser(description="Ingest SDI estimate workbooks into corpus records.")
    ap.add_argument("--glob", help="glob of workbooks, e.g. 'K:/Estimating/**/*.xls'")
    ap.add_argument("--paths", nargs="*", help="explicit workbook paths")
    ap.add_argument("--out", default="corpus.jsonl", help="output JSONL path")
    ap.add_argument("--pretty", action="store_true", help="also print a human summary")
    args = ap.parse_args()

    paths: List[Path] = []
    if args.glob:
        paths += [Path(p) for p in _glob.glob(args.glob, recursive=True)]
    if args.paths:
        paths += [Path(p) for p in args.paths]
    paths = [p for p in paths if p.suffix.lower() in (".xls", ".xlsx", ".xlsm")]
    if not paths:
        ap.error("no workbooks found")

    records = ingest_paths(paths)
    with open(args.out, "w", encoding="utf-8") as fh:
        for rec in records:
            fh.write(json.dumps(rec, ensure_ascii=False) + "\n")

    jobs = [r for r in records if r.get("record_type") == "job"]
    parts = [r for r in records if r.get("record_type") == "part"]
    boughtin = [r for r in records if r.get("record_type") == "bought_in"]
    errs = [r for r in records if r.get("record_type") == "error"]
    print(f"Ingested {len(paths)} workbook(s) -> {len(jobs)} job records, "
          f"{len(parts)} part records, {len(boughtin)} bought-in records, "
          f"{len(errs)} errors -> {args.out}")
    if args.pretty:
        for j in jobs:
            print(f"  {j['job_no']:<14} {str(j['description'])[:30]:<30} qty={j['quantity']} "
                  f"raw=£{j['raw_manufacturing_cost_gbp']} unit=£{j['unit_cost_gbp']} "
                  f"oh={j['overhead_divisor_derived']} {j['warnings'] or ''}")


if __name__ == "__main__":
    main()
