#!/usr/bin/env python3
"""
catalogue_loader.py — SDI Intelligence Phase 1 catalogue loader.

Turns manual-estimate workbooks into the versioned price truth the engine reads.

Pipeline per workbook:
    Estimate sheet "Standard Materials" block
        -> AIEstimating.JobBoughtInMaterials   (raw per-job harvest, provenance)
        -> AIEstimating.Supplier               (upsert by name)
        -> AIEstimating.BoughtInCatalogue      (versioned, effective-dated price rows)

Also sets AIEstimating.CommercialRate values (--set-rate).

USAGE (from C:\\ClaudeVision\\src, venv active; needs pyodbc + xlrd/openpyxl):

    # Dry run (default) — parse + show what WOULD be written:
    python catalogue_loader.py --workbook "K:\\...\\0354158 - ... .xls"

    # Write it:
    python catalogue_loader.py --workbook "K:\\...\\0354158 - ... .xls" --commit

    # Roll rows already landed in JobBoughtInMaterials into the catalogue:
    python catalogue_loader.py --from-landing --commit

    # Set a commercial rate:
    python catalogue_loader.py --set-rate pallet_per_bay=20.00 --commit

Safety: dry-run by default; NVARCHAR-safe inserts (no mojibake); price versioning =
close the current row (effective_to = yesterday) and insert a new one ONLY when the
price actually changed; re-running the same workbook is a no-op.
"""
from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

# ---------------------------------------------------------------- connection ----
SQL_DRIVER = "ODBC Driver 18 for SQL Server"


def connect():
    """Connect using the same credentials as the estimating engine (config.PRICE_SOURCE_CONFIG)."""
    import pyodbc
    try:
        import sys, os
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        import config as _cfg
        c = _cfg.PRICE_SOURCE_CONFIG.get("sqlserver", {})
        server   = c.get("server",   "10.0.0.200")
        database = c.get("database", "SDILive")
        username = c.get("username", "")
        password = c.get("password", "")
    except Exception as e:
        print(f"[warn] Could not load config credentials ({e}), falling back to defaults")
        server, database, username, password = "10.0.0.200", "SDILive", "", ""

    if username and password:
        conn_str = (
            f"DRIVER={{{SQL_DRIVER}}};SERVER={server};DATABASE={database};"
            f"UID={username};PWD={password};"
            "Encrypt=yes;TrustServerCertificate=yes;"
        )
    else:
        # Fallback to Windows auth if no SQL credentials in config
        conn_str = (
            f"DRIVER={{{SQL_DRIVER}}};SERVER={server};DATABASE={database};"
            "Trusted_Connection=yes;TrustServerCertificate=yes;"
        )

    cn = pyodbc.connect(conn_str)
    cn.setdecoding(__import__("pyodbc").SQL_CHAR, encoding="utf-8")
    cn.setencoding(encoding="utf-8")
    return cn


# ---------------------------------------------------------------- categorise ----
def categorise(description: str) -> str:
    """Map a Standard Materials description to a catalogue category."""
    u = (description or "").upper()
    if "TUBE" in u:
        return "tube"
    if "BOX" in u:
        return "box"
    if "PALLET" in u:
        return "pallet"
    if "CARRIAGE" in u or "DELIVERY" in u or "FREIGHT" in u:
        return "delivery"
    if "POWDER" in u:
        return "powder"
    if "EDGING" in u or "EDGEBAND" in u:
        return "edging"
    if ("MFC" in u or "MDF" in u or "PLY" in u) and "EDG" not in u:
        return "board"
    if "STICKER" in u or "LABEL" in u:
        return "label"
    return "fixing"


def catalogue_key(line: Dict[str, Any]) -> str:
    """Stable identity for dedupe/versioning: part code if real, else description slug."""
    pc = str(line.get("part_code") or "").strip()
    if pc and len(pc) >= 3:
        return pc.upper()
    slug = re.sub(r"[^A-Z0-9]+", " ", str(line.get("description") or "").upper()).strip()
    return slug[:120]


# ---------------------------------------------------------------- xls parsing ----
def _open_sheet(path: Path):
    """Return a (nrows, cell(r,c)) accessor for .xls (xlrd) or .xlsx (openpyxl)."""
    if path.suffix.lower() == ".xls":
        import xlrd

        wb = xlrd.open_workbook(str(path))
        sh = wb.sheet_by_name("Estimate")

        def cell(r: int, c: int):
            try:
                return sh.cell_value(r, c)
            except Exception:
                return ""

        return sh.nrows, cell
    else:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), data_only=True, read_only=True)
        ws = wb["Estimate"]
        grid = [list(row) for row in ws.iter_rows(values_only=True)]

        def cell(r: int, c: int):
            try:
                v = grid[r][c]
                return "" if v is None else v
            except Exception:
                return ""

        return len(grid), cell


def parse_estimate_workbook(path: Path) -> Dict[str, Any]:
    """Extract the Standard Materials block from a manual estimate workbook.

    Layout (verified against Tony's Trestle estimate, Rev G):
      r4  c3 = drawing number          r5 c3 = order quantity
      'Bill of Materials' header, then per line:
        c1/c7 part code | c2 description | c8 supplier | c9 price |
        c10 qty per unit | c11 scrap % | c12 total value
      Block ends at the 'Wire' section heading.
    """
    nrows, cell = _open_sheet(path)

    def txt(r: int, c: int) -> str:
        return str(cell(r, c)).strip()

    def num(r: int, c: int) -> Optional[float]:
        v = cell(r, c)
        if isinstance(v, (int, float)):
            return float(v)
        v = re.sub(r"[^\d.\-]", "", str(v))
        return float(v) if v not in ("", "-", ".") else None

    drawing = txt(4, 3) or path.stem
    order_qty = num(5, 3)

    start = None
    for r in range(nrows):
        if txt(r, 2).lower().startswith("bill of materials"):
            start = r + 1
            break
    if start is None:
        raise ValueError(f"'Bill of Materials' header not found in {path.name}")

    lines: List[Dict[str, Any]] = []
    for r in range(start, nrows):
        d = txt(r, 2)
        if d.lower() in ("wire", "sheet steel", "other sheet material"):
            break
        price, qty = num(r, 9), num(r, 10)
        if not d or not price:
            continue
        lines.append(
            {
                "drawing_number": drawing,
                "part_code": (txt(r, 7) or txt(r, 1) or None),
                "description": d,
                "supplier": txt(r, 8) or None,
                "unit_price_gbp": round(price, 4),
                "qty_per_unit": qty,
                "scrap_pct": num(r, 11) or 0.0,
                "total_gbp": round(num(r, 12) or 0.0, 4),
                "category": categorise(d),
            }
        )
    return {"drawing": drawing, "order_qty": order_qty, "lines": lines, "workbook": path.name}


# ---------------------------------------------------------------- db writes -----
def upsert_supplier(cur, name: str, category: Optional[str]) -> Optional[int]:
    if not name:
        return None
    cur.execute("SELECT supplier_id FROM AIEstimating.Supplier WHERE name = ?", name)
    row = cur.fetchone()
    if row:
        return int(row[0])
    cur.execute(
        "INSERT INTO AIEstimating.Supplier(name, category) OUTPUT INSERTED.supplier_id VALUES (?, ?)",
        name, category,
    )
    return int(cur.fetchone()[0])


def land_job_rows(cur, parsed: Dict[str, Any]) -> int:
    """Replace this drawing's rows in JobBoughtInMaterials (idempotent provenance)."""
    cur.execute(
        "DELETE FROM AIEstimating.JobBoughtInMaterials WHERE drawing_number = ?",
        parsed["drawing"],
    )
    n = 0
    for l in parsed["lines"]:
        cur.execute(
            """INSERT INTO AIEstimating.JobBoughtInMaterials
               (drawing_number, part_code, description, supplier, unit_price_gbp,
                qty_per_unit, scrap_pct, total_gbp, source_workbook)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            l["drawing_number"], l["part_code"], l["description"], l["supplier"],
            l["unit_price_gbp"], l["qty_per_unit"], l["scrap_pct"], l["total_gbp"],
            parsed["workbook"],
        )
        n += 1
    return n


def upsert_catalogue(cur, line: Dict[str, Any], source: str, today: dt.date) -> str:
    """Version-aware upsert into BoughtInCatalogue. Returns action taken."""
    key = catalogue_key(line)
    sku = str(line.get("part_code") or "").strip() or None
    desc = line["description"]
    price = float(line["unit_price_gbp"])

    # Current row for this identity (SKU first, else description match)
    if sku:
        cur.execute(
            """SELECT TOP 1 item_id, unit_price_gbp, version FROM AIEstimating.BoughtInCatalogue
               WHERE supplier_sku = ? AND effective_to IS NULL ORDER BY effective_from DESC""",
            sku,
        )
    else:
        cur.execute(
            """SELECT TOP 1 item_id, unit_price_gbp, version FROM AIEstimating.BoughtInCatalogue
               WHERE supplier_sku IS NULL AND description = ? AND effective_to IS NULL
               ORDER BY effective_from DESC""",
            desc,
        )
    row = cur.fetchone()

    if row and abs(float(row[1]) - price) < 0.00005:
        return f"unchanged  {key}"

    supplier_id = upsert_supplier(cur, line.get("supplier"), line.get("category"))
    version = 1
    if row:
        # Close the superseded row the day before the new one starts.
        cur.execute(
            "UPDATE AIEstimating.BoughtInCatalogue SET effective_to = ? WHERE item_id = ?",
            today - dt.timedelta(days=1), int(row[0]),
        )
        version = int(row[2]) + 1

    cur.execute(
        """INSERT INTO AIEstimating.BoughtInCatalogue
           (supplier_id, supplier_sku, description, category, uom, unit_price_gbp,
            effective_from, source, version)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        supplier_id, sku, desc, line.get("category"), "each", price, today, source, version,
    )
    return (f"reprice v{version}  {key}  -> £{price:.4f}" if row
            else f"insert      {key}  £{price:.4f}")


def set_commercial_rate(cur, key: str, value: float, today: dt.date, source: str) -> str:
    cur.execute(
        """SELECT TOP 1 rate_id, value_gbp FROM AIEstimating.CommercialRate
           WHERE rate_key = ? AND effective_to IS NULL ORDER BY effective_from DESC""",
        key,
    )
    row = cur.fetchone()
    if row and abs(float(row[1]) - value) < 0.00005:
        return f"rate unchanged  {key} = £{value:.4f}"
    if row:
        cur.execute(
            "UPDATE AIEstimating.CommercialRate SET effective_to = ? WHERE rate_id = ?",
            today - dt.timedelta(days=1), int(row[0]),
        )
    cur.execute(
        """INSERT INTO AIEstimating.CommercialRate(rate_key, value_gbp, effective_from, source)
           VALUES (?,?,?,?)""",
        key, value, today, source,
    )
    return f"rate set        {key} = £{value:.4f}"


def roll_landing_to_catalogue(cur, today: dt.date) -> List[str]:
    """Mode B: promote everything in JobBoughtInMaterials into the catalogue."""
    cur.execute(
        """SELECT drawing_number, part_code, description, supplier, unit_price_gbp,
                  qty_per_unit, scrap_pct, total_gbp, source_workbook
           FROM AIEstimating.JobBoughtInMaterials
           WHERE unit_price_gbp IS NOT NULL AND unit_price_gbp > 0"""
    )
    actions = []
    for r in cur.fetchall():
        line = {
            "part_code": r[1], "description": r[2], "supplier": r[3],
            "unit_price_gbp": float(r[4]), "category": categorise(r[2] or ""),
        }
        actions.append(upsert_catalogue(cur, line, f"landing:{r[0]}", today))
    return actions


# ---------------------------------------------------------------- main ----------
def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--workbook", action="append", default=[], help="Manual estimate .xls/.xlsx (repeatable)")
    ap.add_argument("--from-landing", action="store_true", help="Roll existing JobBoughtInMaterials rows into the catalogue")
    ap.add_argument("--set-rate", action="append", default=[], metavar="KEY=VALUE",
                    help="Set a CommercialRate, e.g. pallet_per_bay=20.00 (repeatable)")
    ap.add_argument("--commit", action="store_true", help="Write to SQL (default is dry run)")
    args = ap.parse_args()

    if not args.workbook and not args.from_landing and not args.set_rate:
        ap.error("nothing to do: pass --workbook, --from-landing and/or --set-rate")

    today = dt.date.today()
    parsed_books = [parse_estimate_workbook(Path(w)) for w in args.workbook]

    for p in parsed_books:
        print(f"\n=== {p['workbook']}  (drawing {p['drawing']}, order qty {p['order_qty']}) ===")
        by_cat: Dict[str, float] = {}
        for l in p["lines"]:
            by_cat[l["category"]] = by_cat.get(l["category"], 0.0) + l["total_gbp"]
            print(f"  [{l['category']:<8}] {l['description'][:56]:<56} "
                  f"{(l['supplier'] or ''):<16} £{l['unit_price_gbp']:>8.4f} x {l['qty_per_unit']}")
        print("  lane totals/unit: " + "  ".join(f"{k}=£{v:.2f}" for k, v in sorted(by_cat.items())))

    if not args.commit:
        print("\nDRY RUN — nothing written. Re-run with --commit to write to "
              f"{SQL_SERVER}/{SQL_DATABASE}.")
        return

    cn = connect()
    cur = cn.cursor()
    try:
        for p in parsed_books:
            n = land_job_rows(cur, p)
            print(f"\n[landing] {p['drawing']}: {n} row(s) -> JobBoughtInMaterials")
            for l in p["lines"]:
                print("  [catalogue]", upsert_catalogue(cur, l, f"workbook:{p['drawing']}", today))

        if args.from_landing:
            print("\n[landing->catalogue]")
            for a in roll_landing_to_catalogue(cur, today):
                print("  ", a)

        for kv in args.set_rate:
            key, _, val = kv.partition("=")
            print(" ", set_commercial_rate(cur, key.strip(), float(val), today, "manual"))

        cn.commit()
        print("\nCOMMITTED.")
        cur.execute("SELECT COUNT(*) FROM AIEstimating.vCurrentBoughtIn")
        print(f"vCurrentBoughtIn now has {cur.fetchone()[0]} current line(s).")
    except Exception:
        cn.rollback()
        raise
    finally:
        cn.close()


if __name__ == "__main__":
    main()
