"""
SDIAIVision — Estimate xlsx output generator.
Produces a workbook matching the SDI manual estimate sheet format:
  Tab 1: Estimate  — Header + Standard Materials + Sheet Steel + Labour (single sheet)
  Tab 2: Labour    — Department hours/cost summary
  Tab 3: Material Price Break — Quantity break table
"""
from __future__ import annotations
import argparse, json, re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import config


def _packaging_overhead(unit_cost: float, assembly_cost: float = 0.0):
    """E1/E2/E3: per-bay packaging + overhead from config (single source), plus the
    history-derived assembly cost passed in. Returns rows to render + the sell total."""
    pk = getattr(config, "PACKAGING_CONFIG", {}) or {}
    box, pal = pk.get("box", {}) or {}, pk.get("pallet", {}) or {}
    per_box, per_pal, per_del = pk.get("bays_per_box"), pk.get("bays_per_pallet"), pk.get("bays_per_delivery")
    pack, miss = 0.0, []
    if per_box and box.get("price_gbp"): pack += float(box["price_gbp"]) / float(per_box)
    else: miss.append("box")
    if per_pal and pal.get("price_gbp"): pack += float(pal["price_gbp"]) / float(per_pal)
    else: miss.append("pallet")
    if per_del and pk.get("delivery_price_gbp"): pack += float(pk["delivery_price_gbp"]) / float(per_del)
    else: miss.append("delivery")
    pack_costed = not miss
    ovh_pol = getattr(config, "OVERHEAD_POLICY", {}) or {}
    ovh_pct = float(ovh_pol.get("pct", 0.0)) if ovh_pol.get("enabled", False) else 0.0
    pack_val = round(pack, 2) if pack_costed else 0.0
    asm_val = round(float(assembly_cost or 0.0), 2)
    base = unit_cost + asm_val + pack_val
    overhead = round(base * ovh_pct / 100.0, 2)
    sell_total = round(base + overhead, 2)
    return {
        "assembly_gbp": asm_val,
        "packaging_gbp": pack_val,
        "packaging_costed": pack_costed,
        "packaging_note": None if pack_costed else "confirm bays per " + "/".join(miss),
        "overhead_pct": ovh_pct,
        "overhead_gbp": overhead,
        "sell_total_gbp": sell_total,
    }


def _load_bought_in_for(job_number):
    """Bought-in standard materials for this drawing, learned from Tim's BOM
    (job_bought_in_materials.json beside this file). Returns (lines, total_gbp)."""
    import os as _os_bi, json as _json_bi, re as _re_bi
    try:
        p = _os_bi.path.join(_os_bi.path.dirname(_os_bi.path.abspath(__file__)), "job_bought_in_materials.json")
        if not _os_bi.path.exists(p):
            return [], 0.0
        with open(p) as _fh:
            data = (_json_bi.load(_fh) or {}).get("by_drawing", {})
        def _norm(v):
            m = _re_bi.match(r"\s*(\d+)", str(v or ""))
            return m.group(1) if m else str(v or "").strip().upper()
        key = _norm(job_number)
        for k, val in data.items():
            if key and _norm(k) == key:
                return val.get("lines", []), float(val.get("total_gbp") or 0.0)
    except Exception:
        pass
    return [], 0.0

# ── Colours ────────────────────────────────────────────────────────────────────
_SDI_BLUE   = "1F3864"
_SDI_TEAL   = "17375E"
_HDR_YELLOW = "FFC000"   # SDI section header yellow
_COL_YELLOW = "FFFF00"   # column header yellow
_LIGHT_YELL = "FFFACD"   # alternating row tint
_WHITE      = "FFFFFF"
_GREY       = "F2F2F2"
_GREEN      = "E2EFDA"
_RED        = "FCE4D6"
_BLUE_LIGHT = "DEEAF1"
_FONT       = "Arial"

# ── Style helpers ──────────────────────────────────────────────────────────────
def _f(bold=False, size=10, colour="000000", italic=False):
    return Font(name=_FONT, size=size, bold=bold, color=colour, italic=italic)
def _fill(c): return PatternFill("solid", fgColor=c)
def _b():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def _al(h="left"):  return Alignment(horizontal=h, vertical="center", wrap_text=True)
def _safe(v):
    try: return float(v or 0)
    except: return 0.0

def _num(v, default=0.0):
    """float() that tolerates None / non-numeric, falling back to default then 0.0.
    Used for reading the canonical workbook_equivalent_pricing block, where fields
    such as l111_sell_price_gbp are deliberately None until a human sets the margin."""
    try:
        if v is not None:
            return float(v)
    except (TypeError, ValueError):
        pass
    try:
        return float(default) if default is not None else 0.0
    except (TypeError, ValueError):
        return 0.0

def _set(ws, row, col, val, bold=False, fill=None, fmt=None, align="left", size=10, colour="000000", italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = _f(bold=bold, size=size, colour=colour, italic=italic)
    if fill: c.fill = _fill(fill)
    c.border = _b()
    c.alignment = _al(align)
    if fmt: c.number_format = fmt
    return c

def _money(ws, r, col, v, fill=None):
    c = _set(ws, r, col, round(_safe(v),2), fmt='£#,##0.00', align="right")
    if fill: c.fill = _fill(fill)
    return c

def _section_hdr(ws, row, cols, title, merge_to=None):
    """Yellow section header spanning columns."""
    ws.row_dimensions[row].height = 18
    if merge_to:
        ws.merge_cells(start_row=row, start_column=cols, end_row=row, end_column=merge_to)
    c = ws.cell(row=row, column=cols, value=title)
    c.font = _f(bold=True, size=10, colour="000000")
    c.fill = _fill(_HDR_YELLOW)
    c.border = _b()
    c.alignment = _al("center")

def _col_hdrs(ws, row, headers, bg=_COL_YELLOW, height=30):
    ws.row_dimensions[row].height = height
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _f(bold=True, size=9)
        c.fill = _fill(bg)
        c.border = _b()
        c.alignment = _al("center")

def _w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Data helpers ───────────────────────────────────────────────────────────────
def _part_ests(summary): return (summary.get("estimate_summary") or {}).get("part_estimates") or []

def _meta(summary):
    da = summary.get("document_analysis") or {}
    pf = da.get("primary_fields") or {}
    tb = (da.get("title_block") or {}).get("normalized") or {}
    src = summary.get("source_file","").split("\\")[-1].replace(".PDF","").replace(".pdf","")
    wi = (summary.get("estimate_summary") or {}).get("estimate_workbook_inputs") or {}
    return {
        "job_number":  pf.get("job_number") or tb.get("drawing_number") or src,
        "description": pf.get("description") or tb.get("title") or "",
        "customer":    pf.get("customer") or tb.get("customer") or "",
        "quantity":    summary.get("quantity") or summary.get("assumed_job_quantity") or 180,
        "revision":    tb.get("revision") or "",
        "date":        datetime.now().strftime("%d/%m/%Y"),
        "wire_tonne":  _safe(wi.get("wire_cost_per_tonne_gbp") or 1500),
        "steel_tonne": _safe(wi.get("sheet_steel_cost_per_tonne_gbp") or 800),
        "scrap_pct":   _safe(wi.get("scrap_pct") or 4),
        "assembly_pack_labour": (summary.get("bay_estimate") or {}).get("assembly_pack_labour") or {},
    }

def _op_name(op):
    m = {"laser_cutting":"Laser (Metal)","folding":"Fold","welding":"Weld (CO2)",
         "spot_welding":"Spotweld","resistance_welding":"Spotweld",
         "dress_welds":"Dress Welds","powder_coating":"Powder Coat",
         "diamond_polish":"Diamond Polish","cnc_routing":"CNC Route",
         "cnc":"CNC Route","wet_spray":"Wet Spray","assembly":"Assemble",
         "handling":"Handle","packing":"Pack","bench_work":"Bench",
         "hole_machining":"Drill","tapping":"Tap","guillotine":"Guillotine",
         "wire_forming":"Wire Forming","deburring":"Deburr","linisher":"Linish"}
    return m.get(op.lower(), op.replace("_"," ").title())

def _dept(op):
    m = {"laser_cutting":"LASM","laser":"LASM","folding":"FOLD","fold":"FOLD",
         "welding":"WELD","weld":"WELD","spot_welding":"SPOT","resistance_welding":"SPOT",
         "dress_welds":"DRES","powder_coating":"P/C","diamond_polish":"DPOL",
         "cnc_routing":"CNC","cnc":"CNC","wet_spray":"WSPR","assembly":"PACM",
         "handling":"PACM","packing":"PACM","bench_work":"BENC",
         "hole_machining":"DRIL","tapping":"DRIL","guillotine":"GUIL",
         "wire_forming":"WIRE","deburring":"DRES","linisher":"LINS"}
    return m.get(op.lower(), op.upper()[:4])

def _is_sheet_metal(mat):
    return str(mat or "").upper() in {"MILD_STEEL","MILD STEEL","STAINLESS_STEEL",
        "STAINLESS STEEL","ALUMINIUM","ALUMINUM","ZINTEC","BRIGHT_DRAWN"}

def _is_board(mat):
    return str(mat or "").upper() in {"MDF","VENEERED_MDF","OAK_VENEER_MDF",
        "PLYWOOD","BIRCH_PLYWOOD","TIMBER","ACRYLIC","POLYCARBONATE","HDPE_PLASTIC"}

def _labour_rates(le: Dict[str, Any]) -> Dict[str, float]:
    """Hourly rates from labour_estimate; fall back to rate_sources (UDEF/sqlserver path)."""
    rates = {k: _safe(v) for k, v in (le.get("hourly_rates_gbp") or {}).items() if _safe(v)}
    if rates:
        return rates
    for op, rs in (le.get("rate_sources") or {}).items():
        if not isinstance(rs, dict):
            continue
        rate = _safe(rs.get("hourly_rate_gbp"))
        if not rate:
            sel = rs.get("selected") or {}
            rate = _safe(sel.get("price"))
        if rate:
            rates[op] = rate
    return rates

def _labour_time_min(le: Dict[str, Any], pe: Dict[str, Any], op: str) -> float:
    """Run time in minutes for an operation (labour block, then process estimate)."""
    t = _safe((le.get("times_min") or {}).get(op))
    if t:
        return t
    bh = _safe((le.get("batch_hours") or {}).get(op))
    if bh:
        return round(bh * 60.0, 2)
    pe_proc = pe.get("process_estimate") or {}
    t = _safe((pe_proc.get("times_min") or {}).get(op))
    if t:
        return t
    return _safe((pe_proc.get("unit_times_min") or {}).get(op))


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1: ESTIMATE  (matches SDI manual format)
# ══════════════════════════════════════════════════════════════════════════════
def _suppression_banner(summary) -> Optional[str]:
    """Banner text when the headline total is suppressed (INSUFFICIENT DATA), else None.

    Mirrors the Decision Report gate so the Estimate and Material Price Break tabs
    cannot present a confident unit cost / sell price while most of the cost is
    un-credible (e.g. unpriced acrylic). Without this, the Price Break tab showed
    £127.80 with no warning while the Decision Report was flagged NOT FOR QUOTING.
    """
    es = summary.get("estimate_summary") or {}
    ds = es.get("data_sufficiency") or {}
    if not ds.get("suppress_headline_total"):
        return None
    ratio = ds.get("credible_cost_ratio")
    prov = es.get("document_total_provisional_gbp") or ds.get("document_total_provisional_gbp")
    parts = ["\u26a0 INSUFFICIENT DATA \u2014 PROVISIONAL, NOT FOR QUOTING"]
    bits = []
    if isinstance(ratio, (int, float)):
        bits.append(f"credible {ratio:.0%}")
    if isinstance(prov, (int, float)):
        bits.append(f"provisional unit \u00a3{prov:,.2f}")
    if bits:
        parts.append("(" + " \u00b7 ".join(bits) + ")")
    parts.append("\u2014 see Decision Report tab")
    return "   ".join(parts)


def _write_banner(ws, row: int, last_col: str, text: str) -> None:
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = _f(bold=True, size=10, colour=_WHITE)
    c.fill = _fill("C00000")
    c.alignment = _al("center")
    ws.row_dimensions[row].height = 20


def _write_estimate(ws, summary):
    meta  = _meta(summary)
    pes   = _part_ests(summary)
    es    = summary.get("estimate_summary") or {}
    cb    = es.get("cost_breakdown") or {}
    # Sum from individual part estimates — more reliable than aggregate totals
    _mat_parts = sum(
        _safe((pe.get("material_estimate") or {}).get("cost_per_part_gbp", 0)) *
        _safe(pe.get("quantity") or 1)
        for pe in pes
    )
    _lab_parts = sum(
        _safe((pe.get("labour_estimate") or {}).get("total_labour_cost_gbp", 0))
        for pe in pes
    )
    _mat_cb = _safe((cb.get("material") or {}).get("total") or 0)
    _lab_cb = _safe((cb.get("labour") or {}).get("total") or 0)
    # Prefer the reconciled cost_breakdown total (includes bought-in, pack and powder
    # material that the per-part sums omit); max() never drops below the per-part figure,
    # so jobs with an empty/zero cost_breakdown total keep the old behaviour.
    unit_cost = max(_mat_parts, _mat_cb) + max(_lab_parts, _lab_cb)

    # ── Row 1: SDI logo header ─────────────────────────────────────────────
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    # A FALLBACK MUST SAY IT IS ONE.
    #
    # This builder runs only when populate_workbook could not produce the SDI template --
    # an unreachable template share, or a fault inside it. The sheet it writes has always
    # looked exactly like the real thing: same title, same header boxes, same totals block.
    #
    # It is NOT the same thing. It has no route grouping, so it emits one row per part per
    # operation where the template collapses Weld / Dress / P.Coat / Assemble-pack into one
    # row per job. Read as an estimate it therefore shows every assembly operation charged
    # at every level of the tree -- which reads as a cost-model defect and is not one.
    #
    # Two full review cycles on 12120 were spent diagnosing assembly-scope failures from
    # this sheet before anyone noticed which writer produced it. That is the cost of a
    # fallback that cannot be told apart from what it replaced, and it is paid again on
    # every job until the sheet says so itself.
    c.value = ("SDI Displays Limited — AI Estimate  ***  FALLBACK SHEET: the SDI template "
               "could not be populated. Operation rows are NOT route-grouped — weld, dress, "
               "powder and assemble/pack appear once per part instead of once per job, so "
               "labour is OVERSTATED. Do not quote from this sheet.  ***")
    c.font = _f(bold=True, size=12, colour=_WHITE)
    c.fill = _fill(_SDI_BLUE)
    c.alignment = _al("center")
    ws.row_dimensions[1].height = 22

    # Wire/steel cost boxes (top right, cols N-P)
    for r, lbl, val in [
        (1, "Wire Cost Per Tonne =",        f"£  {meta['wire_tonne']:,.2f}"),
        (2, "Sheet Steel Cost Per Tonne =",  f"£  {meta['steel_tonne']:,.2f}"),
    ]:
        _set(ws, r, 14, lbl,  bold=True,  fill=_BLUE_LIGHT, size=9)
        _set(ws, r, 15, val,  bold=True,  fill=_BLUE_LIGHT, size=9, align="right")

    # ── Rows 2-6: Job header block ─────────────────────────────────────────
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 16

    hdr_pairs = [
        (2, "Customer",    meta["customer"]),
        (3, "Description", meta["description"]),
        (4, "Drawing No.", meta["job_number"]),
        (5, "Quantity",    meta["quantity"]),
        (6, "Date",        meta["date"]),
    ]
    for row, lbl, val in hdr_pairs:
        _set(ws, row, 1, lbl, bold=True, fill=_BLUE_LIGHT, size=9)
        _set(ws, row, 2, val, fill=_WHITE, size=9)

    # Headline "Unit Cost" box: show the overhead-absorbed unit cost (M105) so it matches
    # the SDI workbook's Unit Cost cell, not the raw mat+lab. Falls back to the (complete)
    # raw unit_cost if M105 isn't available or the headline is suppressed (low data).
    _wep_box = es.get("workbook_equivalent_pricing") or {}
    _m105_box = _safe(_wep_box.get("m105_total_unit_cost_gbp"))
    _ds_box = es.get("data_sufficiency") or {}
    _box_unit = round(_m105_box, 2) if (_m105_box and not _ds_box.get("suppress_headline_total")) else round(unit_cost, 2)
    rev_pairs = [
        (4, "Rev",         meta["revision"]),
        (5, "Unit Cost",   _box_unit),
        (6, "Prepared By", "SDIAIVision"),
    ]
    for row, lbl, val in rev_pairs:
        _set(ws, row, 3, lbl, bold=True, fill=_BLUE_LIGHT, size=9)
        c = ws.cell(row=row, column=4, value=val)
        c.font = _f(bold=True, size=9)
        c.fill = _fill(_WHITE)
        c.border = _b()
        if isinstance(val, (int, float)) and lbl == "Unit Cost":
            c.number_format = '£#,##0.00'

    # NOT-FOR-QUOTING banner (row 7, between header and Section 1) when the headline
    # total is suppressed — so the Estimate tab can't be read as a quotable number.
    _ban = _suppression_banner(summary)
    if _ban:
        _write_banner(ws, 7, "M", _ban)

    def _part_desc(pe):  # _part_desc_nonefix (2026-07-15)
        """Description cell text: real description if present, else a material+dimension
        fallback. Never prints the literal 'None' (the old f-string leaked None when the
        'description' key existed but was empty)."""
        _pn = str(pe.get("part_number") or "").strip()
        _d = pe.get("description")
        _d = str(_d).strip() if _d is not None else ""
        if _d and _d.lower() != "none":
            return (_pn + "  " + _d).strip()
        # fallback: material + blank dimensions
        _me = pe.get("material_estimate") or {}
        _ng = pe.get("normalized_geometry") or {}
        _mat = str(pe.get("normalized_material") or _me.get("material") or "").replace("_", " ").strip().title()
        _l = _me.get("blank_length_mm") or _ng.get("blank_length_mm")
        _w = _me.get("blank_width_mm") or _ng.get("blank_width_mm")
        _t = pe.get("normalized_thickness_mm") or _me.get("thickness_mm")
        _dims = ""
        try:
            if _l and _w and _t:
                _dims = "%g x %g x %gmm" % (float(_l), float(_w), float(_t))
            elif _l and _w:
                _dims = "%g x %g" % (float(_l), float(_w))
        except (TypeError, ValueError):
            _dims = ""
        _tail = " ".join(x for x in (_dims, _mat) if x).strip()
        return (_pn + "  " + _tail).strip() if _tail else _pn

    # ── SECTION 1: Standard Materials (BOM) ───────────────────────────────
    ROW = 8
    _section_hdr(ws, ROW, 1, "Standard Materials", merge_to=13)
    ROW += 1
    _col_hdrs(ws, ROW, [
        "Bill of Materials (Per Unit)", "", "", "",
        "Part code", "Supplier", "Price", "Qty Per Unit", "Scrap %", "Total Value"
    ])
    ws.merge_cells(f"A{ROW}:D{ROW}")  # merge description cols
    ROW += 1

    bom_start = ROW
    # Bought-in parts (BOUGHT_IN, customer-supplied = £0)
    bom_items = []
    for pe in pes:
        mat = str(pe.get("normalized_material") or "").upper()
        basis = str(pe.get("costing_basis") or "")
        if mat == "BOUGHT_IN" or "customer_supplied" in basis:
            bom_items.append({
                "desc": pe.get("description") or pe.get("part_number") or "",
                "part_code": pe.get("part_number") or "",
                "supplier": "Customer supply",
                "price": 0.0, "qty": _safe(pe.get("quantity") or 1), "scrap": 0, "total": 0.0
            })

    # Powder coat material
    for pe in pes:
        le = pe.get("labour_estimate") or {}
        me = pe.get("material_estimate") or {}
        pc = me.get("powder_consumable") or {}
        if pc and _safe(pc.get("material_cost_gbp")):
            bom_items.append({
                "desc": f"Powder Coat — {pe.get('description','')}", "part_code": "POWDER",
                "supplier": "Valspar", "price": _safe(pc.get("price_per_kg_gbp")),
                "qty": round(_safe(pc.get("kg_per_unit")),4), "scrap": 4,
                "total": _safe(pc.get("material_cost_gbp"))
            })

    # Bought-in standard materials — sourced GENUINELY from the engine's own estimate
    # (drawing-derived): items the engine identified from BOM rows, tube detection, and
    # assembly-note prose, plus the commercial placeholders. We no longer read Tim's
    # manually-transcribed job_bought_in_materials.json — that crutch silently breaks at
    # adoption (no new entries once Tim stops producing manual sheets). Engine-identified
    # bought-ins are marked page_roles=['bought_in']; note-scan items carry source=
    # 'llm_note_scan'; placeholders source='commercial_placeholder'. Powder is excluded
    # (computed as a per-part consumable just above).
    qty = _safe(meta["quantity"])
    _bi_parts = []
    for _pe in _part_ests(summary):
        _roles = _pe.get("page_roles") or []
        _src = str(_pe.get("source") or "")
        _me_chk = _pe.get("material_estimate") or {}
        _is_section = (
            bool(_pe.get("section_stock"))
            or _me_chk.get("stock_form") in ("tube", "section")
            or (_me_chk.get("stock_estimate") or {}).get("section_length_mm") is not None
        )
        _is_bought_in = (
            ("bought_in" in _roles)
            or _src in ("llm_note_scan", "commercial_placeholder")
            or bool(_pe.get("_note_scan"))
            or bool(_pe.get("_commercial_placeholder"))
            or _is_section   # tubes / sections are bought-in stock (Preferred Tubes) — show here, not as flat sheet
        )
        if _is_bought_in:
            _bi_parts.append(_pe)

    for _pe in _bi_parts:
        _me = _pe.get("material_estimate") or {}
        _unit = _safe(
            _pe.get("unit_cost_gbp")
            if _pe.get("unit_cost_gbp") is not None
            else _me.get("cost_per_part_gbp")
        )
        _q = _pe.get("quantity") or 1
        _ext = _safe(
            _pe.get("extended_total_cost_gbp")
            if _pe.get("extended_total_cost_gbp") is not None
            else (_unit * _safe(_q))
        )
        # Surface review flags (AI-identified / estimator-to-price) in the note column.
        _rf = _pe.get("review_flags") or ([] if not _pe.get("review_flag") else ["verify"])
        _note = "; ".join(str(f) for f in _rf)[:90] if _rf else ""
        bom_items.append({
            "desc": _pe.get("description") or _pe.get("part_number") or "",
            "part_code": _pe.get("part_number") or "",
            "supplier": _pe.get("supplier") or (_me.get("price_source") or {}).get("supplier") or "",
            "price": _unit,
            "qty": _q,
            "scrap": 4,
            "total": _ext,
            "note": _note,
        })

    bom_total = 0.0
    for i, item in enumerate(bom_items):
        bg = _WHITE
        for col, val, fmt in [
            (1,item["desc"],    None), (2,"",None), (3,"",None), (4,"",None),
            (5,item["part_code"],None),(6,item["supplier"],None),
        ]:
            _set(ws, ROW, col, val, fill=bg, size=9)
        ws.merge_cells(f"A{ROW}:D{ROW}")
        _money(ws, ROW, 7,  item["price"], fill=bg)
        _set(ws, ROW, 8,  item.get("qty") or None, fill=bg, size=9, align="right")
        _set(ws, ROW, 9,  f'{item["scrap"]}%' if item["scrap"] else "", fill=bg, size=9, align="center")
        note = item.get("note","")
        if note:
            _set(ws, ROW, 10, note, fill=bg, size=9, italic=True, colour="595959")
        else:
            _money(ws, ROW, 10, item["total"], fill=bg)
        bom_total += item["total"]
        ROW += 1

    ROW += 1  # spacer

    # ── SECTION 2: Sheet Steel ────────────────────────────────────────────
    _section_hdr(ws, ROW, 1, "Sheet Steel", merge_to=13)
    ROW += 1
    _col_hdrs(ws, ROW, [
        "Part Description", "Qty\nPer Unit", "Part\nLength", "Part\nWidth",
        "Gauge", "Sheet\nLength", "Sheet\nWidth", "Qty Per\nSheet", "Scrap\n%",
        "Cost Per\nPart"
    ])
    ROW += 1

    steel_total = 0.0
    for pe in pes:
        mat = str(pe.get("normalized_material") or "").upper()
        if not _is_sheet_metal(mat):
            continue
        # Exclude anything already accounted for as a bought-in line (loom, fixings,
        # note-scan items, placeholders) or as linear section stock (tube/RHS/SHS) —
        # those are written in the bought-in section, not as flat sheet, so including
        # them here would double-count. Tubes pass _is_sheet_metal (material=MILD STEEL)
        # so this guard is essential for them specifically.
        _roles = pe.get("page_roles") or []
        _me_chk = pe.get("material_estimate") or {}
        _is_bi = (
            "bought_in" in _roles
            or str(pe.get("source") or "") in ("llm_note_scan", "commercial_placeholder")
            or pe.get("_note_scan") or pe.get("_commercial_placeholder")
        )
        _is_section = bool(pe.get("section_stock")) or _me_chk.get("stock_form") in ("tube", "section") or (_me_chk.get("stock_estimate") or {}).get("section_length_mm") is not None
        if _is_bi or _is_section:
            continue
        me = pe.get("material_estimate") or {}
        se = me.get("stock_estimate") or {}
        ng = pe.get("normalized_geometry") or {}
        blank_l  = _safe(me.get("blank_length_mm") or ng.get("blank_length_mm"))
        blank_w  = _safe(me.get("blank_width_mm")  or ng.get("blank_width_mm"))
        thick    = _safe(pe.get("normalized_thickness_mm") or me.get("thickness_mm"))
        sh_sizes = se.get("candidate_sheet_size_mm") or [2500, 1250]
        sh_l     = sh_sizes[0] if len(sh_sizes) > 0 else 2500
        sh_w     = sh_sizes[1] if len(sh_sizes) > 1 else 1250
        pps      = se.get("parts_per_sheet") or ""
        cost_pp  = _safe(me.get("cost_per_part_gbp"))
        qty_u    = _safe(pe.get("quantity") or 1)
        scrap    = meta["scrap_pct"]
        bg = _LIGHT_YELL

        desc = _part_desc(pe)
        _set(ws, ROW, 1,  desc,   fill=bg, size=9, bold=True)
        _set(ws, ROW, 2,  int(qty_u),  fill=bg, size=9, align="center")
        _set(ws, ROW, 3,  blank_l or None, fill=bg, size=9, align="right", fmt='#,##0')
        _set(ws, ROW, 4,  blank_w or None, fill=bg, size=9, align="right", fmt='#,##0')
        _set(ws, ROW, 5,  thick   or None, fill=bg, size=9, align="center")
        _set(ws, ROW, 6,  sh_l,   fill=bg, size=9, align="right", fmt='#,##0')
        _set(ws, ROW, 7,  sh_w,   fill=bg, size=9, align="right", fmt='#,##0')
        _set(ws, ROW, 8,  pps or None, fill=bg, size=9, align="center")
        _set(ws, ROW, 9,  f'{scrap}%', fill=bg, size=9, align="center")
        _money(ws, ROW, 10, cost_pp, fill=bg)
        steel_total += cost_pp * qty_u
        ROW += 1

    ROW += 1

    # ── SECTION 3: Other Sheet Material (MDF, acrylic, board) ─────────────
    board_parts = [pe for pe in pes if _is_board(pe.get("normalized_material"))]
    if board_parts:
        _section_hdr(ws, ROW, 1, "Other Sheet Material", merge_to=13)
        ROW += 1
        _col_hdrs(ws, ROW, [
            "Part Description", "Qty\nPer Unit", "Part\nLength", "Part\nWidth",
            "Thickness", "Sheet\nLength", "Sheet\nWidth", "Qty Per\nSheet", "Scrap\n%",
            "Cost Per\nPart"
        ])
        ROW += 1
        for pe in board_parts:
            me = pe.get("material_estimate") or {}
            se = me.get("stock_estimate") or {}
            ng = pe.get("normalized_geometry") or {}
            blank_l  = _safe(me.get("blank_length_mm") or ng.get("blank_length_mm"))
            blank_w  = _safe(me.get("blank_width_mm")  or ng.get("blank_width_mm"))
            thick    = _safe(pe.get("normalized_thickness_mm") or me.get("thickness_mm"))
            sh_sizes = se.get("candidate_sheet_size_mm") or [2440, 1220]
            sh_l = sh_sizes[0] if len(sh_sizes) > 0 else 2440
            sh_w = sh_sizes[1] if len(sh_sizes) > 1 else 1220
            pps  = se.get("parts_per_sheet") or ""
            cost_pp = _safe(me.get("cost_per_part_gbp"))
            qty_u   = _safe(pe.get("quantity") or 1)
            bg = _WHITE
            desc = _part_desc(pe)
            _set(ws, ROW, 1, desc, fill=bg, size=9, bold=True)
            _set(ws, ROW, 2, int(qty_u), fill=bg, size=9, align="center")
            _set(ws, ROW, 3, blank_l or None, fill=bg, size=9, align="right", fmt='#,##0')
            _set(ws, ROW, 4, blank_w or None, fill=bg, size=9, align="right", fmt='#,##0')
            _set(ws, ROW, 5, thick or None,   fill=bg, size=9, align="center")
            _set(ws, ROW, 6, sh_l, fill=bg, size=9, align="right", fmt='#,##0')
            _set(ws, ROW, 7, sh_w, fill=bg, size=9, align="right", fmt='#,##0')
            _set(ws, ROW, 8, pps or None, fill=bg, size=9, align="center")
            _set(ws, ROW, 9, f'{meta["scrap_pct"]}%', fill=bg, size=9, align="center")
            _money(ws, ROW, 10, cost_pp, fill=bg)
            ROW += 1
        ROW += 1

    # ── Total Material Cost ───────────────────────────────────────────────
    all_mat = _safe((cb.get("material") or {}).get("total"))
    # Recalculate from individual parts (cost_breakdown total often 0)
    all_mat_recalc = sum(
        _safe((pe.get("material_estimate") or {}).get("extended_material_cost_gbp", 0))
        for pe in pes
    )  # extended_material_cost_gbp = sheet + powder (matches estimator material_total / m59)
    all_mat = all_mat_recalc if all_mat_recalc > 0 else all_mat
    # Bought-in lines (loom, tube, fixings, note-scan items, placeholders) — sourced from the
    # engine estimate (see bought-in section above), NOT Tim's JSON. Sum their extended cost.
    # These parts carry cost in extended_total_cost_gbp / unit_cost_gbp, not material_estimate,
    # so they are NOT already in all_mat_recalc (which sums material_estimate only).
    _bi_total = 0.0
    for _pe in pes:
        _roles = _pe.get("page_roles") or []
        _src = str(_pe.get("source") or "")
        _mec = _pe.get("material_estimate") or {}
        _is_section = (
            bool(_pe.get("section_stock")) or _mec.get("stock_form") in ("tube", "section")
            or (_mec.get("stock_estimate") or {}).get("section_length_mm") is not None
        )
        _is_bi = (
            "bought_in" in _roles
            or _src in ("llm_note_scan", "commercial_placeholder", "prose_recogniser_layer2")
            or _pe.get("_note_scan") or _pe.get("_commercial_placeholder") or _pe.get("_layer2_recognised")
            or _is_section
        )
        if not _is_bi:
            continue
        _u = _pe.get("extended_total_cost_gbp")
        if _u is None:
            _u = _safe(_pe.get("unit_cost_gbp")) * _safe(_pe.get("quantity") or 1)
        _bi_total += _safe(_u)
    all_mat += _bi_total  # engine-sourced bought-in materials (loom, tube, fixings, packaging...)
    _set(ws, ROW, 9, "Total Material Cost", bold=True, fill=_GREY, size=10)
    c = ws.cell(row=ROW, column=10, value=round(all_mat, 2))
    c.font = _f(bold=True, size=10)
    c.fill = _fill(_GREY)
    c.border = _b()
    c.number_format = '£#,##0.00'
    ROW += 2

    # ── SECTION 4: Labour ─────────────────────────────────────────────────
    _section_hdr(ws, ROW, 1, "Labour", merge_to=13)
    ROW += 1
    _col_hdrs(ws, ROW, [
        "Operation", "Part Description", "Dept.", "Qty\nPer Unit",
        "Rate Per\nHour", "Total\nHours", "Labour\nCost",
        "Set Up\n(Mins)", "Total\nValue", ""
    ])
    ROW += 1

    # SRS job number column (col 1 first row per job, like manual)
    srs_written = False
    labour_grand = 0.0
    for pe in pes:
        le = pe.get("labour_estimate") or {}
        costs  = le.get("costs_gbp") or {}
        setups = le.get("setup_times_min") or {}
        rates  = _labour_rates(le)
        ops    = (pe.get("process_estimate") or {}).get("operations") or list(costs.keys())
        if not ops:
            continue

        desc = pe.get("description") or pe.get("part_number") or ""
        qty  = _safe(pe.get("quantity") or 1)

        for i, op in enumerate(ops):
            cost     = _safe(costs.get(op))
            t_min    = _labour_time_min(le, pe, op)
            s_min    = _safe(setups.get(op))
            rate     = _safe(rates.get(op))
            # Back-calculate time if not stored but cost + rate both known
            if not t_min and rate and cost:
                t_min = round((cost / rate) * 60.0, 2)
            hours    = t_min / 60.0
            total_v  = cost + (rate * s_min / 60.0 / max(qty,1))
            bg = _LIGHT_YELL if i % 2 == 0 else _WHITE

            if not srs_written:
                _set(ws, ROW, 1, meta["job_number"], fill=bg, size=8, bold=True, colour="595959")
                srs_written = True
            else:
                _set(ws, ROW, 1, _op_name(op), fill=bg, size=9)

            _set(ws, ROW, 1, _op_name(op),  fill=bg, size=9)
            # DRAWING-DERIVED row description: operation + part + material/thickness the
            # engine extracted, plus op-relevant geometry (bend/hole counts). Shows the
            # engine's own drawing-reading. All fields already on the part estimate.
            _op_verb = {
                "laser_cutting": "Laser (Metal)", "folding": "Fold", "welding": "Weld (CO2)",
                "spot_welding": "Spotweld", "resistance_welding": "Spotweld",
                "dress_welds": "Dress Welds", "powder_coating": "P.Coat",
                "wet_spray": "Wet Spray", "diamond_polish": "Diamond Polish",
                "cnc_routing": "CNC", "cnc": "CNC", "assembly": "Assemble/pack",
                "handling": "Assemble/pack", "packing": "Assemble/pack",
                "hole_machining": "Drill", "tapping": "Tap", "guillotine": "Guillotine",
                "punch": "Punch", "roll": "Roll", "linebend": "Linebend",
                "tube_bending": "Tubebend", "saw": "Saw", "deburring": "Deburr",
            }.get(str(op).lower(), _op_name(op))
            _me   = pe.get("material_estimate") or {}
            _ng   = pe.get("normalized_geometry") or {}
            _matx = str(pe.get("normalized_material") or "").replace("_", " ").strip()
            _thk  = _safe(pe.get("normalized_thickness_mm") or _me.get("thickness_mm"))
            _spec = []
            if _thk:
                _spec.append(("%g" % _thk) + "mm")
            if _matx:
                _spec.append(_matx)
            _detail = ""
            _ol = str(op).lower()
            if _ol == "folding":
                _bn = int(_safe(_ng.get("estimated_bend_line_count")))
                if _bn:
                    _detail = " (%d bend%s)" % (_bn, "" if _bn == 1 else "s")
            elif _ol in ("hole_machining", "drilling", "punch"):
                _hn = int(_safe(_ng.get("estimated_hole_count")))
                if _hn:
                    _detail = " (%d hole%s)" % (_hn, "" if _hn == 1 else "s")
            _base = str(desc).strip().title() if desc else ""
            _row_desc = str(_op_verb)
            if _base:
                _row_desc += " " + "\u2014" + " " + _base
            if _spec:
                _row_desc += ", " + " ".join(_spec)
            _row_desc += _detail
            _set(ws, ROW, 2, _row_desc,      fill=bg, size=9)
            _set(ws, ROW, 3, _dept(op),      fill=bg, size=9, align="center", bold=True)
            _set(ws, ROW, 4, int(qty),        fill=bg, size=9, align="center")
            _money(ws, ROW, 5, rate,          fill=bg)
            c6 = ws.cell(row=ROW, column=6, value=round(hours,4) if hours else None)
            c6.font = _f(size=9); c6.fill = _fill(bg); c6.border = _b()
            c6.number_format = '0.0000'; c6.alignment = _al("right")
            _money(ws, ROW, 7, cost,          fill=bg)
            _set(ws, ROW, 8, round(s_min,1) if s_min else None, fill=bg, size=9, align="center")
            _money(ws, ROW, 9, total_v,       fill=bg)

            labour_grand += cost
            ROW += 1

    # Fabrication labour sub-total (sum of the per-operation rows above; excludes assembly)
    ROW += 1
    _set(ws, ROW, 8, "Fabrication Labour", bold=True, fill=_GREY, size=10)
    c = ws.cell(row=ROW, column=9, value=round(labour_grand,2))
    c.font = _f(bold=True, size=10); c.fill = _fill(_GREY)
    c.border = _b(); c.number_format = '£#,##0.00'

    ROW += 2
    # Recalc fabrication labour from parts if needed
    if labour_grand == 0:
        labour_grand = sum(
            _safe((pe.get("labour_estimate") or {}).get("total_labour_cost_gbp", 0))
            for pe in pes
        )
    # --- Workbook-equivalent pricing rendered STRAIGHT FROM THE CANONICAL JSON block
    #     (estimate_summary.workbook_equivalent_pricing). The JSON is the single source
    #     of truth; the sheet must mirror it, never recompute it, so the two cannot drift.
    #     m105 = ((M59 + M103) / (1 - rebate)) / overhead_factor  (Tim's workbook formula).
    wep = es.get("workbook_equivalent_pricing") or {}
    _asm = meta.get("assembly_pack_labour") or {}
    _asm_cost = round(_num(_asm.get("cost_per_bay_gbp"), 0.0), 2)
    _fab_labour = round(labour_grand, 2)
    m59  = round(_num(wep.get("m59_material_subtotal_gbp"), all_mat), 2)
    m103 = round(_num(wep.get("m103_labour_subtotal_gbp"), _fab_labour + _asm_cost), 2)
    rebate_frac = _num(wep.get("m107_rebate_fraction"), 0.0)
    ovh_factor  = _num(wep.get("overhead_absorption_factor"), 1.0) or 1.0
    margin_frac = _num(wep.get("m109_sell_margin_fraction"), 0.0)
    subtotal = round(m59 + m103, 2)
    _after_rebate = round(subtotal / (1 - rebate_frac), 2) if rebate_frac else subtotal
    # l111 (sell) is often None in the JSON — the estimator leaves margin for a human;
    # fall back to m105 (sell = cost at 0% margin) rather than crashing.
    m105 = round(_num(wep.get("m105_total_unit_cost_gbp"), round(_after_rebate / ovh_factor, 2) if ovh_factor else _after_rebate), 2)
    l111 = round(_num(wep.get("l111_sell_price_gbp"), round(m105 / (1 - margin_frac), 2) if margin_frac else m105), 2)
    total = m105  # headline unit cost (overhead-absorbed) — matches JSON m105 / dashboard
    _summary_rows = [
        ("Total Material Cost (M59)",          m59,       False, None),
        ("  Fabrication labour",               _fab_labour, False, None),
        ("  Assembly / pack (history)",        round(m103 - _fab_labour, 2), False, None),
        ("Total Labour Cost (M103)",           m103,      False, None),
        ("Manufacturing sub-total (M59+M103)", subtotal,  True,  None),
        (f"Rebate gross-up (M107 {rebate_frac*100:.1f}%)", round(_after_rebate - subtotal, 2), False, None),
        (f"Overhead absorption (\u00f7{ovh_factor:g})",    round(m105 - _after_rebate, 2),     False, None),
        ("UNIT COST (M105)",                   m105,      True,  None),
        (f"Sell margin (M109 {margin_frac*100:.1f}%)",     round(l111 - m105, 2),              False, None),
        ("SELL PRICE (L111)",                  l111,      True,  _GREEN),
    ]
    for lbl, val, bold, fillover in _summary_rows:
        fill = fillover or (_HDR_YELLOW if bold else _GREY)
        _set(ws, ROW, 8, lbl, bold=bold, fill=fill, size=10)
        c = ws.cell(row=ROW, column=9, value=round(val, 2))
        c.font = _f(bold=bold, size=10); c.fill = _fill(fill)
        c.border = _b(); c.number_format = '£#,##0.00'
        ROW += 1

    # NOTE: packaging/overhead/rebate/sell are now rendered above, sourced directly from
    # estimate_summary.workbook_equivalent_pricing (single source of truth). The old
    # _packaging_overhead()/rebate recompute was removed because it double-tracked the
    # figures and produced a sell below cost.

    # Column widths (A-J)
    _w(ws, [34, 26, 6, 6, 8, 10, 10, 10, 4, 12])
    ws.freeze_panes = "A8"
    ws.sheet_view.zoomScale = 90


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2: LABOUR  (dept hours summary)
# ══════════════════════════════════════════════════════════════════════════════
def _write_labour(ws, summary):
    meta = _meta(summary)
    pes  = _part_ests(summary)

    ws.merge_cells("A1:F1")
    c = ws["A1"]; c.value = f"Labour Summary — {meta['job_number']}"
    c.font = _f(bold=True, size=12, colour=_WHITE)
    c.fill = _fill(_SDI_BLUE); c.alignment = _al("center")
    ws.row_dimensions[1].height = 22

    _col_hdrs(ws, 3, ["Department", "Hours", "Labour Cost (£)", "Setup (mins)", "Total Cost (£)", "Notes"])

    depts: Dict[str,Dict] = {}
    for pe in pes:
        le    = pe.get("labour_estimate") or {}
        costs = le.get("costs_gbp") or {}
        setups= le.get("setup_times_min") or {}
        rates = _labour_rates(le)
        for op, cost in costs.items():
            d = _dept(op)
            if d not in depts:
                depts[d] = {"name":_op_name(op),"hours":0.0,"cost":0.0,"setup":0.0,"rate":0.0}
            t_min = _labour_time_min(le, pe, op)
            rate = _safe(rates.get(op))
            c_val = _safe(cost)
            if not t_min and rate and c_val:
                t_min = round((c_val / rate) * 60.0, 2)
            depts[d]["hours"]  += t_min / 60.0
            depts[d]["cost"]   += c_val
            depts[d]["setup"]  += _safe(setups.get(op))
            depts[d]["rate"]    = rate

    row = 4
    total_cost = 0.0
    for i, (dept, v) in enumerate(sorted(depts.items())):
        bg = _LIGHT_YELL if i % 2 == 0 else _WHITE
        _set(ws, row, 1, dept, bold=True, fill=bg, size=10)
        c2 = ws.cell(row=row, column=2, value=round(v["hours"],4))
        c2.font = _f(size=10); c2.fill = _fill(bg); c2.border = _b()
        c2.number_format = '0.0000'; c2.alignment = _al("right")
        _money(ws, row, 3, v["cost"], fill=bg)
        _set(ws, row, 4, round(v["setup"],1) if v["setup"] else None, fill=bg, size=10, align="right")
        _money(ws, row, 5, v["cost"], fill=bg)
        _set(ws, row, 6, v["name"], fill=bg, size=9, colour="595959", italic=True)
        total_cost += v["cost"]
        row += 1

    _set(ws, row, 1, "TOTAL", bold=True, fill=_HDR_YELLOW, size=10)
    _money(ws, row, 3, total_cost, fill=_HDR_YELLOW)
    _money(ws, row, 5, total_cost, fill=_HDR_YELLOW)
    _w(ws, [14, 12, 18, 14, 18, 22])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3: MATERIAL PRICE BREAK
# ══════════════════════════════════════════════════════════════════════════════
def _write_price_break(ws, summary):
    meta  = _meta(summary)
    pes   = _part_ests(summary)
    es    = summary.get("estimate_summary") or {}
    pes   = _part_ests(summary)
    cb    = es.get("cost_breakdown") or {}
    _mt = sum(_safe((pe.get("material_estimate") or {}).get("cost_per_part_gbp",0))*_safe(pe.get("quantity") or 1) for pe in pes)
    _lt = sum(_safe((pe.get("labour_estimate") or {}).get("total_labour_cost_gbp",0)) for pe in pes)
    _mcb = _safe((cb.get("material") or {}).get("total") or 0)
    _lcb = _safe((cb.get("labour") or {}).get("total") or 0)
    wep   = es.get("workbook_equivalent_pricing") or {}
    _fallback_unit = (_mt or _mcb) + (_lt or _lcb)
    unit  = round(_num(wep.get("m105_total_unit_cost_gbp"), _fallback_unit), 2)
    sell  = round(_num(wep.get("l111_sell_price_gbp"), unit), 2)

    ws.merge_cells("A1:N1")
    c = ws["A1"]; c.value = f"Material Price Break — {meta['job_number']}"
    c.font = _f(bold=True, size=12, colour=_WHITE)
    c.fill = _fill(_SDI_BLUE); c.alignment = _al("center")
    ws.row_dimensions[1].height = 22

    # NOT-FOR-QUOTING banner (row 2) when the headline total is suppressed — the qty
    # break table must not present confident unit/sell prices on a flagged estimate.
    _ban = _suppression_banner(summary)
    if _ban:
        _write_banner(ws, 2, "N", _ban)

    breaks = [1, 10, 25, 50, 100, 250, 500, 600, 700, 800, 900]
    mults  = [1.0,0.97,0.96,0.94,0.92,0.91,0.90,0.895,0.89,0.885,0.88]

    _col_hdrs(ws, 3, [""] + [f"Qty {q}" for q in breaks])
    # Unit cost = canonical M105 (rebate + overhead already embedded), scaled by qty multiplier.
    _col_hdrs(ws, 4, ["Unit Cost (M105)"] + [f"£{round(unit*m,2):.2f}" for m in mults], bg=_LIGHT_YELL)

    row = 5
    # SELL PRICE = L111 (= M105 / (1 - margin)); rebate is already inside M105, not re-applied here.
    _set(ws, row, 1, "SELL PRICE (L111)", bold=True, fill=_HDR_YELLOW, size=10)
    for i, m in enumerate(mults):
        v = round(sell * m, 2)
        c = ws.cell(row=row, column=2+i, value=v)
        c.font = _f(bold=True, size=10); c.fill = _fill(_HDR_YELLOW)
        c.border = _b(); c.number_format = '£#,##0.00'

    _w(ws, [18] + [12]*11)


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ══════════════════════════════════════════════════════════════════════════════
def write_estimate_xlsx(summary: Dict[str,Any], out_dir=None) -> Path:
    # Default the estimates folder to the SAME absolute output tree as JSON/text/logs
    # (config.OUTPUT_DIR), so spreadsheets always land beside the rest of the output
    # regardless of the launch directory. Was a relative "output/estimates", which
    # followed the cwd and split spreadsheets into src\output when run from src.
    if out_dir is None:
        try:
            import config
            out_dir = config.OUTPUT_DIR / "estimates"
        except Exception:
            out_dir = Path("output/estimates")
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    meta = _meta(summary)
    safe = re.sub(r'[^\w\-]', '_', str(meta["job_number"]))[:60]
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = out_dir / f"{safe}_{ts}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    ws_est = wb.create_sheet("Estimate")
    ws_lab = wb.create_sheet("Labour")
    ws_brk = wb.create_sheet("Material Price Break")

    _write_estimate(ws_est, summary)
    _write_labour(ws_lab, summary)
    _write_price_break(ws_brk, summary)

    # Estimate tab active on open
    ws_est.sheet_view.tabSelected = True
    wb.active = ws_est

    wb.save(path)
    print(f"   -> Estimate xlsx written: {path}")
    return path


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", required=True)
    ap.add_argument("--out", default="output/estimates")
    args = ap.parse_args()
    with open(args.json, encoding="utf-8") as f:
        data = json.load(f)
    write_estimate_xlsx(data, out_dir=args.out)
