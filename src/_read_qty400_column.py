# -*- coding: utf-8 -*-
"""READ-ONLY. Reads the workbook's computed QTY-400 column (the one comparable to Tim's
qty-400 manual sheet) from the latest 1298 output, and lays it beside Tim's figures.

The quantity-break columns are formula-driven (calc-on-load). This tries BOTH:
  - data_only=True  -> cached COMPUTED values (only present if opened+saved in Excel once)
  - data_only=False -> the formula strings (always present)
and reports which we got, so we don't mistake an un-opened file's blank cells for zeros.

The Quantity Breaks tab header row maps columns to quantities:
  D=6  E=50  F=100  G=250  H=400  I=1000   (H is the qty-400 column = Tim's column)

Run from C:\\ClaudeVision\\src :
  C:\\ClaudeVision\\.venv\\Scripts\\python.exe _read_qty400_column.py
"""
import re
from pathlib import Path
import openpyxl

est_dir = Path(r"C:\ClaudeVision\output\estimates")
cands = sorted(est_dir.glob("1298*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
if not cands:
    print("(no 1298 output sheet found)")
    raise SystemExit(0)
latest = cands[0]
print(f"sheet: {latest.name}\n")

# Find the 'Quantity Breaks' sheet/area and the qty-400 column.
def find_breaks_sheet(wb):
    for name in wb.sheetnames:
        ws = wb[name]
        for r in range(1, 20):
            for c in range(1, 16):
                v = ws.cell(row=r, column=c).value
                if v and "quantity break" in str(v).lower():
                    return name, r
    return None, None

for mode in (True, False):
    label = "CACHED COMPUTED VALUES" if mode else "FORMULAS"
    print(f"===== read with data_only={mode}  ({label}) =====")
    try:
        wb = openpyxl.load_workbook(latest, data_only=mode)
    except Exception as e:
        print("  load error:", e); continue
    sheet, hdr = find_breaks_sheet(wb)
    if not sheet:
        print("  (no 'Quantity Breaks' area found by header text; scanning 'Estimate' sheet cols instead)")
    else:
        ws = wb[sheet]
        print(f"  Quantity Breaks on sheet '{sheet}', header row {hdr}")
        # print the header row (quantities) and the next ~20 rows, cols C..I
        for r in range(hdr, hdr+22):
            vals = []
            for c in range(3, 10):  # C..I
                v = ws.cell(row=r, column=c).value
                vals.append("" if v in (None, "") else str(v)[:16])
            if any(vals):
                colL = "  ".join(f"{openpyxl.utils.get_column_letter(c)}:{vals[c-3]}" for c in range(3,10))
                print(f"    r{r}: {colL}")
    print()

print("=== Tim's qty-400 manual figures (for side-by-side) ===")
print("  Material total £0.87  (Bracket £0.34 + Powder £0.25 + polybag £0.05 + fastener £0.13 + pallet £0.03 + delivery £0.08)")
print("  Labour total   £2.02  (Laser £0.29 + Fold £0.86 + P.Coat £0.55 + Pack £0.32)")
print("  Unit total     £3.10")
print()
print("READ: find the qty-400 column (H, header=400) and compare its LASER / labour / total")
print("to Tim's. If cached values are None everywhere, the file hasn't been opened in Excel —")
print("open it once, save, and re-run; OR just open in Excel and read the 400 column directly.")
