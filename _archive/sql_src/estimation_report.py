"""
SDI Intelligence — Estimation Provenance Report
===========================================
Generates a detailed "how we got there" report alongside
every estimate xlsx. Shows exactly where each material,
thickness, cost and operation came from.

Critical for:
  - MD/FD confidence in AI estimates
  - Estimator trust and verification
  - Learning system audit trail
  - Identifying where AI is uncertain

Output: Added as extra sheet "AI Provenance" in estimate xlsx
        AND as a standalone PDF report (optional)

Called from estimator.py after xlsx is written.
"""

import json
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter
    _XLSX_OK = True
except ImportError:
    _XLSX_OK = False

try:
    import corrections_db as db
    _DB_OK = True
except ImportError:
    _DB_OK = False


# ── Colours ────────────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # SDI dark navy
C_HEADER_FG   = "FFFFFF"
C_HIGH        = "C6EFCE"   # green  — high confidence
C_MEDIUM      = "FFEB9C"   # amber  — medium confidence
C_LOW         = "FFC7CE"   # red    — low confidence / needs review
C_KB          = "DDEEFF"   # blue   — from knowledge base
C_HIST        = "E8D5FF"   # purple — from historical data
C_AI          = "FFF2CC"   # yellow — AI inference
C_RULE        = "D9EAD3"   # light green — override rule fired
C_SECTION     = "2F5496"   # section header blue
C_ALT_ROW     = "F5F5F5"   # alternating row


def confidence_colour(confidence: float) -> str:
    if confidence >= 0.85:
        return C_HIGH
    elif confidence >= 0.60:
        return C_MEDIUM
    else:
        return C_LOW


def confidence_label(confidence: float) -> str:
    if confidence >= 0.85:
        return "HIGH ✓"
    elif confidence >= 0.60:
        return "MEDIUM"
    else:
        return "LOW — REVIEW"


def source_label(source: str) -> str:
    """Human-readable explanation of where a value came from."""
    s = str(source or "").lower()
    if "knowledge_base" in s:
        return "SDI Knowledge Base (previously confirmed)"
    elif "override_rule" in s:
        name = re.search(r'override_rule:(.+)', s)
        rule = name.group(1) if name else "learning rule"
        return f"Learning Rule fired: {rule}"
    elif "dxf_flat_pattern" in s:
        return "DXF flat pattern file (exact geometry)"
    elif "dxf_filename" in s:
        return "DXF filename material code (e.g. _MS_, PETG)"
    elif "historical" in s:
        return "Historical SDI estimate match"
    elif "pn_suffix" in s:
        return "Part number suffix (-M=Steel, -A=Acrylic, -T=MDF)"
    elif "pdf" in s:
        return "PDF drawing text extraction"
    elif "ai" in s or "inference" in s:
        return "AI inference (Claude)"
    else:
        return str(source or "AI inference")


def build_provenance(summary: Dict[str, Any]) -> List[Dict]:
    """
    Extract provenance data from a scan summary.
    Returns list of part provenance records.
    """
    parts = (summary.get("manufacturing_writeup") or {}).get("parts") or []
    provenance = []

    # SDI Intelligence — cost lives in estimate_summary.part_estimates, keyed by
    # part_number. Build a lookup so the provenance report shows real costs.
    _est_lookup = {}
    for _pe in (summary.get("estimate_summary") or {}).get("part_estimates", []):
        _pn = _pe.get("part_number")
        if _pn:
            _est_lookup[_pn] = _pe

    for part in parts:
        pn   = part.get("part_number") or "—"
        desc = part.get("description") or "—"
        mat  = part.get("normalized_material") or part.get("material") or "Unknown"
        qty  = part.get("quantity", 1)
        _pe = _est_lookup.get(pn, {})
        unit = float(_pe.get("unit_total_cost_gbp") or 0)
        ext  = float(_pe.get("extended_total_cost_gbp") or 0)
        geo  = str(part.get("geometry_source") or "pdf")
        # Merge textual + inferred ops so auto-assigned laser/handling/powder show.
        ops  = list(part.get("textual_operations") or []) + list(part.get("inferred_operations") or [])

        # ── Material provenance ────────────────────────────────────────────────
        mat_source   = part.get("material_source") or geo
        mat_conf     = 0.9 if "knowledge_base" in mat_source else \
                       0.95 if "dxf_filename" in mat_source or "dxf_flat" in geo else \
                       0.7  if "pn_suffix" in mat_source else \
                       0.6  if "pdf" in geo else 0.5

        # ── Thickness provenance ───────────────────────────────────────────────
        # DXF filename FIRST — most reliable, and avoids real 2mm/3mm acrylic
        # being wrongly stripped as tolerance-table values.
        import re as _re
        thk_val = None
        thk_source = "Not extracted (tolerance table only)"
        _dfn = str(part.get("dxf_source_file") or "")
        _tm = _re.search(r"[_\-\s](\d+\.?\d*)\s*mm", _dfn, _re.IGNORECASE)
        if _tm:
            _tv = float(_tm.group(1))
            if 0.3 <= _tv <= 25.0:
                thk_val = _tv
                thk_source = f"DXF filename ({_dfn})"
        if thk_val is None:
            thicknesses = part.get("thicknesses_mm") or []
            tol_set     = {0.5, 1.0, 1.5, 2.0, 3.0}
            _t_set      = {round(float(t),1) for t in thicknesses if t}
            # Only strip tolerance values if the FULL sequence is present;
            # a standalone 2.0/3.0 is a real thickness.
            if tol_set.issubset(_t_set):
                thk_clean = [t for t in thicknesses
                             if t and round(float(t), 1) not in tol_set]
            else:
                thk_clean = [t for t in thicknesses if t]
            if thk_clean:
                thk_val = thk_clean[0]
                thk_source = "DXF geometry" if "dxf" in geo else "PDF text"
        thk_conf     = 0.95 if "dxf" in thk_source.lower() else \
                       0.7  if thk_val else 0.2

        # ── Geometry provenance ────────────────────────────────────────────────
        geo_data     = part.get("geometry_rollup") or {}
        cut_len      = float(geo_data.get("estimated_cut_length_mm") or 0)
        n_holes      = int(geo_data.get("estimated_hole_count") or 0)
        n_bends      = int(geo_data.get("estimated_bend_line_count") or 0)
        geo_conf_raw = float((geo_data.get("confidence") or {})
                             .get("estimated_cut_length_mm") or
                             geo_data.get("geometry_reliability") or 0)
        geo_source   = "DXF flat pattern (exact)" if "dxf" in geo else \
                       f"PDF vector extraction (reliability {geo_conf_raw:.0%})"

        # ── Cost provenance ────────────────────────────────────────────────────
        hist_match   = None
        if _DB_OK and pn and pn != "—":
            hist_match = db.get_historical_cost(part_number=pn)

        # ── Flags ──────────────────────────────────────────────────────────────
        flags = []
        learning_flag = part.get("_learning_flag") or ""
        if learning_flag:
            flags.extend(learning_flag.split(" | "))
        if unit == 0.0 and mat not in ("BOUGHT_IN",):
            flags.append("Zero cost — thickness or geometry missing")
        if mat in ("Unknown", "UNKNOWN", "LED", "CARD"):
            flags.append(f"Material unresolved: {mat!r}")
        if not thk_val and mat in ("MILD_STEEL", "ACRYLIC", "MDF"):
            flags.append("No thickness extracted — manual review needed")

        # ── Override rules that fired ──────────────────────────────────────────
        overrides_fired = []
        if "override_rule:" in mat_source:
            rule_name = mat_source.split("override_rule:")[-1]
            overrides_fired.append(f"Material rule: {rule_name}")
        if part.get("dxf_source_file"):
            dxf = part["dxf_source_file"].upper()
            if "_MS_" in dxf:
                overrides_fired.append("DXF filename _MS_ → MILD_STEEL")
            elif "PETG" in dxf:
                overrides_fired.append("DXF filename PETG → ACRYLIC")
            elif "JOINERY" in dxf:
                overrides_fired.append("DXF filename JOINERY → MDF")

        # ── Overall confidence ─────────────────────────────────────────────────
        overall_conf = min(mat_conf, thk_conf if thk_val else 0.6,
                           geo_conf_raw if geo_conf_raw > 0 else 0.7)

        provenance.append({
            "part_number":       pn,
            "description":       desc,
            "quantity":          qty,
            "material":          mat,
            "material_source":   source_label(mat_source),
            "material_conf":     mat_conf,
            "thickness_mm":      thk_val,
            "thickness_source":  thk_source,
            "thickness_conf":    thk_conf,
            "cut_length_mm":     cut_len,
            "n_holes":           n_holes,
            "n_bends":           n_bends,
            "geometry_source":   geo_source,
            "geometry_conf":     geo_conf_raw,
            "dxf_file":          part.get("dxf_source_file") or "—",
            "operations":        ", ".join(ops) if ops else "—",
            "unit_cost":         unit,
            "extended_cost":     ext,
            "overall_confidence":overall_conf,
            "flags":             flags,
            "overrides_fired":   overrides_fired,
            "historical_match":  hist_match,
        })

    return provenance


def add_provenance_sheet(wb, summary: Dict[str, Any],
                          scan_meta: Dict[str, Any] = None) -> None:
    """
    Add an 'AI Provenance' sheet to an existing openpyxl workbook.
    Call this after the main estimate sheet is written.
    """
    if not _XLSX_OK:
        return

    provenance = build_provenance(summary)
    if not provenance:
        return

    ws = wb.create_sheet("AI Provenance")
    scan_meta = scan_meta or {}

    def cell(row, col, value="", bold=False, bg=None, fg="000000",
             align="left", wrap=False, size=10, border=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Calibri", bold=bold, color=fg, size=size)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center",
                                 wrap_text=wrap)
        if border:
            thin = Side(style="thin", color="BBBBBB")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return c

    # ── Title block ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:M1")
    cell(1, 1, "SDI Intelligence — Estimate Provenance Report",
         bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, align="center", size=13)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:M2")
    pdf_name = scan_meta.get("pdf_name") or summary.get("source_file") or "—"
    job_no   = scan_meta.get("job_number") or "—"
    scan_dt  = scan_meta.get("scan_date") or datetime.now().strftime("%d/%m/%Y %H:%M")
    cell(2, 1,
         f"Drawing: {pdf_name}   |   Job: {job_no}   |   Scanned: {scan_dt}   |   "
         f"Parts: {len(provenance)}   |   "
         f"Est. Total: £{sum(p['extended_cost'] for p in provenance):.2f}",
         bg="2F5496", fg=C_HEADER_FG, align="center", size=10)
    ws.row_dimensions[2].height = 18

    # ── Legend ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A3:M3")
    cell(3, 1,
         "CONFIDENCE KEY:   🟢 HIGH ≥85%  (Knowledge Base / DXF file)     "
         "🟡 MEDIUM 60-85%  (PDF extraction / PN suffix)     "
         "🔴 LOW <60%  (AI inference only — review recommended)",
         bg="F0F0F0", align="left", size=9)
    ws.row_dimensions[3].height = 16

    ws.row_dimensions[4].height = 6  # spacer

    # ── Column headers ─────────────────────────────────────────────────────────
    headers = [
        ("Part Number",       15), ("Description",     28), ("Qty", 5),
        ("Material",          14), ("Mat. Source",      32), ("Conf.",  8),
        ("Thickness",          9), ("Thk. Source",      22), ("Geometry Source", 26),
        ("Cut (mm)",          10), ("Ops",              22),
        ("Unit £",             9), ("Ext £",             9),
    ]
    for ci, (hdr, width) in enumerate(headers, 1):
        c = cell(5, ci, hdr, bold=True, bg=C_SECTION, fg=C_HEADER_FG,
                 align="center", size=10)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[5].height = 20

    # ── Part rows ──────────────────────────────────────────────────────────────
    row = 6
    for i, p in enumerate(provenance):
        bg = C_ALT_ROW if i % 2 == 0 else "FFFFFF"
        conf_bg = confidence_colour(p["overall_confidence"])

        cell(row, 1,  p["part_number"],        bg=bg,       border=True)
        cell(row, 2,  p["description"],        bg=bg,       border=True, wrap=True)
        cell(row, 3,  p["quantity"],            bg=bg,       align="center", border=True)
        cell(row, 4,  p["material"],            bg=conf_bg,  bold=True, border=True)
        cell(row, 5,  p["material_source"],     bg=conf_bg,  border=True, wrap=True, size=9)
        cell(row, 6,  f"{p['material_conf']:.0%}",
                                                bg=conf_bg,  align="center", border=True)
        cell(row, 7,  f"{p['thickness_mm']}mm" if p["thickness_mm"] else "—",
                                                bg=bg,       align="center", border=True)
        cell(row, 8,  p["thickness_source"],    bg=bg,       border=True, size=9, wrap=True)
        cell(row, 9,  p["geometry_source"],     bg=bg,       border=True, size=9, wrap=True)
        cell(row, 10, f"{p['cut_length_mm']:.0f}" if p["cut_length_mm"] else "—",
                                                bg=bg,       align="right", border=True)
        cell(row, 11, p["operations"],          bg=bg,       border=True, size=9, wrap=True)
        cell(row, 12, f"£{p['unit_cost']:.2f}", bg=bg,       align="right",
             bold=True, border=True)
        cell(row, 13, f"£{p['extended_cost']:.2f}", bg=bg,   align="right",
             bold=True, border=True)
        ws.row_dimensions[row].height = 28
        row += 1

        # ── Flags / warnings ───────────────────────────────────────────────────
        if p["flags"]:
            for flag in p["flags"]:
                ws.merge_cells(f"B{row}:M{row}")
                cell(row, 1, "⚠",              bg=C_LOW, align="center", size=9)
                cell(row, 2, f"REVIEW: {flag}", bg=C_LOW, size=9, wrap=True)
                ws.row_dimensions[row].height = 16
                row += 1

        # ── Override rules that fired ──────────────────────────────────────────
        if p["overrides_fired"]:
            ws.merge_cells(f"B{row}:M{row}")
            cell(row, 1, "🧠",                  bg=C_RULE, align="center", size=9)
            cell(row, 2, "Learning: " + " | ".join(p["overrides_fired"]),
                 bg=C_RULE, size=9)
            ws.row_dimensions[row].height = 14
            row += 1

        # ── Historical matches ─────────────────────────────────────────────────
        if p["historical_match"]:
            for hm in (p["historical_match"] or [])[:2]:
                avg  = hm.get("AvgCost") or hm.get("avg_cost") or 0
                cnt  = hm.get("SampleCount") or hm.get("sample_count") or 0
                hmat = hm.get("Material") or hm.get("material") or "?"
                ws.merge_cells(f"B{row}:M{row}")
                cell(row, 1, "📚",              bg=C_HIST, align="center", size=9)
                cell(row, 2,
                     f"Historical: {cnt} SDI estimate(s) for this part as "
                     f"{hmat} — avg £{avg:.2f}",
                     bg=C_HIST, size=9)
                ws.row_dimensions[row].height = 14
                row += 1

    # ── Summary footer ─────────────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:L{row}")
    total = sum(p["extended_cost"] for p in provenance)
    cell(row, 1,  "ESTIMATE TOTAL", bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG,
         align="right", size=11)
    cell(row, 13, f"£{total:.2f}",  bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG,
         align="right", size=12)
    ws.row_dimensions[row].height = 22

    row += 2
    ws.merge_cells(f"A{row}:M{row}")
    high_count = sum(1 for p in provenance if p["overall_confidence"] >= 0.85)
    med_count  = sum(1 for p in provenance if 0.60 <= p["overall_confidence"] < 0.85)
    low_count  = sum(1 for p in provenance if p["overall_confidence"] < 0.60)
    cell(row, 1,
         f"Confidence summary:  "
         f"🟢 HIGH: {high_count} parts   "
         f"🟡 MEDIUM: {med_count} parts   "
         f"🔴 LOW: {low_count} parts   |   "
         f"Generated by SDI Intelligence  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}",
         bg="F0F0F0", size=9, align="left")

    # Freeze panes below header + column A
    ws.freeze_panes = "B6"

    # Tab colour
    ws.sheet_properties.tabColor = "1F3864"


def generate_standalone_report(summary: Dict[str, Any],
                                output_path: str,
                                scan_meta: Dict[str, Any] = None) -> str:
    """
    Generate a standalone provenance xlsx report.
    Returns path to created file.
    """
    if not _XLSX_OK:
        print("[provenance] openpyxl not available")
        return ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cover"

    # Simple cover page
    ws["A1"] = "SDI Intelligence Estimation Provenance Report"
    ws["A2"] = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"] = f"Drawing: {(scan_meta or {}).get('pdf_name', '—')}"
    ws["A4"] = f"Job: {(scan_meta or {}).get('job_number', '—')}"

    add_provenance_sheet(wb, summary, scan_meta)
    wb.save(output_path)
    print(f"[provenance] Report saved: {output_path}")
    return output_path


if __name__ == "__main__":
    print("SDI Intelligence Provenance Report module ready.")
    print()
    print("Integration into estimator.py:")
    print("  from estimation_report import add_provenance_sheet")
    print("  # After writing main estimate sheet:")
    print("  add_provenance_sheet(wb, summary, scan_meta)")
    print("  wb.save(xlsx_path)")
