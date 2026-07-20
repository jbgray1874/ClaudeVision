"""
SDI Intelligence — Tim rate-card ingester.

Learns the department labour rate card (and per-dept setup minutes) from an SDI
manual estimate workbook and writes it to AIEstimating, so the engine sources
rates by DEPARTMENT CODE instead of hand-mapped per-operation values that drift.

Generalises to any SDI estimate workbook: the "Total Labour Hours By Dept." block
has a stable layout (H=operation label, I=rate £/hr, J=dept code, K=setup mins).
Input must be .xlsx/.xlsm (openpyxl); convert a true .xls first.

Outputs (no DB required):
  - tim_rate_card.json   {by_dept, by_op, setup_min_by_dept, source, ingested}
  - tim_rate_card.sql    MERGE into AIEstimating.LabourRateCard
Optional:
  --write-db             write straight into SDILive (uses config DB creds, needs VPN)
"""
from __future__ import annotations
import argparse, json, datetime
from pathlib import Path
import openpyxl

# Tim's operation labels (dept block, col H) -> the engine's operation names.
# This is the ONE stable mapping we maintain; the RATES come from the sheet.
TIM_LABEL_TO_OP = {
    "punch": "punch", "fold": "folding", "guillotine": "guillotine",
    "laser (metal)": "laser_cutting", "laser (acrylic)": "laser_cutting_acrylic",
    "weld (co2)": "welding", "spotweld": "spot_welding", "dress welds": "dress_welds",
    "roll": "roll", "saw": "saw", "glue": "glue", "tube": "tube", "tubebend": "tube_bend",
    "assemble/pack (metal)": "assembly", "assemble/pack (acrylic)": "assembly_acrylic",
    "manual labour (metal)": "manual_labour_metal", "manual labour (acrylic)": "manual_labour_acrylic",
    "p.coat": "powder_coating", "wet spray": "wet_spray", "cnc": "cnc", "cnc joinery": "cnc_joinery",
    "bench work joinery": "bench_work", "diamond polish": "diamond_polish",
    "drill (acrylic)": "hole_machining", "linebend": "linebend", "pin router": "pin_router",
    "robomac": "robomac", "salvagnini": "salvagnini", "oven": "oven", "edge banding": "edge_banding",
    "packing joinery": "packing_joinery", "machines joinery": "machines_joinery",
}

def parse_rate_card(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Estimate"] if "Estimate" in wb.sheetnames else wb.worksheets[0]
    # locate the dept-block header: H="Labour", I="Rate", J="Dept"
    hdr = None
    for r in range(1, ws.max_row + 1):
        h, i, j = (ws.cell(r, c).value for c in (8, 9, 10))
        if str(h).strip().lower() == "labour" and str(i).strip().lower() == "rate" and str(j).strip().lower() == "dept":
            hdr = r; break
    if hdr is None:
        raise SystemExit("Could not find the 'Labour / Rate / Dept' header block on the Estimate sheet.")
    by_dept, by_op, setup_by_dept, rows = {}, {}, {}, []
    for r in range(hdr + 1, ws.max_row + 1):
        label, rate, dept, setup = (ws.cell(r, c).value for c in (8, 9, 10, 11))
        if not dept or not isinstance(rate, (int, float)):
            continue
        d = str(dept).strip().upper()
        if not d or not d[0].isalpha():
            continue  # past the labour block (e.g. wire price-break rows have numeric J) — stop reading rates
        lab = str(label).strip()
        by_dept[d] = round(float(rate), 4)
        if isinstance(setup, (int, float)):
            setup_by_dept[d] = float(setup)
        op = TIM_LABEL_TO_OP.get(lab.lower())
        if op:
            by_op[op] = round(float(rate), 4)
        rows.append((d, lab, round(float(rate), 4), setup if isinstance(setup, (int, float)) else None))
    return by_dept, by_op, setup_by_dept, rows

def emit_sql(rows, source: str) -> str:
    lines = [
        "IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='AIEstimating') EXEC('CREATE SCHEMA AIEstimating');",
        "IF OBJECT_ID('AIEstimating.LabourRateCard','U') IS NULL",
        "CREATE TABLE AIEstimating.LabourRateCard (",
        "    department_code varchar(10) NOT NULL PRIMARY KEY,",
        "    operation_label varchar(80) NULL,",
        "    hourly_rate_gbp  decimal(10,4) NOT NULL,",
        "    setup_minutes    decimal(8,2) NULL,",
        "    source_workbook  varchar(260) NULL,",
        "    ingested_utc     datetime2 NOT NULL DEFAULT SYSUTCDATETIME());",
        "",
    ]
    for d, lab, rate, setup in rows:
        s = "NULL" if setup is None else f"{setup}"
        labq = lab.replace("'", "''")
        srcq = source.replace("'", "''")
        lines.append(
            "MERGE AIEstimating.LabourRateCard AS t "
            f"USING (SELECT '{d}' dc) AS s ON t.department_code=s.dc "
            f"WHEN MATCHED THEN UPDATE SET operation_label='{labq}', hourly_rate_gbp={rate}, setup_minutes={s}, source_workbook='{srcq}', ingested_utc=SYSUTCDATETIME() "
            f"WHEN NOT MATCHED THEN INSERT(department_code,operation_label,hourly_rate_gbp,setup_minutes,source_workbook) VALUES('{d}','{labq}',{rate},{s},'{srcq}');"
        )
    return "\n".join(lines) + "\n"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", help="SDI manual estimate .xlsx")
    ap.add_argument("--out-dir", default=".")
    ap.add_argument("--write-db", action="store_true", help="also write to SDILive (needs VPN + config creds)")
    a = ap.parse_args()
    by_dept, by_op, setup_by_dept, rows = parse_rate_card(a.workbook)
    payload = {
        "by_dept": by_dept, "by_op": by_op, "setup_min_by_dept": setup_by_dept,
        "source": Path(a.workbook).name, "ingested": datetime.datetime.now(datetime.timezone.utc).isoformat(),
    }
    outd = Path(a.out_dir)
    (outd / "tim_rate_card.json").write_text(json.dumps(payload, indent=2))
    (outd / "tim_rate_card.sql").write_text(emit_sql(rows, Path(a.workbook).name))
    print(f"Parsed {len(rows)} department rates from {payload['source']}")
    print(f"  -> {outd/'tim_rate_card.json'}\n  -> {outd/'tim_rate_card.sql'}")
    if a.write_db:
        try:
            import pyodbc, config
            cn = config.get_connection() if hasattr(config, "get_connection") else None
            cur = cn.cursor()
            cur.execute(open(outd / "tim_rate_card.sql").read())
            cn.commit(); print("  -> written to AIEstimating.LabourRateCard")
        except Exception as exc:
            print(f"  (DB write skipped: {exc})")

if __name__ == "__main__":
    main()
