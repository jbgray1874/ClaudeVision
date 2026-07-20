"""
SDIAIVision — Estimate xlsx output generator.
Produces a workbook matching the SDI manual estimate sheet format:
  Tab 1: Estimate  — Header + Standard Materials + Sheet Steel + Labour (single sheet)
  Tab 2: Labour    — Department hours/cost summary
  Tab 3: Material Price Break — Quantity break table
"""
from __future__ import annotations
import argparse, json, re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Colours ────────────────────────────────────────────────────────────────────
_SDI_BLUE   = "1F3864"
_SDI_TEAL   = "17375E"
_HDR_YELLOW = "FFC000"   # SDI section header yellow
_COL_YELLOW = "FFFF00"   # column header yellow
_LIGHT_YELL = "FFFACD"   # alternating row tint
_WHITE      = "FFFFFF"
_GREY       = "F2F2F2"
_GREEN      = "E2EFDA"
_RED        = "FCE4D6"
_BLUE_LIGHT = "DEEAF1"
_FONT       = "Arial"

# ── Style helpers ──────────────────────────────────────────────────────────────
def _f(bold=False, size=10, colour="000000", italic=False):
    return Font(name=_FONT, size=size, bold=bold, color=colour, italic=italic)
def _fill(c): return PatternFill("solid", fgColor=c)
def _b():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def _al(h="left"):  return Alignment(horizontal=h, vertical="center", wrap_text=True)
def _safe(v):
    try: return float(v or 0)
    except: return 0.0

def _set(ws, row, col, val, bold=False, fill=None, fmt=None, align="left", size=10, colour="000000", italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = _f(bold=bold, size=size, colour=colour, italic=italic)
    if fill: c.fill = _fill(fill)
    c.border = _b()
    c.alignment = _al(align)
    if fmt: c.number_format = fmt
    return c

def _money(ws, r, col, v, fill=None):
    c = _set(ws, r, col, round(_safe(v),2), fmt='£#,##0.00', align="right")
    if fill: c.fill = _fill(fill)
    return c

def _section_hdr(ws, row, cols, title, merge_to=None):
    """Yellow section header spanning columns."""
    ws.row_dimensions[row].height = 18
    if merge_to:
        ws.merge_cells(start_row=row, start_column=cols, end_row=row, end_column=merge_to)
    c = ws.cell(row=row, column=cols, value=title)
    c.font = _f(bold=True, size=10, colour="000000")
    c.fill = _fill(_HDR_YELLOW)
    c.border = _b()
    c.alignment = _al("center")

def _col_hdrs(ws, row, headers, bg=_COL_YELLOW, height=30):
    ws.row_dimensions[row].height = height
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _f(bold=True, size=9)
        c.fill = _fill(bg)
        c.border = _b()
        c.alignment = _al("center")

def _w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Data helpers ───────────────────────────────────────────────────────────────
def _part_ests(summary): return (summary.get("estimate_summary") or {}).get("part_estimates") or []

def _meta(summary):
    da = summary.get("document_analysis") or {}
    pf = da.get("primary_fields") or {}
    tb = (da.get("title_block") or {}).get("normalized") or {}
    src = summary.get("source_file","").split("\\")[-1].replace(".PDF","").replace(".pdf","")
    wi = (summary.get("estimate_summary") or {}).get("estimate_workbook_inputs") or {}
    return {
        "job_number":  pf.get("job_number") or tb.get("drawing_number") or src,
        "description": pf.get("description") or tb.get("title") or "",
        "customer":    pf.get("customer") or tb.get("customer") or "",
        "quantity":    summary.get("quantity") or summary.get("assumed_job_quantity") or 180,
        "revision":    tb.get("revision") or "",
        "date":        datetime.now().strftime("%d/%m/%Y"),
        "wire_tonne":  _safe(wi.get("wire_cost_per_tonne_gbp") or 1500),
        "steel_tonne": _safe(wi.get("sheet_steel_cost_per_tonne_gbp") or 800),
        "scrap_pct":   _safe(wi.get("scrap_pct") or 4),
    }

def _op_name(op):
    m = {"laser_cutting":"Laser (Metal)","folding":"Fold","welding":"Weld (CO2)",
         "spot_welding":"Spotweld","resistance_welding":"Spotweld",
         "dress_welds":"Dress Welds","powder_coating":"Powder Coat",
         "diamond_polish":"Diamond Polish","cnc_routing":"CNC Route",
         "cnc":"CNC Route","wet_spray":"Wet Spray","assembly":"Assemble",
         "handling":"Handle","packing":"Pack","bench_work":"Bench",
         "hole_machining":"Drill","tapping":"Tap","guillotine":"Guillotine",
         "wire_forming":"Wire Forming","deburring":"Deburr","linisher":"Linish"}
    return m.get(op.lower(), op.replace("_"," ").title())

def _dept(op):
    m = {"laser_cutting":"LASM","laser":"LASM","folding":"FOLD","fold":"FOLD",
         "welding":"WELD","weld":"WELD","spot_welding":"SPOT","resistance_welding":"SPOT",
         "dress_welds":"DRES","powder_coating":"P/C","diamond_polish":"DPOL",
         "cnc_routing":"CNC","cnc":"CNC","wet_spray":"WSPR","assembly":"PACM",
         "handling":"PACM","packing":"PACM","bench_work":"BENC",
         "hole_machining":"DRIL","tapping":"DRIL","guillotine":"GUIL",
         "wire_forming":"WIRE","deburring":"DRES","linisher":"LINS"}
    return m.get(op.lower(), op.upper()[:4])

def _is_sheet_metal(mat):
    return str(mat or "").upper() in {"MILD_STEEL","MILD STEEL","STAINLESS_STEEL",
        "STAINLESS STEEL","ALUMINIUM","ALUMINUM","ZINTEC","BRIGHT_DRAWN"}

def _is_board(mat):
    return str(mat or "").upper() in {"MDF","VENEERED_MDF","OAK_VENEER_MDF",
        "PLYWOOD","BIRCH_PLYWOOD","TIMBER","ACRYLIC","POLYCARBONATE","HDPE_PLASTIC"}


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1: ESTIMATE  (matches SDI manual format)
# ══════════════════════════════════════════════════════════════════════════════
def _write_estimate(ws, summary):
    meta  = _meta(summary)
    pes   = _part_ests(summary)
    es    = summary.get("estimate_summary") or {}
    cb    = es.get("cost_breakdown") or {}
    # Sum from individual part estimates — more reliable than aggregate totals
    _mat_parts = sum(
        _safe((pe.get("material_estimate") or {}).get("cost_per_part_gbp", 0)) *
        _safe(pe.get("quantity") or 1)
        for pe in pes
    )
    _lab_parts = sum(
        _safe((pe.get("labour_estimate") or {}).get("total_labour_cost_gbp", 0))
        for pe in pes
    )
    _mat_cb = _safe((cb.get("material") or {}).get("total") or 0)
    _lab_cb = _safe((cb.get("labour") or {}).get("total") or 0)
    unit_cost = (_mat_parts or _mat_cb) + (_lab_parts or _lab_cb)

    # ── Row 1: SDI logo header ─────────────────────────────────────────────
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = "SDI Displays Limited — AI Estimate"
    c.font = _f(bold=True, size=12, colour=_WHITE)
    c.fill = _fill(_SDI_BLUE)
    c.alignment = _al("center")
    ws.row_dimensions[1].height = 22

    # Wire/steel cost boxes (top right, cols N-P)
    for r, lbl, val in [
        (1, "Wire Cost Per Tonne =",        f"£  {meta['wire_tonne']:,.2f}"),
        (2, "Sheet Steel Cost Per Tonne =",  f"£  {meta['steel_tonne']:,.2f}"),
    ]:
        _set(ws, r, 14, lbl,  bold=True,  fill=_BLUE_LIGHT, size=9)
        _set(ws, r, 15, val,  bold=True,  fill=_BLUE_LIGHT, size=9, align="right")

    # ── Rows 2-6: Job header block ─────────────────────────────────────────
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 16

    hdr_pairs = [
        (2, "Customer",    meta["customer"]),
        (3, "Description", meta["description"]),
        (4, "Drawing No.", meta["job_number"]),
        (5, "Quantity",    meta["quantity"]),
        (6, "Date",        meta["date"]),
    ]
    for row, lbl, val in hdr_pairs:
        _set(ws, row, 1, lbl, bold=True, fill=_BLUE_LIGHT, size=9)
        _set(ws, row, 2, val, fill=_WHITE, size=9)

    rev_pairs = [
        (4, "Rev",         meta["revision"]),
        (5, "Unit Cost",   round(unit_cost, 2)),
        (6, "Prepared By", "SDIAIVision"),
    ]
    for row, lbl, val in rev_pairs:
        _set(ws, row, 3, lbl, bold=True, fill=_BLUE_LIGHT, size=9)
        c = ws.cell(row=row, column=4, value=val)
        c.font = _f(bold=True, size=9)
        c.fill = _fill(_WHITE)
        c.border = _b()
        if isinstance(val, (int, float)) and lbl == "Unit Cost":
            c.number_format = '£#,##0.00'

    # ── SECTION 1: Standard Materials (BOM) ───────────────────────────────
    ROW = 8
    _section_hdr(ws, ROW, 1, "Standard Materials", merge_to=13)
    ROW += 1
    _col_hdrs(ws, ROW, [
        "Bill of Materials (Per Unit)", "", "", "",
        "Part code", "Supplier", "Price", "Qty Per Unit", "Scrap %", "Total Value"
    ])
    ws.merge_cells(f"A{ROW}:D{ROW}")  # merge description cols
    ROW += 1

    bom_start = ROW
    # Bought-in parts (BOUGHT_IN, customer-supplied = £0)
    bom_items = []
    for pe in pes:
        mat = str(pe.get("normalized_material") or "").upper()
        basis = str(pe.get("costing_basis") or "")
        if mat == "BOUGHT_IN" or "customer_supplied" in basis:
            bom_items.append({
                "desc": pe.get("description") or pe.get("part_number") or "",
                "part_code": pe.get("part_number") or "",
                "supplier": "Customer supply",
                "price": 0.0, "qty": _safe(pe.get("quantity") or 1), "scrap": 0, "total": 0.0
            })

    # Powder coat material
    for pe in pes:
        le = pe.get("labour_estimate") or {}
        me = pe.get("material_estimate") or {}
        pc = me.get("powder_consumable") or {}
        if pc and _safe(pc.get("material_cost_gbp")):
            bom_items.append({
                "desc": f"Powder Coat — {pe.get('description','')}", "part_code": "POWDER",
                "supplier": "Valspar", "price": _safe(pc.get("price_per_kg_gbp")),
                "qty": round(_safe(pc.get("kg_per_unit")),4), "scrap": 4,
                "total": _safe(pc.get("material_cost_gbp"))
            })

    # Standard consumables (packaging) — show once at job level
    qty = _safe(meta["quantity"])
    bom_items += [
        {"desc":"Poly Bag", "part_code":"POLY_BAG", "supplier":"SDI Packaging",
         "price":0.25, "qty":1, "scrap":0, "total":0.25},
        {"desc":"Carton", "part_code":"BOX_MEDIUM", "supplier":"Harleys",
         "price":1.85, "qty":1, "scrap":4, "total":1.92},
        {"desc":"Pallet (amortised)", "part_code":"PALLET", "supplier":"A Dale",
         "price":round(6.50/max(qty,1),4), "qty":1, "scrap":0,
         "total":round(6.50/max(qty,1),4)},
        {"desc":"Fixings Allowance", "part_code":"FIXINGS", "supplier":"SDI",
         "price":0.35, "qty":1, "scrap":0, "total":0.35},
        {"desc":"Delivery / Installation", "part_code":"", "supplier":"",
         "price":0.0, "qty":0, "scrap":0, "total":0.0, "note":"Quoted separate"},
    ]

    bom_total = 0.0
    for i, item in enumerate(bom_items):
        bg = _WHITE
        for col, val, fmt in [
            (1,item["desc"],    None), (2,"",None), (3,"",None), (4,"",None),
            (5,item["part_code"],None),(6,item["supplier"],None),
        ]:
            _set(ws, ROW, col, val, fill=bg, size=9)
        ws.merge_cells(f"A{ROW}:D{ROW}")
        _money(ws, ROW, 7,  item["price"], fill=bg)
        _set(ws, ROW, 8,  item.get("qty") or None, fill=bg, size=9, align="right")
        _set(ws, ROW, 9,  f'{item["scrap"]}%' if item["scrap"] else "", fill=bg, size=9, align="center")
        note = item.get("note","")
        if note:
            _set(ws, ROW, 10, note, fill=bg, size=9, italic=True, colour="595959")
        else:
            _money(ws, ROW, 10, item["total"], fill=bg)
        bom_total += item["total"]
        ROW += 1

    ROW += 1  # spacer

    # ── SECTION 2: Sheet Steel ────────────────────────────────────────────
    _section_hdr(ws, ROW, 1, "Sheet Steel", merge_to=13)
    ROW += 1
    _col_hdrs(ws, ROW, [
        "Part Description", "Qty\nPer Unit", "Part\nLength", "Part\nWidth",
        "Gauge", "Sheet\nLength", "Sheet\nWidth", "Qty Per\nSheet", "Scrap\n%",
        "Cost Per\nPart"
    ])
    ROW += 1

    steel_total = 0.0
    for pe in pes:
        mat = str(pe.get("normalized_material") or "").upper()
        if not _is_sheet_metal(mat):
            continue
        me = pe.get("material_estimate") or {}
        se = me.get("stock_estimate") or {}
        ng = pe.get("normalized_geometry") or {}
        blank_l  = _safe(me.get("blank_length_mm") or ng.get("blank_length_mm"))
        blank_w  = _safe(me.get("blank_width_mm")  or ng.get("blank_width_mm"))
        thick    = _safe(pe.get("normalized_thickness_mm") or me.get("thickness_mm"))
        sh_sizes = se.get("candidate_sheet_size_mm") or [2500, 1250]
        sh_l     = sh_sizes[0] if len(sh_sizes) > 0 else 2500
        sh_w     = sh_sizes[1] if len(sh_sizes) > 1 else 1250
        pps      = se.get("parts_per_sheet") or ""
        cost_pp  = _safe(me.get("cost_per_part_gbp"))
        qty_u    = _safe(pe.get("quantity") or 1)
        scrap    = meta["scrap_pct"]
        bg = _LIGHT_YELL

        desc = f"{pe.get('part_number','')}  {pe.get('description','')}"
        _set(ws, ROW, 1,  desc,   fill=bg, size=9, bold=True)
        _set(ws, ROW, 2,  int(qty_u),  fill=bg, size=9, align="center")
        _set(ws, ROW, 3,  blank_l or None, fill=bg, size=9, align="right", fmt='#,##0')
        _set(ws, ROW, 4,  blank_w or None, fill=bg, size=9, align="right", fmt='#,##0')
        _set(ws, ROW, 5,  thick   or None, fill=bg, size=9, align="center")
        _set(ws, ROW, 6,  sh_l,   fill=bg, size=9, align="right", fmt='#,##0')
        _set(ws, ROW, 7,  sh_w,   fill=bg, size=9, align="right", fmt='#,##0')
        _set(ws, ROW, 8,  pps or None, fill=bg, size=9, align="center")
        _set(ws, ROW, 9,  f'{scrap}%', fill=bg, size=9, align="center")
        _money(ws, ROW, 10, cost_pp, fill=bg)
        steel_total += cost_pp * qty_u
        ROW += 1

    ROW += 1

    # ── SECTION 3: Other Sheet Material (MDF, acrylic, board) ─────────────
    board_parts = [pe for pe in pes if _is_board(pe.get("normalized_material"))]
    if board_parts:
        _section_hdr(ws, ROW, 1, "Other Sheet Material", merge_to=13)
        ROW += 1
        _col_hdrs(ws, ROW, [
            "Part Description", "Qty\nPer Unit", "Part\nLength", "Part\nWidth",
            "Thickness", "Sheet\nLength", "Sheet\nWidth", "Qty Per\nSheet", "Scrap\n%",
            "Cost Per\nPart"
        ])
        ROW += 1
        for pe in board_parts:
            me = pe.get("material_estimate") or {}
            se = me.get("stock_estimate") or {}
            ng = pe.get("normalized_geometry") or {}
            blank_l  = _safe(me.get("blank_length_mm") or ng.get("blank_length_mm"))
            blank_w  = _safe(me.get("blank_width_mm")  or ng.get("blank_width_mm"))
            thick    = _safe(pe.get("normalized_thickness_mm") or me.get("thickness_mm"))
            sh_sizes = se.get("candidate_sheet_size_mm") or [2440, 1220]
            sh_l = sh_sizes[0] if len(sh_sizes) > 0 else 2440
            sh_w = sh_sizes[1] if len(sh_sizes) > 1 else 1220
            pps  = se.get("parts_per_sheet") or ""
            cost_pp = _safe(me.get("cost_per_part_gbp"))
            qty_u   = _safe(pe.get("quantity") or 1)
            bg = _WHITE
            desc = f"{pe.get('part_number','')}  {pe.get('description','')}"
            _set(ws, ROW, 1, desc, fill=bg, size=9, bold=True)
            _set(ws, ROW, 2, int(qty_u), fill=bg, size=9, align="center")
            _set(ws, ROW, 3, blank_l or None, fill=bg, size=9, align="right", fmt='#,##0')
            _set(ws, ROW, 4, blank_w or None, fill=bg, size=9, align="right", fmt='#,##0')
            _set(ws, ROW, 5, thick or None,   fill=bg, size=9, align="center")
            _set(ws, ROW, 6, sh_l, fill=bg, size=9, align="right", fmt='#,##0')
            _set(ws, ROW, 7, sh_w, fill=bg, size=9, align="right", fmt='#,##0')
            _set(ws, ROW, 8, pps or None, fill=bg, size=9, align="center")
            _set(ws, ROW, 9, f'{meta["scrap_pct"]}%', fill=bg, size=9, align="center")
            _money(ws, ROW, 10, cost_pp, fill=bg)
            ROW += 1
        ROW += 1

    # ── Total Material Cost ───────────────────────────────────────────────
    all_mat = _safe((cb.get("material") or {}).get("total"))
    # Recalculate from individual parts (cost_breakdown total often 0)
    all_mat_recalc = sum(
        _safe((pe.get("material_estimate") or {}).get("cost_per_part_gbp", 0)) *
        _safe(pe.get("quantity") or 1)
        for pe in pes
    )
    all_mat = all_mat_recalc if all_mat_recalc > 0 else all_mat
    _set(ws, ROW, 9, "Total Material Cost", bold=True, fill=_GREY, size=10)
    c = ws.cell(row=ROW, column=10, value=round(all_mat, 2))
    c.font = _f(bold=True, size=10)
    c.fill = _fill(_GREY)
    c.border = _b()
    c.number_format = '£#,##0.00'
    ROW += 2

    # ── SECTION 4: Labour ─────────────────────────────────────────────────
    _section_hdr(ws, ROW, 1, "Labour", merge_to=13)
    ROW += 1
    _col_hdrs(ws, ROW, [
        "Operation", "Part Description", "Dept.", "Qty\nPer Unit",
        "Rate Per\nHour", "Total\nHours", "Labour\nCost",
        "Set Up\n(Mins)", "Total\nValue", ""
    ])
    ROW += 1

    # SRS job number column (col 1 first row per job, like manual)
    srs_written = False
    labour_grand = 0.0
    for pe in pes:
        le = pe.get("labour_estimate") or {}
        costs  = le.get("costs_gbp") or {}
        times  = le.get("times_min") or {}
        setups = le.get("setup_times_min") or {}
        rates  = le.get("hourly_rates_gbp") or {}
        ops    = (pe.get("process_estimate") or {}).get("operations") or list(costs.keys())
        if not ops:
            continue

        desc = pe.get("description") or pe.get("part_number") or ""
        qty  = _safe(pe.get("quantity") or 1)

        for i, op in enumerate(ops):
            cost     = _safe(costs.get(op))
            t_min    = _safe(times.get(op))
            s_min    = _safe(setups.get(op))
            rate     = _safe(rates.get(op))
            hours    = t_min / 60.0
            total_v  = cost + (rate * s_min / 60.0 / max(qty,1))
            bg = _LIGHT_YELL if i % 2 == 0 else _WHITE

            if not srs_written:
                _set(ws, ROW, 1, meta["job_number"], fill=bg, size=8, bold=True, colour="595959")
                srs_written = True
            else:
                _set(ws, ROW, 1, _op_name(op), fill=bg, size=9)

            _set(ws, ROW, 1, _op_name(op),  fill=bg, size=9)
            _set(ws, ROW, 2, desc,           fill=bg, size=9)
            _set(ws, ROW, 3, _dept(op),      fill=bg, size=9, align="center", bold=True)
            _set(ws, ROW, 4, int(qty),        fill=bg, size=9, align="center")
            _money(ws, ROW, 5, rate,          fill=bg)
            c6 = ws.cell(row=ROW, column=6, value=round(hours,4) if hours else None)
            c6.font = _f(size=9); c6.fill = _fill(bg); c6.border = _b()
            c6.number_format = '0.0000'; c6.alignment = _al("right")
            _money(ws, ROW, 7, cost,          fill=bg)
            _set(ws, ROW, 8, round(s_min,1) if s_min else None, fill=bg, size=9, align="center")
            _money(ws, ROW, 9, total_v,       fill=bg)

            labour_grand += cost
            ROW += 1

    # Labour total
    ROW += 1
    _set(ws, ROW, 8, "Total Labour Cost", bold=True, fill=_GREY, size=10)
    c = ws.cell(row=ROW, column=9, value=round(labour_grand,2))
    c.font = _f(bold=True, size=10); c.fill = _fill(_GREY)
    c.border = _b(); c.number_format = '£#,##0.00'

    ROW += 2
    # Unit cost summary
    # Recalc labour total from parts if needed
    if labour_grand == 0:
        labour_grand = sum(
            _safe((pe.get("labour_estimate") or {}).get("total_labour_cost_gbp", 0))
            for pe in pes
        )
    total = all_mat + labour_grand
    for lbl, val in [("Total Material Cost", all_mat),
                     ("Total Labour Cost",   labour_grand),
                     ("UNIT COST",           total)]:
        bold = (lbl == "UNIT COST")
        fill = _HDR_YELLOW if bold else _GREY
        _set(ws, ROW, 8, lbl, bold=bold, fill=fill, size=10)
        c = ws.cell(row=ROW, column=9, value=round(val,2))
        c.font = _f(bold=bold, size=10); c.fill = _fill(fill)
        c.border = _b(); c.number_format = '£#,##0.00'
        ROW += 1

    # Rebate + sell price
    rebate = round(total * 0.066, 2)
    sell   = round(total - rebate, 2)
    ROW += 1
    for lbl, val in [("Rebate (6.6%)", -rebate), ("Net Sell Price (unit)", sell)]:
        _set(ws, ROW, 8, lbl, bold=(lbl!="Rebate (6.6%)"), fill=_GREY, size=10)
        c = ws.cell(row=ROW, column=9, value=val)
        c.font = _f(bold=(lbl!="Rebate (6.6%)"), size=10)
        c.fill = _fill(_GREEN if "Sell" in lbl else _GREY)
        c.border = _b(); c.number_format = '£#,##0.00'
        ROW += 1

    # Column widths (A-J)
    _w(ws, [34, 26, 6, 6, 8, 10, 10, 10, 4, 12])
    ws.freeze_panes = "A8"
    ws.sheet_view.zoomScale = 90


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2: LABOUR  (dept hours summary)
# ══════════════════════════════════════════════════════════════════════════════
def _write_labour(ws, summary):
    meta = _meta(summary)
    pes  = _part_ests(summary)

    ws.merge_cells("A1:F1")
    c = ws["A1"]; c.value = f"Labour Summary — {meta['job_number']}"
    c.font = _f(bold=True, size=12, colour=_WHITE)
    c.fill = _fill(_SDI_BLUE); c.alignment = _al("center")
    ws.row_dimensions[1].height = 22

    _col_hdrs(ws, 3, ["Department", "Hours", "Labour Cost (£)", "Setup (mins)", "Total Cost (£)", "Notes"])

    depts: Dict[str,Dict] = {}
    for pe in pes:
        le    = pe.get("labour_estimate") or {}
        costs = le.get("costs_gbp") or {}
        times = le.get("times_min") or {}
        setups= le.get("setup_times_min") or {}
        rates = le.get("hourly_rates_gbp") or {}
        for op, cost in costs.items():
            d = _dept(op)
            if d not in depts:
                depts[d] = {"name":_op_name(op),"hours":0.0,"cost":0.0,"setup":0.0,"rate":0.0}
            depts[d]["hours"]  += _safe(times.get(op)) / 60.0
            depts[d]["cost"]   += _safe(cost)
            depts[d]["setup"]  += _safe(setups.get(op))
            depts[d]["rate"]    = _safe(rates.get(op))

    row = 4
    total_cost = 0.0
    for i, (dept, v) in enumerate(sorted(depts.items())):
        bg = _LIGHT_YELL if i % 2 == 0 else _WHITE
        _set(ws, row, 1, dept, bold=True, fill=bg, size=10)
        c2 = ws.cell(row=row, column=2, value=round(v["hours"],4))
        c2.font = _f(size=10); c2.fill = _fill(bg); c2.border = _b()
        c2.number_format = '0.0000'; c2.alignment = _al("right")
        _money(ws, row, 3, v["cost"], fill=bg)
        _set(ws, row, 4, round(v["setup"],1) if v["setup"] else None, fill=bg, size=10, align="right")
        _money(ws, row, 5, v["cost"], fill=bg)
        _set(ws, row, 6, v["name"], fill=bg, size=9, colour="595959", italic=True)
        total_cost += v["cost"]
        row += 1

    _set(ws, row, 1, "TOTAL", bold=True, fill=_HDR_YELLOW, size=10)
    _money(ws, row, 3, total_cost, fill=_HDR_YELLOW)
    _money(ws, row, 5, total_cost, fill=_HDR_YELLOW)
    _w(ws, [14, 12, 18, 14, 18, 22])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3: MATERIAL PRICE BREAK
# ══════════════════════════════════════════════════════════════════════════════
def _write_price_break(ws, summary):
    meta  = _meta(summary)
    pes   = _part_ests(summary)
    es    = summary.get("estimate_summary") or {}
    pes   = _part_ests(summary)
    cb    = es.get("cost_breakdown") or {}
    _mt = sum(_safe((pe.get("material_estimate") or {}).get("cost_per_part_gbp",0))*_safe(pe.get("quantity") or 1) for pe in pes)
    _lt = sum(_safe((pe.get("labour_estimate") or {}).get("total_labour_cost_gbp",0)) for pe in pes)
    _mcb = _safe((cb.get("material") or {}).get("total") or 0)
    _lcb = _safe((cb.get("labour") or {}).get("total") or 0)
    unit  = (_mt or _mcb) + (_lt or _lcb)

    ws.merge_cells("A1:N1")
    c = ws["A1"]; c.value = f"Material Price Break — {meta['job_number']}"
    c.font = _f(bold=True, size=12, colour=_WHITE)
    c.fill = _fill(_SDI_BLUE); c.alignment = _al("center")
    ws.row_dimensions[1].height = 22

    breaks = [1, 10, 25, 50, 100, 250, 500, 600, 700, 800, 900]
    mults  = [1.0,0.97,0.96,0.94,0.92,0.91,0.90,0.895,0.89,0.885,0.88]
    rebate = 0.066

    _col_hdrs(ws, 3, [""] + [f"Qty {q}" for q in breaks])
    _col_hdrs(ws, 4, ["Unit Cost"] + [f"£{round(unit*m,2):.2f}" for m in mults], bg=_LIGHT_YELL)

    row = 5
    _set(ws, row, 1, "Less Rebate (6.6%)", bold=True, fill=_GREY, size=9)
    for i, m in enumerate(mults):
        v = round(unit * m * rebate, 2)
        _money(ws, row, 2+i, -v, fill=_GREY)
    row += 1
    _set(ws, row, 1, "NET SELL PRICE", bold=True, fill=_HDR_YELLOW, size=10)
    for i, m in enumerate(mults):
        v = round(unit * m * (1 - rebate), 2)
        c = ws.cell(row=row, column=2+i, value=v)
        c.font = _f(bold=True, size=10); c.fill = _fill(_HDR_YELLOW)
        c.border = _b(); c.number_format = '£#,##0.00'

    _w(ws, [18] + [12]*11)


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ══════════════════════════════════════════════════════════════════════════════
def write_estimate_xlsx(summary: Dict[str,Any], out_dir=None) -> Path:
    out_dir = Path(out_dir or "output/estimates")
    out_dir.mkdir(parents=True, exist_ok=True)

    meta = _meta(summary)
    safe = re.sub(r'[^\w\-]', '_', str(meta["job_number"]))[:60]
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = out_dir / f"{safe}_{ts}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    ws_est = wb.create_sheet("Estimate")
    ws_lab = wb.create_sheet("Labour")
    ws_brk = wb.create_sheet("Material Price Break")

    _write_estimate(ws_est, summary)
    _write_labour(ws_lab, summary)
    _write_price_break(ws_brk, summary)

    # Estimate tab active on open
    ws_est.sheet_view.tabSelected = True
    wb.active = ws_est

    wb.save(path)
    print(f"   -> Estimate xlsx written: {path}")
    return path


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", required=True)
    ap.add_argument("--out", default="output/estimates")
    args = ap.parse_args()
    with open(args.json, encoding="utf-8") as f:
        data = json.load(f)
    write_estimate_xlsx(data, out_dir=args.out)
