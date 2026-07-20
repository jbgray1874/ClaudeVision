from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

import config
from estimate_sheet_discovery import discover_output_cells_map

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover
    load_workbook = None

try:
    import xlrd  # type: ignore
except ImportError:  # pragma: no cover
    xlrd = None


class _XlrdSheetWrapper:
    """Adapter so .xls sheets expose the same cell access API as openpyxl."""

    def __init__(self, xlrd_sheet: Any) -> None:
        self._s = xlrd_sheet

    def cell(self, row: int, column: int) -> Any:
        class _Cell:
            def __init__(self, value: Any) -> None:
                self.value = value

        try:
            return _Cell(self._s.cell_value(row - 1, column - 1))
        except IndexError:
            return _Cell(None)

    def __getitem__(self, cell_ref: str) -> Any:
        import re as _re
        m = _re.match(r"([A-Z]+)(\d+)", cell_ref.upper())
        if not m:
            raise KeyError(cell_ref)
        col = 0
        for ch in m.group(1):
            col = col * 26 + (ord(ch) - ord("A") + 1)
        col -= 1
        row = int(m.group(2)) - 1
        class _Cell:
            def __init__(self, value: Any) -> None:
                self.value = value
        try:
            return _Cell(self._s.cell_value(row, col))
        except IndexError:
            return _Cell(None)


class _XlrdWorkbookWrapper:
    """Adapter so .xls workbooks expose .sheetnames and __getitem__ like openpyxl."""

    def __init__(self, xlrd_book: Any) -> None:
        self._book = xlrd_book
        self.sheetnames: List[str] = xlrd_book.sheet_names()

    def __getitem__(self, name: str) -> _XlrdSheetWrapper:
        return _XlrdSheetWrapper(self._book.sheet_by_name(name))

    def close(self) -> None:
        self._book.release_resources()


def _safe_float(value: Any) -> Optional[float]:
    try:
        if value is None:
            return None
        if isinstance(value, str):
            v = value.strip().replace("£", "").replace(",", "")
            if not v:
                return None
            return float(v)
        return float(value)
    except (TypeError, ValueError):
        return None


def _extract_by_path(payload: Dict[str, Any], dotted_path: str) -> Any:
    cur: Any = payload
    for key in dotted_path.split("."):
        if not isinstance(cur, dict):
            return None
        cur = cur.get(key)
    return cur


def _percent_diff(expected: Optional[float], actual: Optional[float]) -> Optional[float]:
    if expected is None or actual is None:
        return None
    if abs(expected) < 1e-9:
        return None
    return round(abs(actual - expected) / abs(expected) * 100.0, 4)


def _status_from_pct(pct: Optional[float], match: float, warn: float) -> str:
    if pct is None:
        return "review"
    if pct <= match:
        return "match"
    if pct <= warn:
        return "warning"
    return "fail"


def build_report_rows(summary: Dict[str, Any], workbook_path: Path, sheet_name: str = "Estimate") -> List[Dict[str, Any]]:
    suffix = workbook_path.suffix.lower()
    if suffix == ".xls":
        if xlrd is None:
            raise RuntimeError("xlrd is required for .xls files (pip install xlrd).")
        wb = _XlrdWorkbookWrapper(xlrd.open_workbook(str(workbook_path)))
    elif suffix in {".xlsx", ".xlsm"}:
        if load_workbook is None:
            raise RuntimeError("openpyxl is required for workbook parity report")
        wb = load_workbook(workbook_path, data_only=True, rich_text=False)
    else:
        raise RuntimeError(f"Unsupported workbook format '{suffix}' — need .xls, .xlsx, or .xlsm.")
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' not found in workbook")
    ws = wb[sheet_name]

    mapping, _discovery_meta = discover_output_cells_map(ws)
    if not mapping:
        mapping = (config.ESTIMATE_TEMPLATE_WRITEBACK or {}).get("output_cells") or {}

    thresholds = (config.WORKBOOK_EQUIVALENT_PRICING or {}).get("variance_thresholds_pct") or {}
    match_pct = float(thresholds.get("match", 3.0))
    warn_pct = float(thresholds.get("warning", 10.0))

    rows: List[Dict[str, Any]] = []
    for cell_ref, path in mapping.items():
        expected_raw = _extract_by_path(summary, str(path))
        actual_raw = ws[cell_ref].value
        expected = _safe_float(expected_raw)
        actual = _safe_float(actual_raw)
        pct = _percent_diff(expected, actual)
        status = _status_from_pct(pct, match_pct, warn_pct)
        rows.append(
            {
                "cell": cell_ref,
                "summary_path": path,
                "expected_value": expected,
                "actual_value": actual,
                "actual_cell_raw": actual_raw,
                "pct_variance": pct,
                "status": status,
            }
        )
    return rows


def write_reports(rows: List[Dict[str, Any]], csv_path: Path, json_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "cell",
        "summary_path",
        "expected_value",
        "actual_value",
        "actual_cell_raw",
        "pct_variance",
        "status",
    ]
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    json_path.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate cell-level parity report between estimate_summary and workbook cells.")
    parser.add_argument("--summary-json", required=True, help="Path to scan summary JSON containing estimate_summary")
    parser.add_argument("--workbook", required=True, help="Workbook to validate")
    parser.add_argument("--sheet", default="Estimate", help="Sheet name to validate")
    parser.add_argument("--out-csv", default=str(config.CSV_DIR / "estimate_workbook_parity_report.csv"), help="CSV report path")
    parser.add_argument("--out-json", default=str(config.CSV_DIR / "estimate_workbook_parity_report.json"), help="JSON report path")
    args = parser.parse_args()

    summary_path = Path(args.summary_json).resolve()
    workbook_path = Path(args.workbook).resolve()
    summary = json.loads(summary_path.read_text(encoding="utf-8"))

    rows = build_report_rows(summary, workbook_path, sheet_name=args.sheet)
    write_reports(rows, Path(args.out_csv).resolve(), Path(args.out_json).resolve())

    total = len(rows)
    fail = sum(1 for r in rows if r["status"] == "fail")
    warn = sum(1 for r in rows if r["status"] == "warning")
    match = sum(1 for r in rows if r["status"] == "match")
    print(f"Parity rows={total} match={match} warning={warn} fail={fail}")
    print(f"CSV: {Path(args.out_csv).resolve()}")
    print(f"JSON: {Path(args.out_json).resolve()}")


if __name__ == "__main__":
    main()
