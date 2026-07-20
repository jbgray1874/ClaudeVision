"""
SDI Intelligence — Tim standard-materials (bought-in BOM) ingester.

Learns the per-bay BOUGHT-IN bill of materials from an SDI manual estimate
workbook — electrics loom, slotted tube, fixings, vinyl, header, packaging —
and stores it per drawing so the engine can include it in the material total.
These are catalogue lookups, not estimates; the engine's job is the fabricated
sheet + labour, this fills in the bought-in lines Tim sources from the BOM.

POWDER is EXCLUDED on purpose: the engine computes powder consumable itself
(_powder_consumable_estimate), so ingesting it would double-count.

Generalises: the "Bill of Materials (Per Unit)" block has a stable layout
(C=description, H=part code, I=supplier, J=price, K=qty/unit, L=scrap, M=total).

Outputs:
  job_bought_in_materials.json   {by_drawing: {DRAWING: {lines:[...], total_gbp}}}
  job_bought_in_materials.sql    MERGE into AIEstimating.JobBoughtInMaterials
"""
from __future__ import annotations
import argparse, json, datetime, re
from pathlib import Path
import openpyxl

POWDER_CODE_RE = re.compile(r"POWDER", re.I)

def parse(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Estimate"] if "Estimate" in wb.sheetnames else wb.worksheets[0]
    drawing = None
    for r in range(1, 20):
        if str(ws.cell(r, 3).value or "").strip().lower().startswith("drawing"):
            drawing = str(ws.cell(r, 4).value or "").strip()
            break
    start = end = None
    for r in range(1, ws.max_row + 1):
        c = str(ws.cell(r, 3).value or "")
        if "Bill of Materials" in c and start is None:
            start = r + 1
        elif start and c.strip() in ("Wire", "Sheet Steel"):
            end = r; break
    lines, total = [], 0.0
    for r in range(start, end):
        desc = ws.cell(r, 3).value
        code = ws.cell(r, 8).value
        supp = ws.cell(r, 9).value
        price = ws.cell(r, 10).value
        qty = ws.cell(r, 11).value
        scrap = ws.cell(r, 12).value
        tot = ws.cell(r, 13).value
        if not desc or not isinstance(tot, (int, float)) or tot <= 0:
            continue
        if POWDER_CODE_RE.search(str(code or "")) or POWDER_CODE_RE.search(str(desc or "")):
            continue  # engine computes powder consumable
        lines.append({
            "description": str(desc).strip(),
            "part_code": str(code).strip() if code else None,
            "supplier": str(supp).strip() if supp else None,
            "unit_price_gbp": round(float(price), 4) if isinstance(price, (int, float)) else None,
            "qty_per_unit": float(qty) if isinstance(qty, (int, float)) else None,
            "scrap_pct": float(scrap) if isinstance(scrap, (int, float)) else None,
            "total_gbp": round(float(tot), 4),
        })
        total += float(tot)
    return drawing, lines, round(total, 4)

def emit_sql(drawing, lines, source):
    out = [
        "IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='AIEstimating') EXEC('CREATE SCHEMA AIEstimating');",
        "IF OBJECT_ID('AIEstimating.JobBoughtInMaterials','U') IS NULL",
        "CREATE TABLE AIEstimating.JobBoughtInMaterials (",
        "    id int IDENTITY PRIMARY KEY, drawing_number varchar(40) NOT NULL,",
        "    part_code varchar(60) NULL, description varchar(200) NULL, supplier varchar(80) NULL,",
        "    unit_price_gbp decimal(12,4) NULL, qty_per_unit decimal(12,4) NULL,",
        "    scrap_pct decimal(6,4) NULL, total_gbp decimal(12,4) NOT NULL,",
        "    source_workbook varchar(260) NULL, ingested_utc datetime2 NOT NULL DEFAULT SYSUTCDATETIME());",
        f"DELETE FROM AIEstimating.JobBoughtInMaterials WHERE drawing_number='{drawing}';",
    ]
    for ln in lines:
        def q(v): return "NULL" if v is None else "'" + str(v).replace("'", "''") + "'"
        def n(v): return "NULL" if v is None else str(v)
        out.append(
            "INSERT INTO AIEstimating.JobBoughtInMaterials(drawing_number,part_code,description,supplier,unit_price_gbp,qty_per_unit,scrap_pct,total_gbp,source_workbook) "
            f"VALUES('{drawing}',{q(ln['part_code'])},{q(ln['description'])},{q(ln['supplier'])},{n(ln['unit_price_gbp'])},{n(ln['qty_per_unit'])},{n(ln['scrap_pct'])},{n(ln['total_gbp'])},{q(source)});"
        )
    return "\n".join(out) + "\n"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--out-dir", default=".")
    ap.add_argument("--write-db", action="store_true")
    a = ap.parse_args()
    drawing, lines, total = parse(a.workbook)
    src = Path(a.workbook).name
    outd = Path(a.out_dir)
    # merge into an existing by_drawing map if present, so multiple jobs accumulate
    jpath = outd / "job_bought_in_materials.json"
    payload = {"by_drawing": {}}
    if jpath.exists():
        try: payload = json.loads(jpath.read_text())
        except Exception: payload = {"by_drawing": {}}
    payload.setdefault("by_drawing", {})[drawing] = {
        "lines": lines, "total_gbp": total, "source": src,
        "ingested": datetime.datetime.now(datetime.timezone.utc).isoformat(),
    }
    jpath.write_text(json.dumps(payload, indent=2))
    (outd / "job_bought_in_materials.sql").write_text(emit_sql(drawing, lines, src))
    print(f"Drawing {drawing}: {len(lines)} bought-in line(s), total £{total:.2f} (powder excluded)")
    print(f"  -> {jpath}\n  -> {outd/'job_bought_in_materials.sql'}")
    if a.write_db:
        try:
            import pyodbc, config
            cn = config.get_connection(); cur = cn.cursor()
            cur.execute(open(outd / "job_bought_in_materials.sql").read()); cn.commit()
            print("  -> written to AIEstimating.JobBoughtInMaterials")
        except Exception as exc:
            print(f"  (DB write skipped: {exc})")

if __name__ == "__main__":
    main()
