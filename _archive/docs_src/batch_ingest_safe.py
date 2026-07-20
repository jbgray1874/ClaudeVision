"""
SDI Intelligence — Safe Historical Batch Ingest
=================================================
Replaces batch_ingest_historical.py for large file sets.
Adds a per-file timeout — any file taking longer than
TIMEOUT_SECONDS is automatically skipped and logged.

Usage:
    python src\batch_ingest_safe.py `
        --root "K:\Estimating\Completed\Manual Estimates" `
        --out "output\historical_estimates" `
        --timeout 45

Features:
    - 45 second timeout per file (configurable)
    - Skips already-processed files (resume after crash)
    - Logs all skipped/failed files to output\historical_estimates\SKIPPED.log
    - Shows progress + estimated time remaining
    - Windows compatible (uses threading not signal)
"""

import os
import sys
import json
import argparse
import threading
import traceback
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional

# ── Config ─────────────────────────────────────────────────────────────────────
TIMEOUT_SECONDS = 45
SUPPORTED_EXTS  = {'.xls', '.xlsx', '.xlsm'}
SKIP_PATTERNS   = ['.SKIP', '.skip', '~$', 'SKIP_']


def _sanitise_filename(name: str) -> str:
    """Convert xls filename to safe output JSON name."""
    import re
    name = re.sub(r'[^\w\s\-\.]', '_', name)
    name = re.sub(r'\s+', '_', name)
    return name.replace('.xlsx','').replace('.xlsm','').replace('.xls','')


def _already_done(xls_path: Path, out_dir: Path) -> bool:
    """Check if this file has already been processed."""
    stem = _sanitise_filename(xls_path.name)
    candidates = [
        out_dir / f"{stem}.formula_parse.json",
        out_dir / f"{stem}.json",
    ]
    return any(c.exists() for c in candidates)


def _parse_one_file(xls_path: Path, out_dir: Path) -> dict:
    """
    Parse a single workbook and write output JSON.
    Called inside a thread with timeout.
    Returns dict with status and output path.
    """
    try:
        # Try to import the existing parser
        sys.path.insert(0, str(Path(__file__).parent))
        try:
            from estimate_template_parser import write_estimate_template_parse
            stem = _sanitise_filename(xls_path.name)
            out_path = out_dir / f"{stem}.formula_parse.json"
            write_estimate_template_parse(str(xls_path), str(out_path))
            size = out_path.stat().st_size if out_path.exists() else 0
            return {"status": "ok", "path": str(out_path), "size": size}
        except ImportError:
            pass

        # Fallback: use openpyxl directly for basic extraction
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(xls_path),
                                         read_only=True,
                                         data_only=True)
            parts = []
            for sheet_name in wb.sheetnames[:5]:  # max 5 sheets
                ws = wb[sheet_name]
                rows_read = 0
                for row in ws.iter_rows(max_row=200, values_only=True):
                    if rows_read > 200: break
                    if any(row):
                        parts.append(list(str(c) if c is not None else "" for c in row))
                    rows_read += 1
                if rows_read >= 200:
                    break
            wb.close()
            stem = _sanitise_filename(xls_path.name)
            out_path = out_dir / f"{stem}.json"
            data = {
                "source_file": xls_path.name,
                "source_path": str(xls_path),
                "sheets":      sheet_name,
                "rows_sample": parts[:50],
                "parsed_at":   datetime.now().isoformat(),
            }
            out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                                encoding='utf-8')
            size = out_path.stat().st_size
            return {"status": "ok", "path": str(out_path), "size": size}
        except Exception as e:
            return {"status": "error", "error": str(e)}

    except Exception as e:
        return {"status": "error", "error": str(e)}


def parse_with_timeout(xls_path: Path, out_dir: Path,
                       timeout: int = TIMEOUT_SECONDS) -> dict:
    """Run _parse_one_file in a thread with timeout."""
    result = {"status": "timeout", "error": f"Exceeded {timeout}s"}
    lock   = threading.Event()

    def _worker():
        nonlocal result
        try:
            result = _parse_one_file(xls_path, out_dir)
        except Exception as e:
            result = {"status": "error", "error": str(e)}
        finally:
            lock.set()

    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    completed = lock.wait(timeout=timeout)

    if not completed:
        # Thread still running — mark as timeout, move on
        result = {"status": "timeout",
                  "error": f"File parsing exceeded {timeout} seconds — skipped"}
    return result


def find_all_workbooks(root: Path) -> list:
    import os
    clean = []
    print("  (walking directory tree...)", flush=True)
    for dirpath, dirnames, filenames in os.walk(str(root)):
        dirnames[:] = [d for d in dirnames if not d.startswith('~')]
        for filename in filenames:
            if filename.startswith('~'): continue
            if any(p in filename for p in SKIP_PATTERNS): continue
            ext = os.path.splitext(filename)[1].lower()
            if ext in SUPPORTED_EXTS:
                clean.append(Path(os.path.join(dirpath, filename)))
    return sorted(clean)


def run_ingest(root: str, out_dir: str, timeout: int = TIMEOUT_SECONDS) -> None:
    root_path = Path(root)
    out_path  = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    skip_log  = out_path / "SKIPPED.log"
    skip_file = open(str(skip_log), 'a', encoding='utf-8')

    print(f"\nSDI Intelligence — Safe Historical Batch Ingest")
    print(f"=" * 55)
    print(f"Root:     {root_path}")
    print(f"Output:   {out_path}")
    print(f"Timeout:  {timeout}s per file")
    print(f"Scanning for workbooks...")

    all_files = find_all_workbooks(root_path)
    total     = len(all_files)

    # Filter already done
    todo = [f for f in all_files if not _already_done(f, out_path)]
    done_already = total - len(todo)

    print(f"Found:    {total:,} workbooks")
    print(f"Already done: {done_already:,} — skipping")
    print(f"To process:   {len(todo):,}")
    print()

    inserted   = 0
    skipped    = 0
    errors     = 0
    start_time = datetime.now()

    for i, xls_path in enumerate(todo, 1):
        rel = str(xls_path).replace(str(root_path), '').lstrip('\\/')
        pct = (i / len(todo)) * 100

        # ETA
        elapsed = (datetime.now() - start_time).total_seconds()
        rate    = i / elapsed if elapsed > 0 else 0
        remaining = (len(todo) - i) / rate if rate > 0 else 0
        eta = str(timedelta(seconds=int(remaining))) if remaining > 0 else "?"

        print(f"  [{i:>6}/{len(todo):>6}] ({pct:.1f}%) ETA:{eta}  {rel[:70]}")

        result = parse_with_timeout(xls_path, out_path, timeout)

        if result["status"] == "ok":
            size_kb = result.get("size", 0) / 1024
            print(f"           -> {Path(result['path']).name}  ({size_kb:.0f} KB)")
            inserted += 1
        elif result["status"] == "timeout":
            print(f"           -> ⏭  SKIPPED (timeout {timeout}s) — too large/complex")
            skip_file.write(f"TIMEOUT  {rel}\n")
            skip_file.flush()
            skipped += 1
        else:
            err = str(result.get("error",""))[:100]
            print(f"           -> ⚠  ERROR: {err}")
            skip_file.write(f"ERROR    {rel}  |  {err}\n")
            skip_file.flush()
            errors += 1

        # Progress summary every 100 files
        if i % 100 == 0:
            elapsed_min = elapsed / 60
            print()
            print(f"  ── Progress: {inserted} ok | {skipped} timeout | "
                  f"{errors} errors | {elapsed_min:.1f} min elapsed ──")
            print()

    skip_file.close()
    elapsed_total = (datetime.now() - start_time).total_seconds() / 60

    print()
    print("=" * 55)
    print(f"COMPLETE")
    print(f"  Processed:  {inserted:,}")
    print(f"  Skipped:    {skipped:,} (timeout)")
    print(f"  Errors:     {errors:,}")
    print(f"  Time:       {elapsed_total:.1f} minutes")
    print(f"  Skip log:   {skip_log}")
    print()
    print("Now load into SDILive:")
    print("  python src\\historical_ingest_to_sdilive.py --all")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="SDI Intelligence — Safe Historical Batch Ingest with timeout"
    )
    parser.add_argument("--root",    required=True,
                        help="Root folder to scan (e.g. K:\\Estimating\\Completed\\Manual Estimates)")
    parser.add_argument("--out",     required=True,
                        help="Output folder for JSON files")
    parser.add_argument("--timeout", type=int, default=45,
                        help="Seconds before skipping a file (default: 45)")
    args = parser.parse_args()
    run_ingest(args.root, args.out, args.timeout)

