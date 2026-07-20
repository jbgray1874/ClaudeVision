"""
Automated full parity: Estimate workbook (cached Excel values) vs Manufacturing JSON outcome.

Supports .xls (via xlrd), .xlsx, .xlsm (via openpyxl data_only=True, or optional Excel COM on Windows).
For .xls: xlrd reads computed values directly — no recalculation needed.
For .xlsx/.xlsm: prefer ``--full-parity-read-via-excel`` (Windows + pywin32) if openpyxl reads zeros for formulas;
otherwise Excel must recalculate and save so openpyxl reads cached numbers.
Produces JSON bundle + flat CSV — see config.ESTIMATE_FULL_PARITY, JOB_QUOTE_QUANTITY_BREAKS.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import config
from estimate_sheet_discovery import discover_labour_route_row_span, discover_output_cells_map
from operation_normaliser import (
    AI_TO_SDI,
    ai_op_to_primary_sdi_code,
    display_label,
    get_all_sdi_codes_for_ai_op,
    normalise_operation_code,
)

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover
    load_workbook = None

try:
    import xlrd  # type: ignore
except ImportError:  # pragma: no cover
    xlrd = None


class _XlrdSheetWrapper:
    """Thin adapter so .xls sheets expose the same .cell(row, column).value API as openpyxl."""

    def __init__(self, xlrd_sheet: Any) -> None:
        self._s = xlrd_sheet

    def cell(self, row: int, column: int) -> Any:
        class _Cell:
            def __init__(self, value: Any) -> None:
                self.value = value
        try:
            return _Cell(self._s.cell_value(row - 1, column - 1))
        except IndexError:
            return _Cell(None)


class _XlrdWorkbookWrapper:
    """Thin adapter so .xls workbooks expose .sheetnames and __getitem__ like openpyxl."""

    def __init__(self, xlrd_book: Any) -> None:
        self._book = xlrd_book
        self.sheetnames: List[str] = xlrd_book.sheet_names()

    def __getitem__(self, name: str) -> _XlrdSheetWrapper:
        return _XlrdSheetWrapper(self._book.sheet_by_name(name))

    def close(self) -> None:
        self._book.release_resources()


class _ExcelComCell:
    __slots__ = ("value",)

    def __init__(self, value: Any) -> None:
        self.value = value


class _ExcelComSheetWrapper:
    """Adapter so Excel COM worksheets expose .cell(row, column).value like openpyxl."""

    def __init__(self, com_worksheet: Any) -> None:
        self._ws = com_worksheet

    def cell(self, row: int, column: int) -> _ExcelComCell:
        try:
            v = self._ws.Cells(row, column).Value
        except Exception:
            return _ExcelComCell(None)
        return _ExcelComCell(v)


class _ExcelComWorkbookWrapper:
    """Adapter so Excel COM workbooks expose .sheetnames, __getitem__, and .close()."""

    def __init__(self, excel_app: Any, com_workbook: Any) -> None:
        self._excel = excel_app
        self._wb = com_workbook
        n = int(com_workbook.Sheets.Count)
        self.sheetnames: List[str] = [str(com_workbook.Sheets(i).Name) for i in range(1, n + 1)]

    def __getitem__(self, name: str) -> _ExcelComSheetWrapper:
        return _ExcelComSheetWrapper(self._wb.Worksheets(name))

    def close(self) -> None:
        wb, xl = self._wb, self._excel
        self._wb = None
        self._excel = None
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if xl is not None:
                xl.Quit()
        except Exception:
            pass


def _open_workbook_excel_com(path: Path, *, prime_sheet: Optional[str] = None) -> _ExcelComWorkbookWrapper:
    """Open workbook in Excel and force a full calculation so cross-sheet formulas resolve before reading cells."""
    if sys.platform != "win32":
        raise RuntimeError("Excel COM parity reader is only supported on Windows.")
    try:
        import win32com.client  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "Excel COM reader needs pywin32: pip install pywin32"
        ) from e
    p = str(path.resolve())
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    try:
        excel.EnableEvents = False
    except Exception:
        pass
    try:
        com_wb = excel.Workbooks.Open(p, 0, True)
    except Exception:
        try:
            excel.Quit()
        except Exception:
            pass
        raise
    try:
        # xlCalculationAutomatic — cross-sheet refs often stay stale until a full calc.
        try:
            excel.Calculation = -4105
        except Exception:
            pass
        try:
            com_wb.ForceFullCalculation = True
        except Exception:
            pass
        if prime_sheet:
            try:
                com_wb.Worksheets(prime_sheet).Activate()
            except Exception:
                pass
        try:
            excel.CalculateFull()
        except Exception:
            try:
                excel.Calculate()
            except Exception:
                pass
        try:
            excel.CalculateUntilAsyncQueriesDone()
        except Exception:
            pass
    except Exception:
        pass
    return _ExcelComWorkbookWrapper(excel, com_wb)


def _sdi_operation_table() -> List[Dict[str, Any]]:
    rows = getattr(config, "SDI_OPERATION_CODES", None)
    if rows:
        return list(rows)
    return [
        {"code": "LASM", "internal_estimator_op": "laser_cutting"},
        {"code": "FOLD", "internal_estimator_op": "folding"},
        {"code": "WELD", "internal_estimator_op": "welding"},
        {"code": "PC", "internal_estimator_op": "powder_coating"},
        {"code": "SPRY", "internal_estimator_op": "wet_spray"},
        {"code": "PACM", "internal_estimator_op": "assembly"},
        {"code": "HAND", "internal_estimator_op": "handling"},
        {"code": "DRIL", "internal_estimator_op": "drilling"},
        {"code": "COUN", "internal_estimator_op": "countersinking"},
        {"code": "TAP", "internal_estimator_op": "tapping"},
        {"code": "DPOL", "internal_estimator_op": "diamond_polish"},
        {"code": "GLUE", "internal_estimator_op": "glue"},
    ]


def _canonical_sdi_from_internal(internal_op: str) -> Optional[str]:
    """Map AI estimator op name to primary SDI workbook code (parity display)."""
    op = internal_op.strip().lower()
    primary = ai_op_to_primary_sdi_code(op)
    if primary:
        return primary.upper()
    for row in _sdi_operation_table():
        iop = row.get("internal_estimator_op")
        if iop and str(iop).lower() == op:
            return str(row["code"]).strip().upper()
    if op == "hole_machining":
        return "DRIL"
    return None


def _canonical_from_internal(internal_op: str) -> Optional[str]:
    """Map AI estimator op name to canonical operation key for aggregation."""
    op = internal_op.strip().lower()
    if op in AI_TO_SDI:
        return op
    mapped = normalise_operation_code(op)
    if mapped:
        return mapped
    for row in _sdi_operation_table():
        iop = row.get("internal_estimator_op")
        if iop and str(iop).lower() == op:
            return str(iop).lower()
    return None


def _safe_float(value: Any) -> Optional[float]:
    try:
        if value is None:
            return None
        if isinstance(value, str):
            v = value.strip().replace("£", "").replace(",", "")
            if not v:
                return None
            return float(v)
        return float(value)
    except (TypeError, ValueError):
        return None


def _pct_diff(expected: Optional[float], actual: Optional[float]) -> Optional[float]:
    if expected is None or actual is None:
        return None
    if abs(expected) < 1e-9:
        return None
    return round(abs(actual - expected) / abs(expected) * 100.0, 4)


def _status_from_pct(pct: Optional[float], match_pct: float, warn_pct: float) -> str:
    if pct is None:
        return "review"
    if pct <= match_pct:
        return "match"
    if pct <= warn_pct:
        return "warning"
    return "fail"


def _col_idx(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        if not ("A" <= ch <= "Z"):
            raise ValueError("bad column %r" % (letters,))
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _extract_by_path(payload: Dict[str, Any], dotted_path: str) -> Any:
    cur: Any = payload
    for key in dotted_path.split("."):
        if not isinstance(cur, dict):
            return None
        cur = cur.get(key)
    return cur


_ASSUMED_JOB_QTY_PATH = "estimate_summary.estimate_workbook_inputs.assumed_job_quantity"


def _normalize_workbook_quantity_value(val: float) -> Any:
    """Prefer integers for typical whole-number order quantities."""
    if abs(val - round(val)) < 1e-6:
        return int(round(val))
    return round(val, 6)


def _summary_for_parity_with_workbook_order_qty(
    summary: Dict[str, Any],
    workbook_qty: float,
    *,
    qty_cell: str,
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """
    Shallow copy of ``summary`` so money-cell extraction sees the workbook reference order qty.

    The PDF/JSON snapshot often still carries ``config.DEFAULT_JOB_QUANTITY`` because no template
    workbook was merged during ``estimate_document``. Full parity should compare money labels using
    the same reference quantity as the open workbook (Estimate D6 or discovered equivalent).
    """
    prev = _safe_float(_extract_by_path(summary, _ASSUMED_JOB_QTY_PATH))
    norm = _normalize_workbook_quantity_value(float(workbook_qty))
    meta = {
        "workbook_cell": qty_cell,
        "workbook_quantity": norm,
        "json_quantity_before_overlay": prev,
        "note": "assumed_job_quantity in this clone matches the parity workbook reference qty for label comparisons only.",
    }
    out: Dict[str, Any] = dict(summary)
    est = summary.get("estimate_summary")
    if not isinstance(est, dict):
        out["estimate_summary"] = {"estimate_workbook_inputs": {"assumed_job_quantity": norm}}
        return out, meta
    out_est = dict(est)
    ewb = est.get("estimate_workbook_inputs")
    if isinstance(ewb, dict):
        out_est["estimate_workbook_inputs"] = {**ewb, "assumed_job_quantity": norm}
    else:
        out_est["estimate_workbook_inputs"] = {"assumed_job_quantity": norm}
    out["estimate_summary"] = out_est
    return out, meta


def _split_cell(ref: str) -> Tuple[int, int]:
    m = re.match(r"^([A-Za-z]{1,3})(\d+)$", str(ref).strip())
    if not m:
        raise ValueError("bad cell %r" % (ref,))
    return _col_idx(m.group(1)), int(m.group(2))


def _col_letter_from_index(col_idx: int) -> str:
    """1-based column index to Excel letters (1 -> A)."""
    n = col_idx
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _cell_text_lower(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip().lower()
    if isinstance(value, (int, float, bool)):
        return str(value).strip().lower()
    return str(value).strip().lower()


def _parity_money_specs() -> List[Dict[str, Any]]:
    """
    Dynamic label-based money parity targets. No workbook cell addresses.

    ``summary_path`` matches ``estimate_summary`` JSON from ``estimate_document``.
    """
    return [
        {
            "label": "Reference order quantity",
            "summary_path": "estimate_summary.estimate_workbook_inputs.assumed_job_quantity",
            "search_terms": ["quantity", "order qty", "parts per assembly", "assembly qty", "qty"],
            "context_must_include_any": [],
            "exclude_if_any": [],
            "label_col_max": 9,
            "pick": "first_strong",
        },
        {
            "label": "Material subtotal",
            "summary_path": "estimate_summary.workbook_equivalent_pricing.m59_material_subtotal_gbp",
            "search_terms": ["material", "bought in", "bought-in", "bought in total", "sheet", "plate"],
            "context_must_include_any": ["subtotal", "sub total", "total"],
            "exclude_if_any": ["labour", "labor", "laser metal"],
            "pick": "last",
            "fallback_read_columns": ["M", "L"],
            "value_scan_span": 14,
        },
        {
            "label": "Labour subtotal",
            "summary_path": "estimate_summary.workbook_equivalent_pricing.m103_labour_subtotal_gbp",
            "search_terms": ["labour", "labor", "operations", "production time"],
            "context_must_include_any": ["subtotal", "sub total", "total"],
            "exclude_if_any": ["material", "sell"],
            "pick": "last",
            "fallback_read_columns": ["M", "L"],
            "value_scan_span": 14,
        },
        {
            "label": "Unit manufacturing cost (L)",
            "summary_path": "estimate_summary.workbook_equivalent_pricing.l105_total_unit_cost_gbp",
            "search_terms": ["unit manufacturing", "mfg cost", "manufacturing cost", "unit cost", "factory cost"],
            "context_must_include_any": [],
            "exclude_if_any": ["sell", "selling", "customer price", "quoted sell"],
            "pick": "last",
            "read_column": "L",
        },
        {
            "label": "Unit manufacturing cost (M)",
            "summary_path": "estimate_summary.workbook_equivalent_pricing.m105_total_unit_cost_gbp",
            "search_terms": ["unit manufacturing", "mfg cost", "manufacturing cost", "unit cost", "factory cost"],
            "context_must_include_any": [],
            "exclude_if_any": ["sell", "selling", "customer price", "quoted sell"],
            "pick": "last",
            "read_column": "M",
        },
        {
            "label": "Sell / quoted price",
            "summary_path": "estimate_summary.workbook_equivalent_pricing.l111_sell_price_gbp",
            "search_terms": ["sell", "selling", "customer price", "quoted price", "unit price"],
            "context_must_include_any": [],
            "exclude_if_any": ["manufacturing cost", "mfg cost", "excl", "excluding uplift"],
            "pick": "last",
        },
    ]


def _label_match_score(text: str, spec: Dict[str, Any]) -> float:
    if not text:
        return 0.0
    for ex in spec.get("exclude_if_any") or []:
        if ex.lower() in text:
            return 0.0
    hits = 0
    for t in spec.get("search_terms") or []:
        if t.lower() in text:
            hits += 1
    if hits == 0:
        return 0.0
    ctx = spec.get("context_must_include_any") or []
    if ctx and not any(c.lower() in text for c in ctx):
        return 0.0
    score = float(hits)
    for c in ctx:
        if c.lower() in text:
            score += 1.5
    if spec.get("label_col_max"):
        score += 0.01
    return score


def _sheet_scan_bounds(ws: Any) -> Tuple[int, int]:
    """Use worksheet dimensions (openpyxl) or xlrd nrows/ncols; cap for safety."""
    mr = getattr(ws, "max_row", None)
    mc = getattr(ws, "max_column", None)
    if isinstance(mr, int) and mr > 0:
        max_row = min(mr, 800)
    else:
        inner = getattr(ws, "_s", None)
        max_row = min(int(getattr(inner, "nrows", 500) or 500), 800) if inner is not None else 500
    if isinstance(mc, int) and mc > 0:
        max_col = min(mc, 40)
    else:
        inner = getattr(ws, "_s", None)
        max_col = min(int(getattr(inner, "ncols", 30) or 30), 40) if inner is not None else 30
    return max_row, max_col


def _iter_all_cells(ws: Any, max_row: int, max_col: int):
    """Yield (row, col, raw_value, text_lower) across the used grid. Prefer iter_rows when available."""
    if hasattr(ws, "iter_rows") and callable(getattr(ws, "iter_rows")):
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                raw = getattr(cell, "value", None)
                txt = _cell_text_lower(raw)
                if txt:
                    yield int(cell.row), int(cell.column), raw, txt
        return
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            raw = ws.cell(row=r, column=c).value
            txt = _cell_text_lower(raw)
            if txt:
                yield r, c, raw, txt


def _collect_scored_matches(ws: Any, spec: Dict[str, Any], max_row: int, max_col: int) -> List[Tuple[float, int, int, str]]:
    """List of (score, row, col, text) for cells that match this spec."""
    out: List[Tuple[float, int, int, str]] = []
    col_max = int(spec["label_col_max"]) if spec.get("label_col_max") else 999
    for r, c, _raw, text in _iter_all_cells(ws, max_row, max_col):
        if c > col_max:
            continue
        sc = _label_match_score(text, spec)
        if sc > 0:
            out.append((sc, r, c, text))
    return out


def _pick_label_cell(scored: List[Tuple[float, int, int, str]], spec: Dict[str, Any]) -> Optional[Tuple[int, int, float]]:
    if not scored:
        return None
    pick = str(spec.get("pick") or "last")
    if pick == "first_strong":
        best_sc = max(s[0] for s in scored)
        pool = [s for s in scored if s[0] >= best_sc - 0.25]
        pool.sort(key=lambda t: (t[1], t[2]))
        return (pool[0][1], pool[0][2], pool[0][0])
    scored.sort(key=lambda t: (t[0], t[1], t[2]))
    return (scored[-1][1], scored[-1][2], scored[-1][0])


def _a1(row: int, col: int) -> str:
    return "%s%d" % (_col_letter_from_index(col), row)


def _first_numeric_to_right(ws: Any, row: int, label_col: int, *, max_span: int = 8) -> Tuple[Optional[str], Any]:
    last_c = label_col + max_span
    for c in range(label_col + 1, last_c + 1):
        raw = ws.cell(row=row, column=c).value
        v = _safe_float(raw)
        if v is not None:
            return _a1(row, c), raw
    return None, None


def _resolve_workbook_money_value(
    ws: Any,
    row: int,
    label_col: int,
    spec: Dict[str, Any],
) -> Tuple[Optional[str], Any, str]:
    """
    Read a money value on the label row: fixed column, scan right, then M/L fallbacks.

    Returns (cell_ref, raw_value, source_tag).
    """
    read_col = spec.get("read_column")
    if read_col:
        cell_ref, raw_wb = _value_on_row_column(ws, row, str(read_col))
        if _safe_float(raw_wb) is not None:
            return cell_ref, raw_wb, "read_column"
    span = int(spec.get("value_scan_span") or 8)
    cell_ref, raw_wb = _first_numeric_to_right(ws, row, label_col, max_span=span)
    if _safe_float(raw_wb) is not None:
        return cell_ref, raw_wb, "scan_right"
    for col_letter in spec.get("fallback_read_columns") or []:
        cell_ref, raw_wb = _value_on_row_column(ws, row, str(col_letter))
        if _safe_float(raw_wb) is not None:
            return cell_ref, raw_wb, "fallback_column"
    return None, None, "none"


def _cells_by_summary_path(cell_map: Dict[str, str]) -> Dict[str, List[str]]:
    by_path: Dict[str, List[str]] = {}
    for cell_ref, path in cell_map.items():
        by_path.setdefault(str(path), []).append(str(cell_ref).upper())
    return by_path


def _upgrade_money_rows_from_total_discovery(
    ws: Any,
    rows_out: List[Dict[str, Any]],
    *,
    match_pct: float,
    warn_pct: float,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Fill money rows with null workbook values using ESTIMATE_SHEET_TOTAL_DISCOVERY (I:K labels → M)."""
    cell_map, disc_meta = discover_output_cells_map(ws)
    by_path = _cells_by_summary_path(cell_map)
    upgraded = 0
    for row in rows_out:
        if row.get("workbook_cached_numeric") is not None:
            continue
        path = str(row.get("json_path") or "")
        for cell_ref in by_path.get(path) or []:
            col_i, rn = _split_cell(cell_ref)
            raw_wb = ws.cell(row=rn, column=col_i).value
            wb_val = _safe_float(raw_wb)
            if wb_val is None:
                continue
            ej = _safe_float(row.get("json_numeric"))
            pct = _pct_diff(ej, wb_val)
            row["cell"] = cell_ref
            row["workbook_cached_numeric"] = wb_val
            row["workbook_cached_raw"] = raw_wb
            row["pct_variance"] = pct
            row["status"] = _status_from_pct(
                pct if ej is not None and wb_val is not None else None, match_pct, warn_pct
            )
            row["value_source"] = "estimate_sheet_total_discovery"
            upgraded += 1
            break
    return rows_out, {"estimate_sheet_total_discovery": disc_meta, "upgraded_rows": upgraded}


def _value_on_row_column(ws: Any, row: int, col_letter: str) -> Tuple[Optional[str], Any]:
    col = _col_idx(col_letter)
    raw = ws.cell(row=row, column=col).value
    return _a1(row, col), raw


def _read_money_cells(
    ws: Any,
    specs: List[Dict[str, Any]],
    summary: Dict[str, Any],
    *,
    match_pct: float,
    warn_pct: float,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Scan the full Estimate grid (dynamic bounds), match label text, read value from the row
    (next numeric right of label, or fixed ``read_column``). No config cell map.
    """
    max_row, max_col = _sheet_scan_bounds(ws)
    meta: Dict[str, Any] = {
        "mode": "full_sheet_label_scan",
        "max_row": max_row,
        "max_col": max_col,
        "hits": [],
        "warnings": [],
    }
    if os.getenv("PARITY_SCAN_DEBUG", "").lower() in {"1", "true", "yes"}:
        print("[PARITY] Scanning sheet dynamically: %s rows x %s columns" % (max_row, max_col))

    rows_out: List[Dict[str, Any]] = []
    for spec in specs:
        label = str(spec.get("label") or "")
        path = str(spec.get("summary_path") or "")
        scored = _collect_scored_matches(ws, spec, max_row, max_col)
        hit = _pick_label_cell(scored, spec)
        if not hit:
            meta["warnings"].append({"label": label, "path": path, "detail": "no_label_match"})
            if os.getenv("PARITY_SCAN_DEBUG", "").lower() in {"1", "true", "yes"}:
                print("[PARITY] WARN: could not find label for %r" % (label,))
            rows_out.append(
                {
                    "section": "money_cell",
                    "cell": "",
                    "label": label,
                    "json_path": path,
                    "json_numeric": _safe_float(_extract_by_path(summary, path)),
                    "workbook_cached_numeric": None,
                    "workbook_cached_raw": None,
                    "pct_variance": None,
                    "status": "review",
                }
            )
            continue
        r, lc, sc = hit
        cell_ref, raw_wb, value_source = _resolve_workbook_money_value(ws, r, lc, spec)
        wb_val = _safe_float(raw_wb)
        ej = _safe_float(_extract_by_path(summary, path))
        pct = _pct_diff(ej, wb_val)
        st = _status_from_pct(pct if ej is not None and wb_val is not None else None, match_pct, warn_pct)
        meta["hits"].append(
            {
                "label": label,
                "summary_path": path,
                "label_cell": _a1(r, lc),
                "value_cell": cell_ref or "",
                "label_match_score": sc,
                "value_source": value_source,
            }
        )
        if os.getenv("PARITY_SCAN_DEBUG", "").lower() in {"1", "true", "yes"}:
            print(
                "[PARITY] OK: %r label %s -> value %s (wb=%s json=%s)"
                % (label, _a1(r, lc), cell_ref or "?", wb_val, ej)
            )
        rows_out.append(
            {
                "section": "money_cell",
                "cell": cell_ref or "",
                "label": label,
                "json_path": path,
                "json_numeric": ej,
                "workbook_cached_numeric": wb_val,
                "workbook_cached_raw": raw_wb,
                "pct_variance": pct,
                "status": st,
            }
        )

    rows_out, total_disc_meta = _upgrade_money_rows_from_total_discovery(
        ws, rows_out, match_pct=match_pct, warn_pct=warn_pct
    )
    meta["estimate_totals_discovery"] = total_disc_meta
    return rows_out, meta


def _money_specs_from_override(ws: Any) -> Tuple[List[Dict[str, str]], Dict[str, Any]]:
    """Fixed cell list from ESTIMATE_FULL_PARITY.money_cells (optional escape hatch)."""
    full = getattr(config, "ESTIMATE_FULL_PARITY", None) or {}
    specs: Dict[Tuple[str, str], Dict[str, str]] = {}
    for x in full["money_cells"]:  # type: ignore[index]
        c = str(x["cell"]).upper()
        p = str(x["path"])
        specs[(c, p)] = {"cell": c, "summary_path": p, "label": str(x.get("label") or "")}

    extras = getattr(config, "ESTIMATE_FULL_PARITY_EXTRA_MONEY_CELLS", None) or []
    for row in extras:
        c = str(row["cell"]).upper()
        specs[(c, str(row["summary_path"]))] = {
            "cell": c,
            "summary_path": str(row["summary_path"]),
            "label": str(row.get("label") or ""),
        }

    qty_cell, qty_meta = _quantity_cell_from_label_scan(ws)
    q_path = "estimate_summary.estimate_workbook_inputs.assumed_job_quantity"
    if not any(v["summary_path"] == q_path for v in specs.values()):
        specs[(qty_cell, q_path)] = {
            "cell": qty_cell,
            "summary_path": q_path,
            "label": "reference_order_qty_workbook",
        }

    def _sk(spec: Dict[str, str]):
        ci, rn = _split_cell(spec["cell"])
        return (rn, ci)

    return sorted(specs.values(), key=_sk), {
        "mode": "money_cells_override",
        "quantity_cell": qty_cell,
        "quantity_discovery": qty_meta,
        "totals_discovery": {},
    }


def _quantity_cell_from_label_scan(ws: Any) -> Tuple[str, Dict[str, Any]]:
    """Resolve quantity cell using the same label rules as the first money spec."""
    max_r, max_c = _sheet_scan_bounds(ws)
    spec = _parity_money_specs()[0]
    scored = _collect_scored_matches(ws, spec, max_r, max_c)
    hit = _pick_label_cell(scored, spec)
    meta: Dict[str, Any] = {"mode": "label_scan", "max_row": max_r, "max_col": max_c}
    if not hit:
        meta["mode"] = "fallback"
        meta["cell"] = "D6"
        return "D6", meta
    r, lc, _sc = hit
    for c in range(lc + 1, min(lc + 8, max_c + 1)):
        v = _safe_float(ws.cell(row=r, column=c).value)
        if v is not None and v > 0:
            ref = _a1(r, c)
            meta["cell"] = ref
            return ref, meta
    meta["mode"] = "fallback"
    meta["cell"] = "D6"
    return "D6", meta


def _build_money_parity_from_sheet(
    ws: Any,
    summary: Dict[str, Any],
    *,
    match_pct: float,
    warn_pct: float,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Build money_cell_comparisons: label-driven scan, or fixed cells if money_cells override is set.
    """
    full = getattr(config, "ESTIMATE_FULL_PARITY", None) or {}
    overrides = isinstance(full.get("money_cells"), list) and len(full["money_cells"]) > 0

    if overrides:
        money_specs, meta = _money_specs_from_override(ws)
        money_rows: List[Dict[str, Any]] = []
        for spec in money_specs:
            col_i, rn = _split_cell(spec["cell"])
            wb_val = ws.cell(row=rn, column=col_i).value
            raw_json = _extract_by_path(summary, spec["summary_path"])
            ej = _safe_float(raw_json)
            aw = _safe_float(wb_val)
            pct = _pct_diff(ej, aw)
            st = _status_from_pct(pct if ej is not None and aw is not None else None, match_pct, warn_pct)
            money_rows.append(
                {
                    "section": "money_cell",
                    "cell": spec["cell"],
                    "label": spec.get("label") or "",
                    "json_path": spec["summary_path"],
                    "json_numeric": ej,
                    "workbook_cached_numeric": aw,
                    "workbook_cached_raw": wb_val,
                    "pct_variance": pct,
                    "status": st,
                }
            )
        return money_rows, meta

    money_rows, scan_meta = _read_money_cells(
        ws, _parity_money_specs(), summary, match_pct=match_pct, warn_pct=warn_pct
    )
    qty_cell, qty_meta = _quantity_cell_from_label_scan(ws)
    extras = getattr(config, "ESTIMATE_FULL_PARITY_EXTRA_MONEY_CELLS", None) or []
    for row in extras:
        c = str(row["cell"]).upper()
        p = str(row["summary_path"])
        col_i, rn = _split_cell(c)
        wb_val = ws.cell(row=rn, column=col_i).value
        raw_json = _extract_by_path(summary, p)
        ej = _safe_float(raw_json)
        aw = _safe_float(wb_val)
        pct = _pct_diff(ej, aw)
        st = _status_from_pct(pct if ej is not None and aw is not None else None, match_pct, warn_pct)
        money_rows.append(
            {
                "section": "money_cell",
                "cell": c,
                "label": str(row.get("label") or ""),
                "json_path": p,
                "json_numeric": ej,
                "workbook_cached_numeric": aw,
                "workbook_cached_raw": wb_val,
                "pct_variance": pct,
                "status": st,
            }
        )

    discovery_meta = {
        "mode": "label_grid_scan",
        "quantity_cell": qty_cell,
        "quantity_discovery": qty_meta,
        "totals_discovery": scan_meta,
    }
    return money_rows, discovery_meta


def _parts_from_estimate(est: Dict[str, Any]) -> List[Dict[str, Any]]:
    raw = est.get("part_estimates") or est.get("parts") or []
    return [p for p in raw if isinstance(p, dict)]


def _flatten_provenance(est: Dict[str, Any]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for p in _parts_from_estimate(est):
        if not isinstance(p, dict):
            continue
        mat = p.get("material_estimate") or {}
        lab = p.get("labour_estimate") or {}
        proc = p.get("process_estimate") or {}
        out.append(
            {
                "part_number": p.get("part_number"),
                "bom_quantity": p.get("quantity"),
                "normalized_material": p.get("normalized_material") or mat.get("material"),
                "normalized_thickness_mm": p.get("normalized_thickness_mm"),
                "material_extended_cost_gbp": mat.get("extended_material_cost_gbp"),
                "material_price_basis": mat.get("price_source") or {},
                "powder_detail": mat.get("powder_consumable"),
                "labour_operations_cost_gbp": lab.get("costs_gbp") or {},
                "labour_rate_sources": lab.get("rate_sources") or {},
                "labour_missing_rates": lab.get("missing_rate_operations") or [],
                "process_setup_min": proc.get("setup_times_min"),
                "process_run_min_per_unit": proc.get("run_times_min_per_unit"),
                "process_total_min_job": proc.get("times_min"),
                "routing": proc.get("routing"),
                "manufacturing_features": proc.get("manufacturing_features"),
                "risk_flags": p.get("risk_flags") or [],
                "costing_basis": (p.get("cost_breakdown") or {}).get("costing_basis"),
            }
        )
    return out


def _aggregate_labour_by_canonical(part_estimates: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """Roll up JSON labour by canonical operation (laser_cutting, folding, …)."""
    totals: Dict[str, Dict[str, Any]] = {}
    for p in part_estimates:
        if not isinstance(p, dict):
            continue
        proc = p.get("process_estimate") or {}
        lab = p.get("labour_estimate") or {}
        tm = proc.get("times_min") if isinstance(proc.get("times_min"), dict) else {}
        costs = lab.get("costs_gbp") if isinstance(lab.get("costs_gbp"), dict) else {}
        rs = lab.get("rate_sources") if isinstance(lab.get("rate_sources"), dict) else {}

        # pricing_service.py: no per-op breakdown — rebuild from joined_sources or ops list
        if not costs:
            joined = p.get("joined_sources") or {}
            wb_lab = (joined.get("reverse_engineered_workbook") or {}).get("labour") or {}
            wb_costs = wb_lab.get("costs_gbp") if isinstance(wb_lab.get("costs_gbp"), dict) else {}
            if wb_costs:
                costs = wb_costs
                tm = wb_lab.get("times_min") if isinstance(wb_lab.get("times_min"), dict) else tm
                rs = wb_lab.get("rate_sources") if isinstance(wb_lab.get("rate_sources"), dict) else rs
            else:
                ops = list(
                    dict.fromkeys(
                        [str(o) for o in (p.get("textual_operations") or []) if o]
                        + [str(o) for o in (p.get("inferred_operations") or []) if o]
                    )
                )
                total_lab = _safe_float(p.get("labour_cost_gbp"))
                if total_lab is None:
                    total_lab = _safe_float(lab.get("total_labour_cost_gbp"))
                if ops and total_lab and total_lab > 0:
                    share = float(total_lab) / len(ops)
                    costs = {op: share for op in ops}

        for op_internal, cost in costs.items():
            if not isinstance(op_internal, str):
                continue
            cost_f = float(cost or 0.0)
            min_pt = float(tm.get(op_internal, 0.0) or 0.0)

            canonical = _canonical_from_internal(op_internal)
            if canonical is None:
                canonical = normalise_operation_code(op_internal)
            if canonical is None:
                agg = totals.setdefault(
                    "__unmapped_internal_operations__",
                    {"total_minutes": 0.0, "cost_gbp": 0.0, "ops": set()},
                )
                agg["total_minutes"] += min_pt
                agg["cost_gbp"] += cost_f
                agg["ops"].add(op_internal)
                continue

            agg = totals.setdefault(
                canonical,
                {"total_minutes": 0.0, "cost_gbp": 0.0, "ops": set(), "sdi_codes": set()},
            )
            agg["total_minutes"] += min_pt
            agg["cost_gbp"] += cost_f
            agg["ops"].add(op_internal)
            primary = ai_op_to_primary_sdi_code(canonical)
            if primary:
                agg["sdi_codes"].add(primary.upper())
            rss = rs.get(op_internal) if isinstance(rs, dict) else None
            if isinstance(rss, dict) and rss.get("hourly_rate_gbp"):
                lst = agg.setdefault("rate_hints_gbp_per_hour", [])
                fv = _safe_float(rss.get("hourly_rate_gbp"))
                if fv:
                    lst.append(float(fv))

    for agg in totals.values():
        if isinstance(agg.get("ops"), set):
            agg["ops"] = sorted(agg["ops"])
        if isinstance(agg.get("sdi_codes"), set):
            agg["sdi_codes"] = sorted(agg["sdi_codes"])
        agg["hours_decimal"] = round(float(agg.get("total_minutes", 0.0)) / 60.0, 4)
        agg["cost_gbp"] = round(float(agg.get("cost_gbp", 0.0)), 4)
    return totals


def _aggregate_workbook_labour_by_canonical(
    lab_wb: List[Dict[str, Any]],
) -> Dict[str, Dict[str, Any]]:
    """Roll up workbook labour route rows by canonical operation."""
    totals: Dict[str, Dict[str, Any]] = {}
    for row_w in lab_wb:
        code = row_w.get("operation_code")
        if not code:
            continue
        canonical = normalise_operation_code(str(code))
        if not canonical:
            continue
        agg = totals.setdefault(
            canonical,
            {
                "hours_decimal": 0.0,
                "cost_gbp": 0.0,
                "workbook_codes": set(),
                "sheet_rows": [],
            },
        )
        hrs = _safe_float(row_w.get("total_h"))
        cost = _safe_float(row_w.get("line_cost_gbp_workbook"))
        if hrs:
            agg["hours_decimal"] += float(hrs)
        if cost:
            agg["cost_gbp"] += float(cost)
        agg["workbook_codes"].add(str(code).strip().upper())
        agg["sheet_rows"].append(row_w.get("sheet_row"))

    for agg in totals.values():
        if isinstance(agg.get("workbook_codes"), set):
            agg["workbook_codes"] = sorted(agg["workbook_codes"])
        agg["hours_decimal"] = round(float(agg.get("hours_decimal", 0.0)), 4)
        agg["cost_gbp"] = round(float(agg.get("cost_gbp", 0.0)), 4)
    return totals


def _build_labour_route_comparisons(
    agg_json: Dict[str, Dict[str, Any]],
    agg_wb: Dict[str, Dict[str, Any]],
    lab_wb: List[Dict[str, Any]],
    *,
    match_pct: float,
    warn_pct: float,
) -> List[Dict[str, Any]]:
    """Compare JSON vs workbook labour at canonical-operation level."""
    lab_compare: List[Dict[str, Any]] = []

    for row_w in lab_wb:
        code = row_w.get("operation_code")
        if not code:
            lab_compare.append({"section": "labour_route", "status": "blank", **row_w})
            continue
        canonical = normalise_operation_code(str(code))
        if not canonical:
            lab_compare.append(
                {
                    "section": "labour_route",
                    "status": "noise",
                    "operation_code": str(code).strip().upper(),
                    "canonical_operation": None,
                    "display_label": str(code).strip()[:60],
                    "sheet_row": row_w.get("sheet_row"),
                    "json_hours_decimal": None,
                    "workbook_hours_decimal": row_w.get("total_h"),
                    "json_labour_cost_gbp": None,
                    "workbook_line_cost_gbp": row_w.get("line_cost_gbp_workbook"),
                    "detail": "Header/summary row or unrecognised code — excluded from match counts",
                }
            )
            continue

    all_canonical = sorted(
        set(agg_json.keys()) | set(agg_wb.keys()) - {"__unmapped_internal_operations__"}
    )

    for canonical in all_canonical:
        if canonical == "__unmapped_internal_operations__":
            continue
        j = agg_json.get(canonical) or {}
        w = agg_wb.get(canonical) or {}
        hrs_json = j.get("hours_decimal")
        cost_json = j.get("cost_gbp")
        hrs_wb = w.get("hours_decimal")
        cost_wb = w.get("cost_gbp")

        ph = _pct_diff(hrs_json, hrs_wb)
        pc = _pct_diff(cost_json, cost_wb)
        sts = [
            _status_from_pct(ph if hrs_json is not None and hrs_wb is not None else None, match_pct, warn_pct),
            _status_from_pct(pc if cost_json is not None and cost_wb is not None else None, match_pct, warn_pct),
        ]
        if hrs_json is None and cost_json is None:
            worst = "review"
        elif hrs_wb is None and cost_wb is None:
            worst = "review"
        else:
            worst = "fail" if "fail" in sts else ("warning" if "warning" in sts else ("match" if "match" in sts else "review"))

        primary_code = ai_op_to_primary_sdi_code(canonical) or canonical
        lab_compare.append(
            {
                "section": "labour_route",
                "status": worst,
                "comparison_mode": "canonical",
                "operation_code": primary_code,
                "canonical_operation": canonical,
                "display_label": display_label(canonical),
                "workbook_operation_codes": w.get("workbook_codes") or [],
                "json_hours_decimal": hrs_json,
                "workbook_hours_decimal": hrs_wb,
                "hours_pct_variance": ph,
                "json_labour_cost_gbp": cost_json,
                "workbook_line_cost_gbp": cost_wb,
                "cost_pct_variance": pc,
                "json_ops_rolled_up": j.get("ops"),
                "workbook_sheet_rows": w.get("sheet_rows"),
                "json_sdi_codes": j.get("sdi_codes") or get_all_sdi_codes_for_ai_op(canonical)[:6],
            }
        )

    unmapped = agg_json.get("__unmapped_internal_operations__")
    if unmapped:
        lab_compare.append(
            {
                "section": "labour_route_json_only",
                "status": "review",
                "operation_code": "__unmapped_internal_operations__",
                "canonical_operation": None,
                "display_label": "Unmapped AI operations",
                "json_minutes_total": round(float(unmapped.get("total_minutes", 0.0)), 4),
                "json_cost_gbp": round(float(unmapped.get("cost_gbp", 0.0)), 4),
                "internal_ops": unmapped.get("ops"),
                "detail": "Estimator operations with no canonical mapping — extend operation_normaliser.AI_TO_SDI",
            }
        )

    return lab_compare


def _read_labour_route_rows(
    ws: Any,
    *,
    start: int,
    end: int,
    operation_column: str = "B",
) -> List[Dict[str, Any]]:
    ix_code = _col_idx(str(operation_column or "B").strip().upper() or "B")
    ix_k, ix_l, ix_m, ix_n, ix_o = (
        _col_idx("K"),
        _col_idx("L"),
        _col_idx("M"),
        _col_idx("N"),
        _col_idx("O"),
    )
    rows: List[Dict[str, Any]] = []

    for r in range(start, end + 1):
        raw_code = ws.cell(row=r, column=ix_code).value
        if raw_code is None or str(raw_code).strip() == "":
            rows.append({"sheet_row": r, "operation_code": None})
            continue
        rows.append(
            {
                "sheet_row": r,
                "operation_code": str(raw_code).strip().upper(),
                "setup_h": _safe_float(ws.cell(row=r, column=ix_k).value),
                "run_h": _safe_float(ws.cell(row=r, column=ix_l).value),
                "total_h": _safe_float(ws.cell(row=r, column=ix_m).value),
                "hourly_rate_gbp": _safe_float(ws.cell(row=r, column=ix_n).value),
                "line_cost_gbp_workbook": _safe_float(ws.cell(row=r, column=ix_o).value),
            }
        )

    return rows


def build_full_parity_report(
    summary: Dict[str, Any],
    workbook_path: Path,
    *,
    estimate_sheet_name: Optional[str] = None,
    labour_row_start: Optional[int] = None,
    labour_row_end: Optional[int] = None,
    read_via_excel: bool = False,
) -> Dict[str, Any]:
    path = workbook_path.resolve()
    suffix = path.suffix.lower()

    if suffix == ".xls":
        if xlrd is None:
            raise RuntimeError("xlrd is required for .xls files (pip install xlrd).")
    elif suffix in {".xlsx", ".xlsm"}:
        if read_via_excel:
            pass
        elif load_workbook is None:
            raise RuntimeError("openpyxl is required (pip install openpyxl).")
    else:
        raise RuntimeError("Unsupported workbook %r — need .xls, .xlsx, or .xlsm." % (path.name,))

    est = summary.get("estimate_summary")
    if not isinstance(est, dict):
        raise RuntimeError("JSON needs estimate_summary (dict).")

    cfg_full = getattr(config, "ESTIMATE_FULL_PARITY", {}) or {}
    thresh = (getattr(config, "WORKBOOK_EQUIVALENT_PRICING", {}) or {}).get("variance_thresholds_pct") or {}
    match_pct = float(thresh.get("match", 3.0))
    warn_pct = float(thresh.get("warning", 10.0))

    sheet_nm = estimate_sheet_name or str(cfg_full.get("estimate_sheet_name") or "Estimate")

    qb_start = int(cfg_full.get("quantity_break_rows_start") or 115)
    qb_end = int(cfg_full.get("quantity_break_rows_end") or 125)

    qty_std = getattr(config, "JOB_QUOTE_QUANTITY_BREAKS", None) or [1, 2, 4, 6, 10, 20, 30, 40, 50]

    read_mode = "openpyxl"
    if suffix == ".xls":
        wb = _XlrdWorkbookWrapper(xlrd.open_workbook(str(path)))
        read_mode = "xlrd"
    elif read_via_excel:
        wb = _open_workbook_excel_com(path, prime_sheet=sheet_nm)
        read_mode = "excel_com"
    else:
        wb = load_workbook(path, data_only=True, rich_text=False)
        read_mode = "openpyxl"
    try:
        if sheet_nm not in wb.sheetnames:
            raise RuntimeError("Sheet %r not in workbook; have: %s" % (sheet_nm, ", ".join(wb.sheetnames)))

        ws = wb[sheet_nm]

        if labour_row_start is not None or labour_row_end is not None:
            lr0 = int(labour_row_start or cfg_full.get("labour_route_row_start") or 117)
            lr1 = int(labour_row_end or cfg_full.get("labour_route_row_end") or 148)
            labour_route_meta: Dict[str, Any] = {"mode": "kwargs_or_config"}
        elif bool(cfg_full.get("labour_route_discover", True)):
            lr0, lr1, labour_route_meta = discover_labour_route_row_span(ws)
        else:
            lr0 = int(cfg_full.get("labour_route_row_start") or 117)
            lr1 = int(cfg_full.get("labour_route_row_end") or 148)
            labour_route_meta = {"mode": "config_only"}

        ix_f = _col_idx("F")
        qb_values: List[Dict[str, Any]] = []
        for qr in range(qb_start, qb_end + 1):
            v = ws.cell(row=qr, column=ix_f).value
            qb_values.append({"sheet_row": qr, "raw": v, "numeric": _safe_float(v)})

        qty_cell_pre, qty_meta_pre = _quantity_cell_from_label_scan(ws)
        col_qty_pre, row_qty_pre = _split_cell(qty_cell_pre)
        wb_order_qty_pre = _safe_float(ws.cell(row=row_qty_pre, column=col_qty_pre).value)
        json_qty_pre = _safe_float(_extract_by_path(summary, _ASSUMED_JOB_QTY_PATH))
        parity_qty_alignment: Optional[Dict[str, Any]] = None
        summary_for_money = summary
        if (
            wb_order_qty_pre is not None
            and wb_order_qty_pre > 0
            and (json_qty_pre is None or abs(float(json_qty_pre) - float(wb_order_qty_pre)) > 1e-6)
        ):
            summary_for_money, parity_qty_alignment = _summary_for_parity_with_workbook_order_qty(
                summary, float(wb_order_qty_pre), qty_cell=qty_cell_pre
            )

        money_rows, sheet_discovery_meta = _build_money_parity_from_sheet(
            ws, summary_for_money, match_pct=match_pct, warn_pct=warn_pct
        )
        if parity_qty_alignment:
            sheet_discovery_meta = {**sheet_discovery_meta, "parity_quantity_aligned_from_workbook": parity_qty_alignment}
        if qty_meta_pre:
            sheet_discovery_meta = {**sheet_discovery_meta, "quantity_prescan": qty_meta_pre}

        col_qty, row_qty = _split_cell(sheet_discovery_meta["quantity_cell"])
        wb_d6 = _safe_float(ws.cell(row=row_qty, column=col_qty).value)

        part_estimates = _parts_from_estimate(est)
        agg_json = _aggregate_labour_by_canonical(part_estimates)

        op_col = str(labour_route_meta.get("operation_column") or cfg_full.get("labour_route_operation_column") or "B")
        lab_wb = _read_labour_route_rows(ws, start=lr0, end=lr1, operation_column=op_col)
        agg_wb = _aggregate_workbook_labour_by_canonical(lab_wb)
        lab_compare = _build_labour_route_comparisons(
            agg_json,
            agg_wb,
            lab_wb,
            match_pct=match_pct,
            warn_pct=warn_pct,
        )

        doc_total = _safe_float(est.get("document_total_estimated_cost_gbp"))
        implied_unit_from_workbook_qty: Optional[float] = None

        if doc_total is not None and wb_d6 and wb_d6 > 0:
            implied_unit_from_workbook_qty = round(doc_total / wb_d6, 4)

        l105_row = next(
            (r for r in money_rows if str(r.get("json_path", "")).endswith("l105_total_unit_cost_gbp")),
            None,
        )
        wb_l105 = None
        if l105_row and l105_row.get("cell"):
            wb_l105 = l105_row.get("workbook_cached_numeric")
            if wb_l105 is None and l105_row["cell"]:
                col_l105, row_l105 = _split_cell(str(l105_row["cell"]))
                wb_l105 = _safe_float(ws.cell(row=row_l105, column=col_l105).value)
        rollup_l105 = None

        if wb_l105 is not None and implied_unit_from_workbook_qty is not None:
            pcr = _pct_diff(implied_unit_from_workbook_qty, wb_l105)
            rollup_l105 = {
                "label": "unit_cost_workbook_vs_json_total_div_workbook_qty",
                "workbook_unit_cost_cell": (l105_row or {}).get("cell"),
                "workbook_unit_cost_cached": wb_l105,
                "json_implied_unit_using_workbook_qty": implied_unit_from_workbook_qty,
                "pct_variance": pcr,
                "status": _status_from_pct(pcr, match_pct, warn_pct),
            }

        prov = {
            "estimate_policy_manifest": est.get("estimate_policy_manifest"),
            "estimate_review_signals": est.get("estimate_review_signals"),
            "powder_coating_summary": est.get("powder_coating_summary"),
            "estimate_workbook_inputs": est.get("estimate_workbook_inputs"),
            "estimate_source_extract": est.get("estimate_source_extract"),
            "pricing_metadata": (((est.get("cost_breakdown") or {}).get("pricing_metadata")) or {}),
            "parts": _flatten_provenance(est),
        }

        counts = {
            "money_match": sum(1 for r in money_rows if r["status"] == "match"),
            "money_warning": sum(1 for r in money_rows if r["status"] == "warning"),
            "money_fail": sum(1 for r in money_rows if r["status"] == "fail"),
            "labour_route_match": sum(
                1
                for r in lab_compare
                if r.get("section") == "labour_route"
                and r.get("status") == "match"
                and r.get("comparison_mode") == "canonical"
            ),
            "labour_route_issues": sum(
                1
                for r in lab_compare
                if r.get("section") == "labour_route"
                and r.get("comparison_mode") == "canonical"
                and r.get("status") in {"fail", "warning", "review"}
            ),
        }

        note_by_mode = {
            "xlrd": "Values from binary .xls via xlrd (computed results in the file).",
            "openpyxl": (
                "Values from .xlsx via openpyxl data_only=True (Excel cached results). "
                "If formulas show as 0, open in Excel, recalculate, Save, or rerun with --full-parity-read-via-excel."
            ),
            "excel_com": (
                "Values from .xlsx via Excel COM after CalculateFull() and activating the estimate sheet "
                "(cross-sheet formula refs should resolve; if cells are still 0, open once in Excel and Save)."
            ),
        }
        return {
            "schema": "estimate_full_workbook_parity.v1",
            "workbook_path": str(path),
            "estimate_sheet": sheet_nm,
            "workbook_read_mode": read_mode,
            "precalculation_note": note_by_mode.get(read_mode, ""),
            "configured_standard_quantity_break_list": qty_std,
            "workbook_quantity_break_column_F_rows": qb_values,
            "workbook_quantity_cell_ref": sheet_discovery_meta.get("quantity_cell"),
            "estimate_sheet_discovery": sheet_discovery_meta,
            "workbook_cell_D6_quantity": wb_d6,
            "rollup_unit_cost_comparison": rollup_l105,
            "money_cell_comparisons": money_rows,
            "labour_route_comparisons": lab_compare,
            "labour_route_discovery": {
                **labour_route_meta,
                "labour_route_row_start": lr0,
                "labour_route_row_end": lr1,
                "comparison_mode": "canonical",
                "operation_normaliser": "operation_normaliser.py",
            },
            "estimate_provenance": prov,
            "status_counts": counts,
        }
    finally:
        wb.close()


def flatten_rows_for_csv(bundle: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for r in bundle.get("money_cell_comparisons") or []:
        rr = dict(r)
        rr["category"] = "money"
        rows.append(rr)
    for r in bundle.get("labour_route_comparisons") or []:
        rr = dict(r)
        rr["category"] = "labour_route"
        rows.append(rr)
    return rows


def _write_text_preferred(preferred: Path, text: str, *, kind: str) -> Path:
    """Write UTF-8 text; on PermissionError (e.g. file open in Excel) use a timestamped sibling path."""
    preferred.parent.mkdir(parents=True, exist_ok=True)
    try:
        preferred.write_text(text, encoding="utf-8")
        return preferred
    except PermissionError:
        alt = preferred.with_name(f"{preferred.stem}.auto_{int(time.time())}{preferred.suffix}")
        print(
            f"NOTE: Could not write {preferred} ({kind}; permission denied — close it in another app if open). "
            f"Wrote to: {alt}",
            file=sys.stderr,
        )
        alt.write_text(text, encoding="utf-8")
        return alt


def _open_csv_write_preferred(preferred: Path, *, purpose: str = "CSV") -> Tuple[Path, Any]:
    """Open CSV for write; on PermissionError use a timestamped sibling path."""
    preferred.parent.mkdir(parents=True, exist_ok=True)
    try:
        return preferred, preferred.open("w", newline="", encoding="utf-8")
    except PermissionError:
        alt = preferred.with_name(f"{preferred.stem}.auto_{int(time.time())}{preferred.suffix}")
        print(
            f"NOTE: Could not write {preferred} ({purpose}; permission denied — close it in Excel if open). "
            f"Writing to: {alt}",
            file=sys.stderr,
        )
        return alt, alt.open("w", newline="", encoding="utf-8")


def write_full_parity_reports(bundle: Dict[str, Any], out_json: Path, out_csv: Path) -> Tuple[Path, Path]:
    """Write bundle JSON and flat CSV. Returns ``(path_json_written, path_csv_written)`` (CSV path may differ if the preferred file is locked)."""
    json_text = json.dumps(bundle, indent=2, ensure_ascii=False)
    actual_json = _write_text_preferred(out_json, json_text, kind="full parity bundle JSON")
    rows = flatten_rows_for_csv(bundle)
    if not rows:
        return actual_json, out_csv
    hdr = sorted(set().union(*(r.keys() for r in rows)))
    actual_csv, fh = _open_csv_write_preferred(out_csv, purpose="full parity flat CSV")
    try:
        dw = csv.DictWriter(fh, fieldnames=hdr, extrasaction="ignore")
        dw.writeheader()
        dw.writerows(rows)
    finally:
        fh.close()
    return actual_json, actual_csv


def generate_bom_comparison_csv(summary: Dict[str, Any], output_csv: Path) -> Path:
    """
    One row per line in ``estimate_summary.part_estimates`` (BOM-style), AI costs populated.

    ``manual_*`` and ``variance_pct`` are left blank until a workbook BOM extract is wired in.
    """
    est = summary.get("estimate_summary") or {}
    if not isinstance(est, dict):
        est = {}
    parts_est = _parts_from_estimate(est)
    writeup_parts = summary.get("manufacturing_writeup", {}).get("parts") or []
    by_pn: Dict[str, Dict[str, Any]] = {}
    for wp in writeup_parts:
        if isinstance(wp, dict) and wp.get("part_number"):
            by_pn[str(wp["part_number"]).strip().upper()] = wp

    rows: List[Dict[str, Any]] = []
    cols = list(getattr(config, "BOM_COMPARISON_COLUMNS", []) or [])
    if not cols:
        raise RuntimeError("config.BOM_COMPARISON_COLUMNS is empty")

    def _blank_row() -> Dict[str, Any]:
        return {c: "" for c in cols}

    for pe in parts_est:
        pn = str(pe.get("part_number") or "").strip()
        if not pn:
            continue
        wp = by_pn.get(pn.upper(), {})
        mat_e = pe.get("material_estimate") or {}
        lab_e = pe.get("labour_estimate") or {}
        cb = pe.get("cost_breakdown") or {}
        mat_gbp = float(
            pe.get("material_cost_gbp")
            or mat_e.get("extended_material_cost_gbp")
            or 0.0
        )
        lab_gbp = float(
            pe.get("labour_cost_gbp")
            or lab_e.get("total_labour_cost_gbp")
            or 0.0
        )
        ext = cb.get("extended_total_cost_gbp")
        if ext is None:
            ext = float(mat_gbp + lab_gbp)
        else:
            ext = float(ext)

        basis = str(cb.get("costing_basis") or "")
        bought = bool(pe.get("_bought_in_from_text_scan")) or "system_cost" in basis.lower()
        flag = "Bought-in" if bought else "Manufactured"

        costs = lab_e.get("costs_gbp") if isinstance(lab_e.get("costs_gbp"), dict) else {}
        op_bits = [f"{k}:{round(float(v or 0.0), 2)}" for k, v in sorted(costs.items())]
        operations = "; ".join(op_bits[:16])

        rf = pe.get("risk_flags") or []
        if isinstance(rf, list):
            notes = " | ".join(str(x) for x in rf if x)
        else:
            notes = str(rf)

        thk = pe.get("normalized_thickness_mm")
        if thk is None:
            thk = mat_e.get("thickness_mm")
        mat_name = pe.get("normalized_material") or mat_e.get("material") or ""

        r = _blank_row()
        r.update(
            {
                "part_number": pn,
                "description": str(pe.get("description") or wp.get("description") or ""),
                "quantity": pe.get("quantity") if pe.get("quantity") is not None else wp.get("quantity") or "",
                "ai_material": str(mat_name),
                "ai_thickness_mm": thk if thk is not None else "",
                "ai_bought_in_flag": flag,
                "ai_material_cost_gbp": round(mat_gbp, 2),
                "ai_labour_cost_gbp": round(lab_gbp, 2),
                "ai_total_cost_gbp": round(ext, 2),
                "manual_material_cost_gbp": "",
                "manual_labour_cost_gbp": "",
                "manual_total_cost_gbp": "",
                "variance_pct": "",
                "operations": operations,
                "notes": notes,
            }
        )
        rows.append(r)

    if not parts_est:
        tot = _blank_row()
        tot.update(
            {
                "part_number": "(none)",
                "description": "No part_estimates in estimate_summary",
                "notes": "Run scan / estimate_document so estimate_summary.part_estimates is populated",
            }
        )
        rows.append(tot)

    tot = _blank_row()
    data_rows = [r for r in rows if str(r.get("part_number")) != "TOTAL"]
    tot.update(
        {
            "part_number": "TOTAL",
            "description": "AI BOM extended totals",
            "ai_material_cost_gbp": round(sum(float(r.get("ai_material_cost_gbp") or 0.0) for r in data_rows), 2),
            "ai_labour_cost_gbp": round(sum(float(r.get("ai_labour_cost_gbp") or 0.0) for r in data_rows), 2),
            "ai_total_cost_gbp": round(sum(float(r.get("ai_total_cost_gbp") or 0.0) for r in data_rows), 2),
            "notes": "manual_* columns reserved for spreadsheet / ERP extract",
        }
    )
    rows = [r for r in rows if str(r.get("part_number")) != "TOTAL"]
    rows.append(tot)

    output_csv = output_csv.resolve()
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    actual, fh = _open_csv_write_preferred(output_csv, purpose="BOM comparison CSV")
    try:
        writer = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    finally:
        fh.close()
    return actual


def generate_and_write(
    summary_path: Path,
    workbook_path: Path,
    out_json: Path,
    out_csv: Path,
    *,
    read_via_excel: bool = False,
) -> Tuple[Dict[str, Any], Path, Path]:
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    bundle = build_full_parity_report(summary, workbook_path, read_via_excel=read_via_excel)
    written_json, written_csv = write_full_parity_reports(bundle, out_json, out_csv)
    return bundle, written_json, written_csv


def main() -> None:
    parser = argparse.ArgumentParser(description="Full Estimate workbook parity vs Manufacturing JSON")
    parser.add_argument("--summary-json", required=True)
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--out-json", default=str(config.CSV_DIR / "estimate_full_parity_bundle.json"))
    parser.add_argument("--out-csv", default=str(config.CSV_DIR / "estimate_full_parity_flat.csv"))
    parser.add_argument(
        "--read-via-excel",
        action="store_true",
        help="Windows only: read .xlsx/.xlsm through Excel COM (requires pywin32).",
    )
    args = parser.parse_args()
    bundle, out_j, out_c = generate_and_write(
        Path(args.summary_json),
        Path(args.workbook),
        Path(args.out_json),
        Path(args.out_csv),
        read_via_excel=bool(args.read_via_excel),
    )
    print(json.dumps(bundle.get("status_counts") or {}, indent=2))
    print("JSON:", out_j.resolve())
    print("CSV:", out_c.resolve())


if __name__ == "__main__":
    main()
